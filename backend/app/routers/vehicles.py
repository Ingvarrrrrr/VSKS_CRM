"""
Vehicles CRUD router — Plan 29-04, Phase 29 «Имущество → Автотранспорт».

Decisions covered: D-06 (visibility), D-08 (schema), D-10 (field history).

Endpoints:
  GET    /api/vehicles                   — list with multi-tenancy filter
  POST   /api/vehicles                   — create, require_action('vehicle.edit')
  PATCH  /api/vehicles/{vehicle_id}      — partial update + field history hook
  GET    /api/vehicles/{vehicle_id}/field-history — audit log
  GET    /api/vehicles/{vehicle_id}      — detail (LAST — catch-all int)
  DELETE /api/vehicles/{vehicle_id}      — hard delete, require_action('vehicle.delete')

Registration in __init__.py done in separate commit after Wave 1 (per plan constraint).
"""
from urllib.parse import quote as _url_quote
from fastapi import APIRouter, Depends, HTTPException, Query, Body
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy.orm import selectinload
from sqlalchemy.orm.attributes import flag_modified
from datetime import datetime, date, timezone
from typing import Any, List, Optional

from app.database import get_db
from app.auth.jwt import get_current_user, get_org_filter, ADMIN_ROLES
from app.auth.permissions import require_tab, require_action
from app.models.user import User
from app.models.vehicle import Vehicle
from app.models.vehicle_field_history import VehicleFieldHistory
from app.models.vehicle_transfer_history import VehicleTransferHistory
from app.models.organization import Organization
from app.schemas.vehicle import (
    VehicleOut,
    VehicleListItem,
    VehicleCreate,
    VehiclePatch,
    FieldHistoryOut,
    VehicleTransferHistoryOut,
)
from app.services.vehicle_fields import get_field_label, get_string_column_limits

router = APIRouter(prefix="/api/vehicles", tags=["vehicles"])

# ─────────────────────────── Field classification ───────────────────────────

_DATE_FIELDS = {
    "registered_at", "insurance_until", "last_to_date", "tech_inspection_until",
    "assignment_doc_date",  # Phase 29.3
    # Автоблок: новые date-колонки (AUTOBLOCK_FIELDS_SPEC.md §1)
    "ownership_doc_date", "owner_since", "sts_issued_at",
    "tech_inspection_last_date",
    # 2026-09: pass_*_until убраны — пропуска теперь отдельная таблица/API
    # (app/models/vehicle_pass.py, app/routers/vehicle_passes.py), эти колонки
    # больше не пишутся через PATCH /api/vehicles.
    "first_aid_kit_until", "extinguisher_check_date", "tracker_paid_until", "tachograph_check_date",
}
_DATETIME_FIELDS: set = set()
_BOOL_FIELDS = {
    "has_tracker", "akb_ok", "has_radio", "mirrors_ok",
    "has_keys", "has_first_aid_kit", "has_spare_wheel", "has_extinguisher",
    # Автоблок
    "has_spare_tires", "has_mirrors", "has_tachograph", "repair_required",
    "has_branding",  # 2026-09
}
_JSONB_FIELDS = {"props"}

# Fields logged to vehicle_field_history on every PATCH change
_TRACKED_FIELDS_FOR_HISTORY = {
    "state", "plate", "vin", "fuel_type",
    "fuel_norm_summer", "fuel_norm_winter",
    "insurance_until", "next_to_km",
    "assigned_org_id", "owner_org_id",
    "brand", "model", "color",
    "registered_at", "type",
}

# All columns the PATCH endpoint may write (whitelist; current_odometer_km excluded per plan)
_PATCHABLE_FIELDS = {
    "owner_org_id", "assigned_org_id", "assigned_text",
    "brand", "model", "color", "vin", "plate",
    "registered_at", "insurance_until",
    "type", "state",
    "fuel_type", "fuel_norm_summer", "fuel_norm_winter",
    "has_tracker", "akb_ok", "has_radio", "mirrors_ok",
    "has_keys", "has_first_aid_kit", "has_spare_wheel", "has_extinguisher",
    "next_to_km",
    "props",
    # Extended Голичков registry fields (Plan 29-3a2)
    "year_of_manufacture", "last_to_mileage_km", "last_to_date",
    "pts_number", "sts_number", "tech_inspection_until", "purchase_info",
    # Phase 29.3: assignment + engine
    "assignment_basis", "assignment_doc_number", "assignment_doc_date",
    "engine_power_hp", "engine_volume_l",
    # Автоблок: полный реестр полей ТС (AUTOBLOCK_FIELDS_SPEC.md §1)
    "body_type", "pts_category",
    "insurance_company", "insurance_policy_number",
    "ownership_basis", "ownership_doc_number", "ownership_doc_date", "owner_since",
    "location_city", "location_address", "home_base_city", "responsible_name",
    "pts_kind", "sts_issued_at",
    "tech_inspection_status", "tech_inspection_last_date",
    # 2026-09: pass_* убраны из белого списка — единственный источник правды
    # теперь vehicle_passes (app/routers/vehicle_passes.py). Колонки в БД
    # остались (данные перенесены миграцией), но PATCH /api/vehicles их больше
    # не принимает — чтобы не появилось два места записи одних и тех же данных.
    "has_spare_tires", "tires_condition", "has_mirrors",
    "first_aid_kit_until", "extinguisher_check_date", "tracker_paid_until",
    "has_tachograph", "tachograph_check_date",
    "repair_required", "tech_condition_info",
    # 2026-09: брендирование — признак + резина по сезонным комплектам
    "has_branding",
    "tires_summer_radius", "tires_summer_profile", "tires_summer_condition",
    "tires_winter_radius", "tires_winter_profile", "tires_winter_condition",
}


# ─────────────────────────── Helpers ────────────────────────────────────────

_STRING_COLUMN_LIMITS = get_string_column_limits()


def _check_string_length(field: str, value: Any) -> None:
    """HTTP 400 с внятной причиной, если строковое значение превышает лимит
    VARCHAR-колонки Vehicle (coordinator review 2026-08-31).

    Раньше проверки не было вовсе: значение долетало до asyncpg и падало
    500-кой с полным traceback в теле ответа (StringDataRightTruncationError).
    Лимиты — программно из модели (_STRING_COLUMN_LIMITS), подпись поля — из
    реестра app/services/vehicle_fields.py.
    """
    if not isinstance(value, str):
        return
    limit = _STRING_COLUMN_LIMITS.get(field)
    if limit is not None and len(value) > limit:
        label = get_field_label(field) or field
        raise HTTPException(
            status_code=400,
            detail=f"Поле «{label}» — не больше {limit} символов, введено {len(value)}.",
        )


def _validate_doc_dates(ownership_doc_date: Any, assignment_doc_date: Any) -> None:
    """Пункт 6 задания (2026-09): дата документа основания права ЭКСПЛУАТАЦИИ не
    может быть раньше даты документа основания права СОБСТВЕННОСТИ — нельзя было
    начать эксплуатировать по документу до того, как появилось само право
    собственности. Проверяем только когда обе даты заданы; текст ошибки — по-русски,
    с обеими датами (owner should see cause, not a generic 400)."""
    if not ownership_doc_date or not assignment_doc_date:
        return
    if assignment_doc_date < ownership_doc_date:
        raise HTTPException(
            status_code=400,
            detail=(
                f"Дата документа основания права эксплуатации ({assignment_doc_date.strftime('%d.%m.%Y')}) "
                f"не может быть раньше даты документа основания права собственности "
                f"({ownership_doc_date.strftime('%d.%m.%Y')})."
            ),
        )


def _coerce_patch_value(field: str, value: Any) -> Any:
    """Convert ISO string to date for DATE columns; empty string → None.

    Lesson 2026-05-13: per-router _DATE_FIELDS (not global), copy from purchases.py.
    """
    if value == "" or value is None:
        return None
    if field in _DATE_FIELDS and isinstance(value, str):
        try:
            return date.fromisoformat(value[:10])
        except Exception:
            return None
    if field in _DATETIME_FIELDS and isinstance(value, str):
        try:
            return datetime.fromisoformat(value.replace("Z", "+00:00"))
        except Exception:
            try:
                return date.fromisoformat(value[:10])
            except Exception:
                return None
    return value


def _visibility_q(user: User):
    """Base SELECT with multi-tenancy filter applied (D-06).

    owner_org_id OR assigned_org_id must be in user.org_ids.
    superadmin / account_owner: no filter (get_org_filter returns None).
    """
    q = select(Vehicle)
    org_ids = get_org_filter(user)  # None → superadmin/account_owner bypass
    if org_ids is not None:
        q = q.where(
            or_(
                Vehicle.owner_org_id.in_(org_ids),
                Vehicle.assigned_org_id.in_(org_ids),
            )
        )
    return q


def _check_vehicle_visibility(vehicle: Vehicle, user: User) -> None:
    """Raise 403 if user has no access to this specific vehicle (D-06)."""
    if user.role in ("superadmin", "account_owner"):
        return
    org_ids = get_org_filter(user)
    if org_ids is None:
        return
    if (
        vehicle.owner_org_id not in org_ids
        and vehicle.assigned_org_id not in org_ids
    ):
        raise HTTPException(status_code=403, detail="Нет доступа к этому ТС")


def _log_field_changes(
    db: AsyncSession,
    vehicle: Vehicle,
    old_dict: dict,
    new_dict: dict,
    user_id: int,
    comment: Optional[str] = None,
) -> None:
    """Write VehicleFieldHistory rows for actually-changed tracked fields (D-10).

    Compares str(old) vs str(new) to handle date/None comparisons uniformly.
    Only writes rows when value actually changed.
    """
    for field in _TRACKED_FIELDS_FOR_HISTORY:
        old_val = old_dict.get(field)
        new_val = new_dict.get(field)
        old_str = str(old_val) if old_val is not None else None
        new_str = str(new_val) if new_val is not None else None
        if old_str == new_str:
            continue
        db.add(
            VehicleFieldHistory(
                vehicle_id=vehicle.id,
                field_key=field,
                old_value=old_str,
                new_value=new_str,
                changed_at=datetime.now(timezone.utc),
                changed_by_user_id=user_id,
                comment=comment,
            )
        )


async def _fetch_org_names(db: AsyncSession, org_ids: list[int]) -> dict[int, str]:
    """Fetch {org_id: org_name} for given ids. Empty ids → empty dict."""
    unique = [i for i in set(org_ids) if i is not None]
    if not unique:
        return {}
    rows = (await db.execute(
        select(Organization.id, Organization.name).where(Organization.id.in_(unique))
    )).all()
    return {r.id: r.name for r in rows}


async def _fetch_org_data(db: AsyncSession, org_ids: list[int]) -> dict[int, dict]:
    """Fetch {org_id: {name, color, inn}} for given ids. Empty ids → empty dict.

    Автоблок: inn добавлен для вычисляемых полей owner_inn/operator_inn (§1).
    """
    unique = [i for i in set(org_ids) if i is not None]
    if not unique:
        return {}
    rows = (await db.execute(
        select(Organization.id, Organization.name, Organization.color, Organization.inn)
        .where(Organization.id.in_(unique))
    )).all()
    return {r.id: {"name": r.name, "color": r.color, "inn": r.inn} for r in rows}


def _enrich_vehicle_out(vehicle: Vehicle, org_map: dict[int, str], org_data: Optional[dict[int, dict]] = None) -> VehicleOut:
    """Build VehicleOut with owner_org_name / assigned_org_name filled from org_map."""
    out = VehicleOut.model_validate(vehicle)
    out.owner_org_name = org_map.get(vehicle.owner_org_id)
    out.assigned_org_name = org_map.get(vehicle.assigned_org_id) if vehicle.assigned_org_id else None
    if org_data:
        out.owner_org_color = (org_data.get(vehicle.owner_org_id) or {}).get("color")
        out.assigned_org_color = (org_data.get(vehicle.assigned_org_id) or {}).get("color") if vehicle.assigned_org_id else None
        # Автоблок: owner_inn/operator_inn — вычисляемые read-only поля (§1)
        out.owner_inn = (org_data.get(vehicle.owner_org_id) or {}).get("inn")
        out.operator_inn = (org_data.get(vehicle.assigned_org_id) or {}).get("inn") if vehicle.assigned_org_id else None
    # Phase 30: current_station = assigned_org если есть, иначе owner_org
    out.current_station_id = vehicle.assigned_org_id or vehicle.owner_org_id
    out.current_station_name = (
        org_map.get(vehicle.assigned_org_id) if vehicle.assigned_org_id
        else org_map.get(vehicle.owner_org_id)
    )
    return out


def _enrich_list_item(vehicle: Vehicle, org_map: dict[int, str], org_data: Optional[dict[int, dict]] = None) -> VehicleListItem:
    """Build VehicleListItem with org names filled."""
    out = VehicleListItem.model_validate(vehicle)
    out.owner_org_name = org_map.get(vehicle.owner_org_id)
    out.assigned_org_name = org_map.get(vehicle.assigned_org_id) if vehicle.assigned_org_id else None
    if org_data:
        out.owner_org_color = (org_data.get(vehicle.owner_org_id) or {}).get("color")
        out.assigned_org_color = (org_data.get(vehicle.assigned_org_id) or {}).get("color") if vehicle.assigned_org_id else None
    # Phase 30: current_station = assigned_org если есть, иначе owner_org
    out.current_station_id = vehicle.assigned_org_id or vehicle.owner_org_id
    out.current_station_name = (
        org_map.get(vehicle.assigned_org_id) if vehicle.assigned_org_id
        else org_map.get(vehicle.owner_org_id)
    )
    return out


# ─────────────────────────── Endpoints ──────────────────────────────────────

@router.get("", response_model=dict)
async def list_vehicles(
    state: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    owner_org_id: Optional[List[int]] = Query(None),
    assigned_org_id: Optional[List[int]] = Query(None),
    brand: Optional[str] = Query(None),
    plate: Optional[str] = Query(None),
    vin: Optional[str] = Query(None, description="Точный поиск по VIN (case-insensitive)"),
    q: Optional[str] = Query(None, description="Search brand/model/plate/vin"),
    limit: int = Query(50, ge=1, le=500),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_tab("vehicles")),
    db: AsyncSession = Depends(get_db),
):
    """GET /api/vehicles — paginated list with multi-tenancy filter."""
    base_q = _visibility_q(current_user)

    if state:
        base_q = base_q.where(Vehicle.state == state)
    if type:
        base_q = base_q.where(Vehicle.type == type)
    if owner_org_id:
        base_q = base_q.where(Vehicle.owner_org_id.in_(owner_org_id))
    if assigned_org_id:
        base_q = base_q.where(Vehicle.assigned_org_id.in_(assigned_org_id))
    if brand:
        base_q = base_q.where(Vehicle.brand.ilike(f"%{brand}%"))
    if plate:
        base_q = base_q.where(Vehicle.plate.ilike(f"%{plate}%"))
    if vin:
        base_q = base_q.where(func.upper(Vehicle.vin) == vin.upper())
    if q:
        pattern = f"%{q}%"
        base_q = base_q.where(
            or_(
                Vehicle.brand.ilike(pattern),
                Vehicle.model.ilike(pattern),
                Vehicle.plate.ilike(pattern),
                Vehicle.vin.ilike(pattern),
            )
        )

    # Total count
    count_q = select(func.count()).select_from(base_q.subquery())
    total = (await db.execute(count_q)).scalar_one()

    # Items
    items_q = base_q.order_by(Vehicle.id.desc()).limit(limit).offset(offset)
    items = (await db.execute(items_q)).scalars().all()

    # Enrich with org names (Option C: single lookup dict)
    all_org_ids = []
    for v in items:
        if v.owner_org_id:
            all_org_ids.append(v.owner_org_id)
        if v.assigned_org_id:
            all_org_ids.append(v.assigned_org_id)
    org_map = await _fetch_org_names(db, all_org_ids)
    org_data = await _fetch_org_data(db, all_org_ids)

    return {
        "items": [_enrich_list_item(v, org_map, org_data) for v in items],
        "total": total,
    }


_EXPORT_TYPE_LABEL = {
    'car_light': 'Легковой', 'suv': 'Внедорожник', 'pickup': 'Пикап',
    'minivan': 'Минивэн', 'truck_van': 'Фургон', 'truck_board': 'Грузовой',
    'truck_tank': 'Цистерна', 'truck_metal': 'Металловоз', 'bus': 'Автобус',
    'special': 'Спецтехника', 'quadbike': 'Квадроцикл', 'snowmobile': 'Снегоход',
    'boat': 'Лодка', 'boat_motor': 'Лод. мотор', 'trailer': 'Прицеп', 'other': 'Другой',
}
_EXPORT_STATE_LABEL = {
    'working': 'Рабочее', 'in_repair': 'В ремонте', 'broken': 'Сломан',
    'needs_repair': 'Требует ремонта', 'destroyed': 'Уничтожен', 'utilized': 'Утилизирован',
}
_EXPORT_PTS_KIND_LABEL = {'paper': 'Бумажный', 'electronic': 'Электронный'}


def _export_field_value(field: dict, vehicle: Vehicle, org_map: dict, org_inn_map: dict):
    """Автоблок §7: значение одного поля реестра для строки Excel-экспорта."""
    key = field["key"]
    storage = field["storage"]
    ftype = field["type"]

    if storage == "computed":
        if key == "owner_inn":
            return org_inn_map.get(vehicle.owner_org_id) or ''
        if key == "operator_inn":
            return (org_inn_map.get(vehicle.assigned_org_id) or '') if vehicle.assigned_org_id else ''
        return ''

    if storage == "props":
        return (vehicle.props or {}).get(field["props_key"]) or ''

    if key == "owner_org_id":
        return org_map.get(vehicle.owner_org_id) or ''
    if key == "assigned_org_id":
        return (org_map.get(vehicle.assigned_org_id) or '') if vehicle.assigned_org_id else ''

    val = getattr(vehicle, key, None)
    if val is None or val == '':
        # Автоблок (2026-09): гос. номер необязателен — понятная заглушка вместо
        # пустой ячейки в выгрузке, чтобы строка не выглядела как ошибка/пропуск.
        if key == "plate":
            return 'Без номера'
        return ''
    if key == "type":
        return _EXPORT_TYPE_LABEL.get(val, val)
    if key == "state":
        return _EXPORT_STATE_LABEL.get(val, val)
    if key == "pts_kind":
        return _EXPORT_PTS_KIND_LABEL.get(val, val)
    if ftype == "date" and hasattr(val, "strftime"):
        return val.strftime('%d.%m.%Y')
    if ftype == "bool":
        return 'Да' if val else 'Нет'
    return val


@router.get("/export/excel")
async def export_excel(
    state: Optional[str] = Query(None),
    type: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    user: User = Depends(require_tab("vehicles")),
):
    """Excel-выгрузка списка ТС.

    Автоблок §7: выгружает все поля реестра (services/vehicle_fields) в порядке
    групп, пропуская поля, скрытые для организации текущего пользователя.
    """
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from fastapi.responses import StreamingResponse
    from app.services.vehicle_fields import FIELD_GROUPS, get_hidden_field_keys

    q = select(Vehicle).options(
        selectinload(Vehicle.owner_org),
        selectinload(Vehicle.assigned_org),
    )
    if state:
        if state == 'in_repair':
            q = q.where(Vehicle.state.in_(['in_repair', 'broken', 'needs_repair']))
        elif state in ('not_running', 'disposal'):
            q = q.where(Vehicle.state.in_(['destroyed', 'utilized']))
        else:
            q = q.where(Vehicle.state == state)
    if type:
        q = q.where(Vehicle.type == type)
    org_ids = get_org_filter(user)
    if org_ids is not None:
        q = q.where(Vehicle.owner_org_id.in_(org_ids))

    vehicles = (await db.execute(q.order_by(Vehicle.plate))).scalars().all()

    hidden_keys = await get_hidden_field_keys(db, user.org_id)
    visible_fields = [f for g in FIELD_GROUPS for f in g["fields"] if f["key"] not in hidden_keys]

    # org name map (из уже загруженных relationship'ов) + inn map (§1 owner_inn/operator_inn)
    org_map: dict[int, str] = {}
    org_ids_for_inn: set[int] = set()
    for v in vehicles:
        if v.owner_org:
            org_map[v.owner_org_id] = v.owner_org.name
        if v.assigned_org:
            org_map[v.assigned_org_id] = v.assigned_org.name
        org_ids_for_inn.add(v.owner_org_id)
        if v.assigned_org_id:
            org_ids_for_inn.add(v.assigned_org_id)
    org_inn_map: dict[int, str] = {}
    if org_ids_for_inn:
        inn_rows = (await db.execute(
            select(Organization.id, Organization.inn).where(Organization.id.in_(org_ids_for_inn))
        )).all()
        org_inn_map = {r.id: r.inn for r in inn_rows if r.inn}

    wb = Workbook()
    ws = wb.active
    ws.title = "Автотранспорт"
    headers = ['№'] + [f["label"] for f in visible_fields]

    bold = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.font = bold
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r, v in enumerate(vehicles, 2):
        row = [r - 1] + [_export_field_value(f, v, org_map, org_inn_map) for f in visible_fields]
        for c, val in enumerate(row, 1):
            ws.cell(r, c, val)

    for idx, header in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(len(header) + 2, 30))
    ws.freeze_panes = 'A2'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f"attachment; filename*=UTF-8''{_url_quote('Транспорт.xlsx', safe='-_.~')}"},
    )


@router.post("", response_model=VehicleOut, status_code=201)
async def create_vehicle(
    body: VehicleCreate,
    current_user: User = Depends(require_action("vehicle.edit")),
    db: AsyncSession = Depends(get_db),
):
    """POST /api/vehicles — create new vehicle record."""
    # Org visibility check: non-admin may only create for their own orgs
    if current_user.role not in ADMIN_ROLES:
        org_ids = get_org_filter(current_user) or []
        if body.owner_org_id not in org_ids:
            raise HTTPException(
                status_code=403,
                detail="Нет прав создавать ТС для указанной организации-владельца",
            )

    # VIN duplicate check (before insert)
    if body.vin:
        existing_q = await db.execute(
            select(Vehicle.id, Vehicle.plate, Vehicle.brand, Vehicle.model)
            .where(func.upper(Vehicle.vin) == body.vin.upper())
            .limit(1)
        )
        dup = existing_q.first()
        if dup and not body.force:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DUPLICATE_VIN",
                    "vehicle_id": dup[0],
                    "plate": dup[1],
                    "brand": dup[2],
                    "model": dup[3],
                    "message": f"ТС с таким VIN уже существует: {dup[1]} ({dup[2]} {dup[3] or ''})".strip()
                }
            )

    vehicle_data = body.model_dump(exclude={"force"})
    for _field, _value in vehicle_data.items():
        _check_string_length(_field, _value)
    _validate_doc_dates(vehicle_data.get("ownership_doc_date"), vehicle_data.get("assignment_doc_date"))

    vehicle = Vehicle(**vehicle_data)
    db.add(vehicle)
    try:
        await db.commit()
        await db.refresh(vehicle)
    except Exception as exc:
        await db.rollback()
        exc_str = str(exc).lower()
        if "unique" in exc_str and "plate" in exc_str:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DUPLICATE_PLATE",
                    "message": "ТС с таким гос. номером уже существует",
                    "plate": body.plate,
                },
            )
        raise
    _org_ids_c = [vehicle.owner_org_id, vehicle.assigned_org_id]
    org_map = await _fetch_org_names(db, _org_ids_c)
    org_data = await _fetch_org_data(db, _org_ids_c)
    return _enrich_vehicle_out(vehicle, org_map, org_data)


@router.patch("/{vehicle_id}", response_model=VehicleOut)
async def patch_vehicle(
    vehicle_id: int,
    body: dict = Body(...),
    current_user: User = Depends(require_action("vehicle.edit")),
    db: AsyncSession = Depends(get_db),
):
    """PATCH /api/vehicles/{vehicle_id} — partial update with history hook (D-10).

    - _coerce_patch_value for DATE fields (Lesson 2026-05-13)
    - flag_modified for props JSONB (Gotcha 2026-05-13)
    - _log_field_changes writes VehicleFieldHistory rows for changed tracked fields
    - current_odometer_km is IGNORED even if sent (managed by vehicle_odometer.py)
    """
    vehicle = await db.get(Vehicle, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="ТС не найдено")
    _check_vehicle_visibility(vehicle, current_user)

    history_comment: Optional[str] = body.get("_history_comment")

    # Snapshot old values for history BEFORE any setattr
    old_snapshot = {
        f: getattr(vehicle, f)
        for f in _TRACKED_FIELDS_FOR_HISTORY
        if hasattr(vehicle, f)
    }

    # Первый проход: коэрсим и ВАЛИДИРУЕМ длину строк ДО единого setattr (coordinator
    # review 2026-08-31) — если поле №3 превышает лимит, поля №1-2 не должны успеть
    # замутировать ORM-объект перед тем, как запрос упадёт в 400.
    props_mutated = False
    coerced_items: list[tuple[str, Any]] = []
    for k, v in (body or {}).items():
        # Skip internal meta keys and non-patchable fields
        if k.startswith("_") or k not in _PATCHABLE_FIELDS:
            continue
        # current_odometer_km not patchable — managed by odometer endpoint
        if k == "current_odometer_km":
            continue
        if not hasattr(vehicle, k):
            continue

        v = _coerce_patch_value(k, v)
        _check_string_length(k, v)
        coerced_items.append((k, v))

    # Пункт 6 задания (2026-09): проверяем ЭФФЕКТИВНЫЕ значения обеих дат (тело
    # запроса могло менять только одну из двух — вторая берётся из уже
    # сохранённого значения), ДО единого setattr — та же логика, что и у
    # _check_string_length выше: невалидный PATCH не должен успеть замутировать
    # ORM-объект перед тем, как запрос упадёт в 400.
    _coerced_map = dict(coerced_items)
    _eff_ownership_date = (
        _coerced_map["ownership_doc_date"] if "ownership_doc_date" in _coerced_map
        else vehicle.ownership_doc_date
    )
    _eff_assignment_date = (
        _coerced_map["assignment_doc_date"] if "assignment_doc_date" in _coerced_map
        else vehicle.assignment_doc_date
    )
    _validate_doc_dates(_eff_ownership_date, _eff_assignment_date)

    for k, v in coerced_items:
        if getattr(vehicle, k) != v:
            setattr(vehicle, k, v)
            if k in _JSONB_FIELDS:
                props_mutated = True

    # flag_modified for JSONB before commit (Gotcha 2026-05-13)
    if props_mutated:
        flag_modified(vehicle, "props")

    # New snapshot for history comparison
    new_snapshot = {
        f: getattr(vehicle, f)
        for f in _TRACKED_FIELDS_FOR_HISTORY
        if hasattr(vehicle, f)
    }

    # Write history rows for changed tracked fields (D-10)
    _log_field_changes(db, vehicle, old_snapshot, new_snapshot, current_user.id, history_comment)

    # Auto-history hook (Phase 29.3): record transfer if ownership/assignment changed
    _transfer_fields = {"owner_org_id", "assigned_org_id", "assigned_text"}
    if any(k in body for k in _transfer_fields):
        changed_transfer = any(
            str(old_snapshot.get(f) or '') != str(new_snapshot.get(f) or '')
            for f in _transfer_fields
            if f in _TRACKED_FIELDS_FOR_HISTORY or f in {"assigned_text"}
        )
        # Compare explicitly for assigned_text (not in _TRACKED_FIELDS_FOR_HISTORY by default)
        old_assigned_text = old_snapshot.get("assigned_text") if "assigned_text" in old_snapshot else getattr(vehicle, "assigned_text", None)
        # Re-check: compare old owner/assigned vs new
        owner_changed = str(old_snapshot.get("owner_org_id")) != str(new_snapshot.get("owner_org_id"))
        assigned_org_changed = str(old_snapshot.get("assigned_org_id")) != str(new_snapshot.get("assigned_org_id"))
        assigned_text_changed = ("assigned_text" in body) and (
            (body.get("assigned_text") or None) != (getattr(vehicle, "assigned_text", None))
        )
        if owner_changed or assigned_org_changed or assigned_text_changed:
            db.add(VehicleTransferHistory(
                vehicle_id=vehicle.id,
                from_owner_org_id=old_snapshot.get("owner_org_id"),
                to_owner_org_id=new_snapshot.get("owner_org_id"),
                from_assigned_org_id=old_snapshot.get("assigned_org_id"),
                to_assigned_org_id=new_snapshot.get("assigned_org_id"),
                from_assigned_text=vehicle.assigned_text if not assigned_text_changed else None,
                to_assigned_text=body.get("assigned_text") if "assigned_text" in body else None,
                basis=body.get("assignment_basis") or getattr(vehicle, "assignment_basis", None),
                doc_number=body.get("assignment_doc_number") or getattr(vehicle, "assignment_doc_number", None),
                doc_date=_coerce_patch_value("assignment_doc_date", body.get("assignment_doc_date") or getattr(vehicle, "assignment_doc_date", None)),
                comment=history_comment,
                changed_by_user_id=current_user.id,
            ))

    try:
        await db.commit()
        await db.refresh(vehicle)
    except Exception as exc:
        await db.rollback()
        exc_str = str(exc).lower()
        if "unique" in exc_str and "plate" in exc_str:
            new_plate = body.get("plate", vehicle.plate)
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "DUPLICATE_PLATE",
                    "message": "ТС с таким гос. номером уже существует",
                    "plate": new_plate,
                },
            )
        raise

    _org_ids_p = [vehicle.owner_org_id, vehicle.assigned_org_id]
    org_map = await _fetch_org_names(db, _org_ids_p)
    org_data = await _fetch_org_data(db, _org_ids_p)
    return _enrich_vehicle_out(vehicle, org_map, org_data)


@router.get("/distinct/{field}")
async def vehicle_distinct_values(
    field: str,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab("vehicles")),
):
    """GET /api/vehicles/distinct/{field} — уникальные значения колонки для autocomplete.
    Только whitelisted поля (защита от SQL-injection через имя колонки)."""
    ALLOWED = {
        'brand', 'model', 'color', 'assigned_text',
        'assignment_basis', 'assignment_doc_number',
        'pts_number', 'sts_number',
    }
    if field not in ALLOWED:
        raise HTTPException(400, f"Field '{field}' not allowed")
    col = getattr(Vehicle, field)
    res = await db.execute(
        select(col).where(col.isnot(None), col != '').distinct().order_by(col).limit(500)
    )
    values = [v for v in res.scalars().all() if v]
    return values


@router.get("/{vehicle_id}/field-history", response_model=List[FieldHistoryOut])
async def get_vehicle_field_history(
    vehicle_id: int,
    field_key: Optional[str] = Query(None, description="Filter by specific field"),
    limit: int = Query(100, ge=1, le=1000),
    offset: int = Query(0, ge=0),
    current_user: User = Depends(require_tab("vehicles")),
    db: AsyncSession = Depends(get_db),
):
    """GET /api/vehicles/{vehicle_id}/field-history — change audit log."""
    vehicle = await db.get(Vehicle, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="ТС не найдено")
    _check_vehicle_visibility(vehicle, current_user)

    q = (
        select(VehicleFieldHistory)
        .where(VehicleFieldHistory.vehicle_id == vehicle_id)
        .order_by(VehicleFieldHistory.changed_at.desc())
    )
    if field_key:
        q = q.where(VehicleFieldHistory.field_key == field_key)

    q = q.limit(limit).offset(offset)
    rows = (await db.execute(q)).scalars().all()
    return [FieldHistoryOut.model_validate(r) for r in rows]


@router.get("/{vehicle_id}/transfer-history", response_model=List[VehicleTransferHistoryOut])
async def get_vehicle_transfer_history(
    vehicle_id: int,
    current_user: User = Depends(require_tab("vehicles")),
    db: AsyncSession = Depends(get_db),
):
    """GET /api/vehicles/{vehicle_id}/transfer-history — история передач ТС (Phase 29.3)."""
    vehicle = await db.get(Vehicle, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="ТС не найдено")
    _check_vehicle_visibility(vehicle, current_user)

    rows = (await db.execute(
        select(VehicleTransferHistory)
        .where(VehicleTransferHistory.vehicle_id == vehicle_id)
        .order_by(VehicleTransferHistory.changed_at.desc())
    )).scalars().all()
    return [VehicleTransferHistoryOut.model_validate(r) for r in rows]


@router.get("/{vehicle_id}", response_model=VehicleOut)
async def get_vehicle(
    vehicle_id: int,
    current_user: User = Depends(require_tab("vehicles")),
    db: AsyncSession = Depends(get_db),
):
    """GET /api/vehicles/{vehicle_id} — full vehicle card.

    Registered LAST inside this router to avoid swallowing sub-path routes.
    Sub-routers (attachments, repairs, etc.) use /api/vehicle-* prefix (plan 29-05).
    """
    vehicle = await db.get(Vehicle, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="ТС не найдено")
    _check_vehicle_visibility(vehicle, current_user)
    _org_ids_g = [vehicle.owner_org_id, vehicle.assigned_org_id]
    org_map = await _fetch_org_names(db, _org_ids_g)
    org_data = await _fetch_org_data(db, _org_ids_g)
    return _enrich_vehicle_out(vehicle, org_map, org_data)


@router.delete("/{vehicle_id}", status_code=204)
async def delete_vehicle(
    vehicle_id: int,
    current_user: User = Depends(require_action("vehicle.delete")),
    db: AsyncSession = Depends(get_db),
):
    """DELETE /api/vehicles/{vehicle_id} — hard delete, CASCADE via FK constraints.

    Phase 29.3-R3 (Д-3): only admin+ can delete vehicles.
    FK cascades: vehicle_attachments, vehicle_repairs, vehicle_odometer,
                 fuel_logs, trips, vehicle_field_history.
    """
    # Phase 29.3-R3 (Д-3): explicit role check — только admin / superadmin
    if current_user.role not in ("admin", "superadmin", "account_owner"):
        raise HTTPException(
            status_code=403,
            detail="Удалять ТС может только администратор."
        )

    vehicle = await db.get(Vehicle, vehicle_id)
    if not vehicle:
        raise HTTPException(status_code=404, detail="ТС не найдено")
    _check_vehicle_visibility(vehicle, current_user)

    await db.delete(vehicle)
    await db.commit()
