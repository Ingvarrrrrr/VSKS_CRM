import io
from datetime import date, timedelta
from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, case
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.models.purchase import Purchase
from app.models.purchase_item import PurchaseItem
from app.models.event import Event
from app.models.subsidy import Subsidy
from app.models.contractor import Contractor
from app.models.feo_category import FeoCategory
from app.models.user import User
from app.auth.jwt import get_current_user
from app.auth.visibility import get_visible_subsidy_ids
from typing import Optional

router = APIRouter(prefix="/api/reports", tags=["reports"])


def _get_period_dates(period: str, ref_date: date):
    """Calculate start and end dates for a period."""
    if period == "week":
        start = ref_date - timedelta(days=ref_date.weekday())  # Monday
        end = start + timedelta(days=6)
    elif period == "month":
        start = ref_date.replace(day=1)
        # Last day of month
        if start.month == 12:
            end = start.replace(year=start.year + 1, month=1, day=1) - timedelta(days=1)
        else:
            end = start.replace(month=start.month + 1, day=1) - timedelta(days=1)
    elif period == "year":
        start = ref_date.replace(month=1, day=1)
        end = ref_date.replace(month=12, day=31)
    else:
        start = ref_date - timedelta(days=7)
        end = ref_date
    return start, end


def _apply_vis_filter(q, vis):
    """Apply subsidy visibility gate. vis=None means no filter (SaaS), vis=set → filter by subsidy_id."""
    if vis is not None:
        q = q.where(Purchase.subsidy_id.in_(vis))
    return q


def _purchase_to_dict(p, contractors: dict, subsidies: dict, users: dict):
    return {
        "id": p.id,
        "subject": p.subject or p.item_name or "",
        "status": p.status,
        "purchase_number": p.purchase_number,
        "registry_number": p.registry_number,
        "planned_total_price": float(p.planned_total_price or 0),
        "contract_price": float(p.contract_price or 0),
        "payment_amount": float(p.payment_amount or 0),
        "execution_term": str(p.execution_term) if p.execution_term else None,
        "delivery_date": str(p.delivery_date) if p.delivery_date else None,
        "contractor_name": contractors.get(p.contractor_id, ""),
        "subsidy_name": subsidies.get(p.subsidy_id, ""),
        "assigned_user_name": users.get(p.assigned_user_id, ""),
        "task_comment": p.task_comment,
    }


@router.get("/summary")
async def report_summary(
    period: str = Query("week", regex="^(week|month|year)$"),
    ref_date: Optional[str] = Query(None),
    subsidy_id: Optional[int] = Query(None),
    department: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    today = date.fromisoformat(ref_date) if ref_date else date.today()
    start, end = _get_period_dates(period, today)

    # Reports tab: двухуровневая видимость по вкладке «Отчёты».
    _vis = await get_visible_subsidy_ids(current_user, db, "reports")

    # Load lookup tables
    contractors_r = await db.execute(select(Contractor.id, Contractor.name))
    contractors = {r.id: r.name for r in contractors_r}
    subsidies_r = await db.execute(select(Subsidy.id, Subsidy.name))
    subsidies = {r.id: r.name for r in subsidies_r}
    users_r = await db.execute(select(User.id, User.full_name, User.username, User.department))
    users = {r.id: r.full_name or r.username for r in users_r}

    # Department filter: find user IDs belonging to the department
    dept_user_ids = None
    if department:
        dept_user_ids = {r.id for r in users_r if r.department and r.department.lower() == department.lower()}

    def _extra_filters(q):
        """Apply subsidy_id and department filters."""
        if subsidy_id is not None:
            q = q.where(Purchase.subsidy_id == subsidy_id)
        if dept_user_ids is not None:
            q = q.where(Purchase.assigned_user_id.in_(dept_user_ids))
        return q

    # Active purchases (in_progress, contracted)
    active_q = _extra_filters(_apply_vis_filter(
        select(Purchase).where(Purchase.status.in_(["in_progress", "contracted"])),
        _vis
    ))
    active_result = await db.execute(active_q)
    active = [_purchase_to_dict(p, contractors, subsidies, users) for p in active_result.scalars().all()]

    # Completed in period (status=paid, payment_date in range)
    completed_q = _extra_filters(_apply_vis_filter(
        select(Purchase)
        .where(Purchase.status == "paid")
        .where(Purchase.payment_doc_date.between(start, end)),
        _vis
    ))
    completed_result = await db.execute(completed_q)
    completed = [_purchase_to_dict(p, contractors, subsidies, users) for p in completed_result.scalars().all()]

    # Planned (status=planned or work_in_progress, created/updated in period — use id proxy)
    planned_q = _extra_filters(_apply_vis_filter(
        select(Purchase)
        .where(Purchase.status.in_(["planned", "work_in_progress"])),
        _vis
    ))
    planned_result = await db.execute(planned_q)
    planned = [_purchase_to_dict(p, contractors, subsidies, users) for p in planned_result.scalars().all()]

    # Upcoming deadlines (execution_term in next period)
    next_start, next_end = _get_period_dates(period, end + timedelta(days=1))
    upcoming_q = _extra_filters(_apply_vis_filter(
        select(Purchase)
        .where(Purchase.execution_term.between(next_start, next_end))
        .where(Purchase.status.notin_(["paid", "delivered"])),
        _vis
    ))
    upcoming_result = await db.execute(upcoming_q)
    upcoming = [_purchase_to_dict(p, contractors, subsidies, users) for p in upcoming_result.scalars().all()]

    # Overdue
    overdue_q = _extra_filters(_apply_vis_filter(
        select(Purchase)
        .where(Purchase.execution_term < today)
        .where(Purchase.status.notin_(["paid", "delivered"])),
        _vis
    ))
    overdue_result = await db.execute(overdue_q)
    overdue = [_purchase_to_dict(p, contractors, subsidies, users) for p in overdue_result.scalars().all()]

    # Totals
    totals = {
        "planned_count": len(planned),
        "planned_sum": sum(p["planned_total_price"] for p in planned),
        "active_count": len(active),
        "active_sum": sum(p["contract_price"] or p["planned_total_price"] for p in active),
        "completed_count": len(completed),
        "completed_sum": sum(p["payment_amount"] for p in completed),
        "upcoming_count": len(upcoming),
        "overdue_count": len(overdue),
    }

    return {
        "period": period,
        "start_date": str(start),
        "end_date": str(end),
        "planned": planned,
        "active": active,
        "completed": completed,
        "upcoming": upcoming,
        "overdue": overdue,
        "totals": totals,
    }


@router.get("/subsidy/{subsidy_id}/xlsx")
async def export_subsidy_report_xlsx(
    subsidy_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """
    Генерирует Excel-файл Приложение №3 по субсидии.
    21 колонка (A–U), по одной строке на закупку.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        from fastapi import HTTPException
        raise HTTPException(500, "openpyxl не установлен")

    # ── Загрузка данных ───────────────────────────────────────────────────────
    subsidy = await db.get(Subsidy, subsidy_id)
    if not subsidy:
        from fastapi import HTTPException
        raise HTTPException(404, "Субсидия не найдена")

    events_r = await db.execute(select(Event).where(Event.subsidy_id == subsidy_id).order_by(Event.id))
    events = {e.id: e for e in events_r.scalars().all()}

    purchases_r = await db.execute(
        select(Purchase).where(Purchase.subsidy_id == subsidy_id).order_by(Purchase.event_id, Purchase.id)
    )
    purchases = purchases_r.scalars().all()

    contractor_ids = {p.contractor_id for p in purchases if p.contractor_id}
    feo_ids = {p.feo_category_id for p in purchases if p.feo_category_id}

    contractors: dict = {}
    if contractor_ids:
        cr = await db.execute(select(Contractor).where(Contractor.id.in_(contractor_ids)))
        for c in cr.scalars().all():
            contractors[c.id] = c

    feo_cats: dict = {}
    if feo_ids:
        fr = await db.execute(select(FeoCategory).where(FeoCategory.id.in_(feo_ids)))
        for f in fr.scalars().all():
            feo_cats[f.id] = f

    # ── Workbook ──────────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Приложение №3"

    # ── Стили ─────────────────────────────────────────────────────────────────
    thin = Side(style="thin")
    thick = Side(style="medium")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(fill_type="solid", fgColor="BDD7EE")   # голубой
    sub_fill    = PatternFill(fill_type="solid", fgColor="DDEBF7")   # светло-голубой
    title_font  = Font(bold=True, size=11)
    hdr_font    = Font(bold=True, size=9)
    data_font   = Font(size=9)
    wrap_align  = Alignment(wrap_text=True, vertical="top", horizontal="left")
    center_align = Alignment(wrap_text=True, vertical="center", horizontal="center")

    def _border(cell):
        cell.border = border_all

    def _hdr(cell, text, fill=header_fill):
        cell.value = text
        cell.font = hdr_font
        cell.fill = fill
        cell.alignment = center_align
        _border(cell)

    # ── Строка 1 — Заголовок ──────────────────────────────────────────────────
    ws.merge_cells("A1:U1")
    c = ws["A1"]
    c.value = (
        f"Первичные данные о ходе предоставления субсидии\n"
        f"«{subsidy.name}»"
    )
    c.font = Font(bold=True, size=12)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.row_dimensions[1].height = 36

    # ── Строка 2 — Период/дата ────────────────────────────────────────────────
    ws.merge_cells("A2:U2")
    ws["A2"].value = f"Отчётный период: {date.today().strftime('%d.%m.%Y')}"
    ws["A2"].font = Font(italic=True, size=9)
    ws["A2"].alignment = Alignment(horizontal="right")

    # ── Строки 3-4 — Заголовки (2 уровня) ────────────────────────────────────
    # Row 3: main headers (merged across sub-columns where needed)
    COLS = [
        ("A", "A", "Наименование\nмероприятия"),
        ("B", "B", "Регион\nпроведения"),
        ("C", "C", "Период\nпроведения"),
        ("D", "D", "Реквизиты\nприказа"),
        ("E", "E", "Плановые\nпоказатели"),
        ("F", "F", "Планируемые\nзатраты, руб."),
        ("G", "G", "Фактически достигнутые\nпоказатели"),
        ("H", "H", "Освещение\nв СМИ"),
        ("I", "U", "Договоры/контракты"),
    ]
    col_letters_sub = {
        "I": "Наименование\nконтрагента",
        "J": "ИНН",
        "K": "Реквизиты\nдоговора",
        "L": "Сумма\nдоговора, руб.",
        "M": "Оплачено,\nруб.",
        "N": "Предмет\nдоговора",
        "O": "Краткое описание ТЗ",
        "P": "Страна\nпроизводства",
        "Q": "Реквизиты\nакта",
        "R": "Реквизиты\nплатёжного\nпоручения",
        "S": "Казначейский\nкод",
        "T": "Статья расходов\nпо ФЭО",
        "U": "Претензионная\nработа",
    }

    # Row 3 headers
    for (start_col, end_col, label) in COLS:
        if start_col == end_col:
            cell = ws[f"{start_col}3"]
            _hdr(cell, label)
            ws[f"{start_col}4"].value = ""
            ws[f"{start_col}4"].border = border_all
            ws[f"{start_col}4"].fill = sub_fill
            ws.merge_cells(f"{start_col}3:{end_col}4")
        else:
            ws.merge_cells(f"{start_col}3:{end_col}3")
            cell = ws[f"{start_col}3"]
            _hdr(cell, label)
            # row 4 sub-headers
            for col_letter, sub_label in col_letters_sub.items():
                _hdr(ws[f"{col_letter}4"], sub_label, fill=sub_fill)

    ws.row_dimensions[3].height = 32
    ws.row_dimensions[4].height = 42

    # ── Строка 5 — Номера колонок ─────────────────────────────────────────────
    for i, letter in enumerate("ABCDEFGHIJKLMNOPQRSTU", start=1):
        c = ws[f"{letter}5"]
        c.value = i
        c.font = hdr_font
        c.alignment = center_align
        c.fill = sub_fill
        c.border = border_all

    # ── Ширины колонок ────────────────────────────────────────────────────────
    col_widths = {
        "A": 30, "B": 15, "C": 18, "D": 20, "E": 22,
        "F": 16, "G": 22, "H": 22, "I": 25, "J": 14,
        "K": 20, "L": 16, "M": 14, "N": 30, "O": 30,
        "P": 14, "Q": 20, "R": 20, "S": 14, "T": 22, "U": 12,
    }
    for letter, width in col_widths.items():
        ws.column_dimensions[letter].width = width

    # ── Данные ────────────────────────────────────────────────────────────────
    def fmt_date(d):
        return d.strftime("%d.%m.%Y") if d else ""

    def fmt_money(v):
        return f"{float(v):,.2f}".replace(",", " ") if v is not None else ""

    row = 6
    for p in purchases:
        ev = events.get(p.event_id) if p.event_id else None
        contractor = contractors.get(p.contractor_id) if p.contractor_id else None
        feo = feo_cats.get(p.feo_category_id) if p.feo_category_id else None

        # Period string
        if ev and ev.date_from and ev.date_to:
            period_str = f"{fmt_date(ev.date_from)} — {fmt_date(ev.date_to)}"
        elif ev and ev.date_from:
            period_str = fmt_date(ev.date_from)
        else:
            period_str = ""

        # Media links
        media_parts = [l for l in [
            ev.media_link_1 if ev else None,
            ev.media_link_2 if ev else None,
            ev.media_link_3 if ev else None,
        ] if l]
        media_str = "\n".join(media_parts)

        # Items as subject
        items_r = await db.execute(
            select(PurchaseItem).where(PurchaseItem.purchase_id == p.id)
        )
        items_list = items_r.scalars().all()
        subject_str = p.subject or p.item_name or ""
        if not subject_str and items_list:
            subject_str = "; ".join(i.item_name for i in items_list if i.item_name)

        # TZ description — get from products via product_id
        tz_str = ""
        product_ids = [i.product_id for i in items_list if i.product_id]
        if product_ids:
            from app.models.product import Product
            prods_r = await db.execute(select(Product).where(Product.id.in_(product_ids)))
            prods = {pr.id: pr for pr in prods_r.scalars().all()}
            descs = []
            for i in items_list:
                if i.product_id and i.product_id in prods:
                    pr = prods[i.product_id]
                    d = pr.description or ""
                    if d:
                        descs.append(d)
            tz_str = "\n".join(descs[:3])

        # Act & payment
        act_str = ""
        if p.acceptance_doc_name or p.acceptance_doc_number:
            parts = []
            if p.acceptance_doc_name:
                parts.append(p.acceptance_doc_name)
            if p.acceptance_doc_number:
                parts.append(f"№{p.acceptance_doc_number}")
            if p.acceptance_doc_date:
                parts.append(f"от {fmt_date(p.acceptance_doc_date)}")
            act_str = " ".join(parts)

        pay_str = ""
        if p.payment_doc_number:
            parts = [f"ПП №{p.payment_doc_number}"]
            if p.payment_doc_date:
                parts.append(f"от {fmt_date(p.payment_doc_date)}")
            pay_str = " ".join(parts)

        feo_str = ""
        if feo:
            feo_str = f"{feo.code} {feo.name}" if feo.code else feo.name

        row_data = [
            ev.name if ev else "",                                      # A
            ev.region if ev else "",                                    # B
            period_str,                                                 # C
            ev.order_decree if ev else "",                              # D
            ev.planned_indicators if ev else "",                        # E
            fmt_money(p.planned_total_price),                          # F
            ev.actual_indicators if ev else "",                         # G
            media_str,                                                  # H
            contractor.name if contractor else "",                      # I
            contractor.inn if contractor else "",                       # J
            f"№{p.contract_number or ''} от {fmt_date(p.contract_date)}" if p.contract_number else "",  # K
            fmt_money(p.contract_price or p.planned_total_price),      # L
            fmt_money(p.payment_amount),                               # M
            subject_str,                                               # N
            tz_str,                                                    # O
            p.country_origin or "",                                    # P
            act_str,                                                   # Q
            pay_str,                                                   # R
            p.treasury_code or "",                                     # S
            feo_str,                                                   # T
            "Да" if p.has_pretension else "Нет",                      # U
        ]

        for col_idx, value in enumerate(row_data, start=1):
            c = ws.cell(row=row, column=col_idx, value=value)
            c.font = data_font
            c.alignment = wrap_align
            c.border = border_all

        ws.row_dimensions[row].height = 45
        row += 1

    if row == 6:
        # No data — add empty row
        for col_idx in range(1, 22):
            c = ws.cell(row=6, column=col_idx, value="")
            c.border = border_all

    # ── Freeze header ─────────────────────────────────────────────────────────
    ws.freeze_panes = "A6"

    # ── Сохранить в буфер ─────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    import urllib.parse
    filename_ru = f"Приложение_3_{subsidy.name}_{date.today()}.xlsx"
    filename_ascii = f"Appendix3_subsidy{subsidy_id}_{date.today()}.xlsx"
    encoded = urllib.parse.quote(filename_ru)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=\"{filename_ascii}\"; filename*=UTF-8''{encoded}"},
    )
