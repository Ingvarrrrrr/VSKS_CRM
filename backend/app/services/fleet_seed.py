"""
Fleet seed from Голичков xlsx — Phase 30-PR1.

Импортирует данные из Передача и Лист4 листов xlsx.

СТРУКТУРА ФАЙЛА (из inspect 2026-05-21):
  Лист2       — полный реестр на ранние даты (22 ТС), уже импортирован в Phase 29
  09.04.2026  — текущий реестр 51 ТС, уже импортирован в Phase 29 vehicles_seed.py
  20.05.24    — срез пробегов на 20.05.24 (гос.знак, VIN, пробег, дата ТО)
  25.06.2024  — срез пробегов на 25.06.24
  01.07.2024  — срез пробегов на 01.07.24
  2025        — срез пробегов на 2025 (без пробегов, только статус)
  Лист4       — 5 колонок: Штаб, Марка/Модель, Тип, VIN, Гос.рег.знак
               НЕТ явных полей документов — TODO: уточнить назначение листа
  Передача    — Организация, Марка, Модель, Год, VIN, Гос.рег.знак, Статус, доп.данные
               Это реестр текущего распределения по штабам, а НЕ история передач
               TODO: выяснить у Голичкова есть ли отдельная история передач

NOTE: Кириллица в шапках читается в mojibake при read_only=True. Используем позиционный
парсинг по индексу колонки, а не по имени заголовка.
"""
import os
import logging
from typing import Optional

log = logging.getLogger(__name__)

# Путь к xlsx (bundled seed_data копия)
XLSX_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', '..', 'seed_data', 'vehicles_golichkov.xlsx'
)


async def seed_fleet_from_xlsx(db, xlsx_path: Optional[str] = None) -> dict:
    """
    Идемпотентно импортирует данные из листов Передача и Лист4.

    Возвращает dict с ключами:
      - reason: str (если пропущено)
      - transfers_imported: int
      - documents_imported: int
      - lист4_rows: int (количество строк в Лист4 для ревью)
    """
    path = xlsx_path or XLSX_PATH
    if not os.path.exists(path):
        log.warning(f"fleet_seed: xlsx не найден: {path}")
        return {'reason': 'xlsx_not_found', 'path': path}

    try:
        import openpyxl
    except ImportError:
        log.warning("fleet_seed: openpyxl не установлен — пропускаем seed")
        return {'reason': 'openpyxl_not_installed'}

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        log.warning(f"fleet_seed: не удалось открыть xlsx: {e}")
        return {'reason': f'xlsx_open_error: {e}'}

    transfers_imported = 0
    documents_imported = 0
    lист4_rows = 0

    # ── Лист "Передача" (последний лист в sheetnames) ─────────────────────────
    # Структура: номер, Организация(штаб), Марка, Модель, Год, VIN, Гос.рег.знак,
    #            Статус(работает/не работает), доп.значение
    # Это реестр текущего закрепления ТС, НЕ история передач.
    # TODO: уточнить структуру с Голичковым — нужна история передач с датами/основанием.
    # Пока пишем диагностику:
    передача_sheet = None
    for sname in wb.sheetnames:
        # Ищем лист с "Передача" (с учётом mojibake — последний непонятный лист)
        # По позиции в sheetnames: wb.sheetnames[-1] — "Передача"
        pass
    передача_sheet = wb.sheetnames[-1]  # последний лист = "Передача"

    if передача_sheet in wb.sheetnames:
        ws = wb[передача_sheet]
        row_count = 0
        for row in ws.iter_rows(min_row=1, max_row=1000, values_only=True):
            if any(c is not None for c in row):
                row_count += 1
        log.info(
            f"fleet_seed: лист '{передача_sheet}' содержит {row_count} строк. "
            f"Парсер истории передач — TODO (нет полей from_org/to_org/date/basis). "
            f"Ручное ревью: нужна история с датами и основанием."
        )
        # TODO: реализовать парсер когда Голичков предоставит файл с историей передач
        # Пример структуры которую ожидаем:
        # Номер | Дата передачи | Гос.знак | VIN | Откуда (орг) | Куда (орг) | Основание | Документ
        transfers_imported = 0

    # ── Лист "Лист4" (предпоследний лист) ────────────────────────────────────
    # Структура: Штаб | Марка/Модель | Тип | VIN | Гос.рег.знак
    # 5 колонок. Это перечень ТС по штабам без документов.
    # TODO: уточнить у Голичкова — документы ли это? или просто ещё один срез реестра?
    лист4_sheet = None
    for sname in wb.sheetnames:
        pass
    # По позиции в sheetnames: wb.sheetnames[-2] — "Лист4"
    if len(wb.sheetnames) >= 2:
        лист4_sheet = wb.sheetnames[-2]
        ws4 = wb[лист4_sheet]
        rows4 = []
        for row in ws4.iter_rows(min_row=1, max_row=500, values_only=True):
            if any(c is not None for c in row):
                rows4.append(row)
        lист4_rows = len(rows4)
        log.info(
            f"fleet_seed: лист '{лист4_sheet}' содержит {lист4_rows} строк. "
            f"Структура: 5 колонок (Штаб/Марка+Модель/Тип/VIN/Гос.знак). "
            f"Это перечень ТС по штабам — НЕ история документов. "
            f"Импорт документов — TODO."
        )
        documents_imported = 0

    wb.close()

    result = {
        'transfers_imported': transfers_imported,
        'documents_imported': documents_imported,
        'лист4_rows': lист4_rows,
        'note': (
            'Парсеры истории передач и документов — TODO: '
            'структура листов не содержит полей from/to/дата/основание для transfers. '
            'Нужен файл с историческими передачами от Голичкова.'
        ),
    }
    log.info(f"fleet_seed: результат: {result}")
    return result


# ── Путь к ВСКС xlsx (одна копия для обоих seed'ов) ──────────────────────────
WAYBILL_XLSX_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    '..', '..', 'seed_data', 'waybill_template_vsks.xlsx'
)


async def seed_drivers_from_xlsx(db, xlsx_path: Optional[str] = None) -> dict:
    """
    Идемпотентно создаёт users с fleet_role='driver' из листа 'Водители' xlsx ВСКС.

    Структура листа (row 1 = header, rows 2+):
      0: Фамилия Имя Отчество
      1: ФИО (краткое)
      2: Табельный номер
      3: Табельный номер система
      4: СНИЛС
      5: Водительское удостоверение  ("61 33 920143")
      6: Дата выдачи (datetime или None)
      7: Категории  ("B,B1,C,C1,M")

    Возвращает {'created': N, 'updated': N, 'skipped': N} или {'reason': '...'}.
    """
    path = xlsx_path or WAYBILL_XLSX_PATH
    if not os.path.exists(path):
        log.warning(f"drivers_seed: xlsx не найден: {path}")
        return {'reason': 'xlsx_not_found', 'path': path}

    try:
        import openpyxl
    except ImportError:
        return {'reason': 'openpyxl_not_installed'}

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        log.warning(f"drivers_seed: не удалось открыть xlsx: {e}")
        return {'reason': f'xlsx_open_error: {e}'}

    sheet_name = 'Водители'
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {'reason': 'no_sheet', 'available': wb.sheetnames}

    from sqlalchemy import select
    from app.models.user import User
    from app.auth.jwt import hash_password
    from app.services.fio import resolve_user_name_input

    ws = wb[sheet_name]
    created, updated, skipped = 0, 0, 0

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Распаковываем 8 колонок с fallback на None
        cols = (list(row) + [None] * 8)[:8]
        full_name, short_name, tab_num, tab_sys, snils, vu, vu_date, categories = cols

        if not full_name or not str(full_name).strip():
            skipped += 1
            continue
        full_name = str(full_name).strip()

        # Парсим ВУ: "61 33 920143" → series="61 33", number="920143"
        vu_series, vu_number = '', ''
        if vu:
            parts = str(vu).strip().split()
            if len(parts) >= 3:
                vu_series = f"{parts[0]} {parts[1]}"
                vu_number = parts[2]
            elif parts:
                vu_number = ' '.join(parts)

        # Username = driver_{табельный без нулей}, fallback = driver_{первое слово фамилии lower}
        if tab_num:
            suffix = str(tab_num).lstrip('0') or str(tab_num)
        else:
            suffix = full_name.split()[0].lower()
        username = f"driver_{suffix}"

        # Нормализуем дату выдачи ВУ
        license_date = None
        if vu_date is not None:
            if hasattr(vu_date, 'date'):
                license_date = vu_date.date()
            else:
                license_date = vu_date  # уже date или строка

        existing = (await db.execute(
            select(User).where(User.username == username)
        )).scalars().first()

        row_last, row_first, row_middle, row_full = resolve_user_name_input(None, None, None, full_name)

        if existing:
            existing.fleet_role = 'driver'
            existing.can_drive = True
            existing.license_series = vu_series
            existing.license_number = vu_number
            existing.license_categories = str(categories) if categories else ''
            existing.license_issued_at = license_date
            existing.driver_tab_number = str(tab_num) if tab_num else ''
            existing.last_name = row_last
            existing.first_name = row_first
            existing.middle_name = row_middle
            existing.full_name = row_full
            updated += 1
        else:
            user = User(
                username=username,
                password_hash=hash_password('driver123'),  # дефолтный пароль — TODO сменить
                role='employee',
                fleet_role='driver',
                last_name=row_last,
                first_name=row_first,
                middle_name=row_middle,
                full_name=row_full,
                can_drive=True,
                license_series=vu_series,
                license_number=vu_number,
                license_categories=str(categories) if categories else '',
                license_issued_at=license_date,
                driver_tab_number=str(tab_num) if tab_num else '',
                is_email_confirmed=False,
            )
            db.add(user)
            created += 1

    wb.close()
    await db.commit()
    result = {'created': created, 'updated': updated, 'skipped': skipped}
    log.info(f"drivers_seed: {result}")
    return result


async def seed_fuel_norms_from_xlsx(
    db, xlsx_path: Optional[str] = None, *, force: bool = False
) -> dict:
    """
    Парсит лист 'Авто' и обновляет Vehicle.fuel_norm_summer/winter
    + Vehicle.props['fuel_norm_details'] (база, коэффициенты, формула).

    Структура листа:
      - Строки 1-5: заголовки/шапка, пропускаем
      - Строка 6: header ('№ п/п', ...)
      - Далее группы по 3 строки:
          row[i]   — главная: №, марка+номер или только марка, топливо, базовая норма,
                              коэф1, летняя норма, зимняя норма
          row[i+1] — гос.номер в col[1] если его нет в row[i]; коэф2; формула в col[5]
          row[i+2] — коэф3

    Guard (2026-09-02, симметрично vehicles_seed.py): в отличие от
    seed_vehicles_from_xlsx, эта функция не защищена ON CONFLICT — она UPDATE'ит
    уже существующие строки Vehicle безусловно (fuel_norm_summer/winter,
    props['fuel_norm_details']), перезаписывая ручные правки при каждом
    рестарте. Как только реестр `vehicles` наполнен (хотя бы одна запись —
    значит, реестр уже вручную курирован), сидер норм расхода больше не
    запускается, если явно не передан force=True.
    """
    import re

    if not force:
        from sqlalchemy import text as _text
        existing_count = (
            await db.execute(_text("SELECT COUNT(*) FROM vehicles"))
        ).scalar_one()
        if existing_count and existing_count > 0:
            log.info(
                f"fuel_norms_seed: реестр уже содержит {existing_count} записей, "
                f"обновление норм пропущено"
            )
            return {'updated': 0, 'reason': 'already_seeded', 'existing': existing_count}

    path = xlsx_path or WAYBILL_XLSX_PATH
    if not os.path.exists(path):
        log.warning(f"fuel_norms_seed: xlsx не найден: {path}")
        return {'reason': 'xlsx_not_found', 'path': path}

    try:
        import openpyxl
    except ImportError:
        return {'reason': 'openpyxl_not_installed'}

    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return {'reason': f'xlsx_open_error: {e}'}

    sheet_name = 'Авто'
    if sheet_name not in wb.sheetnames:
        wb.close()
        return {'reason': 'no_sheet', 'available': wb.sheetnames}

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    # Ищем строку-header по первой колонке == '№ п/п'
    start = 0
    for i, r in enumerate(rows):
        if r and r[0] == '№ п/п':
            start = i + 1
            break

    # Паттерн для гос.номера кириллицей: Б000ББ000 или Б000ББ00
    PLATE_RE = re.compile(r'[А-ЯЁA-Z]\s*\d{3}\s*[А-ЯЁA-Z]{2}\s*\d{2,3}', re.IGNORECASE)

    def extract_plate(text):
        if not text:
            return None
        m = PLATE_RE.search(str(text))
        if m:
            return re.sub(r'\s+', '', m.group(0))
        return None

    from sqlalchemy import select
    from app.models.vehicle import Vehicle
    from sqlalchemy.orm.attributes import flag_modified

    updated = 0
    i = start
    while i < len(rows):
        main = rows[i]
        if not main or main[0] is None:
            i += 1
            continue

        # Убеждаемся что первая колонка — число (порядковый номер)
        try:
            int(main[0])
        except (TypeError, ValueError):
            i += 1
            continue

        # Колонки главной строки: 0=№, 1=марка+номер, 2=топливо, 3=базовая, 4=коэф1, 5=летняя, 6=зимняя
        cols = (list(main) + [None] * 7)[:7]
        _, brand_plate, fuel_type, base_norm, coef1, summer, winter = cols

        # Берём строки i+1 и i+2 для доп.коэффициентов
        row2 = rows[i + 1] if i + 1 < len(rows) else [None] * 7
        row3 = rows[i + 2] if i + 2 < len(rows) else [None] * 7
        row2 = (list(row2) + [None] * 7)[:7]
        row3 = (list(row3) + [None] * 7)[:7]

        coef2 = row2[4]
        coef3 = row3[4]
        formula = row2[5]   # формула в строке 2, колонка «Итоговая летом»

        coefficients = [str(c) for c in [coef1, coef2, coef3] if c is not None]

        # Гос.номер: сначала ищем в brand_plate (строка 1), потом в row2[1]
        plate_match = extract_plate(brand_plate) or extract_plate(row2[1])

        if plate_match:
            res = await db.execute(
                select(Vehicle).where(Vehicle.plate.ilike(f"%{plate_match}%"))
            )
            veh = res.scalars().first()
            if veh:
                if summer is not None:
                    try:
                        veh.fuel_norm_summer = float(summer)
                    except (TypeError, ValueError):
                        pass
                if winter is not None:
                    try:
                        veh.fuel_norm_winter = float(winter)
                    except (TypeError, ValueError):
                        pass
                if not veh.props:
                    veh.props = {}
                veh.props['fuel_norm_details'] = {
                    'base_l_per_100km': float(base_norm) if base_norm is not None else None,
                    'fuel_type': str(fuel_type) if fuel_type else None,
                    'coefficients': coefficients,
                    'formula': str(formula) if formula else '',
                    'summer_l_per_100km': float(summer) if summer is not None else None,
                    'winter_l_per_100km': float(winter) if winter is not None else None,
                }
                flag_modified(veh, 'props')
                updated += 1

        i += 3  # следующая группа из 3 строк

    await db.commit()
    result = {'updated': updated}
    log.info(f"fuel_norms_seed: {result}")
    return result


async def seed_test_fines(db) -> dict:
    """
    Создаёт ~12 тестовых штрафов на разных водителей/машин для демо /fleet/fines.
    Идемпотентный — пропускает если штрафы уже есть.
    """
    from sqlalchemy import select, func
    from app.models.vehicle_fine import VehicleFine
    from app.models.vehicle import Vehicle
    from app.models.user import User
    from datetime import datetime, timedelta
    import random

    count = (await db.execute(select(func.count(VehicleFine.id)))).scalar()
    if count and count > 0:
        return {'reason': 'already_seeded', 'existing': count}

    veh_res = await db.execute(
        select(Vehicle).where(Vehicle.state == 'working').limit(15)
    )
    vehicles = veh_res.scalars().all()

    drv_res = await db.execute(
        select(User).where(User.fleet_role == 'driver')
    )
    drivers = drv_res.scalars().all()

    if not vehicles or not drivers:
        return {'reason': 'no_vehicles_or_drivers', 'vehicles': len(vehicles), 'drivers': len(drivers)}

    violations = [
        ('Превышение скорости 20-40 км/ч', 500),
        ('Превышение скорости 40-60 км/ч', 1000),
        ('Превышение скорости 60-80 км/ч', 2500),
        ('Проезд на запрещающий сигнал', 1000),
        ('Парковка в неположенном месте', 1500),
        ('Не пропустил пешехода', 1500),
        ('Разговор по мобильному телефону', 1500),
        ('Не пристёгнут ремень безопасности', 1000),
        ('Нарушение дорожной разметки', 500),
        ('Стоянка под запрещающим знаком', 1500),
        ('Тонировка стёкол', 500),
        ('Эксплуатация незарегистрированного ТС', 800),
    ]

    now = datetime.now()
    random.seed(42)  # воспроизводимо
    fines = []
    for _ in range(12):
        veh = random.choice(vehicles)
        drv = random.choice(drivers)
        violation, amount = random.choice(violations)
        issued_at = now - timedelta(days=random.randint(1, 180))
        status = random.choice(['unpaid', 'unpaid', 'paid'])  # 2/3 unpaid
        fines.append(VehicleFine(
            vehicle_id=veh.id,
            driver_user_id=drv.id,
            violation_type=violation,
            amount=amount,
            issued_at=issued_at,
            location='Москва, ул. Тестовая',
            status=status,
            paid_at=(issued_at + timedelta(days=random.randint(1, 30))) if status == 'paid' else None,
            doc_number=f'18810177{random.randint(100000, 999999)}',
        ))

    db.add_all(fines)
    await db.commit()
    result = {'created': len(fines)}
    log.info(f"test_fines_seed: {result}")
    return result
