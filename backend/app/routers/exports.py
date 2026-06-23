import io
import os
import zipfile
from urllib.parse import quote
from fastapi import APIRouter, Body, Depends, Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from app.auth.jwt import get_current_user
from app.database import get_db
from app.models.user import User
from app.services import report_excel, report_pdf
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

router = APIRouter(prefix='/api/exports', tags=['exports'])


def _content_disposition(filename: str) -> str:
    """RFC 5987 — кириллица в имени файла недопустима в latin-1 заголовке."""
    ascii_fallback = filename.encode('ascii', 'ignore').decode('ascii').strip() or 'export'
    return f"attachment; filename=\"{ascii_fallback}\"; filename*=UTF-8''{quote(filename)}"


@router.post('/table.xlsx')
async def export_table_xlsx(
    body: dict = Body(...),
    user: User = Depends(get_current_user),
):
    title = body.get('title') or 'Реестр'
    columns = body.get('columns') or []
    rows = body.get('rows') or []
    meta = {
        'author': getattr(user, 'full_name', None) or getattr(user, 'username', None) or str(user.id),
        'generated_at': datetime.utcnow().strftime('%d.%m.%Y %H:%M'),
    }
    data = report_excel.export_table(title, columns, rows, meta)
    raw_name = body.get('filename') or title
    filename = f'{raw_name}_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.xlsx'
    return Response(
        content=data,
        media_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': _content_disposition(filename)},
    )


@router.post('/tasks.zip')
async def export_tasks_zip(
    body: dict = Body(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Excel + вложения в ZIP. kind='tasks'|'purchases'. Файлы диалогов не включаются."""
    from app.models.purchase_file import PurchaseFile

    title = body.get('title') or 'Реестр'
    columns = body.get('columns') or []
    rows = body.get('rows') or []
    kind = body.get('kind') or 'tasks'
    meta = {
        'author': getattr(user, 'full_name', None) or getattr(user, 'username', None) or str(user.id),
        'generated_at': datetime.utcnow().strftime('%d.%m.%Y %H:%M'),
    }

    # Build base xlsx
    xlsx_bytes = report_excel.export_table(title, columns, rows, meta)

    # Для закупок — подгружаем файлы из БД по id строки
    # id строки берём из rows[].id (для задач — tasks не имеют файлов в БД, пропускаем)
    files_map: dict[int, list] = {}  # row_id -> [PurchaseFile]
    if kind == 'purchases':
        purchase_ids = [r.get('id') for r in rows if r.get('id')]
        if purchase_ids:
            result = await db.execute(
                select(PurchaseFile).where(
                    PurchaseFile.purchase_id.in_(purchase_ids),
                    PurchaseFile.is_active == True,
                    # Исключаем acceptance_doc (чеки/акты из диалогов)
                    PurchaseFile.file_type != 'acceptance_doc',
                )
            )
            for pf in result.scalars().all():
                files_map.setdefault(pf.purchase_id, []).append(pf)

    # Patch xlsx: добавить столбец «Файлы» с гиперссылками
    wb = load_workbook(io.BytesIO(xlsx_bytes))
    ws = wb.active
    has_files = bool(files_map)
    if has_files:
        # Найти индекс столбца id в columns (для сопоставления строк)
        id_col_idx = next((i for i, c in enumerate(columns) if c.get('key') == 'id'), None)
        files_col_idx = len(columns) + 1  # новый столбец после существующих
        files_col_letter = get_column_letter(files_col_idx)
        # Заголовок
        hcell = ws.cell(row=3, column=files_col_idx, value='Файлы')
        hcell.font = Font(color='FFFFFF', bold=True, size=11)
        from openpyxl.styles import PatternFill, Alignment
        hcell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        hcell.alignment = Alignment(horizontal='center', vertical='center')
        ws.column_dimensions[files_col_letter].width = 40

        for r_idx, row in enumerate(rows, start=4):
            row_id = row.get('id')
            pf_list = files_map.get(row_id, [])
            if pf_list:
                # Записываем первое имя, гиперссылку на папку
                first_name = pf_list[0].filename
                zip_path = f'files/{row_id}/{first_name}'
                cell = ws.cell(row=r_idx, column=files_col_idx, value=first_name)
                cell.hyperlink = zip_path
                cell.font = Font(color='0563C1', underline='single')
                if len(pf_list) > 1:
                    cell.value = f'{first_name} (+{len(pf_list)-1})'

    xlsx_out = io.BytesIO()
    wb.save(xlsx_out)
    xlsx_bytes = xlsx_out.getvalue()

    # Build ZIP
    zip_buf = io.BytesIO()
    with zipfile.ZipFile(zip_buf, 'w', zipfile.ZIP_DEFLATED) as zf:
        zf.writestr('report.xlsx', xlsx_bytes)
        for row_id, pf_list in files_map.items():
            for pf in pf_list:
                if pf.filepath and os.path.exists(pf.filepath):
                    arcname = f'files/{row_id}/{pf.filename}'
                    zf.write(pf.filepath, arcname)

    raw_name = body.get('filename') or title
    filename = f'{raw_name}_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.zip'
    return Response(
        content=zip_buf.getvalue(),
        media_type='application/zip',
        headers={'Content-Disposition': _content_disposition(filename)},
    )


@router.post('/kanban.pdf')
async def export_kanban_pdf(
    body: dict = Body(...),
    user: User = Depends(get_current_user),
):
    title = body.get('title') or 'Канбан задач'
    columns = body.get('columns') or []
    meta = {
        'author': getattr(user, 'full_name', None) or getattr(user, 'username', None) or str(user.id),
        'generated_at': datetime.utcnow().strftime('%d.%m.%Y %H:%M'),
    }
    data = report_pdf.export_kanban(title, columns, meta)
    raw_name = body.get('filename') or title
    filename = f'{raw_name}_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.pdf'
    return Response(
        content=data,
        media_type='application/pdf',
        headers={'Content-Disposition': _content_disposition(filename)},
    )


@router.post('/table.pdf')
async def export_table_pdf(
    body: dict = Body(...),
    user: User = Depends(get_current_user),
):
    title = body.get('title') or 'Реестр'
    columns = body.get('columns') or []
    rows = body.get('rows') or []
    meta = {
        'author': getattr(user, 'full_name', None) or getattr(user, 'username', None) or str(user.id),
        'generated_at': datetime.utcnow().strftime('%d.%m.%Y %H:%M'),
    }
    data = report_pdf.export_table(title, columns, rows, meta)
    raw_name = body.get('filename') or title
    filename = f'{raw_name}_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.pdf'
    return Response(
        content=data,
        media_type='application/pdf',
        headers={'Content-Disposition': _content_disposition(filename)},
    )
