"""
Генератор Excel-шаблона для импорта реестра транспорта — GET /api/vehicles/import-template.

Один .xlsx с тремя листами:
  1. «Транспорт»    — рабочий лист: строка 1 = заголовки (подписи из реестра
                       services/vehicle_fields.py), данные пользователь пишет со строки 2.
  2. «Инструкция»   — как заполнять, что обязательно, формат дат, что делает система
                       с уже существующими гос. номерами.
  3. «Справочники»  — по колонке на словарь (полные наборы значений из кода), видимый лист.

Подход к DataValidation/defined names скопирован с уже проверенного
app/routers/purchase_export.py (функция download_import_template):
  - значения справочников пишутся на лист «Справочники» начиная со строки 3;
  - диапазон валидации = 2× текущего числа значений (запас на дописывание своих);
  - formula1 ссылается ТОЛЬКО на defined name уровня книги (не на прямой
    cross-sheet диапазон — Excel иначе молча теряет список);
  - showErrorMessage=False / allow_blank=True — список подсказывает, но не блокирует
    ручной ввод.

КРИТИЧНО: каждый заголовок, который здесь выводится в лист «Транспорт», обязан
распознаваться app.routers.vehicles_import._resolve_header_columns. Алиасы под
точный текст подписей реестра добавлены в _COL_MAP (см. vehicles_import.py) —
только добавлены, ни одна старая запись не переписана.

Единственное сознательное исключение из «все поля реестра, кроме computed»:
поле assigned_org_id («Организация-эксплуатант») в шаблон не включается — его
значение при импорте вычисляется на сервере из текста колонки assigned_text
(«У кого в эксплуатации (текст)») через сопоставление с существующими
организациями (commit_import: assigned_org_id = match(assigned_text)). Отдельная
колонка для него создала бы второй заголовок, резолвящийся в тот же ключ
assigned_text, и была бы автоматически отброшена occupied-tracking'ом в
_resolve_header_columns — то есть попала бы в список "нераспознанных".
"""
from __future__ import annotations

from io import BytesIO
from typing import Any, Dict, List, Optional, Set, Tuple

from app.services.vehicle_enum_labels import (
    FUEL_TYPE_LABELS,
    PTS_KIND_LABELS,
    STATE_LABELS,
    TYPE_LABELS,
    as_dd_list,
)
from app.services.vehicle_fields import FIELD_GROUPS
from app.services.vehicle_sheet_dictionaries import (
    FIELD_OPTIONS as _DICT_OPTIONS,
    NO_DATA_LABEL as _NO_DATA_LABEL,
    PASS_STATUS_OPTIONS as _PASS_STATUS_OPTIONS,
)

try:
    from openpyxl import Workbook
    from openpyxl.comments import Comment
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.workbook.defined_name import DefinedName
    from openpyxl.worksheet.datavalidation import DataValidation
except ImportError:  # pragma: no cover — openpyxl обязателен, но не роняем импорт модуля
    Workbook = None
    Comment = None
    DataValidation = None
    DefinedName = None
    get_column_letter = None

# Поле, чьё значение вычисляется на сервере из assigned_text — см. docstring выше.
_EXCLUDED_DERIVED_KEYS: Set[str] = {"assigned_org_id"}

# Диапазон строк данных, на который распространяются DataValidation/числовые форматы.
_DATA_ROWS_END = 1000
# До какой строки реально проставлять number_format ячеек (баланс размера файла/скорости).
_FORMAT_ROWS_END = 300

_DATE_FMT = "DD.MM.YYYY"
_INT_FMT = "0"
_FLOAT_FMT = "0.00"

# ─────────────────────────── Справочники для выпадающих списков ─────────────────
#
# Автоблок (актуализация 2026-08-31): наборы для полей ниже — НЕ придуманы
# здесь, а берутся из app/services/vehicle_sheet_dictionaries.py — единый
# источник правды, извлечённый из правил проверки данных (x14:dataValidation,
# лист «drop») реального листа владельца. Эти поля — единственные, для
# которых DataValidation в этом файле делается БЛОКИРУЮЩИМ (showErrorMessage=
# True, см. _STRICT_DD_KEYS ниже и место построения DataValidation) — так
# потребовал владелец: только варианты из списка, никакой самодеятельности.

def _dd(options: List[str]) -> List[Tuple[str, Optional[str]]]:
    return [(v, None) for v in options]


_DD_BOOL_YESNO: List[Tuple[str, Optional[str]]] = _dd(_DICT_OPTIONS["has_radio"])  # ["Да", "Нет"]

_DD_PTS_CATEGORY: List[Tuple[str, Optional[str]]] = [
    (v, None) for v in ["A", "B", "BE", "C", "CE", "D", "DE", "M", "Tb", "Tm", _NO_DATA_LABEL]
]

# field_key → (defined_name, заголовок колонки на листе «Справочники», список (label, code))
_FIELD_DD_MAP: Dict[str, Tuple[str, str, List[Tuple[str, Optional[str]]]]] = {
    # sort_alpha=True — владелец (2026-09): «Тип ТС» шёл вразнобой, отсортировать
    # по алфавиту (см. as_dd_list в vehicle_enum_labels.py). state/fuel_type/
    # pts_kind ниже сортировку сознательно НЕ получают — их порядок осмысленный
    # (см. докстринг as_dd_list).
    "type": ("dd_vehicle_type", "Тип ТС", as_dd_list(TYPE_LABELS, sort_alpha=True)),
    "state": ("dd_vehicle_state", "Состояние ТС", as_dd_list(STATE_LABELS)),
    "fuel_type": ("dd_fuel_type", "Вид топлива", as_dd_list(FUEL_TYPE_LABELS)),
    "pts_kind": ("dd_pts_kind", "Вид ПТС", as_dd_list(PTS_KIND_LABELS)),
    "pts_category": ("dd_pts_category", "Категория ТС по ПТС", _DD_PTS_CATEGORY),
    # ── Справочники из правил проверки данных листа владельца (блокирующие) ──
    "paint_condition": ("dd_paint_condition", "Состояние ЛКП", _dd(_DICT_OPTIONS["paint_condition"])),
    "tech_inspection_status": ("dd_tech_inspection_status", "Обязательный техосмотр", _dd(_DICT_OPTIONS["tech_inspection_status"])),
    "tires_type": ("dd_tires_type", "Авторезина", _dd(_DICT_OPTIONS["tires_type"])),
    "tires_condition": ("dd_tires_condition", "Состояние резины", _dd(_DICT_OPTIONS["tires_condition"])),
    "body_type": ("dd_body_type", "Кузов", _dd(_DICT_OPTIONS["body_type"])),
    "tires_summer_condition": ("dd_tires_condition", "Состояние резины", _dd(_DICT_OPTIONS["tires_condition"])),
    "tires_winter_condition": ("dd_tires_condition", "Состояние резины", _dd(_DICT_OPTIONS["tires_condition"])),
    # has_branding сюда НЕ входит — это type="bool" поле, оно уже получает
    # список Да/Нет через общий bool-фолбэк ниже (has_bool_field), как и
    # has_radio/has_mirrors/... (та же логика, что и у них).
    # pass_zo/pass_ho/pass_dnr/pass_lnr/pass_moscow сюда больше НЕ входят —
    # 2026-09: пропуска ушли из FIELD_GROUPS в отдельную таблицу vehicle_passes,
    # для них своя генерация колонок ниже ("Пропуск: <Имя>" / "... — до"),
    # см. _PASS_DD_NAME / блок после основного цикла по fields.
}
_BOOL_DD_NAME = "dd_bool_yesno"
_BOOL_DD_HEADER = "Да / Нет"
_PASS_DD_NAME = "dd_pass_status"
_PASS_DD_HEADER = "Пропуск (статус)"
# Базовый набор названий пропусков для шаблона (2026-09) — организации, которым
# нужны другие зоны, дописывают свои колонки прямо в файле в формате
# "Пропуск: <Название>" / "Пропуск: <Название> — до" (см. _resolve_header_columns
# в app/routers/vehicles_import.py — распознаёт ЛЮБОЕ название по этому шаблону
# заголовка, не только эти пять).
_PASS_TEMPLATE_NAMES: List[str] = ["ЗО", "ХО", "ДНР", "ЛНР", "Москва"]

# Ключи полей, чьи допустимые значения взяты из правил проверки данных листа
# владельца (app.services.vehicle_sheet_dictionaries.FIELD_OPTIONS) — для них
# DataValidation в шаблоне блокирующий (showErrorMessage=True): владелец явно
# потребовал "только варианты ответов из правил проверки". Для всех остальных
# полей со списком (type/state/fuel_type/pts_kind/pts_category и т.п.) список —
# по-прежнему подсказка, свой текст допускается (showErrorMessage=False).
_STRICT_DD_KEYS: Set[str] = set(_DICT_OPTIONS.keys())

# ─────────────────────────── Примеры значений для примечаний ────────────────────

_FIELD_EXAMPLES: Dict[str, str] = {
    "plate": "А123ВС777",
    "brand": "ГАЗ",
    "model": "Соболь",
    "year_of_manufacture": "2019",
    "color": "Белый",
    "vin": "XTA210930K1234567",
    "type": "Легковой",
    "body_type": "Седан",
    "pts_category": "B",
    "engine_power_hp": "150",
    "engine_volume_l": "2.0",
    "fuel_type": "АИ-95",
    "fuel_norm_summer": "9.5",
    "fuel_norm_winter": "11",
    "owner_org_id": "ООО «Ромашка»",
    "ownership_basis": "Договор купли-продажи",
    "ownership_doc_number": "№ 45",
    "ownership_doc_date": "12.01.2022",
    "owner_since": "12.01.2022",
    "purchase_info": "Субсидия Минтруда 2022",
    "assigned_text": "ГКУ Соцзащита",
    "assignment_basis": "Договор безвозмездного пользования",
    "assignment_doc_number": "№ 7",
    "assignment_doc_date": "01.03.2023",
    "location_city": "Москва",
    "location_address": "ул. Ленина, д. 5",
    "home_base_city": "Иркутск",
    "responsible_name": "Иванов Иван Иванович",
    "pts_number": "77ТТ123456",
    "pts_kind": "Бумажный",
    "sts_number": "9911 123456",
    "sts_issued_at": "05.05.2020",
    "registered_at": "10.05.2020",
    "insurance_company": "СОГАЗ",
    "insurance_policy_number": "ХХХ 0123456789",
    "insurance_until": "31.12.2026",
    "current_odometer_km": "45000",
    "last_to_date": "01.06.2026",
    "last_to_mileage_km": "42000",
    "next_to_km": "50000",
    "tech_inspection_status": "Да",
    "tech_inspection_last_date": "01.02.2026",
    "tech_inspection_until": "01.02.2027",
    "pass_zo": "Да",
    "pass_zo_until": "31.12.2026",
    "pass_ho": "Да",
    "pass_ho_until": "31.12.2026",
    "pass_dnr": "Да",
    "pass_dnr_until": "31.12.2026",
    "pass_lnr": "Да",
    "pass_lnr_until": "31.12.2026",
    "pass_moscow": "Да",
    "pass_moscow_until": "31.12.2026",
    "tires_type": "Летняя",
    "has_spare_tires": "Да",
    "tires_condition": "Хорошая",
    "tires_summer_radius": "R15",
    "tires_summer_profile": "195/65",
    "tires_summer_condition": "Хорошая",
    "tires_winter_radius": "R15",
    "tires_winter_profile": "195/65",
    "tires_winter_condition": "Хорошая",
    "has_branding": "Да",
    "has_radio": "Да",
    "has_mirrors": "Да",
    "mirrors_ok": "Да",
    "akb_ok": "Да",
    "branding": "Логотип на бортах",
    "has_keys": "Да",
    "has_first_aid_kit": "Да",
    "first_aid_kit_until": "01.09.2027",
    "has_spare_wheel": "Да",
    "has_extinguisher": "Да",
    "extinguisher_check_date": "01.03.2026",
    "has_tracker": "Да",
    "tracker_paid_until": "31.12.2026",
    "has_tachograph": "Нет",
    "tachograph_check_date": "01.03.2026",
    "state": "Рабочее",
    "paint_condition": "Идеальное",
    "repair_required": "Нет",
    "defect_description": "Скол лобового стекла",
    "tech_condition_info": "Требуется замена тормозных колодок",
    "note": "Передан временно в другое подразделение",
}

# Точечные уточнения для полей, где типового текста по `type` недостаточно.
_FIELD_NOTE_OVERRIDE: Dict[str, str] = {
    "owner_org_id": (
        "Впишите точное название организации-собственника, как оно указано в системе. "
        "Если для этой машины оно неизвестно — напишите «Нет данных»: на шаге импорта "
        "будет предложено выбрать организацию по умолчанию для всех таких строк."
    ),
    "assigned_text": (
        "Организация, которая фактически эксплуатирует ТС (может отличаться от собственника). "
        "Если название не совпадёт ни с одной организацией в системе — на шаге сопоставления "
        "в диалоге импорта можно будет указать организацию вручную или оставить как текст."
    ),
    "plate": (
        "Уникальный идентификатор строки. Если такой гос. номер уже есть в системе — "
        "при импорте можно выбрать «обновить существующую запись» или «пропустить»."
    ),
}


def _field_type_note(ftype: str) -> str:
    if ftype == "date":
        return "Формат даты: ДД.ММ.ГГГГ."
    if ftype == "int":
        return "Целое число."
    if ftype == "float":
        return "Число, дробная часть — через точку."
    if ftype == "bool":
        return "Да или Нет."
    if ftype == "text":
        return "Произвольный текст, можно многострочный."
    return "Текст."


def _build_comment_text(
    field: Dict[str, Any],
    dd_list: Optional[List[Tuple[str, Optional[str]]]],
    strict: bool = False,
) -> str:
    key = field["key"]
    ftype = field["type"]
    parts: List[str] = []

    override = _FIELD_NOTE_OVERRIDE.get(key)
    if override:
        parts.append(override)

    if dd_list:
        sample = ", ".join(lbl for lbl, _ in dd_list[:4])
        more = "…" if len(dd_list) > 4 else ""
        if strict:
            parts.append(
                f"Выберите значение СТРОГО из списка ({sample}{more}) — полный перечень на "
                f"листе «Справочники». Любое другое значение Excel отклонит, а при импорте "
                f"файла со сторонним текстом это поле не будет заполнено."
            )
        else:
            parts.append(
                f"Выберите значение из выпадающего списка ({sample}{more}) или впишите своё — "
                f"полный перечень на листе «Справочники»."
            )
    elif not override:
        parts.append(_field_type_note(ftype))
    elif ftype == "date":
        # override уже что-то сказал по смыслу поля — формат даты всё равно называем явно
        parts.append(_field_type_note(ftype))

    example = _FIELD_EXAMPLES.get(key)
    if example:
        parts.append(f"Например: {example}.")

    if field.get("required"):
        parts.append("Обязательное поле — без него строка не будет импортирована.")
        if dd_list and any(v == _NO_DATA_LABEL for v, _ in dd_list):
            # «Состояние» — required=True, но его список (см. as_dd_list) тоже
            # включает «Нет данных» как осознанный допустимый ответ: required
            # не означает «нельзя выбрать «Нет данных»», это означает «нельзя
            # оставить ячейку пустой без явного ответа». Без этой строки
            # получилось бы противоречие: сам список содержит пункт, о
            # котором примечание не упоминает.
            parts.append(
                f"Если конкретное значение неизвестно — выберите «{_NO_DATA_LABEL}», "
                f"этот пункт тоже входит в список и принимается системой."
            )
    elif dd_list:
        # Список для этого поля (в т.ч. bool Да/Нет) всегда включает пункт
        # «Нет данных» (см. as_dd_list / _DICT_OPTIONS) — владелец прямо
        # запретил разрешать пустоту в примечаниях (2026-09): пустых ячеек в
        # файле быть не должно, при отсутствии данных выбирается этот пункт.
        parts.append(f"Если данных нет — выберите «{_NO_DATA_LABEL}».")
    else:
        parts.append(f"Если данных нет — напишите «{_NO_DATA_LABEL}».")

    return " ".join(parts)


def get_template_fields(hidden_keys: Set[str]) -> List[Dict[str, Any]]:
    """Плоский список полей реестра для шаблона: без computed, без скрытых для
    организации, без осознанно исключённых производных полей (см. docstring модуля)."""
    fields: List[Dict[str, Any]] = []
    for group in FIELD_GROUPS:
        for f in group["fields"]:
            if f["storage"] == "computed":
                continue
            if f["key"] in _EXCLUDED_DERIVED_KEYS:
                continue
            if f["key"] in hidden_keys:
                continue
            fields.append(f)
    return fields


def build_vehicle_import_template(hidden_keys: Set[str]) -> BytesIO:
    """Строит .xlsx шаблон импорта ТС и возвращает готовый буфер (позиция — 0)."""
    if Workbook is None:
        raise RuntimeError("openpyxl не установлен на сервере")

    fields = get_template_fields(hidden_keys)

    wb = Workbook()

    # =========================================================================
    # Лист «Транспорт» (создаётся первым — активный лист книги по умолчанию)
    # =========================================================================
    ws = wb.active
    ws.title = "Транспорт"

    # =========================================================================
    # Лист «Справочники» — строим ДО DataValidation, чтобы defined name уже
    # существовали к моменту, когда мы вешаем списки на «Транспорт».
    # =========================================================================
    ws_ref = wb.create_sheet(title="Справочники")

    ref_hint_font = Font(italic=True, color="6B7280", size=9)
    ref_hdr_font = Font(bold=True, size=10)
    ref_hdr_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    hint_cell = ws_ref.cell(1, 1, (
        "Допишите свои значения в пустые ячейки под списком — они появятся в "
        "выпадающих списках на листе «Транспорт»."
    ))
    hint_cell.font = ref_hint_font
    hint_cell.alignment = Alignment(wrap_text=True)
    ws_ref.row_dimensions[1].height = 28

    # Полный перечень словарей: все из _FIELD_DD_MAP (но только для полей, реально
    # присутствующих в шаблоне после фильтрации hidden_keys) + общий Да/Нет.
    # Дедуп по dn_name: пять полей "Пропуск X" делят один defined_name/одну
    # колонку на листе «Справочники» (см. _FIELD_DD_MAP) — без дедупа сюда
    # попали бы 5 идентичных колонок.
    present_keys = {f["key"] for f in fields}
    dd_columns: List[Tuple[str, str, List[Tuple[str, Optional[str]]]]] = []
    _seen_dn_names: Set[str] = set()
    for key, cfg in _FIELD_DD_MAP.items():
        if key not in present_keys or cfg[0] in _seen_dn_names:
            continue
        _seen_dn_names.add(cfg[0])
        dd_columns.append(cfg)
    has_bool_field = any(f["type"] == "bool" for f in fields)
    if has_bool_field:
        dd_columns.append((_BOOL_DD_NAME, _BOOL_DD_HEADER, _DD_BOOL_YESNO))
    # 2026-09: справочник статусов пропусков — не завязан ни на одно поле реестра
    # (пропуска ушли в отдельную таблицу vehicle_passes), поэтому добавляется
    # безусловно, а не через _FIELD_DD_MAP/present_keys — используется колонками
    # "Пропуск: <Имя>", которые дописываются на лист «Транспорт» ниже.
    dd_columns.append((_PASS_DD_NAME, _PASS_DD_HEADER, _dd(_PASS_STATUS_OPTIONS)))

    # dn_name → имя defined name (совпадает с dn_name, оставлено для читаемости DV-кода)
    dn_ranges: Dict[str, str] = {}

    for col_i, (dn_name, hdr_label, dd_list) in enumerate(dd_columns, 1):
        hdr_cell = ws_ref.cell(2, col_i, hdr_label)
        hdr_cell.font = ref_hdr_font
        hdr_cell.fill = ref_hdr_fill
        hdr_cell.alignment = Alignment(horizontal="center")

        n_values = len(dd_list)
        slot_count = n_values * 2  # запас 2× — пользователь может дописать своё

        for row_j, (label, _code) in enumerate(dd_list, 1):
            ws_ref.cell(2 + row_j, col_i, label)

        col_letter = get_column_letter(col_i)
        first_row = 3
        last_row = 2 + slot_count

        if DefinedName is not None:
            attr_text = f"'Справочники'!${col_letter}${first_row}:${col_letter}${last_row}"
            wb.defined_names[dn_name] = DefinedName(dn_name, attr_text=attr_text)
        dn_ranges[dn_name] = dn_name

        max_len = max((len(lbl) for lbl, _ in dd_list), default=10)
        ws_ref.column_dimensions[col_letter].width = max(max_len + 4, len(hdr_label) + 2, 14)

    ws_ref.freeze_panes = "A3"

    # =========================================================================
    # Лист «Инструкция» — вставляем ВТОРЫМ (между «Транспорт» и «Справочники»)
    # =========================================================================
    ws_instr = wb.create_sheet(title="Инструкция", index=1)
    _fill_instruction_sheet(ws_instr)

    # =========================================================================
    # Заполняем лист «Транспорт»: заголовки, форматы, примечания, DataValidation
    # =========================================================================
    headers = [f["label"] for f in fields]
    ws.append(headers)

    font_hdr = Font(color="FFFFFF", bold=True, size=11)
    fill_req = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    fill_opt = PatternFill(start_color="1976D2", end_color="1976D2", fill_type="solid")
    align_hdr = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for i, f in enumerate(fields, 1):
        c = ws.cell(1, i)
        c.fill = fill_req if f.get("required") else fill_opt
        c.font = font_hdr
        c.alignment = align_hdr

        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = max(14, min(len(f["label"]) + 4, 34))

        # ── Примечание к заголовку ──
        is_strict = f["key"] in _STRICT_DD_KEYS
        if Comment is not None:
            dd_cfg = _FIELD_DD_MAP.get(f["key"])
            dd_list = dd_cfg[2] if dd_cfg else (_DD_BOOL_YESNO if f["type"] == "bool" else None)
            comment_text = _build_comment_text(f, dd_list, strict=is_strict)
            c.comment = Comment(comment_text, "GALA")
            c.comment.width = 260
            c.comment.height = 120

        # ── Числовой/датовый формат ячеек данных ──
        if f["type"] in ("date", "int", "float"):
            fmt = _DATE_FMT if f["type"] == "date" else (_INT_FMT if f["type"] == "int" else _FLOAT_FMT)
            for r in range(2, _FORMAT_ROWS_END + 1):
                ws.cell(r, i).number_format = fmt
            ws.column_dimensions[col_letter].number_format = fmt

        # ── DataValidation на диапазон данных ──
        if DataValidation is not None:
            dd_cfg = _FIELD_DD_MAP.get(f["key"])
            ref_name: Optional[str] = None
            if dd_cfg is not None:
                ref_name = dd_cfg[0]
            elif f["type"] == "bool" and has_bool_field:
                ref_name = _BOOL_DD_NAME

            if ref_name:
                dv_kwargs: Dict[str, Any] = dict(
                    type="list",
                    formula1=f"={ref_name}",
                    allow_blank=True,
                    showErrorMessage=is_strict,
                    showInputMessage=False,
                )
                if is_strict:
                    # Владелец потребовал: для этих колонок — ТОЛЬКО варианты
                    # из правил проверки данных его листа, свободный ввод
                    # блокируется Excel'ем на месте (см. _STRICT_DD_KEYS).
                    dv_kwargs["errorTitle"] = "Недопустимое значение"
                    dv_kwargs["error"] = (
                        f"Допускаются только значения из списка «{f['label']}» "
                        f"(см. лист «Справочники»)."
                    )
                dv = DataValidation(**dv_kwargs)
                dv.sqref = f"{col_letter}2:{col_letter}{_DATA_ROWS_END}"
                ws.add_data_validation(dv)

    # =========================================================================
    # Пропуска (2026-09): "Пропуск: <Имя>" / "Пропуск: <Имя> — до" — по два
    # столбца на базовое название (см. _PASS_TEMPLATE_NAMES). Не входят в
    # FIELD_GROUPS (данные уходят в отдельную таблицу vehicle_passes, не в
    # колонку Vehicle), поэтому дописываются отдельным блоком после основного
    # цикла по полям реестра. Организациям, которым нужны другие названия,
    # достаточно дописать свою пару колонок в этом же формате — импорт
    # (app/routers/vehicles_import.py, _PASS_STATUS_HDR_RE/_PASS_UNTIL_HDR_RE)
    # распознаёт ЛЮБОЕ название по шаблону заголовка, не только эти пять.
    pass_col_start = len(fields) + 1
    pass_col_i = pass_col_start
    for pass_name in _PASS_TEMPLATE_NAMES:
        for suffix, is_date in ((f"Пропуск: {pass_name}", False), (f"Пропуск: {pass_name} — до", True)):
            c = ws.cell(1, pass_col_i, suffix)
            c.fill = fill_opt
            c.font = font_hdr
            c.alignment = align_hdr
            col_letter = get_column_letter(pass_col_i)
            ws.column_dimensions[col_letter].width = max(14, min(len(suffix) + 4, 34))

            if is_date:
                for r in range(2, _FORMAT_ROWS_END + 1):
                    ws.cell(r, pass_col_i).number_format = _DATE_FMT
                if Comment is not None:
                    c.comment = Comment(
                        f"Дата истечения пропуска «{pass_name}». Формат даты: ДД.ММ.ГГГГ. "
                        f"Если данных нет — оставьте пустой ячейкой либо впишите «{_NO_DATA_LABEL}».",
                        "GALA",
                    )
                    c.comment.width = 260
                    c.comment.height = 100
            else:
                if Comment is not None:
                    sample = ", ".join(_PASS_STATUS_OPTIONS[:4])
                    c.comment = Comment(
                        f"Статус пропуска «{pass_name}». Выберите значение СТРОГО из списка "
                        f"({sample}) — полный перечень на листе «Справочники». Если данных нет "
                        f"— выберите «{_NO_DATA_LABEL}».",
                        "GALA",
                    )
                    c.comment.width = 260
                    c.comment.height = 110
                if DataValidation is not None:
                    dv = DataValidation(
                        type="list",
                        formula1=f"={_PASS_DD_NAME}",
                        allow_blank=True,
                        showErrorMessage=True,
                        showInputMessage=False,
                        errorTitle="Недопустимое значение",
                        error=f"Допускаются только значения из списка «Пропуск (статус)» (см. лист «Справочники»).",
                    )
                    dv.sqref = f"{col_letter}2:{col_letter}{_DATA_ROWS_END}"
                    ws.add_data_validation(dv)
            pass_col_i += 1

    ws.row_dimensions[1].height = 40
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _fill_instruction_sheet(ws) -> None:
    """Текстовые пояснения — что обязательно, форматы, поведение при дублях."""
    title_font = Font(bold=True, size=14, color="1E3A5F")
    section_font = Font(bold=True, size=12, color="1976D2")
    body_align = Alignment(wrap_text=True, vertical="top")

    ws.column_dimensions["A"].width = 110

    rows: List[Tuple[str, Optional[Font]]] = [
        ("Инструкция по заполнению шаблона реестра транспорта", title_font),
        ("", None),
        ("1. Что обязательно заполнить", section_font),
        (
            "Гос. рег. знак (колонка «Гос. рег. знак» на листе «Транспорт») — без него строка "
            "не будет импортирована, вся остальная информация в такой строке проигнорируется.",
            None,
        ),
        (
            "Организация-собственник обязательна для новой записи. Если для конкретной машины "
            "она неизвестна — впишите «Нет данных»: при импорте будет предложено выбрать "
            "организацию по умолчанию для всех таких строк.",
            None,
        ),
        ("", None),
        ("2. Формат дат", section_font),
        (
            "Все даты пишите в формате ДД.ММ.ГГГГ, например 15.03.2026. Ячейки колонок с датами "
            "уже отформатированы под этот вид — Excel сам подскажет, если ввести дату иначе.",
            None,
        ),
        ("", None),
        ("3. Если данных нет", section_font),
        (
            "Пустых ячеек в файле быть не должно (кроме гос. номера — без него, см. пункт 1, "
            "строка не импортируется вовсе). Если по конкретной машине для какого-то поля данных "
            "нет — впишите «Нет данных»: для колонок со списком выберите этот пункт из "
            "выпадающего списка, для остальных — впишите текстом. Система распознает «Нет "
            "данных» как «поле не заполнено»: значение не будет записано и предупреждение об "
            "этом не поднимется, дозаполнить такие поля можно будет позже прямо в системе.",
            None,
        ),
        ("", None),
        ("4. Выпадающие списки", section_font),
        (
            "В колонках с ограниченным набором значений при клике по ячейке появляется список "
            "допустимых значений — полный перечень смотрите на листе «Справочники». Для "
            "большинства таких колонок (тип ТС, состояние, топливо, вид ПТС и т.п.) список — "
            "подсказка: можно вписать своё значение, оно тоже будет принято при импорте. "
            "Пункт «Нет данных» есть в каждом таком списке — выбирайте его, если по этому полю "
            "для машины сведений нет.",
            None,
        ),
        (
            "Колонки «Авторезина», «Летняя резина — состояние», «Зимняя резина — состояние», "
            "«Состояние лакокрасочного покрытия», «Кузов», «Пропуск: <Название>», «Обязательный "
            "техосмотр», «Брендирование (Да/Нет)», «Наличие радиостанции», «Наличие запасного "
            "колеса», «Трекер», «Тахограф», «Наличие и исправность зеркал» — исключение: Excel не "
            "даст вписать значение вне списка (в том числе список включает «Нет данных»), а если "
            "сторонний текст всё же попадёт в файл (например, скопирован из другой таблицы), при "
            "импорте он будет отброшен с предупреждением и сохранён в «Примечание».",
            None,
        ),
        ("", None),
        ("5. Пропуска", section_font),
        (
            "Пропуска — произвольный набор, разный у каждой машины: колонки «Пропуск: ЗО», "
            "«Пропуск: ХО», «Пропуск: ДНР», «Пропуск: ЛНР», «Пропуск: Москва» — только пример "
            "базового набора. Если организации нужен пропуск с другим названием — допишите пару "
            "своих колонок в том же формате «Пропуск: <Название>» и «Пропуск: <Название> — до», "
            "система распознает их автоматически по этому шаблону заголовка.",
            None,
        ),
        ("", None),
        ("6. Что будет с уже существующими гос. номерами", section_font),
        (
            "Если гос. номер из файла уже есть в системе, при импорте можно выбрать одну из "
            "двух стратегий: «Пропустить» — существующая запись не изменится, строка файла "
            "будет проигнорирована; «Обновить» — заполненные в файле поля перезапишут значения "
            "в существующей карточке ТС (пустые ячейки при обновлении ничего не затирают).",
            None,
        ),
        ("", None),
        ("7. Куда нести готовый файл", section_font),
        (
            "Сохраните файл и в реестре «Автотранспорт» нажмите кнопку «Импорт» — там же можно "
            "будет сопоставить организации, если название в файле не совпало ни с одной из "
            "имеющихся, и выбрать стратегию для уже существующих гос. номеров.",
            None,
        ),
    ]

    for idx, (text, font) in enumerate(rows, 1):
        cell = ws.cell(idx, 1, text)
        if font:
            cell.font = font
        cell.alignment = body_align
        if text:
            # грубая оценка высоты строки под перенос текста в колонке шириной 110
            ws.row_dimensions[idx].height = max(18, 15 * (len(text) // 100 + 1))
