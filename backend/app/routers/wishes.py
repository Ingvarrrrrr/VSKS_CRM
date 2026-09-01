import re as _re
from fastapi import APIRouter, Depends, Query, HTTPException
from sqlalchemy import select, delete, func, or_, and_
from datetime import date, datetime
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Optional

from app.database import get_db
from app.auth.jwt import (
    get_current_user, get_org_filter, require_role,
    ALL_ROLES, MANAGER_ROLES, ADMIN_ROLES, OWNER_ROLES,
)
from app.auth.permissions import require_tab, has_org_key
from app.auth.visibility import build_visibility_clause, get_visible_user_ids, get_visible_subsidy_ids
from app.models.user import User
from app.models.wish import Wish
from app.models.wish_item import WishItem
from app.models.wish_member import WishMember
from app.schemas.wishes import WishCreate, WishUpdate, WishOut, WishReject, WishConvert, WishItemPatch, WishExecutionPatch, WishItemFeoPatch, WishStatusForce, WishStop, WishItemPurchaseMatch, WishPurchaseSummary
from app.models.purchase import Purchase
from app.models.purchase_item import PurchaseItem
from app.models.feo_category import FeoCategory
from app.services.text_match import normalize as _normalize_name
from app.models.purchase_event import PurchaseMember
from app.routers.purchase_members import _create_assignment_chat_room
from app.models.chat_message import ChatMessage
from app.services.feo_plan import assert_no_unapproved_excess, assert_tz_not_over_plan, assert_tz_batch_not_over_plan, compute_feo_plan_tree
# _auto_assign_planned_items вынесена в app.services.plan_autoassign (владелец,
# 2026-08-12, «закупка сама становится планом»): нужна ТАКЖЕ из purchases.py для
# закупок, созданных/меняемых в обход заявки. Поведение НЕ менялось при переносе —
# см. полный докстринг в app/services/plan_autoassign.py::auto_assign_planned_items.
# Alias сохраняет имя со старым подчёркиванием — все вызовы ниже по файлу не тронуты.
from app.services.plan_autoassign import (
    auto_assign_planned_items as _auto_assign_planned_items,
    backfill_item_type_from_plan as _backfill_item_type_from_plan,
    move_or_detach_planned_item as _move_or_detach_planned_item,
    deactivate_if_orphaned as _deactivate_if_orphaned,
)
from app.models.feo_planned_item import FeoPlannedItem
from decimal import Decimal
from app.services.wish_tabs import WISH_TAB_STATUS_SETS, wish_tab_statuses


def _is_saas(user: User) -> bool:
    """SaaS-роли (superadmin/account_owner) — обходят любые status-guard'ы."""
    return user.role in OWNER_ROLES


router = APIRouter(prefix="/api/wishes", tags=["wishes"])

# Phase 31: fields tracked for diff-highlighting (D-05..D-09)
# estimated_price is the wish amount proxy (Wish has no total_price column)
WISH_TRACKED_FIELDS: set[str] = {
    "title", "description", "status", "subsidy_id", "estimated_price",
}


def _enrich(w: Wish) -> WishOut:
    """Convert Wish ORM object to WishOut, filling computed name fields."""
    d = WishOut.model_validate(w)
    if w.creator:
        d.creator_name = w.creator.full_name or w.creator.username
    if w.approver:
        d.approver_name = w.approver.full_name or w.approver.username
    if w.subsidy:
        d.subsidy_name = w.subsidy.name
    if w.assignee:
        d.assignee_name = w.assignee.full_name or w.assignee.username
        d.assigned_to_name = d.assignee_name  # alias for legacy frontend
    if getattr(w, 'event', None):
        d.event_name = w.event.name
    if getattr(w, 'executor', None):
        d.executor_name = w.executor.full_name or w.executor.username
    if getattr(w, 'stopped_by_user', None):
        d.stopped_by_name = w.stopped_by_user.full_name or w.stopped_by_user.username
    if getattr(w, 'rejected_by_user', None):
        d.rejected_by_name = w.rejected_by_user.full_name or w.rejected_by_user.username
    # Контрагент — из справочника, если contractor_id задан, иначе свободный
    # ввод contractor_name (см. WishOut.contractor_display_name).
    if getattr(w, 'contractor', None):
        d.contractor_display_name = w.contractor.name
    elif w.contractor_name:
        d.contractor_display_name = w.contractor_name
    return d


async def _load_wish(wish_id: int, db: AsyncSession) -> Wish:
    """Load wish with all relationships."""
    result = await db.execute(
        select(Wish)
        .options(
            selectinload(Wish.creator),
            selectinload(Wish.approver),
            selectinload(Wish.assignee),
            selectinload(Wish.subsidy),
            selectinload(Wish.event),
            selectinload(Wish.executor),
            selectinload(Wish.items),
            selectinload(Wish.stopped_by_user),
            selectinload(Wish.contractor),
            selectinload(Wish.rejected_by_user),
        )
        .where(Wish.id == wish_id)
    )
    wish = result.scalar_one_or_none()
    if wish is None:
        raise HTTPException(status_code=404, detail="Заявка не найдена")
    return wish


async def _attach_purchase_matches(wish: Wish, enriched: WishOut, db: AsyncSession) -> None:
    """W-diff (2026-08-13, карточка заявки только): для каждой позиции заявки
    находит её «двойника» в закупке(ах), созданных из этой заявки, и заполняет
    WishItemOut.purchase_match (см. WishItemPurchaseMatch за назначением полей).

    Сопоставление (владелец, 2026-08-13, доразбор неоднозначности заявки №31 —
    «Перчатки … размер L» дважды в заявке, 4 шт/«Финал» и 20 шт/«Окружные»,
    человек различает их количеством):
      1. По PurchaseItem.wish_item_id — надёжная прямая связь.
      2. Иначе — по точному нормализованному имени (app.services.text_match.normalize)
         среди позиций закупок ЭТОЙ ЖЕ заявки, у которых wish_item_id ПУСТ (позиции
         с прямой связью уже разобраны шагом 1 и не участвуют в поиске по имени —
         иначе можно «увести» чужого двойника). Если под этим именем ИЗНАЧАЛЬНО
         (до любых claim'ов) был ровно один кандидат — match_method='item_name'.
      3. Если изначально кандидатов было НЕСКОЛЬКО — сужаем точным совпадением
         quantity. Единственный кандидат с подходящим количеством —
         match_method='item_name_qty' (связь по имени И количеству, честно
         отличается от простого 'item_name'). Это касается КАЖДОЙ позиции
         заявки из такой группы, даже если к моменту её обработки в пуле
         остался только один кандидат (соседняя позиция уже забрала другой) —
         статус группы «была неоднозначна по имени» фиксируется ДО обработки,
         чтобы обе позиции пары «4 шт/20 шт» получили одинаково честную метку,
         а не «первая — по количеству, вторая — как бы просто по имени».
      4. Всё ещё больше одного кандидата с подходящим количеством (или у позиции
         заявки количество не задано) — неоднозначность, наугад не выбираем:
         match_method='item_name_ambiguous' + ambiguous_candidates_count.
         Остальные поля пустые.
      5. Не нашлось вовсе — все поля None.

    Каждый кандидат-двойник закупки достаётся НЕ БОЛЕЕ ЧЕМ одной позиции заявки:
    выбранный кандидат удаляется из пула конкретного имени (list.pop/remove),
    иначе при повторе того же имени в заявке (см. пример выше) вторая позиция
    получила бы того же самого «уже занятого» двойника.

    Один пакетный запрос закупок + один пакетный запрос их позиций (с JOIN на
    feo_categories.name) на всю заявку — без N+1 по позициям.
    """
    purchases_rows = (await db.execute(
        select(Purchase.id, Purchase.purchase_number, Purchase.status, Purchase.stopped_at)
        .where(Purchase.wish_id == wish.id)
    )).all()
    purchase_info = {r[0]: r for r in purchases_rows}  # purchase_id -> (id, number, status, stopped_at)
    purchase_ids = list(purchase_info.keys())

    purchase_items_rows = []
    if purchase_ids:
        purchase_items_rows = (await db.execute(
            select(PurchaseItem, FeoCategory.name)
            .outerjoin(FeoCategory, FeoCategory.id == PurchaseItem.feo_category_id)
            .where(PurchaseItem.purchase_id.in_(purchase_ids))
        )).all()

    direct_by_wish_item_id: dict[int, tuple] = {}
    unclaimed_by_name: dict[str, list[tuple]] = {}
    for pi, feo_name in purchase_items_rows:
        if pi.wish_item_id is not None:
            direct_by_wish_item_id.setdefault(pi.wish_item_id, (pi, feo_name))
        else:
            key = _normalize_name(pi.item_name or "")
            if key:
                unclaimed_by_name.setdefault(key, []).append((pi, feo_name))
    # Кол-во кандидатов под каждое имя ДО начала claim'а (снимок) — нужно, чтобы
    # оба «близнеца» одной пары «одинаковое имя, разное количество» (см. докстринг)
    # были помечены одинаково 'item_name_qty', а не «первый по количеству, второй
    # по остатку» (иначе один и тот же по сути разбор количества выглядел бы для
    # фронта как связь по имени БЕЗ количества у второй позиции).
    original_candidate_count: dict[str, int] = {k: len(v) for k, v in unclaimed_by_name.items()}

    def _build_match(method: str, pi, feo_name) -> WishItemPurchaseMatch:
        p_row = purchase_info.get(pi.purchase_id)
        return WishItemPurchaseMatch(
            match_method=method,
            purchase_item_id=pi.id,
            purchase_id=pi.purchase_id,
            purchase_number=p_row[1] if p_row else None,
            purchase_status=p_row[2] if p_row else None,
            purchase_stopped_at=p_row[3] if p_row else None,
            feo_category_id=pi.feo_category_id,
            feo_category_name=feo_name,
            quantity=float(pi.quantity) if pi.quantity is not None else None,
            unit_price=float(pi.unit_price) if pi.unit_price is not None else None,
            total_price=float(pi.total_price) if pi.total_price is not None else None,
        )

    wish_items = list(wish.items or [])
    for w_item, out_item in zip(wish_items, enriched.items):
        direct = direct_by_wish_item_id.get(w_item.id)
        if direct is not None:
            out_item.purchase_match = _build_match("wish_item_id", direct[0], direct[1])
            continue
        key = _normalize_name(w_item.item_name or "")
        candidates = unclaimed_by_name.get(key, []) if key else []
        was_ambiguous_name = original_candidate_count.get(key, 0) > 1
        if not candidates:
            continue  # не нашлось — purchase_match остаётся None (default)
        if len(candidates) == 1 and not was_ambiguous_name:
            # Имя изначально указывало ровно на одну позицию закупки — количество
            # не потребовалось.
            pi, feo_name = candidates.pop(0)
            out_item.purchase_match = _build_match("item_name", pi, feo_name)
        else:
            # Имя само по себе неоднозначно (сейчас или изначально — включая
            # случай, когда «сосед» по имени уже разобран выше и остался ровно
            # один кандидат: для НЕГО это тоже разбор по количеству, а не по
            # голому имени). Сужаем точным совпадением quantity.
            qty_matches = [
                c for c in candidates
                if w_item.quantity is not None and c[0].quantity == w_item.quantity
            ]
            if len(qty_matches) == 1:
                pi, feo_name = qty_matches[0]
                candidates.remove((pi, feo_name))  # claim — не достанется другой позиции заявки
                out_item.purchase_match = _build_match("item_name_qty", pi, feo_name)
            else:
                out_item.purchase_match = WishItemPurchaseMatch(
                    match_method="item_name_ambiguous",
                    ambiguous_candidates_count=len(candidates),
                )


async def _is_wish_member(wish_id: int, user_id: int, db: AsyncSession) -> bool:
    res = await db.execute(
        select(WishMember.id).where(
            WishMember.wish_id == wish_id,
            WishMember.user_id == user_id,
        ).limit(1)
    )
    return res.scalar_one_or_none() is not None


async def _ensure_no_pending_approvals(
    wish, db: AsyncSession, current_user=None, allow_override: bool = False
) -> None:
    """Конвертация заявки запрещена, пока цепочка согласования не завершена:
    иначе заявка уходит в converted, а pending-согласующие «зависают»
    и не видят её во вкладке «На согласовании мне».

    Собственное pending-согласование текущего пользователя блоком не считается:
    «Быстрое одобрение» согласующим = его решение approved. Оно фиксируется в цепочке,
    и блокировка снимается, если других pending-согласующих не осталось.

    allow_override=True + менеджер/SaaS → одобряет ВСЕ pending от имени current_user,
    не блокирует конвертацию."""
    from datetime import datetime, timezone
    from app.models.wish_approval import WishApproval
    pending = (await db.execute(
        select(WishApproval)
        .where(WishApproval.wish_id == wish.id, WishApproval.status == "pending")
        .order_by(WishApproval.order_num)
    )).scalars().all()

    own = [a for a in pending if current_user is not None and a.user_id == current_user.id]
    others = [a for a in pending if not (current_user is not None and a.user_id == current_user.id)]

    if others:
        # W3: менеджер/SaaS может одобрить чужие pending без блокировки
        is_manager = (
            current_user is not None
            and (
                _is_saas(current_user)
                or getattr(current_user, 'role', None) in MANAGER_ROLES
            )
        )
        if allow_override and is_manager:
            override_comment = (
                f"Одобрено без согласования остальных "
                f"({getattr(current_user, 'full_name', None) or getattr(current_user, 'username', '')})"
            )
            now = datetime.now(timezone.utc)
            for a in others:
                a.status = "approved"
                a.decided_at = now
                a.decided_by_user_id = current_user.id
                a.decided_by_username = current_user.full_name or current_user.username
                a.comment = override_comment
            await db.flush()
        else:
            names = ", ".join(a.approver_full_name or f"пользователь #{a.user_id}" for a in others)
            raise HTTPException(
                status_code=409,
                detail=f"У заявки незавершённое согласование ({names}). "
                       "Дождитесь решения согласующих или удалите цепочку согласования.",
            )

    # Других pending нет (или они одобрены выше) — фиксируем собственное решение согласующего как approved
    for a in own:
        a.status = "approved"
        a.decided_at = datetime.now(timezone.utc)
        a.decided_by_user_id = current_user.id
        a.decided_by_username = current_user.full_name or current_user.username
    if own:
        await db.flush()


def _as_date(value):
    """needed_date приходит из нетипизированного items: list — Pydantic его не приводит."""
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return None
        return date.fromisoformat(value[:10])
    if isinstance(value, datetime):
        return value.date()
    return value


def _eff_date(wish, item):
    """Эффективная плановая дата позиции: приоритет
    позиция(needed_date) → Срок исполнения(execution_deadline) → Желаемая дата(desired_date)."""
    return (
        getattr(item, "needed_date", None)
        or getattr(wish, "execution_deadline", None)
        or getattr(wish, "desired_date", None)
    )


async def _ensure_needed_dates(wish, db, items, context: str = "convert") -> None:
    """W2-гейт: если субсидия требует плановые даты — проверяем все items.
    Бросает HTTPException 409 с error_code='missing_needed_dates' если есть позиции без даты.
    Авансовые отчёты пропускаются вызывающим — этот хелпер не проверяет source.
    context='submit' — сообщение для отправки на согласование; 'convert' (default) — для переноса в план закупок."""
    if not wish.subsidy_id:
        return
    from app.models.subsidy import Subsidy
    subsidy = await db.get(Subsidy, wish.subsidy_id)
    if not (subsidy and subsidy.require_planned_dates):
        return
    items_without_date = [it for it in items if not _eff_date(wish, it)]
    if not items_without_date:
        return
    names_list = ", ".join(f'«{it.item_name}»' for it in items_without_date[:5])
    suffix = f" и ещё {len(items_without_date) - 5} поз." if len(items_without_date) > 5 else ""
    if context == "submit":
        intro = (
            f"Невозможно отправить заявку на согласование: у следующих позиций не указана "
            f"дата потребности (к какой дате планируется закупить): {names_list}{suffix}."
        )
    else:
        intro = (
            f"Невозможно перенести заявку в План закупок: у следующих позиций не указана "
            f"дата потребности (к какой дате планируется закупить): {names_list}{suffix}."
        )
    message = (
        f"{intro} "
        f"Без даты потребности ФЭО не может распределить расходы по месяцам. "
        f"Укажите дату потребности для позиции, либо «Срок исполнения», либо «Желаемую дату поставки/исполнения» "
        f"в заявке. Требование дат можно отключить в настройках субсидии "
        f"(доступно Хозяину аккаунта)."
    )
    raise HTTPException(
        status_code=409,
        detail={
            "message": message,
            "error_code": "missing_needed_dates",
            "missing_item_ids": [it.id for it in items_without_date],
            "missing_item_names": [it.item_name for it in items_without_date],
        },
    )


async def _ensure_feo_categories_assigned(wish, items, db: AsyncSession) -> None:
    """Жёсткий гейт «без категории ФЭО закупка не проходит» (владелец, 2026-08-11).

    Реальный случай с прода: заявка №32 «Приобретение Пикапов» была согласована,
    из неё создалась закупка РЕЕ-2026-00889 (Great Wall POER, 2 шт × 4 000 000), но
    у позиции и у самой заявки feo_category_id был пуст. Автозаведение плановой
    позиции (_auto_assign_planned_items) берёт эффективную категорию, не находит её
    и молча пропускает позицию — закупка осталась сиротой, её сумма не попала ни в
    один план ФЭО.

    Проверяем ДО любых мутаций (при отказе ни одна закупка не создаётся частично):
    у КАЖДОЙ позиции должна быть эффективная категория — item.feo_category_id, а
    если он не задан — фолбэк на wish.feo_category_id. Категория должна не только
    быть задана, но и реально существовать в справочнике (FeoCategory) — ссылка на
    удалённую категорию (структуру ФЭО субсидии пересоздавали) раньше молча
    обнулялась и закупка создавалась без ФЭО (см. историю ветки convert_warning,
    убрана этим гейтом) — именно так обнулилась категория у заявки №32 ещё до
    автозаведения. Теперь это тоже отказ, а не предупреждение постфактум.

    wish.feo_category_id, если задан, тоже обязан существовать — он пишется в
    Purchase.feo_category_id напрямую (не только как фолбэк для позиций без своей
    категории), битая ссылка уронила бы INSERT закупки FK-violation'ом (500 без
    объяснения), если бы её не отловили здесь заранее.

    `items` — WishItem ORM или dict payload (см. _item_field) — вызывающий обязан
    заранее отфильтровать строки-заготовки (_is_meaningful_item), иначе пустая
    заготовка без названия тоже потребует категорию.
    """
    from app.models.feo_category import FeoCategory

    wish_cat_id = wish.feo_category_id
    ref_ids = {i for i in ({wish_cat_id} | {_item_field(it, 'feo_category_id') for it in items}) if i}
    valid_ids: set[int] = set()
    if ref_ids:
        valid_ids = set((await db.execute(
            select(FeoCategory.id).where(FeoCategory.id.in_(ref_ids))
        )).scalars().all())

    wish_cat_ok = wish_cat_id is None or wish_cat_id in valid_ids
    problem_names: list[str] = []
    for it in items:
        item_cat = _item_field(it, 'feo_category_id')
        eff_cat = item_cat or wish_cat_id
        if eff_cat is None or eff_cat not in valid_ids:
            problem_names.append(str(_item_field(it, 'item_name') or 'без названия'))

    if not problem_names and wish_cat_ok:
        return

    names_list = ", ".join(f'«{n}»' for n in problem_names[:10])
    suffix = f" и ещё {len(problem_names) - 10} поз." if len(problem_names) > 10 else ""
    parts: list[str] = []
    if problem_names:
        parts.append(
            f"у следующих позиций нет категории ФЭО, либо она ссылается на удалённую "
            f"из справочника категорию: {names_list}{suffix}"
        )
    if not wish_cat_ok:
        parts.append(
            "категория ФЭО самой заявки была удалена из справочника "
            "(структуру ФЭО субсидии пересоздавали)"
        )
    message = (
        "Невозможно перенести заявку в План закупок: " + "; ".join(parts) + ". "
        "Выберите категорию ФЭО для каждой позиции (или для заявки целиком в разделе "
        "«Категория ФЭО»), а если категория неизвестна — выберите «Не определена». "
        "Без категории закупка не попадёт ни в один план ФЭО и её сумма потеряется."
    )
    raise HTTPException(
        status_code=409,
        detail={
            "message": message,
            "error_code": "missing_feo_category",
            "missing_item_names": problem_names,
        },
    )


# W1: статусы «дошли до договора» — блокируют редактирование привязанной заявки
CONTRACTED_STATUSES = ("contracted", "ordered", "delivered", "paid")


async def _wish_linked_purchases(wish_id: int, db: AsyncSession) -> list:
    """Возвращает список закупок, привязанных к заявке."""
    res = await db.execute(select(Purchase).where(Purchase.wish_id == wish_id))
    return res.scalars().all()


async def _wish_purchase_summaries_map(wish_ids: list, db: AsyncSession) -> dict:
    """Пункт 4 (владелец, 2026-08-13): «из согласованной заявки — переход в её
    закупки; если их несколько — выпадающий список с номером/статусом/суммой».
    purchase_ids (List[int], см. рядом) отдаёт только id — этого мало для
    осмысленного меню. Формат карточки закупки вынесен в
    app.services.purchase_summary.purchase_summaries_by_id (план
    zany-fluttering-mountain.md, 2026-08-13) — переиспользуется и «виновниками»
    превышения плана (excess_plan_items, см. app.services.feo_plan), чтобы
    формат карточки не разъехался на два.
    """
    out: dict = {}
    if not wish_ids:
        return out
    from app.services.purchase_summary import purchase_summaries_by_id

    link_rows = (await db.execute(
        select(Purchase.wish_id, Purchase.id)
        .where(Purchase.wish_id.in_(wish_ids))
        .order_by(Purchase.id)
    )).all()
    if not link_rows:
        return out
    summaries = await purchase_summaries_by_id(db, (pid for _wid, pid in link_rows))
    for wid, pid in link_rows:
        s = summaries.get(pid)
        if s is None:
            continue
        out.setdefault(wid, []).append(WishPurchaseSummary(**s))
    return out


def _item_field(it, name: str, default=None):
    """Универсальный доступ к полю позиции: и pydantic/dict-payload (body.items —
    список dict), и ORM WishItem (атрибуты) должны обрабатываться одним предикатом."""
    if isinstance(it, dict):
        return it.get(name, default)
    return getattr(it, name, default)


def _is_meaningful_item(it) -> bool:
    """Строка-заготовка (без названия, количества и суммы) — фронт создаёт их
    заранее для будущего ввода; в БД, План закупок и в гейты они попадать не должны.
    Принимает как ORM WishItem, так и «сырой» dict из body.items."""
    name = _item_field(it, "item_name") or ""
    quantity = _item_field(it, "quantity") or 0
    total_price = _item_field(it, "total_price") or 0
    return bool(str(name).strip() or float(quantity or 0) or float(total_price or 0))


async def _wish_contracted_locked(wish_id: int, db: AsyncSession) -> bool:
    """True если хотя бы одна привязанная закупка находится в статусе «Договор+».
    Тонкая обёртка над _wish_locked_descr (см. ниже) — держим оба места на одном
    источнике правды, чтобы bool-гейт и текстовое сообщение не разъехались."""
    return bool(await _wish_locked_descr(wish_id, db))


async def _wish_locked_descr(wish_id: int, db: AsyncSession) -> Optional[str]:
    """Человекочитаемое описание закупок, из-за которых заявка заблокирована
    для правки (Договор+) — None, если блокирующих закупок нет.

    Владелец увидел на проде сообщение «на этапе «Договор»» для закупки в
    статусе `ordered` («Заказано») — CONTRACTED_STATUSES шире одного статуса
    «Договор» (contracted/ordered/delivered/paid), а текст называл только его.
    Единый хелпер вместо копипасты формата в двух местах (409 при PUT и при
    удалении позиций) — чтобы формулировка не расходилась. Формат каждой
    закупки: `№{purchase_number or id} «{item_name}» — стадия «{label}»`.
    """
    from app.routers.purchase_export import _STATUS_LABELS
    purchases = await _wish_linked_purchases(wish_id, db)
    locked = [p for p in purchases if p.status in CONTRACTED_STATUSES]
    if not locked:
        return None
    return "; ".join(
        f"№{p.purchase_number or p.id} «{p.item_name or ''}» — стадия «{_STATUS_LABELS.get(p.status, p.status)}»"
        for p in locked
    )


async def _reset_approvals(wish_id: int, db: AsyncSession, keep_user_id: Optional[int] = None) -> None:
    """Сбрасывает решения согласующих цепочки обратно в pending.

    keep_user_id (владелец, 2026-08-19: «нужно, чтобы было видно, кто
    отклонил») — строка ЭТОГО согласующего НЕ трогается: у него остаются
    status='rejected', decided_at, decided_by_user_id, comment, чтобы после
    сброса остальных было видно, кто именно отклонил заявку. Остальные — как
    раньше, в pending."""
    from app.models.wish_approval import WishApproval
    approvals = (await db.execute(
        select(WishApproval).where(WishApproval.wish_id == wish_id)
    )).scalars().all()
    for a in approvals:
        if keep_user_id is not None and a.user_id == keep_user_id:
            continue
        a.status = "pending"
        a.decided_at = None
        a.decided_by_user_id = None
        a.decided_by_username = None
        a.comment = None
    if approvals:
        await db.flush()


async def _sync_wish_items_to_purchases(wish, db: AsyncSession) -> None:
    """УСТАРЕЛО, БОЛЬШЕ НЕ ВЫЗЫВАЕТСЯ (владелец, 2026-08-13): «после согласования
    заявку править нельзя, только в закупке» — единственный вызывающий код
    (_distribute_wish_to_purchases, ветка повторного одобрения уже
    конвертированной заявки) отключён, т.к. правка wish.items теперь
    запрещена на статусе 'converted' (см. гейт в update_wish). Оставлено в
    коде на случай будущего отката решения — НЕ удалять просто так.

    Синхронизирует позиции заявки в связанные закупки (не в TZ_FROZEN_STATUSES).

    Для каждой закупки НЕ в TZ_FROZEN_STATUSES: каждая PurchaseItem с wish_item_id
    находит соответствующий WishItem и копирует item_name/unit/unit_price/quantity/total_price.
    Затем пересчитывает суммы закупки.

    Асимметрия (владелец, 2026-08-07, план zany-fluttering-mountain.md, шаг 5):
    раньше здесь стоял CONTRACTED_STATUSES (contracted/ordered/delivered/paid) —
    на «Ведётся работа» (work_in_progress) правка заявки по-прежнему могла
    двигать цену/кол-во позиции закупки, хотя прямой PATCH той же позиции
    (purchases.py:patch_purchase_item) её уже замораживает по TZ_FROZEN_STATUSES.
    Заморозка ТЗ обходилась правкой заявки. Приведено к одному набору статусов —
    TZ_FROZEN_STATUSES (импорт из purchases.py, локально — во избежание цикла
    роутер↔роутер).
    """
    from app.routers.purchases import TZ_FROZEN_STATUSES as _TZ_FROZEN_STATUSES

    wish_item_map = {wi.id: wi for wi in (wish.items or [])}
    if not wish_item_map:
        return
    purchases = await _wish_linked_purchases(wish.id, db)
    for p in purchases:
        if p.status in _TZ_FROZEN_STATUSES:
            continue
        pitems_res = await db.execute(
            select(PurchaseItem).where(
                PurchaseItem.purchase_id == p.id,
                PurchaseItem.wish_item_id.isnot(None),
            )
        )
        pitems = pitems_res.scalars().all()
        changed = False
        for pi in pitems:
            wi = wish_item_map.get(pi.wish_item_id)
            if wi is None:
                continue
            _new_qty = wi.quantity
            _new_price = wi.unit_price
            _new_total = (wi.unit_price or 0) * (wi.quantity or 0)
            # Шаг 5 «цена ТЗ не выше плановой»: та же позиция, тот же гейт, что и
            # у прямого PATCH — правка через заявку не должна быть лазейкой.
            # over_plan=true — сознательно сверх плана, пропускаем (см. purchases.py).
            if not getattr(wi, 'over_plan', False):
                await assert_tz_not_over_plan(
                    db,
                    feo_planned_item_id=wi.feo_planned_item_id,
                    feo_category_id=wi.feo_category_id,
                    quantity=_new_qty,
                    unit_price=_new_price,
                    total_price=_new_total,
                    item_name=wi.item_name,
                )
            pi.item_name = wi.item_name
            pi.unit = wi.unit
            pi.unit_price = _new_price
            pi.quantity = _new_qty
            pi.total_price = _new_total
            # Снимок плана (Шаг 1): позиция ещё НЕ ушла из плана закупок (проверено
            # выше — p.status not in TZ_FROZEN_STATUSES), поэтому правка заявки
            # по-прежнему двигает и «текущую» цену, и зафиксированный план вместе —
            # план не заморожен, пока закупка не объявлена (Шаг 2).
            pi.planned_quantity = wi.quantity
            pi.planned_unit_price = wi.unit_price
            pi.planned_total = (wi.unit_price or 0) * (wi.quantity or 0)
            pi.feo_category_id = wi.feo_category_id
            pi.feo_planned_item_id = wi.feo_planned_item_id
            pi.over_plan = getattr(wi, 'over_plan', False)
            pi.vat_rate = getattr(wi, 'vat_rate', None)
            changed = True
        if changed:
            # Признак «Товар/Услуга/Работа» (блок 1): наследуется от плановой
            # позиции, если у самой позиции закупки он ещё не задан (уже
            # заполненный — не трогаем).
            await _backfill_item_type_from_plan(pitems, db)
            p.feo_per_item = bool(getattr(wish, 'feo_per_item', False))
            p.vat_mode = getattr(wish, 'vat_mode', None) or 'uniform'
            await db.flush()
            # Пересчёт сумм закупки
            items_sum_res = await db.execute(
                select(func.coalesce(func.sum(PurchaseItem.total_price), 0))
                .where(PurchaseItem.purchase_id == p.id)
            )
            items_sum = items_sum_res.scalar() or 0
            p.total_nmck = items_sum
            p.planned_total_price = items_sum or p.planned_total_price
            await db.flush()


async def _withdraw_wish_from_plan(wish_id: int, db: AsyncSession, *, action_text: str) -> None:
    """Убирает закупки заявки из плана закупок — обратная операция к
    _distribute_wish_to_purchases.

    Жёсткий гейт (владелец, 2026-08-07): если хоть одна связанная закупка ушла
    дальше «Плана закупок» (work_in_progress/contracted/ordered/delivered/paid),
    откат заявки ЗАПРЕЩЁН — HTTPException(409) с перечислением номеров, предметов
    и стадий. Проверка идёт ДО любых мутаций, поэтому при блокировке ни одна
    закупка заявки не трогается (частичного отката не бывает).
    Закупки, оставшиеся в 'plan_schedule' (и только они), возвращаются в скрытый
    статус 'wishes' — не удаляются: сохраняются история, файлы, чаты и связь
    с заявкой; при повторном одобрении гейт вернёт их в план.
    `action_text` — что именно не удаётся сделать («вернуть в черновик»,
    «отклонить», «удалить», ...) — подставляется в текст 409 вызывающей точкой,
    чтобы сообщение звучало по-русски для конкретного действия.
    Commit НЕ делает — это на вызывающем.
    """
    from app.routers.purchase_budget import PLANNED_STATUSES
    from app.routers.purchase_export import _STATUS_LABELS

    purchases = await _wish_linked_purchases(wish_id, db)
    blockers: list[str] = []
    for p in purchases:
        if p.status != "plan_schedule" and p.status in PLANNED_STATUSES:
            # work_in_progress/contracted/ordered/delivered/paid — стадия ушла дальше
            # «Плана закупок», откат запрещён.
            label = _STATUS_LABELS.get(p.status, p.status)
            blockers.append(f"№{p.purchase_number or p.id} «{p.item_name or ''}» на стадии «{label}»")
    if blockers:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Заявку нельзя {action_text}: закупка " + "; ".join(blockers)
                + ". Сначала отмените её в реестре закупок."
            ),
        )
    changed = False
    for p in purchases:
        if p.status == "plan_schedule":
            p.status = "wishes"
            changed = True
        # status == 'wishes' и прочие (например 'cancelled') — не трогаем
    if changed:
        await db.flush()


async def _notify_pending_approvers(wish, db: AsyncSession, requester_name: str) -> None:
    """Уведомляет согласующих из цепочки о необходимости решения.

    sequential → уведомить первого pending; parallel → уведомить всех.
    Повторно используется из submit_wish и update_wish (при повторном согласовании).
    """
    try:
        from app.models.wish_approval import WishApproval
        from app.notifications import notify_wish_approval_step
        from app.models.user import User as _User
        pending = (await db.execute(
            select(WishApproval).where(
                WishApproval.wish_id == wish.id,
                WishApproval.status == "pending",
            ).order_by(WishApproval.order_num)
        )).scalars().all()
        if pending:
            targets = pending[:1] if (wish.approval_mode or "sequential") == "sequential" else pending
            for ap in targets:
                if ap.user_id and ap.user_id != wish.created_by:
                    approver_user = await db.get(_User, ap.user_id)
                    if approver_user:
                        await notify_wish_approval_step(wish, approver_user, requester_name)
    except Exception as e:
        import logging as _log
        _log.getLogger(__name__).warning("notify approvers failed: %s", e)


async def _collect_excess_warnings(
    db: AsyncSession, subsidy_id: Optional[int], cat_items: dict[int, list[dict]],
) -> list[dict]:
    """Владелец (2026-08-12): «Когда заявка идёт с превышением, ... она должна
    уходить [в закупку], но на закупке должен быть значок, что она заблокирована
    из-за превышения» — согласование заявки больше НЕ блокируется несогласованным
    превышением плана ФЭО (assert_no_unapproved_excess убран из путей создания
    закупок из заявки, см. вызывающий код), а вместо отказа собирает
    ПРЕДУПРЕЖДЕНИЯ, по одному на каждую затронутую категорию ФЭО (включая
    ПРЕДКОВ — та же цепочка parent_id, что и в assert_no_unapproved_excess, но
    без исключения).

    cat_items — {feo_category_id (лист, куда попала позиция заявки): [{"name","amount"}, ...]}.
    Вызывать ПОСЛЕ того как закупки/позиции уже созданы и flush()-нуты (до
    commit) — иначе compute_feo_plan_tree не увидит новый план и excess будет
    занижен на сумму этого самого действия.

    Возвращает список (дубли по category_id схлопнуты):
      {"category_id", "category_name", "budget", "plan_after", "excess_amount", "items"}
    Пустой список, если превышения нет или category_ids/subsidy_id отсутствуют.
    """
    if not cat_items or not subsidy_id:
        return []
    tree = await compute_feo_plan_tree(db, [subsidy_id])
    if not tree:
        return []
    from app.models.feo_category import FeoCategory

    warnings_by_cat: dict[int, dict] = {}
    for leaf_cid, items in cat_items.items():
        if leaf_cid not in tree or not items:
            continue
        # Цепочка «узел → предки» через parent_id (снизу вверх) — как в
        # assert_no_unapproved_excess, но собираем предупреждения, а не бросаем 409.
        chain_ids: list[int] = []
        cur_id: Optional[int] = leaf_cid
        seen: set[int] = set()
        while cur_id is not None and cur_id not in seen and cur_id in tree:
            seen.add(cur_id)
            chain_ids.append(cur_id)
            cur_id = tree[cur_id]["parent_id"]
        for cid in chain_ids:
            node = tree[cid]
            excess = node.get("excess_over_feo") or node.get("excess_amount") or 0.0
            if excess <= 0.005 or node.get("excess_approved"):
                continue
            entry = warnings_by_cat.get(cid)
            if entry is None:
                cat_row = await db.get(FeoCategory, cid)
                entry = {
                    "category_id": cid,
                    "category_name": cat_row.name if cat_row else f"#{cid}",
                    "budget": node.get("budget"),
                    "plan_after": float((node.get("plan") or 0.0) + (node.get("over") or 0.0)),
                    "excess_amount": float(excess),
                    "items": [],
                }
                warnings_by_cat[cid] = entry
            entry["items"].extend(items)
    return list(warnings_by_cat.values())


async def _sync_purchase_from_wish(wish, purchases: list, db: AsyncSession) -> Optional[dict]:
    """Повторное согласование заявки (вернули в черновик/отклонили → поправили →
    согласовали заново) приводит УЖЕ СУЩЕСТВУЮЩУЮ закупку заявки к её текущему
    состоянию — предмет и состав позиций (владелец, план crystalline-soaring-
    heron.md, п.2). Прод-находка: РЕЕ-2026-00898 сохранила предмет заявки №40 на
    момент ПЕРВОГО переноса, хотя заявку потом правили (в т.ч. предмет) и
    согласовывали заново — `_distribute_wish_to_purchases` в ветке «защита от
    дублей» раньше только продвигала статус скрытой закупки, ничего не сверяя.

    Вызывается из `_distribute_wish_to_purchases`/`convert_wish` ТОЛЬКО когда у
    заявки уже есть закупка(и) — по построению это повторный проход. Правка
    wish.items вне пути «вернуть в черновик» заблокирована на статусе 'converted'
    (см. гейт в update_wish), так что к моменту вызова любое расхождение состава —
    легитимная правка, сделанная, пока заявка была НЕ converted; правило «после
    согласования правят в закупке» этим не нарушается.

    len(purchases) != 1 — заявка когда-то распределена канбаном на НЕСКОЛЬКО
    закупок (split=True, approve_distribution) — без сохранённого сопоставления
    «группа → закупка» синхронизация неоднозначна, поэтому не выполняется вовсе
    (возвращает None, поведение как до этой задачи). Обычное согласование
    (split=False) — всегда ОДНА закупка, это подавляющее большинство случаев,
    включая описанный на проде.

    Гейт по стадии — тот же принцип и те же тексты, что `_withdraw_wish_from_plan`
    выше: закупка ушла дальше «Плана закупок» (TZ_FROZEN_STATUSES) — состав и
    предмет НЕ трогаются, причина возвращается в `blocked_reason` с номером и
    стадией по-русски (не HTTPException — согласование заявки уже состоялось,
    отказывать в нём из-за стадии закупки владелец не просил).

    Возвращает контракт `purchase_sync` (см. план, зафиксирован для фронта):
    {purchase_id, registry_number, subject_before, subject_after, items_added,
    items_removed, items_changed, items_conflicted, items_kept_manual,
    blocked_reason}. None — нечего/некого синхронизировать (0 или 2+ закупок).
    Commit НЕ делает — это на вызывающем (как и весь _distribute_wish_to_purchases).
    """
    if len(purchases) != 1:
        return None
    from app.routers.purchases import TZ_FROZEN_STATUSES
    from app.routers.purchase_export import _STATUS_LABELS

    p = purchases[0]

    if p.status in TZ_FROZEN_STATUSES:
        label = _STATUS_LABELS.get(p.status, p.status)
        return {
            "purchase_id": p.id,
            "registry_number": p.registry_number,
            "subject_before": p.subject,
            "subject_after": p.subject,
            "items_added": [],
            "items_removed": [],
            "items_changed": [],
            "items_conflicted": [],
            "items_kept_manual": [],
            "blocked_reason": (
                f"Закупка №{p.purchase_number or p.id} «{p.subject or p.item_name or ''}» "
                f"уже на стадии «{label}» — обновить предмет и состав из заявки нельзя. "
                "Дальнейшие изменения вносите прямо в закупке."
            ),
        }

    items_res = await db.execute(select(WishItem).where(WishItem.wish_id == wish.id))
    _all_wish_items = items_res.scalars().all()
    wish_items = [it for it in _all_wish_items if _is_meaningful_item(it)]
    # Дефект 2 (QA): ВСЕ id позиций заявки (включая «незначимые» — пустые
    # заготовки), не только meaningful — по ним отличаем «строка когда-то
    # пришла из заявки» от «заведена вручную прямо в закупке» ниже.
    _all_wish_item_ids = {it.id for it in _all_wish_items}

    # Те же гейты, что при первом переносе (_distribute_wish_to_purchases ниже) —
    # новая/изменившаяся позиция обязана иметь категорию ФЭО и дату потребности
    # прежде чем попасть в закупку.
    await _ensure_feo_categories_assigned(wish, wish_items, db)
    await _ensure_needed_dates(wish, db, wish_items)
    await _auto_assign_planned_items(
        wish_items, wish.feo_category_id, db, note=f"повторным согласованием заявки №{wish.id}",
    )
    await assert_tz_batch_not_over_plan(db, wish_items)

    pitems_res = await db.execute(
        select(PurchaseItem).where(PurchaseItem.purchase_id == p.id)
    )
    existing_items = pitems_res.scalars().all()

    # Дефект 1 (QA, 2026-08-21): сопоставление «старая позиция закупки → текущая
    # позиция заявки» ДВУХСТУПЕНЧАТО. Ступень 1 — по жёсткой связи wish_item_id
    # (переживает переименование позиции в НЕ-draft статусах заявки — update_wish
    # правит WishItem по id in-place там, см. докстринг функции выше). Ступень 2 —
    # по нормализованному имени (app.services.text_match.normalize, та же
    # функция, что и в plan_autoassign.auto_assign_planned_items) — ТОЛЬКО для
    # позиций закупки, не разобранных на ступени 1 (правка в статусе 'draft' идёт
    # «удалить все WishItem → создать заново» — id теряются целиком, имя остаётся
    # единственным якорем). Без ступени 1 переименование позиции читалось как
    # «удалить строку закупки → создать новую» и стирало ручные поля закупки
    # (contractor_id/contractor_name/contractor_inn/receipt_id/match_confirmed/
    # final_unit_price и пр.) — тот самый дефект.
    from app.services.text_match import normalize as _normalize_name

    _existing_by_wid: dict[int, PurchaseItem] = {}
    _unmatched_existing: list[PurchaseItem] = []
    for pi in existing_items:
        if pi.wish_item_id is not None and pi.wish_item_id in _all_wish_item_ids:
            _existing_by_wid[pi.wish_item_id] = pi
        else:
            _unmatched_existing.append(pi)

    _existing_pool: dict[str, list[PurchaseItem]] = {}
    for pi in _unmatched_existing:
        key = _normalize_name(pi.item_name or "")
        _existing_pool.setdefault(key, []).append(pi)

    _wish_contractor_obj = None
    if getattr(wish, 'contractor_id', None):
        from app.models.contractor import Contractor as _Contractor
        _wish_contractor_obj = await db.get(_Contractor, wish.contractor_id)

    items_added: list[dict] = []
    items_removed: list[dict] = []
    items_changed: list[dict] = []
    items_conflicted: list[dict] = []
    items_kept_manual: list[dict] = []
    synced_items: list[PurchaseItem] = []

    def _fmt(qty, price, total) -> str:
        return f"{qty or 0} × {price or 0} ₽ = {float(total or 0):.2f} ₽"

    # Дефект 3 (QA): поля с сохранённым «снимком ТЗ на момент переноса» —
    # planned_* движется вместе с текущим значением ТОЛЬКО пока его не трогали
    # руками в закупке (правило проекта «после согласования правят в закупке»).
    # Признак ручной правки — текущее значение отличается от снимка; тогда поле
    # НЕ перезаписывается из заявки (ни значение, ни сам снимок), а расхождение
    # возвращается в items_conflicted, чтобы фронт показал его человеку вместо
    # молчаливого отката.
    _tracked_fields = (
        ("quantity", "planned_quantity"),
        ("unit_price", "planned_unit_price"),
        ("total_price", "planned_total"),
    )

    for wi in wish_items:
        pi = _existing_by_wid.pop(wi.id, None)
        if pi is None:
            _key = _normalize_name(wi.item_name or "")
            _bucket = _existing_pool.get(_key)
            pi = _bucket.pop(0) if _bucket else None
        if pi is None:
            new_pi = PurchaseItem(
                purchase_id=p.id,
                product_id=wi.product_id,
                item_name=wi.item_name,
                item_type=wi.item_type,
                quantity=wi.quantity,
                unit=wi.unit,
                unit_price=wi.unit_price,
                total_price=wi.total_price,
                planned_quantity=wi.quantity,
                planned_unit_price=wi.unit_price,
                planned_total=wi.total_price,
                country_origin=wi.country_origin,
                feo_category_id=wi.feo_category_id,
                feo_planned_item_id=wi.feo_planned_item_id,
                over_plan=getattr(wi, 'over_plan', False),
                needed_date=_eff_date(wish, wi),
                wish_item_id=wi.id,
                vat_rate=getattr(wi, 'vat_rate', None),
                contractor_id=(_wish_contractor_obj.id if _wish_contractor_obj else None),
                contractor_inn=(_wish_contractor_obj.inn if _wish_contractor_obj else None),
                contractor_name=(
                    _wish_contractor_obj.name if _wish_contractor_obj
                    else getattr(wish, 'contractor_name', None)
                ),
            )
            db.add(new_pi)
            synced_items.append(new_pi)
            items_added.append({
                "name": wi.item_name, "quantity": float(wi.quantity or 0), "amount": float(wi.total_price or 0),
            })
            continue

        _before_qty, _before_price, _before_total = pi.quantity, pi.unit_price, pi.total_price
        _name_changed = (pi.item_name or "") != (wi.item_name or "")

        _any_field_changed = _name_changed
        for field, snap_field in _tracked_fields:
            cur_val = float(getattr(pi, field) or 0)
            snap_val = float(getattr(pi, snap_field) or 0)
            wish_val = float(getattr(wi, field) or 0)
            _manually_edited = abs(cur_val - snap_val) > 0.005
            if _manually_edited:
                if abs(cur_val - wish_val) > 0.005:
                    items_conflicted.append({
                        "name": wi.item_name, "field": field,
                        "in_purchase": cur_val, "in_wish": wish_val,
                    })
                # Не трогаем ни значение, ни снимок — ручная правка остаётся как есть.
                continue
            if abs(cur_val - wish_val) > 0.005:
                _any_field_changed = True
            setattr(pi, field, getattr(wi, field))
            setattr(pi, snap_field, getattr(wi, field))

        if _any_field_changed:
            items_changed.append({
                "name": wi.item_name,
                "was": _fmt(_before_qty, _before_price, _before_total),
                "now": _fmt(pi.quantity, pi.unit_price, pi.total_price),
            })
        pi.item_name = wi.item_name
        pi.item_type = wi.item_type or pi.item_type
        pi.unit = wi.unit
        pi.country_origin = wi.country_origin
        pi.feo_category_id = wi.feo_category_id
        pi.feo_planned_item_id = wi.feo_planned_item_id
        pi.over_plan = getattr(wi, 'over_plan', False)
        pi.vat_rate = getattr(wi, 'vat_rate', None)
        pi.needed_date = _eff_date(wish, wi)
        # Ре-линковка (см. комментарий у _existing_by_wid/_existing_pool выше):
        # правка в 'draft' пересоздаёт WishItem с новым id — восстанавливаем hard
        # link на АКТУАЛЬНЫЙ id, иначе связь «заявка ↔ строка закупки» (W1,
        # используется W-diff/exclude_wish_id/_fpi_reference_keys) обрывается
        # молча при первом же возврате в черновик, даже когда содержимое позиции
        # не менялось.
        pi.wish_item_id = wi.id
        synced_items.append(pi)

    # Дефект 2 (QA): позиции закупки, оставшиеся непарными, удаляются ТОЛЬКО
    # если реально пришли из заявки (wish_item_id указывает на позицию ЭТОЙ
    # заявки — включая «незначимые», см. _all_wish_item_ids выше — просто её
    # больше нет среди текущих значимых позиций заявки, автор убрал/обнулил её).
    # Строки, заведённые закупщиком прямо в закупке (wish_item_id пуст или
    # указывает не на эту заявку), НИКОГДА не трогаются — они возвращаются
    # отдельным списком items_kept_manual, чтобы пользователь видел, что они
    # остались, а не решил, что их «забыли».
    # _existing_by_wid, оставшийся непустым после основного цикла (см. .pop()
    # там) — записи, чей wish_item_id указывает на РЕАЛЬНУЮ позицию этой заявки,
    # но она перестала быть «значимой» (обнулена) и в wish_items не попала;
    # по тому же правилу это тоже «пришла из заявки» → в общий проход удаления.
    _leftover_buckets = list(_existing_pool.values()) + [[pi] for pi in _existing_by_wid.values()]
    for _bucket in _leftover_buckets:
        for pi in _bucket:
            if pi.wish_item_id is not None and pi.wish_item_id in _all_wish_item_ids:
                items_removed.append({
                    "name": pi.item_name, "quantity": float(pi.quantity or 0), "amount": float(pi.total_price or 0),
                })
                await db.delete(pi)
            else:
                items_kept_manual.append({
                    "name": pi.item_name, "quantity": float(pi.quantity or 0), "amount": float(pi.total_price or 0),
                })

    if synced_items:
        await _backfill_item_type_from_plan(synced_items, db)

    _subject_before = p.subject
    title = (wish.title or "").strip() or f"Заявка #{wish.id}"
    p.subject = title
    p.item_name = title
    if getattr(wish, 'contractor_id', None) and not p.contractor_id:
        p.contractor_id = wish.contractor_id

    await db.flush()

    items_sum_res = await db.execute(
        select(func.coalesce(func.sum(PurchaseItem.total_price), 0)).where(PurchaseItem.purchase_id == p.id)
    )
    items_sum = items_sum_res.scalar() or 0
    p.total_nmck = items_sum
    p.planned_total_price = items_sum
    p.nmck = items_sum
    await db.flush()

    return {
        "purchase_id": p.id,
        "registry_number": p.registry_number,
        "subject_before": _subject_before,
        "subject_after": p.subject,
        "items_added": items_added,
        "items_removed": items_removed,
        "items_changed": items_changed,
        "items_conflicted": items_conflicted,
        "items_kept_manual": items_kept_manual,
        "blocked_reason": None,
    }


async def _distribute_wish_to_purchases(wish, db, current_user, purchase_status: str = "plan_schedule", split: bool = True) -> list[int]:
    """Создаёт закупки (status='plan_schedule' — «План закупок») из позиций заявки по группам колонок,
    копирует позиции, добавляет участников и чаты, ставит purchase.wish_id.
    split=False (быстрое одобрение / полное согласование цепочкой): одна закупка со ВСЕМИ позициями,
    без разбиения по категориям — разбиение только при явном распределении через канбан.
    Возвращает список id созданных закупок. Транзакцию/commit НЕ делает — это на вызывающем.
    Защита от дублей: если по заявке уже есть закупки (purchases.wish_id == wish.id) — НИЧЕГО не создаёт и возвращает их id."""
    from sqlalchemy.orm import selectinload as sil
    from app.models.product import Product

    # Защита от дублей: если по заявке уже есть закупки — ничего не создаём,
    # но скрытые до одобрения (status='wishes') продвигаем в целевой статус
    existing = (await db.execute(
        select(Purchase).where(Purchase.wish_id == wish.id)
    )).scalars().all()
    if existing:
        # Авансовый: статус не меняем — закупка живёт в своём статусе независимо.
        # Purchase создаётся РАНЬШЕ самой заявки (purchases.py, ветка is_advance) —
        # обычный путь копирования WishItem → PurchaseItem сюда не доходит, поэтому
        # инвариант «закупки вне плана не бывает» (шаг 3, владелец 2026-08-07)
        # применяем здесь напрямую к PurchaseItem закупки, иначе авансовые отчёты
        # остались бы вне плана навсегда.
        if getattr(wish, 'source', None) == 'advance_report':
            adv_items_res = await db.execute(
                select(PurchaseItem).where(PurchaseItem.purchase_id.in_([p.id for p in existing]))
            )
            adv_items = adv_items_res.scalars().all()
            if adv_items:
                await _auto_assign_planned_items(
                    adv_items, wish.feo_category_id, db,
                    note=f"авансовым отчётом (заявка №{wish.id})",
                )
                await db.flush()
            return [p.id for p in existing]
        # W2-гейт: проверяем даты и категорию ФЭО ПЕРЕД продвижением скрытых закупок
        # в целевой статус — это тоже момент «попадания в План закупок» (владелец,
        # 2026-08-11): скрытые (status='wishes') закупки ещё не в плане.
        wishes_purchases = [p for p in existing if p.status == "wishes"]
        if wishes_purchases:
            items_res = await db.execute(select(WishItem).where(WishItem.wish_id == wish.id))
            items_for_gate = [it for it in items_res.scalars().all() if _is_meaningful_item(it)]
            await _ensure_feo_categories_assigned(wish, items_for_gate, db)
            await _ensure_needed_dates(wish, db, items_for_gate)
        for p in existing:
            if p.status == "wishes":
                p.status = purchase_status
        # Владелец (план crystalline-soaring-heron.md, п.2, 2026-08-21): «после
        # согласования заявку правят в закупке, а не в заявке» — верно, ПОКА
        # заявка остаётся converted (правка wish.items на этом статусе заблокирована,
        # см. гейт в update_wish). Но владелец наблюдал ОБРАТНОЕ на проде: заявку
        # ВОЗВРАЩАЛИ в черновик (там правка снова разрешена), меняли предмет и
        # состав, согласовывали ЗАНОВО — а закупка (РЕЕ-2026-00898) осталась со
        # старым предметом, потому что раньше эта ветка (существующая закупка —
        # "защита от дублей" сработала) молчала: только продвигала статус, ничего
        # не сверяла. Единственный путь СЮДА с изменившимися items — именно
        # «вернули в черновик → поправили → согласовали заново» (иначе items не
        # изменить), так что синхронизация здесь НЕ нарушает правило «после
        # согласования — только в закупке»: она реагирует именно на легитимную
        # правку, сделанную, пока заявка была НЕ converted.
        wish._purchase_sync = await _sync_purchase_from_wish(wish, existing, db)
        return [p.id for p in existing]

    # Preload wish items with products for category resolution
    res = await db.execute(
        select(WishItem)
        .options(sil(WishItem.product))
        .where(WishItem.wish_id == wish.id)
    )
    items_full = res.scalars().all()

    # Задача 3: строки-заготовки (без названия и без сумм/количества) — фронт
    # создаёт их заранее для будущего ввода — не должны попадать в План закупок.
    items_full = [it for it in items_full if _is_meaningful_item(it)]

    # Гейт ФЭО (владелец, 2026-08-11): позиция без категории (или со ссылкой на
    # удалённую из справочника) не должна попасть в План закупок — раньше эта же
    # ветка молча обнуляла битую ссылку и создавала закупку без ФЭО (предупреждением
    # postfactum через wish._convert_warning). Именно так реальная заявка №32
    # «Приобретение Пикапов» лишилась категории, а созданная из неё закупка
    # осталась сиротой вне всех планов ФЭО. См. docstring _ensure_feo_categories_assigned.
    await _ensure_feo_categories_assigned(wish, items_full, db)

    # Backfill product_id + category by item_name for legacy wish_items
    # (created before product_id was persisted on wish_items).
    missing = [it for it in items_full if not it.product_id and (it.item_name or "").strip()]
    name_to_product: dict[str, Product] = {}
    if missing:
        names = list({(it.item_name or "").strip() for it in missing})
        pres = await db.execute(select(Product).where(Product.name.in_(names)))
        for p in pres.scalars().all():
            name_to_product[(p.name or "").strip().lower()] = p
        for it in missing:
            hit = name_to_product.get((it.item_name or "").strip().lower())
            if hit:
                it.product_id = hit.id

    def _resolve_key(it: WishItem) -> str:
        """target_column_key → product.category → name-matched product.category → '__uncategorized__'"""
        if it.target_column_key:
            return it.target_column_key
        if it.product_id and it.product and it.product.category:
            return it.product.category
        hit = name_to_product.get((it.item_name or "").strip().lower())
        if hit and hit.category:
            return hit.category
        return "__uncategorized__"

    groups: dict[str, list] = {}
    if split:
        for it in items_full:
            groups.setdefault(_resolve_key(it), []).append(it)
    elif items_full:
        groups["__all__"] = list(items_full)

    if not groups:
        raise HTTPException(status_code=400, detail="Нет позиций для распределения")

    # W2: Гейт обязательности дат потребности при переносе в План закупок
    await _ensure_needed_dates(wish, db, items_full)

    # Инвариант «закупки вне плана не бывает» (владелец, 2026-08-07, план
    # zany-fluttering-mountain.md шаг 3): раньше автозаведение плановой позиции
    # срабатывало ТОЛЬКО для over_plan=true — обычная позиция без привязки
    # молча «съедала» план листа, не будучи привязанной ни к какой FeoPlannedItem
    # (отсюда осиротевшие строки в дереве ФЭО). Теперь — КАЖДАЯ позиция без явной
    # привязки (feo_planned_item_id ещё не проставлен — ни автоподбором, ни
    # пользователем) при переносе в План закупок порождает НОВУЮ FeoPlannedItem
    # (или находит существующую по точному совпадению имени — дедуп ниже) и
    # привязывается к ней; над_плановая надбавка (over_plan) сбрасывается — сама
    # позиция становится обычной плановой. Явный выбор пользователя (уже
    # проставленный feo_planned_item_id, в т.ч. подтверждённый матчинг) НЕ
    # перебивается — эти позиции сюда не попадают.
    # Дедуп по (категория, нормализованное имя trim+lower) — как в импорте Excel Ур.5
    # (feo_categories.py), но с нормализацией, т.к. источник — свободный ввод в заявке.
    # Эта ветка недостижима при повторном согласовании: закупки уже существуют, и
    # выполнение уходит в ветку `if existing:` выше, до сюда не доходя (идемпотентность).
    await _auto_assign_planned_items(
        items_full, wish.feo_category_id, db, note=f"заявкой №{wish.id}",
    )

    # Шаг 5 «ТЗ не выше плана» (владелец, 2026-08-07, план zany-fluttering-mountain.md):
    # _distribute_wish_to_purchases — это общий путь создания закупок из заявки
    # (approve_wish, force_wish_status('converted'), сам convert_wish воспроизводит
    # тот же порядок отдельно, см. его код), т.е. САМОЕ частое место, где ТЗ
    # превращается в закупку. Раньше здесь проверялось только превышение бюджета
    # ФЭО (assert_no_unapproved_excess ниже) — позиция дороже своей плановой строки
    # проходила без единого отказа. Порядок трёх шагов ОБЯЗАТЕЛЕН и именно такой:
    #   1) автозаведение (_auto_assign_planned_items выше) — ДО проверки цены,
    #      иначе у большинства позиций feo_planned_item_id ещё пуст и проверять
    #      не с чем (assert_tz_not_over_plan — no-op без плана, любая цена прошла бы);
    #   2) «ТЗ не выше плана» (здесь) — точечная причина отказа по КОНКРЕТНОЙ
    #      позиции (кол-во/цена/сумма), самая частная и понятная формулировка;
    #   3) превышение ФЭО (assert_no_unapproved_excess ниже) — по КАТЕГОРИИ
    #      целиком, уже ПОСЛЕ того как автозаведение нарастило план категории
    #      всеми новыми позициями; это самая общая причина отказа, ей место
    #      последней, чтобы пользователь сначала увидел, какая именно позиция
    #      виновата, а не абстрактное «превышен бюджет категории».
    # over_plan=true — сознательно сверх плана (тот же обход, что и в
    # _sync_wish_items_to_purchases/update_wish/purchases.py) — пропускаем.
    # Владелец (2026-08-17, прод-инцидент РЕЕ-2026-00887): позиции, привязанные
    # к ОДНОЙ и той же плановой позиции, накапливаются в пределах ЭТОЙ заявки
    # (assert_tz_batch_not_over_plan), а не проверяются поштучно против ПОЛНОГО
    # плана — иначе две строки в сумме превышают план, каждая проходя поодиночке.
    await assert_tz_batch_not_over_plan(db, items_full)

    # Владелец (2026-08-12): «Когда заявка идёт с превышением, ... она должна
    # уходить [в закупку], но на закупке должен быть значок, что она заблокирована
    # из-за превышения. Должна быть возможность передвижки, уменьшения». Раньше
    # здесь стоял assert_no_unapproved_excess — согласование заявки, создающее
    # закупки целиком отказывало 409, пока превышение (в т.ч. на ПРЕДКЕ категории)
    # не согласовано или не убрано, и никакого способа перенести/уменьшить позиции
    # ПОСЛЕ создания закупки не было. Теперь action не блокируется — вместо этого
    # ниже, ПОСЛЕ фактического создания закупок (когда план уже вырос), собираем
    # excess_warnings по каждой затронутой категории ФЭО (и её предкам) — см.
    # _collect_excess_warnings. Здесь только группируем позиции по листовой
    # категории (per-item feo_category_id, fallback — категория заявки целиком).
    _cat_items: dict[int, list[dict]] = {}
    for _wi in items_full:
        _cid = _wi.feo_category_id or wish.feo_category_id
        if _cid:
            _cat_items.setdefault(_cid, []).append({
                "name": _wi.item_name, "amount": float(_wi.total_price or 0),
            })

    # Контрагент заявки (владелец, 2026-08-17) — резолвим один раз на всю
    # конвертацию (не per-группу), чтобы название контрагента попало и в
    # PurchaseItem.contractor_name каждой позиции (mirrors purchases.py PUT,
    # см. её докстринг «Phase 26-Z» — тот же паттерн проставления контрагента
    # позициям без своего). Только заполняем — сами позиции только что созданы,
    # перетирать нечего.
    _wish_contractor_obj = None
    if getattr(wish, 'contractor_id', None):
        from app.models.contractor import Contractor as _Contractor
        _wish_contractor_obj = await db.get(_Contractor, wish.contractor_id)

    created_purchase_ids: list[int] = []
    for column_key, items_in_col in groups.items():
        total_nmck = sum(float(i.total_price or 0) for i in items_in_col)
        display_key = "Не определено" if column_key == "__uncategorized__" else column_key
        title = (wish.title or "").strip() or f"Заявка #{wish.id}"
        subject = title if column_key == "__all__" else f"{title} — {display_key}"
        total_qty_grp = sum(float(i.quantity or 0) for i in items_in_col)

        # W2: шапочная delivery_date — если все позиции группы имеют одну дату
        effective_dates = {_eff_date(wish, wi) for wi in items_in_col}
        effective_dates.discard(None)
        group_delivery_date = effective_dates.pop() if len(effective_dates) == 1 else None

        # C1: авансовый отчёт → фиксируем тип; обычная заявка → single
        _is_advance_wish = (getattr(wish, 'source', None) == 'advance_report')
        _purchase_method = 'advance' if _is_advance_wish else 'single'
        _payment_basis_type = 'advance_report' if _is_advance_wish else None
        p = Purchase(
            wish_id=wish.id,
            subsidy_id=wish.subsidy_id,
            feo_category_id=wish.feo_category_id,
            event_id=getattr(wish, 'event_id', None),  # «Мероприятие»
            item_name=title,
            subject=subject,
            planned_quantity=total_qty_grp or wish.quantity,
            planned_total_price=total_nmck,
            total_nmck=total_nmck,
            nmck=total_nmck,
            status=purchase_status,
            # B1: исполнитель = executor_id (без фолбэка на инициатора)
            assigned_user_id=getattr(wish, 'executor_id', None),
            # B1: служебка «на чьё имя» = assigned_to заявки
            service_note_to_user_id=wish.assigned_to,
            execution_term=getattr(wish, 'execution_deadline', None),  # B-exec
            service_note_text=wish.justification,
            service_note_by=wish.created_by,
            delivery_date=group_delivery_date,  # W2: единая дата для группы
            purchase_method=_purchase_method,
            payment_basis_type=_payment_basis_type,
            feo_per_item=bool(getattr(wish, 'feo_per_item', False)),
            vat_mode=(getattr(wish, 'vat_mode', None) or 'uniform'),
            # Контрагент заявки (владелец, 2026-08-17) — переезжает в закупку,
            # если указан. Purchase свежесозданный (contractor_id ещё пуст) —
            # «не перетирать уже заданное» тут выполняется автоматически.
            contractor_id=getattr(wish, 'contractor_id', None),
        )
        db.add(p)
        await db.flush()  # get p.id
        created_purchase_ids.append(p.id)

        for wi in items_in_col:
            pi = PurchaseItem(
                purchase_id=p.id,
                product_id=wi.product_id,
                item_name=wi.item_name,
                item_type=wi.item_type,
                quantity=wi.quantity,
                unit=wi.unit,
                unit_price=wi.unit_price,
                total_price=wi.total_price,
                # Снимок плана (Шаг 1 «план ≠ факт»): зафиксировать ТЗ заявки как план
                # позиции ОТДЕЛЬНО от unit_price/total_price (которые дальше могут
                # мутировать при правке цены по итогам закупки) — дерево ФЭО обязано
                # читать planned_*, а не текущую unit_price/total_price.
                planned_quantity=wi.quantity,
                planned_unit_price=wi.unit_price,
                planned_total=wi.total_price,
                country_origin=wi.country_origin,
                feo_category_id=wi.feo_category_id,  # B9: per-item feo
                feo_planned_item_id=wi.feo_planned_item_id,  # расходуем уже запланированную позицию, не задваиваем план
                over_plan=getattr(wi, 'over_plan', False),
                needed_date=_eff_date(wish, wi),  # W2: наследование эффективной даты
                wish_item_id=wi.id,  # W1: hard link to source WishItem
                vat_rate=getattr(wi, 'vat_rate', None),
                # Контрагент заявки — на каждую позицию (см. резолв _wish_contractor_obj
                # выше). Если контрагента в справочнике ещё нет — свободный ввод
                # contractor_name (второе поле у Wish, ровно для этого случая).
                contractor_id=(_wish_contractor_obj.id if _wish_contractor_obj else None),
                contractor_inn=(_wish_contractor_obj.inn if _wish_contractor_obj else None),
                contractor_name=(
                    _wish_contractor_obj.name if _wish_contractor_obj
                    else getattr(wish, 'contractor_name', None)
                ),
            )
            db.add(pi)
        await db.flush()

        # Add wish author as purchase member (viewer role) so they can see the purchase
        if wish.created_by and wish.created_by != current_user.id:
            db.add(PurchaseMember(
                purchase_id=p.id,
                user_id=wish.created_by,
                role="viewer",
                added_by_id=current_user.id,
                consent_pending=False,
            ))
        # Also add assigned_to as member if different from author and current_user
        if wish.assigned_to and wish.assigned_to not in (wish.created_by, current_user.id):
            db.add(PurchaseMember(
                purchase_id=p.id,
                user_id=wish.assigned_to,
                role="viewer",
                added_by_id=current_user.id,
                consent_pending=False,
            ))
        await db.flush()

        # Create chat room per purchase if there is an assignee different from current user
        if wish.assigned_to and wish.assigned_to != current_user.id:
            org_id = getattr(current_user, 'org_id', None) or wish.org_id
            await _create_assignment_chat_room(
                db, current_user.id, wish.assigned_to,
                org_id,
                f"Закупка: {p.subject}",
            )

    if created_purchase_ids and not wish.purchase_id:
        wish.purchase_id = created_purchase_ids[0]

    # После создания закупок (план уже вырос) — собрать предупреждения о
    # превышении вместо блокировки, см. комментарий у _cat_items выше.
    wish._excess_warnings = await _collect_excess_warnings(db, wish.subsidy_id, _cat_items)

    return created_purchase_ids


async def _build_wishes_query(
    q,
    current_user: User,
    db: AsyncSession,
    *,
    mine_only: bool = False,
    assigned_to_me: bool = False,
    subordinates_only: bool = False,
    creator_id: Optional[int] = None,
    assigned_to_id: Optional[int] = None,
    subsidy_id: Optional[int] = None,
    org_id: Optional[int] = None,
    account_org_id: Optional[int] = None,
    created_from: Optional[date] = None,
    created_to: Optional[date] = None,
    deadline_from: Optional[date] = None,
    deadline_to: Optional[date] = None,
):
    """Общая видимость + доп. фильтры для GET /wishes/ и GET /wishes/counts.

    Вынесено из list_wishes (сессия 2026-09-01, жалоба владельца на счётчики
    вкладок), чтобы список и счётчики видели РОВНО одни и те же заявки — иначе
    числа разъедутся между вкладкой и бейджем. НЕ накладывает статус-фильтр,
    order_by, offset/limit — это решает вызывающий (у /counts своя логика
    группировки, у списка — своя пагинация).
    """
    org_ids = get_org_filter(current_user)
    vis = await get_visible_subsidy_ids(current_user, db, "wishes")
    if vis is not None:
        # Two-level gate: rows WITH subsidy → gate by vis; rows WITHOUT subsidy → org gate
        null_branch = (
            and_(Wish.subsidy_id.is_(None), Wish.org_id.in_(org_ids))
            if org_ids is not None
            else Wish.subsidy_id.is_(None)
        )
        q = q.where(or_(Wish.subsidy_id.in_(vis), null_branch))
    elif org_ids is not None:
        # SaaS role but org restriction still applies (e.g. account_owner with org_id set)
        q = q.where(Wish.org_id.in_(org_ids))

    # Preload wish_ids where current user is a participant (for visibility extension)
    member_res = await db.execute(
        select(WishMember.wish_id).where(WishMember.user_id == current_user.id)
    )
    member_wish_ids = {r[0] for r in member_res.all()}

    if assigned_to_me:
        # Explicit shortcut: wishes where I am the designated approver
        # или я — согласующий из цепочки (WishApproval)
        from app.models.wish_approval import WishApproval
        appr_res = await db.execute(
            select(WishApproval.wish_id).where(WishApproval.user_id == current_user.id)
        )
        appr_wish_ids = {r[0] for r in appr_res.all()}
        if appr_wish_ids:
            q = q.where(or_(Wish.assigned_to == current_user.id, Wish.id.in_(appr_wish_ids)))
        else:
            q = q.where(Wish.assigned_to == current_user.id)
    elif mine_only or current_user.role == 'employee':
        # Employee always sees only own + wishes they are a participant in
        base_cond = Wish.created_by == current_user.id
        if member_wish_ids:
            q = q.where(or_(base_cond, Wish.id.in_(member_wish_ids)))
        else:
            q = q.where(base_cond)
    elif subordinates_only:
        # Phase 28: use unified visibility helper (covers SaaS bypass + hierarchy +
        # dept heads + managed orgs + UOA org_admin/manager)
        visible_uids = await get_visible_user_ids(current_user, db)
        if visible_uids is None:
            # SaaS role (superadmin/account_owner) → видят всё
            # (org filter уже применён выше). Дополнительных фильтров не нужно.
            pass
        else:
            # «Заявки сотрудников» = видимые подчинённые + сам пользователь
            # (руководитель — тоже сотрудник, свои заявки видит здесь же)
            sub_ids = set(visible_uids) | {current_user.id}
            q = q.where(Wish.created_by.in_(sub_ids))
    else:
        # Phase 28: unified visibility helper + member visibility
        clause = await build_visibility_clause(current_user, db, 'wish')
        if clause is not None:
            if member_wish_ids:
                q = q.where(or_(clause, Wish.id.in_(member_wish_ids)))
            else:
                q = q.where(clause)

    # Дополнительные фильтры (применяются после visibility — только сужают)
    if creator_id is not None:
        q = q.where(Wish.created_by == creator_id)
    if assigned_to_id is not None:
        q = q.where(Wish.assigned_to == assigned_to_id)
    if subsidy_id is not None:
        q = q.where(Wish.subsidy_id == subsidy_id)
    if org_id is not None:
        q = q.where(Wish.org_id == org_id)
    if account_org_id is not None:
        # Аккаунт = корневая орг + все её дочерние (root_org_id/parent_org_id)
        from app.models.organization import Organization
        acc_orgs = select(Organization.id).where(or_(
            Organization.id == account_org_id,
            Organization.root_org_id == account_org_id,
            Organization.parent_org_id == account_org_id,
        ))
        q = q.where(Wish.org_id.in_(acc_orgs))
    if created_from is not None:
        q = q.where(func.date(Wish.created_at) >= created_from)
    if created_to is not None:
        q = q.where(func.date(Wish.created_at) <= created_to)
    if deadline_from is not None:
        q = q.where(Wish.desired_date >= deadline_from)
    if deadline_to is not None:
        q = q.where(Wish.desired_date <= deadline_to)

    return q


@router.get("/counts")
async def wish_tab_counts(
    mine_only: bool = False,
    assigned_to_me: bool = False,
    subordinates_only: bool = False,
    creator_id: Optional[int] = None,
    assigned_to_id: Optional[int] = None,
    subsidy_id: Optional[int] = None,
    org_id: Optional[int] = None,
    account_org_id: Optional[int] = None,
    created_from: Optional[date] = None,
    created_to: Optional[date] = None,
    deadline_from: Optional[date] = None,
    deadline_to: Optional[date] = None,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Счётчики по вкладкам «Заявки на закупку» — с теми же параметрами
    видимости, что и GET /wishes/, но COUNT()-ом, а не выборкой записей.

    Жалоба владельца (сессия 2026-09-01): список обрезан limit=50, поэтому
    сравнивать «сколько заявок» по длине ответа списка нельзя в принципе —
    нужен отдельный точный счётчик. Вкладки накопительные (см.
    app/services/wish_tabs.py): «Отправленные» включает approved/rejected/
    converted, «Одобренные» включает converted и т.д. — один GROUP BY по
    точному статусу, накопление сумм в Python (без 6 отдельных COUNT-запросов).
    """
    q = await _build_wishes_query(
        select(Wish.status, func.count(Wish.id)),
        current_user, db,
        mine_only=mine_only, assigned_to_me=assigned_to_me, subordinates_only=subordinates_only,
        creator_id=creator_id, assigned_to_id=assigned_to_id, subsidy_id=subsidy_id,
        org_id=org_id, account_org_id=account_org_id,
        created_from=created_from, created_to=created_to,
        deadline_from=deadline_from, deadline_to=deadline_to,
    )
    q = q.group_by(Wish.status)
    result = await db.execute(q)
    exact_counts: dict[str, int] = {row_status: cnt for row_status, cnt in result.all()}

    return {
        tab: sum(exact_counts.get(s, 0) for s in statuses)
        for tab, statuses in WISH_TAB_STATUS_SETS.items()
    }


@router.get("/", response_model=list[WishOut])
async def list_wishes(
    status: Optional[str] = None,
    mine_only: bool = False,
    assigned_to_me: bool = False,
    subordinates_only: bool = False,
    creator_id: Optional[int] = None,
    assigned_to_id: Optional[int] = None,
    subsidy_id: Optional[int] = None,
    org_id: Optional[int] = None,
    account_org_id: Optional[int] = None,
    created_from: Optional[date] = None,
    created_to: Optional[date] = None,
    deadline_from: Optional[date] = None,
    deadline_to: Optional[date] = None,
    skip: int = 0,
    limit: int = Query(50, le=200),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List wishes with unified visibility (Phase 28 Bundle 2B).
    assigned_to_me=true: wishes where current user is the assignee.
    subordinates_only=true: wishes created by direct subordinates (not current user).
    mine_only=true / role==employee: show only own wishes (shortcut filters).

    status: имя вкладки ('draft'/'submitted'/'approved'/'rejected'/'converted'/'all')
    фильтрует НАКОПИТЕЛЬНО через wish_tab_statuses (сессия 2026-09-01, жалоба
    владельца) — 'submitted' включает approved/rejected/converted и т.д.
    Отсутствие параметра сохраняет ПРЕЖНЕЕ поведение (список без converted) —
    на него рассчитывают другие потребители эндпоинта (см. wish_tab_statuses
    докстринг и вызовы GET /wishes/ без status в useRiskScores.ts и
    WishesView.vue::loadWishes/loadIncoming).
    """
    q = select(Wish).options(
        selectinload(Wish.creator),
        selectinload(Wish.approver),
        selectinload(Wish.assignee),
        selectinload(Wish.subsidy),
        selectinload(Wish.event),
        selectinload(Wish.executor),
        selectinload(Wish.items),
        selectinload(Wish.stopped_by_user),
        selectinload(Wish.contractor),
        selectinload(Wish.rejected_by_user),
    )
    q = await _build_wishes_query(
        q, current_user, db,
        mine_only=mine_only, assigned_to_me=assigned_to_me, subordinates_only=subordinates_only,
        creator_id=creator_id, assigned_to_id=assigned_to_id, subsidy_id=subsidy_id,
        org_id=org_id, account_org_id=account_org_id,
        created_from=created_from, created_to=created_to,
        deadline_from=deadline_from, deadline_to=deadline_to,
    )

    if status and status != 'all':
        # Накопительная цепочка (владелец, 2026-09-01): именованная вкладка
        # включает всё, что когда-либо прошло через это состояние.
        q = q.where(Wish.status.in_(wish_tab_statuses(status)))
    elif status == 'all':
        # Явное «Все» — действительно всё, включая draft и converted.
        pass
    else:
        # status отсутствует вовсе: ПРЕЖНЕЕ поведение не меняем — по умолчанию
        # «Заявки» = в работе; распределённые (converted) живут в «Закупках».
        # Другие экраны (useRiskScores.ts, WishesView.vue myWishes/incoming)
        # зовут без status и рассчитывают именно на это.
        q = q.where(Wish.status != 'converted')
    q = q.order_by(Wish.created_at.desc()).offset(skip).limit(limit)
    result = await db.execute(q)
    wishes = result.scalars().all()

    # Phase 31: batch unseen_fields/unseen_changes_count (2 queries, no N+1) (D-05..D-09)
    wish_ids = [w.id for w in wishes]
    unseen_map: dict[int, list[str]] = {}
    try:
        from app.routers.entity_changes import get_unseen_map as _get_unseen_map
        unseen_map = await _get_unseen_map(db, 'wish', wish_ids, current_user.id)
    except Exception as _exc:
        import logging as _log
        _log.getLogger(__name__).warning("unseen wish map failed: %s", _exc)

    # «От кого»: имена участников (WishMember) батчем, без N+1
    members_map: dict[int, list[str]] = {}
    approvers_map: dict[int, list[str]] = {}
    purchases_map: dict[int, list[int]] = {}
    if wish_ids:
        mres = await db.execute(
            select(WishMember.wish_id, User.full_name, User.username)
            .join(User, User.id == WishMember.user_id)
            .where(WishMember.wish_id.in_(wish_ids))
        )
        for m_wid, m_fn, m_un in mres.all():
            members_map.setdefault(m_wid, []).append(m_fn or m_un)

        # «Кому»: цепочка согласующих по order_num (батчем, без N+1)
        from app.models.wish_approval import WishApproval
        ares = await db.execute(
            select(WishApproval.wish_id, User.full_name, User.username)
            .join(User, User.id == WishApproval.user_id)
            .where(WishApproval.wish_id.in_(wish_ids))
            .order_by(WishApproval.wish_id, WishApproval.order_num)
        )
        for a_wid, a_fn, a_un in ares.all():
            approvers_map.setdefault(a_wid, []).append(a_fn or a_un)

        # Конвертация разбивает заявку на несколько закупок — отдаём все id
        pres = await db.execute(
            select(Purchase.wish_id, Purchase.id)
            .where(Purchase.wish_id.in_(wish_ids))
            .order_by(Purchase.id)
        )
        for p_wid, p_id in pres.all():
            purchases_map.setdefault(p_wid, []).append(p_id)

    # Пункт 4 (владелец, 2026-08-13): сводка закупок (номер/статус/сумма) для меню
    # «Перейти в закупку» — батчем, тем же приёмом, что purchases_map выше.
    purchase_summaries_map = await _wish_purchase_summaries_map(wish_ids, db)

    # Владелец: столбец «сумма заявки» на листе /wishes — Σ total_price позиций
    # (WishItem), ОДНИМ агрегирующим запросом на всю страницу (без N+1 и без
    # загрузки всех позиций построчно).
    items_total_map: dict[int, Decimal] = {}
    if wish_ids:
        itres = await db.execute(
            select(WishItem.wish_id, func.coalesce(func.sum(WishItem.total_price), 0))
            .where(WishItem.wish_id.in_(wish_ids))
            .group_by(WishItem.wish_id)
        )
        items_total_map = {wid: total for wid, total in itres.all()}

    out_list = []
    for w in wishes:
        enriched = _enrich(w)
        enriched.member_names = members_map.get(w.id, [])
        enriched.approver_names = approvers_map.get(w.id, [])
        enriched.purchase_ids = purchases_map.get(w.id, [])
        enriched.purchases = purchase_summaries_map.get(w.id, [])
        enriched.items_total = items_total_map.get(w.id, Decimal("0"))
        _unseen = unseen_map.get(w.id, [])
        enriched.unseen_fields = _unseen
        enriched.unseen_changes_count = len(_unseen)
        out_list.append(enriched)
    return out_list


@router.get("/{wish_id}", response_model=WishOut)
async def get_wish(
    wish_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get single wish with items. Creator, assignee, or manager/admin of same org."""
    wish = await _load_wish(wish_id, db)
    org_ids = get_org_filter(current_user)
    if org_ids is not None and wish.org_id not in org_ids:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")
    if current_user.role == 'employee' and wish.created_by != current_user.id and wish.assigned_to != current_user.id:
        # Check if current user is a participant (wish member)
        member_res = await db.execute(
            select(WishMember).where(
                WishMember.wish_id == wish_id,
                WishMember.user_id == current_user.id,
            )
        )
        if member_res.scalar_one_or_none() is None:
            # …или согласующий из цепочки (вкладка «На согласование мне»)
            from app.models.wish_approval import WishApproval
            appr = await db.execute(
                select(WishApproval.id).where(
                    WishApproval.wish_id == wish_id,
                    WishApproval.user_id == current_user.id,
                ).limit(1)
            )
            if appr.scalar_one_or_none() is None:
                raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")

    enriched = _enrich(wish)
    mnames = (await db.execute(
        select(User.full_name, User.username)
        .join(WishMember, WishMember.user_id == User.id)
        .where(WishMember.wish_id == wish_id)
    )).all()
    enriched.member_names = [fn or un for fn, un in mnames]

    from app.models.wish_approval import WishApproval
    anames = (await db.execute(
        select(User.full_name, User.username)
        .join(WishApproval, WishApproval.user_id == User.id)
        .where(WishApproval.wish_id == wish_id)
        .order_by(WishApproval.order_num)
    )).all()
    enriched.approver_names = [fn or un for fn, un in anames]
    enriched.purchase_ids = (await db.execute(
        select(Purchase.id).where(Purchase.wish_id == wish_id).order_by(Purchase.id)
    )).scalars().all()
    # Пункт 4 (владелец, 2026-08-13): номер/статус/сумма для меню «Перейти в закупку».
    enriched.purchases = (await _wish_purchase_summaries_map([wish_id], db)).get(wish_id, [])
    # items уже загружены selectinload'ом в _load_wish — суммируем в памяти,
    # без доп. запроса (см. items_total в list_wishes для батч-версии).
    enriched.items_total = sum((it.total_price or Decimal("0")) for it in (wish.items or [])) or Decimal("0")

    # W1: contracted_locked — есть ли закупка в статусе Договор+
    # Правка владельца (2026-08-18): сообщение о блокировке должно называть
    # конкретную закупку и её реальную стадию, а не всегда «Договор» — см.
    # _wish_locked_descr. contracted_locked_reason — готовая строка для баннера
    # на фронте (frontend/src/views/WishesView.vue), None если не заблокировано.
    _locked_descr = await _wish_locked_descr(wish_id, db)
    enriched.contracted_locked = bool(_locked_descr)
    enriched.contracted_locked_reason = _locked_descr

    # W-diff: «двойник» каждой позиции в закупке — только карточка, не список
    await _attach_purchase_matches(wish, enriched, db)

    # Phase 31: unseen_fields for single wish GET (D-05..D-09)
    try:
        from app.routers.entity_changes import get_unseen_map as _get_unseen_map
        _unseen_single = await _get_unseen_map(db, 'wish', [wish_id], current_user.id)
        _unseen_fields = _unseen_single.get(wish_id, [])
        enriched.unseen_fields = _unseen_fields
        enriched.unseen_changes_count = len(_unseen_fields)
    except Exception as _exc:
        import logging as _log
        _log.getLogger(__name__).warning("unseen wish single failed: %s", _exc)

    return enriched


@router.post("/{wish_id}/copy", response_model=WishOut, status_code=201)
async def copy_wish(
    wish_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Копирование заявки (владелец, 2026-08-13): «при переделывании больших
    заявок не приходилось делать двойную работу — могут быть однотипные заявки
    с небольшим расхождением». Создаёт НОВУЮ заявку-черновик.

    Копируется ТОЛЬКО то, что описывает сам товар — item_name/item_type/
    product_id/quantity/unit/country_origin по каждой позиции — и заголовок
    исходной заявки с пометкой «(копия)».

    НЕ копируется (владелец прямо просил дозаполнять и проверять, «куда что
    идёт»): категории ФЭО у позиций (feo_category_id), привязка к плановым
    позициям (feo_planned_item_id/match_confirmed), НДС-ставка и over_plan,
    цены (unit_price/total_price), даты потребности (needed_date), исполнитель,
    согласующие, статус, привязки к закупкам, вложения. Сама новая заявка
    всегда создаётся в статусе 'draft' (как в create_wish), автор — текущий
    пользователь.

    org_id и subsidy_id копируются «как есть» — это рамка (организация и
    бюджет), в которой человек работает, а не «куда что идёт» внутри неё.

    Права: та же проверка видимости, что в GET /{wish_id} (без доп. ролевых
    ограничений сверх неё — копировать может любой, кто видит заявку).
    """
    wish = await _load_wish(wish_id, db)
    org_ids = get_org_filter(current_user)
    if org_ids is not None and wish.org_id not in org_ids:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")
    if current_user.role == 'employee' and wish.created_by != current_user.id and wish.assigned_to != current_user.id:
        member_res = await db.execute(
            select(WishMember).where(
                WishMember.wish_id == wish_id,
                WishMember.user_id == current_user.id,
            )
        )
        if member_res.scalar_one_or_none() is None:
            from app.models.wish_approval import WishApproval
            appr = await db.execute(
                select(WishApproval.id).where(
                    WishApproval.wish_id == wish_id,
                    WishApproval.user_id == current_user.id,
                ).limit(1)
            )
            if appr.scalar_one_or_none() is None:
                raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")

    from app.schemas.wishes import _clamp_title
    new_title = _clamp_title(f"{wish.title} (копия)")

    new_wish = Wish(
        org_id=wish.org_id,
        title=new_title,
        subsidy_id=wish.subsidy_id,
        status="draft",
        created_by=current_user.id,
    )
    db.add(new_wish)
    await db.flush()

    for it in (wish.items or []):
        db.add(WishItem(
            wish_id=new_wish.id,
            product_id=it.product_id,
            item_name=it.item_name,
            item_type=it.item_type,
            quantity=it.quantity,
            unit=it.unit,
            country_origin=it.country_origin,
        ))

    await db.commit()
    new_wish = await _load_wish(new_wish.id, db)
    enriched = _enrich(new_wish)
    enriched.items_total = sum((it.total_price or Decimal("0")) for it in (new_wish.items or [])) or Decimal("0")
    return enriched


@router.post("/{wish_id}/stop", response_model=WishOut)
async def stop_wish(
    wish_id: int,
    body: WishStop,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Остановка заявки (владелец, 2026-08-13): «Останавливать могут все» — доступно
    ЛЮБОМУ пользователю, видящему заявку, без ролевых ограничений (та же проверка
    видимости, что и GET /{wish_id}).

    Заявка НЕ удаляется — проставляются stopped_at/stopped_by/stopped_reason,
    и она исчезает из плана закупок (см. app/services/feo_plan.py — агрегаты
    исключают Purchase.stopped_at IS NOT NULL / Wish.stopped_at IS NOT NULL).

    Останавливаются все привязанные закупки, ещё НЕ дошедшие до договора:
    обычные — раньше 'contracted' по STATUS_ORDER; рамочные (purchase_contract_type
    в FRAMEWORK_TYPES) — раньше 'ordered' (у рамочного договора отдельная граница,
    т.к. сам договор — отдельная сущность, не по каждой закупке). Если остановились
    не все закупки заявки — wish.stopped_partial=True.

    Исполнителю (Wish.executor_id) уходит уведомление. Обратной операции
    (возобновление) нет — владелец её не просил.
    """
    from app.routers.purchases import STATUS_ORDER
    from app.routers.purchase_budget import FRAMEWORK_TYPES

    wish = await _load_wish(wish_id, db)

    # Видимость — та же проверка, что и в GET /{wish_id} (владелец: «Останавливать
    # могут все» = любой, кто вообще видит заявку, без доп. ролевых ограничений).
    org_ids = get_org_filter(current_user)
    if org_ids is not None and wish.org_id not in org_ids:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")
    if current_user.role == 'employee' and wish.created_by != current_user.id and wish.assigned_to != current_user.id:
        member_res = await db.execute(
            select(WishMember).where(
                WishMember.wish_id == wish_id,
                WishMember.user_id == current_user.id,
            )
        )
        if member_res.scalar_one_or_none() is None:
            from app.models.wish_approval import WishApproval
            appr = await db.execute(
                select(WishApproval.id).where(
                    WishApproval.wish_id == wish_id,
                    WishApproval.user_id == current_user.id,
                ).limit(1)
            )
            if appr.scalar_one_or_none() is None:
                raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")

    if wish.stopped_at is not None:
        raise HTTPException(status_code=409, detail="Заявка уже остановлена")

    from datetime import timezone as _timezone
    now = datetime.now(_timezone.utc)
    wish.stopped_at = now
    wish.stopped_by = current_user.id
    wish.stopped_reason = (body.reason if body else None) or None

    purchases = await _wish_linked_purchases(wish_id, db)
    stopped_count = 0
    for p in purchases:
        if p.stopped_at is not None:
            continue  # уже остановлена ранее (другой заявкой/повторный вызов)
        is_framework = p.purchase_contract_type in FRAMEWORK_TYPES
        threshold_status = "ordered" if is_framework else "contracted"
        cur_idx = STATUS_ORDER.index(p.status) if p.status in STATUS_ORDER else 0
        threshold_idx = STATUS_ORDER.index(threshold_status)
        if cur_idx < threshold_idx:
            p.stopped_at = now
            p.stopped_by = current_user.id
            p.stopped_wish_id = wish.id
            stopped_count += 1

    total = len(purchases)
    partial = total > 0 and stopped_count < total
    wish.stopped_partial = partial

    await db.commit()

    try:
        from app.notifications import notify_wish_stopped
        wish_for_notify = await _load_wish(wish_id, db)
        executor = getattr(wish_for_notify, 'executor', None)
        if executor and executor.id != current_user.id:
            stopper_name = getattr(current_user, 'full_name', None) or current_user.username
            await notify_wish_stopped(wish_for_notify, executor, stopper_name, partial)
    except Exception as _exc:
        import logging as _log
        _log.getLogger(__name__).warning("notify wish stopped failed: %s", _exc)

    wish = await _load_wish(wish_id, db)
    return _enrich(wish)


@router.post("/", response_model=WishOut, status_code=201)
async def create_wish(
    body: WishCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Create a new wish (all roles)."""
    from app.services.subsidy_draft_guard import assert_subsidy_approved_for_binding
    await assert_subsidy_approved_for_binding(db, body.subsidy_id)

    org_ids = get_org_filter(current_user)
    org_id = org_ids[0] if org_ids else current_user.org_id

    wish = Wish(
        org_id=org_id,
        title=body.title,
        category=body.category,
        description=body.description,
        quantity=body.quantity,
        unit=body.unit,
        estimated_price=body.estimated_price,
        link=body.link,
        priority=body.priority,
        desired_date=body.desired_date,
        justification=body.justification,
        subsidy_id=body.subsidy_id,
        feo_category_id=body.feo_category_id,
        event_id=body.event_id,
        assigned_to=body.assigned_to,
        status="draft",
        created_by=current_user.id,
        feo_per_item=body.feo_per_item,
        vat_mode=body.vat_mode or 'uniform',
        contractor_id=body.contractor_id,
        contractor_name=body.contractor_name,
    )
    db.add(wish)
    await db.flush()

    if body.items:
        for item_data in body.items:
            if not _is_meaningful_item(item_data):
                continue  # пустая строка-заготовка — в БД не пишем
            wi = WishItem(
                wish_id=wish.id,
                product_id=item_data.get('product_id'),
                item_name=item_data.get('item_name', ''),
                item_type=item_data.get('item_type', 'товар'),
                quantity=item_data.get('quantity', 1),
                unit=item_data.get('unit', 'шт'),
                unit_price=item_data.get('unit_price', 0),
                total_price=item_data.get('total_price', 0),
                country_origin=item_data.get('country_origin', 'РФ'),
                feo_category_id=item_data.get('feo_category_id'),  # B9
                feo_planned_item_id=item_data.get('feo_planned_item_id'),  # привязка к плановой позиции плана закупок
                over_plan=item_data.get('over_plan', False),
                needed_date=_as_date(item_data.get('needed_date')),  # W2
                vat_rate=item_data.get('vat_rate'),
            )
            db.add(wi)
        await db.flush()

    await db.commit()
    wish = await _load_wish(wish.id, db)
    return _enrich(wish)


@router.put("/{wish_id}", response_model=WishOut)
async def update_wish(
    wish_id: int,
    body: WishUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update a wish.
    Draft/rejected: full edit.
    Submitted/approved/converted: edit allowed (for all non-SaaS if not contracted-locked),
    then resets approval chain and re-notifies approvers.
    B3: contracted-locked wishes (purchase in Договор+) are read-only.
    """
    wish = await _load_wish(wish_id, db)

    if not _is_saas(current_user) and wish.created_by != current_user.id:
        # Участники заявки (WishMember) тоже могут её редактировать
        if not await _is_wish_member(wish_id, current_user.id, db):
            # Дефект 1 (владелец, 2026-08-20): согласующий из цепочки WishApproval
            # сюда попадать не должен вовсе — фронт больше не зовёт PUT для него,
            # ФЭО он правит через PATCH /execution. Но если это всё-таки произошло
            # (старый фронт/прямой запрос) — подсказать конкретно, а не молчать
            # общей фразой (feedback_explain_blocking_reason).
            from app.models.wish_approval import WishApproval
            is_chain_approver = (await db.execute(
                select(WishApproval.id).where(
                    WishApproval.wish_id == wish_id,
                    WishApproval.user_id == current_user.id,
                ).limit(1)
            )).scalar_one_or_none() is not None
            detail = "Редактировать заявку может автор или участник заявки"
            if is_chain_approver:
                detail += ". Состав заявки меняет автор — как согласующий, правьте категорию ФЭО прямо в форме (сохраняется автоматически), решение отправляйте кнопкой «Согласовать»/«Отклонить»"
            raise HTTPException(status_code=403, detail=detail)

    # W2: contracted-lock replaces old draft-only gate
    if not _is_saas(current_user):
        _locked_descr = await _wish_locked_descr(wish.id, db)
        if _locked_descr:
            raise HTTPException(
                status_code=409,
                detail=f"Заявка привязана к закупке — {_locked_descr}. Редактирование запрещено",
            )
        # SaaS is always allowed; for others allow draft/rejected AND submitted/approved/converted

    # Владелец (2026-08-13): «После согласования заявку править нельзя, только в
    # закупке» — на статусе 'converted' заявка уже в Плане закупок, единственный
    # источник правды для позиций — сама закупка (PurchaseItem), а не WishItem.
    # Раньше правка позиций тут ЧАСТИЧНО синхронизировалась в закупку
    # (_sync_wish_items_to_purchases копировал название/кол-во/цену, но НЕ
    # категорию ФЭО) — данные расходились: в заявке позиция в одной категории,
    # в закупке в другой. Теперь правка позиций конвертированной заявки
    # отклоняется явно; несущественные поля заявки (приоритет, срок и т.п.)
    # по-прежнему редактируемы — блокируются только `items`.
    if not _is_saas(current_user) and wish.status == "converted" and body.items is not None:
        _linked_purchases = await _wish_linked_purchases(wish_id, db)
        _purchase_nums = ", ".join(
            f"№{p.purchase_number or p.id}" for p in _linked_purchases
        ) or "не найдена"
        raise HTTPException(
            status_code=409,
            detail=(
                "Заявка уже в плане закупок — изменения вносятся в закупке "
                f"({_purchase_nums})"
            ),
        )

    # Capture old status BEFORE mutation
    old_status = wish.status

    # Phase 31: capture old values for diff-tracking BEFORE any mutation (D-05..D-09)
    _old_wish_values = {f: getattr(wish, f, None) for f in WISH_TRACKED_FIELDS}

    # A1 fix: снимок существенных скалярных полей ДО мутации (Дыра 2)
    # Жалоба владельца 2026-08-13: привязка позиции к категории/плановой позиции
    # ФЭО — это маршрутизация по бюджету, а не смена предмета закупки. Согласующий
    # привязывал позицию к ФЭО и нажимал «Сохранить» уже ПОСЛЕ того, как согласовал —
    # feo_category_id в этом наборе стирал его же согласование («повторное
    # согласование» без реального изменения того, ЧТО закупается). За перерасход
    # отвечают отдельные гейты (assert_no_unapproved_excess / потолок субсидии),
    # поэтому feo_category_id сюда не входит.
    _APPROVAL_SENSITIVE_FIELDS: set[str] = {
        "title", "subsidy_id", "estimated_price",
        "quantity", "unit", "justification",
    }
    _old_sensitive = {f: getattr(wish, f, None) for f in _APPROVAL_SENSITIVE_FIELDS}

    # A1 fix: снимок позиций ДО мутации для точного сравнения (Дыра 1)
    # feo_category_id намеренно НЕ входит в кортеж сравнения — см. комментарий
    # у _APPROVAL_SENSITIVE_FIELDS выше (жалоба 2026-08-13).
    _old_items = {
        wi.id: (
            str(wi.item_name or ''),
            str(wi.unit or ''),
            float(wi.quantity or 0),
            float(wi.unit_price or 0),
            float(wi.total_price or 0),
        )
        for wi in (wish.items or [])
    }

    if body.subsidy_id is not None:
        from app.services.subsidy_draft_guard import assert_subsidy_approved_for_binding
        await assert_subsidy_approved_for_binding(db, body.subsidy_id)

    update_data = body.model_dump(exclude_none=True, exclude={'items'})
    for field, value in update_data.items():
        setattr(wish, field, value)

    # Контрагент — необязательное поле (владелец, 2026-08-17), а «необязательное»
    # значит, что его можно и СНЯТЬ, не только поставить. Общий model_dump(exclude_none=True)
    # выше НЕ трогаем скопом: на нём годами держится поведение остальных Optional-полей
    # формы заявки («ключ отсутствует/null» = «не менять», привычно для частичных PUT —
    # например автосохранение одного поля не должно стирать все остальные) — массовая
    # замена на exclude_unset рисковала бы молча обнулить что-то, что никто не просил
    # чистить. Поэтому точечно, только для contractor_id/contractor_name: используем
    # body.model_fields_set — Pydantic v2 кладёт туда имя поля, если СООТВЕТСТВУЮЩИЙ КЛЮЧ
    # реально присутствовал в JSON тела запроса (даже если значение — null), и не кладёт,
    # если ключ отсутствовал вовсе. Так «прислали null» (пользователь снял контрагента)
    # отличается от «ключ не прислали» (например, автосохранение другого поля заявки —
    # значение контрагента трогать не должно).
    if 'contractor_id' in body.model_fields_set:
        wish.contractor_id = body.contractor_id
    if 'contractor_name' in body.model_fields_set:
        wish.contractor_name = body.contractor_name

    # Плановые позиции следуют за сменой категории (владелец, 2026-08-17):
    # предупреждения, когда привязку пришлось снять вместо переезда (см. ветку
    # non-draft ниже) — возвращаются в ответе, не проглатываются молча.
    _plan_transfer_warnings: list[str] = []

    if body.items is not None:
        if old_status == "draft":
            # Draft: delete+recreate (original behaviour)
            await db.execute(delete(WishItem).where(WishItem.wish_id == wish.id))
            for item_data in body.items:
                if not _is_meaningful_item(item_data):
                    continue  # пустая строка-заготовка — в БД не пишем
                wi = WishItem(
                    wish_id=wish.id,
                    product_id=item_data.get('product_id'),
                    item_name=item_data.get('item_name', ''),
                    item_type=item_data.get('item_type', 'товар'),
                    quantity=item_data.get('quantity', 1),
                    unit=item_data.get('unit', 'шт'),
                    unit_price=item_data.get('unit_price', 0),
                    total_price=item_data.get('total_price', 0),
                    country_origin=item_data.get('country_origin', 'РФ'),
                    feo_category_id=item_data.get('feo_category_id'),  # B9
                    feo_planned_item_id=item_data.get('feo_planned_item_id'),  # план закупок ФЭО
                    over_plan=item_data.get('over_plan', False),
                    needed_date=_as_date(item_data.get('needed_date')),  # W2
                    vat_rate=item_data.get('vat_rate'),
                )
                db.add(wi)
            await db.flush()
        else:
            # Non-draft (submitted/approved/converted/rejected): update existing items in-place by id,
            # and DELETE items that dropped out of the payload (owner removes a duplicate/empty row →
            # it must actually disappear from the DB, not just from the UI). Add is still NOT
            # supported here — out of scope, unchanged from prior behaviour (items without a
            # matching id are silently skipped).
            existing_items = {wi.id: wi for wi in wish.items}
            payload_ids: set[int] = set()
            for item_data in body.items:
                item_id = item_data.get('id') if isinstance(item_data, dict) else getattr(item_data, 'id', None)
                if item_id:
                    payload_ids.add(item_id)
                wi = existing_items.get(item_id) if item_id else None
                if wi is None:
                    continue
                if 'item_name' in item_data:
                    wi.item_name = item_data['item_name']
                if 'unit_price' in item_data:
                    wi.unit_price = item_data['unit_price']
                if 'quantity' in item_data:
                    wi.quantity = item_data['quantity']
                if 'unit' in item_data:
                    wi.unit = item_data['unit']
                if 'total_price' in item_data:
                    wi.total_price = item_data['total_price']
                elif 'unit_price' in item_data or 'quantity' in item_data:
                    wi.total_price = (wi.unit_price or 0) * (wi.quantity or 0)
                _wi_cat_changing = (
                    'feo_category_id' in item_data
                    and item_data['feo_category_id'] != wi.feo_category_id
                )
                if 'feo_category_id' in item_data:
                    wi.feo_category_id = item_data['feo_category_id']
                if 'feo_planned_item_id' in item_data:
                    # Явный выбор плановой позиции с фронта (в т.ч. её очистка) —
                    # доверяем как есть, автоперенос ниже не запускаем.
                    wi.feo_planned_item_id = item_data['feo_planned_item_id']
                elif _wi_cat_changing and wi.feo_planned_item_id is not None:
                    # Плановые позиции следуют за сменой категории (владелец,
                    # 2026-08-17): payload сменил feo_category_id позиции, но не
                    # прислал новый feo_planned_item_id явно — старая привязка
                    # осталась указывать на прежнюю категорию. Если позиция —
                    # единственный владелец своей плановой строки, строка
                    # переезжает вместе с ней; если план общий с другими
                    # закупками/заявками — привязка снимается и предупреждение
                    # копится в _plan_transfer_warnings (возвращается в ответе).
                    if wi.feo_category_id is not None:
                        _w = await _move_or_detach_planned_item(db, wi, wi.feo_category_id)
                        if _w:
                            _plan_transfer_warnings.append(_w)
                    else:
                        # Категория очищена целиком — плановой позиции
                        # переезжать некуда, привязка просто снимается.
                        from app.models.feo_planned_item import FeoPlannedItem as _FPI2
                        from app.services.plan_autoassign import deactivate_if_orphaned as _deactivate_if_orphaned
                        _old_fpi2 = await db.get(_FPI2, wi.feo_planned_item_id)
                        wi.feo_planned_item_id = None
                        wi.over_plan = False
                        await _deactivate_if_orphaned(db, _old_fpi2)
                if 'over_plan' in item_data:
                    wi.over_plan = item_data['over_plan']
                if 'needed_date' in item_data:
                    wi.needed_date = _as_date(item_data['needed_date'])
                if 'vat_rate' in item_data:
                    wi.vat_rate = item_data['vat_rate']

                # Шаг 5 «цена ТЗ не выше плановой» (владелец, 2026-08-07): правка
                # заявки — тот же обход, что и _sync_wish_items_to_purchases выше,
                # проверяем ЗДЕСЬ (на WishItem), т.к. именно отсюда цена/кол-во
                # затем копируются в PurchaseItem. over_plan=true пропускаем —
                # сознательно сверх плана.
                if ('unit_price' in item_data or 'quantity' in item_data) and not getattr(wi, 'over_plan', False):
                    await assert_tz_not_over_plan(
                        db,
                        feo_planned_item_id=wi.feo_planned_item_id,
                        feo_category_id=wi.feo_category_id,
                        quantity=wi.quantity,
                        unit_price=wi.unit_price,
                        total_price=wi.total_price,
                        item_name=wi.item_name,
                    )

            # Позиции, которых больше нет в payload, — удаляем физически.
            ids_to_delete = set(existing_items.keys()) - payload_ids
            if ids_to_delete:
                _locked_descr = await _wish_locked_descr(wish.id, db)
                if _locked_descr:
                    raise HTTPException(
                        status_code=409,
                        detail=(
                            f"Нельзя удалить позицию заявки: закупка {_locked_descr} уже на этапе "
                            "договора. Отмените закупку вручную в реестре закупок, либо не удаляйте позицию."
                        ),
                    )
                # Сироты в закупке: PurchaseItem с FK на удаляемую позицию (ondelete=SET NULL
                # только обнулит связь, но НЕ уберёт строку) — чистим явно, но только пока
                # закупка ещё не ушла дальше «Плана закупок» (иначе выше сработал бы гейт).
                _linked_purchases = await _wish_linked_purchases(wish.id, db)
                _early_stage_purchase_ids = {
                    p.id for p in _linked_purchases if p.status in ("plan_schedule", "wishes")
                }
                if _early_stage_purchase_ids:
                    await db.execute(
                        delete(PurchaseItem).where(
                            PurchaseItem.wish_item_id.in_(ids_to_delete),
                            PurchaseItem.purchase_id.in_(_early_stage_purchase_ids),
                        )
                    )
                # Удаляем через саму relationship-коллекцию (cascade="all, delete-orphan" на
                # Wish.items), а не Core delete(): так wish.items остаётся синхронной в памяти
                # для последующего сравнения _old_items/_new_items (сброс согласования ниже).
                for _del_id in ids_to_delete:
                    wish.items.remove(existing_items[_del_id])

            await db.flush()

    # W2: re-approval trigger for submitted/approved/converted wishes
    # A1 fix: сбрасываем согласование ТОЛЬКО при изменении существенных полей.
    # Существенные поля = те, что определяют предмет закупки и требуют повторного
    # одобрения согласующих. Несущественные (priority, desired_date, event_id,
    # assigned_to, executor_id, execution_deadline, link) НЕ сбрасывают цепочку.
    if old_status in ("submitted", "approved", "converted"):
        # Дыра 2 исправлена: используем _old_sensitive (снят ДО мутации) вместо
        # _old_wish_values (который содержит только WISH_TRACKED_FIELDS, без quantity/unit/justification).
        _sensitive_changed = any(
            str(_old_sensitive.get(f)) != str(getattr(wish, f, None))
            for f in _APPROVAL_SENSITIVE_FIELDS
        )
        # Дыра 1 исправлена: сравниваем реальное содержимое позиций до/после.
        # Фронт всегда шлёт body.items, поэтому «items is not None» — недостаточное условие.
        # feo_category_id намеренно НЕ входит — привязка к ФЭО/плановой позиции не
        # меняет предмет закупки, это маршрутизация по бюджету (жалоба 2026-08-13).
        _new_items = {
            wi.id: (
                str(wi.item_name or ''),
                str(wi.unit or ''),
                float(wi.quantity or 0),
                float(wi.unit_price or 0),
                float(wi.total_price or 0),
            )
            for wi in (wish.items or [])
        }
        _items_changed = (_old_items != _new_items)

        if _sensitive_changed or _items_changed:
            # Проверяем наличие согласующих ДО смены статуса (старые заявки могли быть одобрены без цепочки)
            from app.models.wish_approval import WishApproval as _WA
            _approver_count = (await db.execute(
                select(func.count()).select_from(_WA).where(_WA.wish_id == wish.id)
            )).scalar() or 0
            if _approver_count == 0:
                raise HTTPException(
                    status_code=409,
                    detail="Заявка уйдёт на повторное согласование, но согласующие не выбраны — "
                           "добавьте хотя бы одного согласующего в разделе «Согласующие».",
                )
            # W2: проверяем плановые даты перед сбросом в submitted (авансовые пропускаем)
            if getattr(wish, 'source', None) != 'advance_report':
                await _ensure_needed_dates(wish, db, wish.items or [], context="submit")
            wish.status = "submitted"
            # Повторная отправка на согласование после правки — старое
            # отклонение больше не актуально (владелец, 2026-08-19).
            wish.rejected_by = None
            wish.rejected_at = None
            wish.rejection_reason = None
            await db.flush()
            await _reset_approvals(wish.id, db)
            # Заявка уходит на ПОВТОРНОЕ согласование — позиции могли измениться,
            # закупка в Плане закупок больше не актуальна, убираем её оттуда. Если что-то
            # уже в работе/договоре — откат запрещаем явно (не молчим, коммита ещё не было).
            await _withdraw_wish_from_plan(wish.id, db, action_text="вернуть на повторное согласование")
            requester_name = getattr(current_user, 'full_name', None) or current_user.username
            await _notify_pending_approvers(wish, db, requester_name)

    await db.commit()

    # Phase 31: record EntityChange for each TRACKED_FIELD that changed (D-05..D-09)
    # Only record changes made by OTHER users (D-07: own changes are not highlighted)
    try:
        from app.models.entity_change import EntityChange as _EC
        _changes = []
        for _fname in WISH_TRACKED_FIELDS:
            _old = _old_wish_values.get(_fname)
            _new = getattr(wish, _fname, None)
            _old_s = str(_old) if _old is not None else None
            _new_s = str(_new) if _new is not None else None
            if _old_s != _new_s:
                _changes.append(_EC(
                    entity_type='wish',
                    entity_id=wish.id,
                    field_name=_fname,
                    old_value=_old_s,
                    new_value=_new_s,
                    changed_by_id=current_user.id,
                    changed_by_name=getattr(current_user, 'full_name', None) or current_user.username,
                ))
        if _changes:
            for _c in _changes:
                db.add(_c)
            await db.commit()
    except Exception as _exc:
        import logging as _log
        _log.getLogger(__name__).warning("entity_change record failed for wish: %s", _exc)

    wish = await _load_wish(wish_id, db)
    out = _enrich(wish)
    out.plan_transfer_warnings = _plan_transfer_warnings
    return out


@router.post("/{wish_id}/submit", response_model=WishOut)
async def submit_wish(
    wish_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Submit a draft wish for approval (creator only, draft/rejected -> submitted)."""
    wish = await _load_wish(wish_id, db)

    if not _is_saas(current_user) and wish.created_by != current_user.id:
        # Участники заявки (WishMember) тоже могут отправить её на согласование
        if not await _is_wish_member(wish_id, current_user.id, db):
            raise HTTPException(status_code=403, detail="Отправить на согласование может автор или участник заявки")
    if not _is_saas(current_user) and wish.status not in ("draft", "rejected"):
        raise HTTPException(status_code=400, detail="Заявка должна быть в статусе 'draft' или 'rejected'")

    # Проверяем наличие согласующих: нельзя отправить заявку неизвестно кому
    from app.models.wish_approval import WishApproval as _WA_submit
    _approver_count = (await db.execute(
        select(func.count()).select_from(_WA_submit).where(_WA_submit.wish_id == wish_id)
    )).scalar() or 0
    if _approver_count == 0:
        raise HTTPException(
            status_code=409,
            detail="Нельзя отправить на согласование: не выбраны согласующие. "
                   "Добавьте хотя бы одного согласующего в разделе «Согласующие».",
        )

    # W2: проверяем плановые даты (авансовые пропускаем) — ДО смены статуса
    if getattr(wish, 'source', None) != 'advance_report':
        await _ensure_needed_dates(wish, db, wish.items or [], context="submit")

    _was_rejected = wish.status == "rejected"
    wish.status = "submitted"
    # Повторная отправка на согласование (в т.ч. из 'rejected') — старое
    # отклонение больше не актуально, владелец просил не оставлять его висеть
    # (см. Wish.rejected_by докстринг в models/wish.py).
    wish.rejected_by = None
    wish.rejected_at = None
    wish.rejection_reason = None
    await db.flush()
    if _was_rejected:
        # Без keep_user_id — сброс ПОЛНЫЙ. Иначе строка отклонившего согласующего
        # (см. _reset_approvals в decide()) осталась бы 'rejected' навсегда: при
        # повторном дохождении очереди до неё sequential-гейт увидел бы «уже
        # принято решение» и заблокировал бы согласование заново.
        await _reset_approvals(wish.id, db)

    # Уведомить согласующих из цепочки (вынесено в _notify_pending_approvers)
    requester_name = current_user.full_name or current_user.username
    await _notify_pending_approvers(wish, db, requester_name)

    # Notify approver
    if wish.assigned_to and wish.assigned_to != current_user.id:
        org_id = getattr(current_user, 'org_id', None) or wish.org_id
        room_id = await _create_assignment_chat_room(
            db, current_user.id, wish.assigned_to,
            org_id,
            f"Заявка №{wish.id}: {wish.title or 'без названия'}",
        )
        db.add(ChatMessage(
            room_id=room_id,
            sender_id=current_user.id,
            content=f"📋 Заявка отправлена на согласование: {wish.title or '(без названия)'}",
        ))
        await db.flush()

    await db.commit()
    wish = await _load_wish(wish_id, db)
    return _enrich(wish)


@router.post("/{wish_id}/approve", response_model=WishOut)
async def approve_wish(
    wish_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Approve a submitted wish (manager+ roles OR assigned approver, submitted -> approved)."""
    wish = await _load_wish(wish_id, db)

    if not _is_saas(current_user) and wish.status != "submitted":
        raise HTTPException(status_code=400, detail="Заявка должна быть в статусе 'submitted'")

    # Org isolation
    org_ids = get_org_filter(current_user)
    if org_ids is not None and wish.org_id not in org_ids:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")

    if not _is_saas(current_user) and current_user.role not in MANAGER_ROLES and wish.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="Одобрить заявку может менеджер+ или назначенный согласующий")

    # W3: менеджер/SaaS может одобрить чужие pending без блокировки (allow_override=True)
    await _ensure_no_pending_approvals(wish, db, current_user, allow_override=True)

    wish.status = "approved"
    wish.approved_by = current_user.id
    # Согласованная заявка автоматически уходит в «План закупок» ОДНОЙ закупкой
    # (быстрое одобрение — без разбиения по категориям), сумма попадает в план ФЭО.
    created_ids: list[int] = []
    if wish.items:
        created_ids = await _distribute_wish_to_purchases(wish, db, current_user, split=False)
        wish.status = "converted"
    warning = getattr(wish, "_convert_warning", None)
    excess_warnings = getattr(wish, "_excess_warnings", [])
    purchase_sync = getattr(wish, "_purchase_sync", None)
    await db.commit()
    await db.refresh(wish)
    wish = await _load_wish(wish_id, db)
    out = _enrich(wish)
    out.convert_warning = warning
    out.purchase_ids = created_ids
    out.excess_warnings = excess_warnings
    out.purchase_sync = purchase_sync
    # Владелец (2026-08-20): быстрое одобрение тоже создаёт закупку сразу здесь
    # (см. wish.status = "converted" выше) — отдаём номер закупки в этом же ответе,
    # чтобы фронт не звал следом POST /convert «на всякий случай» (см. тот же фикс
    # в wish_approvals.py::decide и wishes.py::convert_wish).
    if created_ids:
        out.purchases = (await _wish_purchase_summaries_map([wish_id], db)).get(wish_id, [])
    return out


@router.post("/{wish_id}/reject", response_model=WishOut)
async def reject_wish(
    wish_id: int,
    body: WishReject,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Reject a submitted wish with reason (manager+ roles OR assigned approver, submitted -> rejected)."""
    wish = await _load_wish(wish_id, db)

    if not _is_saas(current_user) and wish.status != "submitted":
        raise HTTPException(status_code=400, detail="Заявка должна быть в статусе 'submitted'")

    # Org isolation
    org_ids = get_org_filter(current_user)
    if org_ids is not None and wish.org_id not in org_ids:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")

    if not _is_saas(current_user) and current_user.role not in MANAGER_ROLES and wish.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="Отклонить заявку может менеджер+ или назначенный согласующий")

    # SaaS может отклонить заявку в любом статусе (в т.ч. converted) — если по ней уже
    # есть закупка в Плане закупок, убираем её из плана; закупка дальше стадии
    # «План закупок» отклонение блокирует явным 409 (владелец, 2026-08-07).
    await _withdraw_wish_from_plan(wish.id, db, action_text="отклонить")

    wish.status = "rejected"
    wish.rejection_reason = body.rejection_reason
    # Владелец (2026-08-19): «нужно, чтобы было видно, кто отклонил» — прямое
    # отклонение (эта ветка) часто идёт без единой WishApproval-строки вовсе
    # (менеджер+ отклоняет напрямую, минуя цепочку согласующих), поэтому
    # rejected_by/rejected_at на самой заявке — единственный надёжный источник.
    from datetime import timezone as _tz
    wish.rejected_by = current_user.id
    wish.rejected_at = datetime.now(_tz.utc)
    # Владелец (2026-08-19): «если заявка отклоняется одним из согласующих, она
    # отклоняется у всех, потому что заявку могут изменить именно так, с чем
    # остальные согласующие могут не согласиться. И надо будет повторить заново» —
    # отклонение ОДНИМ согласующим не должно оставлять чужие "approved" решения
    # висеть в цепочке: после правок и повторной отправки согласование обязано
    # начаться с чистого листа, как и при обычном возврате на повторное
    # согласование (см. _reset_approvals выше, строка ~1924).
    await _reset_approvals(wish.id, db)
    await db.commit()
    await db.refresh(wish)
    wish = await _load_wish(wish_id, db)
    out = _enrich(wish)
    return out


@router.patch("/{wish_id}/execution", response_model=WishOut)
async def patch_wish_execution(
    wish_id: int,
    body: WishExecutionPatch,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """B-exec: согласующий ставит исполнителя и срок исполнения.

    Разрешено: assignee (согласующий), admin/manager. Только на submitted/approved.
    """
    wish = await _load_wish(wish_id, db)
    if not _is_saas(current_user) and wish.status not in ("submitted", "approved"):
        raise HTTPException(status_code=400, detail="Срок и исполнителя можно задать только на статусах submitted/approved")
    if not _is_saas(current_user) and current_user.role not in MANAGER_ROLES and wish.assigned_to != current_user.id:
        # Согласующий из цепочки тоже может править (в т.ч. сменить ФЭО, если не согласен)
        from app.models.wish_approval import WishApproval
        in_chain = (await db.execute(
            select(WishApproval.id).where(
                WishApproval.wish_id == wish.id,
                WishApproval.user_id == current_user.id,
            ).limit(1)
        )).scalar_one_or_none()
        if not in_chain:
            raise HTTPException(status_code=403, detail="Только согласующий (в т.ч. из цепочки) или менеджер+ может менять эти поля")

    # Владелец (2026-08-19): «менять позиции может только тот, кто имеет право» —
    # правку ФЭО (категория заявки + построчные feo_category_id/feo_planned_item_id)
    # закрываем ОТДЕЛЬНЫМ правом wish.edit_feo — иначе любой согласующий из цепочки
    # (например, случайный юрист) мог сам перетыкивать позиции. Остальные поля
    # этого эндпоинта (executor_id/execution_deadline/event_id/assigned_to) это
    # право не трогает — их разрешают проверки выше (цепочка/assignee/менеджер+).
    _feo_fields_touched = body.feo_category_id is not None or body.items is not None
    if _feo_fields_touched and current_user.role != "superadmin":
        if not await has_org_key(current_user, db, wish.org_id, "wish.edit_feo", subsidy_id=wish.subsidy_id):
            raise HTTPException(
                status_code=403,
                detail="Право на перераспределение позиций заявки по категориям ФЭО не выдано, обратитесь к администратору",
            )

    if body.executor_id is not None:
        wish.executor_id = body.executor_id
    if body.execution_deadline is not None:
        wish.execution_deadline = body.execution_deadline
    if body.event_id is not None:
        wish.event_id = body.event_id
    if body.feo_category_id is not None:
        wish.feo_category_id = body.feo_category_id
    # W2: assigned_to меняется без сброса цепочки согласования
    if body.assigned_to is not None:
        wish.assigned_to = body.assigned_to

    # Владелец (2026-08-19): построчные ФЭО-правки согласующего (см.
    # WishItemFeoPatch) — состав заявки для него заблокирован на фронте
    # (PurchaseItemsEditor readonly + feoAttrsEditable), но категорию и
    # плановую позицию по каждой строке он вправе перераспределить. Здесь —
    # ТОЛЬКО эти два поля, никогда item_name/quantity/unit_price/total_price/
    # unit/country_origin и без добавления/удаления строк.
    _plan_transfer_warnings: list[str] = []
    if body.items is not None:
        _items_by_id = {wi.id: wi for wi in (wish.items or [])}
        # Валидация ВСЕГО набора до мутации — либо применяем целиком, либо
        # не трогаем ничего (одна транзакция, никаких частичных правок при 400/409).
        for item_patch in body.items:
            if item_patch.id not in _items_by_id:
                raise HTTPException(
                    status_code=400,
                    detail=f"Позиция #{item_patch.id} не принадлежит заявке №{wish.id}",
                )
        for item_patch in body.items:
            wi = _items_by_id[item_patch.id]
            _fields_set = item_patch.model_fields_set
            _cat_changing = (
                'feo_category_id' in _fields_set
                and item_patch.feo_category_id != wi.feo_category_id
            )
            if 'feo_category_id' in _fields_set:
                wi.feo_category_id = item_patch.feo_category_id
            if 'feo_planned_item_id' in _fields_set:
                _old_fpi_id = wi.feo_planned_item_id
                if item_patch.feo_planned_item_id is not None:
                    fpi = await db.get(FeoPlannedItem, item_patch.feo_planned_item_id)
                    if not fpi:
                        raise HTTPException(status_code=409, detail="Плановая позиция не найдена")
                    if not fpi.is_active:
                        raise HTTPException(
                            status_code=409,
                            detail=(
                                f"Плановая позиция «{fpi.name}» деактивирована (удалена из плана), "
                                "привязка к ней невозможна — выберите действующую или создайте новую."
                            ),
                        )
                    if wi.feo_category_id is None or fpi.feo_category_id != wi.feo_category_id:
                        _chosen_cat_name = (await db.execute(
                            select(FeoCategory.name).where(FeoCategory.id == fpi.feo_category_id)
                        )).scalar_one_or_none() or f"#{fpi.feo_category_id}"
                        if wi.feo_category_id is not None:
                            _target_cat_name = (await db.execute(
                                select(FeoCategory.name).where(FeoCategory.id == wi.feo_category_id)
                            )).scalar_one_or_none() or f"#{wi.feo_category_id}"
                        else:
                            _target_cat_name = "без категории ФЭО"
                        raise HTTPException(
                            status_code=409,
                            detail=(
                                f"Плановая позиция «{fpi.name}» относится к категории «{_chosen_cat_name}», "
                                f"а позиция заявки — к категории «{_target_cat_name}». "
                                "Выберите плановую позицию той же категории."
                            ),
                        )
                    wi.feo_planned_item_id = item_patch.feo_planned_item_id
                    wi.over_plan = False
                else:
                    wi.feo_planned_item_id = None
                    wi.over_plan = False
                if _old_fpi_id is not None and _old_fpi_id != wi.feo_planned_item_id:
                    _old_fpi = await db.get(FeoPlannedItem, _old_fpi_id)
                    await _deactivate_if_orphaned(db, _old_fpi)
            elif _cat_changing and wi.feo_planned_item_id is not None:
                # Категория сменилась, а плановую позицию явно не переприслали —
                # та же логика «плановые позиции следуют за сменой категории»,
                # что и в update_wish (см. комментарий там, владелец 2026-08-17).
                if wi.feo_category_id is not None:
                    _w = await _move_or_detach_planned_item(db, wi, wi.feo_category_id)
                    if _w:
                        _plan_transfer_warnings.append(_w)
                else:
                    _old_fpi_id2 = wi.feo_planned_item_id
                    _old_fpi2 = await db.get(FeoPlannedItem, _old_fpi_id2)
                    wi.feo_planned_item_id = None
                    wi.over_plan = False
                    await _deactivate_if_orphaned(db, _old_fpi2)

    await db.commit()
    wish = await _load_wish(wish_id, db)
    out = _enrich(wish)
    out.plan_transfer_warnings = _plan_transfer_warnings
    return out


@router.post("/{wish_id}/status", response_model=WishOut)
async def force_wish_status(
    wish_id: int,
    body: WishStatusForce,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Superadmin/account_owner: force-переключение статуса заявки (минуя workflow-guard'ы).

    Форс «как будто пройден обычный путь»: при переключении в 'converted' реально
    создаются закупки (status='wishes') и заявка уходит из «Заявок» в «Закупки».
    'approved' — только одобрена (остаётся в «Заявках», закупки не создаются).
    """
    if not _is_saas(current_user):
        raise HTTPException(status_code=403, detail="Только superadmin/account_owner может переключать статусы напрямую")
    allowed = {"draft", "submitted", "approved", "rejected", "converted"}
    if body.status not in allowed:
        raise HTTPException(status_code=400, detail=f"Недопустимый статус. Разрешены: {sorted(allowed)}")
    wish = await _load_wish(wish_id, db)
    excess_warnings: list = []
    purchase_sync: Optional[dict] = None

    if body.status == "converted":
        if not wish.items:
            raise HTTPException(status_code=400, detail="Заявка пустая — нечего распределять")
        try:
            # Аварийный рычаг: цепочку не блокируем, но закрываем pending-согласования,
            # чтобы заявка не «зависла» у согласующих
            from app.models.wish_approval import WishApproval
            pending = (await db.execute(
                select(WishApproval).where(
                    WishApproval.wish_id == wish.id,
                    WishApproval.status == "pending",
                )
            )).scalars().all()
            for a in pending:
                a.status = "approved"
                a.comment = "Закрыто принудительным переводом статуса"
                a.decided_by_user_id = current_user.id
                a.decided_by_username = current_user.full_name or current_user.username
                a.decided_at = func.now()
            await _distribute_wish_to_purchases(wish, db, current_user, split=False)
            wish.status = "converted"
            wish.approved_by = wish.approved_by or current_user.id
            excess_warnings = getattr(wish, "_excess_warnings", [])
            purchase_sync = getattr(wish, "_purchase_sync", None)
            await db.commit()
        except HTTPException:
            await db.rollback()
            raise
        except Exception as e:
            await db.rollback()
            raise HTTPException(status_code=500, detail=f"Ошибка при создании закупок — откат: {e}")
    elif body.status == "approved":
        wish.status = "approved"
        if not wish.approved_by:
            wish.approved_by = current_user.id
        await db.commit()
    else:
        # draft / submitted / rejected — просто переключение статуса.
        # Гейт (владелец, 2026-08-07): даже аварийный SaaS-рычаг больше не обходит
        # блокировку — закупка дальше «Плана закупок» останавливает force-переключение
        # так же, как обычный откат.
        _force_action_text = {
            "draft": "вернуть в черновик",
            "submitted": "вернуть на повторное согласование",
            "rejected": "отклонить",
        }.get(body.status, "сменить статус")
        await _withdraw_wish_from_plan(wish.id, db, action_text=_force_action_text)
        wish.status = body.status
        if body.status in ("draft", "submitted"):
            # Возврат в черновик / на повторное согласование форс-рычагом —
            # старое отклонение так же не актуально, как и в обычном пути.
            wish.rejected_by = None
            wish.rejected_at = None
            wish.rejection_reason = None
            await db.flush()
            # Полный сброс (без keep_user_id) — иначе строка отклонившего
            # согласующего (см. decide()) осталась бы 'rejected' навсегда и
            # заблокировала бы sequential-гейт при повторном согласовании.
            await _reset_approvals(wish.id, db)
        await db.commit()

    wish = await _load_wish(wish_id, db)
    out = _enrich(wish)
    out.excess_warnings = excess_warnings
    out.purchase_sync = purchase_sync
    return out


@router.post("/{wish_id}/convert")
async def convert_wish(
    wish_id: int,
    body: WishConvert,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('wishes')),
):
    """Convert an approved wish to a purchase (org_admin+, approved -> converted).
    B4: copies all WishItems to PurchaseItems with quantity/price from wish items.
    B9: carries feo_category_id from wish and per-item feo_category_id.
    B10: backfills product_id by item_name for legacy wish_items.
    """
    from app.models.purchase import Purchase
    from app.models.product import Product
    from sqlalchemy.orm import selectinload as sil

    wish = await _load_wish(wish_id, db)
    # Идемпотентность (владелец, 2026-08-20): последний согласующий цепочки сам
    # переносит заявку в закупку (см. wish_approvals.py::decide, _distribute_wish_to_purchases
    # + wish.status = "converted"). Если фронт после этого всё равно дёргает /convert
    # (например пользователь нажал «Передать в План закупок» по старой памяти) — заявка
    # уже 'converted', и раньше это падало 400 «должна быть в статусе approved», хотя
    # закупка реально была создана. Статус 'converted' здесь пропускаем в общий путь —
    # ниже блок «защита от дублей» (existing purchases) находит её и просто возвращает,
    # вторую закупку не создавая.
    _was_already_converted = wish.status == "converted"

    if not _is_saas(current_user) and wish.status not in ("approved", "converted"):
        from app.routers.feo_planned_items import _WISH_STATUS_LABELS
        _label = _WISH_STATUS_LABELS.get(wish.status, wish.status)
        raise HTTPException(
            status_code=400,
            detail=(
                f"Заявку нельзя перенести в закупку: она в статусе «{_label}». "
                "Перенести можно только согласованную заявку."
            ),
        )

    # Org isolation
    org_ids = get_org_filter(current_user)
    if org_ids is not None and wish.org_id not in org_ids:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")

    await _ensure_no_pending_approvals(wish, db, current_user)

    # Защита от дублей: закупки по заявке уже есть — не создаём вторую,
    # скрытые (status='wishes') продвигаем в План закупок
    existing = (await db.execute(
        select(Purchase).where(Purchase.wish_id == wish.id)
    )).scalars().all()

    if _was_already_converted and not existing:
        # Рассинхрон: заявка помечена converted, но ни одной закупки по ней нет —
        # такого в норме не бывает (единственный путь в 'converted' — либо этот же
        # эндпоинт, либо decide()/approve(), оба создают закупку ДО смены статуса).
        raise HTTPException(
            status_code=409,
            detail=(
                "Заявка помечена как перенесённая в закупку, но сама закупка не найдена "
                "(рассинхронизация данных). Обратитесь к администратору — вручную создавать "
                "закупку заново нельзя, это привело бы к дублю."
            ),
        )

    if existing:
        # W2-гейт: проверяем категорию ФЭО и даты ПЕРЕД продвижением скрытых закупок —
        # это тоже момент «попадания в План закупок» (владелец, 2026-08-11).
        wishes_existing = [ep for ep in existing if ep.status == "wishes"]
        if wishes_existing and getattr(wish, 'source', None) != 'advance_report':
            items_res = await db.execute(select(WishItem).where(WishItem.wish_id == wish.id))
            _all_items_existing = items_res.scalars().all()
            await _ensure_feo_categories_assigned(
                wish, [it for it in _all_items_existing if _is_meaningful_item(it)], db,
            )
            await _ensure_needed_dates(wish, db, _all_items_existing)
        # Задача владельца, план zany-fluttering-mountain.md п.4 (2026-08-10): эта
        # ветка — ОТДЕЛЬНЫЙ путь движения закупки по стадиям (wishes → plan_schedule),
        # не проходящий через POST /api/purchases/{pid}/transition (см. её гейт в
        # purchase_transitions.py) — идемпотентное повторное согласование заявки,
        # у которой уже есть скрытые закупки.
        # Владелец (2026-08-12): продвижение больше НЕ блокируется несогласованным
        # превышением — assert_no_unapproved_excess убран, вместо отказа собираем
        # excess_warnings (по категории и её предкам) ПОСЛЕ фактического
        # продвижения статусов, см. _collect_excess_warnings.
        _cat_items: dict[int, list[dict]] = {}
        if wishes_existing:
            from app.models.purchase_item import PurchaseItem as _PurchaseItem
            _ep_ids = [ep.id for ep in wishes_existing]
            _items_res = await db.execute(
                select(_PurchaseItem).where(_PurchaseItem.purchase_id.in_(_ep_ids))
            )
            _items_by_purchase: dict[int, list] = {}
            for _pit in _items_res.scalars().all():
                _items_by_purchase.setdefault(_pit.purchase_id, []).append(_pit)
            for ep in wishes_existing:
                _ep_items = _items_by_purchase.get(ep.id, [])
                for _pit in _ep_items:
                    _cid = _pit.feo_category_id or ep.feo_category_id
                    if _cid:
                        _cat_items.setdefault(_cid, []).append({
                            "name": _pit.item_name, "amount": float(_pit.total_price or 0),
                        })
                if not any(_pit.feo_category_id for _pit in _ep_items) and ep.feo_category_id:
                    _cat_items.setdefault(ep.feo_category_id, []).append({
                        "name": ep.item_name or ep.subject or f"Закупка №{ep.id}",
                        "amount": float(ep.total_nmck or ep.planned_total_price or 0),
                    })
        for ep in existing:
            if ep.status == "wishes":
                ep.status = "plan_schedule"
            # Контрагент заявки (владелец, 2026-08-17) — заполняем ТОЛЬКО пустое
            # поле закупки, не перетираем уже заданное (могло быть проставлено
            # вручную прямо в закупке).
            if getattr(wish, 'contractor_id', None) and not ep.contractor_id:
                ep.contractor_id = wish.contractor_id
        wish.status = "converted"
        wish.approved_by = wish.approved_by or current_user.id
        wish.purchase_id = wish.purchase_id or existing[0].id
        # Повторный перенос (владелец, план crystalline-soaring-heron.md, п.2) —
        # тот же helper, что и в _distribute_wish_to_purchases: приводит предмет и
        # состав уже существующей закупки к текущей заявке (см. её докстринг).
        _purchase_sync = await _sync_purchase_from_wish(wish, existing, db)
        await db.flush()
        _excess_warnings = await _collect_excess_warnings(db, wish.subsidy_id, _cat_items)
        await db.commit()
        return {
            "wish_id": wish.id, "purchase_id": existing[0].id, "status": "converted",
            "registry_number": existing[0].registry_number,
            "already_converted": _was_already_converted,
            "excess_warnings": _excess_warnings,
            "purchase_sync": _purchase_sync,
        }

    # Preload items with products (B4/B10)
    res = await db.execute(
        select(WishItem).options(sil(WishItem.product)).where(WishItem.wish_id == wish.id)
    )
    items_full = res.scalars().all()

    # Гейт ФЭО (владелец, 2026-08-11): /convert — отдельный путь создания закупок
    # мимо _distribute_wish_to_purchases (см. комментарий про «Дефект» ниже про
    # тот же паразитный дубль для «ТЗ не выше плана»/превышения ФЭО) — тем же
    # общим хелпером закрываем и его. См. docstring _ensure_feo_categories_assigned.
    await _ensure_feo_categories_assigned(wish, [it for it in items_full if _is_meaningful_item(it)], db)

    # W2-гейт в основном пути /convert
    await _ensure_needed_dates(wish, db, items_full)

    # B10: Backfill product_id for legacy items lacking it
    missing = [it for it in items_full if not it.product_id and (it.item_name or "").strip()]
    if missing:
        names = list({(it.item_name or "").strip() for it in missing})
        pres = await db.execute(select(Product).where(Product.name.in_(names)))
        name_to_product = {(p.name or "").strip().lower(): p for p in pres.scalars().all()}
        for it in missing:
            hit = name_to_product.get((it.item_name or "").strip().lower())
            if hit:
                it.product_id = hit.id

    # Дефект (приёмка 2026-08-07): /convert — ОТДЕЛЬНЫЙ путь создания закупок,
    # не проходящий через _distribute_wish_to_purchases (approve_wish и
    # force_wish_status('converted') используют её и получают проверки автоматом,
    # этот эндпоинт — нет, шёл мимо гейта «ТЗ не выше плана» и мимо превышения
    # ФЭО). Воспроизводим ТОТ ЖЕ порядок и теми же общими функциями (никакой
    # второй копии логики): автозаведение → «ТЗ не выше плана» → превышение ФЭО.
    # Обоснование порядка — см. комментарий у аналогичного блока в
    # _distribute_wish_to_purchases.
    await _auto_assign_planned_items(
        items_full, wish.feo_category_id, db, note=f"заявкой №{wish.id} (/convert)",
    )
    # Владелец (2026-08-17, прод-инцидент РЕЕ-2026-00887): накопление по общей
    # плановой позиции в пределах ЭТОЙ заявки — см. assert_tz_batch_not_over_plan.
    await assert_tz_batch_not_over_plan(db, items_full)
    # Владелец (2026-08-12): /convert больше не блокируется несогласованным
    # превышением ФЭО — assert_no_unapproved_excess убран, вместо отказа собираем
    # excess_warnings ПОСЛЕ создания закупки (см. вызов _collect_excess_warnings
    # ниже, уже после flush() позиций).
    _conv_cat_items: dict[int, list[dict]] = {}
    for _wi in items_full:
        _cid = _wi.feo_category_id or wish.feo_category_id
        if _cid:
            _conv_cat_items.setdefault(_cid, []).append({
                "name": _wi.item_name, "amount": float(_wi.total_price or 0),
            })

    # B4: planned_total_price = SUM(items.total_price), fallback to body/wish
    total_nmck = sum(float(i.total_price or 0) for i in items_full)

    # B9: pass feo_category_id from wish-level
    # Backend pre-fill: если в body не пришло approved_quantity/price (= 0/None) — считаем из items
    total_qty = sum(float(i.quantity or 0) for i in items_full)
    eff_qty = body.approved_quantity if (body.approved_quantity and float(body.approved_quantity) > 0) else (total_qty or wish.quantity)
    eff_price = body.approved_price if (body.approved_price and float(body.approved_price) > 0) else (total_nmck or wish.estimated_price)
    conv_dates = {_eff_date(wish, wi) for wi in items_full}
    conv_dates.discard(None)
    conv_delivery_date = conv_dates.pop() if len(conv_dates) == 1 else None

    # C1: авансовый отчёт → фиксируем тип; обычная заявка → single
    _is_advance_conv = (getattr(wish, 'source', None) == 'advance_report')
    _conv_purchase_method = 'advance' if _is_advance_conv else 'single'
    _conv_payment_basis_type = 'advance_report' if _is_advance_conv else None
    p = Purchase(
        wish_id=wish.id,
        subsidy_id=body.subsidy_id or wish.subsidy_id,
        feo_category_id=wish.feo_category_id,  # B9
        event_id=getattr(wish, 'event_id', None),  # «Мероприятие»
        item_name=wish.title,
        subject=wish.title,
        planned_quantity=eff_qty,
        planned_total_price=eff_price,
        total_nmck=total_nmck or float(wish.estimated_price or 0),
        nmck=total_nmck or float(wish.estimated_price or 0),
        status="plan_schedule",
        service_note_text=wish.justification,
        service_note_by=wish.created_by,
        # B1: исполнитель = executor_id (без фолбэка на инициатора)
        assigned_user_id=getattr(wish, 'executor_id', None),
        # B1: служебка «на чьё имя» = assigned_to заявки
        service_note_to_user_id=wish.assigned_to,
        execution_term=getattr(wish, 'execution_deadline', None),  # B-exec: срок исполнения
        delivery_date=conv_delivery_date,
        purchase_method=_conv_purchase_method,
        payment_basis_type=_conv_payment_basis_type,
        feo_per_item=bool(getattr(wish, 'feo_per_item', False)),
        vat_mode=(getattr(wish, 'vat_mode', None) or 'uniform'),
        # Контрагент заявки (владелец, 2026-08-17) — переезжает в закупку, если
        # указан. Purchase свежесозданный, поэтому «не перетирать уже заданное»
        # выполняется автоматически.
        contractor_id=getattr(wish, 'contractor_id', None),
    )
    db.add(p)
    await db.flush()  # get p.id

    # Контрагент заявки — на каждую позицию (см. аналогичный резолв в
    # _distribute_wish_to_purchases выше). Если контрагента в справочнике ещё
    # нет — свободный ввод contractor_name.
    _conv_contractor_obj = None
    if getattr(wish, 'contractor_id', None):
        from app.models.contractor import Contractor as _ConvContractor
        _conv_contractor_obj = await db.get(_ConvContractor, wish.contractor_id)

    # B4/B9/B10: copy all WishItems to PurchaseItems
    for wi in items_full:
        pi = PurchaseItem(
            purchase_id=p.id,
            product_id=wi.product_id,
            item_name=wi.item_name,
            item_type=wi.item_type,
            quantity=wi.quantity,           # B4: «утверждённое кол-во» = из WishItem
            unit=wi.unit,
            unit_price=wi.unit_price,       # B4: «утверждённая цена» = из WishItem
            total_price=wi.total_price,
            country_origin=wi.country_origin,
            feo_category_id=wi.feo_category_id,  # B9: per-item feo
            feo_planned_item_id=wi.feo_planned_item_id,  # расходуем уже запланированную позицию, не задваиваем план
            over_plan=getattr(wi, 'over_plan', False),
            needed_date=_eff_date(wish, wi),  # W2: наследование эффективной даты
            wish_item_id=wi.id,  # W1: hard link to source WishItem
            vat_rate=getattr(wi, 'vat_rate', None),
            contractor_id=(_conv_contractor_obj.id if _conv_contractor_obj else None),
            contractor_inn=(_conv_contractor_obj.inn if _conv_contractor_obj else None),
            contractor_name=(
                _conv_contractor_obj.name if _conv_contractor_obj
                else getattr(wish, 'contractor_name', None)
            ),
        )
        db.add(pi)
    await db.flush()

    wish.purchase_id = p.id
    wish.status = "converted"
    wish.approved_by = current_user.id

    _excess_warnings = await _collect_excess_warnings(db, wish.subsidy_id, _conv_cat_items)
    await db.commit()

    return {
        "wish_id": wish.id, "purchase_id": p.id, "status": "converted",
        "registry_number": p.registry_number,
        "already_converted": False,
        "excess_warnings": _excess_warnings,
        # Первое создание закупки — синхронизировать нечего (см. purchase_sync
        # в ветке «существующая закупка» выше).
        "purchase_sync": None,
    }


@router.delete("/{wish_id}", status_code=204)
async def delete_wish(
    wish_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Delete a draft wish (creator only)."""
    wish = await _load_wish(wish_id, db)

    if not _is_saas(current_user) and wish.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Только автор может удалить заявку")
    if not _is_saas(current_user) and wish.status != "draft":
        raise HTTPException(status_code=400, detail="Можно удалить только черновик")

    # Заявку удаляют — purchases.wish_id обнулится каскадом (ON DELETE SET NULL),
    # закупка осиротеет в плане. Плановые (plan_schedule) убираем из плана заранее;
    # если что-то уже в работе/договоре — удаление блокируем явно (владелец, 2026-08-07).
    await _withdraw_wish_from_plan(wish.id, db, action_text="удалить")

    await db.delete(wish)
    await db.commit()


@router.patch("/{wish_id}/items/{item_id}")
async def patch_wish_item(
    wish_id: int,
    item_id: int,
    body: WishItemPatch,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """D-04: Drag-drop target update. Scoped to wish — cannot move items between wishes.

    Returns 409 if wish is approved (read-only).
    Returns 404 if item does not belong to wish_id.
    """
    wish = await _load_wish(wish_id, db)
    if not _is_saas(current_user) and wish.status not in ("draft", "submitted"):
        raise HTTPException(status_code=409, detail="Заявка уже одобрена — редактирование запрещено")
    # Find item BELONGING TO THIS WISH
    item = next((i for i in wish.items if i.id == item_id), None)
    if item is None:
        raise HTTPException(status_code=404, detail="Позиция не найдена в данной заявке")
    # body.target_column_key may be None (clear) or a non-empty string (override)
    item.target_column_key = body.target_column_key
    await db.commit()
    await db.refresh(item)
    return {"id": item.id, "target_column_key": item.target_column_key}


@router.post("/{wish_id}/approve-distribution")
async def approve_distribution(
    wish_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """D-05/D-06: Atomic all-or-nothing распределение. Creates N purchases (status='wishes'),
    one per distinct resolved column key group, copies wish items to purchase_items,
    creates assignment chat rooms, links purchase.wish_id, then marks wish.status='converted'.

    Распределённая заявка уходит из «Заявок» в «Закупки» (status='converted').
    Rolls back entirely on any failure — zero purchases persist if any step fails.
    Returns 400 if wish is already distributed.
    """
    wish = await _load_wish(wish_id, db)
    org_ids = get_org_filter(current_user)
    if org_ids is not None and wish.org_id not in org_ids:
        raise HTTPException(status_code=403, detail="Заявка не входит в ваши организации")
    if not _is_saas(current_user) and wish.status == "converted":
        raise HTTPException(status_code=400, detail="Заявка уже распределена")
    if not _is_saas(current_user) and wish.status not in ("draft", "submitted", "approved"):
        raise HTTPException(status_code=400, detail=f"Нельзя распределить заявку в статусе {wish.status}")
    if current_user.role not in ADMIN_ROLES and wish.assigned_to != current_user.id:
        raise HTTPException(status_code=403, detail="Распределять заявку может админ или назначенный согласующий")
    if not wish.items:
        raise HTTPException(status_code=400, detail="Заявка пустая — нечего распределять")

    await _ensure_no_pending_approvals(wish, db, current_user)

    try:
        ids = await _distribute_wish_to_purchases(wish, db, current_user)
        wish.status = "converted"
        wish.approved_by = current_user.id
        await db.commit()
    except HTTPException:
        await db.rollback()
        raise
    except Exception as e:
        await db.rollback()
        raise HTTPException(status_code=500, detail=f"Ошибка при создании закупок — откат: {e}")

    return {
        "wish_id": wish.id,
        "purchase_ids": ids,
        "count": len(ids),
        "status": "converted",
        "warning": getattr(wish, "_convert_warning", None),
        "excess_warnings": getattr(wish, "_excess_warnings", []),
    }


@router.get("/{wish_id}/export.xlsx")
async def export_wish_xlsx(
    wish_id: int,
    with_photos: bool = True,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    import httpx
    from io import BytesIO
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.comments import Comment
    from fastapi.responses import StreamingResponse
    import io
    from urllib.parse import quote
    from PIL import Image as PILImage
    try:
        from pillow_heif import register_heif_opener
        register_heif_opener()
    except Exception:
        pass

    wish = await _load_wish(wish_id, db)
    org_ids = get_org_filter(current_user)
    if org_ids is not None and wish.org_id not in org_ids:
        raise HTTPException(status_code=403, detail="Нет доступа к этой заявке")
    # Позиции заявки + связанный товар (фото/категория/вид/описание/ссылки)
    res = await db.execute(
        select(WishItem)
        .where(WishItem.wish_id == wish_id)
        .options(selectinload(WishItem.product))
        .order_by(WishItem.id)
    )
    items = res.scalars().all()

    async def _image_bytes(p):
        if p is not None and getattr(p, "photo_data", None):
            return bytes(p.photo_data)
        url = (p.photo_url or p.photo_link) if p is not None else None
        if url and str(url).startswith("http"):
            try:
                async with httpx.AsyncClient(timeout=8.0, follow_redirects=True) as cl:
                    resp = await cl.get(url)
                    if resp.status_code == 200 and resp.content:
                        return resp.content
            except Exception:
                return None
        return None

    def _thumb_png(raw: bytes):
        try:
            im = PILImage.open(BytesIO(raw))
            if im.mode not in ("RGB", "RGBA"):
                im = im.convert("RGB")
            im.thumbnail((90, 90))
            out = BytesIO()
            im.save(out, format="PNG")
            out.seek(0)
            return out, im.width, im.height
        except Exception:
            return None

    from openpyxl.styles import Border, Side
    from openpyxl.utils import get_column_letter
    wb = Workbook(); ws = wb.active; ws.title = "Заявка"
    thin = Side(style="thin", color="C0C0C0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_fill = PatternFill("solid", fgColor="FB923C")
    header_font = Font(bold=True, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_top = Alignment(horizontal="left", vertical="top", wrap_text=True)
    link_font = Font(color="0563C1", underline="single")

    if with_photos:
        headers = [
            "№ п/п", "Наименование", "Фото", "Категория товара", "Вид",
            "Описание", "Ссылка пример", "Количество", "Ед. изм.",
            "Плановая Цена за ед", "Плановая Сумма",
        ]
        widths = [7, 34, 14, 20, 16, 50, 32, 12, 10, 16, 16]
    else:
        headers = [
            "№ п/п", "Наименование", "Категория товара", "Вид",
            "Описание", "Ссылка пример", "Количество", "Ед. изм.",
            "Плановая Цена за ед", "Плановая Сумма",
        ]
        widths = [7, 34, 20, 16, 50, 32, 12, 10, 16, 16]
    ncols = len(headers)
    for c, (h, w) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.font = header_font; cell.fill = header_fill
        cell.alignment = center; cell.border = border
        ws.column_dimensions[cell.column_letter].width = w
    ws.row_dimensions[1].height = 32

    # Позиции колонок (без фото всё после "Наименование" сдвинуто влево)
    col_name = 2
    col_photo = 3 if with_photos else None
    base = 4 if with_photos else 3  # "Категория товара" и далее
    col_cat = base
    col_kind = base + 1
    col_descr = base + 2
    col_example = base + 3
    col_qty = base + 4
    col_unit = base + 5
    col_price = base + 6
    col_total = base + 7

    if with_photos:
        ws.cell(row=1, column=col_photo).comment = Comment(
            "Изображения встроены в файл и отображаются всегда — настройка безопасности Excel не требуется. "
            "Оригинальные ссылки на фото — на скрытом листе «Ссылки» (правый клик по ярлыку листа → Показать).",
            "GALA")

    def _photo_url(p):
        if not p:
            return ""
        return p.photo_url or p.photo_link or (f"/api/products/{p.id}/photo" if p.photo_data else "")

    def _example_url(p):
        if not p:
            return ""
        links = p.price_links if isinstance(p.price_links, list) else []
        for l in links:
            url = l.get("url") if isinstance(l, dict) else None
            if url and str(url).startswith("http"):
                return url
        cl = p.clarification_link or ""
        return cl if cl.startswith("http") else ""

    # Скрытый лист со ссылками — только в режиме с фото
    if with_photos:
        ws_links = wb.create_sheet("Ссылки")
        ws_links.append(["№ п/п", "Наименование", "Фото URL"])

    r = 2
    for i, it in enumerate(items, start=1):
        p = getattr(it, "product", None)
        name = it.item_name or (p.name if p else "") or ""
        category = (p.category if p else "") or ""
        kind = (p.product_type if p else "") or ""
        descr = (p.description if p else "") or ""
        photo = _photo_url(p)
        example = _example_url(p)
        ws.cell(row=r, column=1, value=i)
        ws.cell(row=r, column=col_name, value=name)
        if with_photos:
            raw = await _image_bytes(p)
            if raw:
                t = _thumb_png(raw)
                if t:
                    bio, w_px, h_px = t
                    img = XLImage(bio)
                    img.width = w_px; img.height = h_px
                    ws.add_image(img, f"{get_column_letter(col_photo)}{r}")
                    ws.row_dimensions[r].height = max(
                        ws.row_dimensions[r].height or 0, h_px * 0.78 + 6
                    )
        ws.cell(row=r, column=col_cat, value=category)
        ws.cell(row=r, column=col_kind, value=kind)
        ws.cell(row=r, column=col_descr, value=descr)
        cex = ws.cell(row=r, column=col_example, value=example)
        if example.startswith("http"):
            cex.hyperlink = example; cex.font = link_font
        qty = float(it.quantity or 0)
        unit_price = float(it.unit_price or 0)
        sum_val = float(it.total_price or 0) or (qty * unit_price)
        ws.cell(row=r, column=col_qty, value=qty)
        ws.cell(row=r, column=col_unit, value=it.unit or "")
        price = ws.cell(row=r, column=col_price, value=unit_price)
        total = ws.cell(row=r, column=col_total, value=sum_val)
        price.number_format = '# ##0.00'
        total.number_format = '# ##0.00'
        left_cols = {col_name, col_descr}
        leftcenter_cols = {col_example}
        if with_photos:
            leftcenter_cols.add(col_photo)
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            if c in left_cols:
                cell.alignment = left_top
            elif c in leftcenter_cols:
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
            else:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if with_photos:
            ws_links.append([i, name, photo or ""])
        r += 1

    if with_photos:
        ws_links.sheet_state = "hidden"
    last_row = max(r - 1, 1)
    ws.auto_filter.ref = f"A1:{get_column_letter(ncols)}{last_row}"
    ws.freeze_panes = "A2"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    _wish_title_raw = (wish.title or "").strip()
    _wish_title_clean = _re.sub(r'[\\/:*?"<>|\r\n]+', "", _wish_title_raw)
    _wish_title_clean = _re.sub(r'\s+', "_", _wish_title_clean)[:50]
    fname = f"Заявка_{_wish_title_clean}_{wish.id}.xlsx" if _wish_title_clean else f"Заявка_{wish.id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(fname, safe='-_.~')}"},
    )
