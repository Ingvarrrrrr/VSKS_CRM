"""
Импорт базового ФЭО и справочников в локальную SQLite БД
из файла сметы (листы WORK и Unique).

Ожидания по структуре:
- Лист WORK:
  - строка 2 содержит заголовки;
  - один из столбцов: "Направление расходования ФЭО"
  - один из столбцов: "Наименование статей затрат"
  - остальные колонки пока игнорируются или могут быть добавлены позже.

- Лист Unique:
  - заголовки (1 или 2 строка, определяется автоматически);
  - колонка с названием товара/услуги;
  - колонка с тех.описанием;
  - опционально колонка "Наименование статей затрат" для привязки к статье.
"""

from __future__ import annotations

from pathlib import Path
from typing import Dict, Optional, Tuple

from openpyxl import load_workbook

from db_sqlite import (
    get_group_id_by_code,
    reset_all_data,
    upsert_direction,
    upsert_cost_item,
    link_direction_to_cost_items,
    add_unique_item,
    update_cost_item_details,
)


def _find_column_indices(headers_row) -> Dict[str, int]:
    """
    Поиск полезных колонок в WORK по заголовкам.

    Возвращает словарь:
    {
        "direction": index,
        "cost_item": index,
    }
    """
    result: Dict[str, int] = {}
    for idx, cell in enumerate(headers_row):
        if not cell.value:
            continue
        text = str(cell.value).strip().lower()
        if not text:
            continue

        # Направление расходования ФЭО
        if "направлен" in text and "фэо" in text and "direction" not in result:
            result["direction"] = idx
        # Наименование статей затрат
        if "наимен" in text and "стат" in text and "затрат" in text and "cost_item" not in result:
            result["cost_item"] = idx

        # Единица измерения
        if (
            ("ед" in text and "изм" in text)
            or "единиц" in text
            or "ед. изм" in text
        ) and "unit" not in result:
            result["unit"] = idx

        # Плановая цена за единицу
        if (
            ("планов" in text and "цена" in text)
            or ("цена" in text and "ед" in text)
        ) and "plan_price" not in result:
            result["plan_price"] = idx

    return result


def import_work_sheet(excel_path: Path, group_code: str = "ZO") -> int:
    """
    Импорт направлений и статей затрат из листа WORK
    для указанной группы (по умолчанию ЗО).

    Возвращает количество обработанных строк с данными.
    """
    if not excel_path.is_file():
        raise FileNotFoundError(f"Файл не найден: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "work":
            sheet_name = name
            break
    if not sheet_name:
        raise ValueError("Лист 'WORK' не найден в файле сметы.")

    ws = wb[sheet_name]

    # Заголовки во второй строке
    header_row = ws[2]
    col_map = _find_column_indices(header_row)
    if "direction" not in col_map or "cost_item" not in col_map:
        raise ValueError(
            "Не удалось найти колонки 'Направление расходования ФЭО' "
            "и 'Наименование статей затрат' в листе WORK."
        )

    group_id = get_group_id_by_code(group_code)
    if not group_id:
        raise ValueError(f"Группа ФЭО с кодом '{group_code}' не найдена в БД.")

    processed = 0
    # Для накопления единиц измерения и цен по каждой статье затрат
    cost_item_meta: Dict[str, Tuple[Optional[str], Optional[float]]] = {}

    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row or not any(row):
            continue

        direction_val = row[col_map["direction"]] if col_map["direction"] < len(row) else None
        cost_item_val = row[col_map["cost_item"]] if col_map["cost_item"] < len(row) else None

        direction = str(direction_val).strip() if direction_val else ""
        cost_item = str(cost_item_val).strip() if cost_item_val else ""

        if not direction or not cost_item:
            continue

        # Пытаемся вытащить ед. измерения и плановую цену из этой строки
        unit_val = None
        price_val: Optional[float] = None
        if "unit" in col_map and col_map["unit"] < len(row):
            unit_cell = row[col_map["unit"]]
            unit_val = str(unit_cell).strip() if unit_cell else None
        if "plan_price" in col_map and col_map["plan_price"] < len(row):
            pv = row[col_map["plan_price"]]
            try:
                if isinstance(pv, str):
                    txt = pv.replace(" ", "").replace(",", ".")
                    price_val = float(txt) if txt else None
                elif pv is not None:
                    price_val = float(pv)
            except (TypeError, ValueError):
                price_val = None

        # Запоминаем последнюю не пустую ед.изм/цену для каждой статьи
        meta_unit, meta_price = cost_item_meta.get(cost_item, (None, None))
        if unit_val:
            meta_unit = unit_val
        if price_val is not None:
            meta_price = price_val
        cost_item_meta[cost_item] = (meta_unit, meta_price)

        direction_id = upsert_direction(group_id, direction)
        cost_item_id = upsert_cost_item(group_id, cost_item)
        link_direction_to_cost_items(direction_id, [cost_item_id])
        processed += 1

    # Обновляем детали по статьям затрат (ед.изм и плановая цена)
    for ci_name, (unit, price) in cost_item_meta.items():
        if not ci_name:
            continue
        cost_item_id = upsert_cost_item(group_id, ci_name)
        update_cost_item_details(cost_item_id, unit, price)

    return processed


def _detect_unique_header(ws) -> Tuple[int, Dict[str, int]]:
    """
    Определение строки заголовков и полезных колонок на листе Unique.
    Возвращает (row_index, col_map).
    """
    for header_row_idx in (1, 2):
        header_row = ws[header_row_idx]
        col_map: Dict[str, int] = {}
        for idx, cell in enumerate(header_row):
            if not cell.value:
                continue
            text = str(cell.value).strip().lower()
            if not text:
                continue
            # Наименование товара / услуги
            if (
                ("наимен" in text and ("товар" in text or "услуг" in text or "позиция" in text))
                or ("наимен" in text and "unique" in text)
            ) and "name" not in col_map:
                col_map["name"] = idx
            # Техническое описание
            if ("описан" in text or "тех" in text) and "description" not in col_map:
                col_map["description"] = idx
            # Наименование статьи затрат (для связи с cost_item)
            if "наимен" in text and "стат" in text and "затрат" in text and "cost_item" not in col_map:
                col_map["cost_item"] = idx

        if "name" in col_map:
            return header_row_idx, col_map

    raise ValueError("Не удалось определить строку заголовков и колонку с наименованием в листе Unique.")


def import_unique_sheet(excel_path: Path, default_group_code: str = "ZO") -> int:
    """
    Импорт листа Unique в таблицу unique_item.

    Если в строке есть колонка с 'Наименование статей затрат',
    пытаемся привязать запись к соответствующей статье затрат (через upsert_cost_item).
    """
    if not excel_path.is_file():
        raise FileNotFoundError(f"Файл не найден: {excel_path}")

    wb = load_workbook(excel_path, data_only=True)

    sheet_name = None
    for name in wb.sheetnames:
        if name.strip().lower() == "unique":
            sheet_name = name
            break
    if not sheet_name:
        raise ValueError("Лист 'Unique' не найден в файле сметы.")

    ws = wb[sheet_name]

    header_row_idx, col_map = _detect_unique_header(ws)
    group_id = get_group_id_by_code(default_group_code)

    imported = 0

    for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
        if not row or not any(row):
            continue

        name_val = row[col_map["name"]] if col_map["name"] < len(row) else None
        name = str(name_val).strip() if name_val else ""
        if not name:
            continue

        desc_val = (
            row[col_map["description"]] if "description" in col_map and col_map["description"] < len(row) else None
        )
        description = str(desc_val).strip() if desc_val else ""

        cost_item_id: Optional[int] = None
        if "cost_item" in col_map and group_id:
            ci_val = row[col_map["cost_item"]] if col_map["cost_item"] < len(row) else None
            ci_name = str(ci_val).strip() if ci_val else ""
            if ci_name:
                cost_item_id = upsert_cost_item(group_id, ci_name)

        add_unique_item(name=name, description=description, cost_item_id=cost_item_id)
        imported += 1

    return imported


def import_feo_to_sqlite(excel_path: str, work_group_code: str = "ZO") -> Tuple[int, int]:
    """
    Полный импорт:
    - очищает текущие данные ФЭО в SQLite;
    - импортирует направления и статьи затрат из WORK (для указанной группы);
    - импортирует элементы Unique (для всех, с привязкой к статьям, если возможно).

    Возвращает кортеж (count_work, count_unique).
    """
    path = Path(excel_path)
    if not path.is_file():
        raise FileNotFoundError(f"Файл не найден: {path}")

    # Полностью очищаем старые данные ФЭО/Unique
    reset_all_data()

    work_count = import_work_sheet(path, group_code=work_group_code)
    unique_count = import_unique_sheet(path, default_group_code=work_group_code)

    return work_count, unique_count

