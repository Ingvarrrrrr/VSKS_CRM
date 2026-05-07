"""Phase 22 Plan 02 — unit tests for bank_statement_parser.

Tests are self-contained (synthetic xlsx in BytesIO), no DB, no async.
Run with: pytest backend/tests/test_bank_parser.py -v
"""
from __future__ import annotations

from datetime import date
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

from app.services.bank_statement_parser import (
    ParsedRow,
    parse_purpose,
    parse_workbook,
)


# ---------------------------------------------------------------------------
# Helpers to build synthetic xlsx in memory
# ---------------------------------------------------------------------------

def _make_workbook(headers: list, data_rows: list[list]) -> bytes:
    """Create an xlsx with one sheet: first row headers, then data rows."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(headers)
    for row in data_rows:
        ws.append(row)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_workbook_merged_headers(
    header_groups: list[tuple[str, int]],
    data_rows: list[list],
) -> bytes:
    """Build xlsx where some headers span multiple columns (merged cells).

    header_groups: list of (header_text, col_span) pairs.
    Physically writes merged cells; openpyxl read_only exposes merged start
    cell with value and continuation cells as None — matching production.
    """
    wb = openpyxl.Workbook()
    ws = wb.active

    col = 1
    for text, span in header_groups:
        ws.cell(row=1, column=col, value=text)
        if span > 1:
            ws.merge_cells(
                start_row=1, start_column=col,
                end_row=1, end_column=col + span - 1,
            )
        col += span

    for row in data_rows:
        ws.append(row)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Test 1 — ScrollerHash-like: full fields, status ИСПОЛНЕН
# ---------------------------------------------------------------------------

SCROLLER_HEADERS = [
    "№ п/п",
    "Номер документа",
    "Дата документа",
    "Статус документа",
    "Сумма",
    "ИНН получателя",
    "КПП получателя",
    "Наименование получателя",
    "Счет получателя",
    "БИК получателя",
    "Банк получателя",
    "ИНН плательщика",
    "КПП плательщика",
    "Наименование плательщика",
    "Счет плательщика",
    "Расшифровка п/п/Контракт (договор)",
    "Дата исполнения операции",
    "Дата и время утверждения ЦС",
]

PURPOSE_TEXT = (
    "(712ZU7L4002;0300022) СОГЛАШЕНИЕ 831-2025-ВСКС ОТ 23.06.2025 "
    "ДОГОВОР 11-26-1 ОТ 04.12.2025, УПД TU-9446 ОТ 18.12.2025 "
    "ЗАКУПКА МАТ.ЗАПАСОВ.:СТЕЛЛАЖИ НДС62223.33"
)

SCROLLER_ROW = [
    1,                      # № п/п
    "143",                  # Номер документа
    date(2026, 4, 13),      # Дата документа
    "ИСПОЛНЕН",             # Статус документа
    12090.25,               # Сумма
    "2315176820",           # ИНН получателя
    "231501001",            # КПП получателя
    "ООО Тест",             # Наименование получателя
    "40702810000000001234",  # Счет получателя
    "040349706",            # БИК получателя
    "ПАО Сбербанк",         # Банк получателя
    "7700000001",           # ИНН плательщика
    "770001001",            # КПП плательщика
    "ВСКС",                 # Наименование плательщика
    "40101810400000010041",  # Счет плательщика
    PURPOSE_TEXT,           # Расшифровка п/п/Контракт (договор)
    date(2026, 4, 14),      # Дата исполнения операции
    "14.04.2026 09:15:00",  # Дата и время утверждения ЦС
]


class TestScrollerHashLike:
    def setup_method(self):
        self.data = _make_workbook(SCROLLER_HEADERS, [SCROLLER_ROW])
        self.sheet_name, self.rows = parse_workbook(self.data)

    def test_returns_one_row(self):
        assert len(self.rows) == 1

    def test_is_executed(self):
        assert self.rows[0].is_executed is True

    def test_amount(self):
        assert self.rows[0].amount == Decimal("12090.25")

    def test_payment_number(self):
        assert self.rows[0].payment_number == "143"

    def test_payment_date(self):
        assert self.rows[0].payment_date == date(2026, 4, 13)

    def test_payee_inn(self):
        assert self.rows[0].payee_inn == "2315176820"

    def test_contract_number_prefers_dogovor_over_soglashenie(self):
        # Приоритет: ДОГОВОР > СОГЛАШЕНИЕ → должен вернуть "11-26-1"
        assert self.rows[0].parsed_contract_number == "11-26-1"

    def test_contract_date(self):
        assert self.rows[0].parsed_contract_date == date(2025, 12, 4)

    def test_kbk(self):
        assert self.rows[0].parsed_kbk == "712ZU7L4002"

    def test_purpose_text_set(self):
        assert self.rows[0].purpose_text == PURPOSE_TEXT

    def test_no_skip_reason_for_executed(self):
        assert self.rows[0].skip_reason is None

    def test_raw_json_has_all_headers(self):
        # Ключевые нормализованные поля должны присутствовать
        rj = self.rows[0].raw_json
        assert "НОМЕР ДОКУМЕНТА" in rj
        assert "СТАТУС ДОКУМЕНТА" in rj


# ---------------------------------------------------------------------------
# Test 2 — Scroller_FADM_2026-like: merged headers
# ---------------------------------------------------------------------------

class TestScrollerFADMLike:
    """Merged headers: 3× 'Реквизиты получателя', 3× 'Реквизиты плательщика'."""

    def setup_method(self):
        # header_groups: (header_text, span)
        header_groups = [
            ("№ п/п", 1),
            ("Номер документа", 1),
            ("Дата документа", 1),
            ("Статус документа", 1),
            ("Сумма", 1),
            # 3 merged cols for payee block
            ("Реквизиты получателя", 3),
            ("Расшифровка п/п/Контракт (договор)", 1),
            ("Дата исполнения операции", 1),
            # 3 merged cols for payer block
            ("Реквизиты плательщика", 3),
        ]
        # data: 12 values (1+1+1+1+1+3+1+1+3)
        # Реквизиты получателя cols: ИНН, банк, счет в одном блоке
        payee_inn = "2315176820"
        payer_inn = "7700000001"
        data_row = [
            1,
            "200",
            date(2026, 4, 20),
            "ИСПОЛНЕН",
            50000.0,
            # 3 cols payee block
            f"ИНН {payee_inn} КПП 231501001",
            "ПАО Сбербанк счет 40702810000",
            "БИК 040349706",
            # purpose
            "ДОГОВОР 55-2026 ОТ 01.01.2026 НДС0.00",
            # execution date
            date(2026, 4, 21),
            # 3 cols payer block
            f"ИНН {payer_inn} КПП 770001001 ВСКС",
            "Казначейство России",
            "л/с 40101810400000010041",
        ]
        self.data = _make_workbook_merged_headers(header_groups, [data_row])
        self.sheet_name, self.rows = parse_workbook(self.data)

    def test_returns_one_row(self):
        assert len(self.rows) == 1

    def test_is_executed(self):
        assert self.rows[0].is_executed is True

    def test_payee_inn_extracted_from_merged_block(self):
        # ИНН должен вытащиться regex'ом из объединённого текстового блока
        assert self.rows[0].payee_inn == "2315176820"

    def test_contract_number(self):
        assert self.rows[0].parsed_contract_number == "55-2026"

    def test_contract_date(self):
        assert self.rows[0].parsed_contract_date == date(2026, 1, 1)


# ---------------------------------------------------------------------------
# Test 3 — статус АННУЛИРОВАН
# ---------------------------------------------------------------------------

class TestAnnulledStatus:
    def setup_method(self):
        headers = [
            "Номер документа",
            "Дата документа",
            "Статус документа",
            "Сумма",
            "Расшифровка п/п/Контракт (договор)",
        ]
        data_row = [
            "999",
            date(2026, 3, 1),
            "АННУЛИРОВАН",
            25000.0,
            "ДОГОВОР 77-2025 ОТ 10.10.2025 НДС0.00",
        ]
        self.data = _make_workbook(headers, [data_row])
        self.sheet_name, self.rows = parse_workbook(self.data)

    def test_returns_one_row(self):
        # Строка сохраняется (для журнала), но с is_executed=False
        assert len(self.rows) == 1

    def test_is_not_executed(self):
        assert self.rows[0].is_executed is False

    def test_skip_reason_set(self):
        assert self.rows[0].skip_reason is not None
        assert "АННУЛИРОВАН" in self.rows[0].skip_reason

    def test_payment_number_still_parsed(self):
        assert self.rows[0].payment_number == "999"

    def test_amount_still_parsed(self):
        assert self.rows[0].amount == Decimal("25000")


# ---------------------------------------------------------------------------
# Test 4 — multi-row split: два purpose_text для одной строки
# ---------------------------------------------------------------------------

class TestMultiRowSplit:
    """Два столбца 'Расшифровка п/п/Контракт (договор)' → 2 ParsedRow."""

    def setup_method(self):
        headers = [
            "Номер документа",
            "Дата документа",
            "Статус документа",
            "Сумма",
            "ИНН плательщика",
            # Два столбца с одинаковым именем — multi-split
            "Расшифровка п/п/Контракт (договор)",
            "Расшифровка п/п/Контракт (договор)",
        ]
        data_row = [
            "301",
            date(2026, 4, 1),
            "ИСПОЛНЕН",
            100000.0,
            "7700000001",
            "ДОГОВОР 11-2026 ОТ 01.02.2026 НДС0.00",
            "ДОГОВОР 22-2026 ОТ 15.03.2026 НДС0.00",
        ]
        self.data = _make_workbook(headers, [data_row])
        self.sheet_name, self.rows = parse_workbook(self.data)

    def test_returns_two_rows(self):
        assert len(self.rows) == 2

    def test_same_source_row_hash(self):
        assert self.rows[0].source_row_hash == self.rows[1].source_row_hash

    def test_different_contract_numbers(self):
        numbers = {r.parsed_contract_number for r in self.rows}
        assert "11-2026" in numbers
        assert "22-2026" in numbers

    def test_both_executed(self):
        assert all(r.is_executed for r in self.rows)

    def test_both_have_same_amount(self):
        assert self.rows[0].amount == self.rows[1].amount == Decimal("100000")


# ---------------------------------------------------------------------------
# Unit tests for parse_purpose directly
# ---------------------------------------------------------------------------

class TestParsePurpose:
    def test_prefers_dogovor_over_soglashenie(self):
        text = (
            "(712ZU7L4002;0300022) СОГЛАШЕНИЕ 831-2025-ВСКС ОТ 23.06.2025 "
            "ДОГОВОР 11-26-1 ОТ 04.12.2025"
        )
        result = parse_purpose(text)
        assert result["contract_number"] == "11-26-1"
        assert result["contract_date"] == date(2025, 12, 4)
        assert result["kbk"] == "712ZU7L4002"

    def test_soglashenie_fallback(self):
        text = "СОГЛАШЕНИЕ 831-2025-ВСКС ОТ 23.06.2025 НДС0.00"
        result = parse_purpose(text)
        assert result["contract_number"] == "831-2025-ВСКС"
        assert result["contract_date"] == date(2025, 6, 23)

    def test_reestr(self):
        text = "РЕЕСТР ДОК.-ОСН. 655 ОТ 29.12.2025 ЗАКУПКА УСЛУГ."
        result = parse_purpose(text)
        assert result["contract_number"] == "655"
        assert result["contract_date"] == date(2025, 12, 29)

    def test_kbk_extracted(self):
        text = "(712ZU7L4002;0300022) что-то тут"
        result = parse_purpose(text)
        assert result["kbk"] == "712ZU7L4002"

    def test_vat_extracted(self):
        text = "ДОГОВОР 1-2025 ОТ 01.01.2025 НДС62223.33"
        result = parse_purpose(text)
        assert result.get("vat") == Decimal("62223.33")

    def test_empty_text_returns_empty_dict(self):
        assert parse_purpose("") == {}

    def test_last_dogovor_wins(self):
        """Два ДОГОВОР в тексте — берём последний."""
        text = "ДОГОВОР 10-2025 ОТ 01.01.2025 ДОГОВОР 20-2025 ОТ 01.06.2025"
        result = parse_purpose(text)
        assert result["contract_number"] == "20-2025"
        assert result["contract_date"] == date(2025, 6, 1)
