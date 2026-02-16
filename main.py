"""
Главное приложение CRM системы для управления договорами
Локальная версия с графическим интерфейсом
"""

import sys
import os
from PyQt6.QtWidgets import (QApplication, QMainWindow, QWidget, QVBoxLayout,
                             QHBoxLayout, QTabWidget, QPushButton, QLabel,
                             QLineEdit, QTextEdit, QComboBox, QDateEdit, QSpinBox,
                             QDoubleSpinBox, QTableWidget, QTableWidgetItem, QCheckBox,
                             QGroupBox, QScrollArea, QMessageBox, QFileDialog, QDialog,
                             QDialogButtonBox, QFormLayout, QHeaderView, QGridLayout,
                             QAbstractItemView, QAbstractScrollArea, QCompleter)
from PyQt6.QtCore import Qt, QDate, QTimer, QStringListModel
from PyQt6.QtGui import QFont, QColor, QPalette, QKeySequence, QShortcut
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import pandas as pd
from import_data import DataImporter
from db_sqlite import (
    init_db as init_sqlite_db,
    get_groups as feo_get_groups,
    get_directions_by_group as feo_get_directions_by_group,
    get_cost_items_for_direction as feo_get_cost_items_for_direction,
    get_unique_items_for_cost_item as feo_get_unique_items_for_cost_item,
    get_all_unique_items as feo_get_all_unique_items,
    get_cost_item_details as feo_get_cost_item_details,
)
from import_feo_sqlite import import_feo_to_sqlite

# Инициализируем локальную БД SQLite для ФЭО и справочников
init_sqlite_db()

class DatabaseManager:
    """Менеджер для работы с базой данных Excel"""
    
    def __init__(self, db_file='CRM_База_Данных.xlsx'):
        # Всегда работаем с файлом базы данных, лежащим рядом с main.py.
        # Это защищает от ситуации, когда скрипт запускают из другой папки
        # и по относительному пути создаётся/читается другой Excel.
        if os.path.isabs(db_file):
            self.db_file = db_file
        else:
            base_dir = os.path.dirname(os.path.abspath(__file__))
            self.db_file = os.path.join(base_dir, db_file)
        self.wb = None
        self.load_database()
    
    def load_database(self):
        """Загрузка базы данных"""
        if os.path.exists(self.db_file):
            self.wb = load_workbook(self.db_file)
        else:
            # Создаем структуру если файла нет
            try:
                from create_excel_structure import create_database_structure
                create_database_structure(self.db_file)
                self.wb = load_workbook(self.db_file)
            except Exception as e:
                print(f"Ошибка при создании структуры БД: {e}")
                raise
        self.ensure_feo_sheets()
    
    def save_database(self):
        """Сохранение базы данных"""
        if self.wb:
            self.wb.save(self.db_file)
    
    def get_sheet_data(self, sheet_name):
        """Получение данных из листа"""
        if sheet_name not in self.wb.sheetnames:
            return []
        
        ws = self.wb[sheet_name]
        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if any(row):  # Пропускаем пустые строки
                data.append(row)
        return data
    
    def get_next_id(self, sheet_name, id_column=1):
        """Получение следующего ID"""
        if sheet_name not in self.wb.sheetnames:
            return 1
        
        ws = self.wb[sheet_name]
        last_row = ws.max_row
        if last_row == 1:
            return 1
        
        last_id = ws.cell(row=last_row, column=id_column).value
        if last_id is None:
            return 1
        
        try:
            return int(last_id) + 1
        except:
            return 1
    
    def add_contract(self, contract_data):
        """Добавление договора"""
        ws = self.wb['GoodsService']
        next_row = ws.max_row + 1
        next_id = self.get_next_id('GoodsService')
        
        # Заполняем данные
        ws.cell(row=next_row, column=1, value=next_id)  # ID
        ws.cell(row=next_row, column=2, value=contract_data.get('contract_number', ''))
        ws.cell(row=next_row, column=3, value=contract_data.get('contract_date'))
        ws.cell(row=next_row, column=4, value=contract_data.get('contract_type', ''))
        ws.cell(row=next_row, column=5, value=contract_data.get('contract_kind', ''))
        ws.cell(row=next_row, column=6, value=contract_data.get('subsidy_id'))
        ws.cell(row=next_row, column=7, value=contract_data.get('purchase_number', ''))
        ws.cell(row=next_row, column=8, value=contract_data.get('order_number', ''))
        ws.cell(row=next_row, column=9, value=contract_data.get('subject', ''))
        ws.cell(row=next_row, column=10, value=contract_data.get('description', ''))
        ws.cell(row=next_row, column=11, value=contract_data.get('contractor_id'))
        ws.cell(row=next_row, column=12, value=contract_data.get('status', 'Плановый'))
        ws.cell(row=next_row, column=13, value=contract_data.get('purchase_status', ''))
        ws.cell(row=next_row, column=14, value=contract_data.get('execution_stage', ''))
        ws.cell(row=next_row, column=15, value=contract_data.get('start_date'))
        ws.cell(row=next_row, column=16, value=contract_data.get('end_date'))
        ws.cell(row=next_row, column=17, value=contract_data.get('execution_days'))
        ws.cell(row=next_row, column=18, value=contract_data.get('nmck', 0))
        ws.cell(row=next_row, column=19, value=contract_data.get('price_without_vat', 0))
        ws.cell(row=next_row, column=29, value=contract_data.get('vat_applied', 'Нет'))
        ws.cell(row=next_row, column=30, value=contract_data.get('vat_rate', 20))
        ws.cell(row=next_row, column=31, value=contract_data.get('payment_method', ''))
        ws.cell(row=next_row, column=32, value=contract_data.get('payment_form', ''))
        ws.cell(row=next_row, column=33, value=contract_data.get('advance_amount', 0))
        ws.cell(row=next_row, column=34, value=contract_data.get('payment_term', 0))
        ws.cell(row=next_row, column=35, value=contract_data.get('feo_direction', ''))
        ws.cell(row=next_row, column=36, value=contract_data.get('feo_type', ''))
        ws.cell(row=next_row, column=37, value=contract_data.get('app_direction', ''))
        ws.cell(row=next_row, column=38, value=contract_data.get('specific_type', ''))
        ws.cell(row=next_row, column=39, value=contract_data.get('responsible', ''))
        ws.cell(row=next_row, column=40, value=contract_data.get('city', ''))
        ws.cell(row=next_row, column=41, value=contract_data.get('comments', ''))
        ws.cell(row=next_row, column=42, value=datetime.now())
        ws.cell(row=next_row, column=43, value=datetime.now())
        ws.cell(row=next_row, column=44, value=contract_data.get('author', ''))
        
        # Добавляем формулы
        from create_excel_structure import add_formulas_to_goods_service
        add_formulas_to_goods_service(ws, next_row)
        
        self.save_database()
        return next_id
    
    def get_subsidies(self):
        """Получение списка субсидий.

        Более надёжная версия:
        - ищет лист с именем 'Субсидии';
        - определяет колонку ID и колонку Наименования по заголовкам в первой строке;
        - возвращает список кортежей (id, name, *остальные_колонки*).
        """
        if not self.wb or 'Субсидии' not in self.wb.sheetnames:
            return []

        ws = self.wb['Субсидии']
        header_row = list(ws.iter_rows(min_row=1, max_row=1, values_only=True))[0]

        id_col_idx = None
        name_col_idx = None

        for idx, value in enumerate(header_row):
            if not value:
                continue
            text = str(value).strip().lower()
            if id_col_idx is None and ('id' == text or text.startswith('id ')):
                id_col_idx = idx
            if name_col_idx is None and ('наимен' in text or 'субсид' in text):
                name_col_idx = idx

        # Если не нашли заголовки, падаем обратно к старой логике
        if name_col_idx is None:
            return self.get_sheet_data('Субсидии')

        data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            # ID может быть пустым — в UI он всё равно используется только как data
            sub_id = row[id_col_idx] if id_col_idx is not None and id_col_idx < len(row) else None
            name = row[name_col_idx] if name_col_idx < len(row) else None
            if not name:
                continue
            # Собираем кортеж: (id, name, *остальное как есть*)
            # Чтобы не ломать остальной код, который ожидает, что name в row[1]
            row_list = list(row)
            # Обеспечиваем, что на позиции 0 и 1 именно id и name
            # (если структура другая, просто подставляем поверх)
            if len(row_list) < 2:
                row_list += [None] * (2 - len(row_list))
            row_list[0] = sub_id
            row_list[1] = name
            data.append(tuple(row_list))

        return data
    
    def get_contractors(self):
        """Получение списка контрагентов"""
        return self.get_sheet_data('Контрагенты')
    
    def get_categories_feo(self):
        """Получение категорий ФЭО"""
        if 'Категории_из_ФЭО' not in self.wb.sheetnames:
            return []
        ws = self.wb['Категории_из_ФЭО']
        categories = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Первый столбец - категория
                categories.append(row[0])
        return categories
    
    def get_categories_app(self):
        """Получение категорий из приложения"""
        if 'Категории_из_приложения' not in self.wb.sheetnames:
            return []
        ws = self.wb['Категории_из_приложения']
        categories = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                categories.append(row[0])
        return categories
    
    def get_specific_types(self):
        """Получение списка конкретизированных типов расходов"""
        # Ищем в листе GoodsService все уникальные значения
        if 'GoodsService' not in self.wb.sheetnames:
            return []
        ws = self.wb['GoodsService']
        specific_types = set()
        # Столбец AL (38) - Тип конкретизированный
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 37 and row[37]:
                value = str(row[37]).strip()
                if value:
                    specific_types.add(value)
        return sorted(list(specific_types))
    
    def get_feo_directions_from_excel(self, group_name=None):
        """Получение направлений ФЭО из справочника Excel"""
        if 'Справочник_Направления_ФЭО' not in self.wb.sheetnames:
            return []
        ws = self.wb['Справочник_Направления_ФЭО']
        directions = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            direction_id = row[0]
            group = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            direction_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if not direction_name:
                continue
            if group_name and group != group_name:
                continue
            directions.append({
                'id': direction_id,
                'group': group,
                'name': direction_name,
            })
        return directions
    
    def get_feo_cost_items_from_excel(self, group_name=None, direction_id=None):
        """Получение статей затрат из справочника Excel"""
        if 'Справочник_Статьи_Затрат' not in self.wb.sheetnames:
            return []
        ws = self.wb['Справочник_Статьи_Затрат']
        cost_items = []
        
        # Если указано направление, получаем связанные статьи из листа связей
        linked_item_ids = set()
        if direction_id:
            if 'Связи_Направление_Статья' in self.wb.sheetnames:
                links_ws = self.wb['Связи_Направление_Статья']
                for row in links_ws.iter_rows(min_row=2, values_only=True):
                    if not row or not any(row):
                        continue
                    link_direction_id = row[1]
                    item_id = row[3]
                    if link_direction_id == direction_id and item_id:
                        linked_item_ids.add(item_id)
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            item_id = row[0]
            group = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            item_name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if not item_name:
                continue
            if group_name and group != group_name:
                continue
            # Если указано направление, показываем только связанные статьи
            if direction_id and item_id not in linked_item_ids:
                continue
            unit = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            plan_price = None
            if len(row) > 4 and row[4]:
                try:
                    plan_price = float(row[4])
                except (TypeError, ValueError):
                    pass
            cost_items.append({
                'id': item_id,
                'group': group,
                'name': item_name,
                'unit': unit,
                'plan_price': plan_price,
            })
        return cost_items
    
    def get_unique_items_from_excel(self, cost_item_id=None):
        """Получение уникальных товаров/услуг из справочника Excel"""
        if 'Справочник_Уникальные_Товары' not in self.wb.sheetnames:
            return []
        ws = self.wb['Справочник_Уникальные_Товары']
        items = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            item_id = row[0]
            name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            if not name:
                continue
            description = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            item_cost_item_id = row[3] if len(row) > 3 and row[3] else None
            # Если указана статья затрат, фильтруем по ней
            if cost_item_id and item_cost_item_id != cost_item_id:
                continue
            items.append({
                'id': item_id,
                'name': name,
                'description': description,
                'cost_item_id': item_cost_item_id,
            })
        return items
    
    def get_feo_groups_from_excel(self):
        """Получение списка групп ФЭО из справочников"""
        groups = set()
        # Получаем группы из направлений
        directions = self.get_feo_directions_from_excel()
        for d in directions:
            if d.get('group'):
                groups.add(d['group'])
        # Получаем группы из статей затрат
        cost_items = self.get_feo_cost_items_from_excel()
        for ci in cost_items:
            if ci.get('group'):
                groups.add(ci['group'])
        return sorted(list(groups))
    
    def get_app_directions_by_cost_item(self, cost_item_name):
        """Получение направлений расходов из приложения по наименованию статьи затрат"""
        if 'Справочник_Направления_ФЭО' not in self.wb.sheetnames:
            return []
        ws = self.wb['Справочник_Направления_ФЭО']
        directions = set()
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            # Столбец "Группа" содержит наименование статьи затрат
            cost_item = str(row[1]).strip() if len(row) > 1 and row[1] else ''
            # Столбец "Направление расходования ФЭО" содержит направление расходов из приложения
            direction = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            if cost_item == cost_item_name and direction:
                directions.add(direction)
        return sorted(list(directions))
    
    def check_purchase_number_exists(self, purchase_number, exclude_contract_id=None):
        """Проверка существования номера закупки"""
        if not purchase_number or not purchase_number.strip():
            return False, None
        if 'GoodsService' not in self.wb.sheetnames:
            return False, None
        ws = self.wb['GoodsService']
        purchase_number = purchase_number.strip()
        # Столбец G (7) - Номер закупки
        for row in ws.iter_rows(min_row=2, values_only=True):
            if len(row) > 6 and row[6]:
                existing_purchase = str(row[6]).strip()
                contract_id = row[0] if row[0] else None
                if existing_purchase == purchase_number:
                    if exclude_contract_id and contract_id == exclude_contract_id:
                        continue
                    return True, contract_id
        return False, None
    
    def get_framework_contracts(self):
        """Получение рамочных договоров с остатками"""
        if 'GoodsService' not in self.wb.sheetnames:
            return []
        ws = self.wb['GoodsService']
        framework_contracts = []
        contractor_map = self._get_reference_map('Контрагенты', key_col=1, value_col=2)
        # Столбец E (5) - Вид договора, столбец AA (27) - Остаток к оплате
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) > 4 and row[4] and 'рамочн' in str(row[4]).lower():
                contract_id = row[0]
                contract_number = row[1] if len(row) > 1 else ''
                subject = row[8] if len(row) > 8 else ''
                contractor_id = row[10] if len(row) > 10 else None
                purchase_number = row[6] if len(row) > 6 else ''
                # Получаем остаток из столбца 27
                try:
                    remainder_cell = ws.cell(row=idx, column=27)
                    remainder = remainder_cell.value if remainder_cell.value else 0
                    if isinstance(remainder, str):
                        remainder = 0
                    remainder = float(remainder) if remainder else 0
                except:
                    remainder = 0
                
                # Получаем название контрагента
                contractor_name = contractor_map.get(contractor_id, '') if contractor_id else ''
                
                framework_contracts.append({
                    'id': contract_id,
                    'number': contract_number,
                    'subject': subject,
                    'contractor': contractor_name,
                    'remainder': remainder,
                    'purchase_number': purchase_number
                })
        return framework_contracts

    def get_contracts_for_registry(self):
        """Получение данных для реестра договоров"""
        ws = self.wb['GoodsService']
        contractor_map = self._get_reference_map('Контрагенты', key_col=1, value_col=2)
        subsidy_map = self._get_reference_map('Субсидии', key_col=1, value_col=2)
        
        contracts = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            contract_number = row[1]
            contractor_id = row[10]
            subsidy_id = row[5]
            contracts.append({
                'number': contract_number or '',
                'contractor': contractor_map.get(contractor_id, ''),
                'subject': row[8] or '',
                'subsidy': subsidy_map.get(subsidy_id, ''),
                'amount': row[20] or 0,
                'contracted': row[23] or 0,
                'delivered': row[24] or 0,
                'paid': row[25] or 0,
                'status': row[11] or '',
                'stage': row[13] or ''
            })
        return contracts

    def _get_reference_map(self, sheet_name, key_col, value_col):
        """Получение словаря справочника по ID"""
        ref_map = {}
        if sheet_name not in self.wb.sheetnames:
            return ref_map
        ws = self.wb[sheet_name]
        for row in ws.iter_rows(min_row=2, values_only=True):
            key = row[key_col - 1]
            value = row[value_col - 1]
            if key:
                ref_map[key] = value
        return ref_map

    @staticmethod
    def _to_float(value):
        if value is None:
            return 0.0
        if isinstance(value, (int, float)):
            return float(value)
        try:
            text = str(value).strip().replace(' ', '').replace(',', '.')
            return float(text) if text else 0.0
        except (ValueError, TypeError):
            return 0.0

    def calculate_feo_amount(self, feo_id):
        """Расчет суммы приложений для ФЭО"""
        if not feo_id:
            return 0.0
        total = 0.0
        for entry in self.get_feo_app_entries(feo_id):
            total += entry.get('total_amount') or 0.0
        return total

    def update_feo_base_amount(self, feo_id):
        """Обновление суммы в таблице базового ФЭО"""
        if not feo_id or 'ФЭО_База' not in self.wb.sheetnames:
            return
        ws = self.wb['ФЭО_База']
        amount = self.calculate_feo_amount(feo_id)
        now = datetime.now()
        for row in range(2, ws.max_row + 1):
            if ws.cell(row, column=1).value == feo_id:
                ws.cell(row=row, column=6, value=amount)
                ws.cell(row=row, column=8, value=now)
                break
        self.save_database()

    def populate_subsidy_feo_links(self):
        """Первичное заполнение листа Связи_Субсидия_ФЭО по данным из ФЭО_База.

        Каждая уникальная пара (Субсидия_ID, Субсидия, Направление расходов)
        превращается в отдельную запись.
        """
        if 'ФЭО_База' not in self.wb.sheetnames or 'Связи_Субсидия_ФЭО' not in self.wb.sheetnames:
            return

        ws_base = self.wb['ФЭО_База']
        ws_links = self.wb['Связи_Субсидия_ФЭО']

        existing_keys = set()
        # Собираем уже существующие связи, чтобы не дублировать при повторных вызовах
        if ws_links.max_row > 1:
            for row in ws_links.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                sub_id = row[1]
                direction = row[4] if len(row) > 4 else None
                if sub_id and direction:
                    existing_keys.add((sub_id, str(direction).strip()))

        new_rows = []
        for row in ws_base.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            feo_id = row[0]
            sub_id = row[1]
            sub_name = row[2]
            direction = row[4]
            if not sub_id or not direction:
                continue
            key = (sub_id, str(direction).strip())
            if key in existing_keys:
                continue
            existing_keys.add(key)
            new_rows.append((sub_id, sub_name, None, direction, feo_id))

        if not new_rows:
            return

        # Вычисляем следующий ID
        next_id = 1
        if ws_links.max_row > 1:
            try:
                last_id = ws_links.cell(row=ws_links.max_row, column=1).value
                if last_id:
                    next_id = int(last_id) + 1
            except Exception:
                next_id = ws_links.max_row

        row_idx = ws_links.max_row + 1
        for sub_id, sub_name, feo_dir_id, direction, feo_id in new_rows:
            ws_links.cell(row=row_idx, column=1, value=next_id)
            ws_links.cell(row=row_idx, column=2, value=sub_id)
            ws_links.cell(row=row_idx, column=3, value=sub_name)
            ws_links.cell(row=row_idx, column=4, value=feo_dir_id)
            ws_links.cell(row=row_idx, column=5, value=direction)
            row_idx += 1
            next_id += 1

        self.save_database()

    def ensure_feo_sheets(self):
        """Создание недостающих листов для работы с ФЭО"""
        created = False
        if 'ФЭО_База' not in self.wb.sheetnames:
            from create_excel_structure import create_feo_base_sheet
            create_feo_base_sheet(self.wb)
            created = True
        if 'ФЭО_Приложения' not in self.wb.sheetnames:
            from create_excel_structure import create_feo_applications_sheet
            create_feo_applications_sheet(self.wb)
            created = True
        if 'ФЭО_Версии' not in self.wb.sheetnames:
            ws = self.wb.create_sheet('ФЭО_Версии')
            headers = [
                'ID', 'Дата', 'Источник файла', 'Действие',
                'Лист ФЭО_База', 'Лист ФЭО_Приложения'
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, size=11)
                cell.fill = PatternFill(
                    start_color='366092', end_color='366092', fill_type='solid'
                )
                cell.font = Font(bold=True, color='FFFFFF', size=11)
                cell.alignment = Alignment(
                    horizontal='center', vertical='center', wrap_text=True
                )
            ws.freeze_panes = 'A2'
            created = True
        # Дополнительные справочные листы и связи для ФЭО
        from create_excel_structure import (
            create_links_subsidy_feo_sheet,
            create_links_feo_cost_items_sheet,
            create_cost_breakdown_sheet,
            create_breakdown_items_sheet,
            create_feo_directions_reference_sheet,
            create_feo_cost_items_reference_sheet,
            create_unique_items_reference_sheet,
            create_direction_cost_item_links_sheet,
        )
        if 'Связи_Субсидия_ФЭО' not in self.wb.sheetnames:
            create_links_subsidy_feo_sheet(self.wb)
            # Пытаемся сразу наполнить связями из листа ФЭО_База (если он уже есть)
            try:
                self.populate_subsidy_feo_links()
            except Exception:
                pass
            created = True
        if 'Связи_ФЭО_Статьи' not in self.wb.sheetnames:
            create_links_feo_cost_items_sheet(self.wb)
            created = True
        if 'Статьи_Разбивка' not in self.wb.sheetnames:
            create_cost_breakdown_sheet(self.wb)
            created = True
        if 'Разбивка_Товары' not in self.wb.sheetnames:
            create_breakdown_items_sheet(self.wb)
            created = True
        # Справочники для ручного заполнения
        if 'Справочник_Направления_ФЭО' not in self.wb.sheetnames:
            create_feo_directions_reference_sheet(self.wb)
            created = True
        if 'Справочник_Статьи_Затрат' not in self.wb.sheetnames:
            create_feo_cost_items_reference_sheet(self.wb)
            created = True
        if 'Справочник_Уникальные_Товары' not in self.wb.sheetnames:
            create_unique_items_reference_sheet(self.wb)
            created = True
        if 'Связи_Направление_Статья' not in self.wb.sheetnames:
            create_direction_cost_item_links_sheet(self.wb)
            created = True
        if created:
            self.save_database()

    def archive_feo_version(self, source_file: str | None = None, action: str = 'import'):
        """Сохранение копии текущего базового ФЭО перед изменениями"""
        if not self.wb:
            return
        if 'ФЭО_Версии' not in self.wb.sheetnames:
            self.ensure_feo_sheets()
        if 'ФЭО_База' not in self.wb.sheetnames and 'ФЭО_Приложения' not in self.wb.sheetnames:
            return

        versions_ws = self.wb['ФЭО_Версии']
        next_id = self.get_next_id('ФЭО_Версии')
        timestamp = datetime.now()

        base_copy_name = ''
        apps_copy_name = ''

        # Создаём копию листа ФЭО_База
        if 'ФЭО_База' in self.wb.sheetnames:
            base_ws = self.wb['ФЭО_База']
            base_copy = self.wb.copy_worksheet(base_ws)
            base_copy_name = f"ФЭО_База_v{next_id}"
            base_copy.title = base_copy_name

        # Создаём копию листа ФЭО_Приложения
        if 'ФЭО_Приложения' in self.wb.sheetnames:
            apps_ws = self.wb['ФЭО_Приложения']
            apps_copy = self.wb.copy_worksheet(apps_ws)
            apps_copy_name = f"ФЭО_Приложения_v{next_id}"
            apps_copy.title = apps_copy_name

        row_idx = versions_ws.max_row + 1
        versions_ws.cell(row=row_idx, column=1, value=next_id)
        versions_ws.cell(row=row_idx, column=2, value=timestamp)
        versions_ws.cell(row=row_idx, column=3, value=source_file or '')
        versions_ws.cell(row=row_idx, column=4, value=action)
        versions_ws.cell(row=row_idx, column=5, value=base_copy_name)
        versions_ws.cell(row=row_idx, column=6, value=apps_copy_name)

        self.save_database()

    def get_feo_base_entries(self):
        """Получение записей базового ФЭО"""
        if 'ФЭО_База' not in self.wb.sheetnames:
            return []
        ws = self.wb['ФЭО_База']
        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            entry_id = row[0]
            amount = self.calculate_feo_amount(entry_id) if entry_id else self._to_float(row[5])
            entries.append({
                'id': entry_id,
                'subsidy_id': row[1],
                'subsidy_name': row[2],
                'application_number': row[3],
                'direction': row[4],
                'amount': amount,
                'created_at': row[6],
                'updated_at': row[7],
            })
        return entries

    def save_feo_base_entries(self, entries):
        """Сохранение записей базового ФЭО"""
        if 'ФЭО_База' not in self.wb.sheetnames:
            return
        ws = self.wb['ФЭО_База']
        existing_ids = [int(e['id']) for e in entries if e.get('id')]
        max_id = max(existing_ids, default=0)
        now = datetime.now()
        for entry in entries:
            if not entry.get('id'):
                max_id += 1
                entry['id'] = max_id
                entry['created_at'] = now
            else:
                entry['created_at'] = entry.get('created_at') or now
            entry['updated_at'] = now
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        row_idx = 2
        for entry in entries:
            if entry.get('id'):
                entry['amount'] = self.calculate_feo_amount(entry['id'])
            else:
                entry['amount'] = entry.get('amount', 0)
            ws.cell(row=row_idx, column=1, value=entry['id'])
            ws.cell(row=row_idx, column=2, value=entry.get('subsidy_id'))
            ws.cell(row=row_idx, column=3, value=entry.get('subsidy_name'))
            ws.cell(row=row_idx, column=4, value=entry.get('application_number'))
            ws.cell(row=row_idx, column=5, value=entry.get('direction'))
            ws.cell(row=row_idx, column=6, value=entry.get('amount'))
            ws.cell(row=row_idx, column=7, value=entry.get('created_at'))
            ws.cell(row=row_idx, column=8, value=entry.get('updated_at'))
            row_idx += 1
        valid_ids = {entry['id'] for entry in entries if entry.get('id')}
        self.cleanup_feo_applications(valid_ids)
        self.save_database()

    def get_feo_app_entries(self, feo_id=None):
        """Получение записей приложений ФЭО"""
        if 'ФЭО_Приложения' not in self.wb.sheetnames:
            return []
        ws = self.wb['ФЭО_Приложения']
        entries = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not any(row):
                continue
            quantity = self._to_float(row[6])
            unit_price = self._to_float(row[7])
            total_amount = self._to_float(row[8])
            if total_amount == 0 and quantity and unit_price:
                total_amount = quantity * unit_price
            entry = {
                'id': row[0],
                'feo_id': row[1],
                'app_name': row[2],
                'direction': row[3],
                'method': row[4],
                'unit': row[5],
                'quantity': quantity,
                'unit_price': unit_price,
                'total_amount': total_amount,
                'created_at': row[9],
                'updated_at': row[10],
            }
            if feo_id and entry['feo_id'] != feo_id:
                continue
            entries.append(entry)
        return entries

    def save_feo_app_entries(self, feo_id, entries):
        """Сохранение приложений для конкретного ФЭО"""
        if 'ФЭО_Приложения' not in self.wb.sheetnames or not feo_id:
            return
        all_entries = self.get_feo_app_entries()
        other_entries = [e for e in all_entries if e.get('feo_id') != feo_id]
        max_id = max([e['id'] for e in all_entries if e.get('id')], default=0)
        now = datetime.now()
        prepared = []
        for entry in entries:
            if not entry.get('id'):
                max_id += 1
                entry['id'] = max_id
                entry['created_at'] = now
            else:
                entry['created_at'] = entry.get('created_at') or now
            entry['updated_at'] = now
            entry['feo_id'] = feo_id
            prepared.append(entry)
        combined = other_entries + prepared
        self._write_feo_app_sheet(combined)
        self.update_feo_base_amount(feo_id)

    def cleanup_feo_applications(self, valid_base_ids):
        """Удаление приложений, которые больше не связаны с ФЭО"""
        if 'ФЭО_Приложения' not in self.wb.sheetnames:
            return
        if not valid_base_ids:
            ws = self.wb['ФЭО_Приложения']
            if ws.max_row > 1:
                ws.delete_rows(2, ws.max_row - 1)
            return
        entries = [entry for entry in self.get_feo_app_entries() if entry.get('feo_id') in valid_base_ids]
        self._write_feo_app_sheet(entries)

    def _write_feo_app_sheet(self, entries):
        """Перезапись листа приложений"""
        ws = self.wb['ФЭО_Приложения']
        if ws.max_row > 1:
            ws.delete_rows(2, ws.max_row - 1)
        row_idx = 2
        for entry in entries:
            ws.cell(row=row_idx, column=1, value=entry.get('id'))
            ws.cell(row=row_idx, column=2, value=entry.get('feo_id'))
            ws.cell(row=row_idx, column=3, value=entry.get('app_name'))
            ws.cell(row=row_idx, column=4, value=entry.get('direction'))
            ws.cell(row=row_idx, column=5, value=entry.get('method'))
            ws.cell(row=row_idx, column=6, value=entry.get('unit'))
            ws.cell(row=row_idx, column=7, value=entry.get('quantity'))
            ws.cell(row=row_idx, column=8, value=entry.get('unit_price'))
            ws.cell(row=row_idx, column=9, value=entry.get('total_amount'))
            ws.cell(row=row_idx, column=10, value=entry.get('created_at'))
            ws.cell(row=row_idx, column=11, value=entry.get('updated_at'))
            row_idx += 1

class MainWindow(QMainWindow):
    CREATED_ROLE = Qt.ItemDataRole.UserRole + 1
    UPDATED_ROLE = Qt.ItemDataRole.UserRole + 2
    """Главное окно приложения"""
    
    def __init__(self):
        super().__init__()
        self.db = DatabaseManager()
        self.registry_table = None
        self.registry_subsidy_tabs = None
        self.registry_subsidy_tab_names = []
        self.current_registry_subsidy_name = None  # None = все субсидии
        self.subsidy_options = []
        self.feo_base_table = None
        self.feo_app_table = None
        self.feo_app_base_combo = None
        self.subsidy_lookup = {}
        self.feo_base_entries = []
        self.feo_version_label = None
        self.feo_subsidy_tabs = None
        self.feo_subsidy_tab_ids = []
        self.current_feo_subsidy_id = None  # None = все субсидии
        self._app_table_updating = False
        self._shortcuts = []
        self.init_ui()
    
    def init_ui(self):
        """Инициализация интерфейса"""
        self.setWindowTitle('CRM Система - Управление договорами')
        self.setGeometry(100, 100, 1200, 750)
        self.setMinimumSize(1000, 600)  # Минимальный размер окна
        
        # Центральный виджет
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Главный layout
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # Заголовок
        title = QLabel('CRM Система - Управление договорами и подрядчиками')
        title.setFont(QFont('Arial', 18, QFont.Weight.Bold))
        title.setAlignment(Qt.AlignmentFlag.AlignCenter)
        title.setStyleSheet('padding: 15px; background-color: #1E3A5F; color: #FFFFFF; border-radius: 5px;')
        main_layout.addWidget(title)
        
        # Вкладки
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane {
                border: 1px solid #1E3A5F;
                background: #E6F2FF;
            }
            QTabBar::tab {
                background: #5A7FA8;
                color: #FFFFFF;
                padding: 10px 20px;
                margin-right: 2px;
                border-top-left-radius: 5px;
                border-top-right-radius: 5px;
            }
            QTabBar::tab:selected {
                background: #1E3A5F;
                color: #FFFFFF;
                font-weight: bold;
            }
            QTabBar::tab:hover {
                background: #4A6F98;
            }
        """)
        
        # Создаем вкладки
        self.create_contract_tab()
        self.create_contractors_tab()
        self.create_dashboard_tab()
        self.create_registry_tab()
        self.create_feo_tab()
        
        main_layout.addWidget(self.tabs)
        
        # Статусная строка
        self.statusBar().showMessage('Готово к работе')
    
    def create_contract_tab(self):
        """Создание вкладки для работы с договорами"""
        tab = QWidget()
        layout = QVBoxLayout()
        tab.setLayout(layout)
        
        # Прокручиваемая область
        scroll = QScrollArea()
        scroll_widget = QWidget()
        scroll_layout = QVBoxLayout()
        scroll_widget.setLayout(scroll_layout)
        
        # Основная информация
        basic_group = self.create_basic_info_group()
        scroll_layout.addWidget(basic_group)
        
        # Контрагент
        contractor_group = self.create_contractor_group()
        scroll_layout.addWidget(contractor_group)
        
        # Состав договора
        items_group = self.create_contract_items_group()
        scroll_layout.addWidget(items_group)
        
        # Финансы
        finance_group = self.create_finance_group()
        scroll_layout.addWidget(finance_group)
        
        # Категории ФЭО (с чекбоксами)
        feo_group = self.create_feo_categories_group()
        scroll_layout.addWidget(feo_group)
        
        # Дополнительно
        additional_group = self.create_additional_group()
        scroll_layout.addWidget(additional_group)
        
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        
        # Кнопки
        buttons_layout = QHBoxLayout()
        save_btn = QPushButton('💾 Сохранить договор')
        save_btn.setStyleSheet('padding: 10px; font-size: 14px; background-color: #4CAF50; color: #FFFFFF; font-weight: bold; border-radius: 5px;')
        save_btn.clicked.connect(self.save_contract)
        
        clear_btn = QPushButton('🗑️ Очистить форму')
        clear_btn.setStyleSheet('padding: 10px; font-size: 14px; background-color: #f44336; color: #FFFFFF; font-weight: bold; border-radius: 5px;')
        clear_btn.clicked.connect(self.clear_contract_form)
        
        buttons_layout.addWidget(save_btn)
        buttons_layout.addWidget(clear_btn)
        buttons_layout.addStretch()
        
        layout.addWidget(scroll)
        layout.addLayout(buttons_layout)
        
        self.tabs.addTab(tab, '📄 Договоры')
    
    def create_basic_info_group(self):
        """Группа основной информации"""
        group = QGroupBox('Основная информация')
        layout = QFormLayout()
        
        self.contract_number = QLineEdit()
        self.contract_date = QDateEdit()
        self.contract_date.setDate(QDate.currentDate())
        self.contract_date.setCalendarPopup(True)
        self.contract_date.setDisplayFormat('dd.MM.yyyy')
        self.contract_date.setButtonSymbols(QDateEdit.ButtonSymbols.UpDownArrows)
        
        self.contract_type = QComboBox()
        self.contract_type.addItems(['', 'Поставка', 'Услуги', 'ГПХ', 'Ремонт ТС'])
        
        self.contract_kind = QComboBox()
        self.contract_kind.addItems([
            '', 'Разовый', 'Рамочный', 'Рамочный накопительный', 
            'Хотят купить', 'Плановый', 'Подтвержденный'
        ])
        
        self.subsidy_combo = QComboBox()
        self.subsidy_combo.setEditable(True)
        self.load_subsidies()
        
        self.purchase_number = QLineEdit()
        self.purchase_number.editingFinished.connect(self.check_purchase_number_duplicate)
        self.order_number = QLineEdit()
        self.subject = QTextEdit()
        self.subject.setMaximumHeight(60)
        self.description = QTextEdit()
        self.description.setMaximumHeight(80)
        
        self.status = QComboBox()
        self.status.addItems(['Плановый', 'Подтвержденный', 'Ведутся работы', 'Исполнен', 'Расторгнут', 'Просрочен'])
        
        self.execution_stage = QComboBox()
        self.execution_stage.setEditable(True)
        self.execution_stage.addItems([
            '', 'Подготовка документов', 'Согласование', 'Подписание',
            'Исполнение', 'Приемка', 'Оплата', 'Завершен', 'Расторгнут'
        ])
        self.execution_stage.setToolTip('Стадия исполнения договора:\n'
                                       'Подготовка документов - формирование документов\n'
                                       'Согласование - процесс согласования\n'
                                       'Подписание - договор подписан\n'
                                       'Исполнение - выполнение обязательств\n'
                                       'Приемка - прием выполненных работ\n'
                                       'Оплата - процесс оплаты\n'
                                       'Завершен - договор полностью исполнен\n'
                                       'Расторгнут - договор расторгнут')
        
        self.start_date = QDateEdit()
        self.start_date.setDate(QDate.currentDate())
        self.start_date.setCalendarPopup(True)
        self.start_date.setDisplayFormat('dd.MM.yyyy')
        
        self.end_date = QDateEdit()
        self.end_date.setDate(QDate.currentDate().addDays(30))
        self.end_date.setCalendarPopup(True)
        self.end_date.setDisplayFormat('dd.MM.yyyy')
        
        layout.addRow('Номер договора:', self.contract_number)
        layout.addRow('Дата договора:', self.contract_date)
        layout.addRow('Тип договора:', self.contract_type)
        layout.addRow('Вид договора:', self.contract_kind)
        layout.addRow('Субсидия:', self.subsidy_combo)
        # Кнопка для просмотра рамочных договоров
        purchase_layout = QHBoxLayout()
        purchase_layout.addWidget(self.purchase_number)
        layout.addRow('Номер закупки:', purchase_layout)
        
        # Кнопки в отдельной строке в два ряда
        buttons_layout = QHBoxLayout()
        framework_btn = QPushButton('📋 Рамочные договоры')
        framework_btn.setToolTip('Просмотр всех рамочных договоров с остатками средств')
        framework_btn.clicked.connect(self.show_framework_contracts_dialog)
        buttons_layout.addWidget(framework_btn)
        layout.addRow('', buttons_layout)
        
        # Поле рамочных договоров
        self.framework_contract_combo = QComboBox()
        self.framework_contract_combo.addItem('')
        self.framework_contract_combo.setToolTip('Выберите рамочный договор для использования')
        self.load_framework_contracts()
        self.framework_contract_combo.currentIndexChanged.connect(self.on_framework_contract_selected)
        layout.addRow('Рамочный договор:', self.framework_contract_combo)
        
        layout.addRow('Номер заказа:', self.order_number)
        layout.addRow('Предмет договора:', self.subject)
        layout.addRow('Детальное описание:', self.description)
        layout.addRow('Статус договора:', self.status)
        layout.addRow('Стадия исполнения:', self.execution_stage)
        layout.addRow('Дата начала:', self.start_date)
        layout.addRow('Дата окончания:', self.end_date)
        
        group.setLayout(layout)
        return group
    
    def create_contractor_group(self):
        """Группа контрагента"""
        group = QGroupBox('Контрагент')
        layout = QFormLayout()
        
        self.contractor_combo = QComboBox()
        self.contractor_combo.setEditable(True)
        self.load_contractors()
        
        contractor_layout = QHBoxLayout()
        contractor_layout.addWidget(self.contractor_combo)
        layout.addRow('Контрагент:', contractor_layout)
        
        # Кнопка добавления контрагента в отдельной строке
        add_contractor_btn = QPushButton('➕ Добавить контрагента')
        add_contractor_btn.clicked.connect(self.show_add_contractor_dialog)
        layout.addRow('', add_contractor_btn)
        
        # Автозаполнение реквизитов (только для просмотра)
        self.contractor_details = QTextEdit()
        self.contractor_details.setMaximumHeight(100)
        self.contractor_details.setReadOnly(True)
        self.contractor_combo.currentIndexChanged.connect(self.update_contractor_details)
        
        layout.addRow('Реквизиты контрагента:', self.contractor_details)
        
        group.setLayout(layout)
        return group
    
    def create_contract_items_group(self):
        """Группа состава договора"""
        group = QGroupBox('Состав договора (товары/услуги)')
        layout = QVBoxLayout()
        
        # Таблица позиций
        self.items_table = QTableWidget()
        self.items_table.setColumnCount(10)
        self.items_table.setHorizontalHeaderLabels([
            'Наименование', 'Количество', 'Ед. изм.', 'Плановая цена за ед.', 
            'Фактическая цена за ед.', 'НДС %', 'Стоимость', 'Категория ФЭО', 'Удалить', ''
        ])
        self.items_table.horizontalHeader().setStretchLastSection(True)
        self.items_table.setAlternatingRowColors(True)
        self.items_table.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;
                color: #000000;
                gridline-color: #1E3A5F;
            }
            QTableWidget::item {
                color: #000000;
            }
            QTableWidget::item:selected {
                background-color: #5A7FA8;
                color: #FFFFFF;
            }
        """)
        # Начальная высота таблицы — в несколько раз больше базовой
        vh = self.items_table.verticalHeader()
        base_row_height = vh.defaultSectionSize()
        initial_height = (
            base_row_height * 5
            + self.items_table.horizontalHeader().height()
            + 2 * self.items_table.frameWidth()
        )
        self.items_table.setMinimumHeight(initial_height)
        self.items_table.setSizeAdjustPolicy(QAbstractScrollArea.SizeAdjustPolicy.AdjustToContents)
        # Реакция на изменение ячеек для автопересчёта стоимости
        self.items_table.itemChanged.connect(self.handle_items_table_item_changed)
        
        # Кнопки управления
        items_buttons = QHBoxLayout()
        add_item_btn = QPushButton('➕ Добавить позицию')
        add_item_btn.clicked.connect(self.add_contract_item_row)
        items_buttons.addWidget(add_item_btn)
        items_buttons.addStretch()
        
        layout.addWidget(self.items_table)
        layout.addLayout(items_buttons)
        
        group.setLayout(layout)
        return group
    
    def create_finance_group(self):
        """Группа финансов"""
        group = QGroupBox('Финансовые условия')
        layout = QFormLayout()
        
        self.nmck = QDoubleSpinBox()
        self.nmck.setMaximum(999999999)
        self.nmck.setPrefix('₽ ')
        
        self.price_without_vat = QDoubleSpinBox()
        self.price_without_vat.setMaximum(999999999)
        self.price_without_vat.setPrefix('₽ ')
        
        self.vat_applied = QComboBox()
        self.vat_applied.addItems(['Нет', 'Да'])
        self.vat_applied.currentTextChanged.connect(self.update_vat_fields)
        
        self.vat_rate = QSpinBox()
        self.vat_rate.setRange(0, 100)
        self.vat_rate.setValue(20)
        self.vat_rate.setSuffix('%')
        
        self.payment_method = QComboBox()
        self.payment_method.addItems(['', 'Безналичный', 'Наличный'])
        
        self.payment_form = QComboBox()
        self.payment_form.addItems(['', 'Предоплата', 'Постоплата', 'Поэтапная'])
        
        self.advance_amount = QDoubleSpinBox()
        self.advance_amount.setMaximum(100)
        self.advance_amount.setSuffix('%')
        
        self.payment_term = QSpinBox()
        self.payment_term.setSuffix(' дней')
        
        layout.addRow('НМЦК:', self.nmck)
        layout.addRow('Цена без НДС:', self.price_without_vat)
        layout.addRow('Применение НДС:', self.vat_applied)
        layout.addRow('Ставка НДС:', self.vat_rate)
        layout.addRow('Способ оплаты:', self.payment_method)
        layout.addRow('Форма оплаты:', self.payment_form)
        layout.addRow('Размер аванса:', self.advance_amount)
        layout.addRow('Срок оплаты:', self.payment_term)
        
        group.setLayout(layout)
        return group
    
    def create_feo_categories_group(self):
        """Группа выбора направлений и статей затрат ФЭО"""
        group = QGroupBox('Категоризация по ФЭО')
        # Делаем заголовок читаемым на светлом фоне
        group.setStyleSheet("""
            QGroupBox {
                font-weight: bold;
                margin-top: 4px;
            }
            QGroupBox::title {
                subcontrol-origin: margin;
                left: 10px;
                padding: 0 5px 0 5px;
                background-color: #E6F2FF;
            }
        """)
        layout = QVBoxLayout()
        
        # Направление расходов ФЭО - выпадающий список (из SQLite)
        feo_label = QLabel('Направление расходования ФЭО:')
        feo_label.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        layout.addWidget(feo_label)
        
        self.feo_direction_combo = QComboBox()
        self.feo_direction_combo.setEditable(True)
        layout.addWidget(self.feo_direction_combo)
        
        # Наименование статьи затрат - зависит от выбранного направления
        feo_type_label = QLabel('Наименование статьи затрат:')
        feo_type_label.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        layout.addWidget(feo_type_label)
        
        self.feo_type_combo = QComboBox()
        self.feo_type_combo.setEditable(True)
        self.feo_type_combo.currentTextChanged.connect(self.on_cost_item_changed)
        layout.addWidget(self.feo_type_combo)
        
        # Направление из приложения - скрыто до выбора статьи затрат
        self.app_label = QLabel('Направление расходов из приложений к ФЭО:')
        self.app_label.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        self.app_label.setVisible(False)
        layout.addWidget(self.app_label)
        
        self.app_checkboxes = {}
        self.app_scroll = QScrollArea()
        self.app_scroll.setVisible(False)
        self.app_widget = QWidget()
        self.app_layout = QGridLayout()  # Изменено на GridLayout для 3 столбцов
        self.app_layout.setColumnMinimumWidth(0, 200)
        self.app_layout.setColumnMinimumWidth(1, 200)
        self.app_layout.setColumnMinimumWidth(2, 200)
        
        self.app_widget.setLayout(self.app_layout)
        self.app_scroll.setWidget(self.app_widget)
        self.app_scroll.setWidgetResizable(True)
        self.app_scroll.setMaximumHeight(120)  # Уменьшена высота
        layout.addWidget(self.app_scroll)
        
        # Тип конкретизированный
        specific_label = QLabel('Тип расходов конкретизированный:')
        specific_label.setFont(QFont('Arial', 10, QFont.Weight.Bold))
        layout.addWidget(specific_label)
        
        # Тип конкретизированный - комбобокс с возможностью ввода своего
        self.specific_type = QComboBox()
        self.specific_type.setEditable(True)
        self.specific_type.setInsertPolicy(QComboBox.InsertPolicy.InsertAtTop)
        # Загружаем из базы данных
        specific_types = self.db.get_specific_types()
        if specific_types:
            self.specific_type.addItems([''] + specific_types)
        layout.addWidget(self.specific_type)
        
        group.setLayout(layout)
        
        # Загружаем направления ФЭО из Excel-справочников
        self.load_feo_from_excel()
        return group
    
    def create_additional_group(self):
        """Группа дополнительной информации"""
        group = QGroupBox('Дополнительно')
        layout = QFormLayout()
        
        self.responsible = QLineEdit()
        self.city = QLineEdit()
        self.comments = QTextEdit()
        self.comments.setMaximumHeight(80)
        
        layout.addRow('Ответственный:', self.responsible)
        layout.addRow('Город:', self.city)
        layout.addRow('Комментарии:', self.comments)
        
        group.setLayout(layout)
        return group
    
    def create_contractors_tab(self):
        """Вкладка контрагентов"""
        tab = QWidget()
        layout = QVBoxLayout()
        tab.setLayout(layout)
        
        # Кнопка добавления
        add_btn = QPushButton('➕ Добавить контрагента')
        add_btn.setStyleSheet('padding: 10px; font-size: 14px; background-color: #2196F3; color: #FFFFFF; font-weight: bold; border-radius: 5px;')
        add_btn.clicked.connect(self.show_add_contractor_dialog)
        layout.addWidget(add_btn)
        
        # Таблица контрагентов
        self.contractors_table = QTableWidget()
        self.apply_standard_table_style(self.contractors_table)
        self.load_contractors_table()
        layout.addWidget(self.contractors_table)
        
        self.tabs.addTab(tab, '👥 Контрагенты')
    
    def create_dashboard_tab(self):
        """Вкладка дашборда"""
        tab = QWidget()
        layout = QVBoxLayout()
        tab.setLayout(layout)
        
        # Выбор субсидии
        subsidy_label = QLabel('Выберите субсидию:')
        subsidy_label.setFont(QFont('Arial', 12, QFont.Weight.Bold))
        layout.addWidget(subsidy_label)
        
        self.dashboard_subsidy_combo = QComboBox()
        self.dashboard_subsidy_combo.addItem('Все субсидии')
        subsidies = self.db.get_subsidies()
        for subsidy in subsidies:
            if subsidy[1]:
                self.dashboard_subsidy_combo.addItem(f"{subsidy[1]} (ID: {subsidy[0]})", subsidy[0])
        self.dashboard_subsidy_combo.currentIndexChanged.connect(self.update_dashboard)
        layout.addWidget(self.dashboard_subsidy_combo)
        
        # Метрики
        metrics_group = QGroupBox('Метрики')
        metrics_layout = QGridLayout()
        
        self.total_subsidy_label = QLabel('Общий объём субсидии: 0 ₽')
        self.contracted_label = QLabel('Законтрактовано: 0 ₽')
        self.planned_label = QLabel('Планируется к контрактации: 0 ₽')
        self.delivered_label = QLabel('Поставлено: 0 ₽')
        self.paid_label = QLabel('Оплачено: 0 ₽')
        
        for label in [self.total_subsidy_label, self.contracted_label, 
                      self.planned_label, self.delivered_label, self.paid_label]:
            label.setFont(QFont('Arial', 11))
            label.setStyleSheet('padding: 10px; background-color: #E6F2FF; border-radius: 5px;')
        
        metrics_layout.addWidget(self.total_subsidy_label, 0, 0)
        metrics_layout.addWidget(self.contracted_label, 0, 1)
        metrics_layout.addWidget(self.planned_label, 1, 0)
        metrics_layout.addWidget(self.delivered_label, 1, 1)
        metrics_layout.addWidget(self.paid_label, 2, 0, 1, 2)
        
        metrics_group.setLayout(metrics_layout)
        layout.addWidget(metrics_group)
        
        # Таблица реестра для дашборда
        registry_label = QLabel('Реестр договоров:')
        registry_label.setFont(QFont('Arial', 12, QFont.Weight.Bold))
        layout.addWidget(registry_label)
        
        self.dashboard_registry_table = QTableWidget()
        self.dashboard_registry_table.setColumnCount(6)
        self.dashboard_registry_table.setHorizontalHeaderLabels([
            'Номер договора', 'Контрагент', 'Предмет', 'Сумма', 'Статус', 'Стадия'
        ])
        self.dashboard_registry_table.setAlternatingRowColors(True)
        layout.addWidget(self.dashboard_registry_table)
        
        # Обновляем дашборд при загрузке
        QTimer.singleShot(100, self.update_dashboard)
        
        self.tabs.addTab(tab, '📊 Дашборд')
    
    def update_dashboard(self):
        """Обновление дашборда"""
        if not hasattr(self, 'dashboard_subsidy_combo'):
            return
        
        subsidy_id = self.dashboard_subsidy_combo.currentData()
        
        # Получаем данные из базы
        if 'GoodsService' not in self.db.wb.sheetnames:
            return
        
        ws = self.db.wb['GoodsService']
        total_subsidy = 0
        contracted = 0
        planned = 0
        delivered = 0
        paid = 0
        
        contracts_data = []
        contractor_map = self.db._get_reference_map('Контрагенты', key_col=1, value_col=2)
        
        for idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            if len(row) < 6:
                continue
            
            row_subsidy_id = row[5] if len(row) > 5 else None
            if subsidy_id and row_subsidy_id != subsidy_id:
                continue
            
            # Получаем суммы
            try:
                nmck = float(row[17]) if len(row) > 17 and row[17] else 0
                contracted_val = float(row[24]) if len(row) > 24 and row[24] else 0
                delivered_val = float(row[25]) if len(row) > 25 and row[25] else 0
                paid_val = float(row[26]) if len(row) > 26 and row[26] else 0
                
                status = row[11] if len(row) > 11 else ''
                
                total_subsidy += nmck
                contracted += contracted_val
                delivered += delivered_val
                paid += paid_val
                
                if status == 'Плановый':
                    planned += nmck
                
                # Добавляем в таблицу
                contractor_id = row[10] if len(row) > 10 else None
                contractor_name = contractor_map.get(contractor_id, '') if contractor_id else ''
                
                contracts_data.append({
                    'number': row[1] if len(row) > 1 else '',
                    'contractor': contractor_name,
                    'subject': row[8] if len(row) > 8 else '',
                    'amount': nmck,
                    'status': status,
                    'stage': row[13] if len(row) > 13 else ''
                })
            except Exception as e:
                continue
        
        # Обновляем метрики
        if hasattr(self, 'total_subsidy_label'):
            self.total_subsidy_label.setText(f'Общий объём субсидии: {self.format_currency(total_subsidy)}')
            self.contracted_label.setText(f'Законтрактовано: {self.format_currency(contracted)}')
            self.planned_label.setText(f'Планируется к контрактации: {self.format_currency(planned)}')
            self.delivered_label.setText(f'Поставлено: {self.format_currency(delivered)}')
            self.paid_label.setText(f'Оплачено: {self.format_currency(paid)}')
        
        # Обновляем таблицу
        if hasattr(self, 'dashboard_registry_table'):
            self.dashboard_registry_table.setRowCount(len(contracts_data))
            for row_idx, contract in enumerate(contracts_data):
                self.dashboard_registry_table.setItem(row_idx, 0, QTableWidgetItem(str(contract['number'])))
                self.dashboard_registry_table.setItem(row_idx, 1, QTableWidgetItem(str(contract['contractor'])))
                self.dashboard_registry_table.setItem(row_idx, 2, QTableWidgetItem(str(contract['subject'])))
                self.dashboard_registry_table.setItem(row_idx, 3, QTableWidgetItem(self.format_currency(contract['amount'])))
                self.dashboard_registry_table.setItem(row_idx, 4, QTableWidgetItem(str(contract['status'])))
                self.dashboard_registry_table.setItem(row_idx, 5, QTableWidgetItem(str(contract['stage'])))
            
            self.dashboard_registry_table.resizeColumnsToContents()
    
    def create_registry_tab(self):
        """Вкладка реестра договоров"""
        tab = QWidget()
        layout = QVBoxLayout()
        tab.setLayout(layout)

        # Вкладки по субсидиям (фильтр реестра)
        self.registry_subsidy_tabs = QTabWidget()
        self.registry_subsidy_tab_names = []
        # Первая вкладка - "Все субсидии"
        all_tab = QWidget()
        self.registry_subsidy_tabs.addTab(all_tab, 'Все субсидии')
        self.registry_subsidy_tab_names.append(None)
        # Остальные вкладки - по субсидиям из справочника
        try:
            for sub in sorted(self.db.get_subsidies(), key=lambda s: str(s[1]).lower() if s[1] else ''):
                name = str(sub[1]) if len(sub) > 1 and sub[1] else str(sub[0])
                w = QWidget()
                self.registry_subsidy_tabs.addTab(w, name)
                self.registry_subsidy_tab_names.append(name)
        except Exception:
            pass
        self.registry_subsidy_tabs.currentChanged.connect(self.on_registry_subsidy_tab_changed)
        layout.addWidget(self.registry_subsidy_tabs)

        self.registry_table = QTableWidget()
        headers = [
            'Номер договора', 'Контрагент', 'Предмет договора', 'Субсидия',
            'Сумма (с НДС)', 'Законтрактовано', 'Поставлено', 'Оплачено',
            'Статус', 'Стадия'
        ]
        self.registry_table.setColumnCount(len(headers))
        self.registry_table.setHorizontalHeaderLabels(headers)
        self.registry_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.apply_standard_table_style(self.registry_table)
        
        layout.addWidget(self.registry_table)
        self.load_registry_table()
        
        self.tabs.addTab(tab, '📋 Реестр')

    def create_feo_tab(self):
        """Вкладка работы с ФЭО"""
        tab = QWidget()
        tab_layout = QVBoxLayout()
        tab.setLayout(tab_layout)

        # Блок управления базовым ФЭО и версиями
        version_group = QGroupBox('Базовое ФЭО')
        version_layout = QHBoxLayout()
        version_group.setLayout(version_layout)

        self.feo_version_label = QLabel()
        self.feo_version_label.setStyleSheet('font-weight: bold;')
        self.update_feo_version_label()

        import_btn = QPushButton('📂 Загрузить ФЭО (Excel БД)')
        import_btn.setToolTip('Загрузить базовое ФЭО и приложения в Excel-базу данных.\n'
                              'Текущая версия в Excel будет сохранена как архивная.')
        import_btn.clicked.connect(self.import_feo_from_excel)

        import_sqlite_btn = QPushButton('📥 Импорт ФЭО из сметы (SQLite)')
        import_sqlite_btn.setToolTip(
            'Импортировать направления и статьи затрат ФЭО в локальную БД SQLite\n'
            'из файла сметы (листы WORK и Unique).'
        )
        import_sqlite_btn.clicked.connect(self.import_feo_from_smeta_sqlite)

        version_layout.addWidget(self.feo_version_label, 1)
        version_layout.addWidget(import_btn)
        version_layout.addWidget(import_sqlite_btn)

        tab_layout.addWidget(version_group)

        # Вкладки по субсидиям (фильтр)
        self.feo_subsidy_tabs = QTabWidget()
        self.feo_subsidy_tab_ids = []
        # Первая вкладка - "Все субсидии"
        all_tab = QWidget()
        self.feo_subsidy_tabs.addTab(all_tab, 'Все субсидии')
        self.feo_subsidy_tab_ids.append(None)
        # Остальные вкладки - по субсидиям из справочника
        try:
            for sub in sorted(self.db.get_subsidies(), key=lambda s: str(s[1]).lower() if s[1] else ''):
                sub_id = sub[0]
                name = str(sub[1]) if len(sub) > 1 and sub[1] else str(sub_id)
                w = QWidget()
                self.feo_subsidy_tabs.addTab(w, name)
                self.feo_subsidy_tab_ids.append(sub_id)
        except Exception:
            pass
        self.feo_subsidy_tabs.currentChanged.connect(self.on_feo_subsidy_tab_changed)
        tab_layout.addWidget(self.feo_subsidy_tabs)

        # Базовые направления ФЭО
        base_group = QGroupBox('Базовые направления ФЭО')
        base_layout = QVBoxLayout()
        base_group.setLayout(base_layout)

        base_actions = QHBoxLayout()
        add_base_btn = QPushButton('➕ Добавить строку')
        add_base_btn.clicked.connect(self.add_feo_base_row)
        remove_base_btn = QPushButton('🗑️ Удалить строки')
        remove_base_btn.clicked.connect(self.remove_feo_base_rows)
        paste_base_btn = QPushButton('📋 Вставить из буфера')
        paste_base_btn.clicked.connect(self.paste_feo_base_rows)
        save_base_btn = QPushButton('💾 Сохранить')
        save_base_btn.clicked.connect(self.save_feo_base_table)
        base_actions.addWidget(add_base_btn)
        base_actions.addWidget(remove_base_btn)
        base_actions.addWidget(paste_base_btn)
        base_actions.addWidget(save_base_btn)
        base_actions.addStretch()
        base_layout.addLayout(base_actions)

        self.feo_base_table = QTableWidget()
        self.feo_base_table.setColumnCount(5)
        self.feo_base_table.setHorizontalHeaderLabels([
            'ID', 'Субсидия', 'Номер приложения', 'Направление расходов', 'Расч. объём (тыс. руб)'
        ])
        self.feo_base_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.apply_standard_table_style(self.feo_base_table)
        self.setup_paste_shortcut(self.feo_base_table, self.paste_feo_base_rows)
        self.feo_base_table.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self.feo_base_table.setSelectionMode(QAbstractItemView.SelectionMode.SingleSelection)
        self.feo_base_table.selectionModel().selectionChanged.connect(self.on_feo_base_selection_changed)
        base_layout.addWidget(self.feo_base_table)

        tab_layout.addWidget(base_group)

        # Приложения
        apps_group = QGroupBox('Приложения к ФЭО')
        apps_layout = QVBoxLayout()
        apps_group.setLayout(apps_layout)

        selector_layout = QHBoxLayout()
        selector_label = QLabel('ФЭО:')
        self.feo_app_base_combo = QComboBox()
        self.feo_app_base_combo.currentIndexChanged.connect(self.on_feo_app_base_changed)
        selector_layout.addWidget(selector_label)
        selector_layout.addWidget(self.feo_app_base_combo, 1)
        apps_layout.addLayout(selector_layout)

        apps_actions = QHBoxLayout()
        add_app_btn = QPushButton('➕ Добавить строку')
        add_app_btn.clicked.connect(self.add_feo_app_row)
        remove_app_btn = QPushButton('🗑️ Удалить строки')
        remove_app_btn.clicked.connect(self.remove_feo_app_rows)
        paste_app_btn = QPushButton('📋 Вставить из буфера')
        paste_app_btn.clicked.connect(self.paste_feo_app_rows)
        save_app_btn = QPushButton('💾 Сохранить')
        save_app_btn.clicked.connect(self.save_feo_app_table)
        apps_actions.addWidget(add_app_btn)
        apps_actions.addWidget(remove_app_btn)
        apps_actions.addWidget(paste_app_btn)
        apps_actions.addWidget(save_app_btn)
        apps_actions.addStretch()
        apps_layout.addLayout(apps_actions)

        self.feo_app_table = QTableWidget()
        self.feo_app_table.setColumnCount(8)
        self.feo_app_table.setHorizontalHeaderLabels([
            'ID', 'Название приложения', 'Направление расходов', 'Методика расчёта',
            'Единица измерения', 'Количество', 'План. стоимость за ед.', 'Плановый объём'
        ])
        self.feo_app_table.horizontalHeader().setSectionResizeMode(QHeaderView.ResizeMode.Stretch)
        self.apply_standard_table_style(self.feo_app_table)
        self.setup_paste_shortcut(self.feo_app_table, self.paste_feo_app_rows)
        self.feo_app_table.itemChanged.connect(self.handle_feo_app_item_changed)
        apps_layout.addWidget(self.feo_app_table)

        tab_layout.addWidget(apps_group)

        self.tabs.addTab(tab, '📑 ФЭО')
        self.load_feo_base_table()
        self.refresh_feo_app_base_combo()

    def load_registry_table(self):
        """Загрузка данных в реестр договоров"""
        if not self.registry_table:
            return
        contracts = self.db.get_contracts_for_registry()
        self.registry_table.setRowCount(len(contracts))
        
        for row_idx, contract in enumerate(contracts):
            values = [
                contract['number'],
                contract['contractor'],
                contract['subject'],
                contract['subsidy'],
                self.format_currency(contract['amount']),
                self.format_currency(contract['contracted']),
                self.format_currency(contract['delivered']),
                self.format_currency(contract['paid']),
                contract['status'],
                contract['stage']
            ]
            for col_idx, value in enumerate(values):
                item = QTableWidgetItem(str(value) if value is not None else '')
                if isinstance(value, str):
                    item.setToolTip(value)
                self.registry_table.setItem(row_idx, col_idx, item)

        # Применяем фильтр по выбранной вкладке субсидии
        self.apply_registry_subsidy_filter()

    def on_registry_subsidy_tab_changed(self, index: int):
        """Смена вкладки субсидии во вкладке Реестр"""
        if not self.registry_subsidy_tabs or index < 0 or index >= len(self.registry_subsidy_tab_names):
            self.current_registry_subsidy_name = None
        else:
            self.current_registry_subsidy_name = self.registry_subsidy_tab_names[index]
        self.apply_registry_subsidy_filter()

    def apply_registry_subsidy_filter(self):
        """Скрывает строки реестра по выбранной субсидии"""
        if not self.registry_table:
            return
        # Колонка 3 (индекс 3) – название субсидии
        for row in range(self.registry_table.rowCount()):
            item = self.registry_table.item(row, 3)
            row_subsidy = item.text().strip() if item else ''
            if self.current_registry_subsidy_name and row_subsidy != str(self.current_registry_subsidy_name):
                self.registry_table.setRowHidden(row, True)
            else:
                self.registry_table.setRowHidden(row, False)

    @staticmethod
    def format_currency(value):
        """Форматирование числовых значений"""
        try:
            value = float(value)
            if value == 0:
                return ''
            return f"{value:,.2f}".replace(',', ' ').replace('.', ',')
        except (TypeError, ValueError):
            return value or ''
    
    def load_feo_base_table(self):
        """Загрузка базовых ФЭО"""
        if not self.feo_base_table:
            return
        entries = self.db.get_feo_base_entries()
        self.feo_base_entries = entries
        self.feo_base_table.setRowCount(0)
        for entry in entries:
            row = self.feo_base_table.rowCount()
            self.feo_base_table.insertRow(row)
            self.set_subsidy_cell(row, entry.get('subsidy_id'), entry.get('subsidy_name') or '')
            id_item = self.create_table_item(entry.get('id'), read_only=True)
            id_item.setData(self.CREATED_ROLE, entry.get('created_at'))
            id_item.setData(self.UPDATED_ROLE, entry.get('updated_at'))
            self.feo_base_table.setItem(row, 0, id_item)
            self.feo_base_table.setItem(row, 2, self.create_table_item(entry.get('application_number') or ''))
            self.feo_base_table.setItem(row, 3, self.create_table_item(entry.get('direction') or ''))
            self.feo_base_table.setItem(row, 4, self.create_table_item(self.format_number(entry.get('amount')), read_only=True))
        # Применяем фильтр по выбранной вкладке субсидии
        self.apply_feo_subsidy_filter()
        self.refresh_feo_app_base_combo()
        if self.feo_base_table.rowCount() and self.feo_base_table.selectionModel() and not self.feo_base_table.selectionModel().selectedRows():
            # Выбираем первую видимую строку
            for row in range(self.feo_base_table.rowCount()):
                if not self.feo_base_table.isRowHidden(row):
                    self.feo_base_table.selectRow(row)
                    break
        self.update_feo_version_label()
    
    def add_feo_base_row(self):
        """Добавление строки в базовую таблицу ФЭО"""
        if not self.feo_base_table:
            return
        row = self.feo_base_table.rowCount()
        self.feo_base_table.insertRow(row)
        self.initialize_table_row(self.feo_base_table, row)
    
    def remove_feo_base_rows(self):
        """Удаление выбранных строк базового ФЭО"""
        if not self.feo_base_table:
            return
        for row in self.get_selected_rows(self.feo_base_table):
            self.feo_base_table.removeRow(row)
    
    def paste_feo_base_rows(self):
        """Вставка строк из буфера в базовое ФЭО"""
        self.paste_from_clipboard(self.feo_base_table, start_column=1, max_columns=4)
    
    def save_feo_base_table(self):
        """Сохранение базовых направлений ФЭО"""
        if not self.feo_base_table:
            return
        # Перед сохранением фиксируем текущую версию базового ФЭО
        self.db.archive_feo_version(action='save')
        entries = []
        for row in range(self.feo_base_table.rowCount()):
            subsidy_id, subsidy_name = self.get_subsidy_from_row(row)
            direction = self.get_table_text(self.feo_base_table, row, 3)
            application_number = self.get_table_text(self.feo_base_table, row, 2)
            if not any([subsidy_id, subsidy_name, direction, application_number]):
                continue
            entry_id = self.get_table_int(self.feo_base_table, row, 0)
            current_amount = self.db.calculate_feo_amount(entry_id) if entry_id else 0
            id_item = self.feo_base_table.item(row, 0)
            entry = {
                'id': entry_id,
                'subsidy_id': subsidy_id,
                'subsidy_name': subsidy_name,
                'application_number': application_number,
                'direction': direction,
                'amount': current_amount,
                'created_at': id_item.data(self.CREATED_ROLE) if id_item else None
            }
            entries.append(entry)
        self.db.save_feo_base_entries(entries)
        QMessageBox.information(self, 'Сохранено', 'Базовые ФЭО обновлены.')
        self.load_feo_base_table()
    
    def refresh_feo_app_base_combo(self):
        """Обновление списка ФЭО для приложений"""
        if not self.feo_app_base_combo:
            return
        current_id = self.feo_app_base_combo.currentData()
        self.feo_app_base_combo.blockSignals(True)
        self.feo_app_base_combo.clear()
        self.feo_app_base_combo.addItem('— Не выбрано —', None)
        for entry in self.feo_base_entries:
            # Если выбран фильтр по субсидии - показываем только соответствующие записи
            if self.current_feo_subsidy_id and entry.get('subsidy_id') != self.current_feo_subsidy_id:
                continue
            label_parts = []
            if entry.get('application_number'):
                label_parts.append(str(entry['application_number']))
            if entry.get('direction'):
                label_parts.append(entry['direction'][:60])
            label = ' | '.join(label_parts) if label_parts else f"ID {entry.get('id')}"
            self.feo_app_base_combo.addItem(label, entry.get('id'))
        self.feo_app_base_combo.blockSignals(False)
        if current_id:
            idx = self.feo_app_base_combo.findData(current_id)
            if idx != -1:
                self.feo_app_base_combo.setCurrentIndex(idx)
            elif self.feo_app_base_combo.count() > 1:
                self.feo_app_base_combo.setCurrentIndex(1)
            else:
                self.feo_app_base_combo.setCurrentIndex(0)
        else:
            self.feo_app_base_combo.setCurrentIndex(1 if self.feo_app_base_combo.count() > 1 else 0)
        self.load_feo_app_table()
    
    def on_feo_app_base_changed(self):
        """Событие выбора ФЭО для приложений"""
        self.load_feo_app_table()

    def on_feo_subsidy_tab_changed(self, index: int):
        """Смена вкладки субсидии во вкладке ФЭО"""
        if not self.feo_subsidy_tabs or index < 0 or index >= len(self.feo_subsidy_tab_ids):
            self.current_feo_subsidy_id = None
        else:
            self.current_feo_subsidy_id = self.feo_subsidy_tab_ids[index]
        # Обновляем все ячейки субсидии в таблице
        if self.feo_base_table:
            for row in range(self.feo_base_table.rowCount()):
                # Получаем текущие данные из строки
                sub_id, sub_name = self.get_subsidy_from_row(row)
                # Если выбрана конкретная вкладка, используем её субсидию
                if self.current_feo_subsidy_id is not None:
                    sub_id = self.current_feo_subsidy_id
                    sub_name = self.get_subsidy_name_by_id(sub_id)
                # Обновляем ячейку
                self.set_subsidy_cell(row, sub_id, sub_name)
        self.apply_feo_subsidy_filter()
        self.refresh_feo_app_base_combo()

    def apply_feo_subsidy_filter(self):
        """Скрывает строки базового ФЭО по выбранной субсидии"""
        if not self.feo_base_table:
            return
        for row in range(self.feo_base_table.rowCount()):
            sub_id, _ = self.get_subsidy_from_row(row)
            if self.current_feo_subsidy_id and sub_id and sub_id != self.current_feo_subsidy_id:
                self.feo_base_table.setRowHidden(row, True)
            else:
                self.feo_base_table.setRowHidden(row, False)

    def on_feo_base_selection_changed(self, selected, deselected):
        """Автоматический выбор ФЭО при выборе строки"""
        if not self.feo_base_table or not self.feo_app_base_combo:
            return
        selection_model = self.feo_base_table.selectionModel()
        if not selection_model:
            return
        indexes = selection_model.selectedRows()
        if not indexes:
            return
        row = indexes[0].row()
        feo_id = self.get_table_int(self.feo_base_table, row, 0)
        if not feo_id:
            return
        idx = self.feo_app_base_combo.findData(feo_id)
        if idx != -1:
            self.feo_app_base_combo.setCurrentIndex(idx)
    
    def get_current_base_direction(self):
        """Направление расходов выбранного ФЭО"""
        base_id = self.feo_app_base_combo.currentData() if self.feo_app_base_combo else None
        entry = self.get_base_entry(base_id)
        return entry.get('direction') if entry else ''
    
    def get_base_entry(self, base_id):
        """Поиск записи ФЭО по ID"""
        if not base_id:
            return None
        for entry in self.feo_base_entries:
            if entry.get('id') == base_id:
                return entry
        return None
    
    def load_feo_app_table(self):
        """Загрузка таблицы приложений"""
        if not self.feo_app_table:
            return
        base_id = self.feo_app_base_combo.currentData() if self.feo_app_base_combo else None
        if not base_id:
            self.feo_app_table.setRowCount(0)
            return
        entries = self.db.get_feo_app_entries(base_id)
        self.feo_app_table.setRowCount(0)
        for entry in entries:
            row = self.feo_app_table.rowCount()
            self.feo_app_table.insertRow(row)
            id_item = self.create_table_item(entry.get('id'), read_only=True)
            id_item.setData(self.CREATED_ROLE, entry.get('created_at'))
            id_item.setData(self.UPDATED_ROLE, entry.get('updated_at'))
            self.feo_app_table.setItem(row, 0, id_item)
            self.feo_app_table.setItem(row, 1, self.create_table_item(entry.get('app_name') or ''))
            self.feo_app_table.setItem(row, 2, self.create_table_item(entry.get('direction') or ''))
            self.feo_app_table.setItem(row, 3, self.create_table_item(entry.get('method') or ''))
            self.feo_app_table.setItem(row, 4, self.create_table_item(entry.get('unit') or ''))
            self.feo_app_table.setItem(row, 5, self.create_table_item(self.format_number(entry.get('quantity'))))
            self.feo_app_table.setItem(row, 6, self.create_table_item(self.format_number(entry.get('unit_price'))))
            self.set_app_row_total(row, entry.get('total_amount'))
    
    def add_feo_app_row(self):
        """Добавление строки в таблицу приложений"""
        if not self.feo_app_table:
            return
        if not self.feo_app_base_combo or not self.feo_app_base_combo.currentData():
            QMessageBox.warning(self, 'Внимание', 'Выберите базовое ФЭО перед добавлением приложения.')
            return
        row = self.feo_app_table.rowCount()
        self.feo_app_table.insertRow(row)
        self.initialize_table_row(self.feo_app_table, row)
    
    def remove_feo_app_rows(self):
        """Удаление строк приложений"""
        if not self.feo_app_table:
            return
        for row in self.get_selected_rows(self.feo_app_table):
            self.feo_app_table.removeRow(row)
    
    def paste_feo_app_rows(self):
        """Вставка строк приложений из буфера"""
        self.paste_from_clipboard(self.feo_app_table, start_column=1, max_columns=7)
    
    def save_feo_app_table(self):
        """Сохранение приложений выбранного ФЭО"""
        if not self.feo_app_table or not self.feo_app_base_combo:
            return
        base_id = self.feo_app_base_combo.currentData()
        if not base_id:
            QMessageBox.warning(self, 'Внимание', 'Выберите базовое ФЭО для сохранения приложений.')
            return
        entries = []
        for row in range(self.feo_app_table.rowCount()):
            app_name = self.get_table_text(self.feo_app_table, row, 1)
            direction = self.get_table_text(self.feo_app_table, row, 2)
            method = self.get_table_text(self.feo_app_table, row, 3)
            unit = self.get_table_text(self.feo_app_table, row, 4)
            quantity = self.parse_number(self.get_table_text(self.feo_app_table, row, 5))
            unit_price = self.parse_number(self.get_table_text(self.feo_app_table, row, 6))
            if not any([app_name, direction, method, unit, quantity, unit_price]):
                continue
            total_amount = quantity * unit_price
            id_item = self.feo_app_table.item(row, 0)
            entry = {
                'id': self.get_table_int(self.feo_app_table, row, 0),
                'app_name': app_name,
                'direction': direction,
                'method': method,
                'unit': unit,
                'quantity': quantity,
                'unit_price': unit_price,
                'total_amount': total_amount,
                'created_at': id_item.data(self.CREATED_ROLE) if id_item else None
            }
            entries.append(entry)
        self.db.save_feo_app_entries(base_id, entries)
        QMessageBox.information(self, 'Сохранено', 'Приложения обновлены.')
        self.load_feo_app_table()
        self.load_feo_base_table()
    
    def paste_from_clipboard(self, table, start_column, max_columns):
        """Вставка данных из буфера обмена в таблицу"""
        if not table:
            return
        clipboard = QApplication.clipboard()
        text = clipboard.text()
        if not text:
            return
        lines = [line for line in text.splitlines() if line.strip()]
        if not lines:
            return
        selection = table.selectedIndexes() if table.selectionModel() else []
        if selection:
            start_row = min(idx.row() for idx in selection)
            start_col = min(idx.column() for idx in selection)
        else:
            start_row = table.rowCount()
            start_col = start_column
        for line_idx, line in enumerate(lines):
            values = [val.strip() for val in line.split('\t')]
            row = start_row + line_idx
            while row >= table.rowCount():
                table.insertRow(table.rowCount())
                self.initialize_table_row(table, table.rowCount() - 1)
            if table is self.feo_base_table and not table.cellWidget(row, 1):
                self.set_subsidy_cell(row, None, '')
            for idx in range(min(len(values), max_columns)):
                col = start_col + idx
                if col >= table.columnCount():
                    break
                value = values[idx]
                if col == 0:
                    continue
                if table is self.feo_base_table and col == 1:
                    combo = table.cellWidget(row, 1)
                    self.set_combo_text(combo, value)
                    continue
                if table is self.feo_base_table and col == 4:
                    continue
                if table is self.feo_app_table and col == 7:
                    continue
                widget = table.cellWidget(row, col)
                if isinstance(widget, QComboBox):
                    self.set_combo_text(widget, value)
                else:
                    item = table.item(row, col)
                    if not item:
                        read_only = (table is self.feo_base_table and col == 4) or (table is self.feo_app_table and col == 7)
                        item = self.create_table_item('', read_only=read_only)
                        table.setItem(row, col, item)
                    item.setText(value)
            if table is self.feo_app_table:
                self.update_feo_app_row_total(row)

    def import_feo_from_excel(self):
        """Загрузка базового ФЭО и приложений из внешнего Excel-файла"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            'Выберите файл с базовым ФЭО',
            '',
            'Excel файлы (*.xlsx *.xlsm *.xltx *.xltm);;Все файлы (*.*)'
        )
        if not file_path:
            return

        # Архивируем текущую версию базового ФЭО
        self.db.archive_feo_version(source_file=file_path, action='import')

        try:
            importer = DataImporter(source_file=file_path, target_file=self.db.db_file)
            if not importer.load_files():
                raise RuntimeError('Не удалось открыть файлы для импорта.')
            base_count = importer.import_feo_base()
            apps_count = importer.import_feo_applications()
            importer.target_wb.save(importer.target_file)

            # Перезагружаем БД и обновляем вкладку ФЭО
            self.db.load_database()
            self.load_feo_base_table()
            self.refresh_feo_app_base_combo()
            self.update_feo_version_label()

            QMessageBox.information(
                self,
                'Импорт ФЭО',
                f'Импорт базового ФЭО завершен.\n'
                f'Базовые записи: {base_count}\n'
                f'Приложения: {apps_count}'
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self,
                'Ошибка импорта ФЭО',
                f'Не удалось импортировать данные ФЭО:\n{exc}'
            )

    def import_feo_from_smeta_sqlite(self):
        """Импорт направлений ФЭО и листа Unique в SQLite из файла сметы"""
        file_path, _ = QFileDialog.getOpenFileName(
            self,
            'Выберите файл сметы (WORK / Unique)',
            '',
            'Excel файлы (*.xlsx *.xlsm *.xltx *.xltm);;Все файлы (*.*)'
        )
        if not file_path:
            return

        try:
            work_count, unique_count = import_feo_to_sqlite(file_path, work_group_code="ZO")
            # После импорта в SQLite теперь обновляем комбобоксы из Excel‑справочников
            self.load_feo_from_excel()
            QMessageBox.information(
                self,
                'Импорт ФЭО',
                'Импорт из сметы завершён.\n'
                f'Строк WORK (направления/статьи ЗО): {work_count}\n'
                f'Записей Unique: {unique_count}'
            )
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self,
                'Ошибка импорта ФЭО',
                f'Не удалось импортировать данные ФЭО:\n{exc}'
            )
    
    def get_selected_rows(self, table):
        """Получение выбранных строк в таблице"""
        if not table or not table.selectionModel():
            return []
        rows = sorted({index.row() for index in table.selectionModel().selectedIndexes()}, reverse=True)
        return rows
    
    def create_table_item(self, value, read_only=False):
        """Утилита для создания ячейки таблицы"""
        text = '' if value is None else str(value)
        item = QTableWidgetItem(text)
        if read_only:
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        return item
    
    def get_table_text(self, table, row, column):
        """Получение текста из ячейки"""
        item = table.item(row, column)
        return item.text().strip() if item and item.text() else ''
    
    def get_table_int(self, table, row, column):
        """Получение целого значения из ячейки"""
        text = self.get_table_text(table, row, column)
        try:
            return int(text)
        except ValueError:
            return None
    
    def parse_number(self, value):
        """Преобразование строки в число"""
        if value is None:
            return 0
        if isinstance(value, (int, float)):
            return float(value)
        text = str(value).strip()
        if not text:
            return 0
        text = text.replace(' ', '').replace(',', '.')
        try:
            return float(text)
        except ValueError:
            return 0
    
    def format_number(self, value):
        """Форматирование числа для отображения"""
        try:
            value = float(value)
            if value.is_integer():
                return f"{int(value)}"
            return f"{value:.2f}"
        except (TypeError, ValueError):
            return value or ''

    def handle_feo_app_item_changed(self, item):
        """Реакция на изменение количества/цены"""
        if not item or self._app_table_updating:
            return
        if item.tableWidget() is not self.feo_app_table:
            return
        if item.column() in (5, 6):
            self.update_feo_app_row_total(item.row())

    def update_feo_app_row_total(self, row):
        """Пересчет планового объема"""
        if not self.feo_app_table or row < 0:
            return
        quantity = self.parse_number(self.get_table_text(self.feo_app_table, row, 5))
        unit_price = self.parse_number(self.get_table_text(self.feo_app_table, row, 6))
        total = quantity * unit_price
        self.set_app_row_total(row, total)

    def set_app_row_total(self, row, value):
        """Установка значения в столбец планового объема"""
        if not self.feo_app_table or row < 0:
            return
        self._app_table_updating = True
        item = self.feo_app_table.item(row, 7)
        if not item:
            item = self.create_table_item('', read_only=True)
            self.feo_app_table.setItem(row, 7, item)
        else:
            item.setFlags(item.flags() & ~Qt.ItemFlag.ItemIsEditable)
        item.setText(self.format_number(value))
        self._app_table_updating = False
    
    def resolve_subsidy(self, name):
        """Поиск ID субсидии по названию"""
        if not name:
            return None, ''
        normalized = name.strip()
        if not normalized:
            return None, ''
        for stored_name, stored_id in self.subsidy_lookup.items():
            if stored_name.lower() == normalized.lower():
                return stored_id, stored_name
        return None, normalized
    
    def get_subsidy_name_by_id(self, subsidy_id):
        """Получение названия субсидии по ID"""
        if not subsidy_id:
            return ''
        for sid, name in self.subsidy_options:
            if sid == subsidy_id:
                return name
        return ''
    
    def apply_standard_table_style(self, table):
        """Единый стиль таблиц"""
        if not table:
            return
        table.setAlternatingRowColors(True)
        table.setStyleSheet("""
            QTableWidget {
                background-color: #FFFFFF;
                color: #000000;
                gridline-color: #1E3A5F;
            }
            QTableWidget::item {
                color: #000000;
                background-color: #FFFFFF;
            }
            QTableWidget::item:alternate {
                background-color: #F5F5F5;
            }
            QTableWidget::item:selected {
                background-color: #5A7FA8;
                color: #FFFFFF;
            }
            QHeaderView::section {
                background-color: #1E3A5F;
                color: #FFFFFF;
                padding: 5px;
                border: none;
            }
        """)
    
    def setup_paste_shortcut(self, widget, handler):
        """Горячие клавиши для вставки"""
        if not widget or not handler:
            return
        shortcut_v = QShortcut(QKeySequence('Ctrl+V'), widget)
        shortcut_v.setContext(Qt.ShortcutContext.WidgetShortcut)
        shortcut_v.activated.connect(handler)
        shortcut_p = QShortcut(QKeySequence('Ctrl+P'), widget)
        shortcut_p.setContext(Qt.ShortcutContext.WidgetShortcut)
        shortcut_p.activated.connect(handler)
        self._shortcuts.extend([shortcut_v, shortcut_p])
    
    def initialize_table_row(self, table, row):
        """Инициализация специальных ячеек"""
        if table is self.feo_base_table:
            self.feo_base_table.setItem(row, 0, self.create_table_item('', read_only=True))
            # Если выбрана вкладка субсидии, автоматически заполняем субсидию
            if self.current_feo_subsidy_id is not None:
                subsidy_name = self.get_subsidy_name_by_id(self.current_feo_subsidy_id)
                self.set_subsidy_cell(row, self.current_feo_subsidy_id, subsidy_name)
            else:
                self.set_subsidy_cell(row, None, '')
            for col in range(2, self.feo_base_table.columnCount()):
                read_only = (col == 4)
                self.feo_base_table.setItem(row, col, self.create_table_item('', read_only=read_only))
        elif table is self.feo_app_table:
            self.feo_app_table.setItem(row, 0, self.create_table_item('', read_only=True))
            for col in range(1, self.feo_app_table.columnCount()):
                read_only = (col == 7)
                self.feo_app_table.setItem(row, col, self.create_table_item('', read_only=read_only))
            default_direction = self.get_current_base_direction()
            if default_direction and self.feo_app_table.item(row, 2):
                self.feo_app_table.item(row, 2).setText(default_direction)
            self.update_feo_app_row_total(row)
        else:
            table.setItem(row, 0, self.create_table_item('', read_only=True))
    
    def set_subsidy_cell(self, row, subsidy_id=None, subsidy_name=''):
        """Установка ячейки субсидии в строке"""
        if not self.feo_base_table:
            return
        
        # Если выбрана вкладка субсидии, используем текстовую ячейку вместо комбобокса
        if self.current_feo_subsidy_id is not None:
            # Используем субсидию из текущей вкладки
            current_subsidy_id = self.current_feo_subsidy_id
            current_subsidy_name = self.get_subsidy_name_by_id(current_subsidy_id)
            if not current_subsidy_name:
                current_subsidy_name = subsidy_name if subsidy_name else ''
            # Удаляем комбобокс, если он был установлен ранее
            widget = self.feo_base_table.cellWidget(row, 1)
            if widget:
                self.feo_base_table.removeCellWidget(row, 1)
            # Создаем текстовую ячейку (только для чтения)
            item = self.create_table_item(current_subsidy_name, read_only=True)
            self.feo_base_table.setItem(row, 1, item)
        else:
            # Если вкладка "Все субсидии", используем комбобокс
            combo = self.create_subsidy_combobox(subsidy_id, subsidy_name)
            self.feo_base_table.setCellWidget(row, 1, combo)
    
    def create_subsidy_combobox(self, selected_id=None, display_text=''):
        """Комбобокс выбора субсидии"""
        combo = QComboBox()
        combo.setEditable(False)
        combo.addItem('— выберите субсидию —', None)
        for sid, name in self.subsidy_options:
            combo.addItem(name, sid)
        if selected_id:
            idx = combo.findData(selected_id)
            if idx != -1:
                combo.setCurrentIndex(idx)
            elif display_text:
                combo.setCurrentText(display_text)
        elif display_text:
            combo.setCurrentText(display_text)
        return combo

    def set_combo_text(self, combo, text):
        """Установка текста в комбобоксе с попыткой подобрать значение"""
        if not combo:
            return
        normalized = text or ''
        idx = combo.findText(normalized, Qt.MatchFlag.MatchFixedString)
        if idx == -1:
            combo.setCurrentText(normalized)
        else:
            combo.setCurrentIndex(idx)

    def get_subsidy_from_row(self, row):
        """Получение выбранной субсидии из строки"""
        # Проверяем, есть ли комбобокс (для вкладки "Все субсидии")
        combo = self.feo_base_table.cellWidget(row, 1) if self.feo_base_table else None
        if combo:
            return combo.currentData(), combo.currentText().strip()
        # Если выбрана конкретная вкладка субсидии, возвращаем её ID и название
        if self.current_feo_subsidy_id is not None:
            subsidy_name = self.get_subsidy_name_by_id(self.current_feo_subsidy_id)
            return self.current_feo_subsidy_id, subsidy_name
        # Иначе пытаемся прочитать из текстовой ячейки
        text = self.get_table_text(self.feo_base_table, row, 1)
        return self.resolve_subsidy(text)
    
    def load_subsidies(self):
        """Загрузка субсидий в комбобокс"""
        raw_subsidies = self.db.get_subsidies()
        # Диагностика: если субсидий нет, показываем понятное сообщение с именем файла и списком листов
        if not raw_subsidies:
            try:
                sheetnames = ', '.join(self.db.wb.sheetnames)
            except Exception:
                sheetnames = 'недоступно'
            db_path = getattr(self.db, "db_file", "(неизвестно)")
            msg = (
                "Не удалось загрузить ни одной субсидии из файла базы данных.\n\n"
                f"Файл БД: {db_path}\n"
                f"Доступные листы: {sheetnames}\n\n"
                "Проверьте, что в этом файле есть лист \"Субсидии\" и в нём с 2-й строки "
                "заполнены колонки: A (ID) и B (Наименование)."
            )
            QMessageBox.warning(self, "Субсидии не найдены", msg)
        subsidies = sorted(
            raw_subsidies,
            key=lambda s: str(s[1]).lower() if s[1] else '',
        )
        self.subsidy_combo.clear()
        self.subsidy_combo.addItem('— выберите субсидию —')
        self.subsidy_lookup = {}
        self.subsidy_options = []
        for subsidy in subsidies:
            if subsidy[1]:  # Наименование
                name = str(subsidy[1])
                # Показываем ровно то название, что в листе "Субсидии"
                display_name = name
                self.subsidy_combo.addItem(display_name, subsidy[0])
                self.subsidy_lookup[name] = subsidy[0]
                self.subsidy_options.append((subsidy[0], name))
        # При смене субсидии обновляем доступные направления/статьи ФЭО
        self.subsidy_combo.currentIndexChanged.connect(self.on_subsidy_changed_for_feo)
        # Настройка автодополнения "как в Google Sheets"
        completer = QCompleter(self.subsidy_combo.model(), self.subsidy_combo)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.subsidy_combo.setCompleter(completer)
        if self.feo_base_table and self.feo_base_table.rowCount():
            self.load_feo_base_table()

    def update_feo_version_label(self):
        """Обновление текста о версии базового ФЭО"""
        if not hasattr(self, 'feo_version_label') or not self.feo_version_label:
            return
        entries = self.db.get_feo_base_entries()
        if not entries:
            self.feo_version_label.setText('Базовое ФЭО ещё не загружено.')
            return
        total = len(entries)
        dates = []
        for e in entries:
            for key in ('updated_at', 'created_at'):
                dt = e.get(key)
                if isinstance(dt, datetime):
                    dates.append(dt)
        info = f'Записей в базовом ФЭО: {total}'
        if dates:
            last_dt = max(dates)
            info += f' | Последнее обновление: {last_dt.strftime("%d.%m.%Y %H:%M")}'
        self.feo_version_label.setText(info)
    
    def load_contractors(self):
        """Загрузка контрагентов в комбобокс"""
        contractors = sorted(
            self.db.get_contractors(),
            key=lambda c: str(c[1]).lower() if c[1] else '',
        )
        self.contractor_combo.clear()
        contractor_names = ['']
        self.contractor_combo.addItem('')
        for contractor in contractors:
            if contractor[1]:  # Контрагент
                name = f"{contractor[1]} (ID: {contractor[0]})"
                self.contractor_combo.addItem(name, contractor[0])
                contractor_names.append(contractor[1])  # Для автодополнения только имя
        
        # Обновляем модель автодополнения
        if hasattr(self, 'contractor_completer_model'):
            self.contractor_completer_model.setStringList(contractor_names)
        completer = QCompleter(self.contractor_combo.model(), self.contractor_combo)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.contractor_combo.setCompleter(completer)
    
    def load_contractors_table(self):
        """Загрузка таблицы контрагентов"""
        contractors = self.db.get_contractors()
        
        headers = ['ID', 'Контрагент', 'ИНН', 'КПП', 'Телефон', 'E-mail', 'Активен']
        self.contractors_table.setColumnCount(len(headers))
        self.contractors_table.setHorizontalHeaderLabels(headers)
        self.contractors_table.setRowCount(len(contractors))
        
        for row, contractor in enumerate(contractors):
            for col, header in enumerate(headers):
                if header == 'ID':
                    value = contractor[0]
                elif header == 'Контрагент':
                    value = contractor[1]
                elif header == 'ИНН':
                    value = contractor[3]
                elif header == 'КПП':
                    value = contractor[4]
                elif header == 'Телефон':
                    value = contractor[21]
                elif header == 'E-mail':
                    value = contractor[23]
                elif header == 'Активен':
                    value = contractor[32]
                else:
                    value = ''
                
                item = QTableWidgetItem(str(value) if value else '')
                self.contractors_table.setItem(row, col, item)
        
        self.contractors_table.resizeColumnsToContents()
    
    def update_contractor_details(self):
        """Обновление реквизитов контрагента"""
        contractor_id = self.contractor_combo.currentData()
        if not contractor_id:
            self.contractor_details.clear()
            return
        
        contractors = self.db.get_contractors()
        for contractor in contractors:
            if contractor[0] == contractor_id:
                details = f"ИНН: {contractor[3] or ''}\n"
                details += f"КПП: {contractor[4] or ''}\n"
                details += f"Расчетный счет: {contractor[14] or ''}\n"
                details += f"БИК: {contractor[16] or ''}\n"
                details += f"Адрес: {contractor[18] or ''}"
                self.contractor_details.setText(details)
                break
    
    def on_subsidy_changed_for_feo(self):
        """Реакция на смену субсидии: обновляем доступные направления/статьи ФЭО"""
        if hasattr(self, 'feo_direction_combo'):
            self.load_feo_from_excel()
    
    def load_feo_from_excel(self):
        """Загрузка направлений ФЭО из листа Справочник_Направления_ФЭО (без привязки к субсидии).

        Логика максимально терпимая к структуре листа:
        - ищем колонку по заголовку, а не по жёсткому индексу;
        - не требуем, чтобы были заполнены все ячейки в строке.
        """
        if not hasattr(self, 'feo_direction_combo') or not self.feo_direction_combo:
            return

        self.feo_direction_combo.blockSignals(True)
        self.feo_direction_combo.clear()

        try:
            if 'Справочник_Направления_ФЭО' not in self.db.wb.sheetnames:
                # Диагностика: показываем, из какого файла и какие листы видим
                db_path = getattr(self.db, "db_file", "(неизвестно)")
                try:
                    sheetnames = ', '.join(self.db.wb.sheetnames)
                except Exception:
                    sheetnames = 'недоступно'
                msg = (
                    "Лист 'Справочник_Направления_ФЭО' не найден в файле базы данных.\n\n"
                    f"Файл БД: {db_path}\n"
                    f"Доступные листы: {sheetnames}\n\n"
                    "Убедитесь, что вы редактируете именно этот файл и что лист называется "
                    "точно 'Справочник_Направления_ФЭО'."
                )
                QMessageBox.warning(self, "Направления ФЭО не найдены", msg)
                self.feo_direction_combo.addItem('⚠️ Лист Справочник_Направления_ФЭО не найден', None)
                self.feo_direction_combo.setEnabled(False)
                self.feo_direction_combo.blockSignals(False)
                return

            ws = self.db.wb['Справочник_Направления_ФЭО']

            # Определяем колонку направления по заголовку первой строки
            header_row = ws[1]
            dir_col_idx = None
            headers_debug = []
            for idx, cell in enumerate(header_row):
                text = str(cell.value).strip() if cell.value is not None else ''
                if text:
                    headers_debug.append(text)
                low = text.lower()
                # Ищем что-то типа "Направление расходования ФЭО"
                if 'направ' in low and 'фэо' in low and dir_col_idx is None:
                    dir_col_idx = idx

            # Если заголовок не нашли, пробуем просто взять третью колонку (как по умолчанию)
            if dir_col_idx is None:
                dir_col_idx = 2  # индекс с нуля -> третья колонка

            directions = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                # Берём значение из найденной колонки, но не вываливаемся, если строка короче
                direction = ''
                if dir_col_idx < len(row) and row[dir_col_idx]:
                    direction = str(row[dir_col_idx]).strip()
                if direction:
                    directions.add(direction)

            if not directions:
                # Диагностика: лист есть, но направлений не нашли
                db_path = getattr(self.db, "db_file", "(неизвестно)")
                msg = (
                    "В листе 'Справочник_Направления_ФЭО' не найдено ни одного направления "
                    "в столбце 'Направление расходования ФЭО' (3-й столбец, колонка C).\n\n"
                    f"Файл БД: {db_path}\n"
                    f"Всего строк (с 2-й): {ws.max_row - 1}\n\n"
                    "Проверьте, что в третьем столбце действительно есть текстовые значения."
                )
                QMessageBox.warning(self, "Направления ФЭО пусты", msg)
                self.feo_direction_combo.addItem('⚠️ Направления ФЭО не найдены', None)
                self.feo_direction_combo.setEnabled(False)
                self.feo_direction_combo.blockSignals(False)
                return

            self.feo_direction_combo.setEnabled(True)
            self.feo_type_combo.setEnabled(True)
            # Заполняем направления. Сразу подставляем первое значение, чтобы пользователь видел результат.
            sorted_dirs = sorted(directions)
            for direction in sorted_dirs:
                self.feo_direction_combo.addItem(direction)
            if sorted_dirs:
                self.feo_direction_combo.setCurrentIndex(0)

            # Отладочное окно, чтобы понять, что именно мы прочитали
            # (Без всплывающего окна отладки, чтобы не раздражать)
        except Exception:
            self.feo_direction_combo.addItem('⚠️ Ошибка загрузки направлений ФЭО', None)
            self.feo_direction_combo.setEnabled(False)

        self.feo_direction_combo.blockSignals(False)
        # При смене направления обновляем статьи затрат
        try:
            self.feo_direction_combo.currentTextChanged.disconnect(self.update_feo_dependencies)
        except Exception:
            pass
        self.feo_direction_combo.currentTextChanged.connect(self.update_feo_dependencies)
        # Инициализируем статьи затрат для текущего (первого) направления
        self.update_feo_dependencies()
        # Автодополнение для направлений
        completer = QCompleter(self.feo_direction_combo.model(), self.feo_direction_combo)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.feo_direction_combo.setCompleter(completer)
    
    def update_feo_dependencies(self):
        """Обновление списка статей затрат для выбранного направления ФЭО из листа Справочник_Направления_ФЭО.

        Поиск колонок делаем по заголовкам, а не по жёстким индексам.
        """
        if not hasattr(self, 'feo_direction_combo') or not self.feo_direction_combo:
            return
        if not hasattr(self, 'feo_type_combo') or not self.feo_type_combo:
            return

        self.feo_type_combo.blockSignals(True)
        self.feo_type_combo.clear()

        direction_text = self.feo_direction_combo.currentText().strip()
        # Если направление не выбрано или это служебное сообщение
        if not direction_text or direction_text.startswith('⚠️'):
            self.feo_type_combo.addItem('')
            self.feo_type_combo.blockSignals(False)
            # Прячем блок направлений из приложений
            if hasattr(self, 'app_label'):
                self.app_label.setVisible(False)
            if hasattr(self, 'app_scroll'):
                self.app_scroll.setVisible(False)
            return

        # Ищем статьи затрат (столбец "Группа") для выбранного направления (столбец "Направление расходования ФЭО")
        try:
            if 'Справочник_Направления_ФЭО' not in self.db.wb.sheetnames:
                self.feo_type_combo.addItem('')
                self.feo_type_combo.blockSignals(False)
                return
            ws = self.db.wb['Справочник_Направления_ФЭО']

            # Определяем индексы колонок по заголовкам
            header_row = ws[1]
            dir_col_idx = None
            cost_col_idx = None
            for idx, cell in enumerate(header_row):
                text = str(cell.value).strip() if cell.value is not None else ''
                low = text.lower()
                if 'направ' in low and 'фэо' in low and dir_col_idx is None:
                    dir_col_idx = idx
                if ('группа' in low or ('наимен' in low and 'стат' in low and 'затрат' in low)) and cost_col_idx is None:
                    cost_col_idx = idx

            if dir_col_idx is None:
                dir_col_idx = 2  # по умолчанию третья колонка
            if cost_col_idx is None:
                cost_col_idx = 1  # по умолчанию вторая колонка

            cost_items = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row:
                    continue
                # Направление
                row_dir = ''
                if dir_col_idx < len(row) and row[dir_col_idx]:
                    row_dir = str(row[dir_col_idx]).strip()
                if row_dir != direction_text:
                    continue
                # Статья затрат
                cost_item = ''
                if cost_col_idx < len(row) and row[cost_col_idx]:
                    cost_item = str(row[cost_col_idx]).strip()
                if cost_item:
                    cost_items.add(cost_item)

            self.feo_type_combo.addItem('')
            for name in sorted(cost_items):
                self.feo_type_combo.addItem(name)
        except Exception:
            self.feo_type_combo.addItem('')

        # Автодополнение для статей затрат
        completer = QCompleter(self.feo_type_combo.model(), self.feo_type_combo)
        completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
        completer.setFilterMode(Qt.MatchFlag.MatchContains)
        completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
        self.feo_type_combo.setCompleter(completer)
        self.feo_type_combo.blockSignals(False)
    
    def load_cost_items_from_directions_sheet(self):
        """Загрузка статей затрат из столбца 'Группа' листа Справочник_Направления_ФЭО"""
        if not hasattr(self, 'feo_type_combo') or not self.feo_type_combo:
            return
        try:
            if 'Справочник_Направления_ФЭО' not in self.db.wb.sheetnames:
                return
            ws = self.db.wb['Справочник_Направления_ФЭО']
            cost_items = set()
            for row in ws.iter_rows(min_row=2, values_only=True):
                if not row or not any(row):
                    continue
                # Столбец "Группа" содержит наименование статьи затрат
                cost_item = str(row[1]).strip() if len(row) > 1 and row[1] else ''
                if cost_item:
                    cost_items.add(cost_item)
            self.feo_type_combo.clear()
            self.feo_type_combo.addItem('')
            for item in sorted(cost_items):
                self.feo_type_combo.addItem(item)
            # Автодополнение для статей затрат
            completer = QCompleter(self.feo_type_combo.model(), self.feo_type_combo)
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
            self.feo_type_combo.setCompleter(completer)
        except Exception as e:
            pass
    
    def on_cost_item_changed(self, cost_item_name):
        """Обработчик изменения статьи затрат - загрузка направлений расходов из приложения"""
        if not hasattr(self, 'app_checkboxes') or not hasattr(self, 'app_layout'):
            return
        
        # Очищаем старые чекбоксы
        for checkbox in list(self.app_checkboxes.values()):
            checkbox.deleteLater()
        self.app_checkboxes.clear()
        
        # Очищаем layout
        while self.app_layout.count():
            item = self.app_layout.takeAt(0)
            if item.widget():
                item.widget().deleteLater()
        
        cost_item_name = cost_item_name.strip()
        if not cost_item_name:
            # Скрываем секцию, если статья затрат не выбрана
            self.app_label.setVisible(False)
            self.app_scroll.setVisible(False)
            return
        
        # Получаем направления расходов из приложения для выбранной статьи затрат
        try:
            directions = self.db.get_app_directions_by_cost_item(cost_item_name)
            if directions:
                # Показываем секцию
                self.app_label.setVisible(True)
                self.app_scroll.setVisible(True)
                
                # Добавляем чекбоксы в 3 столбца
                row = 0
                col = 0
                for direction in directions:
                    cb = QCheckBox(str(direction))
                    self.app_checkboxes[direction] = cb
                    self.app_layout.addWidget(cb, row, col)
                    col += 1
                    if col >= 3:
                        col = 0
                        row += 1
            else:
                # Скрываем секцию, если направлений нет
                self.app_label.setVisible(False)
                self.app_scroll.setVisible(False)
        except Exception as e:
            self.app_label.setVisible(False)
            self.app_scroll.setVisible(False)
    
    def update_vat_fields(self):
        """Обновление полей НДС"""
        enabled = self.vat_applied.currentText() == 'Да'
        self.vat_rate.setEnabled(enabled)
    
    def calculate_item_cost(self, row):
        """Автоматический расчет стоимости позиции"""
        try:
            # Получаем количество
            qty_item = self.items_table.item(row, 1)
            qty = float(qty_item.text() if qty_item and qty_item.text() else 0)
            
            # Получаем фактическую цену, если указана, иначе плановую
            actual_price_item = self.items_table.item(row, 4)
            planned_price_item = self.items_table.item(row, 3)
            
            actual_price = float(actual_price_item.text() if actual_price_item and actual_price_item.text() else 0)
            planned_price = float(planned_price_item.text() if planned_price_item and planned_price_item.text() else 0)
            
            # Используем фактическую цену, если она указана, иначе плановую
            price = actual_price if actual_price > 0 else planned_price
            
            # Получаем НДС
            vat_item = self.items_table.item(row, 5)
            vat_rate = float(vat_item.text() if vat_item and vat_item.text() else 0)
            
            # Рассчитываем стоимость с НДС
            cost_without_vat = qty * price
            cost_with_vat = cost_without_vat * (1 + vat_rate / 100)
            
            # Обновляем ячейку стоимости
            cost_item = self.items_table.item(row, 6)
            if cost_item:
                cost_item.setText(f"{cost_with_vat:.2f}")
        except (ValueError, AttributeError):
            cost_item = self.items_table.item(row, 6)
            if cost_item:
                cost_item.setText('0')
    
    def handle_items_table_item_changed(self, item):
        """Обработчик изменений ячеек таблицы позиций договора"""
        if not item:
            return
        # Пересчитываем только при изменении количества, цен или НДС
        if item.column() in (1, 3, 4, 5):
            self.calculate_item_cost(item.row())

    def update_items_table_height(self):
        """Динамическая подстройка высоты таблицы позиций"""
        if not hasattr(self, 'items_table') or not self.items_table:
            return
        vh = self.items_table.verticalHeader()
        row_h = vh.defaultSectionSize()
        rows = max(5, self.items_table.rowCount())
        new_height = (
            rows * row_h
            + self.items_table.horizontalHeader().height()
            + 2 * self.items_table.frameWidth()
        )
        self.items_table.setMinimumHeight(new_height)
    
    def load_framework_contracts(self):
        """Загрузка рамочных договоров в комбобокс"""
        framework_contracts = self.db.get_framework_contracts()
        self.framework_contract_combo.clear()
        self.framework_contract_combo.addItem('')
        for contract in framework_contracts:
            display_text = f"{contract['number']} - {contract['contractor']} ({self.format_currency(contract['remainder'])})"
            self.framework_contract_combo.addItem(display_text, contract['id'])
    
    def on_framework_contract_selected(self):
        """Обработка выбора рамочного договора"""
        contract_id = self.framework_contract_combo.currentData()
        if contract_id:
            # Можно автоматически заполнить номер закупки и другие поля
            framework_contracts = self.db.get_framework_contracts()
            for contract in framework_contracts:
                if contract['id'] == contract_id:
                    # Заполняем номер закупки из рамочного договора
                    if contract.get('purchase_number'):
                        self.purchase_number.setText(str(contract['purchase_number']))
                    break
    
    def show_framework_contracts_dialog(self):
        """Показ диалога рамочных договоров"""
        dialog = QDialog(self)
        dialog.setWindowTitle('Рамочные договоры с остатками')
        dialog.setModal(True)
        dialog.resize(650, 350)  # Уменьшен размер
        
        layout = QVBoxLayout()
        dialog.setLayout(layout)
        
        label = QLabel('Выберите рамочный договор для использования:')
        label.setFont(QFont('Arial', 11, QFont.Weight.Bold))
        layout.addWidget(label)
        
        table = QTableWidget()
        table.setColumnCount(5)
        table.setHorizontalHeaderLabels(['Номер договора', 'Контрагент', 'Предмет договора', 'Остаток средств', 'Выбрать'])
        table.horizontalHeader().setStretchLastSection(False)
        table.setAlternatingRowColors(True)
        
        framework_contracts = self.db.get_framework_contracts()
        table.setRowCount(len(framework_contracts))
        
        for row_idx, contract in enumerate(framework_contracts):
            table.setItem(row_idx, 0, QTableWidgetItem(str(contract['number'])))
            table.setItem(row_idx, 1, QTableWidgetItem(str(contract['contractor'])))
            table.setItem(row_idx, 2, QTableWidgetItem(str(contract['subject'])))
            table.setItem(row_idx, 3, QTableWidgetItem(self.format_currency(contract['remainder'])))
            
            select_btn = QPushButton('Выбрать')
            select_btn.clicked.connect(lambda checked, c=contract: self.select_framework_contract(c, dialog))
            table.setCellWidget(row_idx, 4, select_btn)
        
        table.resizeColumnsToContents()
        layout.addWidget(table)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Close)
        buttons.rejected.connect(dialog.reject)
        layout.addWidget(buttons)
        
        dialog.exec()
    
    def select_framework_contract(self, contract, dialog):
        """Выбор рамочного договора"""
        # Устанавливаем выбранный договор в комбобокс
        for i in range(self.framework_contract_combo.count()):
            if self.framework_contract_combo.itemData(i) == contract['id']:
                self.framework_contract_combo.setCurrentIndex(i)
                break
        dialog.accept()
    
    def add_contract_item_row(self):
        """Добавление строки в таблицу позиций"""
        try:
            if not hasattr(self, 'items_table') or not self.items_table:
                QMessageBox.critical(self, 'Ошибка', 'Таблица позиций не инициализирована.')
                return

            row = self.items_table.rowCount()
            self.items_table.insertRow(row)

            # Наименование
            name_combo = QComboBox()
            name_combo.setEditable(True)
            # Подгружаем ВСЕ товары/услуги из Excel справочника (без ограничения по статье),
            # фильтрация будет выполняться по введённым буквам
            name_combo.addItem('', None)
            try:
                unique_rows = self.db.get_unique_items_from_excel()
                for u in unique_rows:
                    idx = name_combo.count()
                    name_combo.addItem(u['name'], u['id'])
                    if u.get('description'):
                        name_combo.setItemData(idx, u['description'], Qt.ItemDataRole.ToolTipRole)
            except Exception:
                pass
            # Автодополнение по товарам/услугам: список из всех вариантов, фильтр по вхождению
            completer = QCompleter(name_combo.model(), name_combo)
            completer.setCaseSensitivity(Qt.CaseSensitivity.CaseInsensitive)
            completer.setFilterMode(Qt.MatchFlag.MatchContains)
            completer.setCompletionMode(QCompleter.CompletionMode.PopupCompletion)
            name_combo.setCompleter(completer)
            self.items_table.setCellWidget(row, 0, name_combo)

            # Количество
            qty_item = QTableWidgetItem('0')
            qty_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.items_table.setItem(row, 1, qty_item)

            # Ед. изм. - выпадающий список
            unit_combo = QComboBox()
            unit_combo.addItems([
                'шт', 'кг', 'т', 'м', 'м²', 'м³', 'л', 'компл', 'набор',
                'упак', 'пачка', 'рулон', 'лист', 'блок', 'час', 'день',
                'мес', 'год', 'услуга', 'работа',
            ])
            unit_combo.setCurrentText('шт')
            self.items_table.setCellWidget(row, 2, unit_combo)

            # Плановая цена за ед.
            planned_price_item = QTableWidgetItem('0')
            planned_price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.items_table.setItem(row, 3, planned_price_item)

            # Фактическая цена за ед.
            actual_price_item = QTableWidgetItem('0')
            actual_price_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.items_table.setItem(row, 4, actual_price_item)

            # НДС %
            vat_item = QTableWidgetItem('20')
            vat_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.items_table.setItem(row, 5, vat_item)

            # Стоимость (автоматически) - используем фактическую цену, если она указана, иначе плановую
            cost_item = QTableWidgetItem('0')
            cost_item.setFlags(cost_item.flags() & ~Qt.ItemFlag.ItemIsEditable)
            cost_item.setTextAlignment(Qt.AlignmentFlag.AlignRight | Qt.AlignmentFlag.AlignVCenter)
            self.items_table.setItem(row, 6, cost_item)

            # Категория ФЭО
            category_item = QTableWidgetItem()
            self.items_table.setItem(row, 7, category_item)

            # Кнопка удаления
            delete_btn = QPushButton('🗑️')
            delete_btn.setMaximumWidth(40)
            delete_btn.clicked.connect(lambda: self.items_table.removeRow(row))
            self.items_table.setCellWidget(row, 8, delete_btn)

            # Первичный расчёт стоимости
            self.calculate_item_cost(row)
            # Подстраиваем высоту таблицы под количество строк
            self.update_items_table_height()
        except Exception as exc:  # noqa: BLE001
            QMessageBox.critical(
                self,
                'Ошибка добавления позиции',
                f'Не удалось добавить строку в таблицу позиций:\n{exc}',
            )
    
    def show_add_contractor_dialog(self):
        """Показ диалога добавления контрагента"""
        dialog = AddContractorDialog(self.db, self)
        if dialog.exec():
            self.load_contractors()
            self.load_contractors_table()
    
    def check_purchase_number_duplicate(self):
        """Проверка дубликата номера закупки"""
        purchase_number = self.purchase_number.text().strip()
        if not purchase_number:
            return
        
        contract_kind = self.contract_kind.currentText()
        is_framework = 'рамочн' in contract_kind.lower()
        
        # Проверяем существование
        exists, existing_id = self.db.check_purchase_number_exists(purchase_number)
        
        if exists and not is_framework:
            # Для нерамочных договоров дубликат недопустим
            reply = QMessageBox.question(
                self, 'Дубликат номера закупки',
                f'Номер закупки "{purchase_number}" уже существует!\n\n'
                f'Для рамочных договоров допускается повторение номера закупки.\n'
                f'Для других видов договоров номер должен быть уникальным.\n\n'
                f'Изменить номер закупки?',
                QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No
            )
            if reply == QMessageBox.StandardButton.Yes:
                self.purchase_number.setFocus()
            else:
                self.purchase_number.clear()
    
    def save_contract(self):
        """Сохранение договора"""
        try:
            # Получаем выбранную категорию ФЭО из комбобокса
            feo_direction = self.feo_direction_combo.currentText() if hasattr(self, 'feo_direction_combo') else ''
            
            selected_app = [cat for cat, cb in self.app_checkboxes.items() if cb.isChecked()]
            app_direction = ', '.join(selected_app) if selected_app else ''
            
            contract_data = {
                'contract_number': self.contract_number.text(),
                'contract_date': self.contract_date.date().toPyDate(),
                'contract_type': self.contract_type.currentText(),
                'contract_kind': self.contract_kind.currentText(),
                'subsidy_id': self.subsidy_combo.currentData(),
                'purchase_number': self.purchase_number.text(),
                'order_number': self.order_number.text(),
                'subject': self.subject.toPlainText(),
                'description': self.description.toPlainText(),
                'contractor_id': self.contractor_combo.currentData(),
                'status': self.status.currentText(),
                'execution_stage': self.execution_stage.currentText() if hasattr(self.execution_stage, 'currentText') else self.execution_stage.text(),
                'start_date': self.start_date.date().toPyDate(),
                'end_date': self.end_date.date().toPyDate(),
                'nmck': self.nmck.value(),
                'price_without_vat': self.price_without_vat.value(),
                'vat_applied': self.vat_applied.currentText(),
                'vat_rate': self.vat_rate.value(),
                'payment_method': self.payment_method.currentText(),
                'payment_form': self.payment_form.currentText(),
                'advance_amount': self.advance_amount.value(),
                'payment_term': self.payment_term.value(),
                'feo_direction': feo_direction,
                'feo_type': self.feo_type_combo.currentText(),
                'app_direction': app_direction,
                'specific_type': self.specific_type.currentText() if hasattr(self.specific_type, 'currentText') else self.specific_type.text(),
                'responsible': self.responsible.text(),
                'city': self.city.text(),
                'comments': self.comments.toPlainText(),
                'author': 'Пользователь'  # Можно получить из системы
            }
            
            contract_id = self.db.add_contract(contract_data)
            
            # Сохраняем позиции состава
            self.save_contract_items(contract_id)
            
            # Обновляем реестр
            self.load_registry_table()
            
            QMessageBox.information(self, 'Успех', f'Договор успешно сохранен!\nID: {contract_id}')
            self.statusBar().showMessage(f'Договор сохранен. ID: {contract_id}')
            
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Ошибка при сохранении договора:\n{str(e)}')
    
    def save_contract_items(self, contract_id):
        """Сохранение позиций состава договора"""
        ws = self.db.wb['Состав_договора']
        
        for row in range(self.items_table.rowCount()):
            # Наименование может быть либо в ячейке, либо в комбобоксе
            name_widget = self.items_table.cellWidget(row, 0)
            name_text = ''
            if isinstance(name_widget, QComboBox):
                name_text = name_widget.currentText().strip()
            else:
                name_item = self.items_table.item(row, 0)
                if name_item:
                    name_text = name_item.text().strip()
            if not name_text:
                continue
            
            next_row = ws.max_row + 1
            next_id = self.db.get_next_id('Состав_договора')
            
            ws.cell(row=next_row, column=1, value=next_id)
            ws.cell(row=next_row, column=2, value=contract_id)
            ws.cell(row=next_row, column=3, value=row + 1)
            ws.cell(row=next_row, column=4, value=name_text)
            ws.cell(row=next_row, column=6, value=float(self.items_table.item(row, 1).text() or 0))
            # Единица измерения из комбобокса
            unit_widget = self.items_table.cellWidget(row, 2)
            unit_value = unit_widget.currentText() if unit_widget else 'шт'
            ws.cell(row=next_row, column=7, value=unit_value)
            # Плановая цена за ед.
            planned_price = float(self.items_table.item(row, 3).text() or 0)
            ws.cell(row=next_row, column=9, value=planned_price)
            # Фактическая цена за ед.
            actual_price = float(self.items_table.item(row, 4).text() or 0)
            ws.cell(row=next_row, column=10, value=actual_price)
            # НДС
            vat_rate = float(self.items_table.item(row, 5).text() or 0)
            ws.cell(row=next_row, column=11, value=vat_rate)
            # Стоимость (рассчитанная)
            cost_item = self.items_table.item(row, 6)
            cost_value = float(cost_item.text() or 0) if cost_item else 0
            ws.cell(row=next_row, column=12, value=cost_value)
            
            # Категория ФЭО
            category_item = self.items_table.item(row, 7)
            if category_item:
                ws.cell(row=next_row, column=21, value=category_item.text())
        
        self.db.save_database()
    
    def clear_contract_form(self):
        """Очистка формы договора"""
        self.contract_number.clear()
        self.contract_date.setDate(QDate.currentDate())
        self.contract_type.setCurrentIndex(0)
        self.contract_kind.setCurrentIndex(0)
        self.subsidy_combo.setCurrentIndex(0)
        self.purchase_number.clear()
        self.order_number.clear()
        self.subject.clear()
        self.description.clear()
        self.status.setCurrentIndex(0)
        self.execution_stage.clear()
        self.nmck.setValue(0)
        self.price_without_vat.setValue(0)
        self.items_table.setRowCount(0)
        
        if hasattr(self, 'feo_checkboxes'):
            for cb in self.feo_checkboxes.values():
                cb.setChecked(False)
        for cb in self.app_checkboxes.values():
            cb.setChecked(False)
        
        self.feo_type_combo.clear()
        self.specific_type.clear()
        self.responsible.clear()
        self.city.clear()
        self.comments.clear()

class AddContractorDialog(QDialog):
    """Диалог добавления контрагента"""
    
    def __init__(self, db, parent=None):
        super().__init__(parent)
        self.db = db
        self.setWindowTitle('Добавить контрагента')
        self.setModal(True)
        self.resize(550, 600)  # Уменьшен размер
        
        layout = QVBoxLayout()
        self.setLayout(layout)
        
        scroll = QScrollArea()
        scroll_widget = QWidget()
        form_layout = QFormLayout()
        scroll_widget.setLayout(form_layout)
        
        self.name = QLineEdit()
        self.full_name = QLineEdit()
        self.inn = QLineEdit()
        self.kpp = QLineEdit()
        self.ogrn = QLineEdit()
        self.director_name = QLineEdit()
        self.director_position = QLineEdit()
        self.account = QLineEdit()
        self.bik = QLineEdit()
        self.bank_name = QLineEdit()
        self.address = QTextEdit()
        self.address.setMaximumHeight(60)
        self.legal_address = QTextEdit()
        self.legal_address.setMaximumHeight(60)
        self.actual_address = QTextEdit()
        self.actual_address.setMaximumHeight(60)
        self.phone = QLineEdit()
        self.email = QLineEdit()
        self.contact_person = QLineEdit()
        self.contact_phone = QLineEdit()
        self.signer_name = QLineEdit()  # Подписант
        self.signer_basis = QLineEdit()  # Основание действия
        
        form_layout.addRow('Контрагент *:', self.name)
        form_layout.addRow('Полное наименование:', self.full_name)
        form_layout.addRow('ИНН *:', self.inn)
        form_layout.addRow('КПП:', self.kpp)
        form_layout.addRow('ОГРН:', self.ogrn)
        form_layout.addRow('ФИО руководителя:', self.director_name)
        form_layout.addRow('Должность руководителя:', self.director_position)
        form_layout.addRow('Подписант:', self.signer_name)
        form_layout.addRow('Основание действия:', self.signer_basis)
        form_layout.addRow('Расчетный счет:', self.account)
        form_layout.addRow('БИК:', self.bik)
        form_layout.addRow('Наименование банка:', self.bank_name)
        form_layout.addRow('Юридический адрес:', self.legal_address)
        form_layout.addRow('Фактический адрес:', self.actual_address)
        form_layout.addRow('Адрес (общий):', self.address)
        form_layout.addRow('Телефон:', self.phone)
        form_layout.addRow('E-mail:', self.email)
        form_layout.addRow('Контактное лицо:', self.contact_person)
        form_layout.addRow('Телефон контакта:', self.contact_phone)
        
        scroll.setWidget(scroll_widget)
        scroll.setWidgetResizable(True)
        layout.addWidget(scroll)
        
        buttons = QDialogButtonBox(QDialogButtonBox.StandardButton.Ok | QDialogButtonBox.StandardButton.Cancel)
        buttons.accepted.connect(self.accept)
        buttons.rejected.connect(self.reject)
        layout.addWidget(buttons)
    
    def accept(self):
        if not self.name.text() or not self.inn.text():
            QMessageBox.warning(self, 'Ошибка', 'Заполните обязательные поля: Контрагент и ИНН')
            return
        
        try:
            ws = self.db.wb['Контрагенты']
            next_row = ws.max_row + 1
            next_id = self.db.get_next_id('Контрагенты')
            
            ws.cell(row=next_row, column=1, value=next_id)
            ws.cell(row=next_row, column=2, value=self.name.text())
            ws.cell(row=next_row, column=3, value=self.full_name.text())
            ws.cell(row=next_row, column=4, value=self.inn.text())
            ws.cell(row=next_row, column=5, value=self.kpp.text())
            ws.cell(row=next_row, column=6, value=self.ogrn.text())
            ws.cell(row=next_row, column=9, value=self.director_name.text())
            ws.cell(row=next_row, column=10, value=self.director_position.text())
            ws.cell(row=next_row, column=15, value=self.account.text())
            ws.cell(row=next_row, column=17, value=self.bik.text())
            ws.cell(row=next_row, column=18, value=self.bank_name.text())
            ws.cell(row=next_row, column=19, value=self.address.toPlainText())
            # Добавляем новые поля
            if hasattr(self, 'legal_address'):
                ws.cell(row=next_row, column=20, value=self.legal_address.toPlainText())
            if hasattr(self, 'actual_address'):
                ws.cell(row=next_row, column=21, value=self.actual_address.toPlainText())
            ws.cell(row=next_row, column=22, value=self.phone.text())
            ws.cell(row=next_row, column=24, value=self.email.text())
            if hasattr(self, 'signer_name'):
                ws.cell(row=next_row, column=11, value=self.signer_name.text())
            if hasattr(self, 'signer_basis'):
                ws.cell(row=next_row, column=12, value=self.signer_basis.text())
            ws.cell(row=next_row, column=26, value=self.contact_person.text())
            ws.cell(row=next_row, column=28, value=self.contact_phone.text())
            ws.cell(row=next_row, column=31, value=datetime.now())
            ws.cell(row=next_row, column=32, value=datetime.now())
            ws.cell(row=next_row, column=33, value='Да')
            
            self.db.save_database()
            super().accept()
        except Exception as e:
            QMessageBox.critical(self, 'Ошибка', f'Ошибка при сохранении:\n{str(e)}')

def main():
    app = QApplication(sys.argv)
    
    # Устанавливаем стиль
    app.setStyle('Fusion')
    
    # Синяя тема с темным шрифтом
    palette = QPalette()
    # Фон окна - синий
    palette.setColor(QPalette.ColorRole.Window, QColor(52, 96, 146))  # #346092
    # Фон виджетов - светло-синий
    palette.setColor(QPalette.ColorRole.Base, QColor(240, 248, 255))  # AliceBlue
    # Текст - темный для читаемости
    palette.setColor(QPalette.ColorRole.WindowText, QColor(0, 0, 0))  # Черный
    palette.setColor(QPalette.ColorRole.Text, QColor(0, 0, 0))  # Черный
    palette.setColor(QPalette.ColorRole.ButtonText, QColor(0, 0, 0))  # Черный
    # Кнопки - светлые
    palette.setColor(QPalette.ColorRole.Button, QColor(240, 248, 255))  # AliceBlue
    app.setPalette(palette)
    
    # Глобальный стиль для лучшей читаемости
    app.setStyleSheet("""
        QWidget {
            background-color: #346092;
            color: #000000;
        }
        QGroupBox {
            background-color: #E6F2FF;
            border: 2px solid #1E3A5F;
            border-radius: 5px;
            margin-top: 10px;
            padding-top: 10px;
            font-weight: bold;
            color: #000000;
        }
        QGroupBox::title {
            subcontrol-origin: margin;
            left: 10px;
            padding: 0 5px;
        }
        QLabel {
            color: #000000;
            background-color: transparent;
        }
        QLineEdit, QTextEdit, QComboBox, QSpinBox, QDoubleSpinBox, QDateEdit {
            background-color: #FFFFFF;
            color: #000000;
            border: 1px solid #1E3A5F;
            padding: 3px;
        }
        QTableWidget {
            background-color: #FFFFFF;
            color: #000000;
            gridline-color: #1E3A5F;
        }
        QHeaderView::section {
            background-color: #1E3A5F;
            color: #FFFFFF;
            padding: 5px;
            border: none;
        }
        QScrollArea {
            background-color: #E6F2FF;
            border: 1px solid #1E3A5F;
        }
        QCheckBox {
            color: #000000;
            background-color: transparent;
        }
        QCheckBox::indicator {
            background-color: #FFFFFF;
            border: 1px solid #1E3A5F;
        }
        QCheckBox::indicator:checked {
            background-color: #4CAF50;
        }
    """)
    
    window = MainWindow()
    window.show()
    
    sys.exit(app.exec())

if __name__ == '__main__':
    main()

