"""Phase 22 — парсер Excel-выписок банка/казначейства.

Читает xlsx (ScrollerHash или Scroller_FADM_2026), нормализует заголовки,
маппит известные колонки на нормализованные поля, остальное складывает в
raw_json. Возвращает список ParsedRow для дальнейшего сохранения в
bank_payments через bank_statement_parser_save.

Одна строка xlsx = один платёж по одному договору (multi-row split удалён).

Header dict — словарь нормализованных имён → имя поля BankPayment.
"""
from __future__ import annotations

import hashlib
import json
import re
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import Any, Optional

from openpyxl import load_workbook


# ---------------------------------------------------------------------------
# compute_row_hash — SHA-256 от нормализованного JSON всей строки xlsx
# ---------------------------------------------------------------------------

def compute_row_hash(row_dict: dict) -> str:
    """SHA-256 от нормализованного JSON всей строки xlsx.

    Стабилен: sorted keys, str(value).strip(), пустые значения нормализованы в "".
    Две идентичные строки xlsx → один hash. Любое отличие (включая регистры/whitespace) → разный hash.
    """
    normalized = {}
    for k, v in row_dict.items():
        key = str(k).strip().upper().replace('  ', ' ')  # collapse double spaces in key
        if v is None or v == '':
            normalized[key] = ''
        elif isinstance(v, (int, float)):
            normalized[key] = repr(v)  # deterministic float repr
        else:
            normalized[key] = str(v).strip()
    serialized = json.dumps(normalized, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(serialized.encode('utf-8')).hexdigest()


# ---------------------------------------------------------------------------
# Нормализатор заголовков
# ---------------------------------------------------------------------------

def _norm_header(s: Any) -> str:
    if s is None:
        return ""
    return re.sub(r"\s+", " ", str(s).upper().strip())


# ---------------------------------------------------------------------------
# Header → BankPayment field
# ---------------------------------------------------------------------------

HEADER_MAP: dict[str, str] = {
    "НОМЕР ДОКУМЕНТА": "payment_number",
    "ДАТА ДОКУМЕНТА": "payment_date",
    "ДАТА ПЛАТЕЖА": "payment_date",
    "ДАТА ДОК.": "payment_date",
    "ДАТА ДОК": "payment_date",
    "ДАТА": "payment_date",
    "СТАТУС ДОКУМЕНТА": "status",
    "СУММА": "amount",
    "ДАТА ИСПОЛНЕНИЯ ОПЕРАЦИИ": "execution_date",
    "ДАТА ПРОВОДКИ": "posting_date",
    "ДАТА И ВРЕМЯ УТВЕРЖДЕНИЯ ЦС": "execution_datetime",
    # Назначение платежа — несколько возможных заголовков
    "РАСШИФРОВКА П/П/КОНТРАКТ (ДОГОВОР)": "purpose_text",
    "НАЗНАЧЕНИЕ ПЛАТЕЖА": "purpose_text",
    "ИСТОЧНИК ОПЛАТЫ": "payment_source",
    # Колонка «Документ-основание» (соглашение о субсидии)
    "ДОКУМЕНТ-ОСНОВАНИЕ": "basis_doc_text",
    "ДОКУМЕНТ ОСНОВАНИЕ": "basis_doc_text",
    # Шифр субсидии
    "АНАЛИТИЧЕСКИЙ КОД РАЗДЕЛА ПЛАТЕЛЬЩИКА/КОД СУБСИДИИ (ЦЕЛИ)": "subsidy_code",
    "КОД СУБСИДИИ (ЦЕЛИ)": "subsidy_code",
    "ШИФР СУБСИДИИ": "subsidy_code",
    # Реквизиты плательщика — старый формат (fallback)
    "ИНН ПЛАТЕЛЬЩИКА": "payer_inn",
    "КПП ПЛАТЕЛЬЩИКА": "payer_kpp",
    "НАИМЕНОВАНИЕ ПЛАТЕЛЬЩИКА": "payer_name",
    "СЧЕТ ПЛАТЕЛЬЩИКА": "payer_account",
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА": "payer_block",
    # Реквизиты получателя — старый формат (fallback)
    "ИНН ПОЛУЧАТЕЛЯ": "payee_inn",
    "КПП ПОЛУЧАТЕЛЯ": "payee_kpp",
    "НАИМЕНОВАНИЕ ПОЛУЧАТЕЛЯ": "payee_name",
    "СЧЕТ ПОЛУЧАТЕЛЯ": "payee_account",
    "БИК ПОЛУЧАТЕЛЯ": "payee_bik",
    "БАНК ПОЛУЧАТЕЛЯ": "payee_bank",
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ": "payee_block",
    # Новый универсальный формат "main (sub)" — composite-ключи
    # Реквизиты плательщика
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА (ИНН)": "payer_inn",
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА (КПП)": "payer_kpp",
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА (НАИМЕНОВАНИЕ)": "payer_name",
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА (ЛИЦЕВОЙ СЧЕТ)": "payer_account",
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА (НАИМЕНОВАНИЕ БАНКА)": "payer_bank",
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА (БИК)": "payer_bik",
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА (РАСЧЕТНЫЙ СЧЕТ)": "payer_settlement_account",
    "РЕКВИЗИТЫ ПЛАТЕЛЬЩИКА (КОРР. СЧЕТ)": "payer_corr_account",
    # Реквизиты получателя
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ (ИНН)": "payee_inn",
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ (КПП)": "payee_kpp",
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ (НАИМЕНОВАНИЕ)": "payee_name",
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ (ЛИЦЕВОЙ СЧЕТ)": "payee_account",
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ (НАИМЕНОВАНИЕ БАНКА)": "payee_bank",
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ (БИК)": "payee_bik",
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ (РАСЧЕТНЫЙ СЧЕТ)": "payee_settlement_account",
    "РЕКВИЗИТЫ ПОЛУЧАТЕЛЯ (КОРР. СЧЕТ)": "payee_corr_account",
    # Дополнительная информация
    "ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ (НАЗНАЧЕНИЕ ПЛАТЕЖА)": "purpose_text",
    "ДОПОЛНИТЕЛЬНАЯ ИНФОРМАЦИЯ (УПНО)": "upno",
}

# Финальные ИСПОЛНЕНО-эквивалентные статусы
EXECUTED_STATUSES = {"ИСПОЛНЕН", "ИСПОЛНЕНО", "ОТРАЖЕНО НА Л/С ПЛАТЕЛЬЩИКА"}

# Откатанные/отменённые — пропускаются для матча, но сохраняются
REJECTED_STATUSES = {
    "АННУЛИРОВАН",
    "ОТМЕНЕН",
    "ОТКАЗАНО",
    "НЕ НАПРАВЛЯЛОСЬ В БАНК",
    "КОНТРОЛЬ ФК НЕ ПРОЙДЕН",
}


# ---------------------------------------------------------------------------
# Regex
# ---------------------------------------------------------------------------

# КБК в скобках вида (712ZU7L4002;0300022) — берём первую группу до «;» или «)»
RX_KBK = re.compile(r"\(([0-9]{3}[A-ZА-Я0-9]{6,10})[;)]", re.IGNORECASE)

# ИНН из текстовых блоков (10 или 12 цифр)
RX_INN = re.compile(r"\b(\d{10}|\d{12})\b")

RX_VAT = re.compile(r"НДС\s*([\d.,]+)", re.IGNORECASE)

# ---------------------------------------------------------------------------
# DOC_PATTERNS — все типы документов из назначения платежа
# ---------------------------------------------------------------------------

DOC_PATTERNS: dict[str, re.Pattern] = {
    "contracts": re.compile(
        r"(?P<type>ДОГ(?:ОВОР)?|КОНТРАКТ)"
        r"\.?\s*№?\s*(?P<num>[0-9\-/А-ЯA-Z]+)"
        r"(?:\s+ОТ\s+(?P<date>\d{2}\.\d{2}\.\d{4}))?",
        re.IGNORECASE | re.UNICODE,
    ),
    "agreements": re.compile(
        r"СОГЛАШ(?:ЕНИЕ)?"
        r"\.?\s*№?\s*(?P<num>[0-9\-/А-ЯA-Z]+)"
        r"(?:\s+ОТ\s+(?P<date>\d{2}\.\d{2}\.\d{4}))?",
        re.IGNORECASE | re.UNICODE,
    ),
    "acts": re.compile(
        r"АКТ(?:\s*(?:ПРИ[ЁЕ]МКИ|ВЫПОЛНЕННЫХ\s*РАБОТ))?"
        r"\.?\s*№?\s*(?P<num>[0-9\-/]+)"
        r"(?:\s+ОТ\s+(?P<date>\d{2}\.\d{2}\.\d{4}))?",
        re.IGNORECASE | re.UNICODE,
    ),
    "invoices": re.compile(
        r"СЧ[\.\s]?(?:Ф\.?|-?ФАКТУРА)?"
        r"\s*№?\s*(?P<num>[0-9\-/]+)"
        r"(?:\s+ОТ\s+(?P<date>\d{2}\.\d{2}\.\d{4}))?",
        re.IGNORECASE | re.UNICODE,
    ),
    "upd": re.compile(
        r"УПД\.?\s*№?\s*(?P<num>[0-9\-/А-ЯA-Z]+)"
        r"(?:\s+ОТ\s+(?P<date>\d{2}\.\d{2}\.\d{4}))?",
        re.IGNORECASE | re.UNICODE,
    ),
    "ttn": re.compile(
        r"ТТН\.?\s*№?\s*(?P<num>[0-9\-/]+)"
        r"(?:\s+ОТ\s+(?P<date>\d{2}\.\d{2}\.\d{4}))?",
        re.IGNORECASE | re.UNICODE,
    ),
    "registry": re.compile(
        r"РЕЕСТР\s*(?:ДОК\.?-?ОСН\.?)?\s*№?\s*(?P<num>[0-9\-/]+)"
        r"(?:\s+ОТ\s+(?P<date>\d{2}\.\d{2}\.\d{4}))?",
        re.IGNORECASE | re.UNICODE,
    ),
    "advance_reports": re.compile(
        r"АВАНСОВ(?:ЫЙ)?\s*ОТЧ[ЕЁ]Т\.?\s*№?\s*(?P<num>[0-9\-/]+)"
        r"\s*(?:ОТ\s*(?P<date>\d{2}\.\d{2}\.\d{4}))?",
        re.IGNORECASE | re.UNICODE,
    ),
}

# Для быстрого SQL-фильтра — приоритет: ДОГОВОР > СОГЛАШЕНИЕ > КОНТРАКТ > РЕЕСТР
RX_CONTRACT_PARTS = [
    ("ДОГОВОР",    re.compile(r"ДОГОВОР\s+([0-9А-ЯA-Z/\-\.]+)", re.IGNORECASE | re.UNICODE)),
    ("СОГЛАШЕНИЕ", re.compile(r"СОГЛАШЕНИЕ\s+([0-9А-ЯA-Z/\-\.]+)", re.IGNORECASE | re.UNICODE)),
    ("КОНТРАКТ",   re.compile(r"КОНТРАКТ\s+([0-9А-ЯA-Z/\-\.]+)", re.IGNORECASE | re.UNICODE)),
    ("РЕЕСТР",     re.compile(r"РЕЕСТР\s+ДОК\.?-?ОСН\.?\s+([0-9А-ЯA-Z/\-\.]+)", re.IGNORECASE | re.UNICODE)),
]

RX_DATE = re.compile(r"ОТ\s+(\d{2}\.\d{2}\.\d{4})", re.IGNORECASE)

# Basis doc: «№ xxx ОТ ДД.ММ.ГГГГ» из колонки «Документ-основание»
BASIS_DOC_PATTERN = re.compile(
    r"№?\s*([0-9А-ЯA-Z\-/]+)\s+ОТ\s+(\d{2}\.\d{2}\.\d{4})",
    re.IGNORECASE | re.UNICODE,
)


# ---------------------------------------------------------------------------
# extract_all_documents
# ---------------------------------------------------------------------------

def extract_all_documents(purpose_text: str) -> dict:
    """Возвращает {'contracts': [...], 'acts': [...], ...} — все находки из назначения платежа."""
    if not purpose_text:
        return {}
    result: dict[str, list] = {}
    for key, pattern in DOC_PATTERNS.items():
        matches = []
        for m in pattern.finditer(purpose_text):
            try:
                num = m.group("num")
            except IndexError:
                num = None
            try:
                dt = m.group("date")
            except IndexError:
                dt = None
            if num:
                matches.append({"number": num, "date": dt})
        if matches:
            result[key] = matches
    return result


# ---------------------------------------------------------------------------
# parse_basis_doc
# ---------------------------------------------------------------------------

def parse_basis_doc(basis_doc_text: str) -> tuple[Optional[str], Optional[date]]:
    """Парсит «№ xxx ОТ ДД.ММ.ГГГГ» из колонки «Документ-основание»."""
    if not basis_doc_text:
        return None, None
    m = BASIS_DOC_PATTERN.search(basis_doc_text)
    if not m:
        return None, None
    num = m.group(1).strip()
    try:
        d = datetime.strptime(m.group(2), "%d.%m.%Y").date()
    except ValueError:
        d = None
    return num or None, d


# ---------------------------------------------------------------------------
# parse_purpose — устаревшее имя сохранено для совместимости
# ---------------------------------------------------------------------------

def parse_purpose(text: str) -> dict:
    """Возвращает {kbk, contract_number, contract_date, vat}. Все опциональны.

    Приоритет типа: ДОГОВОР > СОГЛАШЕНИЕ > КОНТРАКТ > РЕЕСТР.
    """
    if not text:
        return {}

    result: dict[str, Any] = {}

    kbk_m = RX_KBK.search(text)
    if kbk_m:
        result["kbk"] = kbk_m.group(1).upper()

    contract_number = None
    contract_match_end = None
    for _label, rx in RX_CONTRACT_PARTS:
        # СОГЛАШЕНИЕ — это субсидия, не договор; не записываем в parsed_contract_number
        if _label == "СОГЛАШЕНИЕ":
            continue
        matches = list(rx.finditer(text))
        if matches:
            m = matches[-1]
            contract_number = m.group(1)
            contract_match_end = m.end()
            break

    if contract_number:
        result["contract_number"] = contract_number
        date_matches = list(RX_DATE.finditer(text))
        for dm in date_matches:
            if dm.start() >= (contract_match_end or 0):
                try:
                    result["contract_date"] = datetime.strptime(dm.group(1), "%d.%m.%Y").date()
                except ValueError:
                    pass
                break

    vat_m = RX_VAT.search(text)
    if vat_m:
        vat_s = vat_m.group(1).replace(",", ".")
        try:
            result["vat"] = Decimal(vat_s)
        except InvalidOperation:
            pass

    return result


# ---------------------------------------------------------------------------
# Вспомогательные преобразователи
# ---------------------------------------------------------------------------

def _to_decimal(v: Any) -> Optional[Decimal]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return Decimal(str(v))
        except InvalidOperation:
            return None
    s = str(v).strip().replace(",", ".").replace(" ", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _to_date(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = str(v).strip()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_datetime(v: Any) -> Optional[datetime]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v).strip()
    for fmt in ("%d.%m.%Y %H:%M:%S", "%d.%m.%Y %H:%M", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def _norm_inn(v: Any) -> Optional[str]:
    """Обрезает .0 от числового представления, strip, None если пусто."""
    if v is None:
        return None
    if isinstance(v, float):
        s = str(int(v))
    else:
        s = str(v).strip()
        if s.endswith(".0"):
            s = s[:-2]
    return s if s else None


# ---------------------------------------------------------------------------
# ParsedRow
# ---------------------------------------------------------------------------

@dataclass
class ParsedRow:
    payment_number: Optional[str] = None
    payment_date: Optional[date] = None
    execution_datetime: Optional[datetime] = None
    status: Optional[str] = None
    amount: Optional[Decimal] = None

    payer_inn: Optional[str] = None
    payer_kpp: Optional[str] = None
    payer_name: Optional[str] = None
    payer_account: Optional[str] = None

    payee_inn: Optional[str] = None
    payee_kpp: Optional[str] = None
    payee_name: Optional[str] = None
    payee_account: Optional[str] = None
    payee_bik: Optional[str] = None
    payee_bank: Optional[str] = None

    purpose_text: Optional[str] = None
    parsed_contract_number: Optional[str] = None
    parsed_contract_date: Optional[date] = None
    parsed_kbk: Optional[str] = None
    parsed_documents: Optional[dict] = None

    basis_doc_text: Optional[str] = None
    basis_doc_number: Optional[str] = None
    basis_doc_date: Optional[date] = None
    subsidy_code: Optional[str] = None

    raw_json: dict = field(default_factory=dict)
    source_row_hash: Optional[str] = None
    is_executed: bool = False
    skip_reason: Optional[str] = None


# ---------------------------------------------------------------------------
# _find_header_row
# ---------------------------------------------------------------------------

import logging as _logging

_parser_log = _logging.getLogger(__name__)

# Ключевые слова, однозначно указывающие на строку заголовков таблицы
_EXPECTED_HEADERS = {
    'НОМЕР ДОКУМЕНТА', 'ДАТА ДОКУМЕНТА', 'СУММА',
    'СТАТУС ДОКУМЕНТА', 'НАИМЕНОВАНИЕ ПЛАТЕЛЬЩИКА',
    'НАИМЕНОВАНИЕ ПОЛУЧАТЕЛЯ', 'НАЗНАЧЕНИЕ ПЛАТЕЖА',
}


def _find_header_row(ws) -> int:
    """Возвращает 1-based номер строки с заголовками таблицы.

    В реальных выписках первые 1-4 строки — metadata («Тип документа»,
    «Дата формирования», пустая строка и т.д.).  Ищем строку, где хотя бы
    одно из ожидаемых ключевых слов совпадает с нормализованным значением
    ячейки.  Fallback — строка 1 (совместимость с тестовыми файлами).
    """
    for row_idx in range(1, 15):  # ищем в первых 14 строках
        try:
            cells = [str(c.value or '').strip().upper() for c in ws[row_idx]]
        except Exception:
            break
        normalized = {re.sub(r'\s+', ' ', c) for c in cells if c}
        if normalized & _EXPECTED_HEADERS:
            _parser_log.info(f"bank_statement_parser: headers found at row {row_idx}")
            return row_idx
    _parser_log.warning("bank_statement_parser: header row not detected, using row 1 as fallback")
    return 1


# ---------------------------------------------------------------------------
# _extract_headers
# ---------------------------------------------------------------------------

def _expand_merged_row(ws, row_idx: int, max_col: int) -> list:
    """Возвращает значения row_idx с разворачиванием merged cells.

    Для merged range — все колонки получают value верхне-левой ячейки.
    """
    values = [None] * max_col
    # Обычные ячейки
    for cell in ws[row_idx]:
        col = cell.column - 1
        if 0 <= col < max_col:
            values[col] = cell.value
    # Развернуть merged-ranges на эту строку
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row_idx <= mr.max_row:
            top_left = ws.cell(row=mr.min_row, column=mr.min_col).value
            for col in range(mr.min_col - 1, mr.max_col):
                if 0 <= col < max_col:
                    values[col] = top_left
    return values


def _extract_headers(ws, header_row: int = 1) -> list[str]:
    """Composite-ключи из ДВУХ строк xlsx.

    Логика:
    - main и sub оба пустые → ""
    - только main → main
    - только sub → sub
    - main == sub (merged на 2 строки одно значение) → main
    - main != sub оба непустые → "main (sub)"  ← композит
    """
    max_col = ws.max_column
    main_values = _expand_merged_row(ws, header_row, max_col)
    try:
        sub_values = _expand_merged_row(ws, header_row + 1, max_col)
    except Exception:
        sub_values = [None] * max_col

    out: list[str] = []
    for i in range(max_col):
        main = _norm_header(main_values[i]) if main_values[i] not in (None, "") else ""
        sub = _norm_header(sub_values[i]) if sub_values[i] not in (None, "") else ""

        if not main and not sub:
            out.append("")
        elif not sub:
            out.append(main)
        elif not main:
            out.append(sub)
        elif main == sub:
            out.append(main)
        else:
            out.append(f"{main} ({sub})")

    _parser_log.info(f"bank_statement_parser: extracted {len(out)} composite headers")
    known = sum(1 for h in out if h in HEADER_MAP)
    _parser_log.info(f"bank_statement_parser: {known}/{len(out)} headers found in HEADER_MAP")
    return out


# ---------------------------------------------------------------------------
# _build_row — заполнить ParsedRow из значений ячеек + headers
# ---------------------------------------------------------------------------

def _build_row(
    headers: list[str],
    values: list[Any],
) -> Optional[ParsedRow]:
    """Строит ParsedRow из одной строки данных.

    Возвращает None если строка полностью пустая.
    """
    if all(v is None or str(v).strip() == "" for v in values):
        return None

    row = ParsedRow()
    raw: dict[str, Any] = {}

    col_groups: dict[str, list[Any]] = {}
    for h, v in zip(headers, values):
        if not h:
            continue
        col_groups.setdefault(h, []).append(v)

    def _json_safe(v: Any) -> Any:
        # JSONB-колонка не сериализует datetime/date/Decimal — конвертим в строки
        if v is None:
            return None
        if isinstance(v, datetime):
            return v.isoformat()
        if isinstance(v, date):
            return v.isoformat()
        if isinstance(v, Decimal):
            return str(v)
        if isinstance(v, (str, int, float, bool)):
            return v
        return str(v)

    for h, vals in col_groups.items():
        non_null = [v for v in vals if v is not None and str(v).strip() != ""]
        if not non_null:
            raw[h] = None
        elif len(non_null) == 1:
            raw[h] = _json_safe(non_null[0])
        else:
            raw[h] = "\n".join(str(_json_safe(v)) for v in non_null)

    row.raw_json = raw
    row.source_row_hash = compute_row_hash(raw)

    for h, field_name in HEADER_MAP.items():
        v = raw.get(h)
        if v is None:
            continue

        if field_name == "payment_number":
            row.payment_number = _norm_inn(v)
        elif field_name == "payment_date":
            row.payment_date = _to_date(v)
        elif field_name == "status":
            row.status = str(v).strip().upper() if v else None
        elif field_name == "amount":
            row.amount = _to_decimal(v)
        elif field_name == "execution_datetime":
            row.execution_datetime = _to_datetime(v)
        elif field_name == "payer_inn":
            row.payer_inn = _norm_inn(v)
        elif field_name == "payer_kpp":
            row.payer_kpp = _norm_inn(v)
        elif field_name == "payer_name":
            row.payer_name = str(v).strip() if v else None
        elif field_name == "payer_account":
            row.payer_account = str(v).strip() if v else None
        elif field_name == "payer_block":
            block_text = str(v)
            if not row.payer_inn:
                inn_m = RX_INN.search(block_text)
                if inn_m:
                    row.payer_inn = inn_m.group(1)
        elif field_name == "payee_inn":
            row.payee_inn = _norm_inn(v)
        elif field_name == "payee_kpp":
            row.payee_kpp = _norm_inn(v)
        elif field_name == "payee_name":
            row.payee_name = str(v).strip() if v else None
        elif field_name == "payee_account":
            row.payee_account = str(v).strip() if v else None
        elif field_name == "payee_bik":
            row.payee_bik = str(v).strip() if v else None
        elif field_name == "payee_bank":
            row.payee_bank = str(v).strip() if v else None
        elif field_name == "payee_block":
            block_text = str(v)
            if not row.payee_inn:
                inn_m = RX_INN.search(block_text)
                if inn_m:
                    row.payee_inn = inn_m.group(1)
        elif field_name == "purpose_text":
            if not row.purpose_text:  # берём первое найденное
                row.purpose_text = str(v).strip() if v else None
        elif field_name == "basis_doc_text":
            row.basis_doc_text = str(v).strip() if v else None
        elif field_name == "subsidy_code":
            row.subsidy_code = str(v).strip() if v else None

    # Парсим purpose_text
    if row.purpose_text:
        pp = parse_purpose(row.purpose_text)
        row.parsed_contract_number = pp.get("contract_number")
        row.parsed_contract_date = pp.get("contract_date")
        row.parsed_kbk = pp.get("kbk")
        row.parsed_documents = extract_all_documents(row.purpose_text)
        # Заполняем parsed_contract_number из parsed_documents.contracts[0] если parse_purpose не нашёл
        if not row.parsed_contract_number and row.parsed_documents.get("contracts"):
            first = row.parsed_documents["contracts"][0]
            row.parsed_contract_number = first.get("number")
            if first.get("date"):
                try:
                    row.parsed_contract_date = datetime.strptime(first["date"], "%d.%m.%Y").date()
                except ValueError:
                    pass

    # Парсим basis_doc_text
    if row.basis_doc_text:
        row.basis_doc_number, row.basis_doc_date = parse_basis_doc(row.basis_doc_text)

    # Статус
    status_norm = (row.status or "").upper().strip()
    if status_norm in EXECUTED_STATUSES:
        row.is_executed = True
    if status_norm in REJECTED_STATUSES:
        row.skip_reason = status_norm

    return row


# ---------------------------------------------------------------------------
# _is_numbering_row — строка нумерации колонок xlsx (1, 2, 3, ... 39)
# ---------------------------------------------------------------------------

def _is_numbering_row(values: list) -> bool:
    """Detect 'numbering row' — все непустые значения короткие числа 1-200.

    В некоторых xlsx после строки заголовков идёт строка вида:
    1, 2, 3, ..., 39 — нумерация позиций для ручного заполнения.
    Такую строку надо пропустить, иначе она попадает в реестр как платёж.
    """
    if not values:
        return False
    numeric_count = 0
    non_empty_count = 0
    for v in values:
        if v is None or str(v).strip() == "":
            continue
        non_empty_count += 1
        s = str(v).strip()
        try:
            n = int(float(s))
            if 1 <= n <= 200:
                numeric_count += 1
            else:
                return False  # число вне диапазона — не нумерация
        except (ValueError, OverflowError):
            return False  # хотя бы одно нечисло — это не строка нумерации
    # Минимум 5 числовых ячеек, и все непустые — числа 1-200
    return non_empty_count >= 5 and numeric_count == non_empty_count


# ---------------------------------------------------------------------------
# _find_first_data_row — пропустить sub-headers, numbering rows, empty
# ---------------------------------------------------------------------------

def _find_first_data_row(ws, header_row: int, headers: list[str]) -> int:
    """Найти первую строку с реальными данными после header_row.

    Пропускает sub-headers, numbering rows (24, 25, 26...), пустые строки.
    Признак data row: в колонке СУММА число >= 1, ИЛИ в НОМЕР ДОКУМЕНТА строка длиной >3.
    """
    sum_idx = next((i for i, h in enumerate(headers) if h.startswith('СУММА')), -1)
    docnum_idx = next((i for i, h in enumerate(headers) if h.startswith('НОМЕР ДОКУМЕНТА')), -1)

    for row_idx in range(header_row + 1, header_row + 12):
        try:
            row_vals = [c.value for c in ws[row_idx]]
        except Exception:
            continue

        if _is_numbering_row(row_vals):
            _parser_log.info(f"bank_statement_parser: skip numbering row at {row_idx}")
            continue

        # Пропустить пустую строку
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        # Hard signal: SUMMA >= 1
        if sum_idx >= 0 and sum_idx < len(row_vals):
            v = row_vals[sum_idx]
            if v is not None:
                try:
                    if float(v) >= 1:
                        _parser_log.info(f"bank_statement_parser: data starts at row {row_idx} (SUMMA={v})")
                        return row_idx
                except (ValueError, TypeError):
                    pass

        # Soft signal: docnum длиннее 3
        if docnum_idx >= 0 and docnum_idx < len(row_vals):
            v = row_vals[docnum_idx]
            if v is not None and len(str(v).strip()) > 3:
                _parser_log.info(f"bank_statement_parser: data starts at row {row_idx} (DOCNUM={v})")
                return row_idx

    fallback = header_row + 2
    _parser_log.warning(f"bank_statement_parser: no clear data row, fallback to {fallback}")
    return fallback


# ---------------------------------------------------------------------------
# parse_workbook — главная функция
# ---------------------------------------------------------------------------

def parse_workbook(
    file_bytes: bytes,
    sheet_name: Optional[str] = None,
) -> tuple[str, list[ParsedRow]]:
    """Главная функция парсера.

    1. Открывает xlsx
    2. Берёт sheet_name или первый активный
    3. Читает headers, нормализует, ищет в HEADER_MAP индексы
    4. Для каждой data row:
       - заполняет ParsedRow (известные поля)
       - сохраняет ВСЕ значения в raw_json по нормализованному ключу
       - парсит purpose_text → contract_number/date/kbk + все parsed_documents
       - парсит basis_doc_text → basis_doc_number/date
       - заполняет subsidy_code из «Аналитический код раздела...»
       - ставит is_executed=True если status IN EXECUTED_STATUSES
    5. Возвращает (sheet_name, list[ParsedRow])
    """
    wb = load_workbook(BytesIO(file_bytes), data_only=True)  # без read_only=True для корректной работы с merged_cells

    if sheet_name and sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        active_name = sheet_name
    else:
        ws = wb.active
        active_name = ws.title

    header_row = _find_header_row(ws)
    headers = _extract_headers(ws, header_row)
    data_start_row = _find_first_data_row(ws, header_row, headers)
    _parser_log.info(f"bank_statement_parser: parsing headers={len(headers)}, data starts at row {data_start_row}")

    parsed: list[ParsedRow] = []

    rows_iter = ws.iter_rows(min_row=data_start_row, values_only=True)
    for raw_values in rows_iter:
        values = list(raw_values)

        while len(values) < len(headers):
            values.append(None)

        if _is_numbering_row(values):  # ещё одна защита если случайно попалось
            _parser_log.info("bank_statement_parser: skip numbering row in data section")
            continue

        row = _build_row(headers, values)
        if row is None:
            continue
        parsed.append(row)

    wb.close()
    return active_name, parsed
