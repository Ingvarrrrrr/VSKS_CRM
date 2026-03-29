from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Body
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, delete, or_
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.models.purchase import Purchase
from app.models.purchase_item import PurchaseItem
from app.models.contractor import Contractor
from app.models.contract import Contract
from app.models.subsidy import Subsidy
from app.models.product import Product
from app.models.feo_category import FeoCategory
from app.schemas.schemas import PurchaseCreate, PurchaseOut, PurchaseOutFull, PurchaseItemOut, PurchaseFileOut, SubsidyAllocationOut
from app.models.subsidy_allocation import PurchaseSubsidyAllocation
from app.auth.jwt import get_current_user, require_role, get_org_filter, get_single_org_id, ADMIN_ROLES, MANAGER_ROLES, ALL_ROLES
from app.models.user import User
from app.routers.contracts import ensure_contract_linked
from typing import List, Optional
from decimal import Decimal
from io import BytesIO
from datetime import datetime, date
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None
    load_workbook = None

try:
    import pdfplumber as _pdfplumber
except ImportError:
    _pdfplumber = None

try:
    from docx import Document as _DocxDocument
except ImportError:
    _DocxDocument = None

router = APIRouter(prefix="/api/purchases", tags=["purchases"])

# Status workflow
STATUS_ORDER = ["wishes", "plan_schedule", "confirmed", "work_in_progress", "contracted", "delivered", "paid"]
VALID_SUBSTATUSES = ("tz_forming", "kp_collecting", "on_platform")

# ---------------------------------------------------------------------------
# Excel export: column registry
# ---------------------------------------------------------------------------

ALL_EXPORT_COLUMNS = {
    "purchase_number":        {"label": "№ п/п",                 "group": "Идентификация"},
    "order_number":           {"label": "Номер заявки",          "group": "Идентификация"},
    "registry_number":        {"label": "Реестровый №",          "group": "Идентификация"},
    "status":                 {"label": "Статус",                "group": "Идентификация"},
    "subsidy":                {"label": "Субсидия",              "group": "Идентификация"},
    "feo_category":           {"label": "Категория ФЭО",         "group": "Идентификация"},
    "item_name":              {"label": "Наименование",          "group": "Позиция"},
    "item_type":              {"label": "Тип",                   "group": "Позиция"},
    "unit":                   {"label": "Ед. изм",               "group": "Позиция"},
    "quantity":               {"label": "Кол-во",                "group": "Позиция"},
    "subject":                {"label": "Предмет закупки",       "group": "Позиция"},
    "country_origin":         {"label": "Страна происхождения",  "group": "Позиция"},
    "planned_unit_price":     {"label": "Плановая цена за ед.",  "group": "Цены"},
    "planned_total_price":    {"label": "Плановая сумма",        "group": "Цены"},
    "nmck":                   {"label": "НМЦК",                  "group": "Цены"},
    "contract_price":         {"label": "Цена договора",         "group": "Цены"},
    "economy":                {"label": "Экономия",              "group": "Цены"},
    "price_increase":         {"label": "Удорожание",            "group": "Цены"},
    "purchase_method":        {"label": "Способ закупки",        "group": "Закупка"},
    "purchase_basis":         {"label": "Основание",             "group": "Закупка"},
    "purchase_contract_type": {"label": "Тип договора",          "group": "Закупка"},
    "framework_seq":          {"label": "№ в рамочном",          "group": "Закупка"},
    "contract_number":        {"label": "№ договора",            "group": "Договор"},
    "contract_date":          {"label": "Дата договора",         "group": "Договор"},
    "execution_term":         {"label": "Срок исполнения",       "group": "Договор"},
    "execution_term_changed": {"label": "Срок (изменён)",        "group": "Договор"},
    "delivery_date":          {"label": "Дата доставки",         "group": "Договор"},
    "contractor":             {"label": "Контрагент",            "group": "Контрагент"},
    "contractor_inn":         {"label": "ИНН контрагента",       "group": "Контрагент"},
    "responsible_person":     {"label": "Ответственное лицо",    "group": "Контрагент"},
    "acceptance_doc_name":    {"label": "Закрывающий документ: наименование", "group": "Исполнение"},
    "acceptance_doc_number":  {"label": "Закрывающий документ: №",           "group": "Исполнение"},
    "acceptance_doc_date":    {"label": "Закрывающий документ: дата",         "group": "Исполнение"},
    "acceptance_doc_amount":  {"label": "Закрывающий документ: сумма",        "group": "Исполнение"},
    "payment_doc_number":     {"label": "ПП: №",                 "group": "Оплата"},
    "payment_doc_date":       {"label": "ПП: дата",              "group": "Оплата"},
    "payment_amount":         {"label": "ПП: сумма",             "group": "Оплата"},
    "payment_federal":        {"label": "В т.ч. фед. бюджет",   "group": "Оплата"},
    "delivery_payment_amount":{"label": "Оплата с доставкой",    "group": "Оплата"},
    "vat_applicable":         {"label": "НДС применяется",       "group": "НДС"},
    "vat_rate":               {"label": "Ставка НДС",            "group": "НДС"},
    "vat_exemption_article":  {"label": "Статья НК РФ",          "group": "НДС"},
}

DEFAULT_EXPORT_COLUMNS = [
    "purchase_number", "registry_number", "item_name", "item_type", "unit", "quantity",
    "nmck", "contract_price", "economy", "purchase_method",
    "contract_number", "contract_date", "contractor", "contractor_inn",
    "execution_term", "country_origin",
    "acceptance_doc_name", "acceptance_doc_number", "acceptance_doc_date", "acceptance_doc_amount",
    "payment_doc_number", "payment_doc_date", "payment_amount", "payment_federal",
    "status",
]

_PURCHASE_METHOD_LABELS = {
    "single": "Единственный поставщик",
    "competitive": "Конкурсная процедура",
    "quote_request": "Запрос котировок",
}
_PURCHASE_BASIS_LABELS = {
    "plan_schedule": "план-график",
    "service_note": "служебная записка",
}
_CONTRACT_TYPE_LABELS = {
    "single": "Разовая поставка",
    "framework_cumulative": "Рамочный (нарастающий итог)",
    "framework_with_amount": "Рамочный (с указанием суммы)",
}
_STATUS_LABELS = {
    "wishes": "Желания сотрудников",
    "plan_schedule": "План-график",
    "confirmed": "Подтверждено руководством",
    "work_in_progress": "Ведётся работа",
    "contracted": "Заключён договор",
    "delivered": "Поставлено",
    "paid": "Оплачено",
}
_SUBSTATUS_LABELS = {
    "tz_forming": "Формируется ТЗ",
    "kp_collecting": "Идёт сбор КП",
    "on_platform": "Выставлено на площадку",
}


def _get_cell_value(key: str, p: Purchase, ctx: dict):
    if key == "purchase_number":         return p.purchase_number or ""
    if key == "order_number":            return p.order_number or ""
    if key == "registry_number":         return p.registry_number or ""
    if key == "status":                  return _STATUS_LABELS.get(p.status, p.status or "")
    if key == "subsidy":                 return ctx["subsidies"].get(p.subsidy_id, "")
    if key == "feo_category":            return ctx["feo_categories"].get(p.feo_category_id, "")
    if key == "item_name":               return p.item_name or ""
    if key == "item_type":               return p.item_type or ""
    if key == "unit":                    return p.unit or ""
    if key == "quantity":                return float(p.planned_quantity) if p.planned_quantity else ""
    if key == "subject":                 return p.subject or ""
    if key == "country_origin":          return p.country_origin or ""
    if key == "planned_unit_price":      return float(p.planned_unit_price) if p.planned_unit_price else ""
    if key == "planned_total_price":     return float(p.planned_total_price) if p.planned_total_price else ""
    if key == "nmck":                    return float(p.nmck or p.planned_total_price or 0) or ""
    if key == "contract_price":          return float(p.contract_price) if p.contract_price else ""
    if key == "economy":                 return float(p.economy) if p.economy else ""
    if key == "price_increase":          return float(p.price_increase) if p.price_increase else ""
    if key == "purchase_method":         return _PURCHASE_METHOD_LABELS.get(p.purchase_method, p.purchase_method or "")
    if key == "purchase_basis":          return _PURCHASE_BASIS_LABELS.get(p.purchase_basis, p.purchase_basis or "")
    if key == "purchase_contract_type":  return _CONTRACT_TYPE_LABELS.get(p.purchase_contract_type, p.purchase_contract_type or "")
    if key == "framework_seq":           return p.framework_seq if p.framework_seq is not None else ""
    if key == "contract_number":         return p.contract_number or ""
    if key == "contract_date":           return str(p.contract_date) if p.contract_date else ""
    if key == "execution_term":          return str(p.execution_term) if p.execution_term else ""
    if key == "execution_term_changed":  return str(p.execution_term_changed) if p.execution_term_changed else ""
    if key == "delivery_date":           return str(p.delivery_date) if p.delivery_date else ""
    if key == "contractor":              return ctx["contractors"].get(p.contractor_id, "")
    if key == "contractor_inn":          return p.contractor.inn if p.contractor else ""
    if key == "responsible_person":      return p.responsible_person or ""
    if key == "acceptance_doc_name":     return p.acceptance_doc_name or ""
    if key == "acceptance_doc_number":   return p.acceptance_doc_number or ""
    if key == "acceptance_doc_date":     return str(p.acceptance_doc_date) if p.acceptance_doc_date else ""
    if key == "acceptance_doc_amount":   return float(p.acceptance_doc_amount) if p.acceptance_doc_amount else ""
    if key == "payment_doc_number":      return p.payment_doc_number or ""
    if key == "payment_doc_date":        return str(p.payment_doc_date) if p.payment_doc_date else ""
    if key == "payment_amount":          return float(p.payment_amount) if p.payment_amount else ""
    if key == "payment_federal":         return float(p.payment_federal) if p.payment_federal else ""
    if key == "delivery_payment_amount": return float(p.delivery_payment_amount) if p.delivery_payment_amount else ""
    if key == "vat_applicable":          return "Да" if p.vat_applicable else ""
    if key == "vat_rate":                return p.vat_rate if p.vat_rate is not None else ""
    if key == "vat_exemption_article":   return p.vat_exemption_article or ""
    return ""


FRAMEWORK_TYPES = {"framework_cumulative", "framework_with_amount"}


async def _assign_framework_seq(p: Purchase, db: AsyncSession, exclude_id: Optional[int] = None) -> None:
    """Auto-assign framework_seq if purchase belongs to a framework contract and seq is not set."""
    if p.purchase_contract_type not in FRAMEWORK_TYPES or not p.contract_id:
        return
    if p.framework_seq is not None:
        return  # already set (manual override)
    q = select(func.coalesce(func.max(Purchase.framework_seq), 0)).where(
        Purchase.contract_id == p.contract_id
    )
    if exclude_id:
        q = q.where(Purchase.id != exclude_id)
    result = await db.execute(q)
    p.framework_seq = (result.scalar() or 0) + 1


async def _check_budget(
    subsidy_id: Optional[int],
    amount: Optional[Decimal],
    exclude_pid: Optional[int],
    db: AsyncSession
):
    """Raises 422 if adding `amount` to subsidy total would exceed its budget."""
    if not subsidy_id or not amount:
        return
    subsidy_r = await db.execute(select(Subsidy).where(Subsidy.id == subsidy_id))
    subsidy = subsidy_r.scalar_one_or_none()
    if not subsidy:
        return
    q = select(func.coalesce(func.sum(Purchase.planned_total_price), 0)).where(
        Purchase.subsidy_id == subsidy_id
    )
    if exclude_pid:
        q = q.where(Purchase.id != exclude_pid)
    total_r = await db.execute(q)
    total = Decimal(str(total_r.scalar() or 0))
    amt = Decimal(str(amount))
    budget = Decimal(str(subsidy.budget))
    if total + amt > budget:
        remaining = budget - total
        raise HTTPException(
            422,
            f"Превышение бюджета субсидии «{subsidy.name}». "
            f"Доступно: {remaining:,.2f} ₽, запрашивается: {amt:,.2f} ₽"
        )

# Fields required for each transition target
TRANSITION_REQUIRED: dict[str, list[str]] = {
    "contracted": ["contract_number", "contract_date"],
    "delivered": ["acceptance_doc_name", "acceptance_doc_date", "acceptance_doc_number", "acceptance_doc_amount"],
    "paid": ["payment_doc_number", "payment_doc_date", "payment_amount"],
}


def _item_to_out(item: PurchaseItem) -> PurchaseItemOut:
    product_name = None
    product_photo_url = None
    product_description = None
    product_description_44fz = None
    if item.product:
        product_name = item.product.name
        product_photo_url = item.product.photo_url
        product_description = item.product.description
        product_description_44fz = item.product.description_44fz
    return PurchaseItemOut(
        id=item.id,
        product_id=item.product_id,
        item_name=item.item_name,
        item_type=item.item_type,
        quantity=item.quantity,
        unit=item.unit,
        unit_price=item.unit_price,
        total_price=item.total_price,
        final_unit_price=item.final_unit_price,
        final_total=item.final_total,
        country_origin=item.country_origin,
        product_name=product_name,
        product_photo_url=product_photo_url,
        product_description=product_description,
        product_description_44fz=product_description_44fz,
    )


def _purchase_to_full(p: Purchase, contractors: dict, subsidies: dict, allocations: list | None = None) -> PurchaseOutFull:
    data = {c.name: getattr(p, c.name) for c in Purchase.__table__.columns}
    items = [_item_to_out(i) for i in (p.items or [])]
    files = [
        PurchaseFileOut(
            id=f.id,
            purchase_id=f.purchase_id,
            filename=f.filename,
            mime_type=f.mime_type,
            size=f.size,
            created_at=str(f.created_at) if f.created_at else None,
        )
        for f in (p.files or [])
    ]
    alloc_out = None
    if allocations is not None:
        alloc_out = [
            SubsidyAllocationOut(
                id=a.id,
                subsidy_id=a.subsidy_id,
                subsidy_name=a.subsidy.name if a.subsidy else subsidies.get(a.subsidy_id),
                amount=a.amount,
            )
            for a in allocations
        ]
    return PurchaseOutFull(
        **data,
        items=items,
        files=files,
        contractor_name=contractors.get(p.contractor_id),
        feo_category_name=p.feo_category.name if p.feo_category else None,
        subsidy_name=subsidies.get(p.subsidy_id),
        event_name=p.event.name if p.event else None,
        subsidy_allocations=alloc_out,
    )


@router.get("/responsible-persons")
async def list_responsible_persons(
    subsidy_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Уникальные ответственные исполнители из закупок (для выпадающего списка)."""
    q = select(Purchase.responsible_person).where(Purchase.responsible_person.isnot(None))
    if subsidy_id:
        q = q.where(Purchase.subsidy_id == subsidy_id)
    q = q.distinct().order_by(Purchase.responsible_person)
    result = await db.execute(q)
    return [row[0] for row in result.fetchall()]


@router.get("/", response_model=List[PurchaseOutFull])
async def list_purchases(
    contract_id: Optional[int] = Query(None),
    feo_category_id: Optional[int] = Query(None),
    subsidy_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    purchase_method: Optional[str] = Query(None),
    purchase_basis: Optional[str] = Query(None),
    limit: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.subsidy import Subsidy
    q = select(Purchase).options(
        selectinload(Purchase.contractor),
        selectinload(Purchase.feo_category),
        selectinload(Purchase.items).selectinload(PurchaseItem.product),
        selectinload(Purchase.files),
        selectinload(Purchase.event),
    )
    org_ids = get_org_filter(current_user)
    if org_ids is not None:
        q = q.join(Subsidy, Purchase.subsidy_id == Subsidy.id).where(Subsidy.org_id.in_(org_ids))
    # Employee: only purchases they participate in
    if current_user.role == 'employee':
        from app.models.purchase_event import PurchaseMember
        member_pids = select(PurchaseMember.purchase_id).where(PurchaseMember.user_id == current_user.id)
        q = q.where(
            (Purchase.assigned_user_id == current_user.id) |
            (Purchase.id.in_(member_pids))
        )
    if contract_id:
        q = q.where(Purchase.contract_id == contract_id)
    if feo_category_id:
        q = q.where(Purchase.feo_category_id == feo_category_id)
    if subsidy_id:
        q = q.where(Purchase.subsidy_id == subsidy_id)
    if status:
        q = q.where(Purchase.status == status)
    if purchase_method:
        q = q.where(Purchase.purchase_method == purchase_method)
    if purchase_basis:
        q = q.where(Purchase.purchase_basis == purchase_basis)
    if search:
        like = f"%{search}%"
        from sqlalchemy import cast, String as SAString
        search_filters = [
            Purchase.item_name.ilike(like),
            Purchase.subject.ilike(like),
            Purchase.registry_number.ilike(like),
            Purchase.contract_number.ilike(like),
            Purchase.order_number.ilike(like),
        ]
        # Search by purchase_number if numeric
        if search.strip().isdigit():
            search_filters.append(Purchase.purchase_number == int(search.strip()))
        q = q.where(or_(*search_filters))
    q = q.order_by(Purchase.id.desc())
    if limit:
        q = q.limit(limit)
    result = await db.execute(q)
    purchases = result.scalars().all()

    contractors_r = await db.execute(select(Contractor))
    contractors = {c.id: c.name for c in contractors_r.scalars().all()}
    subsidies_r = await db.execute(select(Subsidy))
    subsidies = {s.id: s.name for s in subsidies_r.scalars().all()}

    return [_purchase_to_full(p, contractors, subsidies) for p in purchases]


@router.get("/my-tasks")
async def my_tasks(
    include_archive: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Kanban: purchases assigned to current user."""
    q = select(Purchase).options(
        selectinload(Purchase.contractor),
        selectinload(Purchase.feo_category),
    ).where(Purchase.assigned_user_id == current_user.id)
    if not include_archive:
        q = q.where(Purchase.status != "paid")
    q = q.order_by(Purchase.execution_term.asc().nulls_last(), Purchase.id.desc())
    result = await db.execute(q)
    purchases = result.scalars().all()
    return [
        {
            "id": p.id, "subject": p.subject or p.item_name or "",
            "status": p.status, "purchase_number": p.purchase_number,
            "registry_number": p.registry_number,
            "execution_term": str(p.execution_term) if p.execution_term else None,
            "delivery_date": str(p.delivery_date) if p.delivery_date else None,
            "planned_total_price": float(p.planned_total_price or 0),
            "contract_price": float(p.contract_price or 0),
            "contractor_name": p.contractor.name if p.contractor else None,
            "feo_category_name": p.feo_category.name if p.feo_category else None,
            "task_comment": p.task_comment,
            "subsidy_id": p.subsidy_id,
        }
        for p in purchases
    ]


@router.get("/kanban-all")
async def kanban_all(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Kanban: all purchases grouped by assigned user (for managers/admins)."""
    q = select(Purchase).options(
        selectinload(Purchase.contractor),
        selectinload(Purchase.feo_category),
    ).where(Purchase.assigned_user_id.isnot(None))
    org_ids = get_org_filter(current_user)
    if org_ids is not None:
        q = q.join(Subsidy, Purchase.subsidy_id == Subsidy.id).where(Subsidy.org_id.in_(org_ids))
    q = q.order_by(Purchase.assigned_user_id, Purchase.execution_term.asc().nulls_last())
    result = await db.execute(q)
    purchases = result.scalars().all()

    # Group by user
    users_result = await db.execute(select(User))
    users_map = {u.id: u.full_name or u.username for u in users_result.scalars().all()}

    grouped = {}
    for p in purchases:
        uid = p.assigned_user_id
        if uid not in grouped:
            grouped[uid] = {"user_id": uid, "user_name": users_map.get(uid, f"User #{uid}"), "tasks": []}
        grouped[uid]["tasks"].append({
            "id": p.id, "subject": p.subject or p.item_name or "",
            "status": p.status, "purchase_number": p.purchase_number,
            "execution_term": str(p.execution_term) if p.execution_term else None,
            "contractor_name": p.contractor.name if p.contractor else None,
            "planned_total_price": float(p.planned_total_price or 0),
            "task_comment": p.task_comment,
        })
    return list(grouped.values())


@router.get("/{pid}/kp-items")
async def get_purchase_kp_items(pid: int, db: AsyncSession = Depends(get_db)):
    """Items in a purchase with product category info for КП smart sending."""
    from app.models.purchase_item import PurchaseItem
    from app.models.product import Product

    stmt = (
        select(
            PurchaseItem.id,
            PurchaseItem.item_name,
            PurchaseItem.quantity,
            PurchaseItem.unit,
            PurchaseItem.unit_price,
            Product.name.label("product_name"),
            Product.category.label("product_category"),
        )
        .outerjoin(Product, Product.id == PurchaseItem.product_id)
        .where(PurchaseItem.purchase_id == pid)
        .order_by(PurchaseItem.id)
    )
    rows = (await db.execute(stmt)).all()
    return [
        {
            "id": r.id,
            "item_name": r.product_name or r.item_name,
            "quantity": float(r.quantity or 0),
            "unit": r.unit or "шт.",
            "unit_price": float(r.unit_price or 0),
            "category": r.product_category or None,
        }
        for r in rows
    ]


@router.get("/{pid}", response_model=PurchaseOutFull)
async def get_purchase(pid: int, db: AsyncSession = Depends(get_db)):
    result = await db.execute(
        select(Purchase)
        .options(
            selectinload(Purchase.contractor),
            selectinload(Purchase.feo_category),
            selectinload(Purchase.items).selectinload(PurchaseItem.product),
            selectinload(Purchase.files),
            selectinload(Purchase.event),
        )
        .where(Purchase.id == pid)
    )
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    subsidies_r = await db.execute(select(Subsidy))
    subsidies = {s.id: s.name for s in subsidies_r.scalars().all()}
    contractors_r = await db.execute(select(Contractor))
    contractors = {c.id: c.name for c in contractors_r.scalars().all()}
    alloc_r = await db.execute(
        select(PurchaseSubsidyAllocation)
        .options(selectinload(PurchaseSubsidyAllocation.subsidy))
        .where(PurchaseSubsidyAllocation.purchase_id == pid)
    )
    allocations = alloc_r.scalars().all()
    return _purchase_to_full(p, contractors, subsidies, allocations=allocations)


@router.post("/", response_model=PurchaseOut)
async def create_purchase(
    data: PurchaseCreate,
    admin_override: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user)
):
    if admin_override and current_user.role not in ADMIN_ROLES:
        raise HTTPException(403, "Обход бюджетного ограничения доступен только администратору")

    items_data = data.items or []
    # Compute total_nmck from items
    total_nmck = sum((i.total_price or Decimal("0")) for i in items_data) or data.nmck

    if not admin_override and data.purchase_basis != 'service_note':
        await _check_budget(data.subsidy_id, total_nmck or data.planned_total_price, None, db)

    if not data.purchase_number:
        max_result = await db.execute(select(func.coalesce(func.max(Purchase.purchase_number), 0)))
        data.purchase_number = max_result.scalar() + 1

    dump = data.model_dump(exclude={"items", "subsidy_allocations"})
    dump["total_nmck"] = total_nmck
    p = Purchase(**dump)
    db.add(p)
    await db.flush()  # get p.id before commit

    year = date.today().year
    if not p.registry_number:
        p.registry_number = f"РЕЕ-{year}-{p.id:05d}"
    if not p.contract_number:
        p.contract_number = f"{year}/{p.id}"

    await _assign_framework_seq(p, db)

    for item_d in items_data:
        d = item_d.model_dump()
        if not d.get("product_id") and d.get("item_name"):
            from app.models.product import Product as _Prod
            existing = (await db.execute(
                select(_Prod).where(_Prod.name == d["item_name"].strip()).limit(1)
            )).scalar_one_or_none()
            if existing:
                d["product_id"] = existing.id
            else:
                new_prod = _Prod(
                    name=d["item_name"].strip(),
                    product_type=d.get("item_type"),
                    price=d.get("unit_price"),
                    org_id=get_single_org_id(current_user) or current_user.org_id,
                )
                db.add(new_prod)
                await db.flush()
                d["product_id"] = new_prod.id
        item = PurchaseItem(purchase_id=p.id, **d)
        db.add(item)

    # Save subsidy allocations
    if data.subsidy_allocations:
        for alloc in data.subsidy_allocations:
            db.add(PurchaseSubsidyAllocation(
                purchase_id=p.id,
                subsidy_id=alloc.subsidy_id,
                amount=alloc.amount,
            ))

    await db.commit()
    await db.refresh(p)
    return p


@router.put("/{pid}", response_model=PurchaseOut)
async def update_purchase(
    pid: int,
    data: PurchaseCreate,
    admin_override: bool = Query(False),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user)
):
    result = await db.execute(select(Purchase).where(Purchase.id == pid))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    # Employees can save any purchase they have access to (org-level access checked at list level)
    if current_user.role not in MANAGER_ROLES and current_user.role not in ("employee",):
        raise HTTPException(403, "Insufficient permissions")
    if admin_override and current_user.role not in ADMIN_ROLES:
        raise HTTPException(403, "Обход бюджетного ограничения доступен только администратору")

    items_data = data.items or []
    items_sum = sum((i.total_price or Decimal("0")) for i in items_data) or data.nmck

    # НМЦК logic: frozen after "contracted" status
    CONTRACTED_STATUSES = ("contracted", "delivered", "paid")
    is_contracted = p.status in CONTRACTED_STATUSES

    if is_contracted:
        # НМЦК зафиксирована — НЕ пересчитываем, берём из БД
        # Обновляем только цену договора из текущих цен позиций
        pass
    else:
        # До стадии "Договор" — НМЦК = сумма позиций
        p.total_nmck = items_sum
        p.planned_total_price = items_sum or p.planned_total_price

    if not admin_override and data.purchase_basis != 'service_note':
        budget_amount = p.total_nmck if is_contracted else items_sum
        await _check_budget(data.subsidy_id, budget_amount or data.planned_total_price, pid, db)

    # If contract_id or type changed, reset seq so it gets re-assigned
    old_contract_id = p.contract_id
    old_type = p.purchase_contract_type
    for k, v in data.model_dump(exclude={"items", "subsidy_allocations"}, exclude_unset=True).items():
        # Don't overwrite frozen total_nmck
        if is_contracted and k in ("total_nmck", "planned_total_price"):
            continue
        setattr(p, k, v)

    # Contract price: for single purchases = sum of current item prices
    is_single = not p.purchase_contract_type or p.purchase_contract_type == "single"
    if is_single and items_sum:
        p.contract_price = items_sum
    if (p.contract_id != old_contract_id or p.purchase_contract_type != old_type) and data.framework_seq is None:
        p.framework_seq = None  # force re-assignment below
    await _assign_framework_seq(p, db, exclude_id=pid)

    # Replace items (auto-link to catalog if product_id is missing)
    await db.execute(delete(PurchaseItem).where(PurchaseItem.purchase_id == pid))
    for item_d in items_data:
        d = item_d.model_dump()
        if not d.get("product_id") and d.get("item_name"):
            # Try to find existing product by exact name
            from app.models.product import Product as _Prod
            existing = (await db.execute(
                select(_Prod).where(_Prod.name == d["item_name"].strip()).limit(1)
            )).scalar_one_or_none()
            if existing:
                d["product_id"] = existing.id
            else:
                # Auto-create product in catalog
                new_prod = _Prod(
                    name=d["item_name"].strip(),
                    product_type=d.get("item_type"),
                    price=d.get("unit_price"),
                    org_id=get_single_org_id(current_user) or current_user.org_id,
                )
                db.add(new_prod)
                await db.flush()
                d["product_id"] = new_prod.id
        item = PurchaseItem(purchase_id=pid, **d)
        db.add(item)

    # Replace subsidy allocations
    await db.execute(delete(PurchaseSubsidyAllocation).where(PurchaseSubsidyAllocation.purchase_id == pid))
    if data.subsidy_allocations:
        for alloc in data.subsidy_allocations:
            db.add(PurchaseSubsidyAllocation(
                purchase_id=pid,
                subsidy_id=alloc.subsidy_id,
                amount=alloc.amount,
            ))

    # Auto-create/link contract record when contract_number is set
    await ensure_contract_linked(p, db)

    await db.commit()
    await db.refresh(p)
    return p


@router.delete("/bulk")
async def bulk_delete_purchases(
    ids: List[int] = Body(...),
    db: AsyncSession = Depends(get_db),
    _=Depends(require_role(*ADMIN_ROLES)),
):
    deleted, failed = [], []
    for pid in ids:
        try:
            result = await db.execute(select(Purchase).where(Purchase.id == pid))
            p = result.scalar_one_or_none()
            if not p:
                failed.append({"id": pid, "reason": "Не найдено"})
                continue
            await db.delete(p)
            await db.flush()
            deleted.append(pid)
        except Exception as e:
            await db.rollback()
            failed.append({"id": pid, "reason": str(e)[:200]})
    if deleted:
        await db.commit()
    return {"deleted": deleted, "failed": failed}


@router.delete("/{pid}")
async def delete_purchase(pid: int, db: AsyncSession = Depends(get_db), _=Depends(require_role(*ADMIN_ROLES))):
    result = await db.execute(select(Purchase).where(Purchase.id == pid))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Not found")
    await db.delete(p)
    await db.commit()
    return {"ok": True}


@router.get("/by-contract/{contract_id}", response_model=List[PurchaseOutFull])
async def purchases_by_contract(
    contract_id: int,
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    """Все закупки в рамках одного договора, отсортированные по framework_seq."""
    result = await db.execute(
        select(Purchase)
        .where(Purchase.contract_id == contract_id)
        .order_by(
            Purchase.framework_seq.asc().nulls_last(),
            Purchase.id.asc()
        )
    )
    purchases = result.scalars().all()
    out = []
    for p in purchases:
        d = PurchaseOutFull.model_validate(p)
        if p.contractor:
            d.contractor_name = p.contractor.name
        out.append(d)
    return out


@router.post("/{pid}/transition", response_model=PurchaseOutFull)
async def transition_status(
    pid: int,
    target_status: str = Query(..., alias="status"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user)
):
    """
    Forward-only status transition.
    wishes → plan_schedule → confirmed → work_in_progress → contracted → delivered → paid
    """
    if target_status not in STATUS_ORDER:
        raise HTTPException(422, f"Недопустимый статус: {target_status}")

    result = await db.execute(
        select(Purchase)
        .options(
            selectinload(Purchase.contractor),
            selectinload(Purchase.feo_category),
            selectinload(Purchase.items).selectinload(PurchaseItem.product),
            selectinload(Purchase.files),
            selectinload(Purchase.event),
        )
        .where(Purchase.id == pid)
    )
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Закупка не найдена")

    current_idx = STATUS_ORDER.index(p.status) if p.status in STATUS_ORDER else 0
    target_idx = STATUS_ORDER.index(target_status)

    # Role check: only manager+ can transition
    if current_user.role not in MANAGER_ROLES:
        raise HTTPException(403, "Недостаточно прав для смены статуса")

    # Direction check: forward-only for non-admin
    if target_idx <= current_idx:
        if current_user.role not in ADMIN_ROLES:
            raise HTTPException(422, "Откат статуса разрешён только администратору")

    # Approval guard: block contracted transition if approval is pending/rejected
    if target_status == "contracted" and p.approval_status and p.approval_status not in ("approved",):
        raise HTTPException(
            422,
            f"Закупка должна быть согласована перед заключением договора. "
            f"Текущий статус согласования: {p.approval_status}"
        )

    # Catalog guard: warn but don't block (except advance reports)
    # Previously blocked status transition — now just logs a warning
    if target_status == "confirmed" and getattr(p, 'purchase_method', None) != "advance":
        unmatched = [i for i in (p.items or []) if not i.product_id]
        if unmatched:
            import logging
            logging.getLogger(__name__).warning(
                "Purchase %s has %d items not linked to catalog", pid, len(unmatched)
            )

    # Field guards for specific target statuses
    if target_status in TRANSITION_REQUIRED:
        missing = [
            f for f in TRANSITION_REQUIRED[target_status]
            if not getattr(p, f, None)
        ]
        if missing:
            labels = {
                "contract_number": "Номер договора",
                "contract_date": "Дата договора",
                "acceptance_doc_name": "Наименование закрывающего документа",
                "acceptance_doc_date": "Дата закрывающего документа",
                "acceptance_doc_number": "Номер закрывающего документа",
                "acceptance_doc_amount": "Сумма закрывающего документа",
                "payment_doc_number": "Номер платёжного поручения",
                "payment_doc_date": "Дата платёжного поручения",
                "payment_amount": "Сумма платежа",
            }
            missing_labels = [labels.get(f, f) for f in missing]
            raise HTTPException(
                422,
                f"Для перехода в статус «{target_status}» заполните: {', '.join(missing_labels)}"
            )

    old_status = p.status
    p.status = target_status

    # При переходе в contracted — обновить цены в каталоге товаров
    if target_status == "contracted" and p.items:
        sub = await db.get(Subsidy, p.subsidy_id) if p.subsidy_id else None
        contract_org_id = sub.org_id if sub else None
        for item in p.items:
            if not item.product_id:
                continue
            product = await db.get(Product, item.product_id)
            if not product:
                continue
            item_price = item.unit_price
            product.contract_price = item_price
            product.price = item_price
            product.contract_number = p.contract_number
            product.contract_date = p.contract_date
            product.contract_org_id = contract_org_id

    # Auto-create/link contract record when moving to contracted status
    if target_status == "contracted":
        await ensure_contract_linked(p, db)

    await db.commit()

    # Auto-log status change event
    try:
        from app.models.purchase_event import PurchaseEvent
        db.add(PurchaseEvent(
            purchase_id=pid,
            user_id=getattr(current_user, "id", None),
            event_type="status_changed",
            data={"from": old_status, "to": target_status},
        ))
        await db.commit()
    except Exception:
        pass

    # Notify purchase members + linked task assignees about status change
    try:
        from app.notifications import notify_purchase_status_changed
        from app.models.purchase_event import PurchaseMember
        from app.models.task import Task, TaskAssignee

        notify_user_ids: set[int] = set()
        notify_users = []

        # 1. Purchase members
        members_r = await db.execute(
            select(PurchaseMember).where(PurchaseMember.purchase_id == pid)
        )
        for m in members_r.scalars().all():
            if m.user_id != current_user.id:
                notify_user_ids.add(m.user_id)

        # 2. Assigned user of purchase
        if p.assigned_user_id and p.assigned_user_id != current_user.id:
            notify_user_ids.add(p.assigned_user_id)

        # 3. All assignees of tasks linked to this purchase
        linked_tasks_r = await db.execute(
            select(Task.id).where(Task.purchase_id == pid)
        )
        linked_task_ids = [r[0] for r in linked_tasks_r.all()]
        if linked_task_ids:
            ta_r = await db.execute(
                select(TaskAssignee.user_id).where(
                    TaskAssignee.task_id.in_(linked_task_ids)
                )
            )
            for r in ta_r.all():
                if r[0] != current_user.id:
                    notify_user_ids.add(r[0])

        # Load user objects
        for uid in notify_user_ids:
            u = await db.get(User, uid)
            if u:
                notify_users.append(u)

        if notify_users:
            await notify_purchase_status_changed(
                p, current_user.full_name or current_user.username,
                target_status, notify_users
            )
    except Exception:
        pass

    # Re-fetch with eager loads after commit
    result2 = await db.execute(
        select(Purchase)
        .options(
            selectinload(Purchase.contractor),
            selectinload(Purchase.feo_category),
            selectinload(Purchase.items).selectinload(PurchaseItem.product),
            selectinload(Purchase.files),
            selectinload(Purchase.event),
        )
        .where(Purchase.id == pid)
    )
    p = result2.scalar_one()
    subsidies_r = await db.execute(select(Subsidy))
    subsidies = {s.id: s.name for s in subsidies_r.scalars().all()}
    contractors_r = await db.execute(select(Contractor))
    contractors = {c.id: c.name for c in contractors_r.scalars().all()}
    return _purchase_to_full(p, contractors, subsidies)


@router.post("/{pid}/convert-to-order", response_model=PurchaseOutFull)
async def convert_service_note_to_order(
    pid: int,
    db: AsyncSession = Depends(get_db),
    current_user=Depends(require_role(*MANAGER_ROLES)),
):
    """Конвертировать служебную записку на выдачу в закупку (меняет purchase_basis на plan_schedule)."""
    result = await db.execute(
        select(Purchase)
        .options(
            selectinload(Purchase.contractor),
            selectinload(Purchase.feo_category),
            selectinload(Purchase.items).selectinload(PurchaseItem.product),
            selectinload(Purchase.files),
            selectinload(Purchase.event),
        )
        .where(Purchase.id == pid)
    )
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Закупка не найдена")
    if p.purchase_basis != 'service_note':
        raise HTTPException(422, "Конвертация доступна только для служебных записок")
    p.purchase_basis = 'plan_schedule'
    await db.commit()
    subsidies_r = await db.execute(select(Subsidy))
    subsidies = {s.id: s.name for s in subsidies_r.scalars().all()}
    contractors_r = await db.execute(select(Contractor))
    contractors = {c.id: c.name for c in contractors_r.scalars().all()}
    return _purchase_to_full(p, contractors, subsidies)


@router.patch("/{pid}/assign")
async def assign_purchase(
    pid: int,
    user_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Assign a purchase to a user (or unassign if user_id is None)."""
    result = await db.execute(select(Purchase).where(Purchase.id == pid))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Закупка не найдена")
    if user_id is not None:
        user_check = await db.execute(select(User).where(User.id == user_id))
        if not user_check.scalar_one_or_none():
            raise HTTPException(404, "Пользователь не найден")
    p.assigned_user_id = user_id
    await db.commit()
    return {"ok": True, "assigned_user_id": user_id}


@router.get("/{pid}/tasks")
async def list_purchase_tasks(
    pid: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Get all tasks linked to a purchase."""
    from app.models.task import Task
    from app.routers.tasks import _enrich_tasks
    from app.schemas.schemas import TaskOut
    result = await db.execute(
        select(Task).where(Task.purchase_id == pid).order_by(Task.created_at.desc())
    )
    tasks = result.scalars().all()
    return await _enrich_tasks(list(tasks), db, current_user_id=current_user.id)


@router.patch("/{pid}/kanban-status")
async def kanban_status_change(
    pid: int,
    target_status: str = Query(..., alias="status"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Quick status change from kanban board (same rules as transition)."""
    # Map common aliases
    STATUS_ALIASES = {"in_progress": "work_in_progress", "planned": "plan_schedule"}
    target_status = STATUS_ALIASES.get(target_status, target_status)
    if target_status not in STATUS_ORDER:
        STATUS_LABELS_RU = dict(zip(STATUS_ORDER, ["Желания", "План-график", "Подтверждено", "Ведётся работа", "Договор", "Поставлено", "Оплачено"]))
        allowed = ", ".join(f"{k} ({v})" for k, v in STATUS_LABELS_RU.items())
        raise HTTPException(422, f"Недопустимый статус: «{target_status}». Допустимые: {allowed}")
    result = await db.execute(select(Purchase).where(Purchase.id == pid))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Закупка не найдена")

    # Only assigned user, managers or admins can change status
    is_assigned = p.assigned_user_id == current_user.id
    is_admin_or_manager = current_user.role in MANAGER_ROLES
    if not is_assigned and not is_admin_or_manager:
        raise HTTPException(403, "Недостаточно прав")

    old_status = p.status
    p.status = target_status
    await db.commit()

    try:
        from app.models.purchase_event import PurchaseEvent
        db.add(PurchaseEvent(
            purchase_id=pid,
            user_id=current_user.id,
            event_type="status_changed",
            data={"from": old_status, "to": target_status, "via": "kanban"},
        ))
        await db.commit()
    except Exception:
        pass

    return {"ok": True, "status": target_status}


@router.patch("/{pid}/substatus")
async def update_substatus(
    pid: int,
    substatus: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Set substatus for work_in_progress purchases."""
    if substatus and substatus not in VALID_SUBSTATUSES:
        raise HTTPException(422, f"Недопустимый подстатус: {substatus}. Допустимые: {', '.join(VALID_SUBSTATUSES)}")
    result = await db.execute(select(Purchase).where(Purchase.id == pid))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Закупка не найдена")

    is_assigned = p.assigned_user_id == current_user.id
    is_admin_or_manager = current_user.role in MANAGER_ROLES
    if not is_assigned and not is_admin_or_manager:
        raise HTTPException(403, "Недостаточно прав")

    # Auto-set status to work_in_progress if setting a substatus
    if substatus and p.status != "work_in_progress":
        p.status = "work_in_progress"
    p.substatus = substatus
    await db.commit()

    try:
        from app.models.purchase_event import PurchaseEvent
        db.add(PurchaseEvent(
            purchase_id=pid,
            user_id=current_user.id,
            event_type="substatus_changed",
            data={"substatus": substatus},
        ))
        await db.commit()
    except Exception:
        pass

    return {"ok": True, "substatus": substatus, "status": p.status}


@router.patch("/{pid}/comment")
async def update_task_comment(
    pid: int,
    comment: str = Query(""),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Update task comment for kanban card."""
    result = await db.execute(select(Purchase).where(Purchase.id == pid))
    p = result.scalar_one_or_none()
    if not p:
        raise HTTPException(404, "Закупка не найдена")
    p.task_comment = comment
    await db.commit()
    return {"ok": True}


@router.get("/users-list")
async def users_list(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """List users available for assignment."""
    q = select(User.id, User.full_name, User.username, User.role)
    org_ids = get_org_filter(current_user)
    if org_ids is not None:
        q = q.where(User.org_id.in_(org_ids))
    result = await db.execute(q.order_by(User.full_name))
    return [
        {"id": r.id, "name": r.full_name or r.username, "role": r.role}
        for r in result
    ]


@router.get("/export/columns")
async def get_export_columns(_=Depends(get_current_user)):
    """Return available export column definitions."""
    return [
        {"key": k, "label": v["label"], "group": v["group"]}
        for k, v in ALL_EXPORT_COLUMNS.items()
    ]


@router.get("/export/excel")
async def export_purchases_to_excel(
    subsidy_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    columns: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    if Workbook is None:
        raise HTTPException(500, "openpyxl не установлен")

    col_keys = (
        [k.strip() for k in columns.split(",") if k.strip() in ALL_EXPORT_COLUMNS]
        if columns else DEFAULT_EXPORT_COLUMNS
    )
    if not col_keys:
        col_keys = DEFAULT_EXPORT_COLUMNS

    q = select(Purchase).order_by(Purchase.id.desc())
    if subsidy_id:
        q = q.where(Purchase.subsidy_id == subsidy_id)
    if status:
        q = q.where(Purchase.status == status)
    purchases = (await db.execute(q)).scalars().all()

    contractors = {c.id: c.name for c in (await db.execute(select(Contractor))).scalars().all()}
    subsidies_map = {s.id: s.name for s in (await db.execute(select(Subsidy))).scalars().all()}
    feo_map = {f.id: f.name for f in (await db.execute(select(FeoCategory))).scalars().all()}
    ctx = {"contractors": contractors, "subsidies": subsidies_map, "feo_categories": feo_map}

    wb = Workbook()
    ws = wb.active
    ws.title = "Закупки"

    col_headers = [ALL_EXPORT_COLUMNS[k]["label"] for k in col_keys]
    ws.append(col_headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    empty_counts = {k: 0 for k in col_keys}
    for p in purchases:
        row = []
        for k in col_keys:
            val = _get_cell_value(k, p, ctx)
            row.append(val)
            if val == "" or val is None:
                empty_counts[k] += 1
        ws.append(row)

    for i, key in enumerate(col_keys, 1):
        col_letter = ws.cell(1, i).column_letter
        ws.column_dimensions[col_letter].width = max(len(col_headers[i - 1]) + 2, 12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"purchases_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    missing = []
    if len(purchases) >= 5:
        for k in col_keys:
            if empty_counts[k] / len(purchases) > 0.8:
                missing.append(ALL_EXPORT_COLUMNS[k]["label"])

    resp_headers = {
        "Content-Disposition": f"attachment; filename={filename}",
        "Access-Control-Expose-Headers": "X-Missing-Columns",
    }
    if missing:
        resp_headers["X-Missing-Columns"] = ",".join(missing[:5])

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=resp_headers,
    )


@router.get("/import/template")
async def download_import_template():
    """Скачать шаблон Excel для импорта закупок."""
    if Workbook is None:
        raise HTTPException(500, "openpyxl не установлен")
    wb = Workbook()
    ws = wb.active
    ws.title = "Закупки"

    headers = [
        "Наименование", "Субсидия", "Категория ФЭО", "Контрагент", "ИНН контрагента",
        "НМЦК", "Способ закупки", "Реестровый №", "№ договора", "Дата договора",
        "Цена договора", "Срок исполнения", "ПП №", "ПП дата", "Оплачено",
        "Статус", "Год",
    ]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample row
    ws.append([
        "Закупка компьютеров", "ФАДМ_2026", "Техническое оснащение", "ООО Поставщик", "1234567890",
        "500000", "ЕИ", "2026/001", "Д-001", "15.03.2026",
        "490000", "30.06.2026", "", "", "",
        "confirmed", "2026",
    ])

    # Column widths
    col_widths = [35, 20, 30, 25, 15, 12, 15, 14, 14, 14, 14, 14, 12, 12, 12, 14, 6]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    ws.freeze_panes = "A2"

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=import_template.xlsx"}
    )


@router.post("/import")
async def import_purchases_from_excel(
    file: UploadFile = File(...),
    subsidy_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Импорт закупок из Excel. Возвращает {created, skipped, errors}."""
    if load_workbook is None:
        raise HTTPException(500, "openpyxl не установлен")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Поддерживаются только файлы .xlsx и .xls")

    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Файл пустой или содержит только заголовки")

    # Parse headers (case-insensitive)
    raw_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    COLUMN_MAP = {
        "наименование": "item_name", "предмет закупки": "item_name",
        "субсидия": "subsidy_name",
        "категория фэо": "feo_category_name",
        "контрагент": "contractor_name",
        "инн контрагента": "contractor_inn", "инн": "contractor_inn",
        "нмцк": "nmck", "сумма": "nmck", "цена": "nmck",
        "способ закупки": "purchase_method", "способ": "purchase_method",
        "реестровый №": "registry_number", "реестровый номер": "registry_number", "реестр. №": "registry_number",
        "№ договора": "contract_number", "номер договора": "contract_number",
        "дата договора": "contract_date",
        "цена договора": "contract_price",
        "срок исполнения": "execution_term",
        "пп №": "payment_doc_number", "пп номер": "payment_doc_number",
        "пп дата": "payment_doc_date",
        "оплачено": "payment_amount",
        "статус": "status",
        "год": "year",
        "№ п/п": "purchase_number",
    }
    col_idx: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        field = COLUMN_MAP.get(h)
        if field and field not in col_idx:
            col_idx[field] = i

    # Load lookup tables
    from app.models.feo_category import FeoCategory
    subs_rows = (await db.execute(select(Subsidy))).scalars().all()
    subs_by_name = {s.name.lower().strip(): s.id for s in subs_rows}

    contractors_rows = (await db.execute(select(Contractor))).scalars().all()
    cont_by_name = {c.name.lower().strip(): c.id for c in contractors_rows}
    cont_by_inn  = {c.inn.strip(): c.id for c in contractors_rows if c.inn}

    feo_rows = (await db.execute(select(FeoCategory))).scalars().all()
    feo_by_name = {f.name.lower().strip(): f.id for f in feo_rows}

    STATUS_MAP = {
        "wishes": "wishes", "желания": "wishes", "planned": "wishes", "планируется": "wishes", "план": "wishes",
        "plan_schedule": "plan_schedule", "план-график": "plan_schedule",
        "confirmed": "confirmed", "подтверждено": "confirmed",
        "work_in_progress": "work_in_progress", "в работе": "work_in_progress", "in_progress": "work_in_progress",
        "contracted": "contracted", "законтрактовано": "contracted",
        "delivered": "delivered", "исполнено": "delivered",
        "paid": "paid", "оплачено": "paid",
    }
    METHOD_MAP = {
        "еи": "single", "единственный исполнитель": "single", "single": "single",
        "кп": "competitive", "конкурсная процедура": "competitive", "competitive": "competitive",
    }

    def cell(row, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    def to_dec(v):
        if v is None:
            return None
        try:
            return Decimal(str(v).replace(" ", "").replace(",", "."))
        except Exception:
            return None

    def to_date_val(v):
        if v is None:
            return None
        if hasattr(v, "date"):
            return v.date()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(str(v).strip(), fmt).date()
            except Exception:
                pass
        return None

    created = 0
    skipped = 0
    errors: list[dict] = []

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            item_name = cell(row, "item_name")
            if not item_name:
                skipped += 1
                continue

            # Resolve subsidy
            sid = subsidy_id
            if not sid:
                sub_name = cell(row, "subsidy_name")
                if sub_name:
                    sid = subs_by_name.get(sub_name.lower().strip())
            if not sid:
                errors.append({"row": row_num, "name": item_name, "message": "Субсидия не найдена"})
                continue

            # Resolve contractor
            cid = None
            c_name = cell(row, "contractor_name")
            c_inn  = cell(row, "contractor_inn")
            if c_inn:
                cid = cont_by_inn.get(c_inn.strip())
            if not cid and c_name:
                cid = cont_by_name.get(c_name.lower().strip())

            # Resolve FEO
            feo_id = None
            feo_name = cell(row, "feo_category_name")
            if feo_name:
                feo_id = feo_by_name.get(feo_name.lower().strip())

            # Status
            status_raw = (cell(row, "status") or "wishes").lower().strip()
            status = STATUS_MAP.get(status_raw, "wishes")

            # Method
            method_raw = (cell(row, "purchase_method") or "").lower().strip()
            method = METHOD_MAP.get(method_raw)

            nmck = to_dec(cell(row, "nmck"))
            p = Purchase(
                subsidy_id=sid,
                feo_category_id=feo_id,
                contractor_id=cid,
                item_name=item_name,
                status=status,
                nmck=nmck,
                total_nmck=nmck,
                planned_total_price=nmck,
                contract_price=to_dec(cell(row, "contract_price")),
                payment_amount=to_dec(cell(row, "payment_amount")),
                purchase_method=method,
                registry_number=cell(row, "registry_number"),
                contract_number=cell(row, "contract_number"),
                contract_date=to_date_val(cell(row, "contract_date")),
                execution_term=to_date_val(cell(row, "execution_term")),
                payment_doc_number=cell(row, "payment_doc_number"),
                payment_doc_date=to_date_val(cell(row, "payment_doc_date")),
            )
            db.add(p)
            created += 1
        except Exception as e:
            errors.append({"row": row_num, "name": cell(row, "item_name") or "?", "message": str(e)})

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors}


# ---------------------------------------------------------------------------
# Purchase items import from Excel
# ---------------------------------------------------------------------------

@router.get("/items/import/template")
async def items_import_template(_=Depends(require_role(*MANAGER_ROLES))):
    """Download xlsx template for bulk purchase items import."""
    if not Workbook:
        raise HTTPException(500, "openpyxl не установлен")

    wb = Workbook()
    ws = wb.active
    ws.title = "Позиции"

    headers = ["Наименование", "Описание / ТЗ", "Тип (товар/услуга/работа)", "Количество", "Ед. изм.", "Цена за единицу"]
    required = {"Наименование"}

    header_fill = PatternFill("solid", fgColor="1E40AF")
    req_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = req_fill if h in required else header_fill
        cell.alignment = Alignment(horizontal="center")

    example = ["Ноутбук Lenovo ThinkPad", "Технические характеристики...", "товар", "5", "шт", "85000"]
    for ci, val in enumerate(example, 1):
        ws.cell(row=2, column=ci, value=val)

    col_widths = [45, 50, 25, 15, 15, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=items_template.xlsx"},
    )


async def _upsert_product_to_catalog(db, item_name: str, item_type: str, unit_price, description: str = "") -> int:
    """Find or create a product in the global catalog. Returns product.id."""
    norm = item_name.strip().lower()
    existing = (await db.execute(
        select(Product).where(func.lower(Product.name) == norm)
    )).scalar_one_or_none()
    if existing:
        return existing.id
    p = Product(
        name=item_name.strip(),
        description=description or "",
        product_type=item_type or "товар",
        price=Decimal(str(unit_price)) if unit_price else Decimal("0"),
        is_active=True,
    )
    db.add(p)
    await db.flush()
    return p.id


@router.post("/{pid}/items/import")
async def import_items_excel(
    pid: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Bulk import items into a purchase from Excel."""
    fname = (file.filename or '').lower()
    if not fname.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Поддерживаются только файлы .xlsx / .xls")

    purchase = await db.get(Purchase, pid)
    if not purchase:
        raise HTTPException(404, "Закупка не найдена")

    content = await file.read()
    try:
        if fname.endswith('.xls'):
            try:
                import xlrd as _xlrd_mod
            except ImportError:
                raise HTTPException(500, "xlrd не установлен")
            wb_xls = _xlrd_mod.open_workbook(file_contents=content)
            ws_xls = wb_xls.sheet_by_index(0)
            all_rows = [tuple(ws_xls.row_values(i)) for i in range(ws_xls.nrows)]
            header_row = all_rows[0] if all_rows else None
            data_iter = all_rows[1:] if len(all_rows) > 1 else []
        else:
            if not load_workbook:
                raise HTTPException(500, "openpyxl не установлен")
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            all_rows_gen = list(ws.iter_rows(values_only=True))
            header_row = all_rows_gen[0] if all_rows_gen else None
            data_iter = all_rows_gen[1:] if len(all_rows_gen) > 1 else []
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Не удалось прочитать файл: {e}")
    if not header_row:
        raise HTTPException(400, "Файл пустой")

    def _norm(v) -> str:
        return str(v).strip().lower() if v else ''

    # Fuzzy column mapping
    col: dict[str, int] = {}
    for i, h in enumerate(header_row):
        h_str = _norm(h)
        if any(x in h_str for x in ('наименован', 'назван', 'name', 'товар', 'предмет')):
            col.setdefault('item_name', i)
        elif any(x in h_str for x in ('описан', 'description', 'тз', 'техническ', 'specification', 'характерист')):
            col.setdefault('description', i)
        elif any(x in h_str for x in ('тип', 'type', 'вид')):
            col.setdefault('item_type', i)
        elif any(x in h_str for x in ('кол', 'количеств', 'qty', 'quantity')):
            col.setdefault('quantity', i)
        elif any(x in h_str for x in ('ед.', 'единиц', 'unit', 'изм')):
            col.setdefault('unit', i)
        elif any(x in h_str for x in ('цена', 'price', 'стоимость', 'за единиц')):
            col.setdefault('unit_price', i)

    if 'item_name' not in col:
        raise HTTPException(400, "Не найдена колонка с наименованием.")

    def _cell(row, field):
        idx = col.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.lower() in ('none', 'null', '-', '—'):
            return None
        return s

    def _to_dec(v):
        if v is None:
            return None
        try:
            return Decimal(str(v).replace(',', '.').replace(' ', ''))
        except Exception:
            return None

    # Load products for auto-matching by name
    org_id = get_single_org_id(current_user)
    prod_q = select(Product)
    if org_id:
        prod_q = prod_q.where((Product.org_id == org_id) | (Product.org_id.is_(None)))
    prod_result = await db.execute(prod_q)
    products = prod_result.scalars().all()
    # Build name lookup (lowercase → product)
    product_by_name: dict[str, Product] = {}
    for p in products:
        if p.name:
            product_by_name[p.name.lower().strip()] = p

    TYPE_MAP = {
        'товар': 'товар', 'товары': 'товар', 'product': 'товар', 'goods': 'товар',
        'услуга': 'услуга', 'услуги': 'услуга', 'service': 'услуга',
        'работа': 'работа', 'работы': 'работа', 'work': 'работа',
    }

    added = 0
    matched_catalog = 0
    new_in_catalog = 0
    errors_list = []

    for row in data_iter:
        item_name = _cell(row, 'item_name')
        if not item_name:
            continue

        description = _cell(row, 'description')
        item_type_raw = (_cell(row, 'item_type') or 'товар').lower().strip()
        item_type = TYPE_MAP.get(item_type_raw, 'товар')
        quantity = _to_dec(_cell(row, 'quantity')) or Decimal('1')
        unit = _cell(row, 'unit') or 'шт'
        unit_price = _to_dec(_cell(row, 'unit_price'))
        total_price = (quantity * unit_price) if unit_price else None

        # Auto-match or create in catalog
        matched_product = product_by_name.get(item_name.lower().strip())
        if matched_product:
            product_id = matched_product.id
            matched_catalog += 1
            if not unit_price and matched_product.price:
                unit_price = matched_product.price
                total_price = quantity * unit_price
        else:
            product_id = await _upsert_product_to_catalog(db, item_name, item_type, unit_price, description or "")
            product_by_name[item_name.lower().strip()] = type('_P', (), {'id': product_id, 'name': item_name, 'price': unit_price})()
            new_in_catalog += 1

        item = PurchaseItem(
            purchase_id=pid,
            product_id=product_id,
            item_name=item_name,
            item_type=item_type,
            quantity=quantity,
            unit=unit,
            unit_price=unit_price,
            total_price=total_price,
        )
        db.add(item)
        added += 1

    await db.commit()
    return {"added": added, "matched_catalog": matched_catalog, "new_in_catalog": new_in_catalog, "errors": errors_list}


# ---------------------------------------------------------------------------
# Excel import with column mapping (preview + mapped import)
# ---------------------------------------------------------------------------

@router.post("/items/import-preview")
async def import_items_preview(
    file: UploadFile = File(...),
    current_user: User = Depends(get_current_user),
):
    """Read Excel file and return sheets, headers, and sample rows for column mapping."""
    fname = (file.filename or '').lower()
    if not fname.endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Поддерживаются только файлы .xlsx / .xls")

    content = await file.read()

    _NAME_HINTS = ('наименован', 'назван', 'товар', 'предмет', 'name', 'title', 'услуг', 'работ')
    _ALL_HINTS = _NAME_HINTS + ('цена', 'описан', 'кол', 'тип', 'price', 'стоимост', 'ед.', 'единиц', 'катег', 'сумм', 'количеств', 'ед. изм')

    def _detect_hdr(rows):
        best_score, best_idx = 0, 0
        for ri, row in enumerate(rows):
            norm = [str(h).strip().lower() if h is not None else "" for h in row]
            score = sum(1 for h in norm if h and any(x in h for x in _ALL_HINTS))
            if score > best_score:
                best_score = score; best_idx = ri
        return best_idx

    try:
        if fname.endswith('.xls'):
            try:
                import xlrd as _xlrd_mod
            except ImportError:
                raise HTTPException(500, "xlrd не установлен")
            wb_xls = _xlrd_mod.open_workbook(file_contents=content)
            sheets = []
            for sheet_name in wb_xls.sheet_names():
                ws_xls = wb_xls.sheet_by_name(sheet_name)
                all_rows = [list(ws_xls.row_values(i)) for i in range(ws_xls.nrows)]
                if not all_rows:
                    continue
                hdr_idx = _detect_hdr(all_rows)
                hdr_rows = all_rows[hdr_idx:]
                if not hdr_rows:
                    continue
                headers = [str(c).strip() if c else f"Столбец {j+1}" for j, c in enumerate(hdr_rows[0])]
                sample = [[str(c).strip() if c is not None else "" for c in row] for row in hdr_rows[1:min(6, len(hdr_rows))]]
                sheets.append({"name": sheet_name, "headers": headers, "sample": sample,
                               "total_rows": ws_xls.nrows - hdr_idx - 1, "header_row_offset": hdr_idx})
        else:
            if not load_workbook:
                raise HTTPException(500, "openpyxl не установлен")
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            sheets = []
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                all_rows = list(ws.iter_rows(values_only=True))
                if not all_rows:
                    continue
                hdr_idx = _detect_hdr(all_rows)
                hdr_rows = all_rows[hdr_idx:]
                if not hdr_rows:
                    continue
                headers = [str(c).strip() if c else f"Столбец {j+1}" for j, c in enumerate(hdr_rows[0])]
                sample = [[str(c).strip() if c is not None else "" for c in row] for row in hdr_rows[1:min(6, len(hdr_rows))]]
                sheets.append({"name": sheet_name, "headers": headers, "sample": sample,
                               "total_rows": len(all_rows) - hdr_idx - 1, "header_row_offset": hdr_idx})
            wb.close()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Не удалось прочитать файл: {e}")

    if not sheets:
        raise HTTPException(400, "Файл не содержит листов с данными")

    return {"sheets": sheets}


@router.post("/{pid}/items/import-mapped")
async def import_items_mapped(
    pid: int,
    file: UploadFile = File(...),
    sheet_name: str = Query(""),
    col_item_name: int = Query(-1, description="Индекс столбца Наименование (0-based)"),
    col_description: int = Query(-1, description="Индекс столбца Описание"),
    col_quantity: int = Query(-1, description="Индекс столбца Количество"),
    col_unit_price: int = Query(-1, description="Индекс столбца Цена"),
    col_total_price: int = Query(-1, description="Индекс столбца Сумма"),
    col_vat: int = Query(-1, description="Индекс столбца НДС"),
    col_unit: int = Query(-1, description="Индекс столбца Ед. изм."),
    header_row_offset: int = Query(0, description="Сколько строк пропустить до заголовка (авто-определено при preview)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Import items using user-specified column mapping."""
    if col_item_name < 0:
        raise HTTPException(400, "Не указан столбец Наименование")

    purchase = await db.get(Purchase, pid)
    if not purchase:
        raise HTTPException(404, "Закупка не найдена")

    fname = (file.filename or '').lower()
    content = await file.read()

    try:
        if fname.endswith('.xls'):
            try:
                import xlrd as _xlrd_mod
            except ImportError:
                raise HTTPException(500, "xlrd не установлен")
            wb_xls = _xlrd_mod.open_workbook(file_contents=content)
            ws_xls = wb_xls.sheet_by_name(sheet_name) if sheet_name else wb_xls.sheet_by_index(0)
            all_rows = [tuple(ws_xls.row_values(i)) for i in range(ws_xls.nrows)]
            skip = header_row_offset + 1  # skip empty rows + header row itself
            data_iter = all_rows[skip:] if len(all_rows) > skip else []
        else:
            if not load_workbook:
                raise HTTPException(500, "openpyxl не установлен")
            wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
            ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            all_rows_gen = list(ws.iter_rows(values_only=True))
            skip = header_row_offset + 1
            data_iter = all_rows_gen[skip:] if len(all_rows_gen) > skip else []
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(400, f"Не удалось прочитать файл: {e}")

    def _cell(row, idx):
        if idx < 0 or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.lower() in ('none', 'null', '-', '—', '0'):
            return None
        return s

    def _to_dec(v):
        if v is None:
            return None
        try:
            return Decimal(str(v).replace(',', '.').replace(' ', '').replace('\xa0', ''))
        except Exception:
            return None

    # Load products for auto-matching
    org_id = get_single_org_id(current_user)
    prod_q = select(Product)
    if org_id:
        prod_q = prod_q.where((Product.org_id == org_id) | (Product.org_id.is_(None)))
    prod_result = await db.execute(prod_q)
    products = prod_result.scalars().all()
    product_by_name: dict[str, Product] = {}
    for p in products:
        if p.name:
            product_by_name[p.name.lower().strip()] = p

    added = 0
    matched_catalog = 0
    new_in_catalog = 0
    errors_list = []

    for row in data_iter:
        item_name = _cell(row, col_item_name)
        if not item_name:
            continue

        description = _cell(row, col_description) if col_description >= 0 else None
        quantity = _to_dec(_cell(row, col_quantity)) if col_quantity >= 0 else Decimal('1')
        if not quantity:
            quantity = Decimal('1')
        unit_price = _to_dec(_cell(row, col_unit_price)) if col_unit_price >= 0 else None
        total_price = _to_dec(_cell(row, col_total_price)) if col_total_price >= 0 else None
        unit = (_cell(row, col_unit) if col_unit >= 0 else None) or 'шт'

        # Calculate missing values
        if not total_price and unit_price:
            total_price = quantity * unit_price
        elif not unit_price and total_price and quantity:
            unit_price = total_price / quantity

        # VAT info → append to description
        vat_str = _cell(row, col_vat) if col_vat >= 0 else None
        if vat_str and description:
            description = f"{description} (НДС: {vat_str})"
        elif vat_str:
            description = f"НДС: {vat_str}"

        # Auto-match or create in catalog
        matched_product = product_by_name.get(item_name.lower().strip())
        if matched_product:
            product_id = matched_product.id
            matched_catalog += 1
            if not unit_price and matched_product.price:
                unit_price = matched_product.price
                total_price = quantity * unit_price
        else:
            product_id = await _upsert_product_to_catalog(db, item_name, 'товар', unit_price, description or "")
            product_by_name[item_name.lower().strip()] = type('_P', (), {'id': product_id, 'name': item_name, 'price': unit_price})()
            new_in_catalog += 1

        item = PurchaseItem(
            purchase_id=pid,
            product_id=product_id,
            item_name=item_name,
            item_type='товар',
            quantity=quantity,
            unit=unit,
            unit_price=unit_price,
            total_price=total_price,
        )
        db.add(item)
        added += 1

    await db.commit()
    return {"added": added, "matched_catalog": matched_catalog, "new_in_catalog": new_in_catalog, "errors": errors_list}


@router.post("/{pid}/items/import-smart")
async def import_items_smart(
    pid: int,
    file: UploadFile = File(...),
    confirm: bool = Query(default=False),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Smart import: extract items table from PDF / DOCX / XLSX."""
    purchase = await db.get(Purchase, pid)
    if not purchase:
        raise HTTPException(404, "Закупка не найдена")
    # Employees can import to any purchase they have access to
    if current_user.role not in MANAGER_ROLES and current_user.role not in ("employee",):
        raise HTTPException(403, "Insufficient permissions")

    content = await file.read()
    filename = (file.filename or "").lower()

    # --- Extract raw tables from file ---
    raw_tables: list[list[list[str]]] = []
    file_type = "unknown"

    if filename.endswith((".xlsx", ".xls")):
        file_type = "excel"
        if not load_workbook:
            raise HTTPException(500, "openpyxl не установлен")
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        rows = [[str(c) if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
        if rows:
            raw_tables.append(rows)
    elif filename.endswith(".pdf"):
        file_type = "pdf"
        if not _pdfplumber:
            raise HTTPException(500, "pdfplumber не установлен")
        with _pdfplumber.open(BytesIO(content)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables() or []
                for tbl in tables:
                    if tbl:
                        raw_tables.append([[str(c) if c is not None else "" for c in row] for row in tbl])
    elif filename.endswith((".docx", ".doc")):
        file_type = "docx"
        if not _DocxDocument:
            raise HTTPException(500, "python-docx не установлен")
        doc = _DocxDocument(BytesIO(content))
        for table in doc.tables:
            rows = [[cell.text.strip() for cell in row.cells] for row in table.rows]
            if rows:
                raw_tables.append(rows)
    else:
        raise HTTPException(400, "Поддерживаются файлы: PDF, DOCX, XLSX/XLS")

    if not raw_tables:
        raise HTTPException(400, "Таблицы в документе не найдены")

    def _detect_columns(header_row: list[str]) -> dict:
        col: dict = {}
        for i, h in enumerate(header_row):
            h_s = h.strip().lower()
            if any(x in h_s for x in ("наименован", "назван", "name", "товар", "предмет", "описан", "услуг")):
                col.setdefault("item_name", i)
            elif any(x in h_s for x in ("тип", "type", "вид")):
                col.setdefault("item_type", i)
            elif any(x in h_s for x in ("кол", "количеств", "qty", "quantity")):
                col.setdefault("quantity", i)
            elif any(x in h_s for x in ("ед.", "единиц", "unit", "изм")):
                col.setdefault("unit", i)
            elif any(x in h_s for x in ("цена ед", "цена за", "стоимость ед", "price")):
                col.setdefault("unit_price", i)
            elif any(x in h_s for x in ("сумма", "итог", "total", "amount", "всего", "стоимость")):
                col.setdefault("total_price", i)
        return col

    # Pick best table (most matched columns)
    best_table: list[list[str]] = []
    best_col: dict = {}
    best_header_row = 0
    for table in raw_tables:
        for r_idx, row in enumerate(table[:6]):
            col = _detect_columns(row)
            if "item_name" in col and len(col) > len(best_col):
                best_col = col
                best_table = table
                best_header_row = r_idx

    if not best_table or "item_name" not in best_col:
        raise HTTPException(400, "Не удалось найти таблицу с позициями. Убедитесь что документ содержит колонку «Наименование».")

    TYPE_MAP = {
        "товар": "товар", "товары": "товар", "product": "товар",
        "услуга": "услуга", "услуги": "услуга", "service": "услуга",
        "работа": "работа", "работы": "работа",
    }

    def _to_dec(v: str):
        if not v:
            return None
        try:
            cleaned = v.replace(",", ".").replace(" ", "").replace("\xa0", "").replace("–", "").replace("—", "")
            return Decimal(cleaned)
        except Exception:
            return None

    def _parse_row(row: list[str]):
        def _get(field: str) -> str:
            idx = best_col.get(field)
            if idx is None or idx >= len(row):
                return ""
            v = row[idx].strip()
            return "" if v.lower() in ("none", "null", "-", "—", "") else v

        item_name = _get("item_name")
        if not item_name:
            return None
        item_type = TYPE_MAP.get(_get("item_type").lower(), "товар")
        quantity = _to_dec(_get("quantity"))
        unit = _get("unit") or "шт"
        unit_price = _to_dec(_get("unit_price"))
        total_price = _to_dec(_get("total_price"))
        if unit_price is None and total_price is not None and quantity:
            try:
                unit_price = total_price / quantity
            except Exception:
                pass
        if total_price is None and unit_price is not None and quantity:
            total_price = unit_price * (quantity or Decimal("1"))
        return {
            "item_name": item_name,
            "item_type": item_type,
            "quantity": float(quantity) if quantity else None,
            "unit": unit,
            "unit_price": float(unit_price) if unit_price else None,
            "total_price": float(total_price) if total_price else None,
        }

    data_rows = best_table[best_header_row + 1:]
    preview = [r for r in (_parse_row(row) for row in data_rows[:100]) if r]

    if not confirm:
        return {"preview": preview, "total_rows": len(preview), "file_type": file_type, "columns_found": list(best_col.keys())}

    # Save items to DB
    org_id = get_single_org_id(current_user)
    prod_q = select(Product)
    if org_id:
        prod_q = prod_q.where((Product.org_id == org_id) | (Product.org_id.is_(None)))
    products = (await db.execute(prod_q)).scalars().all()
    product_by_name = {(p.name or "").lower().strip(): p for p in products}

    added = matched_catalog = new_in_catalog = 0
    for row_data in preview:
        item_name = (row_data["item_name"] or "")[:500]
        qty = Decimal(str(row_data["quantity"])) if row_data["quantity"] else Decimal("1")
        unit_price = Decimal(str(row_data["unit_price"])) if row_data["unit_price"] else None
        total_price = Decimal(str(row_data["total_price"])) if row_data["total_price"] else None
        matched = product_by_name.get(item_name.lower().strip())
        if matched:
            product_id = matched.id
            matched_catalog += 1
            if not unit_price and matched.price:
                unit_price = matched.price
                total_price = qty * unit_price
        else:
            product_id = await _upsert_product_to_catalog(db, item_name, row_data["item_type"], unit_price)
            new_in_catalog += 1
        if total_price is None and unit_price:
            total_price = qty * unit_price
        db.add(PurchaseItem(
            purchase_id=pid, product_id=product_id,
            item_name=item_name, item_type=row_data["item_type"],
            quantity=qty, unit=row_data["unit"],
            unit_price=unit_price, total_price=total_price,
        ))
        added += 1
    await db.commit()
    return {"added": added, "matched_catalog": matched_catalog, "new_in_catalog": new_in_catalog}


# ---------------------------------------------------------------------------
# FEO-format import (57-column layout, headers in row 6)
# ---------------------------------------------------------------------------

FEO_HEADERS = [
    "№ п.п.", "Вид расходов", "Номер субсидии",
    "Номер строки плана закупок", "Категория расходов (ФЭО направление)",
    "Наименование товара", "Контрагент", "Количество поставлено",
    "Ед. изм.", "Стоимость единицы по смете", "Стоимость по плану всего",
    "Подтверждено", "Контрактная цена за единицу", "Стоимость контракта",
    "НМЦК всего", "НМЦК по субсидии", "КПП поставщика",
    "Номер договора", "Дата договора", "Номер ПП",
    "Описание платёжного поручения", "Дата платежа", "Сумма оплаты",
    "Всего исполнено", "Резерв 25", "Резерв 26", "Резерв 27",
    "Статус закупки", "Соответствие НДС (законодательство)", "НДС включён",
    "Соответствие требованиям НДС", "Страна происхождения", "Модель",
    "Номер категории ФЭО", "Статус исполнения", "Процент исполнения",
    "Остаток по обязательствам", "Срок исполнения", "Ссылка на торговую площадку",
    "НДС-расчёт", "Совместные", "Прямой контракт",
    "Прямой контракт (повторный)", "Малые контракты", "Конкурсные",
    "Дата последнего изменения", "Стоимость единицы по оплате", "Итого к оплате",
    "Субсидия", "Часть расходов (направление)", "Период",
    "Оплата", "Сумма НДС", "Счётчик",
    "Описание контракта", "Номер и дата платежа", "",
]


def _feo_find_header_row(rows: list) -> int:
    """Return 0-based index of the header row (first row with >=5 non-empty cells)."""
    for i, row in enumerate(rows[:10]):
        non_empty = sum(1 for c in row if c is not None and str(c).strip())
        if non_empty >= 5:
            return i
    return 0


@router.get("/import/feo-format/template")
async def download_feo_template(_=Depends(require_role(*MANAGER_ROLES))):
    """Скачать шаблон Excel в ФЭО-формате (57 колонок, заголовки в строке 6)."""
    if Workbook is None:
        raise HTTPException(500, "openpyxl не установлен")

    wb = Workbook()
    ws = wb.active
    ws.title = "ФЭО закупки"

    # Rows 1-5: instruction header (merged across all columns)
    ws.merge_cells("A1:BE5")
    instr_cell = ws["A1"]
    instr_cell.value = (
        "ШАБЛОН ДЛЯ ЗАГРУЗКИ ЗАКУПОК В ФЭО-ФОРМАТЕ\n"
        "Строка 6 — заголовки колонок (не изменять).\n"
        "Начиная со строки 7 — данные.\n"
        "Субсидия определяется автоматически по категории ФЭО (колонка 5).\n"
        "Обязательные колонки: 5 (ФЭО направление), 6 (Наименование товара)."
    )
    instr_cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    instr_cell.font = Font(bold=True, size=11)
    ws.row_dimensions[1].height = 80

    # Row 6: headers
    ws.append(FEO_HEADERS)
    header_fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=10)
    for cell in ws[6]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[6].height = 35

    # Row 7: example
    ws.append([
        1, "Товары", "", "1",
        "Направление 1 - Техническое оснащение",
        "Компьютер офисный", "ООО Поставщик", 5, "шт",
        "50000", "250000", "да",
        "48000", "240000", "260000", "260000",
        "771234567890", "Д-001/2026", "15.03.2026",
        "70", "Оплата по договору Д-001/2026", "20.03.2026", "240000",
        "", "", "", "",
        "исполнено", "", "", "", "Россия", "", "",
        "оплачено", "1", "0",
        "30.06.2026", "", "", "", "", "", "", "",
        "", "", "", "", "", "", "", "2026",
        "240000", "0", "", "70 от 20.03.2026", "",
    ])
    ws.row_dimensions[7].height = 20

    ws.freeze_panes = "A7"

    # Column widths
    col_widths = [6, 12, 10, 12, 40, 35, 30, 10, 8, 14, 14, 12,
                  14, 14, 14, 14, 14, 18, 14, 10, 30, 14, 14, 14,
                  10, 10, 10, 14, 25, 14, 25, 18, 14, 14, 16, 12,
                  14, 14, 30, 12, 12, 16, 18, 14, 12, 16, 14, 14,
                  16, 20, 10, 14, 12, 20, 30, 20, 8]
    for i, w in enumerate(col_widths, 1):
        if i <= ws.max_column:
            ws.column_dimensions[ws.cell(6, i).column_letter].width = w

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=feo_import_template.xlsx"},
    )


@router.post("/import/feo-format")
async def import_feo_format(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_role(*MANAGER_ROLES)),
):
    """Импорт закупок из ФЭО-формата (57 колонок, заголовки в строке 6).
    Субсидия определяется автоматически через feo_category.subsidy_id.
    """
    if load_workbook is None:
        raise HTTPException(500, "openpyxl не установлен")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Поддерживаются только файлы .xlsx и .xls")

    content = await file.read()
    wb = load_workbook(BytesIO(content), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Файл пустой")

    header_idx = _feo_find_header_row(rows)
    header_row = rows[header_idx]

    # Build column index by position keyword matching
    def _norm(v) -> str:
        return str(v).strip().lower() if v is not None else ""

    col: dict[str, int] = {}
    for i, h in enumerate(header_row):
        hn = _norm(h)
        if not hn:
            continue
        if "наименован" in hn or "товар" in hn or "услуг" in hn:
            col.setdefault("item_name", i)
        elif "вид расход" in hn or "вид" == hn:
            col.setdefault("item_type", i)
        elif "категори" in hn and ("фэо" in hn or "расход" in hn or "направлен" in hn):
            col.setdefault("feo_name", i)
        elif "контрагент" in hn:
            col.setdefault("contractor_name", i)
        elif "количеств" in hn or "кол-во" in hn:
            col.setdefault("quantity", i)
        elif "ед. изм" in hn or "единиц" in hn:
            col.setdefault("unit", i)
        elif "стоимость единицы по смете" in hn or ("стоимость" in hn and "единиц" in hn and "смет" in hn):
            col.setdefault("planned_unit_price", i)
        elif "стоимость по плану" in hn or ("стоимость" in hn and "план" in hn):
            col.setdefault("planned_total_price", i)
        elif "подтвержд" in hn:
            col.setdefault("confirmed", i)
        elif "стоимость контракта" in hn or ("стоимость" in hn and "контракт" in hn):
            col.setdefault("contract_price", i)
        elif "нмцк" in hn and "субсидии" not in hn:
            col.setdefault("nmck", i)
        elif "кпп" in hn:
            col.setdefault("kpp", i)
        elif "номер договора" in hn or ("номер" in hn and "договор" in hn):
            col.setdefault("contract_number", i)
        elif "дата договора" in hn or ("дата" in hn and "договор" in hn):
            col.setdefault("contract_date", i)
        elif "номер пп" in hn or ("номер" in hn and "платёжн" in hn) or ("номер" in hn and "поручен" in hn):
            col.setdefault("payment_doc_number", i)
        elif "дата платежа" in hn or ("дата" in hn and "платеж" in hn):
            col.setdefault("payment_doc_date", i)
        elif "сумма оплаты" in hn or ("сумма" in hn and "оплат" in hn):
            col.setdefault("payment_amount", i)
        elif "срок исполнения" in hn or ("срок" in hn and "исполнен" in hn):
            col.setdefault("execution_term", i)
        elif "сумма ндс" in hn:
            col.setdefault("vat_amount", i)

    # Fallback: use positional mapping (column indices 0-based from FEO_HEADERS order)
    # Col 5(idx=4)=feo, 6(5)=item_name, 7(6)=contractor, 8(7)=qty, 9(8)=unit,
    # 10(9)=unit_price, 11(10)=planned_total, 12(11)=confirmed, 14(13)=contract_price,
    # 15(14)=nmck, 17(16)=kpp, 18(17)=contract_num, 19(18)=contract_date,
    # 20(19)=pp_num, 22(21)=pp_date, 23(22)=payment_amount, 38(37)=execution_term, 53(52)=vat
    POSITIONAL = {
        "feo_name": 4, "item_name": 5, "item_type": 1,
        "contractor_name": 6, "quantity": 7, "unit": 8,
        "planned_unit_price": 9, "planned_total_price": 10, "confirmed": 11,
        "contract_price": 13, "nmck": 14, "kpp": 16,
        "contract_number": 17, "contract_date": 18,
        "payment_doc_number": 19, "payment_doc_date": 21, "payment_amount": 22,
        "execution_term": 37, "vat_amount": 52,
    }
    if "item_name" not in col and len(header_row) >= 57:
        # Headers didn't match keywords — use positional
        col = {k: v for k, v in POSITIONAL.items()}

    if "item_name" not in col and "feo_name" not in col:
        raise HTTPException(400, "Не удалось найти колонки наименования или ФЭО. Проверьте формат файла.")

    # Load lookup tables
    feo_rows = (await db.execute(select(FeoCategory))).scalars().all()
    feo_by_name: dict[str, FeoCategory] = {}
    for f in feo_rows:
        feo_by_name[f.name.lower().strip()] = f

    cont_rows = (await db.execute(select(Contractor))).scalars().all()
    cont_by_name: dict[str, int] = {c.name.lower().strip(): c.id for c in cont_rows}
    cont_by_kpp: dict[str, int] = {c.kpp.strip(): c.id for c in cont_rows if c.kpp}

    org_id = get_single_org_id(current_user)

    def _cell(row, field) -> Optional[str]:
        idx = col.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        s = str(v).strip()
        return s if s and s.lower() not in ("none", "null", "-", "—", "") else None

    def _to_dec(v) -> Optional[Decimal]:
        if v is None:
            return None
        try:
            return Decimal(str(v).replace(",", ".").replace(" ", "").replace("\xa0", ""))
        except Exception:
            return None

    def _to_date(v) -> Optional[date]:
        if v is None:
            return None
        if isinstance(v, (date, datetime)):
            return v.date() if isinstance(v, datetime) else v
        s = str(v).strip()
        for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
            try:
                return datetime.strptime(s, fmt).date()
            except ValueError:
                pass
        return None

    created = 0
    skipped = 0
    errors_list = []

    data_start = header_idx + 1
    for row_num, row in enumerate(rows[data_start:], start=data_start + 1):
        item_name = _cell(row, "item_name")
        if not item_name:
            skipped += 1
            continue

        # FEO lookup → subsidy
        feo_name_raw = _cell(row, "feo_name")
        feo_obj = None
        if feo_name_raw:
            feo_obj = feo_by_name.get(feo_name_raw.lower().strip())
            if not feo_obj:
                # Partial match
                for k, v in feo_by_name.items():
                    if feo_name_raw.lower() in k or k in feo_name_raw.lower():
                        feo_obj = v
                        break

        if feo_obj is None:
            errors_list.append({"row": row_num, "name": item_name, "message": f"ФЭО категория не найдена: {feo_name_raw!r}"})
            skipped += 1
            continue

        feo_category_id = feo_obj.id
        subsidy_id = feo_obj.subsidy_id

        # Contractor lookup
        cont_name = _cell(row, "contractor_name")
        kpp_raw = _cell(row, "kpp")
        contractor_id = None
        if cont_name:
            contractor_id = cont_by_name.get(cont_name.lower().strip())
            if contractor_id is None and kpp_raw:
                contractor_id = cont_by_kpp.get(kpp_raw.strip())
            if contractor_id is None:
                # Create new contractor
                new_cont = Contractor(name=cont_name, kpp=kpp_raw, org_id=org_id)
                db.add(new_cont)
                await db.flush()
                contractor_id = new_cont.id
                cont_by_name[cont_name.lower().strip()] = contractor_id
                if kpp_raw:
                    cont_by_kpp[kpp_raw.strip()] = contractor_id

        # Numeric fields
        planned_qty = _to_dec(_cell(row, "quantity"))
        planned_unit = _to_dec(_cell(row, "planned_unit_price"))
        planned_total = _to_dec(_cell(row, "planned_total_price"))
        contract_price = _to_dec(_cell(row, "contract_price"))
        nmck = _to_dec(_cell(row, "nmck"))
        payment_amount = _to_dec(_cell(row, "payment_amount"))
        vat_amount = _to_dec(_cell(row, "vat_amount"))

        # Dates
        contract_date = _to_date(_cell(row, "contract_date") or row[col["contract_date"]] if "contract_date" in col else None)
        payment_doc_date = _to_date(_cell(row, "payment_doc_date") or row[col["payment_doc_date"]] if "payment_doc_date" in col else None)
        execution_term = _to_date(_cell(row, "execution_term") or row[col["execution_term"]] if "execution_term" in col else None)

        # Dates directly from row cells (handles datetime objects from openpyxl)
        def _raw_date(field):
            idx = col.get(field)
            if idx is None or idx >= len(row):
                return None
            return _to_date(row[idx])

        contract_date = _raw_date("contract_date")
        payment_doc_date = _raw_date("payment_doc_date")
        execution_term = _raw_date("execution_term")

        # Status inference
        contract_number = _cell(row, "contract_number")
        payment_doc_number = _cell(row, "payment_doc_number")
        if payment_amount and payment_amount > 0:
            status = "paid"
        elif contract_number:
            status = "contracted"
        else:
            status = "confirmed"

        # Confirmed flag
        conf_raw = (_cell(row, "confirmed") or "").lower()
        confirmed = conf_raw in ("да", "yes", "1", "true", "+")

        # Item type
        item_type_raw = (_cell(row, "item_type") or "").lower()
        item_type = "услуга" if "услуг" in item_type_raw else "товар"

        # VAT
        vat_applicable = bool(vat_amount and vat_amount > 0)

        p = Purchase(
            item_name=item_name,
            item_type=item_type,
            feo_category_id=feo_category_id,
            subsidy_id=subsidy_id,
            contractor_id=contractor_id,
            planned_quantity=planned_qty,
            unit=_cell(row, "unit"),
            planned_unit_price=planned_unit,
            planned_total_price=planned_total,
            confirmed=confirmed,
            contract_price=contract_price,
            nmck=nmck,
            total_nmck=nmck,
            contract_number=contract_number,
            contract_date=contract_date,
            payment_doc_number=payment_doc_number,
            payment_doc_date=payment_doc_date,
            payment_amount=payment_amount,
            execution_term=execution_term,
            vat_applicable=vat_applicable,
            status=status,
        )
        db.add(p)
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors_list}


# ── Purchase members (discussion participants) ───────────────────────────────

def _member_dict(m):
    return {
        "id": m.id,
        "purchase_id": m.purchase_id,
        "user_id": m.user_id,
        "role": m.role,
        "added_by_id": m.added_by_id,
        "username": m.user.username if m.user else "",
        "full_name": m.user.full_name if m.user else None,
        "added_by_name": (m.added_by.full_name or m.added_by.username) if m.added_by else None,
        "consent_pending": bool(getattr(m, "consent_pending", False)),
    }


@router.get("/{pid}/members")
async def list_purchase_members(
    pid: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.purchase_event import PurchaseMember
    result = await db.execute(
        select(PurchaseMember).where(PurchaseMember.purchase_id == pid)
    )
    return [_member_dict(m) for m in result.scalars().all()]


@router.post("/{pid}/members")
async def add_purchase_member(
    pid: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.purchase_event import PurchaseMember, PurchaseEvent
    user_id = int(body.get("user_id", 0))
    role = body.get("role", "viewer")

    existing = await db.execute(
        select(PurchaseMember).where(
            PurchaseMember.purchase_id == pid,
            PurchaseMember.user_id == user_id,
        )
    )
    m = existing.scalar_one_or_none()
    is_new = m is None
    if m:
        m.role = role
    else:
        m = PurchaseMember(
            purchase_id=pid, user_id=user_id, role=role,
            added_by_id=current_user.id, consent_pending=(user_id != current_user.id),
        )
        db.add(m)
    await db.flush()

    u = await db.get(User, user_id)
    ev = PurchaseEvent(
        purchase_id=pid,
        user_id=current_user.id,
        event_type="member_added",
        data={"username": (u.full_name or u.username) if u else str(user_id)},
    )
    db.add(ev)
    await db.commit()
    await db.refresh(m)

    # Notify added user
    if u and u.id != current_user.id and is_new:
        try:
            purchase = await db.get(Purchase, pid)
            if purchase:
                if m.consent_pending:
                    from app.notifications import notify_purchase_consent_required
                    await notify_purchase_consent_required(
                        purchase, u, current_user.full_name or current_user.username
                    )
                else:
                    from app.notifications import notify_purchase_member_added
                    await notify_purchase_member_added(
                        purchase, u, current_user.full_name or current_user.username
                    )
        except Exception as e:
            logger.warning("Member add notify failed: %s", e)

    return _member_dict(m)


@router.delete("/{pid}/members/{user_id}")
async def remove_purchase_member(
    pid: int,
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.purchase_event import PurchaseMember, PurchaseEvent
    result = await db.execute(
        select(PurchaseMember).where(
            PurchaseMember.purchase_id == pid,
            PurchaseMember.user_id == user_id,
        )
    )
    m = result.scalar_one_or_none()
    if m:
        u = await db.get(User, user_id)
        ev = PurchaseEvent(
            purchase_id=pid,
            user_id=current_user.id,
            event_type="member_removed",
            data={"username": (u.full_name or u.username) if u else str(user_id)},
        )
        db.add(ev)
        await db.delete(m)
        await db.commit()
    return {"ok": True}


# ── Purchase comments (chat) ────────────────────────────────────────────────

@router.get("/{pid}/comments")
async def list_purchase_comments(
    pid: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.purchase_comment import PurchaseComment
    result = await db.execute(
        select(PurchaseComment)
        .where(PurchaseComment.purchase_id == pid)
        .order_by(PurchaseComment.created_at.asc())
    )
    return result.scalars().all()


@router.post("/{pid}/comments")
async def add_purchase_comment(
    pid: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.purchase_comment import PurchaseComment
    import re as _re

    text = (body.get("text") or "").strip()
    if not text:
        raise HTTPException(422, "Комментарий не может быть пустым")

    p = await db.get(Purchase, pid)
    if not p:
        raise HTTPException(404, "Закупка не найдена")

    comment = PurchaseComment(
        purchase_id=pid,
        user_id=current_user.id,
        user_name=current_user.full_name or current_user.username,
        text=text,
    )
    db.add(comment)
    await db.commit()
    await db.refresh(comment)

    # Notify @mentioned users via Telegram
    try:
        from app.notifications import notify_user, _esc, _purchase_url
        mentions = _re.findall(r'@(\S+)', text)
        if mentions:
            all_users = (await db.execute(select(User))).scalars().all()
            clean_text = _re.sub(r'@[A-Za-zА-Яа-яёЁ\s]{2,40}', '', text).strip()
            clean_text = _re.sub(r'\s{2,}', ' ', clean_text) or text
            preview = _esc(clean_text[:150])
            subject = _esc(p.subject or f"Закупка №{p.purchase_number}")
            sender = _esc(current_user.full_name or current_user.username)
            msg = (
                f"💬 <b>Вас упомянули в закупке</b>\n\n"
                f"📌 <b>{subject}</b>\n"
                f"👤 <i>{sender}</i>:\n"
                f"{preview}"
            )
            for u in all_users:
                for m in mentions:
                    if (u.username and m.lower() == u.username.lower()) or \
                       (u.full_name and m.lower() in u.full_name.lower()):
                        if u.id != current_user.id:
                            await notify_user(u, msg,
                                               button_url=_purchase_url(p.id),
                                               button_label="Открыть закупку")
                            break
    except Exception:
        pass

    return comment


@router.delete("/{pid}/comments/{comment_id}")
async def delete_purchase_comment(
    pid: int, comment_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    from app.models.purchase_comment import PurchaseComment
    c = await db.get(PurchaseComment, comment_id)
    if not c or c.purchase_id != pid:
        raise HTTPException(404, "Комментарий не найден")
    if c.user_id != current_user.id and current_user.role not in ("superadmin", "org_admin", "admin"):
        raise HTTPException(403, "Нельзя удалить чужой комментарий")
    await db.delete(c)
    await db.commit()
    return {"ok": True}


@router.post("/{pid}/broadcast")
async def broadcast_from_purchase(
    pid: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Broadcast from purchase context."""
    from app.models.organization import Organization
    from app.models.department import Department, DepartmentMember
    from app.models.purchase_comment import PurchaseComment
    from app.notifications import notify_user, _esc, _purchase_url

    BROADCAST_ROLES = ("superadmin", "org_admin", "admin", "manager")
    if current_user.role not in BROADCAST_ROLES:
        raise HTTPException(403, "Рассылка доступна только администраторам и менеджерам")

    p = await db.get(Purchase, pid)
    if not p:
        raise HTTPException(404, "Закупка не найдена")

    text = (body.get("text") or "").strip()
    if not text:
        raise HTTPException(422, "Текст сообщения обязателен")

    scope = body.get("scope", "")
    scope_id = body.get("scope_id")

    q = select(User).where(User.id != current_user.id)
    if scope == "department" and scope_id:
        member_uids = select(DepartmentMember.user_id).where(DepartmentMember.department_id == int(scope_id))
        q = q.where(User.id.in_(member_uids))
    elif scope == "organization" and scope_id:
        q = q.where(User.org_id == int(scope_id))
    elif scope == "all":
        org_ids = get_org_filter(current_user)
        if org_ids is not None:
            q = q.where(User.org_id.in_(org_ids))
    else:
        raise HTTPException(422, "Укажите scope")

    users = (await db.execute(q)).scalars().all()

    sender_name = current_user.full_name or current_user.username
    subject = _esc(p.subject or f"Закупка №{p.purchase_number}")
    msg = (
        f"📢 <b>Рассылка</b>\n\n"
        f"📌 <b>{subject}</b>\n"
        f"👤 <i>{_esc(sender_name)}</i>:\n"
        f"{_esc(text)}"
    )

    sent = 0
    for u in users:
        if getattr(u, "telegram_id", None) or getattr(u, "max_chat_id", None):
            await notify_user(u, msg, button_url=_purchase_url(p.id), button_label="Открыть закупку")
            sent += 1

    # Save as comment
    scope_label = {"department": "отделу", "organization": "организации", "all": "всем"}.get(scope, scope)
    db.add(PurchaseComment(
        purchase_id=pid, user_id=current_user.id, user_name=sender_name,
        text=f"[Рассылка {scope_label}] {text}",
    ))
    await db.commit()

    return {"ok": True, "sent": sent, "total_users": len(users)}
