"""
Модуль для импорта данных из исходного Excel файла "Патриотика 2025 (5).xlsx"
"""

import openpyxl
from openpyxl import load_workbook
import pandas as pd
from datetime import datetime
from typing import Dict, List, Optional

FIELD_KEYWORDS = {
    'name': ['наимен', 'контрагент', 'поставщик', 'получатель', 'организац'],
    'full_name': ['полное наимен'],
    'inn': ['инн'],
    'kpp': ['кпп'],
    'ogrn': ['огрн'],
    'okpo': ['окпо'],
    'oktmo': ['октмо'],
    'director_name': ['руководител', 'директор', 'генеральный'],
    'director_position': ['должность руководител', 'должность директора'],
    'bank_account': ['расчет', 'р/с', 'р\\с', 'р-с'],
    'corr_account': ['корреспондентский', 'к/с', 'к\\с', 'к-с'],
    'bik': ['бик'],
    'bank_name': ['банк'],
    'legal_address': ['юридичес', 'юр. адрес'],
    'mail_address': ['почтовый', 'почт. адрес'],
    'fact_address': ['фактичес', 'факт. адрес'],
    'phone': ['телефон', 'моб. телефон'],
    'fax': ['факс'],
    'email': ['email', 'e-mail', 'электрон'],
    'site': ['сайт', 'web'],
    'contact_person': ['контактное лицо'],
    'contact_position': ['должность контакт'],
    'contact_phone': ['телефон контакт'],
    'contact_email': ['email контакт', 'e-mail контакт'],
    'comments': ['примечан', 'комментар']
}

class DataImporter:
    """Класс для импорта данных из исходного файла"""
    
    def __init__(self, source_file, target_file='CRM_База_Данных.xlsx'):
        self.source_file = source_file
        self.target_file = target_file
        self.source_wb = None
        self.target_wb = None
    
    def load_files(self):
        """Загрузка исходного и целевого файлов"""
        try:
            self.source_wb = load_workbook(self.source_file, data_only=True)
            self.target_wb = load_workbook(self.target_file)
            return True
        except Exception as e:
            print(f"Ошибка при загрузке файлов: {e}")
            return False
    
    def import_subsidies(self):
        """Импорт субсидий из листа GoodsService"""
        if 'GoodsService' not in self.source_wb.sheetnames:
            print("Лист GoodsService не найден в исходном файле")
            return 0
        
        source_ws = self.source_wb['GoodsService']
        target_ws = self.target_wb['Субсидии']
        
        # Ищем столбец с субсидиями
        headers = [cell.value for cell in source_ws[1]]
        subsidy_col = None
        
        for i, header in enumerate(headers, 1):
            if header and isinstance(header, str):
                header_lower = header.lower()
                if any(word in header_lower for word in ['субсидия', 'subsidy', 'ведомство', 'department']):
                    subsidy_col = i
                    break
        
        if not subsidy_col:
            print("Столбец с субсидиями не найден")
            return 0
        
        # Собираем уникальные субсидии
        subsidies_map = {}
        for row in source_ws.iter_rows(min_row=2, values_only=False):
            subsidy_value = row[subsidy_col - 1].value
            if subsidy_value and str(subsidy_value).strip():
                subsidy_name = str(subsidy_value).strip()
                if subsidy_name not in subsidies_map:
                    subsidies_map[subsidy_name] = {
                        'name': subsidy_name,
                        'short_name': self.get_short_name(subsidy_name),
                        'department': self.determine_department(subsidy_name),
                        'year': datetime.now().year
                    }
        
        # Проверяем существующие субсидии
        existing_subsidies = set()
        for row in target_ws.iter_rows(min_row=2, values_only=True):
            if row[1]:  # Наименование
                existing_subsidies.add(str(row[1]).strip())
        
        # Добавляем новые субсидии
        new_count = 0
        next_id = self.get_next_id(target_ws)
        
        for subsidy_name, subsidy_data in subsidies_map.items():
            if subsidy_name not in existing_subsidies:
                next_row = target_ws.max_row + 1
                target_ws.cell(row=next_row, column=1, value=next_id)
                target_ws.cell(row=next_row, column=2, value=subsidy_data['name'])
                target_ws.cell(row=next_row, column=3, value=subsidy_data['short_name'])
                target_ws.cell(row=next_row, column=4, value=subsidy_data['department'])
                target_ws.cell(row=next_row, column=5, value=subsidy_data['year'])
                target_ws.cell(row=next_row, column=12, value='Да')
                next_id += 1
                new_count += 1
        
        print(f"Импортировано субсидий: {new_count}")
        return new_count
    
    def import_categories_feo(self):
        """Импорт категорий ФЭО"""
        # Ищем лист с категориями ФЭО
        source_sheet = None
        for sheet_name in self.source_wb.sheetnames:
            name_lower = sheet_name.lower()
            if 'категории' in name_lower and ('фео' in name_lower or 'feo' in name_lower):
                source_sheet = self.source_wb[sheet_name]
                break
        
        if not source_sheet:
            print("Лист с категориями ФЭО не найден")
            return 0
        
        target_ws = self.target_wb['Категории_из_ФЭО']
        
        # Очищаем существующие данные (кроме заголовков)
        if target_ws.max_row > 1:
            target_ws.delete_rows(2, target_ws.max_row)
        
        # Копируем данные
        source_data = []
        for row in source_sheet.iter_rows(min_row=2, values_only=True):
            if any(row):  # Есть хотя бы одно значение
                source_data.append(row)
        
        # Вставляем данные
        for i, row_data in enumerate(source_data, start=2):
            for j, value in enumerate(row_data, start=1):
                if value:
                    target_ws.cell(row=i, column=j, value=value)
        
        print(f"Импортировано категорий ФЭО: {len(source_data)}")
        return len(source_data)
    
    def import_categories_app(self):
        """Импорт категорий из приложения"""
        source_sheet = None
        for sheet_name in self.source_wb.sheetnames:
            name_lower = sheet_name.lower()
            if 'категории' in name_lower and 'приложения' in name_lower:
                source_sheet = self.source_wb[sheet_name]
                break
        
        if not source_sheet:
            print("Лист с категориями из приложения не найден")
            return 0
        
        target_ws = self.target_wb['Категории_из_приложения']
        
        # Очищаем существующие данные
        if target_ws.max_row > 1:
            target_ws.delete_rows(2, target_ws.max_row)
        
        # Копируем данные
        source_data = []
        for row in source_sheet.iter_rows(min_row=2, values_only=True):
            if any(row):
                source_data.append(row)
        
        # Вставляем данные
        for i, row_data in enumerate(source_data, start=2):
            for j, value in enumerate(row_data, start=1):
                if value:
                    target_ws.cell(row=i, column=j, value=value)
        
        print(f"Импортировано категорий из приложения: {len(source_data)}")
        return len(source_data)
    
    def import_feo_base(self):
        """Импорт базового ФЭО (лист 'ФЭО_База')"""
        if not self.source_wb or not self.target_wb:
            if not self.load_files():
                return 0
        
        if 'ФЭО_База' not in self.source_wb.sheetnames:
            print("Лист 'ФЭО_База' не найден в исходном файле")
            return 0
        if 'ФЭО_База' not in self.target_wb.sheetnames:
            print("Лист 'ФЭО_База' не найден в целевом файле")
            return 0
        
        source_ws = self.source_wb['ФЭО_База']
        target_ws = self.target_wb['ФЭО_База']
        
        # Очищаем существующие данные (кроме заголовков)
        if target_ws.max_row > 1:
            target_ws.delete_rows(2, target_ws.max_row - 1)
        
        imported = 0
        for row_idx, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            for col_idx, value in enumerate(row, start=1):
                if value is not None and value != "":
                    target_ws.cell(row=row_idx, column=col_idx, value=value)
            imported += 1
        
        print(f"Импортировано записей базового ФЭО: {imported}")
        return imported
    
    def import_feo_applications(self):
        """Импорт приложений ФЭО (лист 'ФЭО_Приложения')"""
        if not self.source_wb or not self.target_wb:
            if not self.load_files():
                return 0
        
        if 'ФЭО_Приложения' not in self.source_wb.sheetnames:
            print("Лист 'ФЭО_Приложения' не найден в исходном файле")
            return 0
        if 'ФЭО_Приложения' not in self.target_wb.sheetnames:
            print("Лист 'ФЭО_Приложения' не найден в целевом файле")
            return 0
        
        source_ws = self.source_wb['ФЭО_Приложения']
        target_ws = self.target_wb['ФЭО_Приложения']
        
        # Очищаем существующие данные (кроме заголовков)
        if target_ws.max_row > 1:
            target_ws.delete_rows(2, target_ws.max_row - 1)
        
        imported = 0
        for row_idx, row in enumerate(source_ws.iter_rows(min_row=2, values_only=True), start=2):
            if not row or not any(row):
                continue
            for col_idx, value in enumerate(row, start=1):
                if value is not None and value != "":
                    target_ws.cell(row=row_idx, column=col_idx, value=value)
            imported += 1
        
        print(f"Импортировано приложений ФЭО: {imported}")
        return imported
    
    def import_contractors(self):
        """Импорт контрагентов из всех доступных источников"""
        if 'Контрагенты' not in self.target_wb.sheetnames:
            print("Лист 'Контрагенты' не найден в целевом файле")
            return 0
        
        target_ws = self.target_wb['Контрагенты']
        existing = self._build_existing_contractors(target_ws)
        
        total_new = 0
        total_new += self._import_contractors_from_goodsservice(existing, target_ws)
        total_new += self._import_contractors_from_scroller(existing, target_ws)
        
        print(f"Импортировано контрагентов: {total_new}")
        return total_new
    
    def import_all(self):
        """Импорт всех данных"""
        if not self.load_files():
            return False
        
        print("Начало импорта данных...")
        print("-" * 50)
        
        subsidies = self.import_subsidies()
        categories_feo = self.import_categories_feo()
        categories_app = self.import_categories_app()
        contractors = self.import_contractors()
        
        print("-" * 50)
        print(f"Импорт завершен:")
        print(f"  Субсидии: {subsidies}")
        print(f"  Категории ФЭО: {categories_feo}")
        print(f"  Категории из приложения: {categories_app}")
        print(f"  Контрагенты: {contractors}")
        
        # Сохраняем целевой файл
        self.target_wb.save(self.target_file)
        print(f"\nДанные сохранены в файл: {self.target_file}")
        
        return True

    # --------------------------- Контрагенты --------------------------- #

    def _build_existing_contractors(self, ws):
        existing = {'by_inn': {}, 'by_name': {}}
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            name = self._normalize_text(row[1].value if len(row) > 1 else None)
            inn = self._normalize_text(row[3].value if len(row) > 3 else None)
            if inn:
                existing['by_inn'][inn] = row_idx
            if name:
                existing['by_name'][name] = row_idx
        return existing

    def _normalize_text(self, value):
        if value is None:
            return ''
        if isinstance(value, datetime):
            return value.strftime('%d.%m.%Y')
        if isinstance(value, float):
            if value.is_integer():
                return str(int(value))
            return f"{value:.2f}".replace('.', ',')
        return str(value).strip()

    def _get_cell(self, row, idx):
        if not idx or idx <= 0:
            return None
        if idx - 1 < len(row):
            return row[idx - 1]
        return None

    def _upsert_contractor(self, contractor: Dict[str, str], ws, existing: Dict[str, Dict]) -> bool:
        name = contractor.get('name') or ''
        inn = contractor.get('inn') or ''
        row_idx = None
        if inn and inn in existing['by_inn']:
            row_idx = existing['by_inn'][inn]
        elif name and name in existing['by_name']:
            row_idx = existing['by_name'][name]

        is_new = False
        if row_idx is None:
            row_idx = ws.max_row + 1
            ws.cell(row=row_idx, column=1, value=self.get_next_id(ws))
            ws.cell(row=row_idx, column=31, value=datetime.now())
            is_new = True
            if name:
                existing['by_name'][name] = row_idx
            if inn:
                existing['by_inn'][inn] = row_idx

        ws.cell(row=row_idx, column=32, value=datetime.now())
        ws.cell(row=row_idx, column=33, value='Да')

        mapping = {
            2: contractor.get('name'),
            3: contractor.get('full_name'),
            4: contractor.get('inn'),
            5: contractor.get('kpp'),
            6: contractor.get('ogrn'),
            7: contractor.get('okpo'),
            8: contractor.get('oktmo'),
            9: contractor.get('director_name'),
            10: contractor.get('director_position'),
            15: contractor.get('bank_account'),
            16: contractor.get('corr_account'),
            17: contractor.get('bik'),
            18: contractor.get('bank_name'),
            19: contractor.get('legal_address'),
            20: contractor.get('mail_address'),
            21: contractor.get('fact_address'),
            22: contractor.get('phone'),
            23: contractor.get('fax'),
            24: contractor.get('email'),
            25: contractor.get('site'),
            26: contractor.get('contact_person'),
            27: contractor.get('contact_position'),
            28: contractor.get('contact_phone'),
            29: contractor.get('contact_email'),
            30: contractor.get('comments'),
        }

        for col, value in mapping.items():
            if value:
                ws.cell(row=row_idx, column=col, value=value)

        return is_new

    def _import_contractors_from_goodsservice(self, existing, ws):
        if 'GoodsService' not in self.source_wb.sheetnames:
            return 0
        source_ws = self.source_wb['GoodsService']
        headers = [cell.value for cell in source_ws[1]]
        contractor_col = None
        inn_col = None
        for idx, header in enumerate(headers, start=1):
            if not header or not isinstance(header, str):
                continue
            header_lower = header.lower()
            if not contractor_col and any(word in header_lower for word in ['контрагент', 'поставщик', 'исполнитель']):
                contractor_col = idx
            if not inn_col and 'инн' in header_lower:
                inn_col = idx
        if not contractor_col:
            return 0
        new_count = 0
        for row in source_ws.iter_rows(min_row=2, values_only=True):
            if not row:
                continue
            name = self._normalize_text(row[contractor_col - 1]) if contractor_col <= len(row) else ''
            inn = ''
            if inn_col and inn_col <= len(row):
                inn = self._normalize_text(row[inn_col - 1])
            if not name and not inn:
                continue
            contractor = {
                'name': name or inn,
                'full_name': name or inn,
                'inn': inn
            }
            if self._upsert_contractor(contractor, ws, existing):
                new_count += 1
        return new_count

    def _import_contractors_from_scroller(self, existing, ws):
        sheet_name = None
        for name in self.source_wb.sheetnames:
            if 'scroller_united' in name.lower():
                sheet_name = name
                break
        if not sheet_name:
            print("Лист Scroller_United не найден")
            return 0
        source_ws = self.source_wb[sheet_name]
        header_map = self._extract_header_map(source_ws, header_rows=3)

        indexes = {field: self._find_column(header_map, keywords) for field, keywords in FIELD_KEYWORDS.items()}
        new_count = 0
        start_row = 4
        max_col = source_ws.max_column

        for row in source_ws.iter_rows(min_row=start_row, values_only=True):
            if not row or not any(row):
                continue
            name = self._normalize_text(self._get_cell(row, indexes.get('name')))
            inn = self._normalize_text(self._get_cell(row, indexes.get('inn')))
            if not name and not inn:
                continue
            contractor = {
                'name': name or inn,
                'full_name': self._normalize_text(self._get_cell(row, indexes.get('full_name'))) or name or inn,
                'inn': inn,
                'kpp': self._normalize_text(self._get_cell(row, indexes.get('kpp'))),
                'ogrn': self._normalize_text(self._get_cell(row, indexes.get('ogrn'))),
                'okpo': self._normalize_text(self._get_cell(row, indexes.get('okpo'))),
                'oktmo': self._normalize_text(self._get_cell(row, indexes.get('oktmo'))),
                'director_name': self._normalize_text(self._get_cell(row, indexes.get('director_name'))),
                'director_position': self._normalize_text(self._get_cell(row, indexes.get('director_position'))),
                'bank_account': self._normalize_text(self._get_cell(row, indexes.get('bank_account'))),
                'corr_account': self._normalize_text(self._get_cell(row, indexes.get('corr_account'))),
                'bik': self._normalize_text(self._get_cell(row, indexes.get('bik'))),
                'bank_name': self._normalize_text(self._get_cell(row, indexes.get('bank_name'))),
                'legal_address': self._normalize_text(self._get_cell(row, indexes.get('legal_address'))),
                'mail_address': self._normalize_text(self._get_cell(row, indexes.get('mail_address'))),
                'fact_address': self._normalize_text(self._get_cell(row, indexes.get('fact_address'))),
                'phone': self._normalize_text(self._get_cell(row, indexes.get('phone'))),
                'fax': self._normalize_text(self._get_cell(row, indexes.get('fax'))),
                'email': self._normalize_text(self._get_cell(row, indexes.get('email'))),
                'site': self._normalize_text(self._get_cell(row, indexes.get('site'))),
                'contact_person': self._normalize_text(self._get_cell(row, indexes.get('contact_person'))),
                'contact_position': self._normalize_text(self._get_cell(row, indexes.get('contact_position'))),
                'contact_phone': self._normalize_text(self._get_cell(row, indexes.get('contact_phone'))),
                'contact_email': self._normalize_text(self._get_cell(row, indexes.get('contact_email'))),
                'comments': self._normalize_text(self._get_cell(row, indexes.get('comments')))
            }
            if self._upsert_contractor(contractor, ws, existing):
                new_count += 1
        return new_count

    def _extract_header_map(self, ws, header_rows=3) -> Dict[int, str]:
        header_map = {}
        max_col = ws.max_column
        for col in range(1, max_col + 1):
            header_value = ''
            for row in range(1, header_rows + 1):
                value = ws.cell(row=row, column=col).value
                if value and isinstance(value, str):
                    header_value = value.strip().lower()
                    break
            header_map[col] = header_value
        return header_map

    def _find_column(self, header_map: Dict[int, str], keywords: List[str]) -> Optional[int]:
        for col, header in header_map.items():
            if not header:
                continue
            for keyword in keywords:
                if keyword in header:
                    return col
        return None
    
    def get_next_id(self, ws):
        """Получение следующего ID"""
        if ws.max_row == 1:
            return 1
        
        last_id = ws.cell(row=ws.max_row, column=1).value
        if last_id is None:
            return 1
        
        try:
            return int(last_id) + 1
        except:
            return 1
    
    def get_short_name(self, full_name):
        """Получение короткого названия"""
        words = str(full_name).split()
        if len(words) > 3:
            return ' '.join(words[:3])
        return full_name
    
    def determine_department(self, subsidy_name):
        """Определение ведомства по названию субсидии"""
        name_lower = str(subsidy_name).lower()
        if 'минпрос' in name_lower or 'просвещ' in name_lower:
            return 'Минпрос'
        elif 'минтруд' in name_lower or 'труд' in name_lower:
            return 'Минтруд'
        elif 'фадм' in name_lower:
            return 'ФАДМ'
        elif 'регион' in name_lower:
            return 'Регионы'
        return ''

if __name__ == '__main__':
    import sys
    
    if len(sys.argv) < 2:
        print("Использование: python import_data.py <путь_к_файлу_Патриотика.xlsx>")
        sys.exit(1)
    
    source_file = sys.argv[1]
    importer = DataImporter(source_file)
    importer.import_all()

