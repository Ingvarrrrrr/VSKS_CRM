from datetime import date, timedelta
from typing import Optional
from fastapi import APIRouter, Depends, Query
from sqlalchemy import select, func, case, extract, and_, or_, literal
from sqlalchemy.orm import selectinload
from sqlalchemy.ext.asyncio import AsyncSession
from app.database import get_db
from app.models.feo_category import FeoCategory
from app.models.feo_planned_item import FeoPlannedItem
from app.models.purchase import Purchase
from app.models.subsidy import Subsidy
from app.models.contractor import Contractor
from app.models.contract import Contract
from app.models.user import User
from app.auth.jwt import get_current_user, get_org_filter
from app.auth.visibility import get_visible_subsidy_ids
from app.routers.subsidies import calculate_budgets_bulk, _calculate_spent_bulk, _calculate_planned_amounts_bulk, _calculate_feo_planned_tree_bulk
from app.services.feo_plan import calculate_ceiling_forecasts_bulk
from app.config import settings
from decimal import Decimal

router = APIRouter(prefix="/api/dashboard", tags=["dashboard"])


def _expected_payment_date(p: Purchase):
    """Когда ожидается фактическая выплата."""
    if p.status == 'paid' and p.payment_doc_date:
        return p.payment_doc_date.date() if hasattr(p.payment_doc_date, 'date') else p.payment_doc_date
    # для остальных — service_end_date > execution_term > service_deadline_date > contract_date
    for fld in ('service_end_date', 'execution_term', 'service_deadline_date', 'contract_date'):
        v = getattr(p, fld, None)
        if v:
            return v if isinstance(v, date) else v.date()
    return None


def obligation_date(p: Purchase) -> Optional[date]:
    """Дата возникновения обязательства — НЕ использует contract_date как fallback.

    Приоритет:
    1. is_prepayment + prepayment_date → дата предоплаты
    2. service_deadline_date → срок до даты (mode='deadline')
    3. service_end_date → конец периода оказания услуг
    4. execution_term → срок исполнения
    5. None — «без срока» (попадает в no_deadline bucket)
    """
    if getattr(p, 'is_prepayment', False) and getattr(p, 'prepayment_date', None):
        v = p.prepayment_date
        return v if isinstance(v, date) else v.date()
    for fld in ('service_deadline_date', 'service_end_date', 'execution_term'):
        v = getattr(p, fld, None)
        if v:
            return v if isinstance(v, date) else v.date()
    return None


_UNSET = object()


def _apply_subsidy_org_filter(query, user: User, org_ids=_UNSET, subsidy_ids=_UNSET):
    """Filter subsidies by org_ids — или напрямую по subsidy_ids (приоритет).

    subsidy_ids передаётся для scope=dashboard (двухуровневая видимость по вкладке
    «Дашборд»): None → не фильтровать (SaaS), set → Subsidy.id.in_(set).
    Иначе org_ids (или get_org_filter при _UNSET). Caller различает None и [].
    """
    if subsidy_ids is not _UNSET:
        if subsidy_ids is not None:
            query = query.where(Subsidy.id.in_(subsidy_ids))
        return query
    if org_ids is _UNSET:
        org_ids = get_org_filter(user)
    if org_ids is not None:
        query = query.where(Subsidy.org_id.in_(org_ids))
    return query


def _apply_purchase_org_filter(query, user: User, org_ids=_UNSET, subsidy_ids=_UNSET):
    """Filter purchases via subsidy.org_id — или напрямую по subsidy_ids (приоритет).

    subsidy_ids для scope=dashboard: None → не фильтровать, set →
    Purchase.subsidy_id.in_(set). Иначе org_ids (или get_org_filter при _UNSET).
    """
    if subsidy_ids is not _UNSET:
        if subsidy_ids is not None:
            query = query.where(Purchase.subsidy_id.in_(subsidy_ids))
        return query
    if org_ids is _UNSET:
        org_ids = get_org_filter(user)
    if org_ids is not None:
        query = query.where(Purchase.subsidy_id.in_(
            select(Subsidy.id).where(Subsidy.org_id.in_(org_ids))
        ))
    return query


@router.get("/")
async def dashboard(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    scope: Optional[str] = Query(None),
):
    org_ids = get_org_filter(current_user)
    # scope=dashboard → двухуровневая видимость субсидий по вкладке «Дашборд»
    # (пер-субсидийная галочка перебивает орг-дефолт). Гейтим данные по subsidy_id.
    dash_sids = _UNSET
    if scope == "dashboard":
        dash_sids = await get_visible_subsidy_ids(current_user, db, "dashboard")

    # Get categories filtered by org / dashboard-subsidies
    cat_q = select(FeoCategory).order_by(FeoCategory.level, FeoCategory.id)
    if dash_sids is not _UNSET:
        if dash_sids is not None:
            cat_q = cat_q.where(FeoCategory.subsidy_id.in_(dash_sids))
    elif org_ids is not None:
        cat_q = cat_q.where(FeoCategory.subsidy_id.in_(
            select(Subsidy.id).where(Subsidy.org_id.in_(org_ids))
        ))
    cat_result = await db.execute(cat_q)
    cats = cat_result.scalars().all()

    # Get aggregated purchase data per feo_category_id
    agg_q = select(
        Purchase.feo_category_id,
        func.coalesce(func.sum(Purchase.planned_total_price), 0).label("planned"),
        func.coalesce(func.sum(
            case((Purchase.confirmed == True, Purchase.final_total_amount), else_=0)
        ), 0).label("confirmed"),
        func.coalesce(func.sum(Purchase.delivery_payment_amount), 0).label("payment"),
    ).group_by(Purchase.feo_category_id)
    if dash_sids is not _UNSET:
        agg_q = _apply_purchase_org_filter(agg_q, current_user, subsidy_ids=dash_sids)
    else:
        agg_q = _apply_purchase_org_filter(agg_q, current_user, org_ids)
    agg = await db.execute(agg_q)

    agg_map = {}
    for row in agg:
        agg_map[row.feo_category_id] = {
            "total_planned": float(row.planned),
            "total_confirmed": float(row.confirmed),
            "total_payment": float(row.payment),
        }

    # Build tree
    by_id = {}
    for c in cats:
        by_id[c.id] = {
            "id": c.id, "name": c.name, "level": c.level, "code": c.code,
            "subsidy_id": c.subsidy_id,
            "total_planned": 0, "total_confirmed": 0, "total_payment": 0,
            "children": []
        }
        if c.id in agg_map:
            by_id[c.id].update(agg_map[c.id])

    roots = []
    for c in cats:
        node = by_id[c.id]
        if c.parent_id and c.parent_id in by_id:
            by_id[c.parent_id]["children"].append(node)
        else:
            roots.append(node)

    # Roll up totals
    def rollup(node):
        for child in node["children"]:
            rollup(child)
            node["total_planned"] += child["total_planned"]
            node["total_confirmed"] += child["total_confirmed"]
            node["total_payment"] += child["total_payment"]
    for r in roots:
        rollup(r)

    # Global totals (filtered by org)
    total_obl_q = select(func.coalesce(func.sum(Purchase.final_total_amount), 0)).where(Purchase.confirmed == True)
    total_obl_q = _apply_purchase_org_filter(total_obl_q, current_user, org_ids, subsidy_ids=dash_sids)
    total_obligations = float((await db.execute(total_obl_q)).scalar() or 0)

    total_pay_q = select(func.coalesce(func.sum(Purchase.delivery_payment_amount), 0))
    total_pay_q = _apply_purchase_org_filter(total_pay_q, current_user, org_ids, subsidy_ids=dash_sids)
    total_payments = float((await db.execute(total_pay_q)).scalar() or 0)

    return {
        "subsidy_limit": settings.SUBSIDY_LIMIT,
        "total_obligations": total_obligations,
        "total_payments": total_payments,
        "remaining": settings.SUBSIDY_LIMIT - total_obligations,
        "categories": roots
    }


@router.get("/charts")
async def dashboard_charts(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    scope: Optional[str] = Query(None),
):
    org_ids = get_org_filter(current_user)
    # «Субсидии»: орг-админ-роль + гранты. «Дашборд»: двухуровневая видимость по вкладке dashboard.
    managed = scope == "managed"
    dashboard = scope in ("dashboard", "dashboard.radar", "plan")
    visible_subsidy_ids = None
    if managed:
        visible_subsidy_ids = await get_visible_subsidy_ids(current_user, db)
    elif dashboard:
        visible_subsidy_ids = await get_visible_subsidy_ids(current_user, db, scope)
    use_sids = managed or dashboard

    # Status counts for pie chart (filtered)
    status_q = select(Purchase.status, func.count(Purchase.id).label("cnt")).group_by(Purchase.status)
    if use_sids:
        status_q = _apply_purchase_org_filter(status_q, current_user, subsidy_ids=visible_subsidy_ids)
    else:
        status_q = _apply_purchase_org_filter(status_q, current_user, org_ids)
    status_result = await db.execute(status_q)
    status_counts = {row.status: row.cnt for row in status_result}

    # Per-subsidy aggregated purchase data (filtered)
    subsidy_q = (
        select(
            Subsidy.id, Subsidy.name, Subsidy.year, Subsidy.budget,
            func.coalesce(func.sum(Purchase.planned_total_price), 0).label("total_planned"),
            func.coalesce(func.sum(
                case((Purchase.status.in_(["contracted", "delivered", "paid"]), Purchase.contract_price), else_=None)
            ), 0).label("total_confirmed"),
            func.coalesce(func.sum(
                case((Purchase.status == "paid", Purchase.payment_amount), else_=None)
            ), 0).label("total_paid"),
            func.coalesce(func.sum(
                case((Purchase.status == "work_in_progress", Purchase.planned_total_price), else_=None)
            ), 0).label("total_plan_schedule"),
            # «Заказано» = закупки, реально дошедшие до стадии «Заказано» и дальше
            # (ordered/delivered/paid) — НЕ 'contracted' (договор заключён, но ещё не заказано).
            # Сумма = COALESCE(contract_price, planned_total_price), без ветвления по
            # purchase_contract_type (раньше закупки с purchase_contract_type IS NULL тихо
            # выпадали из суммы — второй дефект старой формулы).
            # Ежемесячные закупки (is_monthly_payment=true) сюда НЕ входят — для них своя
            # помесячная логика начисления (см. monthly_ordered_map ниже), иначе их полная
            # сумма договора задвоилась бы с прогрессивным начислением по месяцам.
            func.coalesce(func.sum(
                case(
                    (
                        and_(
                            Purchase.status.in_(["ordered", "delivered", "paid"]),
                            Purchase.is_monthly_payment.isnot(True),
                        ),
                        func.coalesce(Purchase.contract_price, Purchase.planned_total_price)
                    ),
                    else_=None
                )
            ), 0).label("total_ordered"),
            # Per-subsidy basket mirrors (для total_work / total_delivered / total_delivered_unpaid)
            func.coalesce(func.sum(
                case(
                    (Purchase.status.in_(["contracted", "ordered"]),
                     func.coalesce(Purchase.contract_price, Purchase.planned_total_price)),
                    else_=None
                )
            ), 0).label("w_ordered"),
            # Строгий «Заказано» — ТОЛЬКО статус ordered (без contracted). w_ordered выше
            # (contracted+ordered) НЕ трогаем — он используется в total_work, где заключённые
            # договоры обязаны входить в «Ведётся работа». Эта колонка — только для widget.ordered,
            # чтобы карточка «Заказано» на дашборде означала то же самое, что и total_ordered.
            func.coalesce(func.sum(
                case(
                    (Purchase.status == "ordered",
                     func.coalesce(Purchase.contract_price, Purchase.planned_total_price)),
                    else_=None
                )
            ), 0).label("w_ordered_strict"),
            func.coalesce(func.sum(
                case(
                    (Purchase.status == "delivered",
                     func.coalesce(Purchase.contract_price, Purchase.planned_total_price)),
                    else_=None
                )
            ), 0).label("w_delivered"),
            func.coalesce(func.sum(
                case(
                    (Purchase.status == "paid",
                     func.coalesce(Purchase.payment_amount, Purchase.contract_price, Purchase.planned_total_price)),
                    else_=None
                )
            ), 0).label("w_paid"),
            # Per-subsidy basket counts/amounts for per-subsidy widget
            func.coalesce(func.sum(
                case((Purchase.status == "plan_schedule", Purchase.planned_total_price), else_=None)
            ), 0).label("sp_amt"),
            func.coalesce(func.count(
                case((Purchase.status == "plan_schedule", Purchase.id), else_=None)
            ), 0).label("sp_cnt"),
            func.coalesce(func.count(
                case((Purchase.status == "work_in_progress", Purchase.id), else_=None)
            ), 0).label("sw_cnt"),
            func.coalesce(func.count(
                case((Purchase.status.in_(["contracted", "ordered"]), Purchase.id), else_=None)
            ), 0).label("so_cnt"),
            func.coalesce(func.count(
                case((Purchase.status == "ordered", Purchase.id), else_=None)
            ), 0).label("so_cnt_strict"),
            func.coalesce(func.count(
                case((Purchase.status == "delivered", Purchase.id), else_=None)
            ), 0).label("sd_cnt"),
            func.coalesce(func.count(
                case((Purchase.status == "paid", Purchase.id), else_=None)
            ), 0).label("spd_cnt"),
        )
        .select_from(Subsidy)
        .outerjoin(Purchase, Purchase.subsidy_id == Subsidy.id)
        .group_by(Subsidy.id, Subsidy.name, Subsidy.year, Subsidy.budget)
        .order_by(Subsidy.year.desc(), Subsidy.name)
    )
    if use_sids:
        # Двухуровневая видимость: орг-дефолт + пер-субсидийный override (managed=«Субсидии», dashboard=«Дашборд»).
        if visible_subsidy_ids is not None:
            subsidy_q = subsidy_q.where(Subsidy.id.in_(visible_subsidy_ids))
    elif org_ids is not None:
        subsidy_q = subsidy_q.where(Subsidy.org_id.in_(org_ids))
    subsidy_result = await db.execute(subsidy_q)

    # FEO planned sum per subsidy
    feo_planned_q = (
        select(
            FeoCategory.subsidy_id,
            func.coalesce(func.sum(FeoPlannedItem.amount), 0).label("feo_planned_sum"),
        )
        .join(FeoPlannedItem, FeoPlannedItem.feo_category_id == FeoCategory.id)
        .where(FeoPlannedItem.is_active == True)
        .group_by(FeoCategory.subsidy_id)
    )
    if use_sids:
        if visible_subsidy_ids is not None:
            feo_planned_q = feo_planned_q.where(FeoCategory.subsidy_id.in_(visible_subsidy_ids))
    elif org_ids is not None:
        feo_planned_q = feo_planned_q.where(FeoCategory.subsidy_id.in_(
            select(Subsidy.id).where(Subsidy.org_id.in_(org_ids))
        ))
    feo_planned_result = await db.execute(feo_planned_q)
    feo_planned_map: dict[int, float] = {
        row.subsidy_id: float(row.feo_planned_sum)
        for row in feo_planned_result
    }

    subsidy_rows = subsidy_result.all()
    sid_list = [row.id for row in subsidy_rows]
    # Batch вместо N+1: бюджеты, субсидии и контрагенты одним IN-запросом каждый
    budgets = await calculate_budgets_bulk(db, sid_list)
    # Phase 31-05: canonical spent + planned_amounts for D-17 (единый источник)
    spent_map = await _calculate_spent_bulk(db, sid_list)
    planned_amounts_map = await _calculate_planned_amounts_bulk(db, sid_list)
    # Единый источник «Запланировано»: план дерева ФЭО = ручные позиции + позиции из заявок плана закупок.
    # Совпадает с KPI «Запланировано» на вкладке «Субсидии» (selectedPlannedTotal).
    planned_tree_map = await _calculate_feo_planned_tree_bulk(db, sid_list)
    # Владелец (2026-08-30): предупреждение «сумма заказанного приближается к
    # потолку субсидии» — батчем на весь список (см. app/services/feo_plan.py
    # calculate_ceiling_forecasts_bulk).
    ceiling_forecasts = await calculate_ceiling_forecasts_bulk(db, sid_list)
    sub_objs = {}
    if subsidy_rows:
        sub_objs = {
            s.id: s for s in (await db.execute(
                select(Subsidy).where(Subsidy.id.in_(sid_list))
            )).scalars().all()
        }
    contractor_ids = {s.contractor_id for s in sub_objs.values() if s.contractor_id}
    contractors = {}
    if contractor_ids:
        contractors = {
            c.id: c for c in (await db.execute(
                select(Contractor).where(Contractor.id.in_(contractor_ids))
            )).scalars().all()
        }

    # Виджет «Заключено договоров»: active-договоры, сумма по типу.
    # Правило РАЗНОЕ для двух типов (правка 2026-08-04 — рамочные с суммой не требуют закупок):
    #   single → max_amount договора, ТОЛЬКО если EXISTS хотя бы одна привязанная закупка
    #     (purchases.contract_id) с нужным статусом. Разовый договор подписывается ПОД
    #     конкретную закупку — без неё запись осиротевшая и не должна раздувать виджет
    #     (было замечено на ЦентрПоиск_2026: 53 млн вместо фактических 402 тыс., см. баг-репорт 2026-08).
    #   framework_with_amount → max_amount договора БЕЗУСЛОВНО, просто при status='active'.
    #     Рамочный договор с суммой подписывается заранее, закупки по нему делаются постепенно
    #     позже — поэтому он уже «заключён», даже если ни одной закупки по нему ещё не заведено.
    #   framework_cumulative → SUM(COALESCE(p.contract_price, p.planned_total_price))
    #     по закупкам с contract_id=договор и status in (contracted,ordered,delivered,paid)
    _contracted_purchase_exists = (
        select(literal(1))
        .where(Purchase.contract_id == Contract.id)
        .where(Purchase.status.in_(["contracted", "ordered", "delivered", "paid"]))
    )
    contract_single_q = (
        select(
            Contract.subsidy_id,
            func.coalesce(func.sum(Contract.max_amount), 0).label("amt"),
            func.count(Contract.id).label("cnt"),
        )
        .where(Contract.status == "active")
        .where(
            or_(
                and_(Contract.contract_type == "single", _contracted_purchase_exists.exists()),
                Contract.contract_type == "framework_with_amount",
            )
        )
        .group_by(Contract.subsidy_id)
    )
    if use_sids:
        if visible_subsidy_ids is not None:
            contract_single_q = contract_single_q.where(Contract.subsidy_id.in_(visible_subsidy_ids))
    elif org_ids is not None:
        contract_single_q = contract_single_q.where(Contract.subsidy_id.in_(
            select(Subsidy.id).where(Subsidy.org_id.in_(org_ids))
        ))
    cs_rows = (await db.execute(contract_single_q)).all()

    # framework_cumulative: aggr по закупкам привязанным к таким договорам
    contract_fc_q = (
        select(
            Purchase.subsidy_id,
            func.coalesce(func.sum(
                func.coalesce(Purchase.contract_price, Purchase.planned_total_price)
            ), 0).label("amt"),
            # count distinct contracts (not purchases)
            func.count(func.distinct(Purchase.contract_id)).label("cnt"),
        )
        .join(Contract, Purchase.contract_id == Contract.id)
        .where(Contract.status == "active")
        .where(Contract.contract_type == "framework_cumulative")
        .where(Purchase.status.in_(["contracted", "ordered", "delivered", "paid"]))
        .group_by(Purchase.subsidy_id)
    )
    if use_sids:
        if visible_subsidy_ids is not None:
            contract_fc_q = contract_fc_q.where(Purchase.subsidy_id.in_(visible_subsidy_ids))
    elif org_ids is not None:
        contract_fc_q = contract_fc_q.where(Purchase.subsidy_id.in_(
            select(Subsidy.id).where(Subsidy.org_id.in_(org_ids))
        ))
    cfc_rows = (await db.execute(contract_fc_q)).all()

    # Собрать contracts_map: subsidy_id → float (сумма по обеим частям)
    contracts_map: dict[int, float] = {}
    contracts_cnt_map: dict[int, int] = {}
    for r in cs_rows:
        sid = r.subsidy_id
        contracts_map[sid] = contracts_map.get(sid, 0.0) + float(r.amt)
        contracts_cnt_map[sid] = contracts_cnt_map.get(sid, 0) + int(r.cnt)
    for r in cfc_rows:
        sid = r.subsidy_id
        contracts_map[sid] = contracts_map.get(sid, 0.0) + float(r.amt)
        contracts_cnt_map[sid] = contracts_cnt_map.get(sid, 0) + int(r.cnt)

    # Глобальные итоги — сумма по регруппированным строкам (бит-в-бит эквивалентно прежнему)
    contracts_amt = sum(contracts_map.values())
    contracts_cnt = sum(contracts_cnt_map.values())

    # Накопительный SUM(planned_monthly) по active-договорам, регруппированный по subsidy_id
    mp_q = (
        select(
            Contract.subsidy_id,
            func.coalesce(func.sum(Contract.planned_monthly), 0).label("amt"),
        )
        .where(Contract.status == "active")
        .where(Contract.planned_monthly.isnot(None))
        .group_by(Contract.subsidy_id)
    )
    if use_sids:
        if visible_subsidy_ids is not None:
            mp_q = mp_q.where(Contract.subsidy_id.in_(visible_subsidy_ids))
    elif org_ids is not None:
        mp_q = mp_q.where(Contract.subsidy_id.in_(
            select(Subsidy.id).where(Subsidy.org_id.in_(org_ids))
        ))
    mp_map: dict = {}
    for r in (await db.execute(mp_q)).all():
        mp_map[r.subsidy_id] = mp_map.get(r.subsidy_id, 0.0) + float(r.amt)
    # Глобальный итог — сумма по регруппированным строкам (бит-в-бит эквивалентно прежнему скаляру,
    # включая договоры с subsidy_id IS NULL — они попадают в mp_map под ключом None)
    monthly_payments_total = float(sum(mp_map.values()))

    # ── Ежемесячные закупки: начисление «Заказано» по прошедшим месяцам ──────────────
    # is_monthly_payment=true закупки не имеют помесячного графика дат — период берём из
    # service_start_date/service_end_date/service_deadline_date/contract_date. Правило
    # пользователя: услуга оказывается по 31 мая, сегодня 1 июня → июнь уже считается
    # заказанным (обязательство по месяцу уже возникло, раз услуга в нём оказывается).
    # SQL-агрегат total_ordered выше исключает is_monthly_payment (Purchase.is_monthly_payment.isnot(True)) —
    # поэтому их вклад в «Заказано» приходит РОВНО ОДИН РАЗ, отсюда, начислением. Ниже он
    # СКЛАДЫВАЕТСЯ с SQL-агрегатом в итоговое поле total_ordered (не остаётся сбоку) — так
    # «Заказано» реально включает то, что уже оказывается по ежемесячному договору в этом месяце.
    # Применяется только начиная со статуса «Договор заключён» — без договора обязательства нет.
    from app.services.plan_cashflow import add_months as _add_months

    def _calendar_months(start: date, end: date) -> int:
        """Кол-во календарных месяцев от месяца start до месяца end включительно.

        Считает по месяцу/году, день не учитывается (сверено с примером пользователя:
        услуга по 31 мая, сегодня 1 июня → результат уже включает июнь). 0, если end раньше start.
        """
        start_m = date(start.year, start.month, 1)
        end_m = date(end.year, end.month, 1)
        if end_m < start_m:
            return 0
        k = 0
        cur = start_m
        while cur <= end_m:
            k += 1
            cur = _add_months(start_m, k)
        return k

    monthly_q = (
        select(
            Purchase.subsidy_id, Purchase.contract_price, Purchase.planned_total_price,
            Purchase.monthly_payment_count, Purchase.monthly_payment_amount,
            Purchase.service_start_date, Purchase.service_end_date,
            Purchase.service_deadline_date, Purchase.contract_date,
        )
        .where(Purchase.is_monthly_payment == True)
        .where(Purchase.status.in_(["contracted", "ordered", "delivered", "paid"]))
    )
    if use_sids:
        monthly_q = _apply_purchase_org_filter(monthly_q, current_user, subsidy_ids=visible_subsidy_ids)
    else:
        monthly_q = _apply_purchase_org_filter(monthly_q, current_user, org_ids)
    monthly_rows = (await db.execute(monthly_q)).all()

    monthly_ordered_map: dict[int, float] = {}
    _today = date.today()
    for r in monthly_rows:
        start = r.service_start_date or r.contract_date
        if not start:
            continue  # нет точки отсчёта — начисление не делаем

        months_elapsed = _calendar_months(start, _today)
        if months_elapsed <= 0:
            continue

        count_cap = r.monthly_payment_count
        if not count_cap:
            period_end = r.service_end_date or r.service_deadline_date
            derived = _calendar_months(start, period_end) if period_end else 0
            count_cap = derived or None  # None = потолка по месяцам нет

        months_to_pay = min(months_elapsed, count_cap) if count_cap else months_elapsed
        if months_to_pay <= 0:
            continue

        amount_per_month = r.monthly_payment_amount
        if amount_per_month is None:
            contract_total = r.contract_price if r.contract_price is not None else r.planned_total_price
            if contract_total is not None and r.monthly_payment_count:
                amount_per_month = Decimal(contract_total) / Decimal(r.monthly_payment_count)
            else:
                continue  # ни суммы платежа, ни способа её вывести — начисление не делаем

        accrued = Decimal(amount_per_month) * Decimal(months_to_pay)

        contract_total = r.contract_price if r.contract_price is not None else r.planned_total_price
        if contract_total is not None:
            accrued = min(accrued, Decimal(contract_total))  # итог не превышает сумму договора/плана

        if r.subsidy_id is not None:
            monthly_ordered_map[r.subsidy_id] = monthly_ordered_map.get(r.subsidy_id, 0.0) + float(accrued)

    subsidy_stats = []
    for row in subsidy_rows:
        calc = budgets.get(row.id, 0.0)
        effective_budget = calc if calc > 0 else float(row.budget or 0)
        spent = spent_map.get(row.id, 0.0)
        planned_amt = planned_amounts_map.get(row.id, 0.0)
        # planned_tree = правильное «Запланировано» = план дерева ФЭО (= «Свободно» = budget − planned_tree)
        planned_tree = planned_tree_map.get(row.id, 0.0)
        # «Свободно» = budget − planned_tree (совпадает с панелью ФЭО вкладки «Субсидии»)
        remaining = effective_budget - planned_tree
        # discrepancy-чип показывается только при превышении (planned_tree > budget)
        discrepancy = (effective_budget - planned_tree) if planned_tree > effective_budget else None
        sub_obj = sub_objs.get(row.id)
        contractor_name = None
        contractor_inn = None
        if sub_obj and sub_obj.contractor_id:
            contractor = contractors.get(sub_obj.contractor_id)
            if contractor:
                contractor_name = contractor.name
                contractor_inn = contractor.inn
        subsidy_stats.append({
            "id": row.id,
            "name": row.name,
            "year": row.year,
            "budget": float(row.budget),
            "calculated_budget": effective_budget,
            "total_planned": float(row.total_planned),
            "total_confirmed": float(row.total_confirmed),
            "total_paid": float(row.total_paid),
            "total_plan_schedule": float(row.total_plan_schedule),  # SUM work_in_progress planned_total_price
            # total_ordered = SQL-агрегат по НЕежемесячным (row.total_ordered) + начисление по
            # ежемесячным (monthly_ordered_map) — обязаны складываться в одно число: карточка
            # «Заказано» должна включать месяц, за который услуга уже оказывается прямо сейчас.
            "total_ordered": float(row.total_ordered) + monthly_ordered_map.get(row.id, 0.0),
            # Справочно: какая часть total_ordered выше — из помесячного начисления (для отладки/подписи).
            # УЖЕ ВХОДИТ в total_ordered, не складывать повторно.
            "monthly_ordered_accrued": monthly_ordered_map.get(row.id, 0.0),
            "total_feo_planned": feo_planned_map.get(row.id, 0.0),  # NEW 12-01: SUM FeoPlannedItem.amount
            "planned_tree": planned_tree,  # единый источник: план дерева ФЭО (ручные + заявки)
            "feo_budget_total": effective_budget,
            "feo_filled": calc > 0,
            "contractor_id": sub_obj.contractor_id if sub_obj else None,
            "contractor_name": contractor_name,
            "contractor_inn": contractor_inn,
            # Phase 31-05: canonical budget fields (D-17)
            "remaining": remaining,  # = «Свободно» = budget − planned_tree
            "planned_amount": planned_amt,
            "budget_discrepancy": discrepancy,
            # Per-subsidy work/delivery baskets (зеркала глобальных widgets)
            "total_work": float(row.total_plan_schedule) + float(row.w_ordered) + float(row.w_delivered) + float(row.w_paid),
            "total_contracts": contracts_map.get(row.id, 0.0),
            "total_delivered": float(row.w_delivered) + float(row.w_paid),
            "total_delivered_unpaid": float(row.w_delivered),
            # Per-subsidy widget basket (зеркало формул глобального widgets dict)
            "widget": {
                "plan_schedule": {
                    "amount": float(row.sp_amt) + float(row.total_plan_schedule) + float(row.w_ordered) + float(row.w_delivered) + float(row.w_paid),
                    "count": int(row.sp_cnt) + int(row.sw_cnt) + int(row.so_cnt) + int(row.sd_cnt) + int(row.spd_cnt),
                },
                "work": {
                    "amount": float(row.total_plan_schedule) + float(row.w_ordered) + float(row.w_delivered) + float(row.w_paid),
                    "count": int(row.sw_cnt) + int(row.so_cnt) + int(row.sd_cnt) + int(row.spd_cnt),
                },
                "ordered": {
                    # Строго status='ordered' (без 'contracted') — согласовано с total_ordered/
                    # глобальным widgets["ordered"] (правка 2026-08-04). w_ordered (contracted+ordered)
                    # остаётся в work/plan_schedule/total_work выше — там 'заключён договор' должен входить.
                    # + начисление по ежемесячным (monthly_ordered_map) — та же сумма, что вошла
                    # в total_ordered выше, widget.ordered обязан быть согласован с ним.
                    "amount": float(row.w_ordered_strict) + float(row.w_delivered) + float(row.w_paid)
                        + monthly_ordered_map.get(row.id, 0.0),
                    "count": int(row.so_cnt_strict) + int(row.sd_cnt) + int(row.spd_cnt),
                    "monthly_payments_total": mp_map.get(row.id, 0.0),
                },
                "delivered": {
                    "amount": float(row.w_delivered) + float(row.w_paid),
                    "count": int(row.sd_cnt) + int(row.spd_cnt),
                },
                "delivered_unpaid": {"amount": float(row.w_delivered), "count": int(row.sd_cnt)},
                "paid": {"amount": float(row.w_paid), "count": int(row.spd_cnt)},
                "contracts": {
                    "amount": contracts_map.get(row.id, 0.0),
                    "count": contracts_cnt_map.get(row.id, 0),
                },
            },
        })
        # Владелец (2026-08-30): плашка «субсидии у потолка» — прокидываем расчёт
        # прямо в subsidy_stats[i], чтобы карточка/список субсидий на дашборде
        # могли показать предупреждение без второго запроса.
        subsidy_stats[-1].update(ceiling_forecasts.get(row.id, {}))

    # ── Накопительные виджеты закупок ────────────────────────────────────────
    # Корзины: каждая закупка ровно в одной, по текущему статусу.
    # Исключаем cancelled и wishes.
    basket_q = (
        select(
            # stage_plan  : plan_schedule → planned_total_price
            func.coalesce(func.sum(
                case((Purchase.status == "plan_schedule", Purchase.planned_total_price), else_=None)
            ), 0).label("sp_amt"),
            func.coalesce(func.count(
                case((Purchase.status == "plan_schedule", Purchase.id), else_=None)
            ), 0).label("sp_cnt"),
            # stage_work  : work_in_progress → planned_total_price
            func.coalesce(func.sum(
                case((Purchase.status == "work_in_progress", Purchase.planned_total_price), else_=None)
            ), 0).label("sw_amt"),
            func.coalesce(func.count(
                case((Purchase.status == "work_in_progress", Purchase.id), else_=None)
            ), 0).label("sw_cnt"),
            # stage_ordered : contracted|ordered → COALESCE(contract_price, planned_total_price)
            func.coalesce(func.sum(
                case(
                    (Purchase.status.in_(["contracted", "ordered"]),
                     func.coalesce(Purchase.contract_price, Purchase.planned_total_price)),
                    else_=None
                )
            ), 0).label("so_amt"),
            func.coalesce(func.count(
                case((Purchase.status.in_(["contracted", "ordered"]), Purchase.id), else_=None)
            ), 0).label("so_cnt"),
            # stage_ordered_strict : ТОЛЬКО ordered (без contracted) — для widget «Заказано»,
            # чтобы означать то же самое, что и total_ordered. stage_ordered (so_amt/so_cnt) выше
            # остаётся как было — используется в widget «work», где contracted обязан входить.
            func.coalesce(func.sum(
                case(
                    (Purchase.status == "ordered",
                     func.coalesce(Purchase.contract_price, Purchase.planned_total_price)),
                    else_=None
                )
            ), 0).label("so_amt_strict"),
            func.coalesce(func.count(
                case((Purchase.status == "ordered", Purchase.id), else_=None)
            ), 0).label("so_cnt_strict"),
            # stage_delivered_unpaid : delivered → COALESCE(contract_price, planned_total_price)
            func.coalesce(func.sum(
                case(
                    (Purchase.status == "delivered",
                     func.coalesce(Purchase.contract_price, Purchase.planned_total_price)),
                    else_=None
                )
            ), 0).label("sd_amt"),
            func.coalesce(func.count(
                case((Purchase.status == "delivered", Purchase.id), else_=None)
            ), 0).label("sd_cnt"),
            # stage_paid : paid → COALESCE(payment_amount, contract_price, planned_total_price)
            func.coalesce(func.sum(
                case(
                    (Purchase.status == "paid",
                     func.coalesce(Purchase.payment_amount, Purchase.contract_price, Purchase.planned_total_price)),
                    else_=None
                )
            ), 0).label("spd_amt"),
            func.coalesce(func.count(
                case((Purchase.status == "paid", Purchase.id), else_=None)
            ), 0).label("spd_cnt"),
        )
        .where(Purchase.status.notin_(["cancelled", "wishes"]))
    )
    if use_sids:
        basket_q = _apply_purchase_org_filter(basket_q, current_user, subsidy_ids=visible_subsidy_ids)
    else:
        basket_q = _apply_purchase_org_filter(basket_q, current_user, org_ids)
    basket_row = (await db.execute(basket_q)).one()

    sp_amt  = float(basket_row.sp_amt);  sp_cnt  = int(basket_row.sp_cnt)
    sw_amt  = float(basket_row.sw_amt);  sw_cnt  = int(basket_row.sw_cnt)
    so_amt  = float(basket_row.so_amt);  so_cnt  = int(basket_row.so_cnt)
    so_amt_strict = float(basket_row.so_amt_strict); so_cnt_strict = int(basket_row.so_cnt_strict)
    sd_amt  = float(basket_row.sd_amt);  sd_cnt  = int(basket_row.sd_cnt)
    spd_amt = float(basket_row.spd_amt); spd_cnt = int(basket_row.spd_cnt)
    # Глобальный итог начисления по ежемесячным — СУММА уже посчитанных per-subsidy начислений
    # (monthly_ordered_map), а не отдельный пересчёт: monthly_q использует тот же use_sids/
    # visible_subsidy_ids/org_ids фильтр видимости, что и basket_q выше — числа согласованы.
    monthly_ordered_total = float(sum(monthly_ordered_map.values()))

    # ── Собираем виджеты ─────────────────────────────────────────────────────
    widgets = {
        # накопительно: stage_plan + stage_work + stage_ordered + stage_delivered_unpaid + stage_paid
        "plan_schedule": {
            "amount": sp_amt + sw_amt + so_amt + sd_amt + spd_amt,
            "count":  sp_cnt + sw_cnt + so_cnt + sd_cnt + spd_cnt,
        },
        # накопительно: stage_work + stage_ordered + stage_delivered_unpaid + stage_paid
        # (stage_ordered = contracted+ordered — «заключён договор» уже означает, что работа ведётся)
        "work": {
            "amount": sw_amt + so_amt + sd_amt + spd_amt,
            "count":  sw_cnt + so_cnt + sd_cnt + spd_cnt,
        },
        # накопительно: stage_ordered_strict (ТОЛЬКО status='ordered') + stage_delivered_unpaid + stage_paid
        # + начисление по ежемесячным (monthly_ordered_total). Согласовано с total_ordered
        # (правка 2026-08-04) — раньше здесь была contracted+ordered (so_amt) БЕЗ начисления,
        # из-за чего карточка «Заказано» на дашборде расходилась с той же меткой на
        # вкладке «Субсидии» (total_ordered). НЕ используем so_amt/so_cnt здесь — те остаются
        # в widget «work» выше, где contracted обязан входить.
        "ordered": {
            "amount": so_amt_strict + sd_amt + spd_amt + monthly_ordered_total,
            "count":  so_cnt_strict + sd_cnt + spd_cnt,
            "monthly_payments_total": monthly_payments_total,
        },
        # накопительно: stage_delivered_unpaid + stage_paid
        "delivered": {
            "amount": sd_amt + spd_amt,
            "count":  sd_cnt + spd_cnt,
        },
        # только stage_delivered_unpaid (поставлено, но не оплачено)
        "delivered_unpaid": {
            "amount": sd_amt,
            "count":  sd_cnt,
        },
        # только stage_paid
        "paid": {
            "amount": spd_amt,
            "count":  spd_cnt,
        },
        # договоры: single+framework_with_amount → max_amount;
        # framework_cumulative → SUM(COALESCE(contract_price, planned_total_price)) по закупкам
        "contracts": {
            "amount": contracts_amt,
            "count":  contracts_cnt,
        },
    }

    # Владелец (2026-08-30): блок «субсидии у потолка» — субсидии, где сумма
    # заказанного (включая ежемесячные — весь график) достигла/превысила
    # настроенный порог (ceiling_warn_percent, умолчание 90%). Отсортировано
    # по убыванию процента — сначала самые критичные (превышенные), затем
    # ближе всего к порогу.
    subsidies_near_ceiling = sorted(
        (
            {
                "subsidy_id": s["id"], "name": s["name"],
                "ceiling_total": s.get("ceiling_total", 0.0),
                "ceiling_committed_total": s.get("ceiling_committed_total", 0.0),
                "ceiling_committed_percent": s.get("ceiling_committed_percent", 0.0),
                "ceiling_warn_percent": s.get("ceiling_warn_percent", 90.0),
                "ceiling_exceeded": s.get("ceiling_exceeded", False),
            }
            for s in subsidy_stats
            if s.get("ceiling_near_warning") or s.get("ceiling_exceeded")
        ),
        key=lambda x: x["ceiling_committed_percent"], reverse=True,
    )

    return {
        "status_counts": status_counts,
        "subsidy_stats": subsidy_stats,
        "widgets": widgets,
        "subsidies_near_ceiling": subsidies_near_ceiling,
    }


@router.get("/analytics")
async def analytics(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    subsidy_ids: Optional[str] = Query(None),
    scope: Optional[str] = Query(None),
):
    today = date.today()
    org_ids = get_org_filter(current_user)
    dash_sids = _UNSET
    if scope == "dashboard":
        dash_sids = await get_visible_subsidy_ids(current_user, db, "dashboard")
    sid_filter = [int(x) for x in subsidy_ids.split(",") if x.strip()] if subsidy_ids else None

    def _pf(q):
        """Apply purchase org/dashboard + subsidy filter inline."""
        if dash_sids is not _UNSET:
            if dash_sids is not None:
                q = q.where(Purchase.subsidy_id.in_(dash_sids))
        elif org_ids is not None:
            q = q.where(Purchase.subsidy_id.in_(
                select(Subsidy.id).where(Subsidy.org_id.in_(org_ids))
            ))
        if sid_filter:
            q = q.where(Purchase.subsidy_id.in_(sid_filter))
        return q

    # 1. Purchase funnel
    STATUS_ORDER = ["wishes", "plan_schedule", "work_in_progress", "contracted", "delivered", "paid"]
    funnel_result = await db.execute(_pf(
        select(Purchase.status, func.count(Purchase.id).label("cnt"),
               func.coalesce(func.sum(Purchase.planned_total_price), 0).label("total"))
        .group_by(Purchase.status)
    ))
    funnel_raw = {r.status: {"count": r.cnt, "total": float(r.total)} for r in funnel_result}
    funnel = [{"status": s, "count": funnel_raw.get(s, {}).get("count", 0),
               "total": funnel_raw.get(s, {}).get("total", 0.0)} for s in STATUS_ORDER]

    # 2. Monthly paid amounts (last 12 months)
    monthly_result = await db.execute(_pf(
        select(
            extract("year", Purchase.payment_doc_date).label("y"),
            extract("month", Purchase.payment_doc_date).label("m"),
            func.coalesce(func.sum(Purchase.payment_amount), 0).label("total"),
        )
        .where(Purchase.payment_doc_date != None)
        .group_by("y", "m")
        .order_by("y", "m")
    ))
    monthly = [{"year": int(r.y), "month": int(r.m), "total": float(r.total)} for r in monthly_result]

    # 3. Top 10 contractors by total purchase value
    top_result = await db.execute(_pf(
        select(
            Contractor.name,
            func.count(Purchase.id).label("cnt"),
            func.coalesce(func.sum(Purchase.planned_total_price), 0).label("total"),
        )
        .join(Contractor, Purchase.contractor_id == Contractor.id)
        .group_by(Contractor.name)
        .order_by(func.sum(Purchase.planned_total_price).desc())
        .limit(10)
    ))
    top_contractors = [{"name": r.name, "count": r.cnt, "total": float(r.total)} for r in top_result]

    # 4. Upcoming deliveries (next 30 days)
    upcoming_result = await db.execute(_pf(
        select(func.count(Purchase.id).label("cnt"),
               func.coalesce(func.sum(Purchase.planned_total_price), 0).label("total"))
        .where(Purchase.delivery_date.between(today, today + timedelta(days=30)))
    ))
    upcoming = upcoming_result.one()

    # 5. Economy (saved money)
    economy_result = await db.execute(_pf(
        select(func.coalesce(func.sum(Purchase.planned_total_price - Purchase.contract_price), 0))
        .where(Purchase.contract_price != None)
        .where(Purchase.contract_price > 0)
    ))
    economy = float(economy_result.scalar() or 0)

    # 6. Overdue purchases (execution_term past, not paid/delivered)
    overdue_result = await db.execute(_pf(
        select(func.count(Purchase.id))
        .where(Purchase.execution_term < today)
        .where(Purchase.status.notin_(["paid", "delivered"]))
    ))
    overdue_count = overdue_result.scalar() or 0

    # 7. Upcoming deadlines (next 30 days) — detailed list
    deadlines_result = await db.execute(_pf(
        select(Purchase.id, Purchase.subject, Purchase.item_name,
               Purchase.purchase_number, Purchase.execution_term, Purchase.status)
        .where(Purchase.execution_term.between(today, today + timedelta(days=30)))
        .where(Purchase.status.notin_(["paid", "delivered"]))
        .order_by(Purchase.execution_term)
        .limit(20)
    ))
    upcoming_deadlines = [
        {"id": r.id, "name": r.subject or r.item_name or f"Закупка #{r.purchase_number or r.id}",
         "purchase_number": r.purchase_number, "execution_term": str(r.execution_term), "status": r.status}
        for r in deadlines_result
    ]

    # 8. Purchase method distribution
    method_result = await db.execute(_pf(
        select(Purchase.purchase_method, func.count(Purchase.id).label("cnt"))
        .group_by(Purchase.purchase_method)
    ))
    method_distribution = {(r.purchase_method or "unknown"): r.cnt for r in method_result}

    # 9. Plan vs Fact per subsidy
    pf_q = (
        select(
            Subsidy.name,
            func.coalesce(func.sum(Purchase.planned_total_price), 0).label("plan"),
            func.coalesce(func.sum(
                case((Purchase.status.in_(["contracted", "delivered", "paid"]), Purchase.contract_price), else_=None)
            ), 0).label("contracted"),
            func.coalesce(func.sum(
                case((Purchase.status == "paid", Purchase.payment_amount), else_=None)
            ), 0).label("paid"),
        )
        .select_from(Subsidy)
        .outerjoin(Purchase, Purchase.subsidy_id == Subsidy.id)
        .group_by(Subsidy.id, Subsidy.name)
        .order_by(Subsidy.name)
    )
    if dash_sids is not _UNSET:
        if dash_sids is not None:
            pf_q = pf_q.where(Subsidy.id.in_(dash_sids))
    elif org_ids is not None:
        pf_q = pf_q.where(Subsidy.org_id.in_(org_ids))
    if sid_filter:
        pf_q = pf_q.where(Subsidy.id.in_(sid_filter))
    pf_result = await db.execute(pf_q)
    plan_fact = [
        {"subsidy": r.name, "plan": float(r.plan), "contracted": float(r.contracted), "paid": float(r.paid)}
        for r in pf_result
    ]

    return {
        "funnel": funnel,
        "monthly_payments": monthly,
        "top_contractors": top_contractors,
        "upcoming_deliveries": {"count": upcoming.cnt, "total": float(upcoming.total)},
        "economy": economy,
        "overdue_count": overdue_count,
        "upcoming_deadlines": upcoming_deadlines,
        "method_distribution": method_distribution,
        "plan_fact": plan_fact,
    }


@router.get("/financial-plan")
async def get_financial_plan(
    subsidy_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    scope: Optional[str] = Query(None),
):
    """Возвращает помесячную и поквартальную разбивку ожидаемых выплат.

    Структура ответа:
    - plan / committed — текущий план/обязательства по obligation_date
    - overdue[M] — накопленный долг: закупка с obligation_date < M и не полностью оплачена
      добавляется в каждый месяц от (obligation_month+1) до текущего включительно
    - no_deadline — закупки без obligation_date (не оплачены)
    """
    org_ids = get_org_filter(current_user)
    dash_sids = _UNSET
    if scope in ("dashboard", "plan"):
        dash_sids = await get_visible_subsidy_ids(current_user, db, scope)
    q = select(Purchase)
    q = _apply_purchase_org_filter(q, current_user, org_ids, subsidy_ids=dash_sids)
    if subsidy_id:
        q = q.where(Purchase.subsidy_id == subsidy_id)
    rows = (await db.execute(q)).scalars().all()

    PLAN_STATUSES = {'planned', 'wishes', 'plan_schedule'}
    COMMITTED_STATUSES = {'contracted', 'ordered', 'delivered', 'work_in_progress'}
    # paid — учитывается через _expected_payment_date, в overdue не попадает

    today = date.today()
    current_ym = f"{today.year}-{today.month:02d}"

    # buckets
    by_month_plan: dict = {}        # ym -> {amount, count}
    by_month_committed: dict = {}   # ym -> {amount, count}
    by_month_overdue: dict = {}     # ym -> {accumulated, items_count}
    no_deadline_amount = 0.0
    no_deadline_count = 0

    by_quarter_plan: dict = {}
    by_quarter_committed: dict = {}
    by_quarter_overdue: dict = {}

    def _add_bucket(bkt: dict, key: str, amount: float, count: int = 1):
        if key not in bkt:
            bkt[key] = {"amount": 0.0, "count": 0}
        bkt[key]["amount"] += amount
        bkt[key]["count"] += count

    def _add_overdue(bkt: dict, key: str, amount: float):
        if key not in bkt:
            bkt[key] = {"accumulated": 0.0, "items_count": 0}
        bkt[key]["accumulated"] += amount
        bkt[key]["items_count"] += 1

    for p in rows:
        amount = float(p.contract_price or p.planned_total_price or 0)
        if amount == 0:
            continue

        paid_amount = float(p.delivery_payment_amount or 0)
        is_fully_paid = paid_amount >= amount

        obl = obligation_date(p)

        # paid закупки — используем _expected_payment_date (по payment_doc_date)
        if p.status == 'paid':
            d = _expected_payment_date(p)
            if d:
                ym = f"{d.year}-{d.month:02d}"
                yq = f"{d.year}-Q{(d.month - 1) // 3 + 1}"
                _add_bucket(by_month_committed, ym, amount)
                _add_bucket(by_quarter_committed, yq, amount)
            continue

        # no_deadline bucket
        if obl is None:
            no_deadline_amount += amount
            no_deadline_count += 1
            continue

        obl_ym = f"{obl.year}-{obl.month:02d}"
        obl_yq = f"{obl.year}-Q{(obl.month - 1) // 3 + 1}"

        # overdue: obligation_date < first_day_of_current_month and not fully paid
        # Distribute to every month from obl_month+1 up to current month
        if not is_fully_paid:
            # compute first day of month after obligation month
            if obl.month == 12:
                next_year, next_month = obl.year + 1, 1
            else:
                next_year, next_month = obl.year, obl.month + 1
            overdue_start = date(next_year, next_month, 1)
            first_day_current = date(today.year, today.month, 1)

            if overdue_start <= first_day_current:
                # Distribute the remaining amount to each month from overdue_start to today (inclusive)
                remaining = amount - paid_amount
                cur = overdue_start
                while cur <= first_day_current:
                    m_ym = f"{cur.year}-{cur.month:02d}"
                    m_yq = f"{cur.year}-Q{(cur.month - 1) // 3 + 1}"
                    _add_overdue(by_month_overdue, m_ym, remaining)
                    _add_overdue(by_quarter_overdue, m_yq, remaining)
                    # advance month
                    if cur.month == 12:
                        cur = date(cur.year + 1, 1, 1)
                    else:
                        cur = date(cur.year, cur.month + 1, 1)
                # Also add to plan/committed for original period
                if p.status in PLAN_STATUSES:
                    _add_bucket(by_month_plan, obl_ym, amount)
                    _add_bucket(by_quarter_plan, obl_yq, amount)
                elif p.status in COMMITTED_STATUSES:
                    _add_bucket(by_month_committed, obl_ym, amount)
                    _add_bucket(by_quarter_committed, obl_yq, amount)
                continue

        # Future or not-overdue — add to plan/committed normally
        if p.status in PLAN_STATUSES:
            _add_bucket(by_month_plan, obl_ym, amount)
            _add_bucket(by_quarter_plan, obl_yq, amount)
        elif p.status in COMMITTED_STATUSES:
            _add_bucket(by_month_committed, obl_ym, amount)
            _add_bucket(by_quarter_committed, obl_yq, amount)

    def _fmt_plan(d: dict):
        return [{"period": k, "amount": v["amount"], "count": v["count"]} for k, v in sorted(d.items())]

    def _fmt_overdue(d: dict):
        return [{"period": k, "accumulated": v["accumulated"], "items_count": v["items_count"]}
                for k, v in sorted(d.items())]

    # --- feo_plan: real ФЭО cash-flow from FeoPlannedItem schedule ---
    from app.services.plan_cashflow import cashflow_for_subsidies

    # Determine subsidy id set for feo_plan aggregation
    if subsidy_id:
        feo_sids = [subsidy_id]
    elif dash_sids is not _UNSET:
        # dash_sids is either None (superadmin — no restriction) or a list
        if dash_sids is None:
            # superadmin: all subsidies, filtered by org if applicable
            _sid_q = select(Subsidy.id)
            if org_ids is not None:
                _sid_q = _sid_q.where(Subsidy.org_id.in_(org_ids))
            feo_sids = [r[0] for r in (await db.execute(_sid_q)).all()]
        else:
            feo_sids = list(dash_sids)
    else:
        # scope not in dashboard/plan — fall back to org filter
        _sid_q = select(Subsidy.id)
        if org_ids is not None:
            _sid_q = _sid_q.where(Subsidy.org_id.in_(org_ids))
        feo_sids = [r[0] for r in (await db.execute(_sid_q)).all()]

    by_month_feo: dict[str, float] = {}
    by_quarter_feo: dict[str, float] = {}
    if feo_sids:
        cf = await cashflow_for_subsidies(db, feo_sids)
        for sub_data in cf.values():
            for (y, m), amt in sub_data.items():
                mk = f"{y}-{m:02d}"
                qk = f"{y}-Q{(m - 1) // 3 + 1}"
                by_month_feo[mk] = by_month_feo.get(mk, 0.0) + float(amt)
                by_quarter_feo[qk] = by_quarter_feo.get(qk, 0.0) + float(amt)

    def _fmt_feo(d: dict):
        return [{"period": k, "amount": v} for k, v in sorted(d.items())]

    return {
        "by_month": {
            "plan":        _fmt_plan(by_month_plan),
            "committed":   _fmt_plan(by_month_committed),
            "overdue":     _fmt_overdue(by_month_overdue),
            "no_deadline": {"amount": no_deadline_amount, "items_count": no_deadline_count},
            "feo_plan":    _fmt_feo(by_month_feo),
        },
        "by_quarter": {
            "plan":        _fmt_plan(by_quarter_plan),
            "committed":   _fmt_plan(by_quarter_committed),
            "overdue":     _fmt_overdue(by_quarter_overdue),
            "no_deadline": {"amount": no_deadline_amount, "items_count": no_deadline_count},
            "feo_plan":    _fmt_feo(by_quarter_feo),
        },
    }


@router.get("/financial-plan/details")
async def get_financial_plan_details(
    period: Optional[str] = Query(None, description="'YYYY-MM' для месяца или 'YYYY-Qn' для квартала. Не требуется для category=no_deadline"),
    category: str = Query(..., regex="^(plan|committed|overdue|no_deadline)$"),
    granularity: str = Query("month", regex="^(month|quarter)$"),
    subsidy_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    scope: Optional[str] = Query(None),
):
    """Список закупок попавших в указанный период+категорию.

    category:
    - plan / committed — плановые/принятые обязательства по obligation_date
    - overdue — просроченные (obligation_date в прошлом, не полностью оплачены)
    - no_deadline — без срока исполнения (obligation_date is None)
    """
    PLAN_STATUSES = {"planned", "wishes", "plan_schedule"}
    COMMITTED_STATUSES = {"contracted", "ordered", "delivered", "paid", "work_in_progress"}

    today = date.today()
    org_ids = get_org_filter(current_user)
    dash_sids = _UNSET
    if scope in ("dashboard", "plan"):
        dash_sids = await get_visible_subsidy_ids(current_user, db, scope)

    # period обязателен для всех категорий кроме no_deadline
    if category != "no_deadline" and not period:
        from fastapi import HTTPException
        raise HTTPException(status_code=422, detail="period обязателен для category != 'no_deadline'")

    if category == "no_deadline":
        # Загружаем все не-paid закупки без obligation_date
        q = select(Purchase).where(Purchase.status != "paid")
        if subsidy_id:
            q = q.where(Purchase.subsidy_id == subsidy_id)
        q = _apply_purchase_org_filter(q, current_user, org_ids, subsidy_ids=dash_sids)
        q = q.options(selectinload(Purchase.contractor))
        rows = (await db.execute(q)).scalars().all()

        result = []
        for p in rows:
            obl = obligation_date(p)
            if obl is not None:
                continue
            amount = float(p.contract_price or p.planned_total_price or 0)
            if amount == 0:
                continue
            paid_amount = float(p.delivery_payment_amount or 0)
            result.append({
                "id": p.id,
                "purchase_number": p.purchase_number,
                "subject": p.subject or "",
                "contractor_name": (p.contractor.name if p.contractor else None) or "—",
                "amount": amount,
                "status": p.status,
                "expected_date": None,
                "contract_number": p.contract_number,
                "obligation_date": None,
                "is_likely_needed": p.is_likely_needed if hasattr(p, 'is_likely_needed') else True,
                "is_prepayment": p.is_prepayment if hasattr(p, 'is_prepayment') else False,
                "is_overdue": False,
                "missing_deadline": True,
                "paid_amount": paid_amount,
                "remaining": max(amount - paid_amount, 0),
                "stage_label": p.stage_label if hasattr(p, 'stage_label') else None,
            })
        result.sort(key=lambda r: (-r["amount"],))
        return {"period": period, "category": category, "granularity": granularity, "items": result}

    elif category == "overdue":
        # Закупки у которых obligation_date < first_day_of_period и не полностью оплачены
        q = select(Purchase).where(Purchase.status != "paid")
        if subsidy_id:
            q = q.where(Purchase.subsidy_id == subsidy_id)
        q = _apply_purchase_org_filter(q, current_user, org_ids, subsidy_ids=dash_sids)
        q = q.options(selectinload(Purchase.contractor))
        rows = (await db.execute(q)).scalars().all()

        # Parse period to get first day of period
        if granularity == "month":
            year, month = int(period[:4]), int(period[5:7])
            first_day_of_period = date(year, month, 1)
        else:
            year = int(period[:4])
            q_num = int(period[-1])
            first_month = (q_num - 1) * 3 + 1
            first_day_of_period = date(year, first_month, 1)

        result = []
        for p in rows:
            obl = obligation_date(p)
            if obl is None:
                continue
            amount = float(p.contract_price or p.planned_total_price or 0)
            if amount == 0:
                continue
            paid_amount = float(p.delivery_payment_amount or 0)
            is_fully_paid = paid_amount >= amount
            if is_fully_paid:
                continue
            # obligation must be before first_day_of_period
            if obl >= first_day_of_period:
                continue
            # check that period <= current (overdue only goes up to today)
            if first_day_of_period > date(today.year, today.month, 1):
                continue
            remaining = max(amount - paid_amount, 0)
            result.append({
                "id": p.id,
                "purchase_number": p.purchase_number,
                "subject": p.subject or "",
                "contractor_name": (p.contractor.name if p.contractor else None) or "—",
                "amount": amount,
                "status": p.status,
                "expected_date": obl.isoformat(),
                "contract_number": p.contract_number,
                "obligation_date": obl.isoformat(),
                "is_likely_needed": p.is_likely_needed if hasattr(p, 'is_likely_needed') else True,
                "is_prepayment": p.is_prepayment if hasattr(p, 'is_prepayment') else False,
                "is_overdue": True,
                "missing_deadline": False,
                "paid_amount": paid_amount,
                "remaining": remaining,
                "stage_label": p.stage_label if hasattr(p, 'stage_label') else None,
            })
        result.sort(key=lambda r: (r["obligation_date"], -r["amount"]))
        return {"period": period, "category": category, "granularity": granularity, "items": result}

    else:
        # plan / committed — используем obligation_date
        target_statuses = PLAN_STATUSES if category == "plan" else COMMITTED_STATUSES

        q = select(Purchase).where(Purchase.status.in_(target_statuses))
        if subsidy_id:
            q = q.where(Purchase.subsidy_id == subsidy_id)
        q = _apply_purchase_org_filter(q, current_user, org_ids, subsidy_ids=dash_sids)
        q = q.options(selectinload(Purchase.contractor))
        rows = (await db.execute(q)).scalars().all()

        result = []
        for p in rows:
            if p.status == 'paid':
                d = _expected_payment_date(p)
            else:
                d = obligation_date(p)
            if not d:
                continue
            if granularity == "month":
                row_period = f"{d.year}-{d.month:02d}"
            else:
                row_period = f"{d.year}-Q{(d.month - 1) // 3 + 1}"
            if row_period != period:
                continue

            amount = float(p.contract_price or p.planned_total_price or 0)
            if amount == 0:
                continue
            paid_amount = float(p.delivery_payment_amount or 0)
            obl = obligation_date(p)
            is_overdue_flag = (obl is not None and obl < today and paid_amount < amount)

            result.append({
                "id": p.id,
                "purchase_number": p.purchase_number,
                "subject": p.subject or "",
                "contractor_name": (p.contractor.name if p.contractor else None) or "—",
                "amount": amount,
                "status": p.status,
                "expected_date": d.isoformat(),
                "contract_number": p.contract_number,
                "obligation_date": obl.isoformat() if obl else None,
                "is_likely_needed": p.is_likely_needed if hasattr(p, 'is_likely_needed') else True,
                "is_prepayment": p.is_prepayment if hasattr(p, 'is_prepayment') else False,
                "is_overdue": is_overdue_flag,
                "missing_deadline": obl is None,
                "paid_amount": paid_amount,
                "remaining": max(amount - paid_amount, 0),
                "stage_label": p.stage_label if hasattr(p, 'stage_label') else None,
            })

        result.sort(key=lambda r: (r["expected_date"], -r["amount"]))
        return {"period": period, "category": category, "granularity": granularity, "items": result}


@router.get("/financial-plan/export.xlsx")
async def export_financial_plan_xlsx(
    granularity: str = Query("month", regex="^(month|quarter)$"),
    subsidy_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    scope: Optional[str] = Query(None),
):
    """Полная выгрузка всех закупок с группировкой по периоду и категории plan/committed."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse
    from datetime import datetime as dt

    org_ids = get_org_filter(current_user)
    dash_sids = _UNSET
    if scope in ("dashboard", "plan"):
        dash_sids = await get_visible_subsidy_ids(current_user, db, scope)

    PLAN_STATUSES = {"planned", "wishes", "plan_schedule"}
    COMMITTED_STATUSES = {"contracted", "ordered", "delivered", "paid", "work_in_progress"}

    q = select(Purchase).where(Purchase.status.in_(PLAN_STATUSES | COMMITTED_STATUSES))
    if subsidy_id:
        q = q.where(Purchase.subsidy_id == subsidy_id)
    q = _apply_purchase_org_filter(q, current_user, org_ids, subsidy_ids=dash_sids)
    q = q.options(selectinload(Purchase.contractor))

    rows = (await db.execute(q)).scalars().all()

    STATUS_LABELS = {
        "planned": "Запланирован", "wishes": "Заявка",
        "plan_schedule": "План закупок",
        "contracted": "Заключён договор", "ordered": "Заказано", "delivered": "Поставлено",
        "paid": "Оплачено", "work_in_progress": "Ведётся работа",
    }

    grouped: dict = {}
    for p in rows:
        d = _expected_payment_date(p)
        if not d:
            continue
        amount = float(p.contract_price or p.planned_total_price or 0)
        if amount == 0:
            continue
        if granularity == "month":
            period_key = f"{d.year}-{d.month:02d}"
        else:
            period_key = f"{d.year}-Q{(d.month - 1) // 3 + 1}"
        category = "plan" if p.status in PLAN_STATUSES else "committed"
        grouped.setdefault((period_key, category), []).append({
            "purchase_number": p.purchase_number or "",
            "subject": p.subject or "",
            "contractor_name": (p.contractor.name if p.contractor else "—"),
            "amount": amount,
            "status": STATUS_LABELS.get(p.status, p.status),
            "expected_date": d.isoformat(),
            "contract_number": p.contract_number or "",
        })

    wb = Workbook()
    ws = wb.active
    ws.title = "Финплан"

    ws['A1'] = f"Финансовый план — {'по месяцам' if granularity == 'month' else 'по кварталам'}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:H1')

    headers = ["Период", "Категория", "№ закупки", "Предмет", "Контрагент", "Дата", "Статус", "Сумма, ₽"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    row_idx = 4
    period_totals: dict = {}

    for (period_key, category) in sorted(grouped.keys()):
        items = grouped[(period_key, category)]
        cat_label = "Плановые" if category == "plan" else "Принятые обязательства"
        cat_color = "FEF3C7" if category == "plan" else "D1FAE5"

        for item in items:
            ws.cell(row=row_idx, column=1, value=period_key)
            ws.cell(row=row_idx, column=2, value=cat_label)
            ws.cell(row=row_idx, column=3, value=item["purchase_number"])
            ws.cell(row=row_idx, column=4, value=item["subject"])
            ws.cell(row=row_idx, column=5, value=item["contractor_name"])
            ws.cell(row=row_idx, column=6, value=item["expected_date"])
            ws.cell(row=row_idx, column=7, value=item["status"])
            ws.cell(row=row_idx, column=8, value=item["amount"])
            for col in range(1, 9):
                ws.cell(row=row_idx, column=col).fill = PatternFill(
                    start_color=cat_color, end_color=cat_color, fill_type="solid"
                )
            ws.cell(row=row_idx, column=8).number_format = '#,##0.00 ₽'
            row_idx += 1

            period_totals.setdefault(period_key, {"plan": 0.0, "committed": 0.0})
            period_totals[period_key][category] += item["amount"]

    row_idx += 1
    totals_title = ws.cell(row=row_idx, column=1, value="ИТОГО ПО ПЕРИОДАМ")
    totals_title.font = Font(bold=True, size=12)
    row_idx += 1
    ws.cell(row=row_idx, column=1, value="Период").font = Font(bold=True)
    ws.cell(row=row_idx, column=2, value="Плановые").font = Font(bold=True)
    ws.cell(row=row_idx, column=3, value="Принятые").font = Font(bold=True)
    ws.cell(row=row_idx, column=4, value="Всего").font = Font(bold=True)
    row_idx += 1

    for pk in sorted(period_totals.keys()):
        t = period_totals[pk]
        ws.cell(row=row_idx, column=1, value=pk)
        plan_cell = ws.cell(row=row_idx, column=2, value=t["plan"])
        plan_cell.number_format = '#,##0.00 ₽'
        committed_cell = ws.cell(row=row_idx, column=3, value=t["committed"])
        committed_cell.number_format = '#,##0.00 ₽'
        total_cell = ws.cell(row=row_idx, column=4, value=t["plan"] + t["committed"])
        total_cell.number_format = '#,##0.00 ₽'
        row_idx += 1

    for col_letter, width in [('A', 12), ('B', 22), ('C', 12), ('D', 50), ('E', 30), ('F', 12), ('G', 18), ('H', 16)]:
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = f"Финплан_{granularity}_{dt.now().strftime('%Y-%m-%d')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename, safe='-_.~')}"},
    )


@router.get("/financial-plan/details/export.xlsx")
async def export_financial_plan_details_xlsx(
    period: str = Query(...),
    category: str = Query(..., regex="^(plan|committed)$"),
    granularity: str = Query("month", regex="^(month|quarter)$"),
    subsidy_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    scope: Optional[str] = Query(None),
):
    """Excel выгрузка одной группы (период+категория) — содержимое drill-down диалога."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from io import BytesIO
    from fastapi.responses import StreamingResponse

    org_ids = get_org_filter(current_user)
    dash_sids = _UNSET
    if scope in ("dashboard", "plan"):
        dash_sids = await get_visible_subsidy_ids(current_user, db, scope)

    PLAN_STATUSES = {"planned", "wishes", "plan_schedule"}
    COMMITTED_STATUSES = {"contracted", "ordered", "delivered", "paid", "work_in_progress"}
    target_statuses = PLAN_STATUSES if category == "plan" else COMMITTED_STATUSES

    q = select(Purchase).where(Purchase.status.in_(target_statuses))
    if subsidy_id:
        q = q.where(Purchase.subsidy_id == subsidy_id)
    q = _apply_purchase_org_filter(q, current_user, org_ids, subsidy_ids=dash_sids)
    q = q.options(selectinload(Purchase.contractor))

    rows = (await db.execute(q)).scalars().all()

    STATUS_LABELS = {
        "planned": "Запланирован", "wishes": "Заявка",
        "plan_schedule": "План закупок",
        "contracted": "Заключён договор", "ordered": "Заказано", "delivered": "Поставлено",
        "paid": "Оплачено", "work_in_progress": "Ведётся работа",
    }

    items = []
    for p in rows:
        d = _expected_payment_date(p)
        if not d:
            continue
        if granularity == "month":
            row_period = f"{d.year}-{d.month:02d}"
        else:
            row_period = f"{d.year}-Q{(d.month - 1) // 3 + 1}"
        if row_period != period:
            continue
        amount = float(p.contract_price or p.planned_total_price or 0)
        if amount == 0:
            continue
        items.append({
            "purchase_number": p.purchase_number or "",
            "subject": p.subject or "",
            "contractor_name": (p.contractor.name if p.contractor else "—"),
            "amount": amount,
            "status": STATUS_LABELS.get(p.status, p.status),
            "expected_date": d.isoformat(),
            "contract_number": p.contract_number or "",
        })

    items.sort(key=lambda r: (r["expected_date"], -r["amount"]))

    wb = Workbook()
    ws = wb.active
    cat_label = "Плановые" if category == "plan" else "Принятые обязательства"
    ws.title = f"{cat_label[:20]} {period}"

    ws['A1'] = f"Финплан — {cat_label} — {period}"
    ws['A1'].font = Font(size=14, bold=True)
    ws.merge_cells('A1:G1')

    headers = ["№", "Предмет", "Контрагент", "Дата", "Статус", "№ договора", "Сумма, ₽"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    total = 0.0
    for i, item in enumerate(items, start=4):
        ws.cell(row=i, column=1, value=item["purchase_number"])
        ws.cell(row=i, column=2, value=item["subject"])
        ws.cell(row=i, column=3, value=item["contractor_name"])
        ws.cell(row=i, column=4, value=item["expected_date"])
        ws.cell(row=i, column=5, value=item["status"])
        ws.cell(row=i, column=6, value=item["contract_number"])
        amt_cell = ws.cell(row=i, column=7, value=item["amount"])
        amt_cell.number_format = '#,##0.00 ₽'
        total += item["amount"]

    total_row = 4 + len(items) + 1
    ws.cell(row=total_row, column=1, value="ИТОГО").font = Font(bold=True)
    total_cell = ws.cell(row=total_row, column=7, value=total)
    total_cell.font = Font(bold=True)
    total_cell.number_format = '#,##0.00 ₽'
    total_color = "FEF3C7" if category == "plan" else "D1FAE5"
    total_cell.fill = PatternFill(start_color=total_color, end_color=total_color, fill_type="solid")

    for col_letter, width in [('A', 12), ('B', 50), ('C', 30), ('D', 12), ('E', 18), ('F', 16), ('G', 16)]:
        ws.column_dimensions[col_letter].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    filename = f"Финплан_{period}_{category}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename, safe='-_.~')}"},
    )

