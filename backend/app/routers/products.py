import os
import shutil
import logging
from urllib.parse import quote as _url_quote
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File, Body
from fastapi.responses import FileResponse, StreamingResponse, Response
from pydantic import BaseModel
from sqlalchemy import select
from sqlalchemy.orm import defer
from sqlalchemy.ext.asyncio import AsyncSession
from app.auth.jwt import get_current_user, get_org_filter, get_single_org_id, require_role, ADMIN_ROLES
from app.auth.visibility import get_visible_subsidy_ids
from app.auth.permissions import require_tab
from app.database import get_db
from app.models.product import Product
from app.models.user import User
from app.schemas.schemas import ProductCreate, ProductOut, ProductSummaryGroup, ProductSummaryItem
from app.services.product_matcher import bulk_match, SCORE_AUTO, SCORE_SUGGEST
from typing import List, Optional
from decimal import Decimal
from io import BytesIO

_log = logging.getLogger(__name__)
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
except ImportError:
    Workbook = None
    load_workbook = None

try:
    import xlrd as _xlrd
except ImportError:
    _xlrd = None


def _read_excel_rows(content: bytes, filename: str):
    """Read rows from xlsx or xls file. Returns list of tuples (best sheet)."""
    if filename.lower().endswith(".xls"):
        if _xlrd is None:
            raise HTTPException(500, "xlrd не установлен")
        wb = _xlrd.open_workbook(file_contents=content)
        # Pick the sheet with the most non-empty rows
        best_sheet = wb.sheet_by_index(0)
        best_count = sum(1 for i in range(best_sheet.nrows) if any(v for v in best_sheet.row_values(i)))
        for si in range(1, wb.nsheets):
            sh = wb.sheet_by_index(si)
            cnt = sum(1 for i in range(sh.nrows) if any(v for v in sh.row_values(i)))
            if cnt > best_count:
                best_count = cnt
                best_sheet = sh
        return [tuple(best_sheet.row_values(i)) for i in range(best_sheet.nrows)]
    else:
        if load_workbook is None:
            raise HTTPException(500, "openpyxl не установлен")
        wb = load_workbook(BytesIO(content), data_only=True)
        ws = wb.active
        return list(ws.iter_rows(values_only=True))

PRODUCT_UPLOAD_DIR = "/app/uploads/products"
ALLOWED_IMAGE_MIME = {"image/jpeg", "image/jpg", "image/png", "image/webp", "image/gif"}

router = APIRouter(prefix="/api/products", tags=["products"])

@router.get("/", response_model=List[ProductOut])
async def list_products(
    feo_category_id: Optional[int] = Query(None),
    category: Optional[str] = Query(None),
    is_active: Optional[bool] = Query(None),
    search: Optional[str] = Query(None, description="Полнотекстовый поиск по имени/описанию/типу"),
    limit: Optional[int] = Query(None, ge=1, le=10000, description="Ограничить кол-во результатов (фронт грузит весь каталог в пикер)"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Phase 17.1-08 perf: defer photo_data (bytea, up to 10MB per row).
    # The list endpoint must stay small (~1-2MB); clients fetch bytes via
    # GET /{product_id}/photo when has_photo=True.
    q = select(Product).options(defer(Product.photo_data))
    # Products are global — no org_id filter
    if feo_category_id is not None:
        q = q.where(Product.feo_category_id == feo_category_id)
    if category is not None:
        q = q.where(Product.category == category)
    if is_active is not None:
        q = q.where(Product.is_active == is_active)
    if search:
        from sqlalchemy import or_
        pattern = f"%{search}%"
        q = q.where(or_(
            Product.name.ilike(pattern),
            Product.description.ilike(pattern),
            Product.product_type.ilike(pattern),
            Product.category.ilike(pattern),
        ))
    q = q.order_by(Product.name)
    if limit is not None:
        q = q.limit(limit)
    result = await db.execute(q)
    products = result.scalars().all()

    # Скрыть контрактные цены для чужих организаций (если не shared)
    if current_user.role != "superadmin":
        user_org_id = current_user.org_id
        for p in products:
            if p.contract_org_id and p.contract_org_id != user_org_id and not p.price_shared:
                p.contract_price = None
                p.contract_number = None
                p.contract_date = None
                p.contract_org_id = None

    return products

@router.get("/summary", response_model=List[ProductSummaryGroup])
async def product_summary(
    subsidy_id: Optional[int] = Query(None),
    category: Optional[str] = Query(None),
    product_id: Optional[int] = Query(None),
    search: Optional[str] = Query(None),
    org_id: Optional[int] = Query(None),
    region: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    quarter: Optional[int] = Query(None, ge=1, le=4),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    _perm: User = Depends(require_tab('products.summary')),
):
    """Сводная по продукции — агрегация закупок по продуктам через все субсидии."""
    from app.models.purchase_item import PurchaseItem
    from app.models.purchase import Purchase
    from app.models.subsidy import Subsidy
    from app.models.organization import Organization
    from sqlalchemy.orm import joinedload
    from sqlalchemy import extract
    from datetime import date as _date

    q = (
        select(PurchaseItem)
        .join(Product, PurchaseItem.product_id == Product.id)
        .join(Purchase, PurchaseItem.purchase_id == Purchase.id)
        .join(Subsidy, Purchase.subsidy_id == Subsidy.id)
        .outerjoin(Organization, Subsidy.org_id == Organization.id)
        .where(PurchaseItem.product_id.isnot(None))
        .options(
            joinedload(PurchaseItem.product),
            joinedload(PurchaseItem.purchase).joinedload(Purchase.feo_category),
        )
    )

    if subsidy_id is not None:
        q = q.where(Purchase.subsidy_id == subsidy_id)
    if category is not None:
        q = q.where(Product.category == category)
    if product_id is not None:
        q = q.where(Product.id == product_id)
    if search:
        q = q.where(Product.name.ilike(f"%{search}%"))
    if org_id is not None:
        q = q.where(Subsidy.org_id == org_id)
    if region is not None:
        q = q.where(Purchase.region == region)
    if date_from is not None:
        try:
            df = _date.fromisoformat(date_from)
            q = q.where(Purchase.delivery_date >= df)
        except ValueError:
            pass
    if date_to is not None:
        try:
            dt = _date.fromisoformat(date_to)
            q = q.where(Purchase.delivery_date <= dt)
        except ValueError:
            pass
    if quarter is not None:
        q = q.where(extract("quarter", Purchase.delivery_date) == quarter)

    # Двухуровневая видимость по вкладке «Сводная по продукции».
    vis = await get_visible_subsidy_ids(current_user, db, "products.summary")
    if vis is not None:
        if not vis:
            return []
        q = q.where(Purchase.subsidy_id.in_(vis))

    # Also need subsidy name, org name, org_id — use add_columns
    q = q.add_columns(
        Subsidy.name.label("subsidy_name"),
        Organization.name.label("org_name"),
        Subsidy.org_id.label("s_org_id"),
    )
    q = q.order_by(Product.name, Subsidy.name)

    result = await db.execute(q)
    rows = result.unique().all()

    # Group by product
    from collections import defaultdict
    groups: dict[int, dict] = {}
    for row in rows:
        pi = row[0]       # PurchaseItem
        s_name = row[1]   # subsidy_name
        o_name = row[2]   # org_name
        s_org_id = row[3] # subsidy.org_id
        product = pi.product
        purchase = pi.purchase

        pid = product.id
        if pid not in groups:
            groups[pid] = {
                "product_id": pid,
                "product_name": product.name,
                "category": product.category,
                "product_type": product.product_type,
                "total_quantity": Decimal(0),
                "total_amount": Decimal(0),
                "purchase_count": 0,
                "items": [],
            }

        qty = pi.quantity or Decimal(0)
        amt = pi.total_price or pi.final_total or Decimal(0)
        groups[pid]["total_quantity"] += qty
        groups[pid]["total_amount"] += amt
        groups[pid]["purchase_count"] += 1
        groups[pid]["items"].append(ProductSummaryItem(
            purchase_id=purchase.id,
            subsidy_name=s_name or "",
            org_name=o_name,
            org_id=s_org_id,
            region=purchase.region,
            quantity=pi.quantity,
            unit=pi.unit,
            unit_price=pi.unit_price,
            total_price=pi.total_price or pi.final_total,
            status=purchase.status,
            delivery_date=purchase.delivery_date,
            delivery_address=purchase.delivery_address,
            procurement_planned_date=purchase.procurement_planned_date,
            purchase_method=purchase.purchase_method,
        ))

    return [ProductSummaryGroup(**g) for g in groups.values()]


@router.get("/photos/{filename}")
async def serve_product_photo_legacy(filename: str):
    """Legacy filesystem endpoint. Phase 17.1-08 moved photos to bytea in DB;
    any surviving files on the volume are still served, but most will 404 now
    (prod volume lost its contents). Frontends should prefer the bytea
    endpoint GET /api/products/{product_id}/photo via `has_photo`.
    """
    filepath = os.path.join(PRODUCT_UPLOAD_DIR, filename)
    if not os.path.exists(filepath):
        raise HTTPException(404, "Фото не найдено (хранение перенесено в БД)")
    return FileResponse(filepath)


@router.get("/{product_id}/photo")
async def serve_product_photo_bytea(
    product_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Return the photo bytes stored in products.photo_data.

    Phase 17.1-08 — canonical photo serving path. Returns 404 if the product
    has no cached bytes yet (frontend should fall back to the external
    photo_url in that case).

    No auth dep: <img src> cannot send Authorization headers, and the legacy
    /api/products/photos/{filename} endpoint was already unauthenticated.
    Photo bytes are not sensitive — the product list itself requires auth.
    """
    product = await db.get(Product, product_id)
    if not product or not product.photo_data:
        raise HTTPException(404, "Фото не найдено")
    return Response(
        content=product.photo_data,
        media_type=product.photo_mime or "image/jpeg",
        headers={"Cache-Control": "public, max-age=86400"},
    )


@router.delete("/{product_id}/photo")
async def delete_product_photo(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    _: User = Depends(get_current_user),
):
    """Clear the stored photo bytea + metadata + legacy local photo_url.

    Does NOT touch photo_link (the backup external URL the admin maintains
    as source of truth for re-download).
    """
    product = await db.get(Product, product_id)
    if not product:
        raise HTTPException(404, "Товар не найден")
    product.photo_data = None
    product.photo_mime = None
    product.photo_size = None
    if product.photo_url and product.photo_url.startswith('/api/products/photos/'):
        product.photo_url = None
    await db.commit()
    return {"status": "ok", "product_id": product_id}


# ---------------------------------------------------------------------------
# Token-based product matching (Phase 27.4-25)
# ---------------------------------------------------------------------------

class _MatchCandidate(BaseModel):
    product_id: int
    name: str
    price: Optional[float]
    score: float
    description: Optional[str] = None
    photo_url: Optional[str] = None
    item_type: Optional[str] = None
    category: Optional[str] = None


class _MatchResultItem(BaseModel):
    query: str
    status: str  # 'auto' | 'suggest' | 'create'
    candidates: List[_MatchCandidate]


class _MatchRequest(BaseModel):
    queries: List[str]
    limit: int = 3
    # Интерактивный набор в строке позиции: включает префиксное сопоставление
    # стемов (см. text_match._stem_hits), чтобы подсказки появлялись раньше
    # 6-го символа для слов длиннее 6 букв. По умолчанию выключено — пакетный
    # импорт/дедуп не должен становиться нечётким.
    prefix: bool = False


class _MatchResponse(BaseModel):
    results: List[_MatchResultItem]


@router.post("/match", response_model=_MatchResponse)
async def match_products(
    body: _MatchRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Score a list of product name queries against the catalog using token-based fuzzy matching.

    Returns top-k candidates per query with status: 'auto' (score>=0.95),
    'suggest' (0.60<=score<0.95), or 'create' (no match found).
    """
    org_id = get_single_org_id(current_user)
    # 27.4-29: photo_url из БД ИЛИ /api/products/{id}/photo если фото в bytea (photo_data)
    q = select(
        Product.id,
        Product.name,
        Product.price,
        Product.description,
        Product.photo_url,
        (Product.photo_data.isnot(None)).label('has_bytea_photo'),
        Product.product_type,
        Product.category,
    )
    if org_id:
        q = q.where((Product.org_id == org_id) | (Product.org_id.is_(None)))
    rows = (await db.execute(q)).all()
    catalog = [
        (
            r.id,
            r.name or '',
            float(r.price) if r.price is not None else None,
            r.description,
            r.photo_url or (f'/api/products/{r.id}/photo' if r.has_bytea_photo else None),
            r.product_type,
            r.category,
        )
        for r in rows
    ]

    # Потолок поднят с 10 до 200: инлайновый поиск в строке позиции (InlineProductMatch)
    # должен показывать весь список совпадений с прокруткой, а не top-3/top-10.
    # Пакетный импорт по-прежнему не шлёт limit явно и получает дефолт 3.
    top_k = max(1, min(body.limit, 200))
    results = bulk_match(body.queries, catalog, top_k=top_k, prefix_match=body.prefix)

    # Актуализация цены (владелец, 2026-08-29): bulk_match не трогаем (сигнатура
    # зафиксирована) — донабираем метаданные вторым проходом по product_id
    # уже полученных кандидатов, одним SELECT + один контекст на весь запрос.
    candidate_ids = {c.product_id for r in results for c in r.candidates}
    freshness_by_id: dict[int, dict] = {}
    meta_by_id: dict[int, Product] = {}
    if candidate_ids:
        meta_rows = (await db.execute(
            select(Product).options(defer(Product.photo_data)).where(Product.id.in_(candidate_ids))
        )).scalars().all()
        freshness_ctx = await load_freshness_context(db, org_id)
        for prod in meta_rows:
            meta_by_id[prod.id] = prod
            freshness_by_id[prod.id] = evaluate_freshness(prod, freshness_ctx)

    _log.info(
        "POST /api/products/match: %d queries, catalog_size=%d, "
        "auto=%d suggest=%d create=%d",
        len(body.queries),
        len(catalog),
        sum(1 for r in results if r.status == 'auto'),
        sum(1 for r in results if r.status == 'suggest'),
        sum(1 for r in results if r.status == 'create'),
    )

    return _MatchResponse(results=[
        _MatchResultItem(
            query=r.query,
            status=r.status,
            candidates=[
                _MatchCandidate(
                    product_id=c.product_id,
                    name=c.name,
                    price=c.price,
                    score=c.score,
                    description=c.description,
                    photo_url=c.photo_url,
                    item_type=c.item_type,
                    category=c.category,
                )
                for c in r.candidates
            ],
        )
        for r in results
    ])


@router.get("/{product_id}", response_model=ProductOut)
async def get_product(
    product_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Phase 17.1-08 perf: defer photo_data — detail endpoint doesn't return
    # bytes; frontend uses GET /{product_id}/photo for raw image.
    result = await db.execute(
        select(Product)
        .options(defer(Product.photo_data))
        .where(Product.id == product_id)
    )
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(status_code=404, detail="Product not found")
    return product

def _calc_price_from_links(links: list) -> float | None:
    """Average price using only links from the 3 most recent distinct dates."""
    if not links:
        return None
    # Get all distinct dates that have a price, sorted descending
    dated = [(l.get("collected_at") or "", l["price"]) for l in links if l.get("price") is not None]
    if not dated:
        return None
    top_dates = sorted({d for d, _ in dated if d}, reverse=True)[:3]
    # Include links with no date only if no dated links exist
    if top_dates:
        prices = [p for d, p in dated if d in top_dates]
    else:
        prices = [p for _, p in dated]
    return round(sum(prices) / len(prices), 2) if prices else None


def _apply_price_links(data: dict, target: object) -> None:
    """Auto-calculate price as average of price_links (3 most recent dates)."""
    price = _calc_price_from_links(data.get("price_links") or [])
    if price is not None:
        data["price"] = price


def _price_links_max_collected_at(links: list) -> Optional[str]:
    """Владелец 2026-08-29: source_ref/collected_at для source='monitoring' —
    самая свежая дата среди ссылок сравнения цен."""
    dates = [l.get("collected_at") for l in (links or []) if l.get("collected_at")]
    return max(dates) if dates else None

@router.post("/", response_model=ProductOut)
async def create_product(
    product: ProductCreate,
    force: bool = False,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # Phase 21.x: prevent catalog clutter — if a similar product already exists,
    # return 409 with the suggestion. Frontend asks the user "use existing?"
    # and either links to it or retries with ?force=true.
    if not force and (product.name or '').strip():
        from app.product_matcher import find_matching_product
        org_id = get_single_org_id(current_user) or current_user.org_id
        existing = await find_matching_product(
            db, product.name, org_id=org_id, threshold=0.7,
        )
        if existing is not None:
            raise HTTPException(
                status_code=409,
                detail={
                    "code": "duplicate_product",
                    "message": f'Похожий товар уже есть в каталоге: "{existing.name}"',
                    "existing": ProductOut.model_validate(existing).model_dump(mode='json'),
                },
            )

    data = product.model_dump()
    _apply_price_links(data, None)
    db_product = Product(**data)
    db.add(db_product)
    await db.commit()
    await db.refresh(db_product)
    return db_product

@router.put("/{product_id}", response_model=ProductOut)
async def update_product(
    product_id: int,
    product: ProductCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    db_product = result.scalar_one_or_none()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    old_price = db_product.price
    data = product.model_dump()
    had_links = bool(data.get("price_links"))
    _apply_price_links(data, db_product)
    new_price = data.get("price")
    for key, value in data.items():
        setattr(db_product, key, value)
    from datetime import datetime
    db_product.updated_at = datetime.utcnow()
    db_product.updated_by = current_user.full_name or current_user.username

    # Актуализация цены (владелец, 2026-08-29): price изменился — записать
    # источник + историю. Если пришли price_links и из них посчиталась цена —
    # это автомониторинг ссылок ('monitoring'), иначе ручной ввод ('manual').
    if new_price is not None and new_price != old_price:
        if had_links and _calc_price_from_links(data.get("price_links") or []) is not None:
            collected = _price_links_max_collected_at(data.get("price_links") or [])
            await actualize_product_price(
                db, db_product, price=new_price, source="monitoring",
                collected_at=collected, user=current_user,
            )
        else:
            await actualize_product_price(
                db, db_product, price=new_price, source="manual", user=current_user,
            )

    await db.commit()
    await db.refresh(db_product)
    return db_product

@router.patch("/{product_id}", response_model=ProductOut)
async def patch_product(
    product_id: int,
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Частичное обновление товара (цена, ссылки, категория, вид, ед. изм.)."""
    result = await db.execute(select(Product).where(Product.id == product_id))
    db_product = result.scalar_one_or_none()
    if not db_product:
        raise HTTPException(404, "Product not found")
    if "unit" in data:
        db_product.unit = (data["unit"] or "").strip() or None
    if "category" in data:
        cat = (data["category"] or "").strip()
        if not cat:
            raise HTTPException(422, "Категория не может быть пустой")
        db_product.category = cat
    if "product_type" in data:
        pt = (data["product_type"] or "").strip()
        db_product.product_type = pt or None
    if "category" in data or "product_type" in data:
        from datetime import datetime
        db_product.updated_at = datetime.utcnow()
        db_product.updated_by = current_user.full_name or current_user.username
    old_price = db_product.price
    if "price_links" in data:
        db_product.price_links = data["price_links"]
        from sqlalchemy.orm.attributes import flag_modified
        flag_modified(db_product, "price_links")
        # Auto-update price from 3 most recent dates (if not explicitly provided)
        if "price" not in data:
            calc = _calc_price_from_links(data["price_links"])
            if calc is not None:
                new_price = Decimal(str(calc))
                if new_price != old_price:
                    # Актуализация цены (владелец, 2026-08-29): пересчёт из price_links
                    # — это автомониторинг ('monitoring'), collected_at — самая свежая
                    # дата среди ссылок.
                    collected = _price_links_max_collected_at(data["price_links"])
                    await actualize_product_price(
                        db, db_product, price=new_price, source="monitoring",
                        collected_at=collected, user=current_user,
                    )
                else:
                    db_product.price = new_price
    if "price" in data:
        if data["price"] is not None:
            # Явный ручной ввод цены ('manual') — актуализация с историей.
            await actualize_product_price(
                db, db_product, price=data["price"], source="manual", user=current_user,
            )
        else:
            db_product.price = None
    await db.commit()
    await db.refresh(db_product)
    return db_product


@router.patch("/{product_id}/share-price")
async def toggle_price_sharing(
    product_id: int,
    data: dict = Body(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Включить/отключить sharing контрактной цены для других организаций."""
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Товар не найден")
    # Только владелец организации или superadmin
    if current_user.role != "superadmin":
        if product.contract_org_id != current_user.org_id:
            raise HTTPException(403, "Только организация-владелец контракта может управлять доступом к цене")
    product.price_shared = bool(data.get("shared", False))
    await db.commit()
    await db.refresh(product)
    return product


@router.patch("/{product_id}/verify-tz", response_model=ProductOut)
async def verify_product_tz(
    product_id: int,
    tz_type: str = Query(..., description="'standard' или '44fz'"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Отметить ТЗ товара как проверенное (standard или 44fz)."""
    from datetime import datetime as _dt
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Товар не найден")
    verifier_name = current_user.full_name or current_user.username
    if tz_type == "44fz":
        product.tz_44fz_verified_at = _dt.utcnow()
        product.tz_44fz_verified_by = verifier_name
    else:
        product.tz_verified_at = _dt.utcnow()
        product.tz_verified_by = verifier_name
    await db.commit()
    await db.refresh(product)
    return product


@router.delete("/{product_id}/verify-tz", response_model=ProductOut)
async def unverify_product_tz(
    product_id: int,
    tz_type: str = Query(..., description="'standard' или '44fz'"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Снять отметку проверки ТЗ (только admin/superadmin)."""
    if current_user.role not in ("admin", "superadmin"):
        raise HTTPException(403, "Только администратор может снять отметку проверки ТЗ")
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Товар не найден")
    if tz_type == "44fz":
        product.tz_44fz_verified_at = None
        product.tz_44fz_verified_by = None
    else:
        product.tz_verified_at = None
        product.tz_verified_by = None
    await db.commit()
    await db.refresh(product)
    return product


@router.delete("/{product_id}")
async def delete_product(
    product_id: int,
    db: AsyncSession = Depends(get_db)
):
    result = await db.execute(select(Product).where(Product.id == product_id))
    db_product = result.scalar_one_or_none()
    if not db_product:
        raise HTTPException(status_code=404, detail="Product not found")
    await db.delete(db_product)
    await db.commit()
    return {"message": "Product deleted"}


@router.delete("/bulk/all")
async def delete_all_products(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Удалить все товары. Только для superadmin."""
    if current_user.role != "superadmin":
        raise HTTPException(403, "Только суперадмин может удалить все товары")
    from sqlalchemy import delete as sa_delete
    result = await db.execute(sa_delete(Product))
    await db.commit()
    return {"message": f"Удалено {result.rowcount} товаров"}


@router.get("/import/template")
async def download_products_template(
    _=Depends(get_current_user),
):
    """Шаблон Excel для импорта товаров."""
    if Workbook is None:
        raise HTTPException(500, "openpyxl не установлен")
    wb = Workbook()
    ws = wb.active
    ws.title = "Товары"
    headers = [
        "Наименование", "Описание", "Категория", "Вид", "Ед. изм.",
        "Цена", "Ссылка 1", "Цена ссылки 1", "Ссылка 2", "Цена ссылки 2", "Ссылка 3", "Цена ссылки 3",
        "Фото (URL)", "Многоразовое", "Активен", "Категория ФЭО",
    ]
    ws.append(headers)
    fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    font = Font(color="FFFFFF", bold=True, size=11)
    for cell in ws[1]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.append([
        "Компьютер Dell", "Core i5, 16GB RAM", "Оргтехника", "Рабочая станция", "шт",
        "85000", "https://market.yandex.ru/...", "83000", "https://dns-shop.ru/...", "87000", "", "",
        "", "да", "да", "Техническое оснащение",
    ])
    for i, w in enumerate([30, 30, 20, 20, 10, 12, 35, 14, 35, 14, 35, 14, 30, 12, 10, 30], 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w
    ws.freeze_panes = "A2"
    buf = BytesIO(); wb.save(buf); buf.seek(0)
    return StreamingResponse(buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_url_quote('Шаблон_импорта_товаров.xlsx', safe='-_.~')}"})


@router.post("/import")
async def import_products_from_excel(
    file: UploadFile = File(...),
    purchase_id: Optional[int] = Query(None, description="Если передан — добавить импортированные товары в закупку"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Импорт товаров из Excel. Возвращает {created, skipped, errors}."""
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Поддерживаются только .xlsx и .xls")

    content = await file.read()
    try:
        rows = _read_excel_rows(content, file.filename or "")
    except HTTPException:
        raise
    except Exception:
        raise HTTPException(400, "Не удалось прочитать файл. Убедитесь, что файл не повреждён.")
    if len(rows) < 2:
        raise HTTPException(400, "Файл пустой")

    # Auto-detect header row — scan ALL rows for the one with most recognizable column names
    NAME_HINTS = ('наименован', 'назван', 'товар', 'предмет', 'name', 'title', 'услуг', 'работ')
    ALL_HINTS = NAME_HINTS + ('цена', 'описан', 'кол', 'тип', 'price', 'стоимост', 'ед.', 'единиц', 'катег')
    header_row_idx = 0
    best_score = 0
    for ri, row in enumerate(rows):
        norm = [str(h).strip().lower() if h is not None else "" for h in row]
        score = sum(1 for h in norm if h and any(x in h for x in ALL_HINTS))
        if score > best_score:
            best_score = score
            header_row_idx = ri
    rows = rows[header_row_idx:]  # trim leading rows above header

    raw_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    COLUMN_MAP = {
        "наименование": "name",
        "название": "name",
        "наименование товара": "name",
        "товар": "name",
        "name": "name",
        "описание": "description",
        "description": "description",
        "категория": "category",
        "category": "category",
        "вид": "product_type",
        "тип": "product_type",
        "type": "product_type",
        "цена": "price",
        "цена, руб": "price",
        "цена, ₽": "price",
        "цена (руб)": "price",
        "стоимость": "price",
        "price": "price",
        "фото (url)": "photo_link",
        "фото": "photo_link",
        "photo": "photo_link",
        "ссылка на фото": "photo_link",
        "многоразовое": "is_reusable",
        "активен": "is_active",
        "активна": "is_active",
        "active": "is_active",
        "категория фэо": "feo_category_name",
        "фэо": "feo_category_name",
        "направление фэо": "feo_category_name",
    }
    # Also map Ссылка N / Цена ссылки N
    import re as _re
    col_idx: dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        # Exact match first
        field = COLUMN_MAP.get(h)
        if field and field not in col_idx:
            col_idx[field] = i
        # Fuzzy/partial match for name and common fields
        if 'name' not in col_idx and any(x in h for x in ('наименован', 'назван', 'товар', 'предмет', 'наимен')):
            col_idx['name'] = i
        elif 'description' not in col_idx and any(x in h for x in ('описан', 'техническ', 'характерист', 'specification')):
            col_idx['description'] = i
        elif 'price' not in col_idx and any(x in h for x in ('цена', 'стоимость', 'price')) and 'сумм' not in h:
            col_idx['price'] = i
        elif 'product_type' not in col_idx and any(x in h for x in ('тип', 'вид', 'type')):
            col_idx['product_type'] = i
        elif 'category' not in col_idx and 'категор' in h:
            col_idx['category'] = i
        elif 'quantity' not in col_idx and any(x in h for x in ('кол-во', 'количеств', 'qty', 'кол.')):
            col_idx['quantity'] = i
        elif 'unit' not in col_idx and any(x in h for x in ('ед.', 'ед. изм', 'единиц', 'unit')):
            col_idx['unit'] = i
        # Dynamic link columns
        m = _re.match(r"ссылка (\d+)$", h)
        if m:
            col_idx[f"link_url_{m.group(1)}"] = i
        m2 = _re.match(r"цена ссылки (\d+)$", h)
        if m2:
            col_idx[f"link_price_{m2.group(1)}"] = i

    # Post-map validation: if 'name' column contains numbers in data rows,
    # find the first string-heavy column instead (handles article+name dual-column files)
    if 'name' in col_idx and len(rows) > 1:
        name_col = col_idx['name']
        sample_vals = [rows[i][name_col] for i in range(1, min(4, len(rows))) if name_col < len(rows[i])]
        numeric_count = sum(1 for v in sample_vals if isinstance(v, (int, float)) and v == v)
        if numeric_count >= len(sample_vals) and sample_vals:
            # Mapped name column has only numbers — find the first string column
            for ci in range(len(rows[0])):
                if ci == name_col:
                    continue
                str_vals = [rows[i][ci] for i in range(1, min(4, len(rows))) if ci < len(rows[i])]
                if sum(1 for v in str_vals if isinstance(v, str) and len(v.strip()) > 5) >= len(str_vals) // 2 + 1:
                    col_idx['name'] = ci
                    break

    # FEO lookup
    from app.models.feo_category import FeoCategory
    feo_rows = (await db.execute(select(FeoCategory))).scalars().all()
    feo_by_name = {f.name.lower().strip(): f.id for f in feo_rows}

    def cell(row, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row): return None
        v = row[idx]; return str(v).strip() if v is not None else None

    def to_bool(v):
        if v is None: return True
        return str(v).lower().strip() in ("да", "yes", "true", "1", "+")

    def to_dec(v):
        if v is None: return None
        try: return Decimal(str(v).replace(" ", "").replace(",", "."))
        except: return None

    # Load existing products for dedup check (key → Product)
    def _norm_key(s) -> str:
        return (s or '').replace('\r\n', '\n').replace('\r', '\n').strip().lower()

    existing_result = await db.execute(select(Product))
    existing_by_key: dict[str, Product] = {}
    for ep in existing_result.scalars().all():
        k = _norm_key(ep.name) + '|' + _norm_key(ep.description)
        existing_by_key[k] = ep

    created = 0; skipped = 0; errors: list[dict] = []
    all_products: list[Product] = []   # both new and existing (for purchase items)
    product_row_data: list[dict] = []  # qty/unit per product for PurchaseItem

    from datetime import datetime as _dt
    _user_name = getattr(current_user, 'full_name', None) or getattr(current_user, 'username', '') or ''
    _import_note = (
        f"Импорт каталога из файла «{file.filename}», "
        f"{_user_name}, {_dt.now().strftime('%d.%m.%Y %H:%M')}"
    )

    for row_num, row in enumerate(rows[1:], start=2):
        try:
            name = cell(row, "name")
            if not name: continue  # empty row — don't count as skipped product
            desc_val = cell(row, "description")
            dedup_key = _norm_key(name) + '|' + _norm_key(desc_val)

            # Collect price_links
            price_links = []
            for n in range(1, 10):
                url = cell(row, f"link_url_{n}")
                if not url: break
                price_val = to_dec(cell(row, f"link_price_{n}"))
                price_links.append({"url": url, "price": float(price_val) if price_val else None})

            feo_name = cell(row, "feo_category_name")
            feo_id = feo_by_name.get(feo_name.lower().strip()) if feo_name else None

            price = to_dec(cell(row, "price"))
            if not price and price_links:
                prices = [l["price"] for l in price_links if l["price"]]
                if prices: price = Decimal(str(round(sum(prices) / len(prices), 2)))

            qty_str = cell(row, "quantity")
            unit_raw = cell(row, "unit")  # без дефолта — для бэкфилла Product.unit
            unit_str = unit_raw or "шт."
            row_qty = None
            if qty_str:
                try: row_qty = Decimal(str(qty_str).replace(',', '.').replace(' ', ''))
                except: pass

            if dedup_key in existing_by_key:
                # Product already in catalog — update price + backfill empty fields
                ep = existing_by_key[dedup_key]
                if price and ep.price != price:
                    # Актуализация цены (владелец, 2026-08-29): цена пришла из
                    # импортируемого Excel-файла — source='import'.
                    await actualize_product_price(
                        db, ep, price=price, source="import",
                        source_ref=file.filename, user=current_user,
                    )

                # Fill ONLY empty string-fields on existing product from this row
                def _fill(attr, val):
                    if val and not getattr(ep, attr):
                        setattr(ep, attr, val)

                # Категория: «Прочее» — дефолт, считаем пустым; заполненную в БД не трогаем (БД главнее)
                if cell(row, "category") and (not ep.category or ep.category == 'Прочее'):
                    ep.category = cell(row, "category")
                _fill("product_type", cell(row, "product_type"))
                _fill("photo_link", cell(row, "photo_link"))
                _fill("description", cell(row, "description"))

                ep.import_note = _import_note
                ep.updated_at = _dt.utcnow()
                ep.updated_by = _user_name

                # feo_category_id — numeric, set only if empty
                if feo_id and not ep.feo_category_id:
                    ep.feo_category_id = feo_id

                # price_links — list, fill only if existing empty and new non-empty
                if price_links and not ep.price_links:
                    from sqlalchemy.orm.attributes import flag_modified
                    ep.price_links = price_links
                    flag_modified(ep, "price_links")

                # Единица измерения (владелец, 2026-09-01): не трогаем уже
                # заполненную; иначе — из самого импорта, иначе — из истории
                # закупок этого товара (единственная встречавшаяся).
                await backfill_product_unit(db, ep, import_unit=unit_raw)

                all_products.append(ep)
                product_row_data.append({"qty": row_qty, "unit": unit_str, "price": price or ep.price})
                skipped += 1
                continue

            p = Product(
                name=name,
                description=cell(row, "description"),
                category=cell(row, "category"),
                product_type=cell(row, "product_type"),
                unit=(unit_raw or "").strip() or None,  # брэнд-новый товар — истории покупок ещё нет
                price=price,
                photo_link=cell(row, "photo_link"),
                is_reusable=to_bool(cell(row, "is_reusable")),
                is_active=to_bool(cell(row, "is_active")),
                feo_category_id=feo_id,
                price_links=price_links or [],
                import_note=_import_note,
                updated_at=_dt.utcnow(),
                updated_by=_user_name,
            )
            db.add(p)
            all_products.append(p)
            product_row_data.append({"qty": row_qty, "unit": unit_str, "price": price})
            created += 1
        except Exception as e:
            errors.append({"row": row_num, "name": cell(row, "name") or "?", "message": str(e)})

    # Flush to get product IDs
    await db.flush()

    product_ids: list[int] = [p.id for p in all_products]

    # If purchase_id provided — add ALL products (new + existing) as purchase items
    if purchase_id and all_products:
        from app.models.purchase_item import PurchaseItem
        for idx_p, p in enumerate(all_products):
            rd = product_row_data[idx_p] if idx_p < len(product_row_data) else {}
            qty = rd.get("qty") or Decimal('1')
            unit = rd.get("unit") or 'шт.'
            unit_price = rd.get("price") or p.price
            total = (qty * unit_price) if unit_price else None
            db.add(PurchaseItem(
                purchase_id=purchase_id,
                product_id=p.id,
                item_name=p.name[:500],
                item_type=p.product_type or 'товар',
                quantity=qty,
                unit=unit,
                unit_price=unit_price,
                total_price=total,
            ))

    await db.commit()
    recognized = {field: raw_headers[idx] for field, idx in col_idx.items()}
    return {"created": created, "skipped": skipped, "errors": errors,
            "product_ids": product_ids,
            "headers_found": recognized, "headers_raw": raw_headers[:20]}


@router.post("/deduplicate")
async def deduplicate_products(
    dry_run: bool = False,
    threshold: float = 0.8,
    skip_ids: Optional[str] = None,  # CSV: дубликаты, которые пользователь снял с галочкой
    db: AsyncSession = Depends(get_db),
    _=Depends(require_tab('products')),
):
    """Удалить дубликаты товаров. Группировка через fuzzy-матчинг по имени
    (token-set + char-ratio, то же что и при импорте). Единый общий каталог.

    `dry_run=true` — вернуть найденные группы без удаления (для preview UI).
    `threshold` — порог сходства (по умолчанию 0.8).
    `skip_ids` — CSV id'шников, которые пользователь снял с галочкой и НЕ хочет удалять.
    """
    from app.product_matcher import name_similarity, _normalize, _tokens
    from collections import defaultdict
    from difflib import SequenceMatcher
    from sqlalchemy import update as sa_update
    from app.models.purchase_item import PurchaseItem

    skipped_ids = set()
    if skip_ids:
        try:
            skipped_ids = {int(x) for x in skip_ids.split(',') if x.strip()}
        except ValueError:
            pass

    from sqlalchemy.orm import defer
    result = await db.execute(
        select(Product, Product.photo_data.isnot(None)).options(defer(Product.photo_data))
    )
    _rows = result.all()
    all_products: list[Product] = [r[0] for r in _rows]
    has_photo_blob: dict[int, bool] = {r[0].id: bool(r[1]) for r in _rows}

    def priority_score(p: Product) -> tuple:
        has_price_links = bool(p.price_links)
        has_photo = bool(has_photo_blob.get(p.id) or p.photo_url or p.photo_link)
        has_desc = bool((p.description or '').strip())
        contract_date_ts = p.contract_date.toordinal() if p.contract_date else 0
        return (has_price_links, has_photo, has_desc, contract_date_ts, p.id)

    duplicate_groups: list[dict] = []

    # Единый общий каталог — дедупликация по всей базе без деления по org_id
    org_products = all_products
    n = len(org_products)
    if n >= 2:
        # Union-find для транзитивного объединения
        parent = list(range(n))

        def find(x: int) -> int:
            while parent[x] != x:
                parent[x] = parent[parent[x]]
                x = parent[x]
            return x

        def union(a: int, b: int) -> None:
            ra, rb = find(a), find(b)
            if ra != rb:
                parent[ra] = rb

        # Дубликат = ТОЛЬКО точное совпадение имени (после нормализации регистра/
        # пробелов/ё). Похожие названия с РАЗНЫМИ характеристиками — это разные
        # товары (Гофрокороб д600/Т24 ≠ д500/Т22; ZenBook 14 i7 ≠ ZenBook 13 i5).
        # Никакого fuzzy (token-set / char-ratio): он ложно сливал разные SKU.
        norms = [_normalize(p.name or '') for p in org_products]
        by_norm: dict[str, int] = {}
        for i, key in enumerate(norms):
            if not key:
                continue
            prev = by_norm.get(key)
            if prev is not None:
                union(i, prev)
            else:
                by_norm[key] = i

        groups_map: dict[int, list[Product]] = defaultdict(list)
        for i, p in enumerate(org_products):
            groups_map[find(i)].append(p)

        for ps in groups_map.values():
            if len(ps) < 2:
                continue
            ps.sort(key=priority_score, reverse=True)
            winner = ps[0]
            dups = ps[1:]
            dup_dicts = []
            for p in dups:
                sim = name_similarity(winner.name or '', p.name or '')
                match = "exact" if sim >= 0.999 else "fuzzy"
                score = round(sim * 100)
                dup_dicts.append({
                    "id": p.id,
                    "name": p.name,
                    "category": p.category,
                    "product_type": p.product_type,
                    "has_photo": bool(has_photo_blob.get(p.id) or p.photo_url or p.photo_link),
                    "has_description": bool((p.description or '').strip()),
                    "score": score,
                    "match": match,
                })
            duplicate_groups.append({
                "winner": {
                    "id": winner.id,
                    "name": winner.name,
                    "category": winner.category,
                    "product_type": winner.product_type,
                    "has_photo": bool(has_photo_blob.get(winner.id) or winner.photo_url or winner.photo_link),
                    "has_description": bool((winner.description or '').strip()),
                },
                "duplicates": dup_dicts,
            })

    if dry_run:
        return {
            "dry_run": True,
            "groups": duplicate_groups,
            "total_groups": len(duplicate_groups),
            "total_to_delete": sum(len(g["duplicates"]) for g in duplicate_groups),
            "kept": len(all_products),
        }

    deleted = 0
    for grp in duplicate_groups:
        winner_id = grp["winner"]["id"]
        dup_ids = [dup["id"] for dup in grp["duplicates"] if dup["id"] not in skipped_ids]
        if not dup_ids:
            continue
        await db.execute(
            sa_update(PurchaseItem)
            .where(PurchaseItem.product_id.in_(dup_ids))
            .values(product_id=winner_id)
        )
        await db.execute(
            Product.__table__.delete().where(Product.id.in_(dup_ids))
        )
        deleted += len(dup_ids)

    await db.commit()
    return {
        "dry_run": False,
        "deleted": deleted,
        "kept": len(all_products) - deleted,
        "groups": duplicate_groups,
    }


def _fetch_photo_bytes(url: str) -> tuple[bytes, str]:
    """Download image from URL, return (raw_bytes, mime_type).

    Blocking — meant to be called via asyncio.to_thread. Converts webp to jpeg
    when possible. Enforces a 10MB size cap.
    """
    import urllib.request as _ur, io as _io
    SUPPORTED = ("image/jpeg", "image/jpg", "image/png", "image/gif", "image/bmp", "image/tiff", "image/webp")
    req = _ur.Request(url, headers={'User-Agent': 'Mozilla/5.0'})
    with _ur.urlopen(req, timeout=15) as r:
        ct = r.headers.get("Content-Type", "").split(";")[0].strip().lower() or "image/jpeg"
        raw = r.read()
    is_webp = "webp" in ct or url.lower().endswith(".webp")
    if is_webp:
        try:
            from PIL import Image as _Img
            img = _Img.open(_io.BytesIO(raw)).convert("RGB")
            buf = _io.BytesIO()
            img.save(buf, format="JPEG", quality=85)
            raw = buf.getvalue()
            ct = "image/jpeg"
        except Exception:
            # Pillow unavailable / broken webp — fall through, keep bytes + mime.
            ct = "image/webp"
    elif ct not in SUPPORTED:
        raise ValueError(f"Неподдерживаемый формат: {ct}")
    if len(raw) > 10 * 1024 * 1024:
        raise ValueError("Файл > 10MB")
    return raw, ct


def _pick_external_url(p: Product) -> Optional[str]:
    """Return the best external http(s) URL to download from.

    Priority: photo_url (if http/https) → photo_link (if http/https) → None.

    `photo_link` acts as the "backup external link" — admin maintains it as a
    source of truth, while `photo_url` can get overwritten with legacy local
    `/api/products/photos/...` paths or other non-http values. Local paths and
    any non-http values are treated as invalid sources and skipped.
    """
    def _is_http(v: Optional[str]) -> bool:
        return bool(v and (v.startswith("http://") or v.startswith("https://")))
    if _is_http(p.photo_url):
        return p.photo_url
    if _is_http(p.photo_link):
        return p.photo_link
    return None


async def _download_and_save_photo(product_id: int, url: str, db: AsyncSession) -> tuple[bool, Optional[str]]:
    """Download external URL and persist bytes to product.photo_data.

    Returns (success, error_msg). The existing `photo_url` field is NOT cleared —
    it remains the source of truth for re-downloading the photo later.
    """
    import asyncio
    try:
        raw, mime = await asyncio.to_thread(_fetch_photo_bytes, url)
    except Exception as e:
        return False, str(e)
    product = await db.get(Product, product_id)
    if not product:
        return False, "Товар не найден"
    product.photo_data = raw
    product.photo_mime = mime
    product.photo_size = len(raw)
    # Do NOT clear photo_url — external URL stays as source of truth.
    await db.commit()
    return True, None


@router.post("/download-photos")
async def download_all_photos(
    db: AsyncSession = Depends(get_db),
):
    """Скачать фото для всех активных товаров с внешней ссылкой, ещё не
    закэшированных в БД.

    Phase 17.1-08: cached copies live in `products.photo_data`. The correct
    guard is therefore "no photo_data yet" — NOT "no local filesystem URL"
    (old `/api/products/photos/*` URLs all point to a now-empty volume).
    """
    result = await db.execute(select(Product).where(Product.is_active == True))
    all_products = result.scalars().all()

    updated, skipped, errors = 0, 0, []
    for p in all_products:
        # Already cached in DB → skip (idempotent re-runs are cheap).
        if p.photo_data is not None:
            skipped += 1
            continue
        # Find source URL: prefer photo_url (external), fallback to photo_link.
        # Legacy local `/api/products/photos/...` paths are filtered out by the
        # http(s) prefix check inside _pick_external_url.
        src = _pick_external_url(p)
        if not src:
            skipped += 1
            continue
        ok, err = await _download_and_save_photo(p.id, src, db)
        if ok:
            updated += 1
        else:
            errors.append({"id": p.id, "name": p.name, "error": err})

    return {"updated": updated, "skipped": skipped, "errors": errors}


@router.post("/{product_id}/download-photo", response_model=ProductOut)
async def download_single_photo(
    product_id: int,
    db: AsyncSession = Depends(get_db),
):
    """Скачать фото одного товара по его внешней URL-ссылке в БД."""
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Товар не найден")
    src = _pick_external_url(product)
    if not src:
        raise HTTPException(400, "Нет внешней ссылки для скачивания")
    ok, err = await _download_and_save_photo(product.id, src, db)
    if not ok:
        raise HTTPException(500, f"Ошибка скачивания: {err}")
    await db.refresh(product)
    return product


@router.post("/{product_id}/photo", response_model=ProductOut)
async def upload_product_photo(
    product_id: int,
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
):
    """Загрузка фото товара пользователем. Phase 17.1-08 — сохраняем в БД
    (bytea), не в файловую систему."""
    result = await db.execute(select(Product).where(Product.id == product_id))
    product = result.scalar_one_or_none()
    if not product:
        raise HTTPException(404, "Product not found")
    if file.content_type not in ALLOWED_IMAGE_MIME:
        raise HTTPException(400, f"Недопустимый тип файла: {file.content_type}")
    raw = await file.read()
    if len(raw) > 10 * 1024 * 1024:
        raise HTTPException(400, "Файл > 10MB")
    product.photo_data = raw
    product.photo_mime = file.content_type or "image/jpeg"
    product.photo_size = len(raw)
    # photo_url is left untouched — user-uploaded photo doesn't need an
    # external source URL, but any pre-existing one stays as-is.
    await db.commit()
    await db.refresh(product)
    return product


# import-no-clutter: bulk-add purchase items to catalog
class _BulkFromItemsRequest(BaseModel):
    purchase_item_ids: List[int]


@router.post("/bulk-from-purchase-items")
async def bulk_create_from_items(
    body: _BulkFromItemsRequest,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Массовое создание Product из несвязанных PurchaseItem'ов.

    Для каждого PurchaseItem с product_id=None создаёт или находит Product
    через _upsert_product_to_catalog (идемпотентно).
    Обновляет item.product_id, match_confirmed=True.
    Returns: {created: int, linked: int, errors: list}
    """
    from app.models.purchase_item import PurchaseItem
    from app.routers.purchase_items_import import _upsert_product_to_catalog

    created = 0
    linked = 0
    errors: list[str] = []

    for item_id in body.purchase_item_ids:
        try:
            item = await db.get(PurchaseItem, item_id)
            if not item:
                errors.append(f"PurchaseItem {item_id} не найден")
                continue
            if item.product_id is not None:
                linked += 1
                continue
            from datetime import datetime as _dtb
            _uname = getattr(current_user, 'full_name', None) or getattr(current_user, 'username', '') or ''
            product_id = await _upsert_product_to_catalog(
                db, item.item_name, item.item_type or "товар", item.unit_price,
                import_note=f"Добавлен из позиций закупки (сопоставление), {_uname}, {_dtb.now().strftime('%d.%m.%Y %H:%M')}",
                updated_by=_uname,
            )
            item.product_id = product_id
            item.match_confirmed = True
            created += 1
        except Exception as e:
            errors.append(f"item {item_id}: {e}")

    try:
        await db.commit()
    except Exception as e:
        raise HTTPException(500, f"Ошибка сохранения: {e}")

    return {"created": created, "linked": linked, "errors": errors}