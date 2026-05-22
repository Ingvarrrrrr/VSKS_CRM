"""
ВСКС xlsx-форма путевого листа — Phase 30.1.

Копируем шаблон waybill_vsks_template.xlsx (лист «ПЛ Голичков 275» как канонический образец),
заполняем ключевые ячейки: номер ПЛ, дата, гос.номер, водитель, марка ТС.

TODO: полная разметка через openpyxl coordinates (все поля медосмотра, одометр,
      задание водителю, движение горючего) — задача отдельного PR.

Структура шаблонного листа «ПЛ Голичков 275» (из inspect 2026-05-21):
  R1  C69  — серийный номер бланка (275)
  R1  C78  — «№»
  R3  C34  — «от « »  | C54 — год
  R11 C25  — марка и модель автомобиля
  R12 C36  — гос.номер
  R13 C10  — ФИО водителя
  R13 C57  — СНИЛС
  R15 C13  — номер ВУ  | C39 — дата выдачи  | C56 — категории
"""
import logging
import os
from io import BytesIO
from pathlib import Path

logger = logging.getLogger(__name__)

# Пути к шаблону (работает из Docker /app/... и локально)
_TEMPLATE_CANDIDATES = [
    Path("/app/backend/templates/waybill_vsks_template.xlsx"),
    Path(__file__).parent.parent.parent / "templates" / "waybill_vsks_template.xlsx",
]
_SEED_CANDIDATES = [
    Path("/app/seed_data/waybill_template_vsks.xlsx"),
    Path(__file__).parent.parent.parent / "seed_data" / "waybill_template_vsks.xlsx",
]

# Имя листа-образца
_SHEET_NAME = "ПЛ Голичков 275"


def _get_template_path() -> Path:
    """Возвращает путь к шаблону — сначала из templates/, потом из seed_data/."""
    for p in _TEMPLATE_CANDIDATES + _SEED_CANDIDATES:
        if p.exists():
            return p
    raise FileNotFoundError(
        f"Шаблон waybill_vsks_template.xlsx не найден. "
        f"Проверенные пути: {[str(p) for p in _TEMPLATE_CANDIDATES + _SEED_CANDIDATES]}"
    )


async def render_waybill_xlsx_vsks(waybill, db) -> bytes:
    """
    Рендерит путевой лист ВСКС в формате .xlsx.

    waybill = Trip ORM объект (или SimpleNamespace для тестов).
    db = AsyncSession (может быть None).
    Возвращает bytes .xlsx файла.

    Минимальная разметка (топ-5 полей):
      - Номер ПЛ
      - Дата
      - Гос.номер ТС
      - ФИО водителя
      - Марка и модель ТС

    TODO: полная разметка — координаты ячеек согласованы с Голичковым,
          медосмотр, одометр, топливо, маршрут.
    """
    try:
        import openpyxl
        from openpyxl import load_workbook
        from copy import copy
    except ImportError as e:
        raise RuntimeError(f"openpyxl не установлен: {e}")

    template_path = _get_template_path()

    # Загружаем без read_only — нужно редактировать
    wb = load_workbook(str(template_path), data_only=False, read_only=False)

    # Выбираем лист-образец, fallback — активный лист
    if _SHEET_NAME in wb.sheetnames:
        ws = wb[_SHEET_NAME]
    else:
        ws = wb.active
        logger.warning(
            f"waybill_xlsx: лист '{_SHEET_NAME}' не найден, "
            f"используем активный: {ws.title}"
        )

    # ── Вспомогательные функции ───────────────────────────────────────────────
    def fmt_date(d, fmt="%d.%m.%Y"):
        if d is None:
            return ""
        if hasattr(d, 'strftime'):
            return d.strftime(fmt)
        return str(d)

    def safe_str(v, default=""):
        if v is None:
            return default
        return str(v).strip() or default

    # ── Читаем поля из waybill ────────────────────────────────────────────────
    waybill_number = safe_str(getattr(waybill, 'number', None), str(getattr(waybill, 'id', '')))
    date_start = getattr(waybill, 'date_start', None)
    date_end = getattr(waybill, 'date_end', None)

    vehicle = getattr(waybill, 'vehicle', None)
    plate = safe_str(getattr(vehicle, 'plate', None)) if vehicle else ''
    brand = safe_str(getattr(vehicle, 'brand', None)) if vehicle else ''
    model = safe_str(getattr(vehicle, 'model', None)) if vehicle else ''
    brand_model = f"{brand} {model}".strip() or plate

    driver = getattr(waybill, 'driver_user', None)
    if driver:
        driver_name = safe_str(getattr(driver, 'full_name', None))
        driver_snils = safe_str(getattr(driver, 'snils', None) or getattr(driver, 'inn', None))
        vu_series = safe_str(getattr(driver, 'license_series', None))
        vu_number = safe_str(getattr(driver, 'license_number', None))
        vu_full = f"{vu_series} {vu_number}".strip()
        vu_issued = fmt_date(getattr(driver, 'license_issued_at', None))
        categories = safe_str(getattr(driver, 'license_categories', None))
    else:
        driver_name = safe_str(getattr(waybill, 'driver_name', None))
        driver_snils = ''
        vu_full = ''
        vu_issued = ''
        categories = ''

    # ── Заполняем ключевые ячейки ─────────────────────────────────────────────
    # Номер ПЛ — R1 C69 (col 69 = BQ)
    _set_cell(ws, 1, 69, waybill_number)

    # Дата «от» — R3 C54 (год) и R3 C34 (день+месяц в виде строки)
    if date_start:
        _set_cell(ws, 3, 54, date_start.year if hasattr(date_start, 'year') else '')
        _set_cell(ws, 3, 34, f"от «{fmt_date(date_start, '%d')}» {_month_ru(date_start)}")

    # Марка и модель ТС — R11 C25
    _set_cell(ws, 11, 25, brand_model)

    # Гос.номер — R12 C36
    _set_cell(ws, 12, 36, plate)

    # ФИО водителя — R13 C10
    _set_cell(ws, 13, 10, driver_name)

    # СНИЛС водителя — R13 C57
    if driver_snils:
        _set_cell(ws, 13, 57, driver_snils)

    # ВУ номер — R15 C13
    if vu_full:
        _set_cell(ws, 15, 13, vu_full)
    # Дата выдачи ВУ — R15 C39
    if vu_issued:
        _set_cell(ws, 15, 39, vu_issued)
    # Категории — R15 C56
    if categories:
        _set_cell(ws, 15, 56, categories)

    # ── Сохраняем в BytesIO ───────────────────────────────────────────────────
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def _set_cell(ws, row: int, col: int, value):
    """Безопасно записывает значение в ячейку (row, col) рабочего листа."""
    try:
        ws.cell(row=row, column=col).value = value
    except Exception as e:
        logger.debug(f"waybill_xlsx: не удалось записать R{row}C{col}={value!r}: {e}")


def _month_ru(d) -> str:
    """Возвращает название месяца по-русски (родительный падеж)."""
    months = [
        '', 'января', 'февраля', 'марта', 'апреля', 'мая', 'июня',
        'июля', 'августа', 'сентября', 'октября', 'ноября', 'декабря'
    ]
    try:
        return months[d.month]
    except (AttributeError, IndexError):
        return ''
