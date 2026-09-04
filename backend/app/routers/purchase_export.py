"""Purchase export/import router — extracted from purchases.py (Phase 16-02).

Handles:
  GET  /api/purchases/export/columns   — list available Excel export columns
  GET  /api/purchases/export/excel     — stream .xlsx export of purchases
  GET  /api/purchases/import/template  — download blank import template (1-sheet, inline payments)
  POST /api/purchases/import           — Scroller-format xlsx import
  POST /api/purchases/import/preview   — preview without committing
"""
from fastapi import APIRouter, Depends, Query, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from io import BytesIO
from typing import Optional, List, Dict, Any
from decimal import Decimal
from datetime import datetime
from urllib.parse import quote
from collections import defaultdict

from app.database import get_db
from app.models.purchase import Purchase
from app.models.purchase_item import PurchaseItem
from app.models.subsidy import Subsidy
from app.models.contractor import Contractor
from app.models.feo_category import FeoCategory
from app.models.payment import Payment
from app.services.purchase_payments import recompute_purchase_payments
from app.models.event import Event
from app.routers.events import normalize_event_name
from app.auth.jwt import get_current_user
from app.auth.permissions import has_org_key
from app.auth.visibility import get_visible_subsidy_ids
from app.models.user import User
from app.services.ru_regions import RU_REGIONS

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    from openpyxl.workbook.defined_name import DefinedName
except ImportError:
    Workbook = None
    load_workbook = None
    DataValidation = None
    DefinedName = None

import logging
logger = logging.getLogger(__name__)

router = APIRouter(prefix="/api/purchases", tags=["purchase-export"])

# ---------------------------------------------------------------------------
# Excel export: column registry
# ---------------------------------------------------------------------------

ALL_EXPORT_COLUMNS = {
    "purchase_number":        {"label": "№ п/п",                 "group": "Идентификация"},
    "order_number":           {"label": "Номер заявки",          "group": "Идентификация"},
    "registry_number":        {"label": "Реестровый №",          "group": "Идентификация"},
    "status":                 {"label": "Статус",                "group": "Идентификация"},
    "subsidy":                {"label": "Субсидия",              "group": "Идентификация"},
    "feo_category":           {"label": "Категория ФЭО",         "group": "Идентификация"},
    "item_name":              {"label": "Наименование",          "group": "Позиция"},
    "item_type":              {"label": "Тип",                   "group": "Позиция"},
    "unit":                   {"label": "Ед. изм",               "group": "Позиция"},
    "quantity":               {"label": "Кол-во",                "group": "Позиция"},
    "subject":                {"label": "Предмет закупки",       "group": "Позиция"},
    "country_origin":         {"label": "Страна происхождения",  "group": "Позиция"},
    "planned_unit_price":     {"label": "Плановая цена за ед.",  "group": "Цены"},
    "planned_total_price":    {"label": "Плановая сумма",        "group": "Цены"},
    "nmck":                   {"label": "НМЦК",                  "group": "Цены"},
    "contract_price":         {"label": "Цена договора",         "group": "Цены"},
    "economy":                {"label": "Экономия",              "group": "Цены"},
    "price_increase":         {"label": "Удорожание",            "group": "Цены"},
    "purchase_method":        {"label": "Способ закупки",        "group": "Закупка"},
    "purchase_basis":         {"label": "Основание",             "group": "Закупка"},
    "purchase_contract_type": {"label": "Тип договора",          "group": "Закупка"},
    "framework_seq":          {"label": "№ в рамочном",          "group": "Закупка"},
    "contract_number":        {"label": "№ договора",            "group": "Договор"},
    "contract_date":          {"label": "Дата договора",         "group": "Договор"},
    "execution_term":         {"label": "Срок исполнения",       "group": "Договор"},
    "execution_term_changed": {"label": "Срок (изменён)",        "group": "Договор"},
    "delivery_date":          {"label": "Дата доставки",         "group": "Договор"},
    "contractor":             {"label": "Контрагент",            "group": "Контрагент"},
    "contractor_inn":         {"label": "ИНН контрагента",       "group": "Контрагент"},
    "responsible_person":     {"label": "Ответственное лицо",    "group": "Контрагент"},
    "acceptance_doc_name":    {"label": "Закрывающий документ: наименование", "group": "Исполнение"},
    "acceptance_doc_number":  {"label": "Закрывающий документ: №",           "group": "Исполнение"},
    "acceptance_doc_date":    {"label": "Закрывающий документ: дата",         "group": "Исполнение"},
    "acceptance_doc_amount":  {"label": "Закрывающий документ: сумма",        "group": "Исполнение"},
    "payment_doc_number":     {"label": "ПП: №",                 "group": "Оплата"},
    "payment_doc_date":       {"label": "ПП: дата",              "group": "Оплата"},
    "payment_amount":         {"label": "ПП: сумма",             "group": "Оплата"},
    "payment_federal":        {"label": "В т.ч. фед. бюджет",   "group": "Оплата"},
    "payment_purpose":        {"label": "Назначение платежа",    "group": "Оплата"},
    "delivery_payment_amount":{"label": "Оплата с доставкой",    "group": "Оплата"},
    "vat_applicable":         {"label": "НДС применяется",       "group": "НДС"},
    "vat_rate":               {"label": "Ставка НДС",            "group": "НДС"},
    "vat_exemption_article":  {"label": "Статья НК РФ",          "group": "НДС"},
    "vat_mode":               {"label": "Режим НДС",             "group": "НДС"},
    "etp_url":                {"label": "Ссылка ЭТП",            "group": "Закупка"},
    "region":                 {"label": "Регион мероприятия",     "group": "Позиция"},
    "delivery_region":        {"label": "Регион поставки",        "group": "Позиция"},
    "delivery_location":      {"label": "Место доставки/услуг",  "group": "Позиция"},
    "delivery_address":       {"label": "Адрес доставки",        "group": "Позиция"},
    "final_unit_price":       {"label": "Факт. цена за ед.",     "group": "Цены"},
    "final_total_amount":     {"label": "Факт. сумма",           "group": "Цены"},
    "contract_end_date":      {"label": "Срок действия договора","group": "Договор"},
    "submission_deadline":    {"label": "Окончание приёма заявок","group": "Договор"},
    "commitment_quarter":     {"label": "Квартал обязательств",  "group": "Оплата"},
    "planned_payment_month":  {"label": "План. месяц платежа",   "group": "Оплата"},
    "payment_month":          {"label": "Месяц платежа",         "group": "Оплата"},
    "stage_label":            {"label": "Этап (подпись)",        "group": "Идентификация"},
    "substatus":              {"label": "Подстатус",             "group": "Идентификация"},
}

DEFAULT_EXPORT_COLUMNS = [
    "purchase_number", "registry_number", "item_name", "item_type", "unit", "quantity",
    "region", "delivery_region", "nmck", "planned_total_price", "contract_price", "final_total_amount", "economy",
    "purchase_method", "contract_number", "contract_date", "contract_end_date",
    "contractor", "contractor_inn", "execution_term", "country_origin",
    "acceptance_doc_name", "acceptance_doc_number", "acceptance_doc_date", "acceptance_doc_amount",
    "payment_doc_number", "payment_doc_date", "payment_amount", "payment_federal", "payment_purpose",
    "status",
]

_PURCHASE_METHOD_LABELS = {
    "single":        "Единственный поставщик",
    "competitive":   "Конкурентная процедура",
    "quote_request": "Запрос котировок",
    "advance":       "Авансовый отчёт",
}
_PURCHASE_BASIS_LABELS = {
    "plan_schedule": "план закупок",
    "service_note":  "служебная записка",
}
_CONTRACT_TYPE_LABELS = {
    "single":               "Разовая поставка",
    "framework_cumulative": "Рамочный (нарастающий итог)",
    "framework_with_amount":"Рамочный (с указанием суммы)",
}
_STATUS_LABELS = {
    "wishes":          "Желания сотрудников",
    "plan_schedule":   "План закупок",
    "work_in_progress":"Ведётся работа",
    "contracted":      "Заключён договор",
    "ordered":         "Заказано",
    "delivered":       "Поставлено",
    "paid":            "Оплачено",
}
_SUBSTATUS_LABELS = {
    "tz_forming":              "Формирование ТЗ",
    "kp_collecting":           "Сбор КП",
    "on_platform":             "Размещено на площадке",
    "contractor_negotiations": "Переговоры с поставщиком",
    "contract_signing":        "Подписание договора",
}
_ITEM_TYPE_LABELS = {
    "товар":  "Товар",
    "услуга": "Услуга",
    "работа": "Работа",
}
_PAYMENT_BASIS_LABELS = {
    "contract":          "Договор",
    "invoice":           "Счёт",
    "invoice_contract":  "Счёт-договор",
}

# ---------------------------------------------------------------------------
# Dropdown value sets — used for DataValidation AND «Справочники» sheet.
# Each tuple: (display_label, internal_key or None).
# For fields where we store the display label as-is (free text): only label, key=None.
# ---------------------------------------------------------------------------
_DD_CONTRACT_TYPE   = [  # label → key stored in purchase_contract_type
    ("Разовая поставка",             "single"),
    ("Рамочный (нарастающий итог)", "framework_cumulative"),
    ("Рамочный (с указанием суммы)","framework_with_amount"),
]
_DD_STATUS = [           # label → key stored in status
    ("Желания сотрудников",        "wishes"),
    ("План закупок",                "plan_schedule"),
    ("Ведётся работа",             "work_in_progress"),
    ("Заключён договор",           "contracted"),
    ("Заказано",                   "ordered"),
    ("Поставлено",                 "delivered"),
    ("Поставлено, но не оплачено", "delivered"),
    ("Оплачено",                   "paid"),
]
_DD_SUBSTATUS = [        # label → key stored in substatus
    ("Формирование ТЗ",             "tz_forming"),
    ("Сбор КП",                     "kp_collecting"),
    ("Размещено на площадке",       "on_platform"),
    ("Переговоры с поставщиком",    "contractor_negotiations"),
    ("Подписание договора",         "contract_signing"),
]
_DD_METHOD = [           # label → key stored in purchase_method
    ("Единственный поставщик",  "single"),
    ("Конкурентная процедура",  "competitive"),
    ("Запрос котировок",        "quote_request"),
]
_DD_BASIS = [            # label → key stored in purchase_basis
    ("план закупок",      "plan_schedule"),
    ("служебная записка","service_note"),
]
_DD_VAT_APPLICABLE = [   # just display values, stored as bool
    ("Да",   None),
    ("Нет",  None),
]
_DD_VAT_RATE = [         # stored as integer
    ("0",   None),
    ("10",  None),
    ("20",  None),
    ("22",  None),
]
_DD_PREPAYMENT = [
    ("Да",  None),
    ("Нет", None),
]
_DD_MONTHLY = [
    ("Да",  None),
    ("Нет", None),
]
_DD_ITEM_TYPE = [        # label → stored directly in item_type
    ("Товар",  "товар"),
    ("Услуга", "услуга"),
    ("Работа", "работа"),
]
_DD_PAYMENT_BASIS = [    # label → key stored in payment_basis_type
    ("Договор",      "contract"),
    ("Счёт",         "invoice"),
    ("Счёт-договор", "invoice_contract"),
]
_DD_UNIT = [             # free-text suggestions only
    ("шт",    None), ("пар",  None), ("комп.", None), ("кг",   None),
    ("л",     None), ("м",    None), ("м²",   None),  ("м³",   None),
    ("уп.",   None), ("набор",None), ("усл.", None),
]
_DD_QUARTER = [          # квартал принятия обязательств (1-4)
    ("1", None),
    ("2", None),
    ("3", None),
    ("4", None),
]
_DD_VAT_MODE = [         # режим НДС
    ("Одинаковый",          "uniform"),
    ("Для каждого товара",  "per_item"),
]

# 89 субъектов РФ — для поля delivery_region (регион поставки, без спец-значений)
_DD_DELIVERY_REGION: list[tuple[str, str]] = [
    (name, name) for name in sorted(RU_REGIONS.keys())
]

# 89 субъектов + спец-значения — для поля region (регион проведения мероприятия)
_DD_EVENT_REGION: list[tuple[str, str]] = [
    ("Не определён",         "Не определён"),
    ("Несколько регионов",   "Несколько регионов"),
    ("Федеральное мероприятие", "Федеральное мероприятие"),
] + _DD_DELIVERY_REGION


def _get_cell_value(key: str, p: Purchase, ctx: dict):
    if key == "purchase_number":         return p.purchase_number or ""
    if key == "order_number":            return p.order_number or ""
    if key == "registry_number":         return p.registry_number or ""
    if key == "status":                  return _STATUS_LABELS.get(p.status, p.status or "")
    if key == "subsidy":                 return ctx["subsidies"].get(p.subsidy_id, "")
    if key == "feo_category":            return ctx["feo_categories"].get(p.feo_category_id, "")
    if key == "item_name":               return p.item_name or ""
    if key == "item_type":               return p.item_type or ""
    if key == "unit":                    return p.unit or ""
    if key == "quantity":                return float(p.planned_quantity) if p.planned_quantity else ""
    if key == "subject":                 return p.subject or ""
    if key == "country_origin":          return p.country_origin or ""
    if key == "planned_unit_price":      return float(p.planned_unit_price) if p.planned_unit_price else ""
    if key == "planned_total_price":     return float(p.planned_total_price) if p.planned_total_price else ""
    if key == "nmck":                    return float(p.nmck or p.planned_total_price or 0) or ""
    if key == "contract_price":          return float(p.contract_price) if p.contract_price else ""
    if key == "economy":                 return float(p.economy) if p.economy else ""
    if key == "price_increase":          return float(p.price_increase) if p.price_increase else ""
    if key == "purchase_method":         return _PURCHASE_METHOD_LABELS.get(p.purchase_method, p.purchase_method or "")
    if key == "purchase_basis":          return _PURCHASE_BASIS_LABELS.get(p.purchase_basis, p.purchase_basis or "")
    if key == "purchase_contract_type":  return _CONTRACT_TYPE_LABELS.get(p.purchase_contract_type, p.purchase_contract_type or "")
    if key == "framework_seq":           return p.framework_seq if p.framework_seq is not None else ""
    if key == "contract_number":         return p.contract_number or ""
    if key == "contract_date":           return str(p.contract_date) if p.contract_date else ""
    if key == "execution_term":          return str(p.execution_term) if p.execution_term else ""
    if key == "execution_term_changed":  return str(p.execution_term_changed) if p.execution_term_changed else ""
    if key == "delivery_date":           return str(p.delivery_date) if p.delivery_date else ""
    if key == "contractor":              return ctx["contractors"].get(p.contractor_id, "")
    if key == "contractor_inn":          return ctx["contractor_inns"].get(p.contractor_id, "")
    if key == "responsible_person":      return p.responsible_person or ""
    if key == "acceptance_doc_name":     return p.acceptance_doc_name or ""
    if key == "acceptance_doc_number":   return p.acceptance_doc_number or ""
    if key == "acceptance_doc_date":     return str(p.acceptance_doc_date) if p.acceptance_doc_date else ""
    if key == "acceptance_doc_amount":   return float(p.acceptance_doc_amount) if p.acceptance_doc_amount else ""
    if key == "payment_doc_number":      return p.payment_doc_number or ""
    if key == "payment_doc_date":        return str(p.payment_doc_date) if p.payment_doc_date else ""
    if key == "payment_amount":          return float(p.payment_amount) if p.payment_amount else ""
    if key == "payment_federal":         return float(p.payment_federal) if p.payment_federal else ""
    if key == "payment_purpose":         return ctx["payment_purposes"].get(p.id, "")
    if key == "delivery_payment_amount": return float(p.delivery_payment_amount) if p.delivery_payment_amount else ""
    if key == "vat_applicable":          return "Да" if p.vat_applicable else ""
    if key == "vat_rate":                return p.vat_rate if p.vat_rate is not None else ""
    if key == "vat_exemption_article":   return p.vat_exemption_article or ""
    if key == "vat_mode":
        _vat_mode_labels = {"uniform": "Одинаковый", "per_item": "Для каждого товара"}
        return _vat_mode_labels.get(p.vat_mode or "uniform", p.vat_mode or "")
    if key == "etp_url":                 return "" if getattr(p, 'purchase_method', None) == 'advance' else (p.etp_url or "")
    if key == "region":                  return p.region or ""
    if key == "delivery_region":         return p.delivery_region or ""
    if key == "delivery_location":       return p.delivery_location or ""
    if key == "delivery_address":        return p.delivery_address or ""
    if key == "final_unit_price":        return float(p.final_unit_price) if p.final_unit_price else ""
    if key == "final_total_amount":      return float(p.final_total_amount) if p.final_total_amount else ""
    if key == "contract_end_date":       return str(p.contract_end_date) if p.contract_end_date else ""
    if key == "submission_deadline":
        if p.submission_deadline:
            try:
                return str(p.submission_deadline.date()) if hasattr(p.submission_deadline, 'date') else str(p.submission_deadline)
            except Exception:
                return str(p.submission_deadline)
        return ""
    if key == "commitment_quarter":      return p.commitment_quarter if p.commitment_quarter is not None else ""
    if key == "planned_payment_month":   return str(p.planned_payment_month) if p.planned_payment_month else ""
    if key == "payment_month":
        if p.payment_doc_date:
            try:
                d = p.payment_doc_date
                return f"{d.month:02d}.{d.year}"
            except Exception:
                return ""
        return ""
    if key == "stage_label":             return p.stage_label or ""
    if key == "substatus":               return _SUBSTATUS_LABELS.get(p.substatus or "", p.substatus or "")
    return ""


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------

@router.get("/export/columns")
async def get_export_columns(_=Depends(get_current_user)):
    """Return available export column definitions."""
    return [
        {"key": k, "label": v["label"], "group": v["group"]}
        for k, v in ALL_EXPORT_COLUMNS.items()
    ]


@router.get("/export/excel")
async def export_purchases_to_excel(
    subsidy_id: Optional[int] = Query(None),
    status: Optional[str] = Query(None),
    columns: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
    _=Depends(get_current_user),
):
    if Workbook is None:
        raise HTTPException(500, "openpyxl не установлен")

    col_keys = (
        [k.strip() for k in columns.split(",") if k.strip() in ALL_EXPORT_COLUMNS]
        if columns else DEFAULT_EXPORT_COLUMNS
    )
    if not col_keys:
        col_keys = DEFAULT_EXPORT_COLUMNS

    q = select(Purchase).order_by(Purchase.id.desc())
    if subsidy_id:
        q = q.where(Purchase.subsidy_id == subsidy_id)
    if status:
        q = q.where(Purchase.status == status)
    purchases = (await db.execute(q)).scalars().all()

    contractor_rows = (await db.execute(select(Contractor))).scalars().all()
    contractors = {c.id: c.name for c in contractor_rows}
    contractor_inns = {c.id: (c.inn or "") for c in contractor_rows}
    subsidies_map = {s.id: s.name for s in (await db.execute(select(Subsidy))).scalars().all()}
    feo_map = {f.id: f.name for f in (await db.execute(select(FeoCategory))).scalars().all()}

    purchase_ids = [p.id for p in purchases]
    payment_purposes: dict[int, str] = {}
    if purchase_ids:
        pay_rows = (
            await db.execute(
                select(Payment)
                .where(Payment.purchase_id.in_(purchase_ids))
                .order_by(Payment.payment_date.asc().nullsfirst(), Payment.id.asc())
            )
        ).scalars().all()
        _pp_map: dict[int, list[str]] = defaultdict(list)
        for pay in pay_rows:
            if pay.payment_purpose and pay.payment_purpose.strip():
                _pp_map[pay.purchase_id].append(pay.payment_purpose.strip())
        payment_purposes = {pid: "; ".join(purposes) for pid, purposes in _pp_map.items()}

    ctx = {
        "contractors": contractors,
        "contractor_inns": contractor_inns,
        "subsidies": subsidies_map,
        "feo_categories": feo_map,
        "payment_purposes": payment_purposes,
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Закупки"

    col_headers = [ALL_EXPORT_COLUMNS[k]["label"] for k in col_keys]
    ws.append(col_headers)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", wrap_text=True)

    empty_counts = {k: 0 for k in col_keys}
    for p in purchases:
        row = []
        for k in col_keys:
            val = _get_cell_value(k, p, ctx)
            row.append(val)
            if val == "" or val is None:
                empty_counts[k] += 1
        ws.append(row)

    for i, key in enumerate(col_keys, 1):
        col_letter = ws.cell(1, i).column_letter
        ws.column_dimensions[col_letter].width = max(len(col_headers[i - 1]) + 2, 12)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"Закупки_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    missing = []
    if len(purchases) >= 5:
        for k in col_keys:
            if empty_counts[k] / len(purchases) > 0.8:
                missing.append(ALL_EXPORT_COLUMNS[k]["label"])

    resp_headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename, safe='-_.~')}",
        "Access-Control-Expose-Headers": "X-Missing-Columns",
    }
    if missing:
        # HTTP-заголовки кодируются latin-1; кириллица в названиях колонок → percent-encode.
        resp_headers["X-Missing-Columns"] = quote(",".join(missing[:5]))

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=resp_headers,
    )


# ---------------------------------------------------------------------------
# Import template (single «Закупки» sheet, payments inline)
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Single source of truth: column specification for import template.
# Each entry:
#   header        – text in row 1
#   required      – True → red header (DC2626), False → blue (2563EB)
#   width         – column width in Excel
#   fmt           – format / example value shown in справочник
#   effect        – what this field affects in the system
#   if_empty      – what happens when the cell is blank
#   comment       – 2-3 sentence Excel comment text (shown as cell note)
# ---------------------------------------------------------------------------
_COL_SPEC: list[dict] = [
    {
        "header":   "Тип договора",
        "required": False,
        "width":    28,
        "fmt":      "Разовая поставка / Рамочный (нарастающий итог) / Рамочный (с указанием суммы)",
        "effect":   "Тип договора закупки (purchase_contract_type)",
        "if_empty": "Тип не указывается, закупка сохраняется без типа договора",
        "comment":  (
            "Тип договора закупки. Выберите из выпадающего списка или впишите вручную.\n"
            "Допустимые значения: Разовая поставка, Рамочный (нарастающий итог), "
            "Рамочный (с указанием суммы).\n"
            "Влияет на логику группировки заказов внутри одного договора."
        ),
        "_dd": "_DD_CONTRACT_TYPE",
    },
    {
        "header":   "Номер закупки",
        "required": False,
        "width":    16,
        "fmt":      "20",
        "effect":   "Номер группы закупки (purchase_number) — для нумерации в плане-графике",
        "if_empty": "Номер закупки не назначается",
        "comment":  (
            "Порядковый номер закупки в плане-графике. Используется для группировки строк "
            "одной закупки (если одна закупка содержит несколько позиций).\n"
            "Влияет на отображение в реестре закупок.\n"
            "Если не заполнено — номер не будет присвоен."
        ),
    },
    {
        "header":   "Номер заказа внутри закупки",
        "required": False,
        "width":    26,
        "fmt":      "Заказ 1 (март)",
        "effect":   "Номер/название заказа внутри рамочного договора (order_number)",
        "if_empty": "Заказ не именуется — закупка сохраняется без номера заказа",
        "comment":  (
            "Для рамочных договоров: название или номер конкретного заказа (поставки) "
            "внутри рамочного договора, например «Заказ 1 (март)».\n"
            "Вместе с «Номером закупки» формирует уникальный ключ: одна строка = один заказ.\n"
            "Если не заполнено — закупка не привязывается к конкретному заказу."
        ),
    },
    {
        "header":   "Предмет договора (общий)",
        "required": False,
        "width":    30,
        "fmt":      "Обмундирование к слёту",
        "effect":   "Общий предмет договора (subject) — отображается в карточке закупки",
        "if_empty": "Предметом станет наименование первой позиции",
        "comment":  (
            "Общий предмет договора — краткое описание того, что закупается по договору в целом.\n"
            "Отображается в заголовке карточки закупки и в реестре.\n"
            "Если не заполнено — в качестве предмета будет использовано наименование первой позиции."
        ),
    },
    {
        "header":   "Наименование товара",
        "required": True,
        "width":    32,
        "fmt":      "Кепи камуфляж",
        "effect":   "Наименование позиции (item_name) — обязательно; строка без этого поля пропускается",
        "if_empty": "Строка отклоняется импортом",
        "comment":  (
            "Наименование конкретной позиции (товара, услуги, работы). Обязательное поле.\n"
            "Каждая уникальная позиция попадает в список позиций закупки (PurchaseItem).\n"
            "Строка без наименования будет пропущена с соответствующей ошибкой."
        ),
    },
    {
        "header":   "ФЭО Ур.1",
        "required": True,
        "width":    22,
        "fmt":      "Снаряжение",
        "effect":   "Первый уровень иерархии ФЭО — определяет категорию закупки",
        "if_empty": "Строка отклоняется импортом",
        "comment":  (
            "Первый (корневой) уровень категории ФЭО субсидии. Обязательное поле.\n"
            "Значение ищется в дереве ФЭО выбранной субсидии точным совпадением (или вхождением).\n"
            "Если категория не найдена или поле пустое — строка будет отклонена с ошибкой."
        ),
    },
    {
        "header":   "ФЭО Ур.2",
        "required": False,
        "width":    22,
        "fmt":      "Одежда",
        "effect":   "Второй уровень иерархии ФЭО",
        "if_empty": "Закупка привязывается к категории первого уровня",
        "comment":  (
            "Второй уровень иерархии категории ФЭО. Заполняйте, только если в субсидии "
            "есть второй уровень дерева ФЭО.\n"
            "Система последовательно спускается по уровням: Ур.1 → Ур.2 → Ур.3 и т.д.\n"
            "Если не заполнено — закупка привяжется к категории предыдущего уровня."
        ),
    },
    {
        "header":   "ФЭО Ур.3",
        "required": False,
        "width":    22,
        "fmt":      "Кепи",
        "effect":   "Третий уровень иерархии ФЭО",
        "if_empty": "Закупка привязывается к категории предыдущего уровня",
        "comment":  (
            "Третий уровень иерархии категории ФЭО. Заполняйте, если в субсидии есть третий уровень.\n"
            "Значение ищется среди дочерних узлов категории второго уровня.\n"
            "Если не заполнено — закупка привяжется к категории второго уровня."
        ),
    },
    {
        "header":   "ФЭО Ур.4",
        "required": False,
        "width":    22,
        "fmt":      "",
        "effect":   "Четвёртый уровень иерархии ФЭО",
        "if_empty": "Закупка привязывается к категории третьего уровня",
        "comment":  (
            "Четвёртый уровень иерархии ФЭО. Используется для глубоких деревьев категорий.\n"
            "Заполняйте только при наличии четвёртого уровня в субсидии.\n"
            "Если не заполнено — категория берётся с предыдущего уровня."
        ),
    },
    {
        "header":   "ФЭО Ур.5",
        "required": False,
        "width":    22,
        "fmt":      "",
        "effect":   "Пятый уровень иерархии ФЭО",
        "if_empty": "Закупка привязывается к категории четвёртого уровня",
        "comment":  (
            "Пятый (самый глубокий) уровень иерархии ФЭО. Заполняется крайне редко.\n"
            "Значение ищется среди дочерних узлов категории четвёртого уровня.\n"
            "Если не заполнено — категория берётся с предыдущего уровня."
        ),
    },
    {
        "header":   "Мероприятие",
        "required": False,
        "width":    25,
        "fmt":      "Форма к слёту",
        "effect":   "Название мероприятия (event_name) — привязка закупки к событию",
        "if_empty": "Закупка не привязывается к мероприятию",
        "comment":  (
            "Название мероприятия, в рамках которого производится закупка.\n"
            "Система ищет мероприятие по названию и привязывает закупку (event_id).\n"
            "Если не заполнено или мероприятие не найдено — связь с мероприятием не создаётся."
        ),
    },
    {
        "header":   "Контрагент",
        "required": False,
        "width":    25,
        "fmt":      "ООО Поставщик",
        "effect":   "Название контрагента для поиска или подсказки при создании",
        "if_empty": "Контрагент определяется по ИНН; при совпадении ИНН — без разницы",
        "comment":  (
            "Наименование контрагента. Используется как запасной вариант поиска, "
            "если контрагент не найден по ИНН.\n"
            "При создании нового контрагента (когда ИНН не найден в базе) — становится именем.\n"
            "Если не заполнено, но ИНН есть — контрагент будет назван по ИНН."
        ),
    },
    {
        "header":   "ИНН контрагента",
        "required": True,
        "width":    18,
        "fmt":      "1234567890",
        "effect":   "По ИНН ищется или создаётся контрагент; без него строка отклоняется",
        "if_empty": "Строка отклоняется импортом",
        "comment":  (
            "ИНН контрагента (10 цифр для юридических лиц, 12 для ИП). Обязательное поле.\n"
            "Система ищет контрагента по ИНН в базе; если не найден — создаёт нового.\n"
            "Строка без ИНН будет отклонена."
        ),
    },
    {
        "header":   "Способ закупки",
        "required": False,
        "width":    28,
        "fmt":      "Единственный поставщик / Конкурентная процедура / Запрос котировок",
        "effect":   "Способ проведения закупки (purchase_method)",
        "if_empty": "Способ закупки не указывается",
        "comment":  (
            "Способ закупки. Выберите из выпадающего списка или впишите вручную.\n"
            "Допустимые значения: Единственный поставщик, Конкурентная процедура, Запрос котировок.\n"
            "Влияет на отображение в реестре и фильтрах по способу закупки."
        ),
        "_dd": "_DD_METHOD",
    },
    {
        "header":   "Реестровый №",
        "required": False,
        "width":    16,
        "fmt":      "2026/001",
        "effect":   "Реестровый номер закупки (registry_number) в реестре закупок субсидии",
        "if_empty": "Реестровый номер не присваивается",
        "comment":  (
            "Реестровый номер закупки в реестре по плану-графику субсидии.\n"
            "Отображается в карточке закупки и при экспорте.\n"
            "Если не заполнено — поле останется пустым, это не ошибка."
        ),
    },
    {
        "header":   "№ договора",
        "required": True,
        "width":    16,
        "fmt":      "Д-001",
        "effect":   "Номер договора (contract_number) — ключ группировки строк в одну закупку",
        "if_empty": "Строка отклоняется импортом",
        "comment":  (
            "Номер договора. Обязательное поле — служит ключом группировки: "
            "строки с одинаковым номером договора объединяются в одну закупку.\n"
            "Также используется для связи платёжных документов с закупкой.\n"
            "Строки без номера договора будут отклонены."
        ),
    },
    {
        "header":   "Дата договора",
        "required": True,
        "width":    14,
        "fmt":      "15.03.2026",
        "effect":   "Дата заключения договора (contract_date)",
        "if_empty": "Строка отклоняется импортом",
        "comment":  (
            "Дата заключения договора в формате ДД.ММ.ГГГГ или ГГГГ-ММ-ДД. Обязательное поле.\n"
            "Отображается в карточке закупки, используется при формировании документов.\n"
            "Строки без даты договора будут отклонены."
        ),
    },
    {
        "header":   "Максимальная цена договора",
        "required": False,
        "width":    24,
        "fmt":      "600000",
        "effect":   "Цена договора / НМЦК (contract_price) — максимальный лимит по договору",
        "if_empty": "Цена договора не задаётся; НМЦК рассчитывается из позиций",
        "comment":  (
            "Цена договора (или максимальная цена для рамочного). Для рамочных договоров — "
            "суммарный лимит, не сумма одного заказа.\n"
            "Используется для расчёта экономии и отображения в реестре.\n"
            "Если не заполнено — НМЦК рассчитается как сумма плановых позиций."
        ),
    },
    {
        "header":   "Срок исполнения",
        "required": False,
        "width":    16,
        "fmt":      "30.06.2026",
        "effect":   "Срок исполнения договора (execution_term) — плановая дата поставки",
        "if_empty": "Срок исполнения не задаётся",
        "comment":  (
            "Плановая дата исполнения договора (поставки / оказания услуг) в формате ДД.ММ.ГГГГ.\n"
            "Отображается в карточке закупки и участвует в контроле сроков.\n"
            "Если не заполнено — срок не будет указан."
        ),
    },
    {
        "header":   "Срок исполнения (изменён)",
        "required": False,
        "width":    20,
        "fmt":      "31.08.2026",
        "effect":   "Изменённый срок исполнения (execution_term_changed) — после доп. соглашения",
        "if_empty": "Изменённый срок не указывается",
        "comment":  (
            "Новый срок исполнения, если он был изменён доп. соглашением к договору.\n"
            "При заполнении отображается отдельно от исходного срока — в скобках в карточке.\n"
            "Если не заполнено — будет использован исходный срок исполнения."
        ),
    },
    {
        "header":   "Дата доставки",
        "required": False,
        "width":    14,
        "fmt":      "05.07.2026",
        "effect":   "Фактическая дата доставки (delivery_date)",
        "if_empty": "Дата доставки не фиксируется",
        "comment":  (
            "Фактическая дата доставки товара или выполнения услуги в формате ДД.ММ.ГГГГ.\n"
            "Отображается в карточке закупки в разделе исполнения.\n"
            "Если не заполнено — дата доставки не указывается."
        ),
    },
    {
        "header":   "Этап закупки",
        "required": False,
        "width":    22,
        "fmt":      "Оплачено / Ведётся работа / Заключён договор",
        "effect":   "Этап (статус) закупки в рабочем процессе (status)",
        "if_empty": "Этап определяется автоматически по заполненным полям",
        "comment":  (
            "Этап закупки. Выберите из выпадающего списка или впишите вручную.\n"
            "Допустимые значения: Желания сотрудников, План закупок, Ведётся работа, "
            "Заключён договор, Заказано, Поставлено (= поставлено, но не оплачено), Оплачено.\n"
            "Если не заполнено — этап определяется по данным строки: оплата → Оплачено, "
            "договор → Заключён договор, иначе → Ведётся работа."
        ),
        "_dd": "_DD_STATUS",
    },
    {
        "header":   "Подстатус (для «Ведётся работа»)",
        "required": False,
        "width":    30,
        "fmt":      "Сбор КП / Размещено на площадке",
        "effect":   "Подстатус этапа «Ведётся работа» (substatus) — уточняет стадию проработки",
        "if_empty": "Подстатус не устанавливается",
        "comment":  (
            "Уточняющий подстатус для этапа «Ведётся работа». Выберите из списка.\n"
            "Значения: Формирование ТЗ, Сбор КП, Размещено на площадке, "
            "Переговоры с поставщиком, Подписание договора.\n"
            "Если не заполнено или этап не «Ведётся работа» — подстатус не устанавливается."
        ),
        "_dd": "_DD_SUBSTATUS",
    },
    {
        "header":   "Основание закупки",
        "required": False,
        "width":    22,
        "fmt":      "план закупок / служебная записка",
        "effect":   "Основание для проведения закупки (purchase_basis)",
        "if_empty": "Основание не указывается",
        "comment":  (
            "Основание для проведения закупки. Выберите из выпадающего списка.\n"
            "Допустимые значения: «план закупок» или «служебная записка».\n"
            "Если не заполнено — основание не будет указано."
        ),
        "_dd": "_DD_BASIS",
    },
    {
        "header":   "Ответственное лицо",
        "required": False,
        "width":    25,
        "fmt":      "Иванов И.И.",
        "effect":   "ФИО ответственного за закупку (responsible_person)",
        "if_empty": "Ответственное лицо не указывается",
        "comment":  (
            "ФИО или должность сотрудника, ответственного за проведение закупки.\n"
            "Отображается в карточке закупки и может использоваться в документах.\n"
            "Если не заполнено — поле останется пустым."
        ),
    },
    {
        "header":   "Ссылка ЭТП",
        "required": False,
        "width":    30,
        "fmt":      "https://etpgpb.ru/procedure/12345",
        "effect":   "URL процедуры на торговой площадке (etp_url)",
        "if_empty": "Ссылка на ЭТП не сохраняется",
        "comment":  (
            "Ссылка (URL) на процедуру закупки на электронной торговой площадке (ЭТП).\n"
            "Отображается в карточке закупки как кликабельная ссылка.\n"
            "Если не заполнено — поле останется пустым."
        ),
    },
    {
        "header":   "Номер платёжного документа",
        "required": False,
        "width":    26,
        "fmt":      "ПП-101",
        "effect":   "Номер платёжного поручения (payment_doc_number) — первый платёж",
        "if_empty": "Платёж не регистрируется для этой строки",
        "comment":  (
            "Номер платёжного поручения (ПП). Для нескольких платежей по одному договору "
            "добавляйте отдельную строку с тем же № договора и новым ПП.\n"
            "Используется для формирования реестра оплат и расчёта статуса Оплачено.\n"
            "Если не заполнено — платёж не будет создан для этой строки."
        ),
    },
    {
        "header":   "Дата платёжного документа",
        "required": False,
        "width":    22,
        "fmt":      "01.04.2026",
        "effect":   "Дата платёжного поручения (payment_doc_date)",
        "if_empty": "Дата платежа не фиксируется",
        "comment":  (
            "Дата платёжного поручения в формате ДД.ММ.ГГГГ.\n"
            "Отображается в карточке платежа и используется при расчёте кассовых показателей.\n"
            "Если не заполнено — дата платежа не указывается."
        ),
    },
    {
        "header":   "Сумма оплаты",
        "required": False,
        "width":    16,
        "fmt":      "245000",
        "effect":   "Сумма по данному платёжному документу (payment_amount) — участвует в дашборде",
        "if_empty": "Сумма платежа не учитывается",
        "comment":  (
            "Сумма по данному платёжному поручению в рублях.\n"
            "Участвует в расчёте показателя «Оплачено» на дашборде и виджетах субсидии.\n"
            "Если не заполнено — этот платёж не будет учтён в сводной оплате закупки."
        ),
    },
    {
        "header":   "Назначение платежа",
        "required": False,
        "width":    30,
        "fmt":      "Оплата по Д-001, 1-й платёж",
        "effect":   "Текст назначения платежа (payment_purpose) — для справки",
        "if_empty": "Назначение платежа не указывается",
        "comment":  (
            "Текст назначения платежа из платёжного поручения.\n"
            "Сохраняется в записи платежа для справки и не влияет на расчёты.\n"
            "Если не заполнено — назначение платежа останется пустым."
        ),
    },
    {
        "header":   "В т.ч. федеральный бюджет",
        "required": False,
        "width":    24,
        "fmt":      "120000",
        "effect":   "Доля оплаты из федерального бюджета (payment_federal) — участвует в отчётах по источникам финансирования",
        "if_empty": "Федеральный бюджет не учитывается в разбивке оплаты",
        "comment":  (
            "Сумма в рублях, оплаченная из федерального бюджета в рамках данного платёжного поручения.\n"
            "Влияет на разбивку по источникам финансирования в отчётах по субсидии.\n"
            "Если не заполнено — федеральная доля оплаты не учитывается."
        ),
    },
    {
        "header":   "Количество (план)",
        "required": True,
        "width":    16,
        "fmt":      "100",
        "effect":   "Плановое количество (planned_quantity) — основа расчёта плановой суммы",
        "if_empty": "Строка отклоняется импортом",
        "comment":  (
            "Плановое количество единиц товара или объём услуги. Обязательное поле.\n"
            "Вместе с ценой за единицу формирует плановую сумму закупки.\n"
            "Строка без количества будет отклонена."
        ),
    },
    {
        "header":   "Ед. изм.",
        "required": False,
        "width":    10,
        "fmt":      "шт / пар / м²",
        "effect":   "Единица измерения (unit) — отображается в карточке и документах",
        "if_empty": "Единица измерения не указывается",
        "comment":  (
            "Единица измерения товара или услуги. Выберите из списка или впишите своё.\n"
            "Частые значения: шт, пар, кг, л, м, м², уп., компл., усл.\n"
            "Если не заполнено — единица измерения останется пустой."
        ),
        "_dd": "_DD_UNIT",
    },
    {
        "header":   "Цена за ед. (план)",
        "required": True,
        "width":    18,
        "fmt":      "4900",
        "effect":   "Плановая цена за единицу (planned_unit_price) — основа НМЦК",
        "if_empty": "Строка отклоняется импортом",
        "comment":  (
            "Плановая цена за единицу товара или услуги в рублях. Обязательное поле.\n"
            "Умножается на количество для получения плановой суммы.\n"
            "Строка без цены будет отклонена."
        ),
    },
    {
        "header":   "Сумма план",
        "required": False,
        "width":    14,
        "fmt":      "490000",
        "effect":   "Плановая сумма (planned_total_price) — если не задана, считается из цены × кол-во",
        "if_empty": "Сумма рассчитывается как Цена × Количество",
        "comment":  (
            "Плановая сумма по позиции. Если не задана — рассчитывается автоматически "
            "как Цена за ед. × Количество.\n"
            "Можно оставить пустым — система посчитает сама.\n"
            "Если заполнено, используется как итоговая плановая сумма."
        ),
    },
    {
        "header":   "Кол-во факт",
        "required": False,
        "width":    14,
        "fmt":      "100",
        "effect":   "Фактическое количество поставленного товара (fact_qty)",
        "if_empty": "Фактическое количество не указывается",
        "comment":  (
            "Фактически поставленное количество единиц товара (после поставки).\n"
            "Используется при формировании закрывающих документов.\n"
            "Если не заполнено — фактическое количество не будет указано."
        ),
    },
    {
        "header":   "Цена за ед. (факт)",
        "required": False,
        "width":    18,
        "fmt":      "4900",
        "effect":   "Фактическая цена за единицу (fact_unit_price)",
        "if_empty": "Фактическая цена не указывается",
        "comment":  (
            "Фактическая цена за единицу товара по договору (может отличаться от плановой).\n"
            "Используется при расчёте фактической суммы и формировании документов.\n"
            "Если не заполнено — фактическая цена не будет задана."
        ),
    },
    {
        "header":   "Сумма факт",
        "required": True,
        "width":    14,
        "fmt":      "490000",
        "effect":   "Итоговая фактическая сумма поставки (final_total_amount) — участвует в дашборде",
        "if_empty": "Строка отклоняется импортом",
        "comment":  (
            "Итоговая фактическая сумма поставки по данной позиции в рублях. Обязательное поле.\n"
            "Участвует в расчёте показателей «Поставлено» и «Оплачено» на дашборде.\n"
            "Строка без фактической суммы будет отклонена."
        ),
    },
    {
        "header":   "Страна происхождения",
        "required": False,
        "width":    22,
        "fmt":      "РФ",
        "effect":   "Страна происхождения товара (country_origin) — для реестра и документов",
        "if_empty": "Страна происхождения не указывается",
        "comment":  (
            "Страна происхождения товара (например: РФ, Китай, Германия).\n"
            "Отображается в реестре позиций и используется при формировании документов.\n"
            "Если не заполнено — поле останется пустым."
        ),
    },
    {
        "header":   "НДС применяется",
        "required": False,
        "width":    16,
        "fmt":      "Да / (пусто)",
        "effect":   "Признак применения НДС (vat_applicable) — участвует в документах",
        "if_empty": "НДС считается неприменимым (по умолчанию)",
        "comment":  (
            "Признак применения НДС к закупке. Выберите «Да» из списка, если НДС применяется.\n"
            "Влияет на расчёт суммы НДС в документах (договоре, счёте-фактуре).\n"
            "Если не заполнено — НДС считается неприменимым."
        ),
        "_dd": "_DD_VAT_APPLICABLE",
    },
    {
        "header":   "Ставка НДС",
        "required": False,
        "width":    12,
        "fmt":      "0 / 10 / 20",
        "effect":   "Ставка НДС в процентах (vat_rate) — 0, 10 или 20",
        "if_empty": "Ставка НДС не указывается",
        "comment":  (
            "Ставка НДС в процентах. Выберите из списка: 0, 10 или 20.\n"
            "Используется при формировании документов (счёт-фактура, договор с НДС).\n"
            "Если не заполнено — ставка НДС не задаётся."
        ),
        "_dd": "_DD_VAT_RATE",
    },
    {
        "header":   "Статья НК РФ",
        "required": False,
        "width":    22,
        "fmt":      "ст. 149 п. 3 пп. 17",
        "effect":   "Статья НК РФ об освобождении от НДС (vat_exemption_article)",
        "if_empty": "Основание для освобождения от НДС не указывается",
        "comment":  (
            "Статья НК РФ, на основании которой закупка освобождена от НДС "
            "(заполняется вместо ставки, если НДС не применяется по закону).\n"
            "Используется при формировании договора и документов с признаком «без НДС».\n"
            "Если не заполнено — поле освобождения от НДС останется пустым."
        ),
    },
    {
        "header":   "Закрывающий документ: наименование",
        "required": False,
        "width":    30,
        "fmt":      "УПД",
        "effect":   "Наименование закрывающего документа (acceptance_doc_name / acceptance_docs)",
        "if_empty": "Закрывающий документ не регистрируется",
        "comment":  (
            "Наименование закрывающего документа (УПД, Акт, Накладная и т.п.).\n"
            "Сохраняется в поле acceptance_doc_name и в JSONB-массиве acceptance_docs.\n"
            "Если не заполнено — закрывающий документ не будет зарегистрирован."
        ),
    },
    {
        "header":   "Закрывающий документ: №",
        "required": False,
        "width":    20,
        "fmt":      "145",
        "effect":   "Номер закрывающего документа (acceptance_doc_number / acceptance_docs)",
        "if_empty": "Номер закрывающего документа не сохраняется",
        "comment":  (
            "Номер закрывающего документа (УПД, акта, накладной).\n"
            "Сохраняется в поле acceptance_doc_number и в JSONB-массиве acceptance_docs.\n"
            "Если не заполнено — номер закрывающего документа не указывается."
        ),
    },
    {
        "header":   "Закрывающий документ: дата",
        "required": False,
        "width":    20,
        "fmt":      "08.05.2026",
        "effect":   "Дата закрывающего документа (acceptance_doc_date / acceptance_docs)",
        "if_empty": "Дата закрывающего документа не сохраняется",
        "comment":  (
            "Дата закрывающего документа в формате ДД.ММ.ГГГГ.\n"
            "Сохраняется в поле acceptance_doc_date и в JSONB-массиве acceptance_docs.\n"
            "Если не заполнено — дата закрывающего документа не указывается."
        ),
    },
    {
        "header":   "Закрывающий документ: сумма",
        "required": False,
        "width":    22,
        "fmt":      "490000",
        "effect":   "Сумма по закрывающему документу (acceptance_doc_amount / acceptance_docs)",
        "if_empty": "Сумма закрывающего документа не сохраняется",
        "comment":  (
            "Сумма по закрывающему документу в рублях.\n"
            "Сохраняется в поле acceptance_doc_amount и в JSONB-массиве acceptance_docs.\n"
            "Если не заполнено — сумма закрывающего документа не указывается."
        ),
    },
    {
        "header":   "Предоплата",
        "required": False,
        "width":    14,
        "fmt":      "Да / (пусто)",
        "effect":   "Признак предоплаты (is_prepayment) — влияет на порядок проведения платежей",
        "if_empty": "Признак предоплаты не устанавливается",
        "comment":  (
            "Признак того, что закупка предполагает предоплату. Выберите «Да» из списка.\n"
            "Влияет на логику отображения в разделе платежей и дашборде.\n"
            "Если не заполнено — предоплата не устанавливается."
        ),
        "_dd": "_DD_PREPAYMENT",
    },
    {
        "header":   "Ежемесячный платёж",
        "required": False,
        "width":    18,
        "fmt":      "Да / (пусто)",
        "effect":   "Признак ежемесячного платежа (is_monthly_payment) — для услуг с абонплатой",
        "if_empty": "Признак ежемесячного платежа не устанавливается",
        "comment":  (
            "Признак того, что оплата производится ежемесячными платежами. Выберите «Да».\n"
            "Используется для корректного отображения графика платежей по договору.\n"
            "Если не заполнено — закупка считается единовременной оплатой."
        ),
        "_dd": "_DD_MONTHLY",
    },
    {
        "header":   "Тип позиции (Товар/Услуга)",
        "required": False,
        "width":    24,
        "fmt":      "Товар / Услуга / Работа",
        "effect":   "Тип позиций закупки (item_type) — влияет на аналитику и документы",
        "if_empty": "Тип позиции не указывается (по умолчанию «Товар»)",
        "comment":  (
            "Тип позиции: Товар, Услуга или Работа. Выберите из выпадающего списка.\n"
            "Влияет на аналитику (разбивка товары/услуги) и шаблоны документов.\n"
            "Если не заполнено — тип считается «Товар»."
        ),
        "_dd": "_DD_ITEM_TYPE",
    },
    {
        "header":   "Основание для оплаты",
        "required": False,
        "width":    22,
        "fmt":      "Договор / Счёт / Счёт-договор",
        "effect":   "Основание для оплаты по договору (payment_basis_type) — для документов",
        "if_empty": "Используется основание «Договор» по умолчанию",
        "comment":  (
            "Основание для оплаты: Договор, Счёт или Счёт-договор. Выберите из списка.\n"
            "Используется при генерации договора и платёжных документов.\n"
            "Если не заполнено — применяется значение «Договор»."
        ),
        "_dd": "_DD_PAYMENT_BASIS",
    },
    {
        "header":   "Год",
        "required": False,
        "width":     8,
        "fmt":       "2026",
        "effect":    "Год закупки — используется при фильтрации в реестре",
        "if_empty":  "Год определяется по дате договора или текущему году",
        "comment":   (
            "Год проведения закупки (4 цифры). Используется для фильтрации по году в реестре.\n"
            "Если не заполнено — год определяется по дате договора или текущему году.\n"
            "Это справочное поле, не влияет на сохранение закупки."
        ),
    },
    {
        "header":   "Регион проведения мероприятия",
        "required": False,
        "width":    30,
        "fmt":      "Москва / Федеральное мероприятие",
        "effect":   "Регион проведения мероприятия / оказания услуг (region) — отображается в карточке",
        "if_empty": "Регион мероприятия не указывается",
        "comment":  (
            "Субъект РФ, в котором проводится мероприятие, или спец-значение.\n"
            "Допустимы: «Федеральное мероприятие», «Несколько регионов», «Не определён» или один из 89 субъектов.\n"
            "Если не заполнено — регион мероприятия останется пустым."
        ),
        "_dd": "_DD_EVENT_REGION",
    },
    {
        "header":   "Регион поставки (субъект РФ)",
        "required": False,
        "width":    28,
        "fmt":      "Москва / Московская область",
        "effect":   "Субъект РФ места поставки (delivery_region) — используется для ОКАТО и федерального округа при публикации на ЭТП Фабрикант",
        "if_empty": "Подставится адрес организации субсидии",
        "comment":  (
            "Субъект РФ, куда осуществляется поставка товара или оказывается услуга.\n"
            "Используется для определения ОКАТО и федерального округа при публикации на ЭТП Фабрикант.\n"
            "Если не заполнено — подставится регион из адреса организации субсидии."
        ),
        "_dd": "_DD_DELIVERY_REGION",
    },
    {
        "header":   "Место оказания услуг / доставки",
        "required": False,
        "width":    30,
        "fmt":      "г. Москва, ул. Ленина, 1",
        "effect":   "Место оказания услуг или пункт доставки (delivery_location)",
        "if_empty": "Место оказания услуг не указывается",
        "comment":  (
            "Адрес или описание места, где оказывается услуга или доставляется товар.\n"
            "Используется в документах (договор, ТЗ) в разделе «Место оказания услуг».\n"
            "Если не заполнено — место доставки/услуг останется пустым."
        ),
    },
    {
        "header":   "Адрес доставки",
        "required": False,
        "width":    30,
        "fmt":      "197022, г. Санкт-Петербург, пр. Большой, 5",
        "effect":   "Почтовый адрес доставки (delivery_address)",
        "if_empty": "Адрес доставки не указывается",
        "comment":  (
            "Точный почтовый адрес доставки товара. Используется при формировании товарной накладной.\n"
            "Отображается в карточке закупки в блоке доставки.\n"
            "Если не заполнено — адрес доставки останется пустым."
        ),
    },
    {
        "header":   "Экономия по результатам конкурентных закупок",
        "required": False,
        "width":    32,
        "fmt":      "15000",
        "effect":   "Сумма экономии по результатам конкурентных процедур (economy)",
        "if_empty": "Экономия не фиксируется",
        "comment":  (
            "Сумма экономии в рублях, полученная по результатам конкурентной процедуры.\n"
            "Рассчитывается как разница НМЦК и цены договора; можно указать вручную.\n"
            "Если не заполнено — экономия не будет указана."
        ),
    },
    {
        "header":   "Срок действия договора",
        "required": False,
        "width":    22,
        "fmt":      "31.12.2026",
        "effect":   "Дата окончания срока действия договора (contract_end_date)",
        "if_empty": "Срок действия договора не указывается",
        "comment":  (
            "Дата окончания срока действия договора в формате ДД.ММ.ГГГГ.\n"
            "Отображается в карточке закупки в блоке «Договор и оплата».\n"
            "Если не заполнено — срок действия договора не будет указан."
        ),
    },
    {
        "header":   "Квартал принятия обязательств",
        "required": False,
        "width":    22,
        "fmt":      "1 / 2 / 3 / 4",
        "effect":   "Квартал, в котором принимаются финансовые обязательства (commitment_quarter)",
        "if_empty": "Квартал обязательств не указывается",
        "comment":  (
            "Квартал принятия финансовых обязательств по договору (1, 2, 3 или 4).\n"
            "Используется для планирования финансирования по кварталам.\n"
            "Если не заполнено — квартал обязательств не будет указан."
        ),
        "_dd": "_DD_QUARTER",
    },
    {
        "header":   "Планируемый месяц платежа",
        "required": False,
        "width":    22,
        "fmt":      "01.04.2026",
        "effect":   "Планируемая дата (месяц) платежа по договору (planned_payment_month)",
        "if_empty": "Планируемый месяц платежа не указывается",
        "comment":  (
            "Планируемая дата платежа — укажите первое число нужного месяца (ДД.ММ.ГГГГ).\n"
            "Используется для планирования кассового разрыва и графика платежей.\n"
            "Если не заполнено — планируемый месяц платежа не будет задан."
        ),
    },
    {
        "header":   "Дата окончания приёма заявок",
        "required": False,
        "width":    26,
        "fmt":      "20.03.2026",
        "effect":   "Дата окончания приёма заявок от участников (submission_deadline)",
        "if_empty": "Срок приёма заявок не указывается",
        "comment":  (
            "Дата окончания приёма заявок от участников конкурентной процедуры (ДД.ММ.ГГГГ).\n"
            "Отображается в карточке закупки и используется при публикации на ЭТП.\n"
            "Если не заполнено — срок приёма заявок не будет указан."
        ),
    },
    {
        "header":   "Режим НДС",
        "required": False,
        "width":    24,
        "fmt":      "Одинаковый / Для каждого товара",
        "effect":   "Режим применения НДС к позициям (vat_mode)",
        "if_empty": "Используется режим «Одинаковый» по умолчанию",
        "comment":  (
            "Режим применения НДС: «Одинаковый» — одна ставка для всех позиций; "
            "«Для каждого товара» — ставка задаётся per-позиционно.\n"
            "Влияет на формирование документов со спецификацией.\n"
            "Если не заполнено — применяется режим «Одинаковый»."
        ),
        "_dd": "_DD_VAT_MODE",
    },
    {
        "header":   "Подпись этапа",
        "required": False,
        "width":    20,
        "fmt":      "Февраль 2026",
        "effect":   "Текстовая подпись этапа (stage_label) — отображается в карточке",
        "if_empty": "Подпись этапа не указывается",
        "comment":  (
            "Произвольная текстовая подпись текущего этапа, например «Февраль 2026» или «1-й транш».\n"
            "Отображается рядом со статусом в карточке закупки.\n"
            "Если не заполнено — подпись этапа будет пустой."
        ),
    },
]

# Derived tuples for backward-compatible access
_TEMPLATE_COLUMNS = [(s["header"], s["required"], s["width"]) for s in _COL_SPEC]

# Helper: index of column by header
def _col_spec_index(header: str) -> int:
    for i, s in enumerate(_COL_SPEC, 1):
        if s["header"] == header:
            return i
    return 1


# Example row 1: закупка 20, Заказ 1 (рамочный накопительный, кепи)
_TEMPLATE_EXAMPLE_ROW = [
    "Рамочный (нарастающий итог)",          # Тип договора
    "20",                                   # Номер закупки
    "Заказ 1 (март)",                       # Номер заказа внутри закупки
    "Обмундирование к слёту",               # Предмет договора (общий)
    "Кепи камуфляж",                        # Наименование товара
    "Снаряжение",                           # ФЭО Ур.1
    "Одежда",                               # ФЭО Ур.2
    "Кепи",                                 # ФЭО Ур.3
    "",                                     # ФЭО Ур.4
    "",                                     # ФЭО Ур.5
    "Форма к слёту",                        # Мероприятие
    "ООО Поставщик",                        # Контрагент
    "1234567890",                           # ИНН контрагента
    "Единственный поставщик",               # Способ закупки
    "2026/001",                             # Реестровый №
    "Д-001",                                # № договора
    "15.03.2026",                           # Дата договора
    "600000",                               # Максимальная цена договора
    "30.06.2026",                           # Срок исполнения
    "",                                     # Срок исполнения (изменён)
    "",                                     # Дата доставки
    "Оплачено",                             # Этап закупки
    "",                                     # Подстатус (для «Ведётся работа»)
    "план закупок",                          # Основание закупки
    "Петров П.П.",                          # Ответственное лицо
    "",                                     # Ссылка ЭТП
    "ПП-101",                               # Номер платёжного документа
    "01.04.2026",                           # Дата платёжного документа
    "245000",                               # Сумма оплаты
    "Оплата по Д-001, 1-й платёж",          # Назначение платежа
    "120000",                               # В т.ч. федеральный бюджет
    "100",                                  # Количество (план)
    "шт",                                   # Ед. изм.
    "4900",                                 # Цена за ед. (план)
    "490000",                               # Сумма план
    "100",                                  # Кол-во факт
    "4900",                                 # Цена за ед. (факт)
    "490000",                               # Сумма факт
    "РФ",                                    # Страна происхождения
    "Да",                                   # НДС применяется
    "20",                                   # Ставка НДС
    "",                                     # Статья НК РФ
    "УПД",                                  # Закрывающий документ: наименование
    "145",                                  # Закрывающий документ: №
    "08.05.2026",                           # Закрывающий документ: дата
    "490000",                               # Закрывающий документ: сумма
    "",                                     # Предоплата
    "",                                     # Ежемесячный платёж
    "Товар",                                # Тип позиции (Товар/Услуга)
    "Договор",                              # Основание для оплаты
    "2026",                                 # Год
    "Москва",                               # Регион проведения мероприятия
    "Москва",                               # Регион поставки (субъект РФ)
    "",                                     # Место оказания услуг / доставки
    "",                                     # Адрес доставки
    "10000",                                # Экономия по результатам конкурентных закупок
    "31.12.2026",                           # Срок действия договора
    "2",                                    # Квартал принятия обязательств
    "01.04.2026",                           # Планируемый месяц платежа
    "10.03.2026",                           # Дата окончания приёма заявок
    "Одинаковый",                           # Режим НДС
    "Март 2026",                            # Подпись этапа
]

# Example row 2: закупка 21, разовый — услуга, «Ведётся работа» с подстатусом, без НДС
_TEMPLATE_EXAMPLE_ROW_2 = [
    "Разовая поставка",                     # Тип договора
    "21",                                   # Номер закупки
    "",                                     # Номер заказа внутри закупки
    "Обучение персонала",                   # Предмет договора (общий)
    "Тренинг по охране труда",              # Наименование товара
    "Обучение",                             # ФЭО Ур.1
    "Охрана труда",                         # ФЭО Ур.2
    "",                                     # ФЭО Ур.3
    "",                                     # ФЭО Ур.4
    "",                                     # ФЭО Ур.5
    "",                                     # Мероприятие
    "ИП Сидоров",                           # Контрагент
    "123456789012",                         # ИНН контрагента
    "Запрос котировок",                     # Способ закупки
    "2026/002",                             # Реестровый №
    "Д-002",                                # № договора
    "01.06.2026",                           # Дата договора
    "85000",                                # Максимальная цена договора
    "30.08.2026",                           # Срок исполнения
    "",                                     # Срок исполнения (изменён)
    "",                                     # Дата доставки
    "Ведётся работа",                       # Этап закупки
    "Сбор КП",                              # Подстатус (для «Ведётся работа»)
    "служебная записка",                    # Основание закупки
    "Иванова А.А.",                         # Ответственное лицо
    "",                                     # Ссылка ЭТП
    "",                                     # Номер платёжного документа
    "",                                     # Дата платёжного документа
    "",                                     # Сумма оплаты
    "",                                     # Назначение платежа
    "",                                     # В т.ч. федеральный бюджет
    "1",                                    # Количество (план)
    "усл.",                                 # Ед. изм.
    "85000",                                # Цена за ед. (план)
    "85000",                                # Сумма план
    "1",                                    # Кол-во факт
    "85000",                                # Цена за ед. (факт)
    "85000",                                # Сумма факт
    "",                                     # Страна происхождения
    "",                                     # НДС применяется
    "",                                     # Ставка НДС
    "ст. 149 п. 3 пп. 15",                 # Статья НК РФ
    "",                                     # Закрывающий документ: наименование
    "",                                     # Закрывающий документ: №
    "",                                     # Закрывающий документ: дата
    "",                                     # Закрывающий документ: сумма
    "",                                     # Предоплата
    "",                                     # Ежемесячный платёж
    "Услуга",                               # Тип позиции (Товар/Услуга)
    "Счёт",                                 # Основание для оплаты
    "2026",                                 # Год
    "Санкт-Петербург",                      # Регион проведения мероприятия
    "Санкт-Петербург",                      # Регион поставки (субъект РФ)
    "г. Санкт-Петербург, пр. Большой, 5",  # Место оказания услуг / доставки
    "",                                     # Адрес доставки
    "",                                     # Экономия по результатам конкурентных закупок
    "31.08.2026",                           # Срок действия договора
    "3",                                    # Квартал принятия обязательств
    "01.08.2026",                           # Планируемый месяц платежа
    "",                                     # Дата окончания приёма заявок
    "Одинаковый",                           # Режим НДС
    "",                                     # Подпись этапа
]
def _build_dv_prompt(spec: dict) -> tuple[str, str]:
    """Строит (promptTitle, prompt) для DataValidation, соблюдая лимиты Excel:
    promptTitle <= 32 символа, prompt <= 255 символов (обрезка по границе слова)."""

    # --- promptTitle: заголовок колонки, макс 32 символа ---
    title_raw = spec["header"]
    prompt_title = title_raw if len(title_raw) <= 32 else title_raw[:32]

    # --- prompt: короткий текст из полей spec ---
    parts: list[str] = []
    if spec.get("required"):
        parts.append("Обязательное поле.")

    # Первое предложение/строка comment
    comment_raw = spec.get("comment", "")
    first_sentence = comment_raw.split("\n")[0].split(". ")[0]
    if first_sentence and not first_sentence.endswith("."):
        first_sentence += "."
    if first_sentence:
        parts.append(first_sentence)

    if spec.get("if_empty"):
        parts.append(f"Если не заполнено: {spec['if_empty']}")

    if spec.get("_dd"):
        parts.append("Можно выбрать из списка или вписать своё.")

    full_prompt = "\n".join(parts)

    # Обрезка до 255 символов по границе слова
    MAX_PROMPT = 255
    if len(full_prompt) > MAX_PROMPT:
        cut = full_prompt[:MAX_PROMPT - 1]
        # ищем последний пробел или перенос
        boundary = max(cut.rfind(" "), cut.rfind("\n"))
        if boundary > 0:
            cut = cut[:boundary]
        full_prompt = cut.rstrip() + "…"

    return prompt_title, full_prompt


@router.get("/import/template")
async def download_import_template(
    subsidy_id: Optional[int] = Query(None, description="ID субсидии для каскадных ФЭО-списков"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Скачать шаблон Excel для импорта закупок (лист «Закупки» + «Справочники» + «Справочник колонок»).
    При subsidy_id — добавляет дерево ФЭО субсидии с выпадающими списками на листе «Закупки»."""
    if Workbook is None:
        raise HTTPException(500, "openpyxl не установлен")

    # Проверяем доступ к субсидии, если она указана
    if subsidy_id is not None:
        vis = await get_visible_subsidy_ids(current_user, db, "purchases")
        if vis is not None and subsidy_id not in vis:
            raise HTTPException(
                403,
                f"Субсидия #{subsidy_id} вам недоступна или не существует. "
                "Проверьте права доступа к субсидии."
            )

    subsidy_name: Optional[str] = None
    if subsidy_id is not None:
        subsidy_name = (await db.execute(select(Subsidy.name).where(Subsidy.id == subsidy_id))).scalar_one_or_none()
    # Причина, по которой каскада ФЭО в файле нет (если такое случится) — покажем её прямо в файле.
    feo_warning: Optional[str] = None
    if subsidy_id is None:
        feo_warning = (
            "Шаблон скачан без выбранной субсидии — направления расходов (ФЭО) в него не подставлены "
            "и связанные списки Ур.1→Ур.5 не работают. Выберите субсидию в диалоге импорта и скачайте шаблон заново."
        )

    wb = Workbook()

    fill_req  = PatternFill(start_color="DC2626", end_color="DC2626", fill_type="solid")
    fill_opt  = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    font_hdr  = Font(color="FFFFFF", bold=True, size=11)
    align_c   = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # =========================================================================
    # Sheet «Справочники» — создаём ДО «Закупки», чтобы потом добавить DataValidation
    # =========================================================================
    # Карта: имя переменной → (defined_name, header_label, dd_list)
    _DD_REGISTRY = {
        "_DD_CONTRACT_TYPE":   ("dd_contract_type",  "Тип договора",             _DD_CONTRACT_TYPE),
        "_DD_STATUS":          ("dd_status",          "Этап закупки",             _DD_STATUS),
        "_DD_SUBSTATUS":       ("dd_substatus",       "Подстатус",                _DD_SUBSTATUS),
        "_DD_METHOD":          ("dd_method",          "Способ закупки",           _DD_METHOD),
        "_DD_BASIS":           ("dd_basis",           "Основание закупки",        _DD_BASIS),
        "_DD_VAT_APPLICABLE":  ("dd_vat_applicable",  "НДС применяется",         _DD_VAT_APPLICABLE),
        "_DD_VAT_RATE":        ("dd_vat_rate",        "Ставка НДС",               _DD_VAT_RATE),
        "_DD_PREPAYMENT":      ("dd_prepayment",      "Предоплата",               _DD_PREPAYMENT),
        "_DD_MONTHLY":         ("dd_monthly",         "Ежемесячный платёж",       _DD_MONTHLY),
        "_DD_ITEM_TYPE":       ("dd_item_type",       "Тип позиции",              _DD_ITEM_TYPE),
        "_DD_PAYMENT_BASIS":   ("dd_payment_basis",   "Основание для оплаты",     _DD_PAYMENT_BASIS),
        "_DD_UNIT":            ("dd_unit",            "Ед. изм.",                 _DD_UNIT),
        "_DD_QUARTER":         ("dd_quarter",         "Квартал обязательств",     _DD_QUARTER),
        "_DD_VAT_MODE":        ("dd_vat_mode",        "Режим НДС",                _DD_VAT_MODE),
        "_DD_DELIVERY_REGION": ("dd_delivery_region", "Регион поставки",          _DD_DELIVERY_REGION),
        "_DD_EVENT_REGION":    ("dd_event_region",    "Регион мероприятия",       _DD_EVENT_REGION),
    }

    # Добавляем лист «Справочники» как второй (будет Sheet2)
    # Сначала создаём активный лист, потом переименуем
    wb_ref_ws = wb.create_sheet(title="Справочники")

    ref_hdr_font  = Font(bold=True, size=10)
    ref_hdr_fill  = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    ref_val_align = Alignment(wrap_text=False, vertical="top")
    ref_hint_font = Font(italic=True, color="6B7280", size=9)

    # Строка-подсказка в строке 1
    hint_cell = wb_ref_ws.cell(1, 1,
        "Допишите свои значения в пустые ячейки под списком — они появятся в выпадающих списках на листе «Закупки»")
    hint_cell.font = ref_hint_font
    hint_cell.alignment = Alignment(wrap_text=True)
    wb_ref_ws.row_dimensions[1].height = 28

    # Словарь: dd_var_name → (col_letter, first_data_row, last_data_row)
    # для построения формул DataValidation
    _dd_ref_ranges: dict[str, str] = {}

    # Каждый справочник — отдельная колонка, начиная со строки 2
    for col_i, (dd_var_name, (dn_name, hdr_label, dd_list)) in enumerate(_DD_REGISTRY.items(), 1):
        # Заголовок колонки
        hdr_cell = wb_ref_ws.cell(2, col_i, hdr_label)
        hdr_cell.font = ref_hdr_font
        hdr_cell.fill = ref_hdr_fill
        hdr_cell.alignment = Alignment(horizontal="center")

        n_values = len(dd_list)
        # Диапазон = 2× числа значений (пользователь может дописать в пустые ячейки)
        slot_count = n_values * 2

        # Записываем значения начиная со строки 3
        for row_j, (label, _key) in enumerate(dd_list, 1):
            wb_ref_ws.cell(2 + row_j, col_i, label).alignment = ref_val_align

        col_letter = wb_ref_ws.cell(1, col_i).column_letter
        first_data_row = 3               # строка 3 = первое значение
        last_data_row  = 2 + slot_count  # = строка 3 + (slot_count-1)

        # Именованный диапазон уровня книги (надёжнее прямой ссылки на лист в DV)
        if DefinedName is not None:
            attr_text = f"'Справочники'!${col_letter}${first_data_row}:${col_letter}${last_data_row}"
            wb.defined_names[dn_name] = DefinedName(dn_name, attr_text=attr_text)

        # Ключ → имя именованного диапазона для DataValidation
        _dd_ref_ranges[dd_var_name] = dn_name

        # Ширина колонки
        max_len = max((len(lbl) for lbl, _ in dd_list), default=10)
        wb_ref_ws.column_dimensions[col_letter].width = max(max_len + 4, len(hdr_label) + 2, 16)

    wb_ref_ws.freeze_panes = "A3"

    # =========================================================================
    # Sheet 1: «Закупки»
    # =========================================================================
    ws = wb.active
    ws.title = "Закупки"

    headers = [s["header"] for s in _COL_SPEC]
    ws.append(headers)

    for i, spec in enumerate(_COL_SPEC, 1):
        c = ws.cell(1, i)
        c.fill = fill_req if spec["required"] else fill_opt
        c.font = font_hdr
        c.alignment = align_c

    ws.append(_TEMPLATE_EXAMPLE_ROW)
    ws.append(_TEMPLATE_EXAMPLE_ROW_2)

    for i, spec in enumerate(_COL_SPEC, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = spec["width"]

    ws.row_dimensions[1].height = 38
    ws.freeze_panes = "A2"

    # ---- DataValidation: два DV на каждую колонку — шапка (строка 1) и данные (2:1000) ----
    # Два непересекающихся диапазона не конфликтуют в Excel.
    # Шапка: только подсказка (type=None), данные: подсказка + список (если есть).
    if DataValidation is not None:
        for col_i, spec in enumerate(_COL_SPEC, 1):
            col_letter = ws.cell(1, col_i).column_letter
            dd_var = spec.get("_dd")
            ref_formula = _dd_ref_ranges.get(dd_var) if dd_var else None

            prompt_title, prompt_text = _build_dv_prompt(spec)

            # DV на шапку (строка 1): только подсказка, без списка
            dv_hdr = DataValidation(
                type=None,
                showInputMessage=True,
            )
            dv_hdr.promptTitle = prompt_title
            dv_hdr.prompt      = prompt_text
            dv_hdr.sqref       = f"{col_letter}1"
            ws.add_data_validation(dv_hdr)

            # DV на строки данных (2:1000): подсказка + список если есть
            if ref_formula:
                dv_data = DataValidation(
                    type="list",
                    formula1=f"={ref_formula}",
                    allow_blank=True,
                    showErrorMessage=False,
                    showInputMessage=True,
                )
                dv_data.error      = "Значение не из списка — будет принято как есть"
                dv_data.errorTitle = "Нестандартное значение"
            else:
                dv_data = DataValidation(
                    type=None,
                    showInputMessage=True,
                )
            dv_data.promptTitle = prompt_title
            dv_data.prompt      = prompt_text
            dv_data.sqref       = f"{col_letter}2:{col_letter}1000"
            ws.add_data_validation(dv_data)

    # =========================================================================
    # Sheet «Справочник колонок»
    # =========================================================================
    ref_ws = wb.create_sheet(title="Справочник колонок")

    ref_headers = ["Колонка", "Обязательная", "Формат/пример", "На что влияет", "Что будет, если не заполнить"]
    ref_ws.append(ref_headers)

    ref_hdr_fill2 = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    for cell in ref_ws[1]:
        cell.font = Font(color="FFFFFF", bold=True, size=11)
        cell.fill = ref_hdr_fill2
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    fill_req_row = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    align_wrap = Alignment(wrap_text=True, vertical="top")

    # Строки колонок «ФЭО Ур.N» на этом листе — запоминаем, чтобы после построения каскада
    # (ниже по коду) дописать в «Формат/пример», что это связанный список.
    _feo_fmt_rows: dict[int, int] = {}
    _feo_headers_set = {"ФЭО Ур.1", "ФЭО Ур.2", "ФЭО Ур.3", "ФЭО Ур.4", "ФЭО Ур.5"}

    for spec in _COL_SPEC:
        fmt_val = spec["fmt"]
        # Если у колонки есть dd — добавляем ключи в формат
        if spec.get("_dd"):
            fmt_val += " (выпадающий список)"
        row_vals = [
            spec["header"],
            "Да" if spec["required"] else "Нет",
            fmt_val,
            spec["effect"],
            spec["if_empty"],
        ]
        ref_ws.append(row_vals)
        row_idx = ref_ws.max_row
        if spec["header"] in _feo_headers_set:
            _feo_fmt_rows[int(spec["header"][-1])] = row_idx
        if spec["required"]:
            for col_idx2 in range(1, 6):
                ref_ws.cell(row_idx, col_idx2).fill = fill_req_row
        for col_idx2 in range(1, 6):
            ref_ws.cell(row_idx, col_idx2).alignment = align_wrap

    # Column widths on reference sheet
    ref_col_widths = [34, 13, 34, 50, 45]
    for i, w in enumerate(ref_col_widths, 1):
        ref_ws.column_dimensions[ref_ws.cell(1, i).column_letter].width = w

    ref_ws.row_dimensions[1].height = 28
    ref_ws.freeze_panes = "A2"

    # =========================================================================
    # Каскадные ФЭО-списки (только при subsidy_id)
    # =========================================================================
    # Реализован настоящий INDIRECT-каскад через скрытые хелпер-колонки:
    #   Ур.1 = feo_roots (корни субсидии)
    #   Ур.2 = INDIRECT("feo_" & helper1)  — дети выбранного Ур.1
    #   Ур.3 = INDIRECT("feo_" & helper2)  — дети выбранного Ур.2
    #   Ур.4 = INDIRECT("feo_" & helper3)  — дети выбранного Ур.3
    #   Ур.5 = INDIRECT("feo_" & helper4)  — дети выбранного Ур.4
    # Хелпер L для строки r: =IFERROR(INDEX(lvlL_ids, MATCH(<ФЭО-ячейка L>, lvlL_names, 0)), "")
    # Примечание: при дублях имён внутри уровня MATCH возьмёт первое совпадение (допустимое ограничение).
    feo_dd_levels: set[int] = set()  # уровни, для которых реально построен связанный список (для «Справочник колонок»)
    if subsidy_id is not None and DataValidation is not None and DefinedName is not None:
        try:
            feo_q = select(FeoCategory).where(FeoCategory.subsidy_id == subsidy_id).order_by(
                FeoCategory.sort_order.is_(None), FeoCategory.sort_order, FeoCategory.id
            )
            feo_all_cats = (await db.execute(feo_q)).scalars().all()

            if not feo_all_cats:
                feo_warning = (
                    f"У субсидии «{subsidy_name or subsidy_id}» не заполнено дерево направлений расходов (ФЭО), "
                    "поэтому связанных списков в шаблоне нет. Заполните направления в карточке субсидии."
                )
            else:
                # Уровень считаем по цепочке родителей, а не по колонке level (она может быть
                # NULL/0-based/рассинхронизирована с parent_id — тогда фильтр по level молчаливо
                # ничего не находит и каждый уровень тихо получает fallback на корни).
                by_id = {c.id: c for c in feo_all_cats}

                def _depth_of(cat) -> int:
                    d, cur, guard = 1, cat, 0
                    while cur.parent_id in by_id and guard < 12:
                        cur = by_id[cur.parent_id]
                        d += 1
                        guard += 1
                    return d

                cats_by_depth: dict[int, list] = {}
                for cat in feo_all_cats:
                    cats_by_depth.setdefault(_depth_of(cat), []).append(cat)

                roots = cats_by_depth.get(1, [])
                root_names = [c.name for c in roots]

                if not roots:
                    feo_warning = (
                        f"В дереве направлений расходов (ФЭО) субсидии «{subsidy_name or subsidy_id}» не найдено "
                        "ни одного корневого узла — связанные списки в шаблоне не построены."
                    )
                else:
                    max_depth = max(cats_by_depth)

                    # Строим карту детей: parent_id → [children] (для служебных колонок «ФЭО дети <id>»)
                    children_map: dict[int, list] = {}
                    for cat in feo_all_cats:
                        if cat.parent_id is not None and cat.parent_id in by_id:
                            children_map.setdefault(cat.parent_id, []).append(cat)

                    # Находим последнюю занятую колонку «Справочники»
                    feo_ref_start_col = len(_DD_REGISTRY) + 1

                    # ---------------------------------------------------------------
                    # 1. Колонка корней (Ур.1) + defined name feo_roots — видимая, понятная пользователю
                    # ---------------------------------------------------------------
                    feo_root_col = feo_ref_start_col
                    feo_root_col_letter = wb_ref_ws.cell(1, feo_root_col).column_letter
                    root_hdr_label = f"Направления расходов — {(subsidy_name or '')[:40]} (Ур.1)"
                    wb_ref_ws.cell(2, feo_root_col, root_hdr_label).font = ref_hdr_font
                    wb_ref_ws.cell(2, feo_root_col).fill = ref_hdr_fill
                    for ri, name in enumerate(root_names, 1):
                        wb_ref_ws.cell(2 + ri, feo_root_col, name).alignment = ref_val_align
                    # Диапазон без пустого хвоста — ровно столько строк, сколько корневых направлений
                    feo_root_last = 2 + len(root_names)
                    feo_root_attr = f"'Справочники'!${feo_root_col_letter}$3:${feo_root_col_letter}${feo_root_last}"
                    wb.defined_names["feo_roots"] = DefinedName("feo_roots", attr_text=feo_root_attr)
                    wb_ref_ws.column_dimensions[feo_root_col_letter].width = 36

                    # ---------------------------------------------------------------
                    # 2. Для каждого узла с детьми — отдельная СЛУЖЕБНАЯ (скрытая) колонка + defined name feo_<id>
                    # ---------------------------------------------------------------
                    feo_node_next_col = feo_ref_start_col + 1  # следующая свободная колонка
                    for parent_cat in feo_all_cats:
                        kids = children_map.get(parent_cat.id)
                        if not kids:
                            continue
                        kid_names = [k.name for k in kids]
                        nc = feo_node_next_col
                        nc_letter = wb_ref_ws.cell(1, nc).column_letter
                        hdr_label_node = f"ФЭО дети {parent_cat.id}"
                        wb_ref_ws.cell(2, nc, hdr_label_node).font = ref_hdr_font
                        wb_ref_ws.cell(2, nc).fill = ref_hdr_fill
                        for ki, kname in enumerate(kid_names, 1):
                            wb_ref_ws.cell(2 + ki, nc, kname).alignment = ref_val_align
                        node_last = 2 + len(kid_names)
                        node_attr = f"'Справочники'!${nc_letter}$3:${nc_letter}${node_last}"
                        dn_node = f"feo_{parent_cat.id}"
                        wb.defined_names[dn_node] = DefinedName(dn_node, attr_text=node_attr)
                        wb_ref_ws.column_dimensions[nc_letter].width = 32
                        wb_ref_ws.column_dimensions[nc_letter].hidden = True
                        feo_node_next_col += 1

                    # ---------------------------------------------------------------
                    # 3. Для каждой глубины 1..4: колонка имён (видимая) + id (служебная, скрытая)
                    #    Нужны для хелперов INDEX/MATCH: по имени выбранной ячейки → id узла
                    # ---------------------------------------------------------------
                    lvl_cols: dict[int, dict] = {}  # глубина → {names_col, ids_col, names_letter, ids_letter}
                    for lvl in range(1, 5):
                        cats_at_lvl = cats_by_depth.get(lvl)
                        if not cats_at_lvl:
                            continue
                        names_col = feo_node_next_col
                        ids_col   = feo_node_next_col + 1
                        names_letter = wb_ref_ws.cell(1, names_col).column_letter
                        ids_letter   = wb_ref_ws.cell(1, ids_col).column_letter

                        wb_ref_ws.cell(2, names_col, f"Направления Ур.{lvl} (все)").font = ref_hdr_font
                        wb_ref_ws.cell(2, names_col).fill = ref_hdr_fill
                        wb_ref_ws.cell(2, ids_col, f"Ур.{lvl} id (служ.)").font = ref_hdr_font
                        wb_ref_ws.cell(2, ids_col).fill = ref_hdr_fill

                        for ri, cat in enumerate(cats_at_lvl, 1):
                            wb_ref_ws.cell(2 + ri, names_col, cat.name).alignment = ref_val_align
                            wb_ref_ws.cell(2 + ri, ids_col, cat.id).alignment = ref_val_align

                        lvl_last = 2 + len(cats_at_lvl)
                        names_attr = f"'Справочники'!${names_letter}$3:${names_letter}${lvl_last}"
                        ids_attr   = f"'Справочники'!${ids_letter}$3:${ids_letter}${lvl_last}"
                        wb.defined_names[f"lvl{lvl}_names"] = DefinedName(f"lvl{lvl}_names", attr_text=names_attr)
                        wb.defined_names[f"lvl{lvl}_ids"]   = DefinedName(f"lvl{lvl}_ids",   attr_text=ids_attr)

                        wb_ref_ws.column_dimensions[names_letter].width = 30
                        wb_ref_ws.column_dimensions[ids_letter].width = 10
                        wb_ref_ws.column_dimensions[ids_letter].hidden = True

                        lvl_cols[lvl] = {
                            "names_col": names_col, "ids_col": ids_col,
                            "names_letter": names_letter, "ids_letter": ids_letter,
                            "lvl_last": lvl_last,
                        }
                        feo_node_next_col += 2

                    # ---------------------------------------------------------------
                    # 4. Скрытые хелпер-колонки на листе «Закупки» (col 70+)
                    #    helperL → id выбранного узла уровня L в этой строке
                    # ---------------------------------------------------------------
                    # Находим колонки ФЭО Ур.1..5 на листе «Закупки»
                    feo_headers_map: dict[int, str] = {}  # уровень → col_letter
                    for feo_col_i, spec in enumerate(_COL_SPEC, 1):
                        hdr = spec["header"]
                        if hdr in ("ФЭО Ур.1", "ФЭО Ур.2", "ФЭО Ур.3", "ФЭО Ур.4", "ФЭО Ур.5"):
                            lvl_num = int(hdr[-1])
                            feo_headers_map[lvl_num] = ws.cell(1, feo_col_i).column_letter

                    # Хелпер-колонки начинаем с col=70 (заведомо за _COL_SPEC)
                    helper_start_col = 70
                    helper_cols: dict[int, str] = {}  # уровень → col_letter хелпера
                    from openpyxl.utils import get_column_letter

                    for lvl in range(1, 5):
                        # Хелпер уровня lvl нужен только чтобы вести список СЛЕДУЮЩЕГО уровня —
                        # если lvl уже листовой (глубже дерева нет), строить его не нужно: он бы
                        # ссылался на defined name feo_<id> для узла без детей, которого не существует.
                        if lvl not in lvl_cols or lvl not in feo_headers_map or lvl >= max_depth:
                            continue
                        h_col = helper_start_col + (lvl - 1)
                        h_letter = get_column_letter(h_col)
                        helper_cols[lvl] = h_letter

                        feo_letter = feo_headers_map[lvl]
                        names_dn = f"lvl{lvl}_names"
                        ids_dn   = f"lvl{lvl}_ids"

                        # Формулы для строк 2..1000
                        for r in range(2, 1001):
                            formula = (
                                f"=IFERROR(INDEX({ids_dn},"
                                f"MATCH({feo_letter}{r},{names_dn},0)),\"\")"
                            )
                            ws.cell(r, h_col, formula)

                        # Скрываем хелпер-колонку
                        ws.column_dimensions[h_letter].hidden = True

                    # ---------------------------------------------------------------
                    # 5. DataValidation на листе «Закупки»
                    # Заменяем ранее добавленный DV (из основного цикла) на каскадный.
                    # Ровно ОДИН DataValidation на каждую ФЭО-колонку.
                    # Ур.1 → всегда feo_roots; Ур.N>1 → INDIRECT от хелпера предыдущего уровня,
                    # ЕСЛИ хелпер есть; иначе (в т.ч. N > глубины дерева субсидии) — списка НЕТ,
                    # только подсказка, что уровня в дереве субсидии нет. Никакого fallback на корни.
                    # ---------------------------------------------------------------
                    feo_headers_list = ["ФЭО Ур.1", "ФЭО Ур.2", "ФЭО Ур.3", "ФЭО Ур.4", "ФЭО Ур.5"]
                    for feo_col_i, spec in enumerate(_COL_SPEC, 1):
                        if spec["header"] not in feo_headers_list:
                            continue
                        level_num = int(spec["header"][-1])
                        col_letter = ws.cell(1, feo_col_i).column_letter

                        dn_formula = None
                        if level_num == 1:
                            dn_formula = "=feo_roots"
                        else:
                            prev_lvl = level_num - 1
                            h_letter = helper_cols.get(prev_lvl)
                            if h_letter:
                                # Относительная ссылка на хелпер (без фиксации строки)
                                # Excel применит её построчно в диапазоне sqref
                                dn_formula = f'=INDIRECT("feo_"&${h_letter}2)'

                        # Удаляем ранее добавленный DV данных этой колонки из основного цикла
                        # (sqref данных основного цикла = "{col_letter}2:{col_letter}1000")
                        # DV шапки ("{col_letter}1") не трогаем — он нужен пользователю.
                        old_data_sqref = f"{col_letter}2:{col_letter}1000"
                        ws.data_validations.dataValidation = [
                            dv for dv in ws.data_validations.dataValidation
                            if str(dv.sqref) != old_data_sqref
                        ]

                        base_prompt_title, base_prompt_text = _build_dv_prompt(spec)

                        if dn_formula:
                            # Строим prompt: базовый из _build_dv_prompt + строка о ФЭО
                            extra_line = f"Выберите категорию ФЭО уровня {level_num}."
                            candidate_prompt = (
                                (base_prompt_text + "\n" + extra_line).strip() if base_prompt_text else extra_line
                            )
                            final_prompt = candidate_prompt if len(candidate_prompt) <= 255 else base_prompt_text

                            feo_dv = DataValidation(
                                type="list",
                                formula1=dn_formula,
                                allow_blank=True,
                                showErrorMessage=False,
                                showInputMessage=True,
                            )
                            feo_dv.promptTitle = base_prompt_title
                            feo_dv.prompt      = final_prompt
                            feo_dv.sqref       = f"{col_letter}2:{col_letter}1000"
                            ws.add_data_validation(feo_dv)
                            feo_dd_levels.add(level_num)
                        else:
                            # В дереве субсидии нет такого уровня — списка нет, только подсказка
                            extra_line = (
                                f"В дереве субсидии «{subsidy_name or subsidy_id}» нет уровня {level_num} — "
                                "колонку можно оставить пустой."
                            )
                            candidate_prompt = (
                                (base_prompt_text + "\n" + extra_line).strip() if base_prompt_text else extra_line
                            )
                            final_prompt = candidate_prompt if len(candidate_prompt) <= 255 else base_prompt_text

                            feo_dv = DataValidation(
                                type=None,
                                showInputMessage=True,
                            )
                            feo_dv.promptTitle = base_prompt_title
                            feo_dv.prompt      = final_prompt
                            feo_dv.sqref       = f"{col_letter}2:{col_letter}1000"
                            ws.add_data_validation(feo_dv)

        except Exception as exc:
            logger.exception("Не удалось построить каскад ФЭО для субсидии %s", subsidy_id)
            feo_warning = (
                f"Связанные списки ФЭО не построены из-за ошибки на сервере: {type(exc).__name__}: {exc}. "
                "Сообщите администратору — в логах бэкенда есть подробности."
            )

    # Дописываем в «Справочник колонок», что колонки ФЭО с реально построенным списком — связанные
    if feo_dd_levels:
        note = " (выпадающий список направлений субсидии, привязан к предыдущему уровню)"
        for lvl_num, row_idx in _feo_fmt_rows.items():
            if lvl_num in feo_dd_levels:
                cell = ref_ws.cell(row_idx, 3)
                cell.value = (cell.value or "") + note

    # Если каскада ФЭО в файле нет (или он не построился из-за ошибки) — делаем это заметным
    # прямо в файле, а не только в API/логах, чтобы пользователь не листал пустые списки молча.
    if feo_warning:
        logger.warning("Шаблон импорта закупок отдан без каскада ФЭО (subsidy_id=%s): %s", subsidy_id, feo_warning)

        # На листе «Справочники»: дописываем предупреждение к подсказке в A1
        warn_cell = wb_ref_ws.cell(1, 1)
        base_hint = (
            "Допишите свои значения в пустые ячейки под списком — они появятся в выпадающих списках "
            "на листе «Закупки»"
        )
        warn_cell.value = base_hint + "\n⚠ " + feo_warning
        warn_cell.font = Font(bold=True, color="B91C1C", size=10)
        warn_cell.alignment = Alignment(wrap_text=True)
        wb_ref_ws.row_dimensions[1].height = 46

        # На листе «Закупки»: серая заливка шапки + предупреждение в подсказке DV колонок ФЭО Ур.1..5
        fill_warn = PatternFill(start_color="9CA3AF", end_color="9CA3AF", fill_type="solid")
        feo_headers_list = ["ФЭО Ур.1", "ФЭО Ур.2", "ФЭО Ур.3", "ФЭО Ур.4", "ФЭО Ур.5"]
        for feo_col_i, spec in enumerate(_COL_SPEC, 1):
            if spec["header"] not in feo_headers_list:
                continue
            col_letter = ws.cell(1, feo_col_i).column_letter
            ws.cell(1, feo_col_i).fill = fill_warn
            for dv in ws.data_validations.dataValidation:
                if str(dv.sqref) != f"{col_letter}2:{col_letter}1000":
                    continue
                combined = (dv.prompt + "\n⚠ " + feo_warning) if dv.prompt else ("⚠ " + feo_warning)
                if len(combined) > 255:
                    cut = combined[:254]
                    boundary = max(cut.rfind(" "), cut.rfind("\n"))
                    if boundary > 0:
                        cut = cut[:boundary]
                    combined = cut.rstrip() + "…"
                dv.prompt = combined

    # Сохраняем в BytesIO
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{quote('Шаблон_импорта_закупок.xlsx', safe='-_.~')}"}
    )


# ---------------------------------------------------------------------------
# Import helpers & COLUMN_MAP
# ---------------------------------------------------------------------------

# Full column map: lower-stripped header → internal field name.
# New template keys + old keys for backward compat.
_COLUMN_MAP: Dict[str, str] = {
    # --- New template: «Закупки» sheet ---
    "тип договора":                                              "contract_type_raw",
    "тип договора (разовый/рамочный)":                          "contract_type_raw",
    "номер закупки":                                             "purchase_group_num",
    "номер закупки (группа)":                                   "purchase_group_num",
    "номер заказа внутри закупки":                              "order_number",
    "номер заказа внутри закупки или период оказания услуг":   "order_number",
    "номер заказа":                                              "order_number",
    "предмет договора (общий)":                                 "subject",
    "предмет договора":                                         "subject",
    "наименование товара":                                       "item_name",
    "максимальная цена договора":                               "contract_price",
    "максимальная цена договора (предел рамочного)":           "contract_price",
    "наименование позиции":         "item_name",
    "наименование позиции *":       "item_name",
    "тип (товар/услуга/работа)":    "item_type",
    "тип":                          "item_type",
    # New 5-level FEO columns
    "фэо ур.1":                     "feo_l1",
    "фэо ур.2":                     "feo_l2",
    "фэо ур.3":                     "feo_l3",
    "фэо ур.4":                     "feo_l4",
    "фэо ур.5":                     "feo_l5",
    # Old single-path FEO column (backward compat)
    "категория фэо (полный путь) *": "feo_path",
    "категория фэо (полный путь)":  "feo_path",
    "категория фэо *":              "feo_path",
    "категория фэо":                "feo_category_name",
    "мероприятие":                  "event_name",
    "контрагент":                   "contractor_name",
    "инн контрагента *":            "contractor_inn",
    "инн контрагента":              "contractor_inn",
    "способ закупки (еи/кп)":       "purchase_method",
    "способ закупки":               "purchase_method",
    "реестровый №":                 "registry_number",
    "реестровый номер":             "registry_number",
    "реестр. №":                    "registry_number",
    "№ договора *":                 "contract_number",
    "№ договора":                   "contract_number",
    "дата договора *":              "contract_date",
    "дата договора":                "contract_date",
    "цена договора (итого)":        "contract_price",
    "цена договора":                "contract_price",
    "срок исполнения":              "execution_term",
    "статус":                       "status",
    "этап закупки":                 "status",   # новый заголовок (обратная совместимость через оба)
    "этап":                         "status",
    "подстатус (для «ведётся работа»)": "substatus",
    "подстатус":                    "substatus",
    "тип позиции (товар/услуга)":   "item_type",
    "основание для оплаты":         "payment_basis_type",
    "количество (план) *":          "plan_qty",
    "количество (план)":            "plan_qty",
    "ед. изм.":                     "unit",
    "ед. изм":                      "unit",
    "цена за ед. (план) *":         "plan_unit_price",
    "цена за ед. (план)":           "plan_unit_price",
    "сумма план":                   "plan_total",
    "кол-во факт":                  "fact_qty",
    "цена за ед. (факт)":           "fact_unit_price",
    "сумма факт *":                 "fact_total",
    "сумма факт":                   "fact_total",
    "страна происхождения":         "country_origin",
    "ставка ндс":                   "vat_rate",
    "год":                          "year",
    # --- Inline payment columns (new template) ---
    "номер платёжного документа":   "payment_doc_number",
    "номер платежного документа":   "payment_doc_number",
    "дата платёжного документа":    "payment_doc_date",
    "дата платежного документа":    "payment_doc_date",
    "назначение платежа":           "payment_purpose",
    # --- Removed columns from old template (keep for backward compat) ---
    "№ пп *":                       "payment_doc_number",
    "№ пп":                         "payment_doc_number",
    "дата платежа *":               "payment_doc_date",
    "дата платежа":                 "payment_doc_date",
    "сумма оплаты *":               "payment_amount",
    "сумма оплаты":                 "payment_amount",
    "в т.ч. федеральный бюджет":   "payment_federal",
    "в т.ч. фед. бюджет":          "payment_federal",
    # --- Old 17-col backward compat ---
    "наименование":                 "item_name",
    "предмет закупки":              "item_name",
    "субсидия":                     "subsidy_name",
    "инн":                          "contractor_inn",
    "нмцк":                         "nmck",
    "сумма":                        "nmck",
    "цена":                         "nmck",
    "способ":                       "purchase_method",
    "номер договора":               "contract_number",
    "пп №":                         "payment_doc_number",
    "пп номер":                     "payment_doc_number",
    "пп дата":                      "payment_doc_date",
    "оплачено":                     "payment_amount",
    "ссылка этп":                   "etp_url",
    "этп":                          "etp_url",
    # --- New template extended columns ---
    "срок исполнения (изменён)":    "execution_term_changed",
    "срок (изменён)":               "execution_term_changed",
    "дата доставки":                "delivery_date",
    "основание закупки":            "purchase_basis",
    "основание":                    "purchase_basis",
    "ответственное лицо":           "responsible_person",
    "ответственный":                "responsible_person",
    "ндс применяется":              "vat_applicable",
    "статья нк рф":                 "vat_exemption_article",
    "статья нк":                    "vat_exemption_article",
    "закрывающий документ: наименование": "acceptance_doc_name",
    "закрывающий документ: наим":   "acceptance_doc_name",
    "наименование закрывающего документа": "acceptance_doc_name",
    "закрывающий документ: №":      "acceptance_doc_number",
    "закрывающий документ: номер":  "acceptance_doc_number",
    "номер закрывающего документа": "acceptance_doc_number",
    "закрывающий документ: дата":   "acceptance_doc_date",
    "дата закрывающего документа":  "acceptance_doc_date",
    "закрывающий документ: сумма":  "acceptance_doc_amount",
    "сумма закрывающего документа": "acceptance_doc_amount",
    "предоплата":                   "is_prepayment",
    "ежемесячный платёж":           "is_monthly_payment",
    "ежемесячный платеж":           "is_monthly_payment",
    "процедура этп":                "etp_url",
    "№ п/п":                        "purchase_number",
    # --- New extended columns (plan §3) ---
    "регион":                                               "region",
    "регион проведения мероприятия":                        "region",
    "регион мероприятия":                                   "region",
    "регион поставки":                                      "delivery_region",
    "регион поставки (субъект рф)":                         "delivery_region",
    "субъект рф (место поставки)":                          "delivery_region",
    "субъект рф":                                           "delivery_region",
    "место оказания услуг / доставки":                      "delivery_location",
    "место оказания услуг":                                 "delivery_location",
    "место доставки":                                       "delivery_location",
    "адрес доставки":                                       "delivery_address",
    "экономия по результатам конкурентных закупок":         "economy",
    "экономия":                                             "economy",
    "срок действия договора":                               "contract_end_date",
    "квартал принятия обязательств":                        "commitment_quarter",
    "планируемый месяц платежа":                            "planned_payment_month",
    "дата окончания приёма заявок":                         "submission_deadline",
    "режим ндс":                                            "vat_mode",
    "подпись этапа":                                        "stage_label",
}

# «Платежи» sheet column map
_PAYMENTS_COLUMN_MAP: Dict[str, str] = {
    "№ договора":           "contract_number",
    "№ пп":                 "document_number",
    "дата платежа":         "payment_date",
    "сумма":                "amount",
    "назначение платежа":   "payment_purpose",
}

_STATUS_MAP = {
    # Русские лейблы из выпадающего списка (новый шаблон)
    "желания сотрудников": "wishes",
    "план закупок":         "plan_schedule",
    "ведётся работа":      "work_in_progress",
    "заключён договор":    "contracted",
    "поставлено":          "delivered",
    "оплачено":            "paid",
    # Технические ключи (backward compat)
    "wishes": "wishes", "желания": "wishes", "planned": "wishes", "планируется": "wishes", "план": "wishes",
    "plan_schedule": "plan_schedule",
    "confirmed": "work_in_progress", "подтверждено": "work_in_progress",
    "work_in_progress": "work_in_progress", "в работе": "work_in_progress", "in_progress": "work_in_progress",
    "contracted": "contracted", "законтрактовано": "contracted",
    "ordered": "ordered", "заказано": "ordered",
    "delivered": "delivered", "исполнено": "delivered",
    "поставлено, но не оплачено": "delivered", "поставлено, не оплачено": "delivered",
    "paid": "paid",
}

_SUBSTATUS_MAP = {
    # Русские лейблы из выпадающего списка
    "формирование тз":             "tz_forming",
    "сбор кп":                     "kp_collecting",
    "размещено на площадке":       "on_platform",
    "переговоры с поставщиком":    "contractor_negotiations",
    "подписание договора":         "contract_signing",
    # Технические ключи
    "tz_forming": "tz_forming",
    "kp_collecting": "kp_collecting",
    "on_platform": "on_platform",
    "contractor_negotiations": "contractor_negotiations",
    "contract_signing": "contract_signing",
}

_CONTRACT_TYPE_MAP = {
    # Русские лейблы из выпадающего списка
    "разовая поставка":              "single",
    "рамочный (нарастающий итог)":  "framework_cumulative",
    "рамочный (с указанием суммы)": "framework_with_amount",
    # Короткие псевдонимы (backward compat)
    "рамочный":           "framework_cumulative",
    "разовый":            "single",
    "рамочный накопительный": "framework_cumulative",
    # Технические ключи
    "single":               "single",
    "framework_cumulative": "framework_cumulative",
    "framework_with_amount":"framework_with_amount",
}

_ITEM_TYPE_MAP = {
    # Русские лейблы из выпадающего списка
    "товар":    "товар",
    "услуга":   "услуга",
    "работа":   "работа",
    # Множественное число (из файла пользователя «Товары», «Услуги»)
    "товары":   "товар",
    "услуги":   "услуга",
    "работы":   "работа",
}

_PAYMENT_BASIS_MAP = {
    # Русские лейблы из выпадающего списка
    "договор":      "contract",
    "счёт":         "invoice",
    "счёт-договор": "invoice_contract",
    # Технические ключи
    "contract":          "contract",
    "invoice":           "invoice",
    "invoice_contract":  "invoice_contract",
}

_METHOD_MAP = {
    # Русские лейблы из выпадающего списка
    "единственный поставщик":   "single",
    "единственный исполнитель": "single",
    "единый поставщик":         "single",
    "конкурентная процедура":   "competitive",
    "конкурсная процедура":     "competitive",
    "запрос котировок":         "quote_request",
    # Аббревиатуры (backward compat)
    "еи":  "single",
    "ед":  "single",
    "кп":  "competitive",
    "зк":  "quote_request",
    # Технические ключи
    "single":        "single",
    "competitive":   "competitive",
    "quote_request": "quote_request",
}

# Required fields for new-format validation (field_name, display_label)
_REQUIRED_FIELDS = [
    ("item_name",       "Наименование товара"),
    ("feo_l1",          "ФЭО Ур.1"),
    ("contractor_inn",  "ИНН контрагента"),
    ("contract_number", "№ договора"),
    ("contract_date",   "Дата договора"),
    ("plan_qty",        "Количество (план)"),
    ("plan_unit_price", "Цена за ед. (план)"),
    ("fact_total",      "Сумма факт"),
]

# Marker fields that distinguish new template from old legacy template
_NEW_TEMPLATE_MARKER_FIELDS = {"feo_l1", "plan_qty", "plan_unit_price", "fact_total"}
# Old-template marker (single path column)
_OLD_TEMPLATE_MARKER_FIELDS = {"feo_path", "feo_category_name"}


def _make_cell_helper(col_idx: Dict[str, int]):
    def cell(row, field):
        idx = col_idx.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None
    return cell


def _to_dec(v) -> Optional[Decimal]:
    if v is None:
        return None
    try:
        return Decimal(str(v).replace(" ", "").replace(",", ".").replace("\xa0", ""))
    except Exception:
        return None


def _to_date_val(v):
    if v is None:
        return None
    if hasattr(v, "date"):
        return v.date()
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(str(v).strip(), fmt).date()
        except Exception:
            pass
    return None


def _build_feo_index(feo_rows: List[FeoCategory], sid: int) -> Dict[int, FeoCategory]:
    """Return {id: FeoCategory} for categories belonging to given subsidy."""
    return {f.id: f for f in feo_rows if f.subsidy_id == sid}


import re as _re

def _norm_feo(s: str) -> str:
    """Нормализовать имя ФЭО для матчинга: срезать ведущий числовой префикс, lower+strip."""
    if not s:
        return ""
    # Срезаем «1. », «2.1 », «3) », «1.2.3. » и т.п.
    s2 = _re.sub(r'^\s*\d+([.\)]\d+)*[.\)]?\s*', '', s)
    s2 = s2.lower().strip()
    # Схлопываем повторяющиеся пробелы
    s2 = _re.sub(r'\s+', ' ', s2)
    return s2


async def _resolve_feo_levels(
    levels: List[str],
    sid: int,
    feo_index: Dict[int, FeoCategory],
    db=None,
    feo_rows_all=None,
    create_missing: bool = False,
    simulate: bool = False,
    pending_created: list | None = None,
    _sim_id_counter: list | None = None,
) -> tuple:
    """
    Walk the FEO tree of THIS subsidy using ordered level values.
    levels = non-empty strings in order (e.g. ['Снаряжение', 'Одежда', 'Кепи']).

    - Схлопывает соседние дубли (нормализованные): если levels[i] == levels[i-1], выбрасывает дубль.
    - Матчинг: сначала точное совпадение по нормализованным именам, затем вхождение.
    - При create_missing=True и наличии db: создаёт отсутствующие узлы автоматически.
    - При simulate=True: НЕ трогает БД, создаёт узлы-заглушки с отрицательными id,
      собирает в pending_created список {level, name, path} для отображения в превью.
    Returns (feo_category_id, error_message_or_None).
    """
    if not levels:
        return None, "Не указан ни один уровень ФЭО"

    # Схлопнуть соседние дубли по нормализованному значению
    deduped: List[str] = [levels[0]]
    for lv in levels[1:]:
        if _norm_feo(lv) != _norm_feo(deduped[-1]):
            deduped.append(lv)
    levels = deduped

    roots = [f for f in feo_index.values() if f.parent_id is None or f.parent_id not in feo_index]
    current_candidates = roots
    current_node = None
    # Track path for pending_created
    path_parts: List[str] = []

    for level_idx, part in enumerate(levels):
        nn_needle = _norm_feo(part)
        matched = None
        # exact match по нормализованным
        for candidate in current_candidates:
            if _norm_feo(candidate.name) == nn_needle:
                matched = candidate
                break
        # substring fallback по нормализованным
        if matched is None:
            for candidate in current_candidates:
                nn_cand = _norm_feo(candidate.name)
                if nn_needle and nn_needle in nn_cand:
                    matched = candidate
                    break
                if nn_cand and nn_cand in nn_needle:
                    matched = candidate
                    break
        # Автосоздание / симуляция, если не найдено
        if matched is None:
            if create_missing and db is not None:
                parent_id = current_node.id if current_node is not None else None
                # level = 1-based depth
                depth = level_idx + 1
                new_node = FeoCategory(
                    subsidy_id=sid,
                    parent_id=parent_id,
                    level=depth,
                    name=part,  # оригинальное имя из файла (с префиксом)
                    sort_order=None,
                    is_active=True,
                )
                db.add(new_node)
                await db.flush()
                # Добавляем в индекс и в общий список
                feo_index[new_node.id] = new_node
                if feo_rows_all is not None:
                    feo_rows_all.append(new_node)
                matched = new_node
            elif simulate:
                # Режим симуляции: создаём заглушку в памяти с отрицательным id
                if _sim_id_counter is None:
                    _sim_id_counter = [-1]
                sim_id = _sim_id_counter[0]
                _sim_id_counter[0] -= 1
                parent_id = current_node.id if current_node is not None else None
                depth = level_idx + 1
                stub_node = FeoCategory(
                    subsidy_id=sid,
                    parent_id=parent_id,
                    level=depth,
                    name=part,
                    sort_order=None,
                    is_active=True,
                )
                # Присваиваем отрицательный id без записи в БД
                stub_node.id = sim_id  # type: ignore[assignment]
                feo_index[sim_id] = stub_node
                if feo_rows_all is not None:
                    feo_rows_all.append(stub_node)
                # Собираем в коллектор (без дублей по полному пути)
                full_path = " / ".join(path_parts + [part])
                if pending_created is not None:
                    existing_paths = {e["path"] for e in pending_created}
                    if full_path not in existing_paths:
                        pending_created.append({
                            "level": depth,
                            "name": part,
                            "path": full_path,
                        })
                matched = stub_node
            else:
                return None, f"ФЭО не найдено на уровне {level_idx + 1}: '{part}'"
        path_parts.append(part)
        current_node = matched
        current_candidates = [f for f in feo_index.values() if f.parent_id == matched.id]

    if current_node is None:
        return None, "ФЭО не найдено"
    return current_node.id, None


async def _resolve_feo_path(
    path_str: str,
    feo_index: Dict[int, FeoCategory],
    sid: int = 0,
    db=None,
    feo_rows_all=None,
    create_missing: bool = False,
    simulate: bool = False,
    pending_created: list | None = None,
    _sim_id_counter: list | None = None,
) -> tuple:
    """
    Backward-compat: resolve old single path string (parts separated by ' / ' or '>').
    Returns (feo_category_id, error_message_or_None).
    """
    parts = [p.strip() for p in _re.split(r"\s*/\s*|\s*>\s*", path_str) if p.strip()]
    if not parts:
        return None, "Пустой путь ФЭО"
    return await _resolve_feo_levels(
        parts, sid, feo_index,
        db=db, feo_rows_all=feo_rows_all, create_missing=create_missing,
        simulate=simulate, pending_created=pending_created, _sim_id_counter=_sim_id_counter,
    )


def _find_payments_sheet(wb):
    """Find the «Платежи» worksheet (case-insensitive match on 'платеж')."""
    for sheet in wb.worksheets:
        if "платеж" in sheet.title.lower():
            return sheet
    return None


# ---------------------------------------------------------------------------
# Core parse-and-group logic (shared by /import and /import/preview)
# ---------------------------------------------------------------------------

async def _parse_and_group(
    content: bytes,
    sid: int,
    db: AsyncSession,
    commit: bool,
) -> Dict[str, Any]:
    """
    Parse Excel content (single «Закупки» sheet with inline payments; legacy 2-sheet
    workbooks with a «Платежи» sheet are still supported), validate rows, group into purchases,
    optionally commit. Handles both new (5-level FEO, 2-sheet) and old (path, 1-sheet) templates.

    Returns dict with keys:
      created_purchases, created_items, created_payments, skipped, errors   (commit=True)
      purchases (list of preview dicts), payments_errors, skipped, errors   (commit=False)
    """
    wb = load_workbook(BytesIO(content), data_only=True)

    # Find «Закупки» sheet — use first sheet (active) as primary
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        raise HTTPException(400, "Файл пустой или содержит только заголовки")

    # --- Build col_idx for «Закупки» ---
    raw_headers = [str(h).strip().lower() if h is not None else "" for h in rows[0]]
    col_idx: Dict[str, int] = {}
    for i, h in enumerate(raw_headers):
        field = _COLUMN_MAP.get(h)
        if field and field not in col_idx:
            col_idx[field] = i

    cell = _make_cell_helper(col_idx)

    # Detect template format: prefer new (feo_l1) over old (feo_path)
    has_new_feo = bool(_NEW_TEMPLATE_MARKER_FIELDS & set(col_idx.keys()))
    has_old_feo = bool(_OLD_TEMPLATE_MARKER_FIELDS & set(col_idx.keys()))
    # is_new_template: new format with level cols, OR both — prefer levels
    is_new_template = has_new_feo

    # --- Load lookup tables ---
    subs_rows = (await db.execute(select(Subsidy))).scalars().all()
    subs_by_name = {s.name.lower().strip(): s.id for s in subs_rows}

    contractors_all = (await db.execute(select(Contractor))).scalars().all()
    cont_by_inn  = {c.inn.strip(): c.id for c in contractors_all if c.inn}
    cont_by_name = {c.name.lower().strip(): c.id for c in contractors_all}

    feo_rows_all = (await db.execute(select(FeoCategory))).scalars().all()
    feo_index = _build_feo_index(feo_rows_all, sid)

    # Мероприятия (Приложение №3): единственная точка ввода — карточка
    # субсидии; импорт только резолвит event_name → event_id среди
    # мероприятий ЭТОЙ субсидии, ничего не создаёт (требование владельца
    # 2026-08-19, «иначе я потом никогда ничего не посчитаю»).
    event_rows_all = (await db.execute(select(Event))).scalars().all()
    events_by_subsidy: Dict[int, Dict[str, tuple]] = defaultdict(dict)
    for _ev in event_rows_all:
        events_by_subsidy[_ev.subsidy_id][normalize_event_name(_ev.name)] = (_ev.id, _ev.name)

    # Anti-dup: existing (contract_number, order_number) pairs for this subsidy
    existing_q = await db.execute(
        select(Purchase.contract_number, Purchase.order_number).where(
            Purchase.subsidy_id == sid,
            Purchase.contract_number.isnot(None),
            Purchase.contract_number != "",
        )
    )
    existing_keys = {(r[0], r[1]) for r in existing_q.fetchall()}

    # Duplicate-purchase index: существующие разовые (НЕ ежемесячные) закупки субсидии,
    # ключ (contractor_id, сумма). Сумма = ЛЮБАЯ из {НМЦК, цена договора, платёж} +
    # суммы отдельных платежей (Payment). Совпадение по любой сумме = возможный повтор.
    existing_dup_q = await db.execute(
        select(
            Purchase.id, Purchase.purchase_number, Purchase.item_name,
            Purchase.subject, Purchase.status, Purchase.contract_date,
            Purchase.contractor_id, Purchase.total_nmck,
            Purchase.contract_price, Purchase.payment_amount,
        ).where(
            Purchase.subsidy_id == sid,
            Purchase.is_monthly_payment.isnot(True),
            Purchase.contractor_id.isnot(None),
        )
    )
    existing_rows = existing_dup_q.fetchall()
    _exist_ids = [r.id for r in existing_rows]
    exist_pay_amounts: dict = defaultdict(list)
    if _exist_ids:
        _pay_q = await db.execute(
            select(Payment.purchase_id, Payment.amount).where(
                Payment.purchase_id.in_(_exist_ids),
                Payment.amount.isnot(None),
            )
        )
        for _pr in _pay_q.fetchall():
            exist_pay_amounts[_pr.purchase_id].append(_pr.amount)

    existing_dup_index: dict = defaultdict(list)
    for r in existing_rows:
        _base = {
            "source": "db",
            "id": r.id,
            "purchase_number": r.purchase_number,
            "name": r.item_name or r.subject or "",
            "status": r.status,
            "contract_date": r.contract_date.isoformat() if r.contract_date else None,
        }
        _pairs = [("НМЦК", r.total_nmck), ("цена договора", r.contract_price), ("платёж", r.payment_amount)]
        _pairs += [("платёж", a) for a in exist_pay_amounts.get(r.id, [])]
        for _reason, _val in _pairs:
            if _val is None:
                continue
            _fv = float(_val)
            if _fv <= 0:
                continue
            existing_dup_index[(r.contractor_id, round(_fv, 2))].append({**_base, "amount": _fv, "match_reason": _reason})

    errors: list[dict] = []
    warnings: list[dict] = []
    skipped = 0

    # --- Simulate mode: коллектор создаваемых ФЭО и счётчик отрицательных id ---
    simulate = not commit
    pending_created: list = []          # список {level, name, path} без дублей
    sim_id_counter: list = [-1]         # изменяемый счётчик [текущее_значение]
    subsidy_has_feo: bool = bool(feo_index)  # были ли категории ДО разбора

    # --- Parse «Закупки» rows ---
    parsed_rows = []
    for row_num, row in enumerate(rows[1:], start=2):
        # Blank row guard
        non_empty = any(
            v is not None and str(v).strip() != ""
            for v in row
        )
        if not non_empty:
            skipped += 1
            continue

        item_name = cell(row, "item_name")
        if not item_name:
            skipped += 1
            continue

        # ---- Resolve FEO ----
        feo_id = None
        feo_levels_display = None  # list of non-empty level strings, for preview display

        if has_new_feo and any(cell(row, f"feo_l{n}") for n in range(1, 6)):
            # New: collect non-empty level values in order
            raw_levels = [cell(row, f"feo_l{n}") or "" for n in range(1, 6)]
            levels = [v for v in raw_levels if v.strip()]
            feo_levels_display = levels

            if is_new_template:
                # feo_l1 is required — already validated below in required-field check,
                # but guard here too
                if not levels:
                    errors.append({"row": row_num, "name": item_name, "message": "ФЭО Ур.1 обязателен"})
                    continue
                feo_id, feo_err = await _resolve_feo_levels(
                    levels, sid, feo_index,
                    db=db, feo_rows_all=feo_rows_all, create_missing=commit,
                    simulate=simulate, pending_created=pending_created,
                    _sim_id_counter=sim_id_counter,
                )
                if feo_err:
                    # В режиме симуляции ФЭО-ошибок не должно быть (simulate обработал),
                    # но если всё же есть — добавляем ошибку, НЕ выбрасываем строку
                    if commit:
                        errors.append({"row": row_num, "name": item_name, "message": feo_err})
                        continue
                    else:
                        errors.append({"row": row_num, "name": item_name, "message": feo_err})
                        # Не делаем continue — строка уходит в превью как есть
            else:
                if levels:
                    feo_id, _ = await _resolve_feo_levels(
                        levels, sid, feo_index,
                        db=db, feo_rows_all=feo_rows_all, create_missing=commit,
                        simulate=simulate, pending_created=pending_created,
                        _sim_id_counter=sim_id_counter,
                    )
        elif has_old_feo:
            # Old: single path column (backward compat)
            feo_path_raw = cell(row, "feo_path") or cell(row, "feo_category_name")
            if feo_path_raw:
                feo_id, feo_err = await _resolve_feo_path(
                    feo_path_raw, feo_index,
                    sid=sid, db=db, feo_rows_all=feo_rows_all, create_missing=commit,
                    simulate=simulate, pending_created=pending_created,
                    _sim_id_counter=sim_id_counter,
                )
                if feo_err and not has_new_feo:
                    # only hard-error in old-style-only templates if it was new-enough format
                    pass  # leave feo_id as None
                feo_levels_display = [p.strip() for p in feo_path_raw.replace(">", "/").split("/") if p.strip()]

        # ---- Required-field validation (new template) ----
        if is_new_template:
            missing_labels = []
            for field, label in _REQUIRED_FIELDS:
                val = cell(row, field)
                if not val:
                    missing_labels.append(label)
            if missing_labels:
                errors.append({"row": row_num, "name": item_name, "missing": missing_labels})
                continue

        # ---- Contract number required (all modes) ----
        contract_num = cell(row, "contract_number")
        if is_new_template and not contract_num:
            errors.append({"row": row_num, "name": item_name, "message": "№ договора обязателен"})
            continue

        # ---- Subsidy ----
        row_sid = sid
        sub_name = cell(row, "subsidy_name")
        if not row_sid and sub_name:
            row_sid = subs_by_name.get(sub_name.lower().strip())
        if not row_sid:
            errors.append({"row": row_num, "name": item_name, "message": "Субсидия не указана"})
            continue

        # ---- Event (Приложение №3) ----
        # НЕ обязательно (требование владельца, 2026-08-19): «не всегда
        # известно». Пустая ячейка — молча, без предупреждения. Название, не
        # найденное среди мероприятий субсидии, тоже больше НЕ роняет строку —
        # закупка создаётся без event_id, а пользователь видит предупреждение
        # (preview/импорт) и оранжевый значок в реестре закупок. Ничего не
        # создаётся автоматически — единственная точка ввода мероприятий это
        # карточка субсидии.
        event_id_val = None
        event_name_raw = cell(row, "event_name")
        if event_name_raw:
            _ev_match = events_by_subsidy.get(row_sid, {}).get(normalize_event_name(event_name_raw))
            if _ev_match:
                event_id_val = _ev_match[0]
            else:
                warnings.append({
                    "row": row_num,
                    "name": item_name,
                    "message": (
                        f"Строка {row_num}: мероприятие «{event_name_raw}» не найдено среди "
                        "мероприятий субсидии — закупка создана без привязки. Заведите "
                        "мероприятие в карточке субсидии (Приложение №3) и привяжите вручную."
                    ),
                })

        # ---- Contractor ----
        c_inn  = cell(row, "contractor_inn")
        c_name = cell(row, "contractor_name")
        cont_id = None
        if c_inn:
            cont_id = cont_by_inn.get(c_inn.strip())
        if not cont_id and c_name:
            cont_id = cont_by_name.get(c_name.lower().strip())
        if not cont_id and c_inn:
            if commit:
                new_cont = Contractor(name=c_name or c_inn, inn=c_inn.strip())
                db.add(new_cont)
                await db.flush()
                cont_id = new_cont.id
                cont_by_inn[c_inn.strip()] = cont_id
                if c_name:
                    cont_by_name[c_name.lower().strip()] = cont_id

        # ---- Status ----
        status_raw = (cell(row, "status") or "").lower().strip()
        status = _STATUS_MAP.get(status_raw)
        if not status:
            # old-style payment columns on this sheet (backward compat)
            pay_amt = _to_dec(cell(row, "payment_amount"))
            if pay_amt and pay_amt > 0:
                status = "paid"
            elif contract_num:
                status = "contracted"
            else:
                status = "work_in_progress"

        # ---- Substatus ----
        substatus_raw = (cell(row, "substatus") or "").lower().strip()
        substatus_val = _SUBSTATUS_MAP.get(substatus_raw) if substatus_raw else None

        # ---- Contract type (purchase_contract_type) ----
        ct_raw = (cell(row, "contract_type_raw") or "").lower().strip()
        contract_type_val = _CONTRACT_TYPE_MAP.get(ct_raw) if ct_raw else None

        # ---- Item type ----
        it_raw = (cell(row, "item_type") or "").lower().strip()
        item_type_val = _ITEM_TYPE_MAP.get(it_raw) if it_raw else (it_raw or None)

        # ---- Payment basis type ----
        pb_raw = (cell(row, "payment_basis_type") or "").lower().strip()
        payment_basis_type_val = _PAYMENT_BASIS_MAP.get(pb_raw) if pb_raw else None

        # ---- Method ----
        method_raw = (cell(row, "purchase_method") or "").lower().strip()
        method = _METHOD_MAP.get(method_raw)

        # ---- New fields: purchase group number, order number, contract subject ----
        pg = cell(row, "purchase_group_num")
        order_no = cell(row, "order_number")
        subject_val = cell(row, "subject")

        # ---- Group key: prioritise purchase+order, then contract, then row ----
        if pg:
            group_key = f"pg:{pg}|{order_no or ''}"
        elif contract_num:
            group_key = f"contract:{contract_num}"
        else:
            group_key = f"row:{row_num}"

        # ---- Collect raw date cells for post-processing ----
        raw_contract_date           = row[col_idx["contract_date"]]           if "contract_date"           in col_idx else None
        raw_execution_term          = row[col_idx["execution_term"]]          if "execution_term"          in col_idx else None
        raw_execution_term_changed  = row[col_idx["execution_term_changed"]]  if "execution_term_changed"  in col_idx else None
        raw_delivery_date           = row[col_idx["delivery_date"]]           if "delivery_date"           in col_idx else None
        raw_payment_doc_date        = row[col_idx["payment_doc_date"]]        if "payment_doc_date"        in col_idx else None
        raw_acceptance_doc_date     = row[col_idx["acceptance_doc_date"]]     if "acceptance_doc_date"     in col_idx else None
        raw_contract_end_date       = row[col_idx["contract_end_date"]]       if "contract_end_date"       in col_idx else None
        raw_planned_payment_month   = row[col_idx["planned_payment_month"]]   if "planned_payment_month"   in col_idx else None
        raw_submission_deadline     = row[col_idx["submission_deadline"]]     if "submission_deadline"     in col_idx else None

        # ---- Boolean flags ----
        vat_applicable_raw  = (cell(row, "vat_applicable") or "").lower().strip()
        is_prepayment_raw   = (cell(row, "is_prepayment") or "").lower().strip()
        is_monthly_raw      = (cell(row, "is_monthly_payment") or "").lower().strip()

        # ---- Commitment quarter ----
        _cq_raw = cell(row, "commitment_quarter")
        commitment_quarter_val = None
        if _cq_raw:
            try:
                _cq_int = int(str(_cq_raw).strip())
                if 1 <= _cq_int <= 4:
                    commitment_quarter_val = _cq_int
            except Exception:
                pass

        # ---- VAT mode ----
        vat_mode_raw = (cell(row, "vat_mode") or "").lower().strip()
        _vat_mode_map = {"одинаковый": "uniform", "для каждого товара": "per_item",
                         "uniform": "uniform", "per_item": "per_item"}
        vat_mode_val = _vat_mode_map.get(vat_mode_raw, "uniform") if vat_mode_raw else None

        parsed_rows.append({
            "row_num":                  row_num,
            "group_key":                group_key,
            "purchase_group_num":       pg,
            "order_number":             order_no,
            "subject":                  subject_val,
            "item_name":                item_name,
            "item_type":                item_type_val,
            "feo_id":                   feo_id,
            "feo_levels":               feo_levels_display or [],
            "cont_id":                  cont_id,
            "cont_inn":                 c_inn,
            "cont_name":                c_name,
            "event_name":               event_name_raw,
            "event_id":                 event_id_val,
            "status":                   status,
            "substatus":                substatus_val,
            "contract_type_val":        contract_type_val,
            "payment_basis_type":       payment_basis_type_val,
            "method":                   method,
            "registry_number":          cell(row, "registry_number"),
            "contract_number":          contract_num,
            "contract_date":            _to_date_val(raw_contract_date),
            "contract_price":           _to_dec(cell(row, "contract_price")),
            "execution_term":           _to_date_val(raw_execution_term),
            "execution_term_changed":   _to_date_val(raw_execution_term_changed),
            "delivery_date":            _to_date_val(raw_delivery_date),
            "purchase_basis":           cell(row, "purchase_basis"),
            "responsible_person":       cell(row, "responsible_person"),
            "etp_url":                  cell(row, "etp_url"),
            "vat_applicable":           vat_applicable_raw in ("да", "yes", "1", "true"),
            "vat_exemption_article":    cell(row, "vat_exemption_article"),
            # Closing document
            "acceptance_doc_name":      cell(row, "acceptance_doc_name"),
            "acceptance_doc_number":    cell(row, "acceptance_doc_number"),
            "acceptance_doc_date":      _to_date_val(raw_acceptance_doc_date),
            "acceptance_doc_amount":    _to_dec(cell(row, "acceptance_doc_amount")),
            # Boolean flags
            "is_prepayment":            is_prepayment_raw in ("да", "yes", "1", "true"),
            "is_monthly_payment":       is_monthly_raw in ("да", "yes", "1", "true"),
            # Inline payment fields (new template) + legacy backward compat
            "payment_doc_number":       cell(row, "payment_doc_number"),
            "payment_doc_date":         _to_date_val(raw_payment_doc_date),
            "payment_amount":           _to_dec(cell(row, "payment_amount")),
            "payment_federal":          _to_dec(cell(row, "payment_federal")),
            "payment_purpose":          cell(row, "payment_purpose"),
            "plan_qty":                 _to_dec(cell(row, "plan_qty")),
            "unit":                     cell(row, "unit"),
            "plan_unit_price":          _to_dec(cell(row, "plan_unit_price")),
            "plan_total":               _to_dec(cell(row, "plan_total")),
            "fact_qty":                 _to_dec(cell(row, "fact_qty")),
            "fact_unit_price":          _to_dec(cell(row, "fact_unit_price")),
            "fact_total":               _to_dec(cell(row, "fact_total")),
            "country_origin":           cell(row, "country_origin"),
            "vat_rate":                 cell(row, "vat_rate"),
            "nmck":                     _to_dec(cell(row, "nmck")),
            "sid":                      row_sid,
            # New extended fields
            "region":                   cell(row, "region"),
            "delivery_region":          cell(row, "delivery_region"),
            "delivery_location":        cell(row, "delivery_location"),
            "delivery_address":         cell(row, "delivery_address"),
            "economy":                  _to_dec(cell(row, "economy")),
            "contract_end_date":        _to_date_val(raw_contract_end_date),
            "commitment_quarter":       commitment_quarter_val,
            "planned_payment_month":    _to_date_val(raw_planned_payment_month),
            "submission_deadline":      _to_date_val(raw_submission_deadline),
            "vat_mode":                 vat_mode_val,
            "stage_label":              cell(row, "stage_label"),
        })

    # --- Parse «Платежи» sheet ---
    pay_sheet = _find_payments_sheet(wb)
    parsed_payments: List[Dict[str, Any]] = []
    payments_errors: List[dict] = []

    if pay_sheet is not None:
        pay_rows = list(pay_sheet.iter_rows(values_only=True))
        if len(pay_rows) >= 2:
            raw_pay_headers = [str(h).strip().lower() if h is not None else "" for h in pay_rows[0]]
            pay_col_idx: Dict[str, int] = {}
            for i, h in enumerate(raw_pay_headers):
                field = _PAYMENTS_COLUMN_MAP.get(h)
                if field and field not in pay_col_idx:
                    pay_col_idx[field] = i

            pay_cell = _make_cell_helper(pay_col_idx)

            for pay_row_num, pay_row in enumerate(pay_rows[1:], start=2):
                # Skip blank rows and note rows (e.g. the hint text we put in row 3)
                non_empty = any(
                    v is not None and str(v).strip() != ""
                    for v in pay_row
                )
                if not non_empty:
                    continue

                p_contract = pay_cell(pay_row, "contract_number")
                if not p_contract:
                    continue
                # Skip the hint-text row if it landed in col 1
                if "одна строка" in p_contract.lower() or "платёж" in p_contract.lower():
                    continue

                raw_pay_date = pay_row[pay_col_idx["payment_date"]] if "payment_date" in pay_col_idx else None

                parsed_payments.append({
                    "pay_row_num":      pay_row_num,
                    "contract_number":  p_contract,
                    "document_number":  pay_cell(pay_row, "document_number"),
                    "payment_date":     _to_date_val(raw_pay_date),
                    "amount":           _to_dec(pay_cell(pay_row, "amount")),
                    "payment_purpose":  pay_cell(pay_row, "payment_purpose"),
                })

    # --- Group rows ---
    groups: Dict[str, List[dict]] = defaultdict(list)
    for pr in parsed_rows:
        groups[pr["group_key"]].append(pr)

    created_purchases = 0
    created_items = 0
    created_payments = 0
    preview_list = []
    batch_seen_dup: dict = defaultdict(list)
    duplicates_count = 0
    without_event = 0  # закупки (созданные/preview), у которых event_id не проставлен

    # Map contract_number → Purchase (built during commit, used for payment linking)
    contract_to_purchase: Dict[str, Purchase] = {}

    for group_key, group_rows in groups.items():
        first = group_rows[0]
        contract_num = first["contract_number"]

        # Anti-dup check (keyed on contract_number + order_number for рамочные)
        order_no = first.get("order_number")
        if contract_num and (contract_num, order_no) in existing_keys:
            skipped += 1
            if not commit:
                preview_list.append({
                    "group_key": group_key,
                    "contract_number": contract_num,
                    "purchase_group": first.get("purchase_group_num") or "",
                    "order_number": first.get("order_number") or "",
                    "contractor": first["cont_name"] or first["cont_inn"] or "",
                    "feo_path": " / ".join(first["feo_levels"]) if first["feo_levels"] else "",
                    "items_count": len(group_rows),
                    "plan_total": None,
                    "fact_total": None,
                    "status": first["status"],
                    "payments_count": 0,
                    "payments_total": None,
                    "skipped": True,
                    "skip_reason": "Дублирующийся договор",
                })
            continue

        # Aggregate sums
        plan_total_sum = sum(
            (pr["plan_total"] or (
                (pr["plan_qty"] or Decimal(0)) * (pr["plan_unit_price"] or Decimal(0))
            ))
            for pr in group_rows
        )
        fact_total_sum = sum(
            (pr["fact_total"] or Decimal(0))
            for pr in group_rows
        )

        # --- Collect inline payments from group rows ---
        inline_payments = []
        for pr in group_rows:
            if pr.get("payment_doc_number") or pr.get("payment_amount") or pr.get("payment_doc_date"):
                inline_payments.append({
                    "document_number": pr.get("payment_doc_number"),
                    "payment_date":    pr.get("payment_doc_date"),
                    "amount":          pr.get("payment_amount"),
                    "payment_purpose": pr.get("payment_purpose"),
                    "row_num":         pr["row_num"],
                })

        # Payments for preview: inline first, then sheet (backward compat)
        sheet_payments = [pp for pp in parsed_payments if pp["contract_number"] == contract_num] if contract_num else []

        # Deduplicate inline payments by fingerprint
        def _pay_fp_local(dn, pd, am):
            return (
                str(dn or "").strip(),
                str(pd or ""),
                round(float(am), 2) if am is not None else None,
            )
        seen_fps: set = set()
        unique_inline: list = []
        for ip in inline_payments:
            fp = _pay_fp_local(ip["document_number"], ip["payment_date"], ip["amount"])
            if fp not in seen_fps:
                seen_fps.add(fp)
                unique_inline.append(ip)

        # Total payments for preview = unique inline + sheet (deduped against inline)
        all_preview_payments = list(unique_inline)
        for sp in sheet_payments:
            fp = _pay_fp_local(sp.get("document_number"), sp.get("payment_date"), sp.get("amount"))
            if fp not in seen_fps:
                seen_fps.add(fp)
                all_preview_payments.append(sp)

        pay_count = len(all_preview_payments)
        pay_total = sum((ip.get("amount") or Decimal(0)) for ip in all_preview_payments)

        # Deduped item count (same key as commit-side PurchaseItem dedup)
        unique_item_keys = {
            ((pr["item_name"] or "").lower().strip(), pr["plan_qty"], pr["plan_unit_price"])
            for pr in group_rows
        }

        if not commit:
            # --- Duplicate-purchase detection (preview only) ---
            dup_matches = []
            if (
                not first.get("is_monthly_payment")
                and first.get("cont_id")
                and first["cont_id"] != -1
            ):
                _cand = [("НМЦК", first.get("nmck")), ("цена договора", plan_total_sum if plan_total_sum else None)]
                _cand += [("платёж", ip.get("amount")) for ip in all_preview_payments]
                _seen_db: set = set()
                _seen_file: set = set()
                for _reason, _val in _cand:
                    if _val is None:
                        continue
                    _fv = float(_val)
                    if _fv <= 0:
                        continue
                    dkey = (first["cont_id"], round(_fv, 2))
                    for m in existing_dup_index.get(dkey, []):
                        if m["id"] not in _seen_db:
                            _seen_db.add(m["id"])
                            dup_matches.append(m)
                    for prev in batch_seen_dup.get(dkey, []):
                        _fk = (prev["name"], prev["amount"])
                        if _fk not in _seen_file:
                            _seen_file.add(_fk)
                            dup_matches.append(prev)
                    batch_seen_dup[dkey].append({
                        "source": "file",
                        "id": None,
                        "purchase_number": None,
                        "name": (first.get("cont_name") or "") + " — " + (contract_num or group_key),
                        "amount": _fv,
                        "status": None,
                        "contract_date": None,
                        "match_reason": _reason,
                    })
            if dup_matches:
                duplicates_count += 1
            if not first.get("event_id"):
                without_event += 1
            preview_list.append({
                "group_key": group_key,
                "contract_number": contract_num or "",
                "purchase_group": first.get("purchase_group_num") or "",
                "order_number": first.get("order_number") or "",
                "contractor": first["cont_name"] or first["cont_inn"] or "",
                "feo_path": " / ".join(first["feo_levels"]) if first["feo_levels"] else "",
                "items_count": len(unique_item_keys),
                "plan_total": float(plan_total_sum) if plan_total_sum else None,
                "fact_total": float(fact_total_sum) if fact_total_sum else None,
                "status": first["status"],
                "payments_count": pay_count,
                "payments_total": float(pay_total) if pay_total else None,
                "skipped": False,
                "duplicate_matches": dup_matches,
                "event_id": first.get("event_id"),
                "event_name": first.get("event_name"),
            })
            continue

        # --- Create Purchase ---
        # For inline payments: sum all unique amounts, first doc_number, min date
        inline_amounts  = [ip["amount"] for ip in unique_inline if ip["amount"] is not None]
        inline_dates    = [ip["payment_date"] for ip in unique_inline if ip["payment_date"] is not None]
        inline_doc_nums = [ip["document_number"] for ip in unique_inline if ip["document_number"]]

        # Resolve purchase_basis (backward compat: accept both key and label)
        basis_raw = (first.get("purchase_basis") or "").lower().strip()
        basis_val = None
        if basis_raw in ("план закупок", "план_график", "plan_schedule"):
            basis_val = "plan_schedule"
        elif basis_raw in ("служебная записка", "служебная_записка", "service_note"):
            basis_val = "service_note"

        # Build acceptance_docs JSONB (first closing document, if any data given)
        acc_name   = first.get("acceptance_doc_name")
        acc_number = first.get("acceptance_doc_number")
        acc_date   = first.get("acceptance_doc_date")
        acc_amount = first.get("acceptance_doc_amount")
        acceptance_docs_val = []
        if any([acc_name, acc_number, acc_date, acc_amount]):
            acceptance_docs_val = [{
                "name":   acc_name or "",
                "number": acc_number or "",
                "date":   str(acc_date) if acc_date else "",
                "amount": float(acc_amount) if acc_amount else None,
            }]

        p = Purchase(
            subsidy_id=first["sid"],
            feo_category_id=first["feo_id"],
            event_id=first.get("event_id"),
            contractor_id=first["cont_id"] if (first["cont_id"] and first["cont_id"] != -1) else None,
            item_name=(first.get("subject") or first["item_name"]),
            purchase_number=int(first["purchase_group_num"]) if (first.get("purchase_group_num") or "").strip().isdigit() else None,
            order_number=first.get("order_number"),
            subject=first.get("subject"),
            status=first["status"],
            substatus=first.get("substatus"),
            purchase_method=first["method"],
            purchase_contract_type=first.get("contract_type_val"),
            payment_basis_type=first.get("payment_basis_type"),
            registry_number=first["registry_number"],
            contract_number=contract_num,
            contract_date=first["contract_date"],
            contract_price=first["contract_price"],
            execution_term=first["execution_term"],
            execution_term_changed=first.get("execution_term_changed"),
            delivery_date=first.get("delivery_date"),
            purchase_basis=basis_val,
            responsible_person=first.get("responsible_person"),
            etp_url=first.get("etp_url"),
            # НДС
            vat_applicable=first.get("vat_applicable") or False,
            # `first.get("vat_rate") and ...` терял явный 0 (falsy int/float)
            # — та же ошибка, что и «or 20» в documents.py: ставка 0%
            # молчаливо превращалась в «не указана». None-check сохраняет 0.
            vat_rate=(
                int(first["vat_rate"])
                if (first.get("vat_rate") is not None and str(first["vat_rate"]).strip() != ""
                    and str(first["vat_rate"]).isdigit())
                else None
            ),
            vat_exemption_article=first.get("vat_exemption_article"),
            # Closing document — JSONB-first + legacy write-through
            acceptance_docs=acceptance_docs_val if acceptance_docs_val else [],
            acceptance_doc_name=acc_name,
            acceptance_doc_number=acc_number,
            acceptance_doc_date=acc_date,
            acceptance_doc_amount=acc_amount,
            # Flags
            is_prepayment=first.get("is_prepayment") or False,
            is_monthly_payment=first.get("is_monthly_payment") or False,
            # Summary payment fields from inline rows (or legacy single-row if no inline)
            payment_doc_number=inline_doc_nums[0] if inline_doc_nums else first.get("payment_doc_number"),
            payment_doc_date=min(inline_dates) if inline_dates else first.get("payment_doc_date"),
            payment_amount=sum(inline_amounts) if inline_amounts else first.get("payment_amount"),
            payment_federal=first.get("payment_federal") if first.get("payment_federal") else None,
            nmck=first["nmck"],
            total_nmck=first["nmck"],
            planned_quantity=first["plan_qty"],
            planned_unit_price=first["plan_unit_price"],
            planned_total_price=plan_total_sum if plan_total_sum else first["nmck"],
            final_total_amount=fact_total_sum if fact_total_sum else None,
            country_origin=first["country_origin"],
            # Item type — from first row (group-level attribute)
            item_type=first.get("item_type"),
            # New extended fields
            region=first.get("region"),
            delivery_region=first.get("delivery_region"),
            delivery_location=first.get("delivery_location"),
            delivery_address=first.get("delivery_address"),
            economy=first.get("economy"),
            contract_end_date=first.get("contract_end_date"),
            commitment_quarter=first.get("commitment_quarter"),
            planned_payment_month=first.get("planned_payment_month"),
            submission_deadline=first.get("submission_deadline"),
            vat_mode=first.get("vat_mode") or "uniform",
            stage_label=first.get("stage_label"),
        )

        # --- Create PurchaseItems (dedup by key within group) ---
        items = []
        seen_item_keys: set = set()
        for pr in group_rows:
            # Normalize item key to avoid duplicates for monthly-payment rows
            norm_name = (pr["item_name"] or "").lower().strip()
            item_key = (norm_name, pr["plan_qty"], pr["plan_unit_price"])
            if item_key in seen_item_keys:
                continue
            seen_item_keys.add(item_key)

            item_plan_total = pr["plan_total"] or (
                (pr["plan_qty"] or Decimal(0)) * (pr["plan_unit_price"] or Decimal(0))
            ) or None
            pi = PurchaseItem(
                item_name=pr["item_name"],
                item_type=pr["item_type"],
                quantity=pr["plan_qty"],
                unit=pr["unit"],
                unit_price=pr["plan_unit_price"],
                total_price=item_plan_total,
                final_unit_price=pr["fact_unit_price"],
                final_total=pr["fact_total"],
                country_origin=pr["country_origin"],
                feo_category_id=pr["feo_id"],
                contractor_id=pr["cont_id"] if (pr["cont_id"] and pr["cont_id"] != -1) else None,
                contractor_inn=pr["cont_inn"],
                contractor_name=pr["cont_name"],
                # Аналогично Purchase.vat_rate выше: truthy-check терял 0.
                vat_rate=str(pr["vat_rate"]) if pr["vat_rate"] not in (None, "") else None,
            )
            items.append(pi)

        p.items = items
        db.add(p)
        await db.flush()  # get p.id

        if contract_num:
            existing_keys.add((contract_num, order_no))
            contract_to_purchase[contract_num] = p

        created_purchases += 1
        created_items += len(items)
        if not first.get("event_id"):
            without_event += 1

        # --- Create Payment records from inline rows ---
        if unique_inline:
            # Anti-dup against already-existing payments (freshly flushed purchase has no payments yet)
            existing_pays_q = (await db.execute(
                select(Payment.document_number, Payment.payment_date, Payment.amount)
                .where(Payment.purchase_id == p.id)
            )).all()

            def _pay_fp(dn, pd, am):
                return (
                    str(dn or "").strip(),
                    str(pd or ""),
                    round(float(am), 2) if am is not None else None,
                )

            existing_fps_inline = {_pay_fp(dn, pd, am) for dn, pd, am in existing_pays_q}

            for ip in unique_inline:
                fp = _pay_fp(ip["document_number"], ip["payment_date"], ip["amount"])
                if fp in existing_fps_inline:
                    continue
                existing_fps_inline.add(fp)
                pay_obj = Payment(
                    purchase_id=p.id,
                    document_number=ip["document_number"],
                    payment_date=ip["payment_date"],
                    amount=ip["amount"],
                    payment_purpose=ip["payment_purpose"],
                )
                db.add(pay_obj)
                created_payments += 1

    # --- Process «Платежи» sheet (commit mode) ---
    if commit and parsed_payments:
        pay_by_contract: Dict[str, List[dict]] = defaultdict(list)
        for pp in parsed_payments:
            pay_by_contract[pp["contract_number"]].append(pp)

        for contract_num, pay_list in pay_by_contract.items():
            # Find the Purchase — may be newly created or already existing
            purchase = contract_to_purchase.get(contract_num)
            if purchase is None:
                # Check DB for existing purchases with this contract_number
                existing_p = (await db.execute(
                    select(Purchase).where(
                        Purchase.subsidy_id == sid,
                        Purchase.contract_number == contract_num,
                    )
                )).scalars().first()
                purchase = existing_p

            if purchase is None:
                for pp in pay_list:
                    payments_errors.append({
                        "row": pp["pay_row_num"],
                        "name": contract_num,
                        "message": f"Платёж: закупка с № договора '{contract_num}' не найдена",
                    })
                continue

            # Anti-dup: fingerprints of payments already stored for this purchase
            existing_pays = (await db.execute(
                select(Payment.document_number, Payment.payment_date, Payment.amount)
                .where(Payment.purchase_id == purchase.id)
            )).all()
            def _pay_fp(dn, pd, am):
                return (
                    str(dn or "").strip(),
                    str(pd or ""),
                    round(float(am), 2) if am is not None else None,
                )
            existing_fps = {_pay_fp(dn, pd, am) for dn, pd, am in existing_pays}

            payment_objects = []
            for pp in pay_list:
                fp = _pay_fp(pp["document_number"], pp["payment_date"], pp["amount"])
                if fp in existing_fps:
                    continue
                existing_fps.add(fp)
                pay_obj = Payment(
                    purchase_id=purchase.id,
                    document_number=pp["document_number"],
                    payment_date=pp["payment_date"],
                    amount=pp["amount"],
                    payment_purpose=pp["payment_purpose"],
                )
                db.add(pay_obj)
                payment_objects.append(pay_obj)
                created_payments += 1

            # Владелец (2026-08-19): платежи из Excel-импорта — это тоже «по нашим
            # данным» (payment_source='manual' по умолчанию на модели,
            # confirmed_by_statement=False), НЕ подтверждённая казначейством
            # оплата. Раньше здесь payment_amount проставлялся напрямую суммой
            # ИМПОРТИРОВАННЫХ платежей — что и выдавало заявленное за
            # подтверждённое. Теперь агрегаты (payment_amount = подтверждено,
            # payment_amount_declared = заявлено) считает recompute_purchase_payments
            # по тем же правилам, что и форма в карточке закупки.
            if payment_objects:
                await recompute_purchase_payments(db, purchase.id)

    elif not commit and parsed_payments:
        # Preview mode: validate payment links only
        for pp in parsed_payments:
            contract_num = pp["contract_number"]
            # Check if any parsed group has this contract_number
            matched = any(
                pr["contract_number"] == contract_num
                for pr in parsed_rows
            )
            if not matched and not any(k[0] == contract_num for k in existing_keys):
                payments_errors.append({
                    "row": pp["pay_row_num"],
                    "name": contract_num,
                    "message": f"Платёж: закупка с № договора '{contract_num}' не найдена",
                })

    if commit:
        await db.commit()
        return {
            "created_purchases": created_purchases,
            "created_items": created_items,
            "created_payments": created_payments,
            "skipped": skipped,
            "errors": errors + payments_errors,
            "warnings": warnings,
            "without_event": without_event,
        }
    else:
        return {
            "purchases": preview_list,
            "payments_errors": payments_errors,
            "skipped": skipped,
            "errors": errors,
            "warnings": warnings,
            "without_event": without_event,
            "duplicates_count": duplicates_count,
            "feo_to_create": sorted(pending_created, key=lambda x: x["path"]),
            "subsidy_has_feo": subsidy_has_feo,
        }


async def _check_purchases_import_permission(
    subsidy_id: int, db: AsyncSession, current_user
) -> Subsidy:
    """Заливать закупки из Excel может только тот, у кого есть право
    редактировать ИМЕННО эту субсидию (subsidy.edit) — просто «залогинен»
    недостаточно (требование владельца, 2026-08-19). Тот же паттерн, что
    events._get_subsidy_for_events(edit=True) и
    subsidy_approvers._get_subsidy_or_404(edit=True)."""
    result = await db.execute(select(Subsidy).where(Subsidy.id == subsidy_id))
    subsidy = result.scalar_one_or_none()
    if not subsidy:
        raise HTTPException(404, "Субсидия не найдена")
    if not await has_org_key(current_user, db, subsidy.org_id, 'subsidy.edit', subsidy_id=subsidy_id):
        raise HTTPException(
            403,
            "Импортировать закупки в эту субсидию может только тот, у кого есть право "
            "её редактирования",
        )
    return subsidy


# ---------------------------------------------------------------------------
# POST /import
# ---------------------------------------------------------------------------

@router.post("/import")
async def import_purchases_from_excel(
    file: UploadFile = File(...),
    subsidy_id: int = Query(..., description="ID субсидии (обязательно)"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Импорт закупок из Excel. Возвращает {created_purchases, created_items, created_payments, skipped, errors}."""
    await _check_purchases_import_permission(subsidy_id, db, current_user)
    if load_workbook is None:
        raise HTTPException(500, "openpyxl не установлен")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Поддерживаются только файлы .xlsx и .xls")

    content = await file.read()
    return await _parse_and_group(content, subsidy_id, db, commit=True)


# ---------------------------------------------------------------------------
# POST /import/preview
# ---------------------------------------------------------------------------

@router.post("/import/preview")
async def preview_purchases_import(
    file: UploadFile = File(...),
    subsidy_id: int = Query(..., description="ID субсидии (обязательно)"),
    db: AsyncSession = Depends(get_db),
    current_user=Depends(get_current_user),
):
    """Превью импорта без сохранения. Возвращает {purchases, payments_errors, skipped, errors}."""
    await _check_purchases_import_permission(subsidy_id, db, current_user)
    if load_workbook is None:
        raise HTTPException(500, "openpyxl не установлен")
    if not (file.filename or "").lower().endswith((".xlsx", ".xls")):
        raise HTTPException(400, "Поддерживаются только файлы .xlsx и .xls")

    content = await file.read()
    return await _parse_and_group(content, subsidy_id, db, commit=False)
