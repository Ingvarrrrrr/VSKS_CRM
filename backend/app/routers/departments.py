from datetime import date
from io import BytesIO
from typing import List, Optional
from urllib.parse import quote as _url_quote

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import select, func, update as sa_update
from sqlalchemy.ext.asyncio import AsyncSession

from app.auth.jwt import (
    get_current_user, require_role, get_org_filter, get_single_org_id,
    ADMIN_ROLES, MANAGER_ROLES,
)
from app.auth.permissions import require_tab
from app.database import get_db
from app.models.department import Department, TaskEditDelegate
from app.models.organization import Organization
from app.models.user import User
from app.models.user_organization import UserOrganization
from app.services.org_assignment_dates import (
    first_row_assignment_dates,
    dept_transfer_date,
    position_change_date,
)

router = APIRouter(prefix="/api/departments", tags=["departments"])


# ── Schemas ──────────────────────────────────────────────────────────────────

class DepartmentCreate(BaseModel):
    name: str
    subsidy_id: Optional[int] = None
    head_user_id: Optional[int] = None
    deputy_head_user_id: Optional[int] = None
    curator_user_id: Optional[int] = None
    parent_id: Optional[int] = None
    org_id: Optional[int] = None  # Override org assignment (superadmin only)

class DepartmentUpdate(BaseModel):
    name: Optional[str] = None
    head_user_id: Optional[int] = None
    deputy_head_user_id: Optional[int] = None
    curator_user_id: Optional[int] = None
    parent_id: Optional[int] = None
    subsidy_id: Optional[int] = None

class DepartmentOut(BaseModel):
    id: int
    name: str
    org_id: int
    subsidy_id: Optional[int] = None
    head_user_id: Optional[int] = None
    head_user_name: Optional[str] = None
    deputy_head_user_id: Optional[int] = None
    curator_user_id: Optional[int] = None
    parent_id: Optional[int] = None
    member_count: int = 0
    class Config:
        from_attributes = True

class MemberAdd(BaseModel):
    user_id: int
    position: Optional[str] = None
    dept_assigned_at: Optional[date] = None  # дата назначения в отдел; по умолчанию — сегодня

class MemberOut(BaseModel):
    id: int
    department_id: int
    user_id: int
    user_name: Optional[str] = None
    user_role: Optional[str] = None
    position: Optional[str] = None
    dept_assigned_at: Optional[date] = None
    position_assigned_at: Optional[date] = None
    class Config:
        from_attributes = True

class DelegateAdd(BaseModel):
    target_user_id: int
    delegate_user_id: int

class DelegateOut(BaseModel):
    id: int
    target_user_id: int
    target_user_name: Optional[str] = None
    delegate_user_id: int
    delegate_user_name: Optional[str] = None
    class Config:
        from_attributes = True


# ── Helpers ──────────────────────────────────────────────────────────────────

async def _enrich_dept(d: Department, db: AsyncSession) -> DepartmentOut:
    head_name = None
    if d.head_user_id:
        u = await db.get(User, d.head_user_id)
        head_name = (u.full_name or u.username) if u else None
    count_res = await db.execute(
        select(func.count(UserOrganization.id)).where(UserOrganization.dept_id == d.id)
    )
    member_count = count_res.scalar() or 0
    return DepartmentOut(
        id=d.id, name=d.name, org_id=d.org_id, subsidy_id=d.subsidy_id,
        head_user_id=d.head_user_id, head_user_name=head_name,
        deputy_head_user_id=d.deputy_head_user_id, curator_user_id=d.curator_user_id,
        parent_id=d.parent_id, member_count=member_count,
    )


# ── CRUD Departments ─────────────────────────────────────────────────────────

@router.get("/", response_model=List[DepartmentOut])
async def list_departments(
    subsidy_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = select(Department).order_by(Department.name)
    org_ids = get_org_filter(current_user)
    if org_ids is not None:
        q = q.where(Department.org_id.in_(org_ids))
    if subsidy_id is not None:
        q = q.where(Department.subsidy_id == subsidy_id)
    depts = (await db.execute(q)).scalars().all()
    return [await _enrich_dept(d, db) for d in depts]


@router.get("/tree")
async def department_tree(
    subsidy_id: Optional[int] = Query(None),
    org_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Дерево отделов с вложенными сотрудниками."""
    q = select(Department).order_by(Department.name)
    org_ids = get_org_filter(current_user)
    if org_ids is not None:
        q = q.where(Department.org_id.in_(org_ids))
    if subsidy_id is not None:
        q = q.where(Department.subsidy_id == subsidy_id)
    if org_id is not None:
        q = q.where(Department.org_id == org_id)
    depts = (await db.execute(q)).scalars().all()

    dept_ids = [d.id for d in depts]
    # Load all members from user_organizations.dept_id (single source of truth)
    uo_dept_rows = []
    if dept_ids:
        uo_dept_rows = (await db.execute(
            select(UserOrganization).where(
                UserOrganization.dept_id.in_(dept_ids),
            )
        )).scalars().all()

    # Load user names + positions
    user_ids = {uo.user_id for uo in uo_dept_rows}
    for d in depts:
        if d.head_user_id:
            user_ids.add(d.head_user_id)
    users_map = {}
    if user_ids:
        for u in (await db.execute(select(User).where(User.id.in_(user_ids)))).scalars().all():  # superadmin-bypass-ok: lookup by pre-computed IDs for enrichment
            users_map[u.id] = {"id": u.id, "name": u.full_name or u.username, "role": u.role, "position": u.position}

    # Load org names
    all_org_ids = {d.org_id for d in depts if d.org_id}
    orgs_map: dict = {}
    if all_org_ids:
        for o in (await db.execute(select(Organization).where(Organization.id.in_(all_org_ids)))).scalars().all():
            orgs_map[o.id] = o.name

    # Build tree
    by_id = {}
    for d in depts:
        head_u = users_map.get(d.head_user_id, {})
        by_id[d.id] = {
            "id": d.id, "name": d.name, "org_id": d.org_id,
            "org_name": orgs_map.get(d.org_id),
            "subsidy_id": d.subsidy_id, "parent_id": d.parent_id,
            "head_user_id": d.head_user_id,
            "head_user_name": head_u.get("name"),
            "members": [],
            "children": [],
        }

    # Build members list from user_organizations (single source of truth)
    added_pairs: set = set()

    for uo in uo_dept_rows:
        if uo.dept_id not in by_id:
            continue
        if (uo.dept_id, uo.user_id) in added_pairs:
            continue
        u = users_map.get(uo.user_id, {})
        entry = {
            "member_id": uo.id, "user_id": uo.user_id,
            "name": u.get("name", "?"), "role": u.get("role"),
            "position": uo.position or u.get("position"),
        }
        by_id[uo.dept_id]["members"].append(entry)
        added_pairs.add((uo.dept_id, uo.user_id))

    roots = []
    for d in depts:
        node = by_id[d.id]
        if d.parent_id and d.parent_id in by_id:
            by_id[d.parent_id]["children"].append(node)
        else:
            roots.append(node)
    return roots


@router.post("/", response_model=DepartmentOut)
async def create_department(
    data: DepartmentCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    # Superadmin can assign to any org via body.org_id; others use their org
    if data.org_id and current_user.role in ('superadmin', 'account_owner'):
        org_id = data.org_id
    else:
        org_id = get_single_org_id(current_user) or current_user.org_id
    # 27.4-02: НЕ применяем .title() — он ломает аббревиатуры («отдел МТО» → «Отдел Мто»).
    # Сохраняем регистр как ввёл пользователь.
    norm_name = data.name.strip() if data.name else data.name
    dept = Department(
        name=norm_name, org_id=org_id,
        subsidy_id=data.subsidy_id, head_user_id=data.head_user_id,
        deputy_head_user_id=data.deputy_head_user_id, curator_user_id=data.curator_user_id,
        parent_id=data.parent_id,
    )
    db.add(dept)
    await db.commit()
    await db.refresh(dept)
    return await _enrich_dept(dept, db)


@router.patch("/{dept_id}", response_model=DepartmentOut)
async def update_department(
    dept_id: int,
    data: DepartmentUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    dept = await db.get(Department, dept_id)
    if not dept:
        raise HTTPException(404, "Отдел не найден")
    old_name = dept.name
    old_head = dept.head_user_id
    old_deputy = dept.deputy_head_user_id
    update = data.dict(exclude_unset=True)
    if "name" in update and update["name"]:
        update["name"] = update["name"].strip()
    for k, v in update.items():
        setattr(dept, k, v)
    await db.commit()
    await db.refresh(dept)
    # If name changed, sync users.department for all members
    if "name" in update and dept.name != old_name:
        member_ids_q = select(UserOrganization.user_id).where(UserOrganization.dept_id == dept_id)
        await db.execute(
            sa_update(User).where(User.id.in_(member_ids_q)).values(department=dept.name)
        )
        await db.commit()
    # If head changed, sync hierarchy
    if dept.head_user_id and dept.head_user_id != old_head:
        from app.routers.users import _sync_head_hierarchy
        await _sync_head_hierarchy(dept, db)
        await db.commit()
    # Двусторонняя синхронизация: head/deputy отдела → должность в членстве
    if dept.head_user_id != old_head or dept.deputy_head_user_id != old_deputy:
        from app.services.dept_role_sync import sync_position_from_head
        await sync_position_from_head(db, dept, old_head, old_deputy)
        await db.commit()
    return await _enrich_dept(dept, db)


@router.delete("/{dept_id}")
async def delete_department(
    dept_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    dept = await db.get(Department, dept_id)
    if not dept:
        raise HTTPException(404, "Отдел не найден")
    # Check for members
    member_count = (await db.execute(
        select(func.count()).select_from(UserOrganization).where(UserOrganization.dept_id == dept_id)
    )).scalar() or 0
    if member_count > 0:
        raise HTTPException(400, f"Нельзя удалить отдел с сотрудниками ({member_count} чел.). Сначала уберите всех сотрудников из отдела.")
    # Check for child departments
    child_count = (await db.execute(
        select(func.count()).select_from(Department).where(Department.parent_id == dept_id)
    )).scalar() or 0
    if child_count > 0:
        raise HTTPException(400, f"Нельзя удалить отдел с дочерними отделами ({child_count} шт.). Сначала удалите или переместите дочерние отделы.")
    await db.delete(dept)
    await db.commit()
    return {"ok": True}


# ── Members ──────────────────────────────────────────────────────────────────

@router.get("/{dept_id}/members", response_model=List[MemberOut])
async def list_members(
    dept_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    # user_organizations is now the single source of truth for dept membership.
    uo_rows = (await db.execute(
        select(UserOrganization).where(UserOrganization.dept_id == dept_id)
    )).scalars().all()

    user_ids = {r.user_id for r in uo_rows}
    users_map: dict[int, User] = {}
    if user_ids:
        for u in (await db.execute(select(User).where(User.id.in_(user_ids)))).scalars().all():  # superadmin-bypass-ok: lookup by pre-computed IDs for department member enrichment
            users_map[u.id] = u

    out: list[MemberOut] = []
    seen_users: set[int] = set()
    for r in uo_rows:
        if r.user_id in seen_users:
            continue
        seen_users.add(r.user_id)
        u = users_map.get(r.user_id)
        out.append(MemberOut(
            id=r.id, department_id=dept_id, user_id=r.user_id,
            user_name=(u.full_name or u.username) if u else None,
            user_role=u.role if u else None,
            position=r.position or (u.position if u else None),
            dept_assigned_at=r.dept_assigned_at,
            position_assigned_at=r.position_assigned_at,
        ))
    return out


@router.post("/{dept_id}/members", response_model=MemberOut)
async def add_member(
    dept_id: int,
    data: MemberAdd,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    dept = await db.get(Department, dept_id)
    if not dept:
        raise HTTPException(404, "Отдел не найден")
    # Multi-dept-per-org разрешён: сотрудник может состоять одновременно в нескольких
    # отделах одной и той же организации (Цыганов в Бухгалтерии+Отделе МТО+«1»).
    # user_organizations is the single source of truth — look for an existing row
    # for this exact (user, org, dept) triple.
    user = await db.get(User, data.user_id)
    uo_exact = (await db.execute(
        select(UserOrganization).where(
            UserOrganization.user_id == data.user_id,
            UserOrganization.org_id == dept.org_id,
            UserOrganization.dept_id == dept_id,
        )
    )).scalar_one_or_none()
    # Строка-заглушка «Без отдела» для этой же пары (user, org) — источник hired_at/
    # позиции/ставки для новой dept-строки, и то, что нужно удалить после успешного
    # назначения (иначе она продолжает висеть в карточке как «Без отдела», хотя
    # человек уже определён в отдел — баг owner 2026-09-01).
    null_row = (await db.execute(
        select(UserOrganization).where(
            UserOrganization.user_id == data.user_id,
            UserOrganization.org_id == dept.org_id,
            UserOrganization.dept_id.is_(None),
        )
    )).scalar_one_or_none()
    if uo_exact:
        # Row already correct — just sync position if needed
        old_position = uo_exact.position
        position_changed = bool(data.position) and data.position != old_position
        if data.position:
            uo_exact.position = data.position
        uo_exact.position_assigned_at = position_change_date(
            position_changed=position_changed,
            current=uo_exact.position_assigned_at,
            explicit=None,
        )
        if uo_exact.dept_assigned_at is None:
            uo_exact.dept_assigned_at = dept_transfer_date(data.dept_assigned_at)
        await db.commit()
        m = uo_exact
    else:
        # No row for this dept yet — insert a new one (multi-dept allowed).
        # hired_at (дата трудоустройства) ОБЩАЯ на пару (user, org) — переносим её
        # из заглушки/любой другой строки этой пары, а НЕ ставим now() (баг owner:
        # дата трудоустройства подменялась датой назначения в отдел).
        new_pos = data.position or (user.position if user else None)
        base_row = null_row or (await db.execute(
            select(UserOrganization)
            .where(
                UserOrganization.user_id == data.user_id,
                UserOrganization.org_id == dept.org_id,
            )
            .order_by(UserOrganization.id.asc())
            .limit(1)
        )).scalars().first()
        kwargs = dict(user_id=data.user_id, org_id=dept.org_id, dept_id=dept_id, position=new_pos)
        if base_row is not None and base_row.hired_at is not None:
            kwargs["hired_at"] = base_row.hired_at
        if base_row is None:
            # Первая-ЕВЕР строка user_organizations для этой пары (user, org) —
            # это приём на работу, обе даты назначения = дате приёма (владелец,
            # 2026-09-01), а не «сегодня».
            kwargs.update(first_row_assignment_dates(
                kwargs.get("hired_at"),
                has_position=bool(new_pos),
                has_dept=True,
                explicit_dept_assigned_at=data.dept_assigned_at,
            ))
        else:
            # Перевод в другой отдел уже трудоустроенного человека — дата
            # перевода явная или сегодня; должность, если не менялась,
            # наследует прежнюю дату назначения (не обнуляем её).
            prior_position = base_row.position
            position_changed = bool(new_pos) and new_pos != prior_position
            kwargs["dept_assigned_at"] = dept_transfer_date(data.dept_assigned_at)
            pos_date = position_change_date(
                position_changed=position_changed,
                current=base_row.position_assigned_at,
                explicit=None,
            )
            if pos_date is not None:
                kwargs["position_assigned_at"] = pos_date
        if base_row is not None:
            if base_row.salary_amount is not None:
                kwargs["salary_amount"] = base_row.salary_amount
            if base_row.employment_percent is not None:
                kwargs["employment_percent"] = base_row.employment_percent
        m = UserOrganization(**kwargs)
        db.add(m)
        await db.commit()
        await db.refresh(m)
    # Заглушка «Без отдела» больше не нужна — человек определён в отдел.
    if null_row is not None and null_row.id != m.id:
        await db.delete(null_row)
        await db.commit()
    # Должность члена → head/deputy отдела (двусторонняя синхронизация)
    from app.services.dept_role_sync import sync_head_from_position
    await sync_head_from_position(db, dept, data.user_id, m.position)
    await db.commit()
    # Also update user.department string
    if user:
        user.department = dept.name
        await db.commit()
    # Auto-create hierarchy: new member becomes subordinate of dept head
    if dept.head_user_id and dept.head_user_id != data.user_id:
        from app.models.user_hierarchy import UserHierarchy
        existing_uh = (await db.execute(
            select(UserHierarchy).where(
                UserHierarchy.manager_id == dept.head_user_id,
                UserHierarchy.subordinate_id == data.user_id,
            )
        )).scalar_one_or_none()
        if not existing_uh:
            db.add(UserHierarchy(manager_id=dept.head_user_id, subordinate_id=data.user_id))
            await db.commit()
    u = await db.get(User, m.user_id)
    return MemberOut(
        id=m.id, department_id=dept_id, user_id=m.user_id,
        user_name=(u.full_name or u.username) if u else None,
        user_role=u.role if u else None,
        position=m.position or (u.position if u else None),
        dept_assigned_at=m.dept_assigned_at,
        position_assigned_at=m.position_assigned_at,
    )


@router.patch("/{dept_id}/members/{user_id}")
async def update_member(
    dept_id: int,
    user_id: int,
    data: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    m = (await db.execute(
        select(UserOrganization).where(
            UserOrganization.dept_id == dept_id,
            UserOrganization.user_id == user_id,
        )
    )).scalar_one_or_none()
    if not m:
        raise HTTPException(404, "Сотрудник не найден в отделе")
    if "position" in data:
        m.position = data["position"]
        await db.commit()
        dept = await db.get(Department, dept_id)
        if dept is not None:
            from app.services.dept_role_sync import sync_head_from_position
            await sync_head_from_position(db, dept, user_id, m.position)
            await db.commit()
    else:
        await db.commit()
    return {"ok": True}


@router.delete("/{dept_id}/members/{user_id}")
async def remove_member(
    dept_id: int,
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    uo_rows = (await db.execute(
        select(UserOrganization).where(
            UserOrganization.user_id == user_id,
            UserOrganization.dept_id == dept_id,
        )
    )).scalars().all()

    if not uo_rows:
        raise HTTPException(404, "Сотрудник не найден в отделе")

    # Phase 30 fix: учитываем закупки ТОЛЬКО в той же организации, что и отдел.
    # Раньше считались закупки глобально → удаление из АНО блокировалось задачами из ВСКС.
    from app.models.purchase import Purchase
    from app.models.subsidy import Subsidy
    dept = await db.get(Department, dept_id)
    dept_org_id = dept.org_id if dept else None
    active_count_q = (
        select(func.count()).select_from(Purchase)
        .join(Subsidy, Subsidy.id == Purchase.subsidy_id, isouter=True)
        .where(
            Purchase.assigned_user_id == user_id,
            Purchase.status.notin_(['paid', 'cancelled']),
        )
    )
    if dept_org_id is not None:
        active_count_q = active_count_q.where(Subsidy.org_id == dept_org_id)
    active_count = (await db.execute(active_count_q)).scalar() or 0
    # Суперадмин переносит/выводит сотрудника несмотря на активные закупки
    # (ссылки не блокируют операции суперадмина — модель Wave 1/3).
    if active_count > 0 and current_user.role != 'superadmin':
        org_name = (await db.execute(
            select(Organization.name).where(Organization.id == dept_org_id)
        )).scalar() if dept_org_id else 'этой организации'
        raise HTTPException(400, f"У сотрудника {active_count} активных задач (закупок) в «{org_name}». Сначала перераспределите задачи.")

    # Человек выведен из отдела → снять с него роль начальника/зама этого отдела
    # и очистить должность в снимаемых членствах (человек в разных отделах может
    # занимать разные должности — трогаем только это членство).
    if dept is not None:
        from app.services.dept_role_sync import clear_role_on_removal
        await clear_role_on_removal(db, dept, user_id)
    for uo in uo_rows:
        if uo.position in ("Начальник отдела", "Заместитель начальника отдела"):
            uo.position = None

    for uo in uo_rows:
        # Multi-dept: если у пары (user, org) остаётся ЕЩЁ хотя бы один отдел (не
        # считая текущей строки) — человек не выпадает из организации, эта строка
        # просто больше не нужна. Заглушку dept_id=NULL заводить НЕ надо — иначе
        # ровно тот же баг, который эта миграция/фича лечит (лишняя «Без отдела»).
        other_dept = (await db.execute(
            select(UserOrganization.id).where(
                UserOrganization.user_id == uo.user_id,
                UserOrganization.org_id == uo.org_id,
                UserOrganization.dept_id.isnot(None),
                UserOrganization.id != uo.id,
            )
        )).first()
        if other_dept is not None:
            await db.delete(uo)
            continue
        # Это был последний отдел пары (user, org) — превращаем строку в «Без
        # отдела», сохранив hired_at/должность/ставку (иначе человек выпадает
        # из организации целиком).
        conflict = (await db.execute(
            select(UserOrganization).where(
                UserOrganization.user_id == uo.user_id,
                UserOrganization.org_id == uo.org_id,
                UserOrganization.dept_id.is_(None),
                UserOrganization.id != uo.id,
            )
        )).scalar_one_or_none()
        if conflict is not None:
            # Заглушка уже существует (не должно происходить в норме, но на
            # всякий случай) — удаляем дубль, членство сохранено в заглушке.
            await db.delete(uo)
        else:
            uo.dept_id = None
            uo.dept_assigned_at = None
            uo.position_assigned_at = None

    await db.commit()
    # Синк карточки: если поле «Отдел» указывало на снимаемый отдел —
    # перенаправить на оставшееся членство (приоритет — родная орга карточки), иначе очистить.
    user = await db.get(User, user_id)
    if user is not None:
        removed_name = dept.name if dept is not None else None
        if user.department == removed_name:
            # Сначала ищем в «родной» орге карточки, затем в любой.
            remaining = None
            for org_filter in (
                [UserOrganization.org_id == user.org_id] if user.org_id else [],
                [],
            ):
                q = (
                    select(UserOrganization, Department)
                    .join(Department, Department.id == UserOrganization.dept_id)
                    .where(
                        UserOrganization.user_id == user_id,
                        UserOrganization.dept_id.isnot(None),
                        UserOrganization.dept_id != dept_id,
                        *org_filter,
                    )
                    .order_by(UserOrganization.id.asc())
                    .limit(1)
                )
                remaining = (await db.execute(q)).first()
                if remaining:
                    break
            user.department = remaining[1].name if remaining else None
            await db.commit()
    return {"ok": True}


# ── Task Edit Delegates ──────────────────────────────────────────────────────

@router.get("/delegates", response_model=List[DelegateOut])
async def list_delegates(
    target_user_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    q = select(TaskEditDelegate)
    org_ids = get_org_filter(current_user)
    if org_ids is not None:
        q = q.where(TaskEditDelegate.org_id.in_(org_ids))
    if target_user_id is not None:
        q = q.where(TaskEditDelegate.target_user_id == target_user_id)
    rows = (await db.execute(q)).scalars().all()
    user_ids = set()
    for r in rows:
        user_ids.add(r.target_user_id)
        user_ids.add(r.delegate_user_id)
    users_map = {}
    if user_ids:
        for u in (await db.execute(select(User).where(User.id.in_(user_ids)))).scalars().all():  # superadmin-bypass-ok: lookup by pre-computed IDs for delegate enrichment
            users_map[u.id] = u.full_name or u.username
    return [
        DelegateOut(
            id=r.id,
            target_user_id=r.target_user_id,
            target_user_name=users_map.get(r.target_user_id),
            delegate_user_id=r.delegate_user_id,
            delegate_user_name=users_map.get(r.delegate_user_id),
        )
        for r in rows
    ]


@router.post("/delegates", response_model=DelegateOut)
async def add_delegate(
    data: DelegateAdd,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    org_id = get_single_org_id(current_user) or current_user.org_id
    existing = (await db.execute(
        select(TaskEditDelegate).where(
            TaskEditDelegate.target_user_id == data.target_user_id,
            TaskEditDelegate.delegate_user_id == data.delegate_user_id,
        )
    )).scalar_one_or_none()
    if existing:
        raise HTTPException(409, "Такое делегирование уже существует")
    d = TaskEditDelegate(
        target_user_id=data.target_user_id,
        delegate_user_id=data.delegate_user_id,
        org_id=org_id,
    )
    db.add(d)
    await db.commit()
    await db.refresh(d)
    users_map = {}
    for uid in (data.target_user_id, data.delegate_user_id):
        u = await db.get(User, uid)
        if u:
            users_map[uid] = u.full_name or u.username
    return DelegateOut(
        id=d.id,
        target_user_id=d.target_user_id,
        target_user_name=users_map.get(d.target_user_id),
        delegate_user_id=d.delegate_user_id,
        delegate_user_name=users_map.get(d.delegate_user_id),
    )


@router.delete("/delegates/{delegate_id}")
async def remove_delegate(
    delegate_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    d = await db.get(TaskEditDelegate, delegate_id)
    if not d:
        raise HTTPException(404, "Делегирование не найдено")
    await db.delete(d)
    await db.commit()
    return {"ok": True}


# ── Permission check helper ──────────────────────────────────────────────────

async def can_edit_task_of_user(
    editor: User, task_owner_id: int, db: AsyncSession
) -> bool:
    """Check if editor can edit tasks of task_owner_id."""
    # Admin can edit anything
    if editor.role in ADMIN_ROLES:
        return True
    # Own task
    if editor.id == task_owner_id:
        return True
    # Department head?
    head_check = await db.execute(
        select(Department.id).join(
            UserOrganization, UserOrganization.dept_id == Department.id
        ).where(
            Department.head_user_id == editor.id,
            UserOrganization.user_id == task_owner_id,
        )
    )
    if head_check.first():
        return True
    # Custom delegate?
    delegate_check = await db.execute(
        select(TaskEditDelegate.id).where(
            TaskEditDelegate.target_user_id == task_owner_id,
            TaskEditDelegate.delegate_user_id == editor.id,
        )
    )
    if delegate_check.first():
        return True
    # ManagerDepartment: editor is manager of a dept containing task_owner?
    from app.models.manager_department import ManagerDepartment
    md_check = await db.execute(
        select(ManagerDepartment.id).join(
            UserOrganization, UserOrganization.dept_id == ManagerDepartment.dept_id
        ).where(
            ManagerDepartment.manager_user_id == editor.id,
            UserOrganization.user_id == task_owner_id,
        )
    )
    if md_check.first():
        return True
    # ManagerOrganization: editor manages entire org that task_owner belongs to?
    from app.models.manager_organization import ManagerOrganization
    from app.models.user_organization import UserOrganization
    task_owner = await db.get(User, task_owner_id)
    if task_owner:
        mo_check = await db.execute(
            select(ManagerOrganization.id).where(
                ManagerOrganization.manager_user_id == editor.id,
                ManagerOrganization.org_id == task_owner.org_id,
            )
        )
        if mo_check.first():
            return True
        # Also check extra org memberships
        mo_extra = await db.execute(
            select(ManagerOrganization.id).join(
                UserOrganization, UserOrganization.org_id == ManagerOrganization.org_id
            ).where(
                ManagerOrganization.manager_user_id == editor.id,
                UserOrganization.user_id == task_owner_id,
            )
        )
        if mo_extra.first():
            return True
    return False


# ── Excel import ─────────────────────────────────────────────────────────────

@router.get("/import/template")
async def dept_import_template(_=Depends(require_tab('staff'))):
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl не установлен")

    wb = Workbook()
    ws = wb.active
    ws.title = "Отделы и сотрудники"

    headers = ["Отдел", "Родительский отдел", "ФИО сотрудника", "Должность", "Начальник (да/нет)", "Субсидия"]
    hdr_fill = PatternFill("solid", fgColor="1E40AF")
    hdr_font = Font(bold=True, color="FFFFFF")
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")

    examples = [
        ["Отдел закупок", "", "Иванов Иван", "Начальник отдела", "да", ""],
        ["Отдел закупок", "", "Петров Пётр", "Менеджер", "нет", ""],
        ["Склад", "", "Сидоров Сидор", "Кладовщик", "да", ""],
        ["ИТ-отдел", "", "Козлов Андрей", "Разработчик", "нет", ""],
        ["Сектор мониторинга", "Отдел закупок", "Фёдорова Мария", "Аналитик", "да", "ФАДМ_2026"],
    ]
    for ri, row in enumerate(examples, 2):
        for ci, val in enumerate(row, 1):
            ws.cell(row=ri, column=ci, value=val)

    for ci, w in enumerate([30, 25, 30, 25, 18, 20], 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename*=UTF-8''{_url_quote('Шаблон_импорта_отделов.xlsx', safe='-_.~')}"},
    )


@router.post("/import/excel")
async def import_departments_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    """Import department tree from Excel."""
    if not (file.filename or '').lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Только .xlsx / .xls файлы")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl не установлен")

    content = await file.read()
    wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    ws = wb.active

    org_id = get_single_org_id(current_user) or current_user.org_id

    # Load existing users by full_name for matching
    users_res = await db.execute(select(User).where(User.org_id == org_id))  # superadmin-bypass-ok: Excel import lookup by name, not a user-list endpoint
    users_by_name = {}
    for u in users_res.scalars().all():
        if u.full_name:
            users_by_name[u.full_name.strip().lower()] = u

    # Load existing subsidies for matching
    from app.models.subsidy import Subsidy
    subs_res = await db.execute(select(Subsidy))
    subs_by_name = {s.name.strip().lower(): s for s in subs_res.scalars().all() if s.name}

    # Parse header
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(400, "Файл пустой")

    def _norm(v):
        return str(v).strip().lower() if v else ''

    col = {}
    for i, h in enumerate(header_row):
        hs = _norm(h)
        if any(x in hs for x in ('отдел', 'department', 'подразделен')):
            col.setdefault('dept', i)
        if any(x in hs for x in ('родител', 'parent')):
            col.setdefault('parent', i)
        if any(x in hs for x in ('фио', 'сотрудник', 'имя', 'full_name')):
            col.setdefault('name', i)
        if any(x in hs for x in ('должност', 'position', 'позиц')):
            col.setdefault('position', i)
        if any(x in hs for x in ('начальник', 'head', 'руковод')):
            col.setdefault('head', i)
        if any(x in hs for x in ('субсид', 'subsidy')):
            col.setdefault('subsidy', i)

    if 'dept' not in col:
        raise HTTPException(400, "Не найдена колонка «Отдел»")

    def _cell(row, field):
        idx = col.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v else None

    # First pass: collect departments
    dept_cache = {}  # name -> Department
    created_depts = 0
    created_members = 0
    errors = []

    rows = list(ws.iter_rows(min_row=2, values_only=True))

    # Create departments first
    for ri, row in enumerate(rows, 2):
        dept_name = _cell(row, 'dept')
        if not dept_name:
            continue
        key = dept_name.lower()
        if key not in dept_cache:
            subsidy_name = _cell(row, 'subsidy')
            subsidy_id = None
            if subsidy_name:
                sub = subs_by_name.get(subsidy_name.lower())
                if sub:
                    subsidy_id = sub.id

            # Check if already exists in DB
            existing = (await db.execute(
                select(Department).where(Department.org_id == org_id, func.lower(Department.name) == key)
            )).scalar_one_or_none()
            if existing:
                dept_cache[key] = existing
            else:
                dept = Department(name=dept_name.strip(), org_id=org_id, subsidy_id=subsidy_id)
                db.add(dept)
                await db.flush()
                dept_cache[key] = dept
                created_depts += 1

    # Set parent_id for departments
    for ri, row in enumerate(rows, 2):
        dept_name = _cell(row, 'dept')
        parent_name = _cell(row, 'parent')
        if dept_name and parent_name:
            dept = dept_cache.get(dept_name.lower())
            parent = dept_cache.get(parent_name.lower())
            if dept and parent and dept.id != parent.id:
                dept.parent_id = parent.id

    # Second pass: add members
    for ri, row in enumerate(rows, 2):
        dept_name = _cell(row, 'dept')
        user_name = _cell(row, 'name')
        if not dept_name or not user_name:
            continue

        dept = dept_cache.get(dept_name.lower())
        if not dept:
            errors.append({"row": ri, "error": f"Отдел «{dept_name}» не найден"})
            continue

        user = users_by_name.get(user_name.lower())
        if not user:
            errors.append({"row": ri, "error": f"Сотрудник «{user_name}» не найден в системе"})
            continue

        position = _cell(row, 'position')
        is_head = _cell(row, 'head')
        is_head_bool = is_head and is_head.lower() in ('да', 'yes', '1', 'true', 'head', 'начальник')

        # Add member via user_organizations (single source of truth)
        existing_m = (await db.execute(
            select(UserOrganization).where(
                UserOrganization.user_id == user.id,
                UserOrganization.org_id == dept.org_id,
                UserOrganization.dept_id == dept.id,
            )
        )).scalar_one_or_none()
        if not existing_m:
            # Заглушка «Без отдела» для этой же пары (user, org) — переносим её
            # hired_at/ставку в новую dept-строку и удаляем саму заглушку (та же
            # болезнь, что и в add_member: иначе остаётся висеть «Без отдела»).
            null_row = (await db.execute(
                select(UserOrganization).where(
                    UserOrganization.user_id == user.id,
                    UserOrganization.org_id == dept.org_id,
                    UserOrganization.dept_id.is_(None),
                )
            )).scalar_one_or_none()
            base_row = null_row or (await db.execute(
                select(UserOrganization)
                .where(
                    UserOrganization.user_id == user.id,
                    UserOrganization.org_id == dept.org_id,
                )
                .order_by(UserOrganization.id.asc())
                .limit(1)
            )).scalars().first()
            new_uo = UserOrganization(user_id=user.id, org_id=dept.org_id, dept_id=dept.id, position=position)
            if base_row is not None and base_row.hired_at is not None:
                new_uo.hired_at = base_row.hired_at
            if base_row is None:
                # Первая-ЕВЕР строка пары (user, org) — приём на работу, обе даты
                # назначения = дате приёма (владелец, 2026-09-01).
                dates = first_row_assignment_dates(
                    new_uo.hired_at, has_position=bool(position), has_dept=True,
                )
                new_uo.dept_assigned_at = dates.get("dept_assigned_at")
                new_uo.position_assigned_at = dates.get("position_assigned_at")
            else:
                # Уже трудоустроен — импорт добавляет его в новый отдел (перевод).
                new_uo.dept_assigned_at = dept_transfer_date(None)
                position_changed = bool(position) and position != base_row.position
                new_uo.position_assigned_at = position_change_date(
                    position_changed=position_changed,
                    current=base_row.position_assigned_at,
                    explicit=None,
                )
            if base_row is not None:
                if base_row.salary_amount is not None:
                    new_uo.salary_amount = base_row.salary_amount
                if base_row.employment_percent is not None:
                    new_uo.employment_percent = base_row.employment_percent
            if null_row is not None:
                await db.delete(null_row)
            db.add(new_uo)
            created_members += 1
        elif position:
            position_changed = position != existing_m.position
            existing_m.position = position
            existing_m.position_assigned_at = position_change_date(
                position_changed=position_changed,
                current=existing_m.position_assigned_at,
                explicit=None,
            )

        if is_head_bool:
            dept.head_user_id = user.id

        # Update user.department
        user.department = dept.name

    await db.commit()
    return {
        "created_departments": created_depts,
        "created_members": created_members,
        "errors": errors,
    }
