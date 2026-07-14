from io import BytesIO
from datetime import date as _date_type

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

# Phase 29 D-04: date fields added to users PATCH (Lesson 2026-05-13)
_DATE_FIELDS = {
    "license_issued_at",
    "license_expires_at",
    "medical_cert_expires_at",
    "tachograph_card_expires_at",
    "psych_cert_expires_at",
    "periodic_medical_expires_at",
}
from app.database import get_db
from app.models.user import User
from app.auth.jwt import hash_password, require_role, get_current_user, get_org_filter, get_single_org_id, ADMIN_ROLES, ALL_ROLES
from app.schemas.schemas import UserCreate, UserUpdate, UserOut, PermissionsOut
from app.auth.permissions import (
    get_effective_tabs,
    get_effective_actions,
    require_action,
    require_tab,
    ensure_user_org_access,
    _active_org,
)
from typing import List, Optional

router = APIRouter(prefix="/api/users", tags=["users"])

@router.get("/", response_model=List[UserOut])
async def list_users(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
    can_drive: Optional[bool] = Query(None, description="Phase 30.2: фильтр водителей — can_drive=true"),
    fleet_role: Optional[str] = Query(None, description="Фильтр по fleet_role (driver/mechanic/doctor/...)"),
    limit: int = Query(500, ge=1, le=1000),
    org_id: Optional[int] = Query(None, description="Вернуть всех членов этой орг: primary User.org_id ИЛИ членство в user_organizations ИЛИ user_org_access"),
    subsidy_id: Optional[int] = Query(None, description="Вернуть всех сотрудников орг(а), связанных с субсидией: subsidy.org_id + орг по contractor_id + орг с тем же ИНН (учёт дублей орг)"),
):
    q = select(User).order_by(User.full_name)
    # D-09: hide superadmin from non-superadmin callers
    if current_user.role != "superadmin":
        q = q.where(User.role != "superadmin")
    org_ids = get_org_filter(current_user)
    explicit_assign_mode = org_id is not None or subsidy_id is not None
    # Контур вызывающего применяем НИЖЕ — только если НЕ передан явный org_id/subsidy_id.
    # В режиме назначения (явная цель) список = сотрудники орг субсидии
    # ∪ те, кому вызывающий может ставить задачи (подчинённые/управляемые орг и
    # отделы, в т.ч. из ДРУГИХ организаций). Контур НЕ применяем, иначе он резал бы
    # union до активной орг вызывающего (баг: super/owner с активной орг=5 видел 2 чел).
    # Резолвим целевой набор орг: явный org_id + (по субсидии) орг контрагента и
    # орг с тем же ИНН — чтобы дубли организаций (одно юрлицо = несколько org-строк)
    # не урезали список сотрудников.
    target_org_ids: set[int] = set()
    if org_id is not None:
        target_org_ids.add(org_id)
    if subsidy_id is not None:
        from app.models.subsidy import Subsidy
        from app.models.organization import Organization
        from app.models.contractor import Contractor
        sub = await db.get(Subsidy, subsidy_id)
        if sub is not None:
            inns: set[str] = set()
            if sub.org_id:
                target_org_ids.add(sub.org_id)
                o = await db.get(Organization, sub.org_id)
                if o is not None and o.inn:
                    inns.add(o.inn)
            if sub.contractor_id:
                c = await db.get(Contractor, sub.contractor_id)
                if c is not None and c.inn:
                    inns.add(c.inn)
                rows = (await db.execute(
                    select(Organization.id, Organization.inn).where(
                        Organization.contractor_id == sub.contractor_id
                    )
                )).all()
                for oid, oinn in rows:
                    target_org_ids.add(oid)
                    if oinn:
                        inns.add(oinn)
            if inns:
                inn_org_ids = (await db.execute(
                    select(Organization.id).where(Organization.inn.in_(list(inns)))
                )).scalars().all()
                target_org_ids.update(inn_org_ids)
    if explicit_assign_mode:
        from sqlalchemy import or_
        from app.models.user_organization import UserOrganization

        def _org_member_conds(target_list: list[int]) -> list:
            member_q = select(UserOrganization.user_id).where(UserOrganization.org_id.in_(target_list))
            conds = [User.org_id.in_(target_list), User.id.in_(member_q)]
            from app.models.user_org_access import UserOrgAccess
            uoa_q = select(UserOrgAccess.user_id).where(UserOrgAccess.org_id.in_(target_list))
            conds.append(User.id.in_(uoa_q))
            return conds

        if subsidy_id is not None:
            # СТРОГО по субсидии (для всех ролей, включая super): исполнителями и
            # согласующими могут быть только сотрудники орг(а) субсидии или люди
            # с персональным доступом к ней — иначе не закрыться по документам.
            # Никакого union с «кому могу ставить задачи» и контуром.
            from app.models.user_subsidy_access import UserSubsidyAccess
            conds = _org_member_conds(list(target_org_ids)) if target_org_ids else []
            conds.append(User.id.in_(
                select(UserSubsidyAccess.user_id).where(UserSubsidyAccess.subsidy_id == subsidy_id)
            ))
            q = q.where(or_(*conds))
        else:
            from app.services.consent import compute_assignable_user_ids
            # Кому вызывающий может ставить задачи. None = super-роль → ВСЕ (без ограничений).
            assignable = await compute_assignable_user_ids(current_user, db)
            if assignable is not None:
                conds = _org_member_conds(list(target_org_ids)) if target_org_ids else []
                if assignable:
                    conds.append(User.id.in_(list(assignable)))
                # Пустой conds (нет ни орг, ни assignable) → вернуть пусто, не весь список.
                q = q.where(or_(*conds)) if conds else q.where(User.id == -1)
    elif org_ids is not None:
        # Обычный режим (без явной цели) — ограничиваем орг вызывающего.
        # СТРОГИЙ скоуп (как в subsidies scope=strict): орг, где вызывающий реально
        # состоит (primary/членство/UOA) или которыми управляет. Контур аккаунта
        # (root+дети) НЕ применяется: менеджер ВСКС не должен видеть сотрудников
        # дочернего Центрпоиска только из-за дерева орг — наследования по дереву нет.
        from sqlalchemy import or_
        from app.models.user_organization import UserOrganization
        from app.models.user_org_access import UserOrgAccess
        if current_user.role not in ('superadmin', 'account_owner'):
            from app.models.manager_organization import ManagerOrganization
            strict: set[int] = set()
            if current_user.org_id:
                strict.add(int(current_user.org_id))
            strict |= {int(r[0]) for r in (await db.execute(
                select(UserOrganization.org_id).where(UserOrganization.user_id == current_user.id)
            )).all() if r[0]}
            strict |= {int(r[0]) for r in (await db.execute(
                select(UserOrgAccess.org_id).where(UserOrgAccess.user_id == current_user.id)
            )).all() if r[0]}
            strict |= {int(r[0]) for r in (await db.execute(
                select(ManagerOrganization.org_id).where(ManagerOrganization.manager_user_id == current_user.id)
            )).all() if r[0]}
            org_ids = list(strict)
        # Bugfix: членство в орг определяется не только primary User.org_id, но и
        # записями user_organizations (отдел/иерархия) и user_org_access. Иначе
        # сотрудник, состоящий в отделе через user_organizations (виден в Иерархии),
        # но с другим/пустым User.org_id, пропадал из списка «Сотрудники».
        conds = [
            User.org_id.in_(org_ids),
            User.id.in_(select(UserOrganization.user_id).where(UserOrganization.org_id.in_(org_ids))),
            User.id.in_(select(UserOrgAccess.user_id).where(UserOrgAccess.org_id.in_(org_ids))),
        ]
        q = q.where(or_(*conds)) if org_ids else q.where(User.id == -1)
    # Phase 30.2: driver-only filter — can_drive=True OR fleet_role='driver'
    if can_drive is True:
        from sqlalchemy import or_
        q = q.where(or_(User.can_drive == True, User.fleet_role == "driver"))
    elif fleet_role:
        q = q.where(User.fleet_role == fleet_role)
    q = q.limit(limit)
    result = await db.execute(q)
    return result.scalars().all()


@router.get("/assignable-ids")
async def assignable_user_ids(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """ID пользователей, которым текущий может ставить задачи БЕЗ согласия.

    Нужен фронту, чтобы в пикере участников заявки/задачи показать пометку
    «требуется согласование» у тех, кто не в подчинении. all=true → SaaS-роль,
    согласование не нужно ни для кого.
    """
    from app.services.consent import compute_assignable_user_ids
    assignable = await compute_assignable_user_ids(current_user, db)
    if assignable is None:
        return {"all": True, "ids": []}
    return {"all": False, "ids": sorted(assignable)}


@router.post("/", response_model=UserOut)
async def create_user(
    data: UserCreate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_action('user.manage')),
):
    # Email uniqueness check
    existing_email = (await db.execute(select(User).where(User.email == data.email))).scalar_one_or_none()
    if existing_email:
        raise HTTPException(400, "Пользователь с таким email уже существует")

    # Auto-generate username from email if not provided
    username = data.username
    if not username:
        base = data.email.split('@')[0].lower()
        username = base
        suffix = 1
        while (await db.execute(select(User).where(User.username == username))).scalar_one_or_none():
            username = f"{base}{suffix}"
            suffix += 1
    else:
        existing = (await db.execute(select(User).where(User.username == username))).scalar_one_or_none()
        if existing:
            raise HTTPException(400, "Пользователь с таким логином уже существует")

    # Куда создаём сотрудника. org_admin привязан к своей орг через user_org_access
    # (напр. Артеева → Донецкое, org 26), а его базовый User.org_id может быть пуст.
    # Поэтому принимаем org_id из запроса, если он входит в доступные пользователю
    # орг (get_org_filter учитывает JWT-контур + UOA). SaaS-роли — без ограничений.
    from app.auth.jwt import get_org_filter
    allowed = get_org_filter(current_user)  # None = все орг (SaaS), иначе список
    requested = data.org_id
    if current_user.role in ('superadmin', 'account_owner'):
        org_id = requested or current_user.org_id or (allowed[0] if allowed else None)
    elif requested and (allowed is None or requested in allowed):
        org_id = requested
    else:
        org_id = current_user.org_id or (allowed[0] if allowed else None)
    if not org_id:
        raise HTTPException(400, "Необходимо указать организацию")

    # Орг могла быть удалена при мерже дублей по ИНН — тогда id из localStorage/JWT
    # «протух». Без этой проверки insert падал FK-violation → INTERNAL_ERROR без
    # внятного сообщения. Отдаём чистую 400 с просьбой выбрать орг из списка.
    from app.models.organization import Organization
    if not await db.get(Organization, org_id):
        raise HTTPException(400, f"Организация (id={org_id}) не найдена — возможно, была объединена. Выберите организацию из списка.")

    norm_dept = data.department.strip().title() if data.department else data.department
    user = User(
        username=username,
        password_hash=hash_password(data.password),
        role=data.role,
        full_name=data.full_name,
        city=data.city,
        department=norm_dept,
        position=data.position,
        phone=data.phone,
        telegram_id=__import__('re').sub(r'[^0-9]', '', str(data.telegram_id)) if data.telegram_id else None,
        email=data.email,
        avatar=data.avatar,
        is_email_confirmed=True,
        org_id=org_id,
        inn=data.inn,
    )
    db.add(user)
    await db.commit()
    await db.refresh(user)
    # Phase 17.1-05: sync primary org access entry
    if user.org_id:
        await ensure_user_org_access(user.id, user.org_id, user.role, db)
        await db.commit()
    # Sync to department if set
    if user.department:
        await _sync_user_department(user, db)
    return user


@router.get("/in-my-orgs", response_model=List[UserOut])
async def list_users_in_my_orgs(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """
    Возвращает пользователей из всех организаций текущего юзера для autocomplete
    «Исполнитель» в MyTasksView (без hierarchy-фильтра).
    Для не-подчинённых назначаемых create_task выставит consent_needed=true.

    Источники org_id (объединение):
      1. user_organizations.user_id == current_user.id
      2. users.org_id (primary org — legacy для юзеров без user_organizations row)
      3. user_org_access.user_id == current_user.id (per-org role/permissions)
    Если ни одного орг не нашлось — возвращаем всех (fallback, ранний бутстрап).
    Суперадмины скрыты от не-суперадминов (бизнес-правило: суперадмин невидим).
    """
    from app.models.user_organization import UserOrganization
    own_orgs: set[int] = set()

    # 1. user_organizations
    res = await db.execute(
        select(UserOrganization.org_id).where(UserOrganization.user_id == current_user.id)
    )
    own_orgs.update(r[0] for r in res.all() if r[0])

    # 2. primary org_id
    if current_user.org_id:
        own_orgs.add(current_user.org_id)

    # 3. user_org_access
    try:
        from app.models.user_org_access import UserOrgAccess
        res = await db.execute(
            select(UserOrgAccess.org_id).where(UserOrgAccess.user_id == current_user.id)
        )
        own_orgs.update(r[0] for r in res.all() if r[0])
    except Exception:
        pass

    if own_orgs:
        # Все user_id из этих орг (через любой источник)
        user_ids: set[int] = set()
        res = await db.execute(
            select(UserOrganization.user_id).where(UserOrganization.org_id.in_(own_orgs)).distinct()
        )
        user_ids.update(r[0] for r in res.all() if r[0])
        res = await db.execute(
            select(User.id).where(User.org_id.in_(own_orgs))
        )
        user_ids.update(r[0] for r in res.all() if r[0])
        # Сам current_user всегда включён
        user_ids.add(current_user.id)
        if user_ids:
            q = select(User).where(User.id.in_(user_ids)).order_by(User.full_name)
            if current_user.role != "superadmin":
                q = q.where(User.role != "superadmin")
            res = await db.execute(q)
            return res.scalars().all()

    # Fallback: ни одной орг не нашлось — вернуть всех (ранний бутстрап / data-issue)
    q_fallback = select(User).order_by(User.full_name)
    if current_user.role != "superadmin":
        q_fallback = q_fallback.where(User.role != "superadmin")
    res = await db.execute(q_fallback)
    return res.scalars().all()


@router.get("/i-can-act-for", response_model=List[UserOut])
async def list_users_i_can_act_for(
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Возвращает себя + всех видимых через _get_visible_user_ids (подчинённые
    по иерархии + начальник отдела/орг). Используется для picker'а инициатора
    служебной записки: за другого человека делать СЗ может только тот, кому
    он подчинён. См. бизнес-правило в documents.py:generate_document."""
    from app.routers.task_visibility import _get_visible_user_ids
    visible = await _get_visible_user_ids(current_user, db)
    if visible is None:
        # SaaS-wide — отдаём всех; суперадмины скрыты от не-суперадминов
        q_all = select(User).order_by(User.full_name)
        if current_user.role != "superadmin":
            q_all = q_all.where(User.role != "superadmin")
        res = await db.execute(q_all)
        return res.scalars().all()
    # Самого себя гарантированно включаем
    visible = set(visible)
    visible.add(current_user.id)
    if not visible:
        return [current_user]
    q_vis = select(User).where(User.id.in_(visible)).order_by(User.full_name)
    if current_user.role != "superadmin":
        q_vis = q_vis.where(User.role != "superadmin")
    res = await db.execute(q_vis)
    return res.scalars().all()


@router.get("/me", response_model=UserOut)
async def get_me(
    org_id: Optional[int] = Query(None),
    current_user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    """Текущий пользователь. С org_id возвращает effective tabs+actions для этой орг (D-08).

    Superadmin (D-05.3) всегда получает полный набор tabs+actions, даже если у него
    нет org_id — иначе фронтенд гейтит все вкладки через authStore.hasTab() = false.
    """
    out = UserOut.model_validate(current_user)
    out.has_license_scan = bool(current_user.license_scan)
    from app.models.permission import PermissionTab, PermissionAction

    if current_user.role == "superadmin":
        # D-05.3: superadmin sees all tabs + actions, независимо от org_id
        tabs_rows = await db.execute(select(PermissionTab.tab_key))
        actions_rows = await db.execute(select(PermissionAction.action_key))
        tabs = sorted(r for r, in tabs_rows)
        actions = sorted(r for r, in actions_rows)
        out.permissions = PermissionsOut(tabs=tabs, actions=actions)
        return out

    # D-08: resolve permissions for the given org_id (or active org from JWT)
    effective_org_id = org_id or getattr(current_user, "_active_org_id", None) or current_user.org_id

    tabs_rows = await db.execute(select(PermissionTab.tab_key))
    all_tab_keys = {r for r, in tabs_rows}
    actions_rows = await db.execute(select(PermissionAction.action_key))
    all_action_keys = {r for r, in actions_rows}

    async def _resolve_for_org(target_org_id: int) -> tuple[list[str], list[str]]:
        eff = await get_effective_tabs(current_user, db, target_org_id)
        return sorted(eff & all_tab_keys), sorted(eff & all_action_keys)

    chosen_tabs: list[str] = []
    chosen_actions: list[str] = []
    chosen_org_id = effective_org_id

    if effective_org_id is not None:
        chosen_tabs, chosen_actions = await _resolve_for_org(effective_org_id)

    # Phase 30.6 fix: если НЕТ активной орг вообще (effective_org_id is None) и
    # tabs пуст — попробовать другие доступные орг и выбрать где tabs больше всего.
    # ВАЖНО (Filippov fix 2026-06): эскалация НЕ должна срабатывать при явно
    # выбранной активной орг. Иначе employee в ЦЕНТРПОИСК получает org_admin-листы
    # из ВСКС, т.к. там прав больше. Явная активная орг = строгие права этой орг.
    if not chosen_tabs and effective_org_id is None:
        from app.models.user_organization import UserOrganization
        candidate_org_ids: set[int] = set()
        try:
            res = await db.execute(
                select(UserOrganization.org_id).where(
                    UserOrganization.user_id == current_user.id
                )
            )
            candidate_org_ids.update(r[0] for r in res.all() if r[0])
        except Exception:
            pass
        # 24.06-fix: кандидаты — ТОЛЬКО орг с реальным членством. Осиротевшие UOA-орг
        # (org_admin без членства) иначе подсовывали employee admin-набор вкладок
        # (Филиппов получал admin.billing). Скоуп прав = членство, как в jwt.py.
        if current_user.org_id:
            candidate_org_ids.add(current_user.org_id)
        candidate_org_ids.discard(effective_org_id)  # уже пробовали

        best_tabs: list[str] = []
        best_actions: list[str] = []
        best_org_id: Optional[int] = None
        for cand in candidate_org_ids:
            t, a = await _resolve_for_org(cand)
            if len(t) > len(best_tabs):
                best_tabs, best_actions, best_org_id = t, a, cand
        if best_tabs:
            chosen_tabs = best_tabs
            chosen_actions = best_actions
            chosen_org_id = best_org_id

    out.permissions = PermissionsOut(tabs=chosen_tabs, actions=chosen_actions)
    # Подсказка фронту: какой org_id реально использовали (для записи в active_org_id)
    if chosen_org_id is not None and chosen_org_id != effective_org_id:
        # Дополнительное поле в response — фронт может его прочитать после json()
        out.__dict__['effective_org_id'] = chosen_org_id
    return out


@router.get("/{user_id}", response_model=UserOut)
async def get_user(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    if user.role == "superadmin" and current_user.role != "superadmin":
        raise HTTPException(404, "Пользователь не найден")
    return user


@router.patch("/{user_id}", response_model=UserOut)
async def update_user(
    user_id: int,
    data: UserUpdate,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_action('user.manage')),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    if current_user.role == 'account_owner' and user.org_id != current_user.org_id:
        raise HTTPException(403, "Нет доступа")

    update_data = data.dict(exclude_unset=True)
    if "password" in update_data:
        pwd = update_data.pop("password")
        if pwd:
            if current_user.role not in ("superadmin", "account_owner"):
                raise HTTPException(403, "Изменение пароля доступно только владельцу аккаунта и выше")
            user.password_hash = hash_password(pwd)

    # Normalize department name to Title Case
    if "department" in update_data and update_data["department"]:
        update_data["department"] = update_data["department"].strip().title()

    # Strip non-digits from telegram_id
    import re as _re
    if "telegram_id" in update_data and update_data["telegram_id"]:
        update_data["telegram_id"] = _re.sub(r'[^0-9]', '', str(update_data["telegram_id"]))

    for k, v in update_data.items():
        # Phase 29 D-04: coerce ISO string → date for driver date fields (Lesson 2026-05-13)
        if k in _DATE_FIELDS and isinstance(v, str):
            try:
                v = _date_type.fromisoformat(v[:10]) if v else None
            except Exception:
                v = None
        setattr(user, k, v)

    await db.commit()
    await db.refresh(user)

    # Phase 17.1-05: sync primary org access when org/role changes
    if ("org_id" in update_data or "role" in update_data) and user.org_id:
        await ensure_user_org_access(user.id, user.org_id, user.role, db)
        await db.commit()

    # Sync department membership
    if "department" in update_data or "position" in update_data:
        await _sync_user_department(user, db)

    return user


async def _sync_user_department(user: User, db: AsyncSession):
    """Sync user.department to user_organizations table and auto-set hierarchy."""
    from app.models.department import Department
    from app.models.user_organization import UserOrganization as _UO_sync
    from app.models.user_hierarchy import UserHierarchy

    if not user.department or not user.org_id:
        return

    # Find or create department
    dept = (await db.execute(
        select(Department).where(
            Department.org_id == user.org_id,
            Department.name == user.department,
        )
    )).scalar_one_or_none()
    if not dept:
        dept = Department(name=user.department, org_id=user.org_id)
        db.add(dept)
        await db.flush()

    # Ensure UO membership (single source of truth)
    existing = (await db.execute(
        select(_UO_sync).where(
            _UO_sync.user_id == user.id,
            _UO_sync.org_id == user.org_id,
            _UO_sync.dept_id == dept.id,
        )
    )).scalar_one_or_none()
    if existing:
        existing.position = user.position
    else:
        db.add(_UO_sync(user_id=user.id, org_id=user.org_id, dept_id=dept.id, position=user.position))
    await db.flush()

    # Должность в карточке пользователя → head/deputy отдела (двусторонняя синхронизация)
    from app.services.dept_role_sync import sync_head_from_position
    await sync_head_from_position(db, dept, user.id, user.position)

    # If user is head of this dept — auto-create hierarchy for all members
    if dept.head_user_id == user.id:
        await _sync_head_hierarchy(dept, db)

    await db.commit()


async def _sync_head_hierarchy(dept, db: AsyncSession):
    """Make all department members subordinates of the head."""
    from app.models.user_organization import UserOrganization as _UO_hier
    from app.models.user_hierarchy import UserHierarchy

    if not dept.head_user_id:
        return
    members = (await db.execute(
        select(_UO_hier.user_id).where(
            _UO_hier.dept_id == dept.id,
            _UO_hier.user_id != dept.head_user_id,
        )
    )).scalars().all()

    for uid in members:
        existing = (await db.execute(
            select(UserHierarchy).where(
                UserHierarchy.manager_id == dept.head_user_id,
                UserHierarchy.subordinate_id == uid,
            )
        )).scalar_one_or_none()
        if not existing:
            db.add(UserHierarchy(manager_id=dept.head_user_id, subordinate_id=uid))

# ---------------------------------------------------------------------------
# Sync user to contractors
# ---------------------------------------------------------------------------

@router.post("/{user_id}/sync-contractor")
async def sync_user_to_contractor(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_action('user.manage')),
):
    """Create or update a Contractor record from user data (same org, filtered accordingly)."""
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    if not user.full_name and not user.inn:
        raise HTTPException(400, "У сотрудника нет ни ФИО, ни ИНН — нечего записывать")

    from app.models.contractor import Contractor

    # Try to find existing contractor by INN or by name within same org
    contractor = None
    if user.inn:
        contractor = (await db.execute(
            select(Contractor).where(
                Contractor.inn == user.inn,
                Contractor.org_id == user.org_id,
            )
        )).scalar_one_or_none()
    if not contractor and user.full_name:
        contractor = (await db.execute(
            select(Contractor).where(
                Contractor.name == user.full_name,
                Contractor.org_id == user.org_id,
            )
        )).scalar_one_or_none()

    if contractor:
        if user.full_name: contractor.name = user.full_name
        if user.inn: contractor.inn = user.inn
        if user.phone: contractor.phone = user.phone
        if user.email: contractor.email = user.email
        contractor.contact_person = user.full_name
        await db.commit()
        return {"ok": True, "action": "updated", "contractor_id": contractor.id}
    else:
        c = Contractor(
            name=user.full_name or user.username,
            inn=user.inn,
            phone=user.phone,
            email=user.email,
            contact_person=user.full_name,
            org_type="Физ.лицо",
            org_id=user.org_id,
        )
        db.add(c)
        await db.commit()
        await db.refresh(c)
        return {"ok": True, "action": "created", "contractor_id": c.id}


# ---------------------------------------------------------------------------
# Signature (подпись пользователя)
# ---------------------------------------------------------------------------

@router.get("/me/signature")
async def get_my_signature(current_user: User = Depends(get_current_user)):
    return {"signature": current_user.signature_image}


@router.put("/me/signature")
async def save_my_signature(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    sig = body.get("signature", "")
    if not sig or not sig.startswith("data:image/"):
        raise HTTPException(422, "Подпись должна быть в формате data:image/png;base64,...")
    if len(sig) > 500_000:
        raise HTTPException(422, "Подпись слишком большая (макс 500 КБ)")
    current_user.signature_image = sig
    await db.commit()
    return {"ok": True}


@router.delete("/me/signature")
async def delete_my_signature(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    current_user.signature_image = None
    await db.commit()
    return {"ok": True}


@router.get("/me/visibility-debug")
async def visibility_debug(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """27.4-11: диагностика «почему вижу X / могу Y» для current_user.

    Возвращает: JWT-источники org_id, user_organizations memberships,
    resolved org_ids фильтра, метаданные всех орг (name+inn+root+owner),
    список всех видимых субсидий с org meta, effective tabs+actions.

    Для диагностики чужих видимых субсидий — смотреть visible_subsidies[].org_id
    vs legacy_users_org_id: если совпадает → data issue (Subsidy.org_id ошибочно
    указывает на чужую org).
    """
    from app.models.organization import Organization
    from app.models.subsidy import Subsidy
    from app.models.user_organization import UserOrganization
    from app.auth.permissions import get_effective_actions, get_effective_tabs, _active_org

    jwt_org_id = getattr(current_user, '_active_org_id', None)
    jwt_org_ids = getattr(current_user, '_active_org_ids', None)
    legacy_org_id = current_user.org_id

    uo_rows = (await db.execute(
        select(UserOrganization.org_id, UserOrganization.position, UserOrganization.dept_id)
        .where(UserOrganization.user_id == current_user.id)
    )).all()

    resolved_org_ids = get_org_filter(current_user)

    all_org_ids = set()
    for v in (jwt_org_id, legacy_org_id):
        if v:
            all_org_ids.add(v)
    if jwt_org_ids:
        all_org_ids.update(jwt_org_ids)
    if resolved_org_ids:
        all_org_ids.update(resolved_org_ids)
    all_org_ids.update(r[0] for r in uo_rows if r[0])

    orgs_meta = {}
    if all_org_ids:
        rows = (await db.execute(
            select(
                Organization.id, Organization.name, Organization.inn,
                Organization.root_org_id, Organization.owner_user_id,
            ).where(Organization.id.in_(all_org_ids))
        )).all()
        orgs_meta = {
            r[0]: {"name": r[1], "inn": r[2], "root_org_id": r[3], "owner_user_id": r[4]}
            for r in rows
        }

    q = select(Subsidy.id, Subsidy.name, Subsidy.org_id).order_by(Subsidy.id)
    if resolved_org_ids is not None:
        q = q.where(Subsidy.org_id.in_(resolved_org_ids))
    visible_subs_raw = (await db.execute(q)).all()

    sub_org_ids = {s[2] for s in visible_subs_raw if s[2]}
    sub_org_meta = {}
    if sub_org_ids:
        rows = (await db.execute(
            select(Organization.id, Organization.name, Organization.inn)
            .where(Organization.id.in_(sub_org_ids))
        )).all()
        sub_org_meta = {r[0]: {"name": r[1], "inn": r[2]} for r in rows}

    visible_subsidies = [
        {
            "id": s[0],
            "name": s[1],
            "org_id": s[2],
            "org_name": sub_org_meta.get(s[2], {}).get("name") if s[2] else None,
            "org_inn": sub_org_meta.get(s[2], {}).get("inn") if s[2] else None,
        }
        for s in visible_subs_raw
    ]

    active_org = _active_org(current_user)
    tabs = await get_effective_tabs(current_user, db, active_org)
    actions = await get_effective_actions(current_user, db, active_org)

    return {
        "user": {
            "id": current_user.id,
            "username": current_user.username,
            "full_name": current_user.full_name,
            "role": current_user.role,
            "legacy_users_org_id": legacy_org_id,
        },
        "jwt": {
            "_active_org_id": jwt_org_id,
            "_active_org_ids": jwt_org_ids,
        },
        "user_organizations": [
            {"org_id": r[0], "position": r[1], "dept_id": r[2]} for r in uo_rows
        ],
        "resolved_org_ids_for_filter": resolved_org_ids,
        "organizations_meta": orgs_meta,
        "visible_subsidies": visible_subsidies,
        "visible_subsidies_count": len(visible_subsidies),
        "effective": {
            "tabs": sorted(tabs),
            "actions": sorted(actions),
            "purchase_status_change_granted": "purchase.status_change" in actions,
        },
    }


@router.get("/me/photo")
async def get_my_photo(current_user: User = Depends(get_current_user)):
    return {"photo_url": current_user.profile_photo}


@router.put("/me/photo")
async def save_my_photo(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    photo = body.get("photo_url", "")
    if not photo or not photo.startswith("data:image/"):
        raise HTTPException(422, "Фото должно быть в формате data:image/...;base64,...")
    if len(photo) > 2_000_000:
        raise HTTPException(422, "Фото слишком большое (макс 2 МБ)")
    current_user.profile_photo = photo
    await db.commit()
    return {"ok": True}


@router.delete("/me/photo")
async def delete_my_photo(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    current_user.profile_photo = None
    await db.commit()
    return {"ok": True}


@router.get("/{user_id}/photo")
async def get_user_photo(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    return {"photo_url": user.profile_photo}


@router.put("/{user_id}/photo")
async def set_user_photo(
    user_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    photo_url = body.get("photo_url")
    if not photo_url or not isinstance(photo_url, str):
        raise HTTPException(422, "photo_url required")
    user.profile_photo = photo_url
    await db.commit()
    return {"ok": True}


@router.delete("/{user_id}/photo")
async def delete_user_photo(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    user.profile_photo = None
    await db.commit()
    return {"ok": True}


# ── Phase 30.3: скан водительского удостоверения ──────────────────────────────
@router.get("/me/license-scan")
async def get_my_license_scan(
    current_user: User = Depends(get_current_user),
):
    return {"license_scan": current_user.license_scan}


@router.put("/me/license-scan")
async def save_my_license_scan(
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    scan = body.get("license_scan", "")
    if not scan or not scan.startswith("data:image/"):
        raise HTTPException(422, "Скан должен быть data:image/...;base64,...")
    if len(scan) > 3_000_000:
        raise HTTPException(422, "Скан слишком большой (макс 3 МБ)")
    current_user.license_scan = scan
    await db.commit()
    return {"ok": True}


@router.delete("/me/license-scan")
async def delete_my_license_scan(
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    current_user.license_scan = None
    await db.commit()
    return {"ok": True}


@router.get("/{user_id}/license-scan")
async def get_user_license_scan(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    return {"license_scan": user.license_scan}


@router.put("/{user_id}/license-scan")
async def set_user_license_scan(
    user_id: int,
    body: dict,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    scan = body.get("license_scan")
    if not scan or not isinstance(scan, str) or not scan.startswith("data:image/"):
        raise HTTPException(422, "Скан должен быть data:image/...;base64,...")
    if len(scan) > 3_000_000:
        raise HTTPException(422, "Скан слишком большой (макс 3 МБ)")
    user.license_scan = scan
    await db.commit()
    return {"ok": True}


@router.delete("/{user_id}/license-scan")
async def delete_user_license_scan(
    user_id: int,
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_tab('staff')),
):
    user = await db.get(User, user_id)
    if not user:
        raise HTTPException(404, "Пользователь не найден")
    user.license_scan = None
    await db.commit()
    return {"ok": True}


async def _count_user_references(user_id: int, db: AsyncSession) -> dict:
    """Счётчики записей, которые ссылаются на пользователя (для предупреждения
    перед удалением: при удалении исполнитель/автор станут пустыми)."""
    from sqlalchemy import func
    from app.models.purchase import Purchase
    from app.models.task import Task
    from app.models.subsidy_approver import SubsidyApprover
    from app.models.user_subsidy_access import UserSubsidyAccess

    async def _cnt(q):
        return (await db.execute(q)).scalar() or 0

    refs = {
        "purchases": await _cnt(
            select(func.count()).select_from(Purchase)
            .where(Purchase.assigned_user_id == user_id)
        ),
        "active_purchases": await _cnt(
            select(func.count()).select_from(Purchase).where(
                Purchase.assigned_user_id == user_id,
                Purchase.status.notin_(['paid', 'cancelled']),
            )
        ),
        "tasks": await _cnt(
            select(func.count()).select_from(Task).where(
                (Task.assigned_user_id == user_id) | (Task.created_by_id == user_id)
            )
        ),
        "subsidy_approver": await _cnt(
            select(func.count()).select_from(SubsidyApprover)
            .where(SubsidyApprover.user_id == user_id)
        ),
        "subsidy_access": await _cnt(
            select(func.count()).select_from(UserSubsidyAccess)
            .where(UserSubsidyAccess.user_id == user_id)
        ),
    }
    return refs


@router.delete("/{user_id}")
async def delete_user(
    user_id: int,
    confirm: bool = Query(False, description="true — подтверждённое удаление несмотря на связанные записи"),
    org_id: Optional[int] = Query(None, description="Контекст орг: если у пользователя есть другие орги — открепить только от этой, аккаунт НЕ удалять"),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_action('user.manage')),
):
    result = await db.execute(select(User).where(User.id == user_id))
    user = result.scalar_one_or_none()
    if not user:
        raise HTTPException(404, "User not found")
    # Org-admin ограничен своей орг; account_owner/admin/superadmin — по всей учётке.
    if current_user.role == 'org_admin' and user.org_id != current_user.org_id:
        raise HTTPException(403, "Удалять можно только сотрудников своей организации.")

    # Удаление карточки в контексте орг = открепление от этой орг, а не
    # глобальное удаление аккаунта (инцидент с Филипповым 2026-07-06).
    # Глобальный hard delete — только когда это его последняя организация.
    if org_id is not None:
        from app.models.user_organization import UserOrganization
        from app.models.user_org_access import UserOrgAccess
        from app.auth.permissions import remove_user_org_access

        user_orgs: set[int] = set()
        if user.org_id:
            user_orgs.add(user.org_id)
        uo_orgs = (await db.execute(
            select(UserOrganization.org_id).where(UserOrganization.user_id == user_id)
        )).scalars().all()
        uoa_orgs = (await db.execute(
            select(UserOrgAccess.org_id).where(UserOrgAccess.user_id == user_id)
        )).scalars().all()
        user_orgs.update(uo_orgs)
        user_orgs.update(uoa_orgs)

        remaining = user_orgs - {org_id}
        if remaining:
            uo_rows = (await db.execute(
                select(UserOrganization).where(
                    UserOrganization.user_id == user_id,
                    UserOrganization.org_id == org_id,
                )
            )).scalars().all()
            for row in uo_rows:
                await db.delete(row)
            await remove_user_org_access(user_id, org_id, db)
            if user.org_id == org_id:
                user.org_id = min(remaining)
            await db.commit()
            return {"ok": True, "detached": True, "org_id": org_id}
        # Последняя орг — падаем в обычный confirm-flow глобального удаления.

    refs = await _count_user_references(user_id, db)
    if not confirm and (any(refs.values()) or org_id is not None):
        parts = []
        if refs["purchases"]:
            act = f" (из них активных: {refs['active_purchases']})" if refs["active_purchases"] else ""
            parts.append(f"закупок (исполнитель): {refs['purchases']}{act}")
        if refs["tasks"]:
            parts.append(f"задач (исполнитель/автор): {refs['tasks']}")
        if refs["subsidy_approver"]:
            parts.append(f"согласующий в субсидиях: {refs['subsidy_approver']}")
        if refs["subsidy_access"]:
            parts.append(f"персональных доступов к субсидиям: {refs['subsidy_access']}")
        msg = ""
        if org_id is not None:
            msg += ("Это единственная организация сотрудника — аккаунт будет удалён "
                    "ПОЛНОСТЬЮ из системы, а не только из этой организации. ")
        if parts:
            msg += (
                "С сотрудником связаны записи: " + "; ".join(parts) + ". "
                "При удалении исполнитель/автор в закупках и задачах станет пустым, "
                "согласующие потеряют привязку к пользователю, персональные доступы будут удалены. "
            )
        msg += "Подтвердите удаление повторно."
        raise HTTPException(409, detail={
            "code": "CONFIRM_DELETE_USER",
            "message": msg,
            "references": refs,
        })

    await db.delete(user)
    await db.commit()
    return {"ok": True}


# ---------------------------------------------------------------------------
# Dictionaries: departments & positions
# ---------------------------------------------------------------------------

DEFAULT_DEPARTMENTS = [
    "Отдел закупок", "Бухгалтерия", "Юридический отдел", "Склад",
    "Отдел кадров", "ИТ-отдел", "Отдел продаж", "Административный отдел",
    "Финансовый отдел", "Отдел логистики", "Отдел маркетинга",
    "Производственный отдел", "Служба безопасности", "Канцелярия",
]

DEFAULT_POSITIONS = [
    "Начальник отдела", "Заместитель начальника отдела", "Ведущий специалист",
    "Главный специалист", "Специалист", "Менеджер", "Старший менеджер",
    "Бухгалтер", "Главный бухгалтер", "Юрист", "Кладовщик",
    "Администратор", "Секретарь", "Инженер", "Аналитик",
    "Руководитель направления", "Директор", "Заместитель директора",
]


@router.get("/dictionaries/departments")
async def get_department_names(
    org_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Default + custom departments used in this org (optionally filtered by ?org_id=N)."""
    from app.models.department import Department
    if org_id is not None:
        # Только primary org: у мульти-орг (UOA) юзеров User.department — строка
        # родной орги, через UOA-join чужие отделы утекали в список другой орги.
        q = (
            select(User.department)
            .where(
                User.department.isnot(None),
                User.department != "",
                User.org_id == org_id,
            )
            .distinct()
        )
        dq = select(Department.name).where(Department.org_id == org_id, Department.name.isnot(None))
    else:
        org_ids = get_org_filter(current_user)
        q = select(User.department).where(User.department.isnot(None), User.department != "").distinct()
        dq = select(Department.name).where(Department.name.isnot(None))
        if org_ids is not None:
            q = q.where(User.org_id.in_(org_ids))
            dq = dq.where(Department.org_id.in_(org_ids))
    result = await db.execute(q)
    custom = {r[0] for r in result.all() if r[0]}
    dept_table = {r[0] for r in (await db.execute(dq)).all() if r[0]}
    if org_id is not None:
        # Конкретная орга → только её реальные отделы, без generic-шаблонов:
        # тестировщики принимали DEFAULT_DEPARTMENTS за «отделы всех организаций».
        return sorted(custom | dept_table)
    all_depts = sorted(set(DEFAULT_DEPARTMENTS) | custom | dept_table)
    return all_depts


@router.get("/dictionaries/positions")
async def get_position_names(
    org_id: Optional[int] = Query(None),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    """Default + custom positions used in this org (optionally filtered by ?org_id=N)."""
    from app.models.user_org_access import UserOrgAccess
    if org_id is not None:
        # Users whose primary org is org_id OR who have access via user_org_access
        q = (
            select(User.position)
            .outerjoin(UserOrgAccess, UserOrgAccess.user_id == User.id)
            .where(
                User.position.isnot(None),
                User.position != "",
                (User.org_id == org_id) | (UserOrgAccess.org_id == org_id),
            )
            .distinct()
        )
    else:
        org_ids = get_org_filter(current_user)
        q = select(User.position).where(User.position.isnot(None), User.position != "").distinct()
        if org_ids is not None:
            q = q.where(User.org_id.in_(org_ids))
    result = await db.execute(q)
    custom = {r[0] for r in result.all()}
    all_positions = sorted(set(DEFAULT_POSITIONS) | custom)
    return all_positions


# ---------------------------------------------------------------------------
# Excel import
# ---------------------------------------------------------------------------

VALID_ROLES = ("employee", "manager", "admin", "account_owner")


@router.get("/import/template")
async def users_import_template(_=Depends(require_action('user.manage'))):
    """Download xlsx template for bulk user import."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise HTTPException(500, "openpyxl не установлен")

    wb = Workbook()
    ws = wb.active
    ws.title = "Пользователи"

    headers = ["ФИО", "Email", "Логин", "Пароль", "Роль", "Город"]
    required = {"ФИО", "Email", "Логин", "Пароль"}

    header_fill = PatternFill("solid", fgColor="1E40AF")
    req_fill = PatternFill("solid", fgColor="1D4ED8")
    header_font = Font(bold=True, color="FFFFFF")

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = header_font
        cell.fill = req_fill if h in required else header_fill
        cell.alignment = Alignment(horizontal="center")

    # Example row
    example = [
        "Иванов Иван Иванович", "ivanov@example.com", "ivanov",
        "Password123", "employee", "Москва",
    ]
    for ci, val in enumerate(example, 1):
        ws.cell(row=2, column=ci, value=val)

    # Role hint row
    ws.cell(row=3, column=5, value="Допустимые роли: employee, manager, admin, account_owner")

    col_widths = [35, 30, 20, 20, 20, 20]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=ci).column_letter].width = w

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=users_template.xlsx"},
    )


@router.post("/import/excel")
async def import_users_excel(
    file: UploadFile = File(...),
    db: AsyncSession = Depends(get_db),
    current_user: User = Depends(require_action('user.manage')),
):
    """Bulk import users from Excel file."""
    if not (file.filename or '').lower().endswith(('.xlsx', '.xls')):
        raise HTTPException(400, "Поддерживаются только файлы .xlsx / .xls")

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(500, "openpyxl не установлен")

    content = await file.read()
    try:
        wb = load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"Не удалось прочитать файл: {e}")

    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise HTTPException(400, "Файл пустой")

    def _norm(v) -> str:
        return str(v).strip().lower() if v else ''

    # Fuzzy column mapping
    col: dict[str, int] = {}
    for i, h in enumerate(header_row):
        h_str = _norm(h)
        if any(x in h_str for x in ('фио', 'фамил', 'имя', 'full_name', 'full name', 'ф.и.о')):
            col.setdefault('full_name', i)
        elif any(x in h_str for x in ('email', 'e-mail', 'почта', 'mail')):
            col.setdefault('email', i)
        elif any(x in h_str for x in ('логин', 'login', 'username', 'пользовател')):
            col.setdefault('username', i)
        elif any(x in h_str for x in ('пароль', 'password', 'pass')):
            col.setdefault('password', i)
        elif any(x in h_str for x in ('роль', 'role', 'должност')):
            col.setdefault('role', i)
        elif any(x in h_str for x in ('город', 'city', 'населённ')):
            col.setdefault('city', i)

    if 'full_name' not in col:
        raise HTTPException(400, "Не найдена колонка ФИО. Убедитесь что заголовок содержит «ФИО» или «full_name».")

    def _cell(row, field):
        idx = col.get(field)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        if v is None:
            return None
        s = str(v).strip()
        if not s or s.lower() in ('none', 'null', '-', '—'):
            return None
        return s

    # Collect existing emails and usernames for dedup
    email_result = await db.execute(select(User.email).where(User.email.isnot(None)))
    existing_emails = {r[0].lower() for r in email_result if r[0]}
    username_result = await db.execute(select(User.username))
    existing_usernames = {r[0].lower() for r in username_result if r[0]}

    org_id = get_single_org_id(current_user)

    created = 0
    skipped = 0
    errors_list = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        full_name = _cell(row, 'full_name')
        if not full_name:
            skipped += 1
            continue

        email = _cell(row, 'email')
        username = _cell(row, 'username')
        password = _cell(row, 'password')
        role = _cell(row, 'role') or 'employee'
        city = _cell(row, 'city')

        # Validate required fields
        if not email:
            errors_list.append({"row": row_idx, "error": f"Нет email для «{full_name}»"})
            continue
        if not username:
            # Auto-generate username from email
            username = email.split('@')[0]
        if not password:
            errors_list.append({"row": row_idx, "error": f"Нет пароля для «{full_name}»"})
            continue

        # Validate role
        role = role.lower().strip()
        if role not in VALID_ROLES:
            errors_list.append({"row": row_idx, "error": f"Недопустимая роль «{role}» для «{full_name}». Допустимые: {', '.join(VALID_ROLES)}"})
            continue

        # Dedup
        if email.lower() in existing_emails:
            skipped += 1
            continue
        if username.lower() in existing_usernames:
            skipped += 1
            continue

        user = User(
            username=username,
            password_hash=hash_password(password),
            role=role,
            full_name=full_name,
            city=city,
            email=email,
            is_email_confirmed=True,
            org_id=org_id,
        )
        db.add(user)
        existing_emails.add(email.lower())
        existing_usernames.add(username.lower())
        created += 1

    await db.commit()
    return {"created": created, "skipped": skipped, "errors": errors_list}
