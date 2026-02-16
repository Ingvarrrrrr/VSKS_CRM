"""
Скрипт для создания структуры Excel файла с базой данных
Запускать один раз для создания файла базы данных
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from datetime import datetime
import sys
import io

# Настройка кодировки для Windows
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')

def create_database_structure(filename='CRM_База_Данных.xlsx'):
    """Создание структуры базы данных в Excel"""
    
    wb = openpyxl.Workbook()
    
    # Удаляем лист по умолчанию
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Создаем листы
    create_goods_service_sheet(wb)
    create_contract_items_sheet(wb)
    create_contractors_sheet(wb)
    create_subsidies_sheet(wb)
    create_categories_feo_sheet(wb)
    create_categories_app_sheet(wb)
    create_payments_sheet(wb)
    create_feo_base_sheet(wb)
    create_feo_applications_sheet(wb)
    create_documents_sheet(wb)
    create_dashboard_sheet(wb)
    create_registry_sheet(wb)
    # Справочники и связи для ФЭО/категорий/товаров
    create_links_subsidy_feo_sheet(wb)
    create_links_feo_cost_items_sheet(wb)
    create_cost_breakdown_sheet(wb)
    create_breakdown_items_sheet(wb)
    # Справочники для заполнения вручную
    create_feo_directions_reference_sheet(wb)
    create_feo_cost_items_reference_sheet(wb)
    create_unique_items_reference_sheet(wb)
    create_direction_cost_item_links_sheet(wb)
    
    # Сохраняем файл
    wb.save(filename)
    print(f"[OK] Структура базы данных создана в файле: {filename}")
    return filename

def create_goods_service_sheet(wb):
    """Создание листа GoodsService (основная таблица)"""
    ws = wb.create_sheet('GoodsService', 0)
    
    headers = [
        'ID', 'Номер договора', 'Дата договора', 'Тип договора', 'Вид договора',
        'Субсидия_ID', 'Номер закупки', 'Номер заказа', 'Предмет договора', 'Детальное описание',
        'Контрагент_ID', 'Статус договора', 'Статус закупки', 'Стадия исполнения',
        'Дата начала', 'Дата окончания', 'Срок исполнения', 'НМЦК', 'Цена без НДС',
        'Сумма НДС', 'Цена с НДС', 'Экономия', 'Процент экономии', 'Законтрактовано',
        'Поставлено', 'Оплачено', 'Остаток к оплате', 'Остаток к поставке',
        'Применение НДС', 'Ставка НДС', 'Способ оплаты', 'Форма оплаты',
        'Размер аванса', 'Срок оплаты', 'Направление расходов ФЭО',
        'Тип расходов ФЭО', 'Направление из приложения', 'Тип конкретизированный',
        'Ответственный', 'Город', 'Комментарии', 'Дата создания',
        'Дата изменения', 'Автор', 'Редактор'
    ]
    
    # Заголовки
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Замораживаем первую строку
    ws.freeze_panes = 'A2'
    
    # Настраиваем ширину столбцов
    column_widths = [50, 120, 100, 120, 100, 80, 120, 120, 200, 300, 80, 120, 120, 150, 
                     100, 100, 80, 120, 120, 100, 120, 100, 100, 120, 120, 120, 120, 120,
                     100, 80, 120, 120, 100, 100, 200, 200, 200, 200, 150, 100, 300, 120, 120, 150, 150]
    for col, width in enumerate(column_widths, 1):
        if col <= len(headers):
            ws.column_dimensions[get_column_letter(col)].width = width / 7
    
    # Выпадающие списки
    setup_data_validation(ws, 4, ['Поставка', 'Услуги', 'ГПХ', 'Ремонт ТС'], start_row=2)
    setup_data_validation(ws, 5, ['Разовый', 'Рамочный'], start_row=2)
    setup_data_validation(ws, 12, ['Плановый', 'Подтвержденный', 'Ведутся работы', 'Исполнен', 'Расторгнут', 'Просрочен'], start_row=2)
    setup_data_validation(ws, 13, ['Плановый', 'Подтвержденный', 'Ведутся работы'], start_row=2)
    setup_data_validation(ws, 29, ['Да', 'Нет'], start_row=2)
    setup_data_validation(ws, 31, ['Безналичный', 'Наличный'], start_row=2)
    setup_data_validation(ws, 32, ['Предоплата', 'Постоплата', 'Поэтапная'], start_row=2)
    
    # Формулы (для строки 2)
    add_formulas_to_goods_service(ws, 2)
    
    # Условное форматирование
    setup_conditional_formatting(ws, 12)  # Столбец статуса

def create_contract_items_sheet(wb):
    """Создание листа Состав_договора"""
    ws = wb.create_sheet('Состав_договора')
    
    headers = [
        'ID', 'Договор_ID', 'Номер позиции', 'Наименование', 'Артикул',
        'Количество', 'Ед. изм.', 'Код ОКЕИ', 'Цена за единицу',
        'Итоговая цена за единицу', 'НДС %', 'Сумма НДС', 'Стоимость позиции',
        'Поставлено количество', 'Оплачено', 'Страна происхождения',
        'Производитель', 'Гарантийный срок', 'Технические характеристики',
        'Описание услуги', 'Направление расходов ФЭО', 'Тип расходов ФЭО',
        'Направление из приложения', 'Тип конкретизированный', 'Комментарии'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'

def create_contractors_sheet(wb):
    """Создание листа Контрагенты"""
    ws = wb.create_sheet('Контрагенты')
    
    headers = [
        'ID', 'Контрагент', 'Полное наименование', 'ИНН', 'КПП', 'ОГРН',
        'ОКПО', 'ОКТМО', 'ФИО руководителя', 'Должность руководителя',
        'Основание действия', 'Номер доверенности', 'Дата доверенности',
        'Кем выдана доверенность', 'Расчётный счёт', 'Кореспондентский счёт',
        'БИК банка', 'Наименование банка', 'Юридический адрес',
        'Почтовый адрес', 'Фактический адрес', 'Телефон организации',
        'Факс', 'E-mail организации', 'Веб-сайт', 'Контактное лицо',
        'Должность контактного лица', 'Телефон контактного лица',
        'E-mail контактного лица', 'Дополнительные контакты',
        'Дата создания', 'Дата изменения', 'Активен'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    setup_data_validation(ws, 33, ['Да', 'Нет'], start_row=2)

def create_subsidies_sheet(wb):
    """Создание листа Субсидии"""
    ws = wb.create_sheet('Субсидии')
    
    headers = [
        'ID', 'Наименование', 'Краткое наименование', 'Ведомство', 'Год',
        'Общий объём', 'Законтрактовано', 'Планируется', 'Поставлено',
        'Оплачено', 'Остаток', 'Активна'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    setup_data_validation(ws, 4, ['Минпрос', 'Минтруд', 'ФАДМ', 'Регионы'], start_row=2)
    setup_data_validation(ws, 12, ['Да', 'Нет'], start_row=2)

def create_categories_feo_sheet(wb):
    """Создание листа Категории_из_ФЭО"""
    ws = wb.create_sheet('Категории_из_ФЭО')
    
    headers = ['Категория'] + [f'Подкатегория_{i}' for i in range(1, 11)]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'

def create_categories_app_sheet(wb):
    """Создание листа Категории_из_приложения"""
    ws = wb.create_sheet('Категории_из_приложения')
    
    headers = ['Категория'] + [f'Подкатегория_{i}' for i in range(1, 11)]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'

def create_feo_base_sheet(wb):
    """Создание листа ФЭО_База"""
    ws = wb.create_sheet('ФЭО_База')
    
    headers = [
        'ID', 'Субсидия_ID', 'Субсидия', 'Номер приложения',
        'Направление расходов', 'Расчетный объем (тыс. руб)',
        'Дата создания', 'Дата изменения'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'

def create_feo_applications_sheet(wb):
    """Создание листа ФЭО_Приложения"""
    ws = wb.create_sheet('ФЭО_Приложения')
    
    headers = [
        'ID', 'ФЭО_ID', 'Название приложения', 'Направление расходов',
        'Методика расчета', 'Единица измерения', 'Количество единиц',
        'Плановая стоимость за единицу', 'Плановый объем расходов',
        'Дата создания', 'Дата изменения'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'


def create_links_subsidy_feo_sheet(wb):
    """Создание листа Связи_Субсидия_ФЭО (связь субсидий и направлений ФЭО)"""
    ws = wb.create_sheet('Связи_Субсидия_ФЭО')

    headers = [
        'ID',
        'Субсидия_ID',
        'Субсидия',
        'ФЭО_Направление_ID',
        'ФЭО_Направление',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'


def create_links_feo_cost_items_sheet(wb):
    """Создание листа Связи_ФЭО_Статьи (связь направлений ФЭО и статей затрат)"""
    ws = wb.create_sheet('Связи_ФЭО_Статьи')

    headers = [
        'ID',
        'ФЭО_Направление_ID',
        'ФЭО_Направление',
        'Статья_ID',
        'Наименование статьи затрат',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'


def create_cost_breakdown_sheet(wb):
    """Создание листа Статьи_Разбивка (разбивка статей затрат)"""
    ws = wb.create_sheet('Статьи_Разбивка')

    headers = [
        'ID',
        'Статья_ID',
        'Наименование статьи затрат',
        'Разбивка_ID',
        'Разбивка_статьи',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'


def create_breakdown_items_sheet(wb):
    """Создание листа Разбивка_Товары (связь разбивки статей и товаров/услуг)"""
    ws = wb.create_sheet('Разбивка_Товары')

    headers = [
        'ID',
        'Разбивка_ID',
        'Разбивка_статьи',
        'Товар_ID',
        'Наименование товара/услуги',
        'Техническое описание',
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.freeze_panes = 'A2'

def create_payments_sheet(wb):
    """Создание листа Платежи"""
    ws = wb.create_sheet('Платежи')
    
    headers = [
        'ID', 'Договор_ID', 'Дата платежа', 'Номер платежа', 'Сумма',
        'Назначение', 'Статус сверки', 'Источник файла', 'Дата загрузки', 'Комментарии'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    setup_data_validation(ws, 7, ['Не сверен', 'Сверен', 'Ошибка'], start_row=2)

def create_documents_sheet(wb):
    """Создание листа Документы"""
    ws = wb.create_sheet('Документы')
    
    headers = [
        'ID', 'Договор_ID', 'Тип документа', 'Название', 'Номер',
        'Дата', 'Ссылка на файл', 'ID файла', 'Комментарии'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    
    doc_types = ['Протокол', 'ТЗ', 'Спецификация', 'Акт', 'Счет', 'Счет-фактура', 'УПД', 'Накладная', 'Платежное поручение', 'Выписка']
    setup_data_validation(ws, 3, doc_types, start_row=2)

def create_dashboard_sheet(wb):
    """Создание листа Дашборд"""
    ws = wb.create_sheet('Дашборд')
    
    cell = ws.cell(row=1, column=1, value='Дашборд')
    cell.font = Font(bold=True, size=16)
    cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    cell.font = Font(bold=True, color='FFFFFF', size=16)

def create_registry_sheet(wb):
    """Создание листа Реестр_договоров"""
    ws = wb.create_sheet('Реестр_договоров')
    
    headers = [
        'Номер договора', 'Контрагент', 'Предмет договора', 'Субсидия',
        'Сумма договора', 'Законтрактовано', 'Поставлено', 'Оплачено', 'Статус', 'Стадия'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'

def setup_data_validation(ws, column, values, start_row=2, end_row=10000):
    """Настройка выпадающих списков"""
    if not values:
        return
    
    dv = DataValidation(type="list", formula1=f'"{",".join(values)}"', allow_blank=True)
    col_letter = get_column_letter(column)
    dv.add(f"{col_letter}{start_row}:{col_letter}{end_row}")
    ws.add_data_validation(dv)

def add_formulas_to_goods_service(ws, row):
    """Добавление формул в GoodsService"""
    # Сумма НДС (столбец T, 20)
    ws.cell(row=row, column=20).value = f'=IF(AC{row}="Да"; S{row}*AD{row}/100; 0)'
    
    # Цена с НДС (столбец U, 21)
    ws.cell(row=row, column=21).value = f'=S{row}+T{row}'
    
    # Экономия (столбец V, 22)
    ws.cell(row=row, column=22).value = f'=R{row}-U{row}'
    
    # Процент экономии (столбец W, 23)
    ws.cell(row=row, column=23).value = f'=IF(R{row}>0; V{row}/R{row}*100; 0)'
    
    # Остаток к оплате (столбец AA, 27)
    ws.cell(row=row, column=27).value = f'=U{row}-Z{row}'

def setup_conditional_formatting(ws, status_column):
    """Настройка условного форматирования по статусам"""
    from openpyxl.formatting.rule import CellIsRule
    
    status_col_letter = get_column_letter(status_column)
    
    # Плановый - желтый
    yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    ws.conditional_formatting.add(f'{status_col_letter}2:{status_col_letter}10000',
                                   CellIsRule(operator='equal', formula=['"Плановый"'], fill=yellow_fill))
    
    # Подтвержденный - синий
    blue_fill = PatternFill(start_color='4A86E8', end_color='4A86E8', fill_type='solid')
    ws.conditional_formatting.add(f'{status_col_letter}2:{status_col_letter}10000',
                                   CellIsRule(operator='equal', formula=['"Подтвержденный"'], fill=blue_fill))
    
    # Ведутся работы - зеленый
    green_fill = PatternFill(start_color='6AA84F', end_color='6AA84F', fill_type='solid')
    ws.conditional_formatting.add(f'{status_col_letter}2:{status_col_letter}10000',
                                   CellIsRule(operator='equal', formula=['"Ведутся работы"'], fill=green_fill))
    
    # Исполнен - серый
    gray_fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
    ws.conditional_formatting.add(f'{status_col_letter}2:{status_col_letter}10000',
                                   CellIsRule(operator='equal', formula=['"Исполнен"'], fill=gray_fill))
    
    # Расторгнут - красный
    red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
    ws.conditional_formatting.add(f'{status_col_letter}2:{status_col_letter}10000',
                                   CellIsRule(operator='equal', formula=['"Расторгнут"'], fill=red_fill))
    
    # Просрочен - оранжевый
    orange_fill = PatternFill(start_color='FF9900', end_color='FF9900', fill_type='solid')
    ws.conditional_formatting.add(f'{status_col_letter}2:{status_col_letter}10000',
                                   CellIsRule(operator='equal', formula=['"Просрочен"'], fill=orange_fill))


def create_feo_directions_reference_sheet(wb):
    """Создание справочника направлений ФЭО для ручного заполнения"""
    ws = wb.create_sheet('Справочник_Направления_ФЭО')
    
    headers = [
        'ID',
        'Группа',
        'Направление расходования ФЭО',
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    
    # Настройка выпадающего списка для групп
    groups = ['ФАДМ', 'Минпросвет', 'ЗО', 'ХО', 'ДНР', 'ЛНР', 'Минтруд', 'КОС']
    setup_data_validation(ws, 2, groups, start_row=2)
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60


def create_feo_cost_items_reference_sheet(wb):
    """Создание справочника статей затрат для ручного заполнения"""
    ws = wb.create_sheet('Справочник_Статьи_Затрат')
    
    headers = [
        'ID',
        'Группа',
        'Наименование статьи затрат',
        'Единица измерения',
        'Плановая цена за единицу',
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    
    # Настройка выпадающего списка для групп
    groups = ['ФАДМ', 'Минпросвет', 'ЗО', 'ХО', 'ДНР', 'ЛНР', 'Минтруд', 'КОС']
    setup_data_validation(ws, 2, groups, start_row=2)
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25


def create_unique_items_reference_sheet(wb):
    """Создание справочника уникальных товаров/услуг для ручного заполнения"""
    ws = wb.create_sheet('Справочник_Уникальные_Товары')
    
    headers = [
        'ID',
        'Наименование товара/услуги',
        'Техническое описание',
        'Статья_Затрат_ID',
        'Наименование статьи затрат',
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 80
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60


def create_direction_cost_item_links_sheet(wb):
    """Создание листа связей между направлениями ФЭО и статьями затрат"""
    ws = wb.create_sheet('Связи_Направление_Статья')
    
    headers = [
        'ID',
        'Направление_ID',
        'Направление расходования ФЭО',
        'Статья_ID',
        'Наименование статьи затрат',
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    ws.freeze_panes = 'A2'
    
    # Настройка ширины столбцов
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 60

if __name__ == '__main__':
    print("Создание структуры базы данных...")
    filename = create_database_structure()
    print(f"\n[OK] Готово! Файл создан: {filename}")
    print("\nСледующий шаг: Запустите main.py для работы с приложением")

