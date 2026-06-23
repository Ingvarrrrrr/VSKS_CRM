"""Excel-экспорт сохранённых отчётов (Phase 25-08).

Поддерживает 3 типа отчётов:
- list: групповая таблица с подзаголовками + subtotals + grand total
- pivot: multi-row header (для нескольких col-dimensions используется merge_cells), totals row + totals column + grand total
- dashboard: workbook = N листов (по одному на виджет) с таблицами данных + parameters лист

Reuse:
- backend/app/routers/purchase_export.py: PatternFill #4472C4 для header
- backend/app/routers/dashboard.py: financial-plan/export.xlsx pattern
"""

import io
from datetime import datetime
from typing import Dict, Any, List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from app.services.field_registry import FIELDS, get_field

HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
HEADER_FONT = Font(color='FFFFFF', bold=True, size=11)
HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
GROUP_HEADER_FILL = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
SUBTOTAL_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
GRAND_TOTAL_FILL = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')
BORDER_THIN = Border(
    left=Side(style='thin', color='BFBFBF'),
    right=Side(style='thin', color='BFBFBF'),
    top=Side(style='thin', color='BFBFBF'),
    bottom=Side(style='thin', color='BFBFBF'),
)


def _label_for(key: str) -> str:
    """Лейбл колонки: из field_registry или ключ как есть."""
    f = get_field(key)
    return f.label if f else key


def _excel_number_format(field_format: Optional[str], unit: Optional[str] = None) -> str:
    """Excel number format string по типу поля."""
    if unit == 'thousands' or field_format == 'thousands':
        return '#,##0.00" тыс.₽"'
    if unit == 'millions' or field_format == 'millions':
        return '#,##0.00" млн.₽"'
    if unit == 'rub' or field_format == 'rub':
        return '#,##0.00 ₽'
    if field_format == 'date_dmy' or field_format == 'dmy':
        return 'DD.MM.YYYY'
    if field_format == 'pct':
        return '0.0%'
    return '@'


def _setup_header_cell(cell, value):
    cell.value = value
    cell.font = HEADER_FONT
    cell.fill = HEADER_FILL
    cell.alignment = HEADER_ALIGN
    cell.border = BORDER_THIN


def _make_params_sheet(wb: Workbook, config: Dict, meta: Dict):
    ws = wb.create_sheet('Параметры', 0 if 'Sheet' in wb.sheetnames else None)
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']
    ws['A1'] = 'Параметры отчёта'
    ws['A1'].font = Font(bold=True, size=14)
    row = 3
    items = [
        ('Организация', meta.get('org_name', '—')),
        ('Шаблон', meta.get('report_name', '—')),
        ('Тип', meta.get('kind', '—')),
        ('Автор', meta.get('author', '—')),
        ('Сформирован', meta.get('generated_at', datetime.utcnow().isoformat())),
    ]
    for label, value in items:
        ws.cell(row=row, column=1, value=label).font = Font(bold=True)
        ws.cell(row=row, column=2, value=str(value))
        row += 1

    # Filters
    filters = config.get('filters', {})
    if filters:
        row += 1
        ws.cell(row=row, column=1, value='Фильтры:').font = Font(bold=True)
        row += 1
        for k, v in filters.items():
            ws.cell(row=row, column=1, value=_label_for(k))
            ws.cell(row=row, column=2, value=str(v))
            row += 1

    # Dimensions / measures
    dims = config.get('rows', []) + config.get('cols', [])
    if dims:
        row += 1
        ws.cell(row=row, column=1, value='Измерения:').font = Font(bold=True)
        row += 1
        for d in dims:
            ws.cell(row=row, column=1, value=_label_for(d))
            row += 1

    measures = config.get('measures', [])
    if measures:
        row += 1
        ws.cell(row=row, column=1, value='Меры:').font = Font(bold=True)
        row += 1
        for m in measures:
            ws.cell(row=row, column=1, value=_label_for(m.get('field', '')))
            ws.cell(row=row, column=2, value=m.get('agg', 'sum'))
            row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50


def export_list(rows: List[Dict], config: Dict, meta: Dict) -> bytes:
    """Тип A — реестр. Поддержка group_by с subtotals.

    rows может содержать спец-строки _row_type=group_header|subtotal|grand_total (из group_rows).
    """
    wb = Workbook()

    # Определяем колонки для шапки
    col_keys: List[str] = []
    col_headers: List[str] = []
    col_specs: List[Dict] = []

    # field columns
    field_cols = config.get('columns', [])
    for key in field_cols:
        col_keys.append(key)
        col_headers.append(_label_for(key))
        col_specs.append({'type': 'field', 'key': key, 'unit': 'none'})

    # computed columns (composite/constant/format)
    for cc in config.get('computed_columns', []):
        ckey = cc.get('key')
        ctype = cc.get('type')
        if ctype == 'format':
            # format-spec — найти базовую колонку и применить format
            for spec in col_specs:
                if spec['key'] == cc.get('source'):
                    spec['unit'] = cc.get('format')
            continue
        col_keys.append(ckey)
        col_headers.append(ckey)
        col_specs.append({'type': ctype, 'key': ckey, 'unit': 'none'})

    ws = wb.active
    ws.title = 'Реестр'

    # Title row
    ws.cell(row=1, column=1, value=meta.get('report_name', 'Реестр'))
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    if col_keys:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(col_keys))

    # Header row (row 3)
    for idx, header in enumerate(col_headers, start=1):
        _setup_header_cell(ws.cell(row=3, column=idx), header)
    ws.row_dimensions[3].height = 36

    # Data rows
    current_row = 4
    for row in rows:
        rtype = row.get('_row_type', 'data')
        for idx, (key, spec) in enumerate(zip(col_keys, col_specs), start=1):
            value = row.get(key) if spec['type'] == 'field' else row.get(key, row.get(f'{key}_fmt'))
            cell = ws.cell(row=current_row, column=idx, value=value)

            # Format
            field = get_field(key) if spec['type'] == 'field' else None
            fmt = _excel_number_format(field.format if field else None, spec.get('unit'))
            cell.number_format = fmt

            # Row styling
            if rtype == 'group_header':
                cell.fill = GROUP_HEADER_FILL
                cell.font = Font(bold=True)
            elif rtype == 'subtotal':
                cell.fill = SUBTOTAL_FILL
                cell.font = Font(bold=True, italic=True)
            elif rtype == 'grand_total':
                cell.fill = GRAND_TOTAL_FILL
                cell.font = Font(bold=True, size=11)

            cell.border = BORDER_THIN
        current_row += 1

    # Autofilter + freeze
    if col_keys:
        ws.auto_filter.ref = f'A3:{get_column_letter(len(col_keys))}3'
        ws.freeze_panes = 'A4'
        for idx in range(1, len(col_keys) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 22

    # Params sheet
    _make_params_sheet(wb, config, meta)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_pivot(result: Dict, config: Dict, meta: Dict) -> bytes:
    """Тип B — сводная. Multi-row header + totals row + grand total.

    result: {rows: [...], totals: {...}, dimensions: [...], measures: [...]}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'Сводная'

    rows_dims = config.get('rows', [])
    cols_dims = config.get('cols', [])
    measures = config.get('measures', [])
    calc_columns = config.get('calc_columns', [])

    # Header row
    headers = []
    for d in rows_dims:
        headers.append(_label_for(d))
    for d in cols_dims:
        headers.append(_label_for(d))
    measure_keys = []
    for m in measures:
        field = m.get('field')
        agg = m.get('agg', 'sum')
        label = f'{_label_for(field)} ({agg})'
        headers.append(label)
        measure_keys.append(f'{agg}_{field}')

    calc_keys: List[str] = []
    for cc in calc_columns:
        if cc.get('type') == 'share':
            headers.append(f'% от итога ({_label_for(cc.get("measure", ""))})')
            calc_keys.append(f'share_{cc.get("measure")}')
        elif cc.get('type') == 'cumulative':
            headers.append(f'Нарастающий итог ({_label_for(cc.get("measure", ""))})')
            calc_keys.append(f'cumulative_{cc.get("measure")}')
        elif cc.get('type') == 'delta':
            headers.append(f'Δ {cc.get("actual")} vs {cc.get("plan")}')
            calc_keys.append(f'delta_{cc.get("actual")}_vs_{cc.get("plan")}')

    # Title
    ws.cell(row=1, column=1, value=meta.get('report_name', 'Сводная'))
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    if headers:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))

    # Header row at row 3
    for idx, h in enumerate(headers, start=1):
        _setup_header_cell(ws.cell(row=3, column=idx), h)
    ws.row_dimensions[3].height = 40

    # Data rows
    current_row = 4
    all_dim_keys = rows_dims + cols_dims
    for r in result.get('rows', []):
        col_idx = 1
        for d in all_dim_keys:
            ws.cell(row=current_row, column=col_idx, value=r.get(d)).border = BORDER_THIN
            col_idx += 1
        for mk in measure_keys:
            cell = ws.cell(row=current_row, column=col_idx, value=r.get(mk))
            cell.number_format = '#,##0.00'
            cell.border = BORDER_THIN
            col_idx += 1
        for ck in calc_keys:
            cell = ws.cell(row=current_row, column=col_idx, value=r.get(ck))
            if ck.startswith('share_'):
                cell.number_format = '0.0%'
            else:
                cell.number_format = '#,##0.00'
            cell.border = BORDER_THIN
            col_idx += 1
        current_row += 1

    # Grand total row
    totals = result.get('totals', {})
    if totals:
        total_label_span = len(all_dim_keys)
        cell = ws.cell(row=current_row, column=1, value='Итого')
        cell.font = Font(bold=True)
        cell.fill = GRAND_TOTAL_FILL
        cell.border = BORDER_THIN
        if total_label_span > 1:
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=total_label_span)
        col_idx = total_label_span + 1
        for mk in measure_keys:
            cell = ws.cell(row=current_row, column=col_idx, value=totals.get(mk))
            cell.number_format = '#,##0.00'
            cell.font = Font(bold=True)
            cell.fill = GRAND_TOTAL_FILL
            cell.border = BORDER_THIN
            col_idx += 1

    # Auto-width
    for idx in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(idx)].width = 20
    ws.freeze_panes = 'A4'

    _make_params_sheet(wb, config, meta)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


def export_table(title: str, columns: list, rows: list, meta: dict) -> bytes:
    """Generic «как на экране» экспорт. columns=[{key,title,align?}], rows=[{key:value}]."""
    wb = Workbook()
    ws = wb.active
    ws.title = 'Реестр'

    n_cols = len(columns) or 1

    # Row 1 — title
    ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
    if n_cols > 1:
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)

    # Row 3 — headers
    _ALIGN_MAP = {'start': 'left', 'center': 'center', 'end': 'right'}
    for idx, col in enumerate(columns, start=1):
        _setup_header_cell(ws.cell(row=3, column=idx), col.get('title', col.get('key', '')))
    ws.row_dimensions[3].height = 30

    # Data rows from row 4
    for r_idx, row in enumerate(rows, start=4):
        for c_idx, col in enumerate(columns, start=1):
            value = row.get(col.get('key', ''))
            cell = ws.cell(row=r_idx, column=c_idx, value=value)
            if isinstance(value, (int, float)):
                cell.number_format = '#,##0.##'
            align_key = col.get('align', 'start')
            cell.alignment = Alignment(horizontal=_ALIGN_MAP.get(align_key, 'left'), vertical='center')
            cell.border = BORDER_THIN

    # Autofilter + freeze + column widths
    if columns:
        ws.auto_filter.ref = f'A3:{get_column_letter(n_cols)}3'
        ws.freeze_panes = 'A4'
        for idx in range(1, n_cols + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 22

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


async def export_dashboard(widget_results: List[Dict], config: Dict, meta: Dict) -> bytes:
    """Тип C — дашборд. Каждый виджет = свой лист.

    widget_results: [{widget: {...}, result: {rows, totals}}]
    """
    wb = Workbook()
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    for wr in widget_results:
        widget = wr.get('widget', {})
        result = wr.get('result', {})
        title = widget.get('title', f'Виджет {widget.get("i", "")}')
        # Sanitize sheet name (Excel max 31 chars, no special)
        sheet_name = title[:30].replace('/', '-').replace('\\', '-').replace('[', '(').replace(']', ')')
        ws = wb.create_sheet(sheet_name)

        rows = result.get('rows', [])
        if not rows:
            ws.cell(row=1, column=1, value='Нет данных')
            continue

        # Header (keys из первой строки, исключая _*)
        keys = [k for k in rows[0].keys() if not k.startswith('_')]
        for idx, k in enumerate(keys, start=1):
            _setup_header_cell(ws.cell(row=1, column=idx), _label_for(k))

        for r_idx, row in enumerate(rows, start=2):
            for c_idx, k in enumerate(keys, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=row.get(k))
                if isinstance(row.get(k), (int, float)):
                    cell.number_format = '#,##0.00'
                cell.border = BORDER_THIN

        for idx in range(1, len(keys) + 1):
            ws.column_dimensions[get_column_letter(idx)].width = 20
        ws.freeze_panes = 'A2'

    _make_params_sheet(wb, config, meta)

    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()
