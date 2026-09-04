"""
vehicles_seed.py — Idempotent UPSERT of vehicles from xlsx Голичкова.

Reads seed_data/vehicles_golichkov.xlsx, sheet '09.04.2026' (52 ТС, header row 1).
Uses ON CONFLICT (plate) DO UPDATE SET ... COALESCE — fills NULL columns, never
overwrites admin edits.

Guard (2026-09-02): runs ONLY on a truly empty `vehicles` table (primary
initialization of a clean DB). Once the registry has at least one row —
including a hand-curated registry with rows deliberately deleted — the
seeder no longer runs, because ON CONFLICT can't protect a *deleted* plate:
with no row left to conflict against, a restart would silently re-insert it.

Plan 29-3a2. Decision D-09.
"""
import json
import logging
import os
import re
from datetime import date, datetime
from pathlib import Path

from sqlalchemy import select, text
from sqlalchemy.ext.asyncio import AsyncSession

log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Mapping tables (per RESEARCH R-5)
# ---------------------------------------------------------------------------

TYPE_MAP = {
    'минивэн': 'minivan',
    'легковой': 'car_light',
    'легковая': 'car_light',
    'грузовой (фургон)': 'truck_van',
    'фургон': 'truck_van',
    'грузовой': 'truck_board',
    'бортовой': 'truck_board',
    'цистерна': 'truck_tank',
    'цельнометаллический': 'truck_metal',
    'автобус': 'bus',
    'спецтехника': 'special',
    'специальный': 'special',
    'квадроцикл': 'quadbike',
    'снегоход': 'snowmobile',
    'лодка': 'boat',
    'катер': 'boat',
    'лодочный мотор': 'boat_motor',
    'мотор лодочный': 'boat_motor',
    'прицеп': 'trailer',
}

STATE_MAP = {
    'рабочее': 'working',
    'исправно': 'working',
    'исправный': 'working',
    'неисправно': 'broken',
    'неисправен': 'broken',
    'в ремонте': 'in_repair',
    'требует ремонта': 'needs_repair',
    'утилизировано': 'utilized',
    'утилизирован': 'utilized',
}

FUEL_MAP = {
    'аи-92': 'AI-92', 'ai-92': 'AI-92',
    'аи-95': 'AI-95', 'ai-95': 'AI-95',
    'аи-98': 'AI-98', 'ai-98': 'AI-98',
    'дт': 'DT', 'дизель': 'DT', 'diesel': 'DT',
    'газ': 'GAS',
}


# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def _str(val) -> str:
    """Safe str(), strips whitespace, returns '' for None."""
    if val is None:
        return ''
    return str(val).strip()


def _bool_from_cell(val):
    """'да'/'есть'/'норм'/'+' → True; 'нет'/'отсутствует'/'-' → False; else None."""
    if val is None:
        return None
    s = _str(val).lower()
    if s in ('да', 'есть', 'исправно', 'норм', '+', 'true', 'yes', '1'):
        return True
    if s in ('нет', 'отсутствует', '-', 'false', 'no', '0', ''):
        return False
    return None


def _date_from_cell(val) -> date | None:
    """datetime/date value or DD.MM.YYYY string → date. None on failure."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = _str(val)
    if not s:
        return None
    # Try DD.MM.YYYY
    try:
        return datetime.strptime(s, '%d.%m.%Y').date()
    except ValueError:
        pass
    # Try ISO 8601 first 10 chars
    try:
        return date.fromisoformat(s[:10])
    except ValueError:
        pass
    log.debug(f"vehicles_seed: cannot parse date '{s}', skipping")
    return None


def _map_type(val: str) -> str:
    """Map raw «Кузов» cell value → VehicleType enum string."""
    if not val:
        return 'other'
    s = val.strip().lower()
    # Exact match first (longest key wins — iterate by key length desc)
    for k in sorted(TYPE_MAP, key=len, reverse=True):
        if k in s:
            return TYPE_MAP[k]
    log.warning(f"vehicles_seed: unknown vehicle type '{val}', mapping to 'other'")
    return 'other'


def _map_state(val: str) -> str:
    """Map raw «Состояние» cell value → VehicleState enum string."""
    if not val:
        return 'working'
    s = val.strip().lower()
    for k in sorted(STATE_MAP, key=len, reverse=True):
        if k in s:
            return STATE_MAP[k]
    return 'working'


def _map_fuel(val: str) -> str | None:
    """Map raw «Топливо» cell value → FuelType enum string or None."""
    if not val:
        return None
    s = val.strip().lower()
    for k, v in FUEL_MAP.items():
        if k in s:
            return v
    return None


def _split_brand_model(val: str) -> tuple[str, str]:
    """'Митсубиши Делика' → ('Митсубиши', 'Делика'). One word → (word, '')."""
    if not val:
        return ('', '')
    parts = val.strip().split(None, 1)
    return (parts[0], parts[1] if len(parts) > 1 else '')


def _parse_int_no_spaces(s: str) -> int | None:
    """'151 601' → 151601; '3 000' → 3000; 'abc' → None."""
    cleaned = re.sub(r'\s', '', s)
    cleaned = re.sub(r'[^\d]', '', cleaned)
    if cleaned:
        try:
            return int(cleaned)
        except ValueError:
            pass
    return None


def _parse_last_to(val) -> tuple[int | None, date | None]:
    """Parse «Пробег и дата последнего ТО» cell.

    Examples:
      "151 601км/22.04.2026"  → (151601, date(2026, 4, 22))
      "3 000км до ТО"         → (3000, None)   # note: «до ТО» — treated as remaining
      "151601км"              → (151601, None)
      empty / None            → (None, None)

    Spaces inside numbers are stripped (Russian locale formatting).
    """
    if val is None:
        return (None, None)
    if isinstance(val, (int, float)):
        return (int(val), None)
    s = _str(val)
    if not s:
        return (None, None)

    mileage: int | None = None
    to_date: date | None = None

    # Extract mileage: digits (possibly with spaces) followed by «км»
    km_match = re.search(r'([\d\s]+)\s*[кКkK][мМmM]', s)
    if km_match:
        mileage = _parse_int_no_spaces(km_match.group(1))

    # Extract date: DD.MM.YYYY after «/»
    date_match = re.search(r'/\s*(\d{2}\.\d{2}\.\d{4})', s)
    if date_match:
        try:
            to_date = datetime.strptime(date_match.group(1), '%d.%m.%Y').date()
        except ValueError:
            pass

    # If no date found, also try bare DD.MM.YYYY anywhere in string (no «/» prefix)
    if to_date is None:
        bare_date = re.search(r'\b(\d{2}\.\d{2}\.\d{4})\b', s)
        if bare_date:
            try:
                to_date = datetime.strptime(bare_date.group(1), '%d.%m.%Y').date()
            except ValueError:
                pass

    return (mileage, to_date)


def _parse_tech_inspection(val) -> tuple[date | None, str | None]:
    """Parse «Техосмотр» cell.

    Returns:
        (date, None)       — if a valid date was found
        (None, text_note)  — if text note (e.g. "Отсутствует, надо делать")
        (None, None)       — if empty
    """
    if val is None:
        return (None, None)
    if isinstance(val, datetime):
        return (val.date(), None)
    if isinstance(val, date):
        return (val, None)
    s = _str(val)
    if not s:
        return (None, None)

    # Try ISO
    try:
        return (date.fromisoformat(s[:10]), None)
    except ValueError:
        pass

    # Try DD.MM.YYYY
    try:
        return (datetime.strptime(s, '%d.%m.%Y').date(), None)
    except ValueError:
        pass

    # Try YYYY
    if re.fullmatch(r'\d{4}', s.strip()):
        try:
            return (date(int(s.strip()), 1, 1), None)
        except ValueError:
            pass

    # Fallback: text note
    return (None, s[:500])


# ---------------------------------------------------------------------------
# Owner org lookup
# ---------------------------------------------------------------------------

async def _lookup_owner_org_id(db: AsyncSession, raw_owner: str) -> int | None:
    """Try to resolve owner text → org id via case-insensitive name match.

    Returns org.id if found, None otherwise.
    Fallback (None) means caller should use default_owner_org_id=1 (ВСКС).
    """
    s = raw_owner.strip() if raw_owner else ''
    if not s:
        return None

    try:
        from app.models.organization import Organization
    except ImportError:
        try:
            from .app.models.organization import Organization
        except ImportError:
            # Absolute import path for container context
            import importlib
            mod = importlib.import_module('app.models.organization')
            Organization = mod.Organization

    # Exact ILIKE match
    result = await db.execute(
        select(Organization).where(Organization.name.ilike(s))
    )
    org = result.scalar_one_or_none()
    if org:
        return org.id

    # Substring match for "ВСКС"
    if 'вскс' in s.lower():
        result = await db.execute(
            select(Organization).where(Organization.name.ilike('%ВСКС%'))
        )
        org = result.scalar_one_or_none()
        if org:
            return org.id

    return None


# ---------------------------------------------------------------------------
# Main seed function
# ---------------------------------------------------------------------------

# Sheet name in the updated Голичков xlsx
_SHEET_NAME = '09.04.2026'

# Column index → field mapping (0-based, per task spec)
_COL_TYPE         = 1   # Кузов → vehicle.type
_COL_BRAND_MODEL  = 2   # Марка и модель ТС
_COL_YEAR         = 3   # Год вып.
_COL_PLATE        = 4   # Гос. рег. знак
_COL_LAST_TO      = 5   # Пробег и дата последнего ТО (freeform)
_COL_ODOMETER     = 6   # Пробег км
_COL_COLOR        = 7   # Цвет
_COL_GOST         = 8   # ГОСТ → props.gost_state
_COL_VIN          = 9   # VIN
# col 10 = empty, skip
_COL_RESPONSIBLE  = 11  # Ответственный → props.responsible_person
_COL_OWNER        = 12  # Собственник → owner_org_id lookup
_COL_ASSIGNED     = 13  # У кого в эксплуатации
_COL_PTS          = 14  # ПТС
_COL_STS          = 15  # СТС
_COL_TECH_INSP    = 16  # Техосмотр
_COL_INSURANCE    = 17  # Страховка
_COL_PURCHASE     = 18  # ПОКУПКА
_COL_STATE        = 19  # Состояние
_COL_PAINT        = 20  # Состояние лакокрасочного покрытия
_COL_TECH_COND    = 21  # Сведения о техническом состоянии


def _cell(row, idx):
    """Safe cell accessor — returns None if row too short."""
    return row[idx] if len(row) > idx else None


async def seed_vehicles_from_xlsx(
    db: AsyncSession,
    xlsx_path: str | None = None,
    *,
    force: bool = False,
) -> dict:
    """Idempotent UPSERT seed from xlsx Голичкова (sheet '09.04.2026').

    Args:
        db: AsyncSession — caller manages commit lifecycle.
        xlsx_path: Override default path. Defaults to seed_data/ at backend root.
        force: Unused (kept for API compatibility; ON CONFLICT UPSERT handles idempotency).

    Returns:
        {
          'total_in_xlsx': N,
          'inserted': N,
          'updated': N,
          'errors': [...],
          'reason': 'ok' | 'xlsx_not_found' | 'no_sheet',
        }
    """
    # ------------------------------------------------------------------
    # Guard: only ever run on a truly empty registry (primary init).
    # ON CONFLICT (plate) DO NOTHING/UPDATE prevents duplicate rows, but a
    # vehicle deleted on purpose simply has no row/plate left to conflict
    # with — so on every restart the seeder would silently resurrect it.
    # A non-empty `vehicles` table means the registry has already been
    # curated by hand; the xlsx must not be touched again after that,
    # unless the caller explicitly passes force=True.
    # ------------------------------------------------------------------
    if not force:
        existing_count = (
            await db.execute(text("SELECT COUNT(*) FROM vehicles"))
        ).scalar_one()
        if existing_count and existing_count > 0:
            log.info(
                f"vehicles_seed: реестр уже содержит {existing_count} записей, "
                f"первичное наполнение пропущено"
            )
            return {
                'total_in_xlsx': 0,
                'inserted': 0,
                'updated': 0,
                'errors': [],
                'reason': 'already_seeded',
            }

    if xlsx_path is None:
        # In Docker container: /app/seed_data/vehicles_golichkov.xlsx
        # Local dev: backend/seed_data/vehicles_golichkov.xlsx
        backend_root = Path(__file__).parent.parent.parent  # /app
        xlsx_path = str(backend_root / 'seed_data' / 'vehicles_golichkov.xlsx')
        # Fallback на legacy путь Доработки/ для local dev backward compat
        if not os.path.exists(xlsx_path):
            repo_root = Path(__file__).parent.parent.parent.parent
            legacy_path = str(repo_root / 'Доработки' / 'реестр_транспорта_от_Голичкова_обновление_042026.xlsx')
            if os.path.exists(legacy_path):
                xlsx_path = legacy_path

    if not os.path.exists(xlsx_path):
        log.info(f"vehicles_seed: xlsx not found at '{xlsx_path}', seeding skipped (non-fatal)")
        return {
            'total_in_xlsx': 0,
            'inserted': 0,
            'updated': 0,
            'errors': [],
            'reason': 'xlsx_not_found',
        }

    try:
        import openpyxl
        wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    except Exception as exc:
        log.warning(f"vehicles_seed: failed to open xlsx: {exc}")
        return {
            'total_in_xlsx': 0,
            'inserted': 0,
            'updated': 0,
            'errors': [str(exc)],
            'reason': 'xlsx_open_error',
        }

    if _SHEET_NAME not in wb.sheetnames:
        log.warning(
            f"vehicles_seed: sheet '{_SHEET_NAME}' not found in xlsx "
            f"(sheets: {wb.sheetnames})"
        )
        return {
            'total_in_xlsx': 0,
            'inserted': 0,
            'updated': 0,
            'errors': [f'sheet {_SHEET_NAME} not found'],
            'reason': 'no_sheet',
        }

    ws = wb[_SHEET_NAME]
    # min_row=2: skip header row (row 1 = headers)
    rows = list(ws.iter_rows(min_row=2, values_only=True))

    DEFAULT_OWNER_ORG_ID = 1  # ВСКС (INN 7731178803, Gotcha 2026-04-23)

    inserted = 0
    updated = 0
    errors = []

    for idx, row in enumerate(rows, start=2):  # row number for error messages
        try:
            # ----------------------------------------------------------------
            # Required field: plate (col 4)
            # ----------------------------------------------------------------
            plate_raw = _cell(row, _COL_PLATE)
            if not plate_raw:
                continue  # skip empty rows silently
            plate = _str(plate_raw).upper()
            # Remove internal spaces in plate number
            plate = re.sub(r'\s+', '', plate)[:20]
            if not plate:
                continue

            # ----------------------------------------------------------------
            # VIN (col 9) — "отсутствует" or empty → None
            # ----------------------------------------------------------------
            vin_raw = _str(_cell(row, _COL_VIN))
            if vin_raw.lower() in ('отсутствует', 'нет', ''):
                vin = None
            else:
                vin = vin_raw.upper()[:17]

            # ----------------------------------------------------------------
            # Brand / Model (col 2)
            # ----------------------------------------------------------------
            brand, model = _split_brand_model(_str(_cell(row, _COL_BRAND_MODEL)))
            brand = brand[:100]
            model = model[:100]

            # ----------------------------------------------------------------
            # Color (col 7)
            # ----------------------------------------------------------------
            color_raw = _str(_cell(row, _COL_COLOR))
            color = color_raw[:50] if color_raw else None

            # ----------------------------------------------------------------
            # Year of manufacture (col 3)
            # ----------------------------------------------------------------
            year_raw = _cell(row, _COL_YEAR)
            year_of_manufacture: int | None = None
            if year_raw is not None:
                try:
                    year_of_manufacture = int(float(str(year_raw).strip()))
                except (ValueError, TypeError):
                    pass

            # ----------------------------------------------------------------
            # Current odometer (col 6)
            # ----------------------------------------------------------------
            odometer_raw = _cell(row, _COL_ODOMETER)
            current_odometer_km: int | None = None
            if odometer_raw is not None:
                odo_str = _str(odometer_raw)
                current_odometer_km = _parse_int_no_spaces(odo_str)

            # ----------------------------------------------------------------
            # Last TO mileage + date (col 5) — freeform parse
            # ----------------------------------------------------------------
            last_to_mileage_km, last_to_date = _parse_last_to(_cell(row, _COL_LAST_TO))
            last_to_raw = _str(_cell(row, _COL_LAST_TO))

            # ----------------------------------------------------------------
            # Owner / org (col 12 «Собственник»)
            # ----------------------------------------------------------------
            owner_raw = _str(_cell(row, _COL_OWNER))
            owner_org_id = await _lookup_owner_org_id(db, owner_raw)
            if owner_org_id is None:
                if owner_raw:
                    log.warning(
                        f"vehicles_seed row {idx}: owner '{owner_raw}' not found in orgs, "
                        f"fallback to ВСКС id={DEFAULT_OWNER_ORG_ID}"
                    )
                owner_org_id = DEFAULT_OWNER_ORG_ID

            # ----------------------------------------------------------------
            # Assigned text (col 13 «У кого в эксплуатации»)
            # ----------------------------------------------------------------
            assigned_text = _str(_cell(row, _COL_ASSIGNED))[:100]

            # ----------------------------------------------------------------
            # Vehicle type (col 1 «Кузов»)
            # ----------------------------------------------------------------
            v_type = _map_type(_str(_cell(row, _COL_TYPE)))

            # ----------------------------------------------------------------
            # Insurance until (col 17 «Страховка»)
            # ----------------------------------------------------------------
            insurance_until = _date_from_cell(_cell(row, _COL_INSURANCE))

            # ----------------------------------------------------------------
            # State (col 19 «Состояние»)
            # ----------------------------------------------------------------
            v_state = _map_state(_str(_cell(row, _COL_STATE)))

            # ----------------------------------------------------------------
            # PTS (col 14) — cast to str, strip
            # ----------------------------------------------------------------
            pts_raw = _cell(row, _COL_PTS)
            pts_number: str | None = None
            if pts_raw is not None:
                pts_str = _str(pts_raw)
                if pts_str:
                    pts_number = pts_str[:50]

            # ----------------------------------------------------------------
            # STS (col 15) — may be int (164303033558884) or str
            # ----------------------------------------------------------------
            sts_raw = _cell(row, _COL_STS)
            sts_number: str | None = None
            if sts_raw is not None:
                sts_str = _str(sts_raw)
                if sts_str:
                    sts_number = sts_str[:50]

            # ----------------------------------------------------------------
            # Tech inspection (col 16 «Техосмотр»)
            # ----------------------------------------------------------------
            tech_inspection_until, tech_inspection_note = _parse_tech_inspection(
                _cell(row, _COL_TECH_INSP)
            )

            # ----------------------------------------------------------------
            # Purchase info (col 18 «ПОКУПКА»)
            # ----------------------------------------------------------------
            purchase_raw = _str(_cell(row, _COL_PURCHASE))
            purchase_info = purchase_raw[:200] if purchase_raw else None

            # ----------------------------------------------------------------
            # JSONB props
            # ----------------------------------------------------------------
            props: dict = {}

            gost_raw = _str(_cell(row, _COL_GOST))
            if gost_raw:
                props['gost_state'] = gost_raw[:200]

            responsible_raw = _str(_cell(row, _COL_RESPONSIBLE))
            if responsible_raw:
                props['responsible_person'] = responsible_raw[:200]

            paint_raw = _str(_cell(row, _COL_PAINT))
            if paint_raw:
                props['paint_condition'] = paint_raw[:200]

            tech_cond_raw = _str(_cell(row, _COL_TECH_COND))
            if tech_cond_raw:
                props['technical_condition'] = tech_cond_raw

            # Store raw last_to text if mileage/date couldn't be parsed
            if last_to_raw and (last_to_mileage_km is None and last_to_date is None):
                props['last_to_raw'] = last_to_raw[:500]

            # Store tech_inspection_note if date couldn't be parsed
            if tech_inspection_note:
                props['tech_inspection_note'] = tech_inspection_note[:500]

            # ----------------------------------------------------------------
            # UPSERT: ON CONFLICT (plate) DO UPDATE SET COALESCE(...)
            # — only fills NULL columns, never overwrites admin edits.
            # — props: merge via concat (vehicles.props || EXCLUDED.props)
            # — xmax trick: xmax=0 → new row (inserted), else → updated
            # ----------------------------------------------------------------
            try:
                async with db.begin_nested():
                    result = await db.execute(
                        text("""
                            INSERT INTO vehicles (
                                owner_org_id, assigned_org_id, assigned_text,
                                brand, model, color, vin, plate,
                                type, state, insurance_until,
                                year_of_manufacture, current_odometer_km,
                                last_to_mileage_km, last_to_date,
                                pts_number, sts_number,
                                tech_inspection_until,
                                purchase_info,
                                props, created_at, updated_at
                            ) VALUES (
                                :owner_org_id, NULL, :assigned_text,
                                :brand, :model, :color, :vin, :plate,
                                :type, :state, :insurance_until,
                                :year_of_manufacture, :current_odometer_km,
                                :last_to_mileage_km, :last_to_date,
                                :pts_number, :sts_number,
                                :tech_inspection_until,
                                :purchase_info,
                                CAST(:props AS jsonb), NOW(), NOW()
                            )
                            ON CONFLICT (plate) DO UPDATE SET
                                brand                  = COALESCE(vehicles.brand,                  EXCLUDED.brand),
                                model                  = COALESCE(vehicles.model,                  EXCLUDED.model),
                                color                  = COALESCE(vehicles.color,                  EXCLUDED.color),
                                vin                    = COALESCE(vehicles.vin,                    EXCLUDED.vin),
                                assigned_text          = COALESCE(vehicles.assigned_text,          EXCLUDED.assigned_text),
                                type                   = COALESCE(vehicles.type,                   EXCLUDED.type),
                                state                  = COALESCE(vehicles.state,                  EXCLUDED.state),
                                insurance_until        = COALESCE(vehicles.insurance_until,        EXCLUDED.insurance_until),
                                year_of_manufacture    = COALESCE(vehicles.year_of_manufacture,    EXCLUDED.year_of_manufacture),
                                current_odometer_km    = COALESCE(vehicles.current_odometer_km,    EXCLUDED.current_odometer_km),
                                last_to_mileage_km     = COALESCE(vehicles.last_to_mileage_km,     EXCLUDED.last_to_mileage_km),
                                last_to_date           = COALESCE(vehicles.last_to_date,           EXCLUDED.last_to_date),
                                pts_number             = COALESCE(vehicles.pts_number,             EXCLUDED.pts_number),
                                sts_number             = COALESCE(vehicles.sts_number,             EXCLUDED.sts_number),
                                tech_inspection_until  = COALESCE(vehicles.tech_inspection_until,  EXCLUDED.tech_inspection_until),
                                purchase_info          = COALESCE(vehicles.purchase_info,          EXCLUDED.purchase_info),
                                props                  = vehicles.props || EXCLUDED.props,
                                updated_at             = NOW()
                            RETURNING xmax, id
                        """),
                        {
                            'owner_org_id':         owner_org_id,
                            'assigned_text':         assigned_text,
                            'brand':                 brand,
                            'model':                 model,
                            'color':                 color,
                            'vin':                   vin,
                            'plate':                 plate,
                            'type':                  v_type,
                            'state':                 v_state,
                            'insurance_until':       insurance_until,
                            'year_of_manufacture':   year_of_manufacture,
                            'current_odometer_km':   current_odometer_km,
                            'last_to_mileage_km':    last_to_mileage_km,
                            'last_to_date':          last_to_date,
                            'pts_number':            pts_number,
                            'sts_number':            sts_number,
                            'tech_inspection_until': tech_inspection_until,
                            'purchase_info':         purchase_info,
                            'props':                 json.dumps(props, ensure_ascii=False),
                        },
                    )
                row_res = result.fetchone()
                if row_res is not None:
                    xmax_val = row_res[0]
                    # xmax = 0 → fresh INSERT; xmax != 0 → UPDATE
                    if xmax_val == 0:
                        inserted += 1
                    else:
                        updated += 1
            except Exception as sql_exc:
                err_msg = f"row {idx} (plate={plate_raw!r}): SQL: {sql_exc}"
                log.warning(f"vehicles_seed: SQL error row {idx} — savepoint rolled back: {sql_exc}")
                errors.append(err_msg)

        except Exception as row_exc:
            err_msg = f"row {idx}: parse: {row_exc}"
            log.warning(f"vehicles_seed: skipping malformed row — {err_msg}")
            errors.append(err_msg)
            continue

    await db.commit()

    total = len(rows)
    log.info(
        f"vehicles_seed: total_in_xlsx={total}, inserted={inserted}, "
        f"updated={updated}, errors={len(errors)}"
    )
    return {
        'total_in_xlsx': total,
        'inserted': inserted,
        'updated': updated,
        'errors': errors,
        'reason': 'ok',
    }
